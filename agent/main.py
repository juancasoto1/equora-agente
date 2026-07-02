import os
import re
import io
import csv
import json
import hmac
import hashlib
import base64
import logging
import asyncio
import random
import time
import secrets
import httpx
import bcrypt
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from fastapi import FastAPI, Request, HTTPException, Cookie, UploadFile, File, Form
from fastapi.responses import PlainTextResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv

from agent.brain import generar_respuesta
from agent.markers import MarkerContext, aplicar_marcadores  # noqa: F401  # registra handlers
from agent.memory import (
    inicializar_db, guardar_mensaje, obtener_historial, limpiar_historial,
    guardar_cliente, obtener_cliente, guardar_pedido_pendiente, limpiar_pedido_pendiente,
    obtener_agent_ids_por_telefono, obtener_descuento_promo, obtener_mensaje,
    construir_contexto_placeholders,
    guardar_carrito_activo, limpiar_carrito_activo, obtener_carrito_activo,
    carrito_es_fresco_para_merge,
    puede_enviar_checkout_abandono, marcar_checkout_abandono_enviado,
    registrar_mensaje_usuario, registrar_mensaje_asistente,
    marcar_followup_enviado, marcar_cierre_enviado,
    conversaciones_para_followup, conversaciones_para_cierre,
    clientes_con_carrito_abandonado,
    guardar_checkout_url, limpiar_checkout_url, clientes_con_checkout_abandonado,
    verificar_cierre_enviado,
    get_modo_humano, set_modo_humano,
    obtener_todas_conversaciones, obtener_historial_con_timestamps,
    registrar_difusion, obtener_difusiones,
    guardar_borrador_plantilla, obtener_borradores_plantillas, eliminar_borrador_plantilla,
    guardar_mensaje_difusion, actualizar_status_difusion, obtener_detalle_campana,
    actualizar_wamid_mensaje,
    obtener_metricas_internas, obtener_campana_reciente_para_telefono,
    marcar_opt_out, verificar_opt_out, revertir_opt_out, obtener_opt_outs,
    obtener_clientes_con_estado,
    get_config_value, set_config_value, get_all_config_values, cargar_config_en_env,
    guardar_cliente_import, editar_cliente,
    # Sprint 1 — SaaS multi-tenant
    crear_usuario, obtener_usuario_por_email, obtener_usuario_por_id,
    crear_sesion, verificar_sesion, cerrar_sesion,
    listar_usuarios, actualizar_usuario, obtener_agentes_de_usuario,
    # Sprint 1 — escalación multi-agente
    crear_usuario_interno, autenticar_usuario_interno, obtener_usuarios_internos,
    actualizar_usuario_interno, obtener_usuario_interno_por_id, ping_usuario_interno,
    contar_agentes_activos, obtener_o_crear_admin_interno,
    crear_ticket, obtener_tickets, contar_tickets,
    tomar_ticket, marcar_ticket_pendiente, resolver_ticket,
    obtener_ticket_activo_por_telefono, transferir_ticket,
    # Sprint 2 — notas y templates
    crear_nota_interna, obtener_notas_ticket,
    obtener_templates_rapidos, crear_template_rapido, eliminar_template_rapido,
    # Sprint 3 — auditoría, supervisor, round-robin
    registrar_evento_ticket, obtener_eventos_ticket,
    obtener_stats_equipo, obtener_siguiente_agente_roundrobin,
    # Sprint A — Pipeline
    obtener_pipeline_activo, actualizar_stages_pipeline,
    listar_deals, crear_deal, actualizar_deal, eliminar_deal,
    listar_actividades_deal, agregar_actividad_deal,
    # Pipeline Fase 2 — Calendly/integraciones
    obtener_integration_config, listar_integration_configs, guardar_integration_config,
    # Pipeline Fase 2 — flujo conversacional Calendly
    obtener_calendly_pendiente, guardar_calendly_pendiente, limpiar_calendly_pendiente,
    crear_appointment_pendiente, confirmar_appointment,
    tiene_appointment_confirmado,
)
from agent.calendly import (
    obtener_usuario as calendly_usuario,
    obtener_event_types as calendly_event_types,
    obtener_horarios_disponibles as calendly_horarios,
    crear_cita as calendly_crear_cita,
)
from agent.hubspot import obtener_portal_info as hubspot_portal_info
from agent.inbox import obtener_inbox_html, obtener_login_html, obtener_global_html, obtener_register_html
from agent.providers import obtener_proveedor
from agent.capi import capi_lead, capi_view_content, capi_add_to_cart, capi_initiate_checkout
from agent.tools import (
    crear_checkout_shopify,
    obtener_catalogo_shopify,
    obtener_catalogo_json,
    obtener_secciones_catalogo,
    obtener_producto_por_retailer_id,
    obtener_url_producto,
    cargar_tarifas_envio,
    obtener_costo_envio,
    obtener_umbral_envio_gratis,
)

load_dotenv()

ENVIRONMENT = os.getenv("ENVIRONMENT", "development")
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "")
log_level = logging.DEBUG if ENVIRONMENT == "development" else logging.INFO
logging.basicConfig(level=log_level)
logger = logging.getLogger("agentkit")

proveedor = obtener_proveedor()
PORT = int(os.getenv("PORT", 8000))

# ── Cache de agentes por phone_number_id ─────────────────────────────────────
# Evita consultar la BD en cada mensaje entrante para resolver el agente
_phone_agent_cache: dict[str, dict] = {}
# Sandbox hub: teléfono → agent_id del agente que validó (in-memory, se resetea al reiniciar)
_sandbox_sessions: dict[str, int] = {}


def _sandbox_code_para_agente(slug: str, agent_id: int) -> str:
    """VOCO-001 style sequential code — solo usa el agent_id como consecutivo."""
    return f"VOCO-{agent_id:03d}"


def _sandbox_codigo_a_agent_id(code: str) -> int | None:
    """'VOCO-001' → 1 (agent_id). Acepta también formato legado 'VOCO-EQ001'."""
    import re
    s = code.strip().upper().replace(" ", "")
    # Formato nuevo: VOCO-001
    m = re.match(r"^VOCO-(\d{3})$", s)
    if m:
        return int(m.group(1))
    # Formato legado: VOCO-EQ001
    m = re.match(r"^VOCO-[A-Z]{2}(\d{3})$", s)
    return int(m.group(1)) if m else None


async def _resolver_agente(phone_number_id: str) -> dict:
    """Retorna el agente para este phone_number_id. Fallback: agent_id=1 (Equora)."""
    from agent.memory import obtener_agente_por_phone_id, obtener_agente
    if phone_number_id and phone_number_id in _phone_agent_cache:
        return _phone_agent_cache[phone_number_id]
    agent = None
    if phone_number_id:
        agent = await obtener_agente_por_phone_id(phone_number_id)
        if agent:
            _phone_agent_cache[phone_number_id] = agent
    if not agent:
        agent = await obtener_agente(1)  # fallback Equora
    return agent or {"id": 1, "slug": "equora", "agent_name": "Andrea", "status": "active"}


async def _get_meta_para_agente(agent: dict) -> "ProveedorMeta":
    """Instancia ProveedorMeta con credenciales del agente (BD) o env vars (fallback)."""
    from agent.memory import get_config_value
    from agent.providers.meta import ProveedorMeta
    agent_id = agent.get("id", 1)
    access_token    = await get_config_value("META_ACCESS_TOKEN",    agent_id) or os.getenv("META_ACCESS_TOKEN", "")
    phone_number_id = await get_config_value("META_PHONE_NUMBER_ID", agent_id) or os.getenv("META_PHONE_NUMBER_ID", "")
    verify_token    = await get_config_value("META_VERIFY_TOKEN",    agent_id) or os.getenv("META_VERIFY_TOKEN", "equora-andrea-2024")
    catalog_id      = await get_config_value("META_CATALOG_ID",      agent_id) or os.getenv("META_CATALOG_ID", "")
    return ProveedorMeta(
        access_token=access_token,
        phone_number_id=phone_number_id,
        verify_token=verify_token,
        catalog_id=catalog_id,
    )


async def _get_proveedor_panel(agent_id: int) -> "ProveedorMeta":
    """Proveedor para envíos manuales desde el panel (Conversaciones / Escalaciones).

    Si el agente tiene META_PHONE_NUMBER_ID propio en BD → usa sus credenciales.
    Si NO tiene (agente demo, tenant sin WABA propia) → usa el WABA sandbox de
    Voco (VOCO_SANDBOX_PHONE_NUMBER_ID / VOCO_SANDBOX_ACCESS_TOKEN) para que
    la respuesta llegue desde el mismo número con que el cliente inició la
    conversación de prueba.
    """
    from agent.memory import get_config_value, obtener_agente
    from agent.providers.meta import ProveedorMeta

    db_phone = await get_config_value("META_PHONE_NUMBER_ID", agent_id)
    db_token = await get_config_value("META_ACCESS_TOKEN",    agent_id)

    if db_phone and db_token:
        verify = await get_config_value("META_VERIFY_TOKEN", agent_id) or os.getenv("META_VERIFY_TOKEN", "")
        cat    = await get_config_value("META_CATALOG_ID",   agent_id) or ""
        return ProveedorMeta(access_token=db_token, phone_number_id=db_phone,
                             verify_token=verify, catalog_id=cat)

    # Sin WABA propia → intentar sandbox de Voco
    sb_phone = os.getenv("VOCO_SANDBOX_PHONE_NUMBER_ID", "").strip()
    sb_token = os.getenv("VOCO_SANDBOX_ACCESS_TOKEN",    "").strip()
    if sb_phone and sb_token:
        return ProveedorMeta(access_token=sb_token, phone_number_id=sb_phone,
                             verify_token=os.getenv("META_VERIFY_TOKEN", ""), catalog_id="")

    # Fallback final: comportamiento anterior (env vars genéricas)
    _agente = await obtener_agente(agent_id) or {"id": agent_id}
    return await _get_meta_para_agente(_agente)


# ──────────────────────────────────────────────────────────────────────────
# #79 — Chulitos de confirmación WhatsApp (✓ enviado / ✓✓ entregado / ✓✓ leído)
#
# Bug original: el wamid (id de Meta del mensaje enviado) se leía ANTES de
# que el envío ocurriera, o se llamaba guardar_mensaje() sin pasar wamid en
# absoluto — en ~25 puntos distintos del código. Resultado: ningún mensaje
# mostraba chulitos.
#
# Para que esto NO se repita cada vez que se agregue un seguimiento, webhook
# o canal de envío nuevo, SIEMPRE usar uno de estos dos helpers en vez de
# llamar guardar_mensaje() directamente para mensajes role="assistant":
#
#   1. _guardar_con_wamid(...)        — caso normal: el envío YA ocurrió
#      (proveedor.enviar_mensaje / enviar_botones / enviar_catalogo_productos
#      / etc. ya se llamó) y ahora hay que persistir el mensaje en historial.
#
#   2. _vincular_wamid_post_envio(...) — caso especial: el mensaje ya se
#      guardó en BD ANTES del envío real (ej. el texto de Andrea se fusiona
#      con un mensaje de catálogo/CTA que se arma más abajo). Llamar esto
#      justo después del envío real para vincular el wamid al row existente.
# ──────────────────────────────────────────────────────────────────────────

async def _guardar_con_wamid(
    proveedor_local, telefono: str, contenido: str, agent_id: int = 1,
    role: str = "assistant",
) -> int:
    """Guarda un mensaje en el historial capturando el wamid del envío que
    se acaba de hacer con proveedor_local (lee proveedor_local.ultimo_wamid,
    poblado por CUALQUIER método de envío del provider — enviar_mensaje,
    enviar_botones, enviar_catalogo_productos, enviar_producto, etc.).

    Llamar SIEMPRE justo después de un envío exitoso en vez de invocar
    guardar_mensaje() a mano — evita repetir el bug de #79. Retorna el id
    de la fila guardada (útil si luego hace falta _vincular_wamid_post_envio)."""
    wamid = getattr(proveedor_local, "ultimo_wamid", "") or ""
    return await guardar_mensaje(
        telefono, role, contenido, agent_id=agent_id,
        wamid=wamid, status="sent" if wamid else "",
    )


async def _vincular_wamid_post_envio(proveedor_local, msg_id: int) -> None:
    """Vincula el wamid del último envío al row de historial que YA se
    guardó antes de que el envío ocurriera (caso: texto absorbido por un
    mensaje de catálogo/CTA armado más abajo). No hace nada si no hay
    wamid (envío falló) o msg_id es 0 (guardar_mensaje no se llamó)."""
    if not msg_id:
        return
    wamid = getattr(proveedor_local, "ultimo_wamid", "") or ""
    if wamid:
        await actualizar_wamid_mensaje(msg_id, wamid)

# URL pública del servidor (se usa para el link de la mini-tienda)
_railway_domain = os.getenv("RAILWAY_PUBLIC_DOMAIN", "")
APP_URL = os.getenv("APP_URL") or (f"https://{_railway_domain}" if _railway_domain else f"http://localhost:{PORT}")

# Monto mínimo de pedido configurado en la tienda Lovable
PEDIDO_MINIMO = int(os.getenv("PEDIDO_MINIMO", 50000))

# ── Tiempos de seguimiento — todos configurables desde Railway ───────────────
# Estado 6: follow-up genérico (sin carrito, sin checkout)
FOLLOWUP_MIN         = int(os.getenv("FOLLOWUP_MIN", 10))         # min inactivo → primer mensaje
CIERRE_MIN           = int(os.getenv("CIERRE_MIN", 5))            # min tras follow-up → cierre
FOLLOWUP_MAX_HORAS   = int(os.getenv("FOLLOWUP_MAX_HORAS", 12))   # ventana máxima

# Estado 3: carrito bajo el pedido mínimo ($50K)
CARRITO_MIN_MIN      = int(os.getenv("CARRITO_MIN_MIN", 15))      # min antes del aviso de pedido mínimo
# Estados 4-5: cross-sell (entre $50K y $80K) y carrito listo (≥$80K)
CROSSSELL_MIN_MIN    = int(os.getenv("CROSSSELL_MIN_MIN", 15))    # min antes del aviso de cross-sell/listo
# Compartidos por estados 3-5
CARRITO_MAX_MIN      = int(os.getenv("CARRITO_MAX_MIN", 120))     # min máximo de ventana
CARRITO_COOLDOWN_MIN = int(os.getenv("CARRITO_COOLDOWN_MIN", 120))# min entre avisos al mismo cliente
CARRITO_UNIF_COOLDOWN_SEG = CARRITO_COOLDOWN_MIN * 60

# Estado 2: checkout abandonado (cliente llegó a Shopify pero no terminó)
CHECKOUT_ABANDONO_MIN      = int(os.getenv("CHECKOUT_ABANDONO_MIN", 20))   # min antes del aviso
CHECKOUT_ABANDONO_MAX      = int(os.getenv("CHECKOUT_ABANDONO_MAX", 120))  # ventana máxima
CHECKOUT_COOLDOWN_MIN      = int(os.getenv("CHECKOUT_COOLDOWN_MIN", 360))  # min entre avisos (6h)
# Nota: el cooldown vive en memoria (_checkout_abandono_notif). Si Railway
# reinicia entre avisos, el dict se borra y puede duplicar. Para evitar
# eso completamente habría que persistirlo en BD — backlog Sprint A día 2.
CHECKOUT_ABANDONO_COOLDOWN_SEG = CHECKOUT_COOLDOWN_MIN * 60

# Loop general
CHECK_INTERVAL_SEG   = int(os.getenv("CHECK_INTERVAL_SEG", 30))   # segundos entre cada ciclo del loop

# #77 — Watchdog del loop de seguimientos. Si _loop_seguimientos() no
# arranca un ciclo nuevo en WATCHDOG_UMBRAL_SEG, se asume colgado (ej. una
# llamada externa sin timeout que nunca retorna) y se fuerza un reinicio
# del proceso (Railway lo reinicia automáticamente al detectar el crash).
WATCHDOG_UMBRAL_SEG = max(300, CHECK_INTERVAL_SEG * 10)  # al menos 5 min
_ultimo_ciclo_seguimiento: float = 0.0  # time.monotonic() del último ciclo arrancado

# Legacy (no usados activamente, se conservan por compatibilidad)
ABANDONO_MIN_INACTIVO = CARRITO_MIN_MIN
ABANDONO_MAX_INACTIVO = CARRITO_MAX_MIN
_abandono_notif: dict[str, float] = {}
ABANDONO_COOLDOWN_SEG = CARRITO_UNIF_COOLDOWN_SEG
_carrito_unif_cooldown: dict[str, float] = {}
_carrito_ultimo_estado: dict[str, int] = {}   # phone → último estado enviado (3, 4 ó 5)
_checkout_abandono_notif: dict[str, float] = {}

# Mensajes del sistema — ahora configurables por agente desde el panel.
# Defaults en agent/mensajes.py:MENSAJES. Cada agente puede personalizarlos
# desde Configuración → Mensajes sin tocar código.
# Keys:
#   system.followup          ← era MENSAJE_FOLLOWUP
#   system.cierre            ← era MENSAJE_CIERRE
#   system.checkout_abandono ← era MENSAJE_CHECKOUT_ABANDONO


async def _resolver_agent_id_principal(telefono: str) -> int:
    """Resuelve el agent_id principal de un cliente para enviar mensajes.

    Si el cliente existe bajo varios agents, retorna el primero (estable
    por ID). Si no existe en BD aún, fallback a 1 (legacy/seed). En
    flujos verdaderamente multi-tenant, el caller debería pasar agent_id
    explícito en lugar de invocar este helper.
    """
    aids = await obtener_agent_ids_por_telefono(telefono)
    return aids[0] if aids else 1


async def _total_carrito_fmt(telefono: str, agent_id: int) -> str:
    """Suma los subtotales del carrito_activo del cliente y devuelve string
    formateado tipo COP ('45.700'), o vacío si no hay carrito.

    Útil para inyectar {total} en el contexto de mensajes que se envían
    desde handlers de botón (donde no tenemos el total dinámico calculado
    de antemano como sí lo tenemos en los Estados 4/5 del flujo principal).
    """
    try:
        from agent.memory import obtener_carrito_activo
        items = await obtener_carrito_activo(telefono, agent_id=agent_id)
    except Exception:
        return ""
    if not items:
        return ""
    total = 0
    for it in items:
        sub = it.get("subtotal")
        if sub is None:
            qty = int(it.get("cantidad", it.get("quantity", 1)) or 1)
            pu  = int(it.get("precio_unitario", 0) or 0)
            sub = qty * pu
        try:
            total += int(sub)
        except (TypeError, ValueError):
            pass
    if total <= 0:
        return ""
    return f"{total:,}".replace(",", ".")

# ── Cross-selling desde mini-tienda ─────────────────────────────────────────
# Las tarifas se cargan desde Shopify Admin API al arrancar (lifespan).
# Acceder siempre via obtener_umbral_envio_gratis() y obtener_costo_envio()
# para que reflejen el valor actual (puede venir de Shopify o de env vars).
# In-memory: phone → timestamp del último cross-sell enviado (cooldown 20 min)
# _crosssell_cooldown reemplazado por _carrito_unif_cooldown (ver más abajo)
# Productos estrella para sugerir (se filtran los que ya están en carrito)
PRODUCTOS_ESTRELLA = [
    "lavaloza", "limpiavidrios", "limpiador", "desengrasante",
    "desmanchador", "ambientador", "multiusos", "detergente",
]


@asynccontextmanager
async def lifespan(app: FastAPI):
    await inicializar_db()
    await cargar_config_en_env()   # carga credenciales guardadas en BD → os.environ

    # Pre-cargar defaults de tools.py en os.environ para que el panel de configuración
    # muestre el estado real aunque los valores vengan del código y no de Railway env vars
    from agent.tools import SHOPIFY_STORE
    if SHOPIFY_STORE and not os.environ.get("SHOPIFY_STORE"):
        os.environ["SHOPIFY_STORE"] = SHOPIFY_STORE
        logger.debug("[config] SHOPIFY_STORE pre-cargado desde defaults de tools.py")

    # Verificar que agente Equora (agent_id=1) existe
    from agent.memory import obtener_agente
    equora = await obtener_agente(1)
    if equora:
        logger.info(f"Agente Equora activo: {equora['agent_name']} ({equora['status']})")
    logger.info("Base de datos inicializada y configuración cargada")
    # Pre-calentar catálogo Shopify al arrancar para que _variant_map esté listo
    # antes de que llegue cualquier petición a /tienda/confirmar
    try:
        await obtener_catalogo_shopify(1)  # agent_id=1 (Equora) al arrancar
        logger.info("Catálogo Shopify pre-cargado al arrancar ✅")
    except Exception as e:
        logger.warning(f"No se pudo pre-cargar catálogo Shopify al arrancar: {e}")
    # Pre-cargar SINCRÓNICAMENTE el catálogo de Facebook (retailer_ids) para
    # que product_list funcione desde el primer mensaje. Antes se lanzaba como
    # background task y la primera petición podía llegar con _fb_items vacío.
    try:
        from agent.tools import _cargar_fb_catalog, _fb_items
        await _cargar_fb_catalog(1)  # agent_id=1 (Equora) al arrancar
        logger.info(f"Catálogo Facebook pre-cargado: {len(_fb_items)} items ✅")
    except Exception as e:
        logger.warning(f"No se pudo pre-cargar catálogo Facebook al arrancar: {e}")
    try:
        costo, gratis = await cargar_tarifas_envio()
        logger.info(f"Tarifas de envío: costo=${costo:,} / gratis desde=${gratis:,} ✅")
    except Exception as e:
        logger.warning(f"No se pudieron cargar tarifas de envío: {e}")
    logger.info(f"Servidor AgentKit corriendo en puerto {PORT}")
    logger.info(f"Proveedor de WhatsApp: {proveedor.__class__.__name__}")
    logger.info(
        f"Tiempos de seguimiento — "
        f"follow-up: {FOLLOWUP_MIN} min | cierre: {CIERRE_MIN} min | "
        f"carrito mínimo: {CARRITO_MIN_MIN} min | cross-sell/listo: {CROSSSELL_MIN_MIN} min | ventana max: {CARRITO_MAX_MIN} min (cooldown {CARRITO_COOLDOWN_MIN} min) | "
        f"checkout: {CHECKOUT_ABANDONO_MIN}-{CHECKOUT_ABANDONO_MAX} min (cooldown {CHECKOUT_COOLDOWN_MIN} min) | "
        f"loop cada: {CHECK_INTERVAL_SEG} seg"
    )
    seguimiento_task = asyncio.create_task(_loop_seguimientos())
    watchdog_task = asyncio.create_task(_watchdog_seguimientos())
    try:
        yield
    finally:
        seguimiento_task.cancel()
        watchdog_task.cancel()
        for t in (seguimiento_task, watchdog_task):
            try:
                await t
            except asyncio.CancelledError:
                pass


async def _loop_seguimientos():
    """Cada CHECK_INTERVAL_SEG revisa conversaciones inactivas.
    Prioridad: cierre > checkout abandonado > carrito (estados 3-5) > follow-up genérico."""
    global _ultimo_ciclo_seguimiento
    while True:
        try:
            await asyncio.sleep(CHECK_INTERVAL_SEG)
            # Marcar el ciclo ANTES de procesar — si alguna de las 4 llamadas
            # se cuelga (ej. await sin timeout que nunca retorna), este
            # timestamp deja de avanzar y _watchdog_seguimientos lo detecta.
            _ultimo_ciclo_seguimiento = time.monotonic()
            await _procesar_abandono_checkout()     # prioridad 2: checkout sin completar
            await _procesar_carrito_unificado()     # prioridad 3-5: carrito activo
            await _procesar_followups()             # prioridad 6: sin carrito, sin checkout
            await _procesar_cierres()               # cierre tras follow-up genérico
        except asyncio.CancelledError:
            raise
        except Exception as e:
            logger.error(f"Error en loop de seguimientos: {e}")


async def _watchdog_seguimientos():
    """#77 — vigila que _loop_seguimientos() siga ticando. Revisa cada
    minuto; si pasó WATCHDOG_UMBRAL_SEG sin que arranque un ciclo nuevo,
    el proceso se asume colgado y se fuerza su salida con os._exit(1) —
    Railway reinicia automáticamente el contenedor al detectar el crash."""
    while True:
        await asyncio.sleep(60)
        if _ultimo_ciclo_seguimiento == 0:
            continue  # aún no corrió el primer ciclo (deploy recién hecho)
        inactivo_seg = time.monotonic() - _ultimo_ciclo_seguimiento
        if inactivo_seg > WATCHDOG_UMBRAL_SEG:
            logger.critical(
                f"[watchdog] _loop_seguimientos sin ticar hace {inactivo_seg:.0f}s "
                f"(umbral {WATCHDOG_UMBRAL_SEG}s) — forzando reinicio del proceso"
            )
            os._exit(1)


def _calcular_total_carrito(items: list[dict]) -> int:
    """Calcula el total del carrito usando subtotal o precio_unitario × cantidad."""
    total = 0
    for p in items:
        subtotal = p.get("subtotal", 0)
        if subtotal:
            total += int(subtotal)
        else:
            total += int(p.get("precio_unitario", 0)) * int(p.get("cantidad", 1))
    return total


def _sugerir_productos(nombres_en_carrito: set[str], agent_id: int = 1, max_items: int = 3) -> list[str]:
    """Devuelve hasta max_items productos aleatorios del catálogo del agente que no estén en el carrito."""
    catalogo = obtener_catalogo_json(agent_id)
    sugeridos: list[str] = []
    vistos: set[str] = set()
    # Mezclar categorías para variar el orden en cada mensaje
    categorias = random.sample(PRODUCTOS_ESTRELLA, len(PRODUCTOS_ESTRELLA))
    for estrella in categorias:
        if len(sugeridos) >= max_items:
            break
        # Mezclar el catálogo para no elegir siempre la misma variante
        items_mezclados = random.sample(catalogo, len(catalogo))
        for item in items_mezclados:
            nombre = item.get("producto", "")
            nombre_lower = nombre.lower()
            if (estrella in nombre_lower
                    and nombre not in vistos
                    and not any(en in nombre_lower for en in nombres_en_carrito)):
                sugeridos.append(nombre)
                vistos.add(nombre)
                break
    return sugeridos


async def _procesar_carrito_unificado():
    """
    Maneja estados 3, 4 y 5 con UN solo mensaje por cliente según su total de carrito:
      Estado 3: total < PEDIDO_MINIMO  → empuja a llegar al mínimo       (CARRITO_MIN_MIN)
      Estado 4: PEDIDO_MINIMO ≤ total < ENVIO_GRATIS → cross-sell gratis  (CROSSSELL_MIN_MIN)
      Estado 5: total ≥ ENVIO_GRATIS  → recuerda confirmar                (CROSSSELL_MIN_MIN)
    Cada estado tiene su propia ventana de tiempo configurable.
    Cooldown compartido: CARRITO_COOLDOWN_MIN por cliente.
    """
    ahora = time.time()
    umbral_gratis = obtener_umbral_envio_gratis()
    gratis_fmt = f"{umbral_gratis:,}".replace(",", ".")

    async def _resolver_pedido_minimo(tel: str) -> int:
        """Lee PEDIDO_MINIMO del config_value del agente principal del cliente.
        Fallback a la env var global. Esto evita usar el 50000 hardcoded
        cuando el merchant configuró otro valor en el panel."""
        try:
            from agent.memory import get_config_value
            aid = await _resolver_agent_id_principal(tel)
            raw = await get_config_value("PEDIDO_MINIMO", aid)
            if raw:
                return int(str(raw).strip() or 0) or PEDIDO_MINIMO
        except Exception:
            pass
        return PEDIDO_MINIMO

    # Dos queries con tiempos distintos: estado 3 vs estados 4-5
    clientes_min   = await clientes_con_carrito_abandonado(min_inactivo=CARRITO_MIN_MIN,   max_inactivo=CARRITO_MAX_MIN)
    clientes_cross = await clientes_con_carrito_abandonado(min_inactivo=CROSSSELL_MIN_MIN, max_inactivo=CARRITO_MAX_MIN)

    # Unificar: estado 3 usa clientes_min, estados 4-5 usan clientes_cross
    # Construir dict telefono→items para cada grupo
    map_min   = {t: i for t, i in clientes_min}
    map_cross = {t: i for t, i in clientes_cross}

    # Todos los teléfonos candidatos (sin duplicados)
    todos = set(map_min) | set(map_cross)

    if todos:
        logger.info(
            f"[carrito-loop] {len(todos)} candidatos — "
            f"map_min={len(map_min)} ({CARRITO_MIN_MIN}-{CARRITO_MAX_MIN} min) | "
            f"map_cross={len(map_cross)} ({CROSSSELL_MIN_MIN}-{CARRITO_MAX_MIN} min)"
        )

    for telefono in todos:
        # No molestar si el cliente cerró la conversación explícitamente
        if await verificar_cierre_enviado(telefono):
            continue

        # Si la conversación fue escalada a un humano (modo_humano=1), NO enviar
        # seguimientos automáticos — el agente humano se encarga.
        if await get_modo_humano(telefono):
            logger.info(f"[seguimiento-carrito] {telefono} en modo humano — saltando")
            continue

        # Determinar items y estado según qué ventana aplica
        # Estado 3 requiere estar en map_min; estados 4-5 requieren estar en map_cross
        en_min   = telefono in map_min
        en_cross = telefono in map_cross
        items = map_min.get(telefono) or map_cross.get(telefono)
        if not items:
            continue

        total = _calcular_total_carrito(items)
        if total <= 0:
            continue

        # Resolver PEDIDO_MINIMO del agente (por config_value, no el global).
        # Esto fixa el bug donde el seguimiento decía "$50.000" aún cuando el
        # merchant configuró otro valor (ej. 25000) en el panel.
        pedido_minimo_local = await _resolver_pedido_minimo(telefono)
        minimo_fmt = f"{pedido_minimo_local:,}".replace(",", ".")

        # Verificar que el estado corresponde a la ventana de tiempo correcta
        # Estado 3 (< mínimo): solo si pasó CARRITO_MIN_MIN
        # Estados 4-5 (≥ mínimo): solo si pasó CROSSSELL_MIN_MIN
        if total < pedido_minimo_local and not en_min:
            logger.info(f"Carrito {telefono}: ${total:,} < mínimo ${pedido_minimo_local}, aún no han pasado {CARRITO_MIN_MIN} min — esperando")
            continue  # Aún no es tiempo para el aviso de pedido mínimo
        if total >= pedido_minimo_local and not en_cross:
            logger.info(f"Carrito {telefono}: ${total:,} ≥ mínimo ${pedido_minimo_local}, aún no han pasado {CROSSSELL_MIN_MIN} min — esperando")
            continue  # Aún no es tiempo para el cross-sell

        # Calcular estado actual para detectar cambios
        estado_actual = 3 if total < pedido_minimo_local else (4 if total < umbral_gratis else 5)

        # Si el estado cambió (ej. cliente agregó productos: 3→4), resetear cooldown
        # para notificar de inmediato el nuevo estado
        estado_anterior = _carrito_ultimo_estado.get(telefono)
        if estado_anterior != estado_actual:
            _carrito_unif_cooldown[telefono] = 0
            logger.info(
                f"Carrito {telefono}: estado cambió "
                f"{estado_anterior or '—'} → {estado_actual} "
                f"(${total:,}) — cooldown reseteado"
            )
            # CAPI: InitiateCheckout — cliente alcanzó el mínimo de pedido (3→4 ó 3→5)
            # Pasamos items con retailer_id para que Meta haga match con catálogo
            if estado_anterior == 3 and estado_actual in (4, 5):
                asyncio.create_task(capi_initiate_checkout(telefono, total, productos=items))

        # Cooldown: no re-notificar el mismo estado en CARRITO_COOLDOWN_MIN minutos
        elapsed_min = (ahora - _carrito_unif_cooldown.get(telefono, 0)) / 60
        if ahora - _carrito_unif_cooldown.get(telefono, 0) < CARRITO_UNIF_COOLDOWN_SEG:
            logger.info(
                f"Carrito {telefono}: estado {estado_actual} (${total:,}) — "
                f"cooldown activo ({elapsed_min:.1f}/{CARRITO_COOLDOWN_MIN} min) — saltando"
            )
            continue

        # ── Verificación VIVA del carrito antes de mandar ───────────────
        # El loop captura items en `clientes_con_carrito_abandonado` y luego
        # los procesa. Entre esos 2 momentos el cliente puede haber vaciado
        # con [[VACIAR_CARRITO]] o confirmado pedido. Re-leer en vivo evita
        # mandar mensajes obsoletos tipo "tienes $42k" cuando ya vació.
        items_vivos = await obtener_carrito_activo(telefono)
        if not items_vivos:
            logger.info(f"[carrito-loop] {telefono} carrito vacío en BD viva — cancelando seguimiento")
            continue
        items = items_vivos
        total = _calcular_total_carrito(items)
        if total <= 0:
            continue
        # Re-calcular estado actual con el total vivo (puede haber cambiado)
        estado_actual = 3 if total < pedido_minimo_local else (4 if total < umbral_gratis else 5)

        nombres_en_carrito = {p.get("producto", "").lower() for p in items}
        total_fmt = f"{total:,}".replace(",", ".")

        # ── Detectar flujo activo del cliente ───────────────────────────
        # Si algún item tiene retailer_id, el cliente está en flujo WhatsApp
        # catálogo nativo → seguimientos deben respetar ese flujo.
        flujo_wa = any(p.get("retailer_id") for p in items)
        costo_envio_loc = obtener_costo_envio()

        # Si no es flujo_wa pero llegó hasta acá → no tiene retailer_ids,
        # probablemente carrito antiguo o de origen web. NO mandamos nada:
        # mejor silencioso que mandar link a tienda y romper la coherencia.
        if not flujo_wa:
            logger.info(f"[carrito-loop] {telefono} sin retailer_ids (no nativo) — cancelando")
            continue

        # Provider del agente real del cliente — NUNCA el singleton global
        # `proveedor` (solo tiene credenciales del agente por defecto).
        # Usarlo enviaría el seguimiento desde el número de WhatsApp
        # equivocado para cualquier otro agente (Meta acepta el envío,
        # pero el cliente nunca lo recibe en su chat real).
        _aid_seguimiento = await _resolver_agent_id_principal(telefono)
        proveedor = await _get_meta_para_agente({"id": _aid_seguimiento})

        async def _crear_checkout_para_seguimiento() -> str | None:
            """Crea un checkout de Shopify a partir de los items del carrito
            y guarda la URL en el cliente. Retorna la URL o None si falla."""
            if not flujo_wa:
                return None
            try:
                productos_check = []
                for p in items:
                    qty = int(p.get("cantidad", p.get("quantity", 1)))
                    productos_check.append({
                        "producto":      p.get("producto", ""),
                        "presentacion":  p.get("presentacion", ""),
                        "cantidad":      qty,
                        "precio_unitario": p.get("precio_unitario", 0),
                        "subtotal":      p.get("subtotal", 0),
                    })
                checkout_url_seg = await crear_checkout_shopify(
                    telefono, {"productos": productos_check, "total": total}
                )
                if checkout_url_seg:
                    await guardar_checkout_url(telefono, checkout_url_seg)
                    logger.info(f"[seguimiento] Checkout creado para {telefono}: {checkout_url_seg}")
                return checkout_url_seg
            except Exception as e:
                logger.error(f"[seguimiento] No pude crear checkout: {e}")
                return None

        try:
            if total < pedido_minimo_local:
                # ── Estado 3: bajo el mínimo ──────────────────────────────
                # Texto configurable por agente (puede apagarse si el negocio
                # no maneja pedido mínimo). Si está desactivado, no enviamos
                # nada — el cliente no es interrumpido.
                falta = pedido_minimo_local - total
                falta_fmt = f"{falta:,}".replace(",", ".")
                sugeridos = _sugerir_productos(nombres_en_carrito, _aid_seguimiento)
                lineas = "\n".join(f"✅ {s}" for s in sugeridos)
                from agent.mensajes import format_seguro as _fmt_est3
                _aid_est3 = await _resolver_agent_id_principal(telefono)
                _ctx_est3 = await construir_contexto_placeholders(_aid_est3)
                _ctx_est3["total"]  = total_fmt
                _ctx_est3["falta"]  = falta_fmt
                _ctx_est3["minimo"] = minimo_fmt
                msg = _fmt_est3(
                    await obtener_mensaje(_aid_est3, "cart.estado3_falta_minimo"),
                    _ctx_est3,
                )
                if not msg:
                    logger.info(f"[carrito-est3] {telefono} desactivado para agent={_aid_est3} — saltando")
                    continue
                if lineas:
                    msg += f"\n\nMuchos clientes agregan:\n{lineas}"

                enviado_seg = False
                if hasattr(proveedor, "enviar_catalogo_productos"):
                    try:
                        from agent.tools import obtener_secciones_catalogo_async
                        secciones = await obtener_secciones_catalogo_async(None)
                        if secciones:
                            # TODO multi-tenant: hoy seguimiento automatico asume
                            # agent_id=1 (Equora). En Sprint B agregar agent_id a
                            # clientes_con_carrito_abandonado y propagar.
                            header_cat = await _catalogo_header_for_agent(1)
                            enviado_seg = await proveedor.enviar_catalogo_productos(
                                telefono, header_cat, msg[:1024], secciones,
                            )
                    except Exception as e:
                        logger.warning(f"[carrito-est3] product_list falló: {e}")
                if not enviado_seg:
                    # Silencio: NO mandar link a tienda web (rompe el flujo
                    # del catálogo nativo). El cliente puede seguir interactuando
                    # con Andrea por iniciativa propia.
                    logger.info(f"[carrito-est3] {telefono} product_list no disponible — cancelando silencioso")
                    continue
                # Persistir el mensaje del seguimiento (panel lo necesita).
                await _guardar_con_wamid(proveedor, telefono, msg)

                # Mensaje complementario con botones Ver/Vaciar carrito —
                # SIEMPRE accesibles para el cliente (acciones determinísticas).
                # (En Estado 3 no incluimos "Confirmar pedido" porque aún no
                # alcanza el mínimo.)
                if hasattr(proveedor, "enviar_botones_reply"):
                    try:
                        texto_botones = "¿Quieres gestionar tu carrito?"
                        await proveedor.enviar_botones_reply(
                            telefono, texto_botones,
                            _botones_carrito_estandar(ofrece_confirmar=False),
                        )
                    except Exception as e:
                        logger.warning(f"[carrito-est3] botones complementarios fallaron: {e}")

            elif total < umbral_gratis:
                # ── Estado 4 (timer seguimiento): sobre el mínimo, recordar ──
                # Mensaje configurable por agente con placeholders dinámicos.
                # Eliminamos las 3 ramas condicionales por ciudad — el cliente
                # personaliza su template según su modelo de negocio (zonas,
                # envío gratis, transportadora, etc.). El default es neutral
                # equivalente al caso "sin ciudad confirmada" del código previo.
                falta = umbral_gratis - total
                falta_fmt = f"{falta:,}".replace(",", ".")
                sugeridos = _sugerir_productos(nombres_en_carrito, _aid_seguimiento)
                lineas = "\n".join(f"✅ {s}" for s in sugeridos)
                from agent.mensajes import format_seguro as _fmt_est4t
                _aid_est4t = await _resolver_agent_id_principal(telefono)
                _ctx_est4t = await construir_contexto_placeholders(_aid_est4t)
                _ctx_est4t["total"] = total_fmt
                _ctx_est4t["falta_envio_gratis"] = falta_fmt
                msg = _fmt_est4t(
                    await obtener_mensaje(_aid_est4t, "cart.estado4_timer"),
                    _ctx_est4t,
                )
                if not msg:
                    logger.info(f"[carrito-est4-timer] {telefono} desactivado para agent={_aid_est4t} — saltando")
                    continue
                if lineas:
                    msg += f"\n\nMuchos clientes también llevan:\n{lineas}"

                # Solo flujo nativo (validado al inicio del loop con flujo_wa).
                # Estrategia con carrito determinístico:
                #   1) Crear checkout en Shopify para tener URL de pago lista.
                #   2) Mandar 3 reply buttons SIEMPRE iguales: Confirmar + Ver + Vaciar.
                #      (Quitamos "Envío gratis" porque "Ver carrito" da info equivalente
                #       y queremos que SIEMPRE el cliente tenga botones para operar el carrito.)
                #   3) Si fallan los botones → product_list (catálogo nativo).
                #   4) Si product_list también falla → silencio (no tienda web).
                checkout_url_seg = await _crear_checkout_para_seguimiento()
                botones_seg = False
                if checkout_url_seg and hasattr(proveedor, "enviar_botones_reply"):
                    try:
                        botones_seg = await proveedor.enviar_botones_reply(
                            telefono, msg,
                            _botones_carrito_estandar(ofrece_confirmar=True),
                        )
                    except Exception as e:
                        logger.warning(f"[carrito-est4] botones reply fallaron: {e}")
                if not botones_seg:
                    enviado_pl = False
                    if hasattr(proveedor, "enviar_catalogo_productos"):
                        try:
                            from agent.tools import obtener_secciones_catalogo_async
                            secciones = await obtener_secciones_catalogo_async(None)
                            if secciones:
                                header_cat = await _catalogo_header_for_agent(1)
                                enviado_pl = await proveedor.enviar_catalogo_productos(
                                    telefono, header_cat, msg[:1024], secciones,
                                )
                        except Exception as e:
                            logger.warning(f"[carrito-est4] product_list falló: {e}")
                    if not enviado_pl:
                        logger.info(f"[carrito-est4] {telefono} ningún canal nativo funcionó — cancelando silencioso")
                        continue
                await _guardar_con_wamid(proveedor, telefono, msg)

            else:
                # ── Estado 5 (timer de seguimiento): pedido alto, invitar a pagar ──
                # Mismo enfoque que el flujo en vivo: mensaje configurable por agente.
                # Eliminamos las 3 ramas condicionales por ciudad — el cliente personaliza
                # su template según su modelo (si quiere distinguir por zona, usa
                # placeholders en el texto; si no, mensaje único).
                from agent.mensajes import format_seguro
                aid_seg = await _resolver_agent_id_principal(telefono)
                _ctx_seg = await construir_contexto_placeholders(aid_seg)
                _ctx_seg["total"] = total_fmt
                msg = format_seguro(
                    await obtener_mensaje(aid_seg, "cart.checkout_listo_texto"),
                    _ctx_seg,
                )
                # Estado 5: cliente ya está sobre el envío gratis o ya alcanzó
                # el mínimo (no Cali). Crear checkout y mandarle el botón.
                # El checkout_url ES Shopify pago — flujo legítimo, no es "tienda web".
                checkout_url_seg = await _crear_checkout_para_seguimiento()
                if not checkout_url_seg:
                    # Sin checkout_url no podemos completar el pago — escalar
                    logger.warning(f"[carrito-est5] {telefono} no se pudo crear checkout — escalando")
                    await _escalar_meta_fallo(
                        telefono,
                        motivo_corto="no se pudo crear checkout",
                        contexto_extra=f"Cliente con carrito de ${total_fmt} listo para pagar. Shopify checkout falló.",
                    )
                    continue
                enviado_seg = False
                if hasattr(proveedor, "enviar_cta_url"):
                    try:
                        enviado_seg = await proveedor.enviar_cta_url(
                            telefono, msg, "Confirmar pedido", checkout_url_seg
                        )
                    except Exception as e:
                        logger.warning(f"[carrito-est5] cta_url falló: {e}")
                if not enviado_seg:
                    # Fallback texto con la URL del checkout (sigue siendo Shopify pago)
                    await proveedor.enviar_mensaje(telefono, f"{msg}\n\n👉 {checkout_url_seg}")

                # Persistir el mensaje del seguimiento (panel lo necesita).
                await _guardar_con_wamid(proveedor, telefono, msg)

                # Mensaje complementario con botones Ver/Vaciar — el cliente
                # SIEMPRE debe poder revisar/limpiar su carrito antes de pagar.
                if hasattr(proveedor, "enviar_botones_reply"):
                    try:
                        await proveedor.enviar_botones_reply(
                            telefono, "¿Quieres revisar antes de confirmar?",
                            _botones_carrito_estandar(ofrece_confirmar=False),
                        )
                    except Exception as e:
                        logger.warning(f"[carrito-est5] botones complementarios fallaron: {e}")

            _carrito_unif_cooldown[telefono] = ahora
            _carrito_ultimo_estado[telefono] = estado_actual
            logger.info(f"Seguimiento carrito estado {estado_actual} → {telefono} (${total_fmt})")

        except Exception as e:
            logger.error(f"Error en seguimiento carrito {telefono}: {e}")


# Keywords que indican que la conversación se cerró naturalmente. Se
# matchean SOLO contra los últimos 1-2 mensajes (no historial completo)
# para evitar falsos positivos de saludos viejos. Lowercase + sin tildes.
_CIERRE_USER_KW = (
    "gracias", "muchas gracias", "ok gracias", "perfecto", "listo",
    "vale", "ya recibí", "ya recibi", "todo bien", "nos vemos",
    "chao", "chau", "adiós", "adios", "hasta luego", "hasta pronto",
    "que tengas", "buen dia", "buena tarde", "buena noche",
)
_CIERRE_ASSISTANT_KW = (
    "que disfrutes", "estamos para servirte", "para servirte",
    "cualquier cosa", "vuelve cuando", "hasta pronto", "hasta luego",
    "que tengas", "feliz dia", "feliz día", "feliz tarde", "feliz noche",
    "un gusto", "un placer", "fue un gusto", "que tengas un",
)


def _texto_normalizado(s: str) -> str:
    """Lowercase + quita tildes para matching case/accent-insensitive."""
    import unicodedata as _u
    s = _u.normalize("NFKD", s or "")
    return "".join(c for c in s if not _u.combining(c)).lower()


async def _conversacion_cerrada_naturalmente(telefono: str, aid: int) -> bool:
    """True si los últimos mensajes indican una despedida natural — para
    NO disparar el follow-up genérico. Heurística defensiva: el backend no
    confía en que el LLM emita un marcador de cierre.

    Reglas:
      · El último mensaje es de assistant Y contiene keyword de despedida
      · El penúltimo es user con gracias/cierre Y el último es assistant
        con respuesta corta (acuse) — patrón "gracias / de nada"
    """
    try:
        hist = await obtener_historial(telefono, limite=4, agent_id=aid)
    except Exception:
        return False
    if not hist:
        return False
    ultimo = hist[-1]
    ultimo_rol = ultimo.get("role", "")
    ultimo_txt = _texto_normalizado(ultimo.get("content", ""))
    if ultimo_rol == "assistant" and any(kw in ultimo_txt for kw in _CIERRE_ASSISTANT_KW):
        return True
    # Patrón "user dice gracias → assistant responde" (sin keyword fuerte de despedida)
    if ultimo_rol == "assistant" and len(hist) >= 2:
        penultimo = hist[-2]
        if penultimo.get("role") == "user":
            pen_txt = _texto_normalizado(penultimo.get("content", ""))
            if any(kw in pen_txt for kw in _CIERRE_USER_KW):
                return True
    return False


async def _procesar_followups():
    """Estado 6: sin carrito, sin checkout. Follow-up genérico una sola vez."""
    telefonos = await conversaciones_para_followup(FOLLOWUP_MIN, FOLLOWUP_MAX_HORAS)
    for telefono in telefonos:
        # No enviar followups si la conversación está escalada a humano
        if await get_modo_humano(telefono):
            logger.info(f"[followup] {telefono} en modo humano — saltando")
            continue
        try:
            aid = await _resolver_agent_id_principal(telefono)
            # Si hay cita confirmada reciente, Calendly ya notificó al cliente —
            # no insistir con follow-up de ventas encima.
            if await tiene_appointment_confirmado(telefono, aid):
                logger.info(f"[followup] {telefono} tiene cita confirmada — saltando follow-up")
                await marcar_cierre_enviado(telefono)
                continue
            # Heurística defensiva: si la conversación cerró natural
            # (gracias / despedida) NO insistir con follow-up. Marcamos
            # como cerrado para que el timer no vuelva a evaluarlo.
            if await _conversacion_cerrada_naturalmente(telefono, aid):
                logger.info(f"[followup] {telefono} cerrada naturalmente — marcando y saltando")
                await marcar_cierre_enviado(telefono)
                continue
            texto = await obtener_mensaje(aid, "system.followup")
            if not texto:
                # El agente desactivó este mensaje desde el panel — no insistir
                logger.info(f"[followup] {telefono} mensaje desactivado para agent={aid} — saltando")
                continue
            _prov = await _get_meta_para_agente({"id": aid})
            await _prov.enviar_mensaje(telefono, texto)
            await _guardar_con_wamid(_prov, telefono, texto, agent_id=aid)
            await marcar_followup_enviado(telefono)
            logger.info(f"Follow-up enviado a {telefono}")
        except Exception as e:
            logger.error(f"Error enviando follow-up a {telefono}: {e}")


async def _procesar_cierres():
    """Estado 6: cierre tras follow-up genérico sin respuesta."""
    telefonos = await conversaciones_para_cierre(CIERRE_MIN)
    for telefono in telefonos:
        # No enviar cierres si la conversación está escalada a humano
        if await get_modo_humano(telefono):
            logger.info(f"[cierre] {telefono} en modo humano — saltando")
            continue
        try:
            aid = await _resolver_agent_id_principal(telefono)
            # Si hay cita confirmada reciente, no enviar cierre de ventas
            if await tiene_appointment_confirmado(telefono, aid):
                logger.info(f"[cierre] {telefono} tiene cita confirmada — saltando cierre")
                await marcar_cierre_enviado(telefono)
                continue
            texto = await obtener_mensaje(aid, "system.cierre")
            if not texto:
                logger.info(f"[cierre] {telefono} mensaje desactivado para agent={aid} — saltando")
                # Aún así marcar como cerrado para que el timer no vuelva a evaluarlo
                await marcar_cierre_enviado(telefono)
                continue
            _prov = await _get_meta_para_agente({"id": aid})
            await _prov.enviar_mensaje(telefono, texto)
            await _guardar_con_wamid(_prov, telefono, texto, agent_id=aid)
            await marcar_cierre_enviado(telefono)
            logger.info(f"Cierre enviado a {telefono}")
        except Exception as e:
            logger.error(f"Error enviando cierre a {telefono}: {e}")


async def _procesar_abandono_carrito_UNUSED():
    """REEMPLAZADO por _procesar_carrito_unificado — se deja comentado como referencia."""
    ahora = time.time()
    clientes = await clientes_con_carrito_abandonado(
        min_inactivo=ABANDONO_MIN_INACTIVO,
        max_inactivo=ABANDONO_MAX_INACTIVO,
    )
    for telefono, _ in clientes:
        ultimo = _abandono_notif.get(telefono, 0)
        if ahora - ultimo < ABANDONO_COOLDOWN_SEG:
            continue
        try:
            tienda_url = f"{EQUORA_BASE}/catalogo"
            enviado = False
            msg_abandono = _mensaje_abandono()
            if hasattr(proveedor, "enviar_cta_url"):
                try:
                    enviado = await proveedor.enviar_cta_url(
                        telefono, msg_abandono, "Ver productos 🛒", tienda_url
                    )
                except Exception:
                    pass
            if not enviado:
                texto_con_url = f"{msg_abandono}\n\n👉 {tienda_url}"
                await proveedor.enviar_mensaje(telefono, texto_con_url)
            await guardar_mensaje(telefono, "assistant", msg_abandono)
            _abandono_notif[telefono] = ahora
            logger.info(f"Recuperación de carrito enviada a {telefono}")
        except Exception as e:
            logger.error(f"Error enviando recuperación de carrito a {telefono}: {e}")


async def _procesar_abandono_checkout():
    """
    Detecta clientes que iniciaron el checkout de Shopify (formulario de pago)
    pero no completaron el pedido. Reenvía el link de checkout para que terminen.
    Shopify limpia pedido_pendiente cuando dispara orders/create o orders/paid.
    """
    import urllib.parse
    clientes = await clientes_con_checkout_abandonado(
        min_min=CHECKOUT_ABANDONO_MIN,
        max_min=CHECKOUT_ABANDONO_MAX,
    )
    for telefono, checkout_url_raw in clientes:
        # Defensa: si el cliente ya no tiene carrito activo (vació explícito
        # o el TTL del carrito expiró), NO mandar recuperación de checkout
        # — es incoherente con su intención. Marcamos como enviado para que
        # el timer no reintente.
        carrito_vivo = await obtener_carrito_activo(telefono)
        if not carrito_vivo:
            logger.info(f"[checkout-abandono] {telefono} sin carrito activo — saltando y marcando")
            await limpiar_checkout_url(telefono)
            await marcar_checkout_abandono_enviado(telefono)
            continue

        # AUTOREPARAR si el URL guardado es corrupto (ej. solo la home).
        # Antes: saltábamos silencioso, lo que dejaba al cliente sin botón.
        # Ahora: recreamos checkout desde carrito_activo si está disponible.
        if not _es_url_checkout_valida(checkout_url_raw):
            logger.warning(
                f"[checkout-abandono] URL corrupta para {telefono} — intentando recrear"
            )
            checkout_url = await _obtener_o_recrear_checkout_url(telefono)
            if not checkout_url:
                logger.warning(f"[checkout-abandono] {telefono} sin checkout válido — saltando")
                continue
        else:
            checkout_url = checkout_url_raw
        # Cooldown PERSISTENTE (BD, no memoria) — sobrevive deploys de Railway
        if not await puede_enviar_checkout_abandono(telefono, CHECKOUT_COOLDOWN_MIN):
            continue
        try:
            aid = await _resolver_agent_id_principal(telefono)
            texto = await obtener_mensaje(aid, "system.checkout_abandono")
            if not texto:
                logger.info(f"[checkout-abandono] {telefono} mensaje desactivado para agent={aid} — saltando")
                # Marcar como enviado para evitar reintentos dentro del cooldown
                await marcar_checkout_abandono_enviado(telefono)
                continue
            _prov = await _get_meta_para_agente({"id": aid})
            enviado = False
            if hasattr(_prov, "enviar_cta_url"):
                try:
                    enviado = await _prov.enviar_cta_url(
                        telefono,
                        texto,
                        "Terminar pedido ✅",
                        checkout_url,
                    )
                except Exception:
                    pass
            if not enviado:
                await _prov.enviar_mensaje(
                    telefono,
                    f"{texto}\n\n👉 {checkout_url}"
                )
            await _guardar_con_wamid(_prov, telefono, texto, agent_id=aid)
            await marcar_checkout_abandono_enviado(telefono)
            logger.info(f"Recuperación de checkout enviada a {telefono}")
        except Exception as e:
            logger.error(f"Error enviando recuperación de checkout a {telefono}: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Helper: escalar cuando Meta product_list falla en un flujo que requiere respuesta
# ──────────────────────────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────
# Helper SaaS: header del catálogo dinámico por agente
# ──────────────────────────────────────────────────────────────────────────────
# Cache simple en memoria — el nombre del agente no cambia en runtime normal.
# Si cambia (rename desde panel), se refresca cuando el proceso se reinicia.
_catalogo_header_cache: dict[int, str] = {}

async def _catalogo_header_for_agent(agent_id: int) -> str:
    """Devuelve el texto que va en el header del product_list de WhatsApp.

    Prioridad:
      1. config_value 'CATALOGO_HEADER' del agente (si el cliente lo personalizó)
      2. Nombre del negocio del agente (Agent.name)
      3. Genérico "Catálogo 🌿"

    Esto reemplaza el "Catálogo Biotú 🌿" hardcoded que rompía SaaS multi-tenant.
    """
    if agent_id in _catalogo_header_cache:
        return _catalogo_header_cache[agent_id]
    try:
        # 1. Override explícito desde config_values
        custom = await get_config_value("CATALOGO_HEADER", agent_id)
        if custom and custom.strip():
            _catalogo_header_cache[agent_id] = custom.strip()[:60]
            return _catalogo_header_cache[agent_id]
        # 2. Nombre del negocio
        from agent.memory import obtener_agente
        agent = await obtener_agente(agent_id)
        if agent:
            nombre = (agent.get("name") or "").strip()
            if nombre:
                # "Catálogo {Nombre del negocio} 🌿" — corto a 60 chars (límite Meta)
                header = f"Catálogo {nombre} 🌿"[:60]
                _catalogo_header_cache[agent_id] = header
                return header
    except Exception as e:
        logger.warning(f"[catalogo-header] fallback genérico (error: {e})")
    # 3. Fallback genérico
    _catalogo_header_cache[agent_id] = "Catálogo 🌿"
    return "Catálogo 🌿"


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de carrito determinístico (backend-managed, no LLM)
# ──────────────────────────────────────────────────────────────────────────────
def _botones_carrito_estandar(ofrece_confirmar: bool = False) -> list[dict]:
    """Devuelve los reply buttons que SIEMPRE acompañan mensajes del sistema
    cuando hay carrito activo. Garantiza que el cliente NUNCA dependa del LLM
    para ver/vaciar — son acciones determinísticas del backend.

    Si ofrece_confirmar=True (Estados 4-5) incluye el botón Confirmar pedido
    como primero (mayor visibilidad para cerrar venta). Si False (Estado 3),
    solo Ver + Vaciar.

    Meta limita a 3 reply buttons por mensaje — respetamos ese tope.
    """
    botones = []
    if ofrece_confirmar:
        botones.append({"id": "act_confirmar_pedido", "title": "Confirmar pedido ✅"})
    botones.append({"id": "act_ver_carrito",   "title": "Ver carrito 🛒"})
    botones.append({"id": "act_vaciar_carrito", "title": "Vaciar carrito 🗑"})
    return botones


def _botones_post_pedido() -> list[dict]:
    """Botones estilo Jelou que se envían cuando el cliente confirma pedido
    y el checkout está listo. 3 opciones:
      💳 Ir al pago    → enviará el invoice_url del Draft Order
      🔍 Agregar más   → reabre el catálogo para sumar productos
      🛒 Ver mi carrito → reenvía resumen del carrito actual
    """
    return [
        {"id": "act_ir_pago",      "title": "💳 Ir al pago"},
        {"id": "act_agregar_mas",  "title": "🔍 Agregar más"},
        {"id": "act_ver_carrito",  "title": "🛒 Ver mi carrito"},
    ]


async def _enviar_resumen_carrito(telefono: str, agent_id: int = 1) -> bool:
    """Lee el carrito_activo de BD y envía un resumen determinístico al cliente.
    Se llama desde el handler act_ver_carrito. NUNCA delega al LLM — el backend
    es el único source of truth del estado del carrito.

    Returns: True si envió, False si carrito vacío (manda mensaje aparte).
    """
    # Provider del agente correcto — NUNCA el singleton global `proveedor`,
    # que solo tiene las credenciales del agente por defecto (id=1). Usarlo
    # acá enviaría desde el número de WhatsApp equivocado para cualquier
    # otro agente: Meta acepta el envío (200 OK, wamid válido → chulito ✓
    # en el panel) pero el cliente nunca lo recibe en su chat real.
    proveedor = await _get_meta_para_agente({"id": agent_id})
    carrito = await obtener_carrito_activo(telefono, agent_id=agent_id)
    if not carrito:
        msg_vacio = "🛒 Tu carrito está vacío en este momento.\n\nCuando agregues productos te lo confirmo aquí."
        await proveedor.enviar_mensaje(telefono, msg_vacio)
        await _guardar_con_wamid(proveedor, telefono, msg_vacio, agent_id=agent_id)
        return False

    lineas_items = []
    total = 0
    for item in carrito:
        cantidad = int(item.get("cantidad", item.get("quantity", 1)))
        producto = item.get("producto", "Producto")
        presentacion = item.get("presentacion", "")
        subtotal = int(item.get("subtotal", 0))
        precio_u = int(item.get("precio_unitario", 0))
        total += subtotal
        if presentacion:
            lineas_items.append(f"  • {cantidad}x {producto} / {presentacion} → ${subtotal:,}".replace(",", "."))
        else:
            lineas_items.append(f"  • {cantidad}x {producto} → ${subtotal:,}".replace(",", "."))

    total_fmt = f"{total:,}".replace(",", ".")
    resumen = (
        "🛒 *Tu carrito actual:*\n\n"
        + "\n".join(lineas_items)
        + f"\n\n💰 *Total: ${total_fmt}*"
    )
    # Reply buttons (Confirmar si supera el mínimo, sino solo Ver/Vaciar)
    pedido_min = 0
    try:
        pedido_min_str = await get_config_value("PEDIDO_MINIMO", agent_id) or os.getenv("PEDIDO_MINIMO", "0")
        pedido_min = int(float(pedido_min_str)) if pedido_min_str else 0
    except Exception:
        pedido_min = 0
    ofrece_confirmar = total >= pedido_min if pedido_min > 0 else True

    enviado = False
    if hasattr(proveedor, "enviar_botones_reply"):
        try:
            # Para "Ver carrito" mostramos solo 2 botones (sin duplicar "Ver" que ya están viendo)
            botones = []
            if ofrece_confirmar:
                botones.append({"id": "act_confirmar_pedido", "title": "Confirmar pedido ✅"})
            botones.append({"id": "act_vaciar_carrito", "title": "Vaciar carrito 🗑"})
            enviado = await proveedor.enviar_botones_reply(telefono, resumen, botones)
        except Exception as e:
            logger.warning(f"[ver-carrito] botones fallaron: {e}")
    if not enviado:
        await proveedor.enviar_mensaje(telefono, resumen)
    await _guardar_con_wamid(proveedor, telefono, resumen, agent_id=agent_id)
    return True


async def _enviar_horarios_calendly(telefono: str, agent_id: int = 1) -> None:
    """Obtiene horarios reales de Calendly y los envía como lista interactiva de WA.

    Guarda el mapeo título→ISO en calendly_pendiente para que la intercepción
    del webhook pueda recuperarlo cuando el cliente seleccione un horario.
    """
    from zoneinfo import ZoneInfo
    _DIAS_ES = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    _MESES_ES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
    _TZ_BOGOTA = ZoneInfo("America/Bogota")

    proveedor = await _get_meta_para_agente({"id": agent_id})

    async def _fallback(msg: str) -> None:
        await proveedor.enviar_mensaje(telefono, msg)
        await _guardar_con_wamid(proveedor, telefono, msg, agent_id=agent_id)

    # 1) Obtener user_uri de Calendly (valida el token también)
    usuario = await calendly_usuario(agent_id)
    if not usuario.get("ok"):
        await _fallback(
            "Lo siento, no pude conectarme a Calendly en este momento 😔\n"
            "Por favor intenta más tarde o contáctanos directamente."
        )
        return

    user_uri = usuario["uri"]

    # 2) Obtener tipos de evento activos
    event_types = await calendly_event_types(agent_id, user_uri)
    if not event_types:
        await _fallback(
            "No encontré tipos de cita configurados en Calendly. "
            "Por favor contáctanos para coordinar tu cita."
        )
        return

    # Usar el primer tipo de evento activo
    et = event_types[0]
    event_type_uri = et["uri"]
    event_type_nombre = et["nombre"]
    duracion_min = et.get("duracion_min", 0)
    location_kinds = et.get("location_kinds", [])
    location_kind = location_kinds[0] if location_kinds else None

    # 3) Obtener horarios disponibles (próximos 7 días)
    slots = await calendly_horarios(agent_id, event_type_uri, dias=7)
    if not slots:
        await _fallback(
            "No encontré horarios disponibles para los próximos días 😔\n"
            "Por favor escríbenos para coordinar tu cita directamente."
        )
        return

    # 4) Convertir UTC → Bogota y construir filas de lista (máx 8 slots)
    opciones: dict[str, str] = {}
    filas = []
    for slot in slots[:8]:
        iso = slot.get("start_time", "")
        if not iso:
            continue
        try:
            dt_utc = datetime.fromisoformat(iso.replace("Z", "+00:00"))
            dt = dt_utc.astimezone(_TZ_BOGOTA)
        except Exception:
            continue

        dia_abr  = _DIAS_ES[dt.weekday()]
        mes_abr  = _MESES_ES[dt.month - 1]
        hora_str = dt.strftime("%I:%M %p").lstrip("0")
        titulo = f"{dia_abr} {dt.day} {mes_abr}, {hora_str}"[:24]

        desc_parts = []
        if duracion_min:
            desc_parts.append(f"{duracion_min} min")
        if location_kind == "google_conference":
            desc_parts.append("Google Meet")
        elif location_kind == "physical":
            desc_parts.append("Presencial")
        elif location_kind:
            desc_parts.append(location_kind.replace("_", " ").title())
        descripcion = (" · ".join(desc_parts))[:72]

        if titulo not in opciones:
            _slot_id = f"slot_{len(filas)}"
            opciones[titulo] = iso
            filas.append({"id": _slot_id, "titulo": titulo, "descripcion": descripcion})

    if not filas:
        await _fallback("No pude preparar la lista de horarios. Por favor intenta más tarde.")
        return

    # id_map: slot_id → ISO (match por ID, más robusto que match por título)
    id_map = {f["id"]: opciones[f["titulo"]] for f in filas}
    titulos_map = {f["id"]: f["titulo"] for f in filas}

    # 5) Guardar estado en BD para la intercepción del webhook
    await guardar_calendly_pendiente(telefono, {
        "paso": "esperando_seleccion",
        "event_type_uri": event_type_uri,
        "event_type_nombre": event_type_nombre,
        "location_kind": location_kind,
        "opciones": opciones,
        "id_map": id_map,
        "titulos_map": titulos_map,
        "seleccion": None,
        "seleccion_titulo": None,
        "appointment_id": None,
        "intentos": 0,
    }, agent_id)

    # 6) Enviar lista interactiva
    texto_intro = (
        f"📅 *Horarios disponibles — {event_type_nombre}*\n\n"
        "Toca el horario que prefieras:"
    )
    secciones = [{"titulo": "Horarios disponibles", "items": filas}]
    enviado = False
    if hasattr(proveedor, "enviar_lista"):
        try:
            enviado = await proveedor.enviar_lista(telefono, texto_intro, "Ver horarios", secciones,
                                                header_text="📅 Elige tu horario")
        except Exception as e:
            logger.warning(f"[calendly] enviar_lista falló: {e}")
    if not enviado:
        # Fallback texto plano si el proveedor no soporta lista
        opciones_txt = "\n".join(f"• {t}" for t in opciones)
        fallback_txt = f"{texto_intro}\n\n{opciones_txt}\n\nResponde con el horario que prefieras."
        await proveedor.enviar_mensaje(telefono, fallback_txt)
        await _guardar_con_wamid(proveedor, telefono, fallback_txt, agent_id=agent_id)
    else:
        await _guardar_con_wamid(proveedor, telefono, texto_intro, agent_id=agent_id)


# ──────────────────────────────────────────────────────────────────────────────
# Helper: obtener o recrear checkout URL válido
# ──────────────────────────────────────────────────────────────────────────────
def _es_url_checkout_valida(url: str) -> bool:
    """True si la URL es un checkout real de Shopify (no la home).

    Acepta los formatos actuales del Storefront API, Admin API y webhooks:
      - '/checkouts/cn/...' y '/checkouts/...'  (webhook checkouts/create, formato clásico)
      - '/checkout/...'                          (alias usado por algunos endpoints)
      - '/cart/c/{id}?key=...&_s=...&_y=...'    (Storefront cartCreate — legacy)
      - '/invoices/...'                          (Admin Draft Order invoice_url — actual)
    """
    if not url:
        return False
    return (("/checkouts/" in url) or ("/checkout/" in url)
            or ("/cart/c/" in url) or ("/invoices/" in url))


async def _obtener_o_recrear_checkout_url(
    telefono: str, agent_id: int = 1, forzar_recrear: bool = False,
) -> str | None:
    """Devuelve un checkout_url VÁLIDO para el cliente.

    - Si forzar_recrear=True: SIEMPRE recrea desde carrito_activo actual.
      Úsalo cuando el cliente toca "Confirmar pedido" — el carrito puede
      haber cambiado desde el último checkout creado (nuevos items, etc.).
    - Si forzar_recrear=False (default): usa el URL guardado si tiene
      /checkouts/. Solo recrea si está corrupto. Úsalo en timers como el
      recordatorio de checkout abandonado.

    Returns: URL válida del checkout, o None si no se pudo (sin carrito).
    """
    cliente_data = await obtener_cliente(telefono, agent_id=agent_id)
    if not cliente_data:
        return None
    url_guardado = (cliente_data.get("pedido_checkout_url") or "").strip()

    # Si NO forzamos y el URL guardado es válido, usarlo (caso normal del timer)
    if not forzar_recrear and url_guardado and _es_url_checkout_valida(url_guardado):
        return url_guardado

    # Recrear desde carrito_activo actual:
    #   - forzar_recrear=True (cliente activamente confirmando)
    #   - O URL corrupta/vacía
    if url_guardado and not forzar_recrear:
        logger.warning(
            f"[checkout-url] URL corrupta para {telefono} ({url_guardado[:60]}) — recreando"
        )
    elif forzar_recrear:
        logger.info(
            f"[checkout-url] Recreando checkout para {telefono} (forzado por confirmación de pedido)"
        )
    carrito = await obtener_carrito_activo(telefono, agent_id=agent_id)
    if not carrito:
        logger.warning(f"[checkout-url] {telefono} sin carrito_activo — no puedo recrear checkout")
        return None
    logger.info(f"[checkout-url] carrito_activo de {telefono} tiene {len(carrito)} items")
    # Convertir items del carrito al formato esperado por crear_checkout_shopify
    productos = []
    total = 0
    descartados = []
    for item in carrito:
        producto = item.get("producto", "")
        presentacion = item.get("presentacion", "")
        cantidad = int(item.get("cantidad", item.get("quantity", 1)))
        precio_u = int(item.get("precio_unitario", 0))
        subtotal = int(item.get("subtotal") or precio_u * cantidad)
        if not producto or not precio_u:
            descartados.append(f"{producto or '?'}/{presentacion or '?'} (precio={precio_u})")
            continue
        productos.append({
            "producto": producto,
            "presentacion": presentacion,
            "cantidad": cantidad,
            "precio_unitario": precio_u,
            "subtotal": subtotal,
        })
        total += subtotal
    if descartados:
        logger.warning(f"[checkout-url] items descartados de {telefono}: {descartados}")
    logger.info(f"[checkout-url] {telefono} → {len(productos)} productos válidos, total ${total:,}")
    if not productos:
        logger.warning(f"[checkout-url] carrito_activo de {telefono} sin productos válidos")
        # Fallback: si tenemos URL guardada válida, usarla mejor que None
        if url_guardado and _es_url_checkout_valida(url_guardado):
            logger.info(f"[checkout-url] usando URL guardada como fallback")
            return url_guardado
        return None
    try:
        nuevo_url = await crear_checkout_shopify(telefono, {"productos": productos, "total": total})
        if nuevo_url and _es_url_checkout_valida(nuevo_url):
            await guardar_checkout_url(telefono, nuevo_url, agent_id=agent_id)
            logger.info(f"[checkout-url] Recreado para {telefono}: {nuevo_url[:80]}")
            return nuevo_url
        # ── Recreación falló — usar URL guardada como FALLBACK ──
        # Esto evita que el cliente quede sin botón cuando crear_checkout_shopify
        # no encuentra alguna variante (ej. "Default Title" que no matchea).
        # Mejor un checkout que ya existía con casi todos los items que NINGUNO.
        logger.error(
            f"[checkout-url] crear_checkout_shopify devolvió URL inválida: {nuevo_url} — "
            f"usando URL guardada como fallback (si existe)"
        )
        if url_guardado and _es_url_checkout_valida(url_guardado):
            logger.info(f"[checkout-url] fallback OK: {url_guardado[:80]}")
            return url_guardado
        return None
    except Exception as e:
        logger.error(f"[checkout-url] Error recreando checkout para {telefono}: {e}")
        # Mismo fallback ante excepción
        if url_guardado and _es_url_checkout_valida(url_guardado):
            logger.info(f"[checkout-url] fallback tras excepción: {url_guardado[:80]}")
            return url_guardado
        return None


async def _escalar_meta_fallo(
    telefono: str,
    motivo_corto: str,
    contexto_extra: str = "",
    agent_id: int = 1,
) -> None:
    """Escala una conversación a humano cuando el catálogo nativo de WhatsApp
    falla y el cliente está esperando una respuesta concreta AHORA.

    Hace 3 cosas:
      1. Manda mensaje cálido al cliente: 'Te conecto con un asesor...'
      2. Crea ticket via _notificar_escalacion (que también pausa el bot
         y avisa al equipo por WhatsApp + panel).
      3. Persiste el mensaje en BD para que aparezca en panel.

    Solo se usa en flujos donde NO mandar nada sería pésima UX:
      - Cliente terminó pedido nativo bajo mínimo → product_list falló
      - Andrea emitió [[TIENDA:]] → product_list falló
    NO se usa en seguimientos automáticos (timers) — esos cancelan silencioso.
    """
    try:
        # 1. Mensaje cálido al cliente
        msg_cliente = (
            "Para ayudarte mejor te conecto con un asesor "
            "que te va a mostrar los productos. Ya viene 🌿"
        )
        _prov = await _get_meta_para_agente({"id": agent_id})
        await _prov.enviar_mensaje(telefono, msg_cliente)
        await _guardar_con_wamid(_prov, telefono, msg_cliente, agent_id=agent_id)

        # 2. Obtener nombre del cliente para el ticket (mejora UX panel)
        try:
            cli = await obtener_cliente(telefono, agent_id=agent_id)
        except Exception:
            cli = None
        nombre = (cli.get("nombre") if cli else "") or ""

        # 3. Crear ticket + pausar bot + notificar equipo
        contexto_full = (
            f"Cliente esperando catálogo. {contexto_extra}".strip()
            if contexto_extra else
            "Cliente esperando catálogo de productos."
        )
        await _notificar_escalacion(
            telefono,
            {
                "motivo":         f"Catálogo no disponible — {motivo_corto}",
                "urgencia":       "normal",
                "nombre_cliente": nombre,
                "contexto":       contexto_full,
            },
            agent_id=agent_id,
        )
        logger.info(f"[escalacion-meta] {telefono} escalado por: {motivo_corto}")
    except Exception as e:
        logger.error(f"[escalacion-meta] fallo escalando {telefono}: {e}")


app = FastAPI(
    title="Voco — Plataforma multi-tenant de agentes WhatsApp",
    version="1.0.0",
    lifespan=lifespan
)

# CORS — orígenes permitidos para llamadas externas (mini-tienda Lovable,
# checkout custom, etc.). Multi-tenant: cada merchant configura sus dominios
# en CORS_ALLOWED_ORIGINS (env var, separados por coma). Fallback a Equora
# legacy mientras hay un solo tenant en producción.
_cors_env = os.getenv("CORS_ALLOWED_ORIGINS", "").strip()
_cors_origins = [o.strip() for o in _cors_env.split(",") if o.strip()] or [
    "https://equoradistribuciones.com",
    "https://www.equoradistribuciones.com",
    "https://equora-6.myshopify.com",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)


@app.get("/health")
async def health_check():
    return {"status": "ok", "service": "voco"}


@app.get("/", response_class=HTMLResponse)
async def landing_page():
    """Landing page pública de Voco — requerida por Meta Tech Provider verification."""
    html = """<!doctype html>
<html lang="es"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="description" content="Voco — Plataforma SaaS que automatiza la atención al cliente, ventas y seguimiento de pedidos por WhatsApp usando inteligencia artificial.">
<title>Voco · Automatización de WhatsApp con IA</title>
<style>
:root{color-scheme:light}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;
  background:#f6f7f9;color:#1f2430;line-height:1.6;-webkit-font-smoothing:antialiased}
a{color:#16a34a;text-decoration:none}
a:hover{text-decoration:underline}

/* NAV */
nav{background:#fff;border-bottom:1px solid #e5e7eb;padding:0 32px;
  display:flex;align-items:center;justify-content:space-between;height:60px}
.logo{display:flex;align-items:center;gap:10px;font-weight:700;font-size:1.1rem;color:#1f2430}
.logo-dot{width:28px;height:28px;border-radius:8px;background:#16a34a;
  display:flex;align-items:center;justify-content:center;color:#fff;font-weight:800;font-size:.95rem}
.nav-links{display:flex;gap:28px;font-size:.9rem}

/* HERO */
.hero{text-align:center;padding:96px 24px 72px;max-width:720px;margin:0 auto}
.badge{display:inline-block;background:#dcfce7;color:#15803d;font-size:.78rem;
  font-weight:600;padding:4px 12px;border-radius:20px;margin-bottom:20px;letter-spacing:.02em}
h1{font-size:2.6rem;font-weight:800;letter-spacing:-.03em;line-height:1.15;margin-bottom:18px}
h1 span{color:#16a34a}
.sub{font-size:1.1rem;color:#4b5563;max-width:540px;margin:0 auto 36px}
.cta{display:inline-block;background:#16a34a;color:#fff;font-weight:600;
  padding:13px 30px;border-radius:10px;font-size:1rem;transition:background .15s}
.cta:hover{background:#15803d;text-decoration:none}

/* FEATURES */
.features{max-width:960px;margin:0 auto;padding:0 24px 80px;
  display:grid;grid-template-columns:repeat(auto-fill,minmax(270px,1fr));gap:20px}
.feat{background:#fff;border:1px solid #e5e7eb;border-radius:14px;padding:28px;
  box-shadow:0 1px 3px rgba(0,0,0,.04)}
.feat-icon{font-size:1.8rem;margin-bottom:12px}
.feat h3{font-size:1rem;font-weight:700;margin-bottom:6px}
.feat p{font-size:.9rem;color:#6b7280}

/* HOW */
.how{background:#fff;border-top:1px solid #e5e7eb;border-bottom:1px solid #e5e7eb;
  padding:64px 24px;text-align:center}
.how h2{font-size:1.7rem;font-weight:800;letter-spacing:-.02em;margin-bottom:8px}
.how .sub{margin-bottom:48px}
.steps{display:flex;flex-wrap:wrap;justify-content:center;gap:32px;max-width:860px;margin:0 auto}
.step{flex:1;min-width:180px;max-width:220px}
.step-num{width:36px;height:36px;border-radius:50%;background:#16a34a;color:#fff;
  font-weight:700;font-size:.95rem;display:flex;align-items:center;justify-content:center;margin:0 auto 12px}
.step h4{font-weight:700;margin-bottom:4px;font-size:.95rem}
.step p{font-size:.85rem;color:#6b7280}

/* FOOTER */
footer{text-align:center;padding:40px 24px;font-size:.85rem;color:#9ca3af}
footer a{color:#9ca3af}
</style>
</head><body>

<nav>
  <div class="logo">
    <div class="logo-dot" style="background:transparent;padding:0"><img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAABOYAAATmCAYAAACF/K4qAABTmmNhQlgAAFOaanVtYgAAAB5qdW1kYzJwYQARABCAAACqADibcQNjMnBhAAAAU3RqdW1iAAAAR2p1bWRjMm1hABEAEIAAAKoAOJtxA3VybjpjMnBhOmU0NGJiODNmLTBkMjktNGM1MS04MGNjLTUyNzI3NTBjNDA4NwAAAAw9anVtYgAAAClqdW1kYzJhcwARABCAAACqADibcQNjMnBhLmFzc2VydGlvbnMAAAAGtGp1bWIAAABLanVtZEDLDDK7ikidpwsq1vR/Q2kTYzJwYS50aHVtYm5haWwuaW5ncmVkaWVudAAAAAAYYzJzaCVuD5q0tLtO7mxwjfZMdZAAAAAUYmZkYgBpbWFnZS9qcGVnAAAABk1iaWRi/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCABkAGQDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD5/ooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiinIjyOqRqzOxwFUZJoAbRXV6T4A1jUdr3CCyhPeb734L1/PFdzpPgTRtM2vJEbyYfxz8j8F6fnmk2jzcRmuGo6X5n5HjhVgASCAehx1pK+gp7SzvrdraaGGaEcGNlBA/DtXE6z8NYJd0ukT+S3XyZSSv4N1H45pKRz4fO6NR2qLl/FHmdFXtS0fUNIm8q+tZISTwxGVb6Hoao1R7EZRmuaLugooooKCiiigAooooAK9Q+GVnANJubzy1Nw05j3kchQqnA/M15fXq/wz/wCRbuP+vtv/AEBKUtjys5bWFdu6LXiTxvbaBcmzS2e4ugoYjO1Fz0yev6V5/qnjTW9U3K1ybeE/8s7f5R+J6n86sfEH/kbp/wDrnH/6DXLUJCy7A0I0YVOW7aT1LVlqN5ptx9os7mSGXuVPX6jv+Nd7ovxKU7YdYg29vPhHH4r/AIflXnFFDVzsxOCo4hfvFr36nv8AFLYa1YbkMN3ayjB4DKfYivDdXto7PWr62iBEcNw6ICewYgV6V8Mv+RbuP+vtv/QErzrxF/yMuqf9fcv/AKEaS3PKyqn7HE1aKeiM2iiiqPfCiiigAooooAK9X+Gf/It3H/X23/oCV5RW54e8U33h13WBUlgkOXifpn1B7Gk1dHDmOHniKDhDc6/xp4O1HVNSfUrApLuRQ0JO1sgY4zwa86urS4spjDdQSQyDqsikGvWdK+IGj3+1Lhmspj2l+7+DD+uK1vEX2D+wbq4voIp4YoiyhwDk9sHtk45FJNo8jD4/E4VxoVoeS7/5M8KoroNH8G6vrO2RIPItz/y2m+UEew6mu+0rwJo2koJ7vF3KoyXnwEX/AID0/PNNs9XE5nh6Gjd32RD8NY3Tw1MWUqHumZSR1G1RkfiDXnPiL/kZdU/6+5f/AEI16Vq/j/SdMUw2X+mTKMARcRr/AMC/wzXlF1cSXd3NcykGSZ2kfHqTk0kcuW06sq1TETjyqWxFRRRVHtBRRRQAUUUUAFFFFACqCzBR1JxX0FP9lhtM3RiEEYBJlxtGOh5r5+jcxyK4xlSCM+1XdT1rUdYm8y+unlwcqnRV+gHApNXPMzDAyxcoWdkr/oeiaz8R7G03RaZGbuUceY2VjH9T+n1rgNW8Q6prT5vbpmjzkRL8qD8B/WsuihKxthsvoYfWC17vcKKKKZ2hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQB//ZAAABqmp1bWIAAABEanVtZGNib3IAEQAQgAAAqgA4m3ETYzJwYS5pbmdyZWRpZW50LnYzAAAAABhjMnNo4fpTWdxhdAGYByRiH9mlfAAAAV5jYm9ypGxyZWxhdGlvbnNoaXBrY29tcG9uZW50T2ZoZGM6dGl0bGV4sWRvY3VtZW50LmFzaHg/cGF0aD0lMkYyZWZkNDY0Yy04ZDA3LTRjMjgtOWE4Yi01NDdiNTRkZmRkNTglMkZEb2N1bWVudENhY2hlJTJGZGVzaWduLWRiNWUzOTIxLTVmNjktNGEyMi05ZjIxLWQxOWJjYTZjMTc0YSUyRm1lZGlhJTJGYXNzZXQtMzE0ZWVjMzQtMGYyMi00NTE1LWE1Y2MtY2RjNTU1NTI0ZWM0LnBuZ2lkYzpmb3JtYXRpaW1hZ2UvcG5naXRodW1ibmFpbKJjdXJseDRzZWxmI2p1bWJmPWMycGEuYXNzZXJ0aW9ucy9jMnBhLnRodW1ibmFpbC5pbmdyZWRpZW50ZGhhc2hYIBQ56AQ/bVN+YKDpOWMFv2rpssWv8H9ElvBvEgKhDZmdAAAB1Gp1bWIAAABBanVtZGNib3IAEQAQgAAAqgA4m3ETYzJwYS5hY3Rpb25zLnYyAAAAABhjMnNo6kqeUtvUa1oM72i5QsbmdwAAAYtjYm9yomdhY3Rpb25zgqNmYWN0aW9ua2MycGEuZWRpdGVkcWRpZ2l0YWxTb3VyY2VUeXBleEZodHRwOi8vY3YuaXB0Yy5vcmcvbmV3c2NvZGVzL2RpZ2l0YWxzb3VyY2V0eXBlL3RyYWluZWRBbGdvcml0aG1pY01lZGlha2Rlc2NyaXB0aW9ueCVFZGl0ZWQgdGhlIGltYWdlIHRvIHJlbW92ZSBiYWNrZ3JvdW5kpGZhY3Rpb25wYzJwYS53YXRlcm1hcmtlZGR3aGVueCEyMDI2LTA2LTE3VDAwOjM0OjM4LjMwMzQ5ODYrMDA6MDBtc29mdHdhcmVBZ2VudL9kbmFtZXgjTWljcm9zb2Z0IFJlc3BvbnNpYmxlIEFJIFByb3ZlbmFuY2VndmVyc2lvbmMxLjD/a2Rlc2NyaXB0aW9ueC9Db250ZW50IHdhdGVybWFya2VkIGJ5IE1pY3Jvc29mdCBSZXNwb25zaWJsZSBBSXJhbGxBY3Rpb25zSW5jbHVkZWT1AAABF2p1bWIAAABDanVtZGNib3IAEQAQgAAAqgA4m3ETYzJwYS5zb2Z0LWJpbmRpbmcAAAAAGGMyc2jFB5/SZhbNekCjTef2tho2AAAAzGNib3KjY2FsZ3gZY29tLm1pY3Jvc29mdC5pbnZpc21hcmsuMWNwYWSAZmJsb2Nrc4GiZXNjb3BloWZyZWdpb26hZnJlZ2lvboGiZHR5cGVnc3BhdGlhbGVzaGFwZaVkdHlwZWlyZWN0YW5nbGVkdW5pdGpwZXJjZW50YWdlZXdpZHRoGGRmaGVpZ2h0GGRmb3JpZ2luomF4AGF5AGV2YWx1ZXgkNzg4MGZhNzktMzZkYS00MzFjLWE5MWMtODU4OTBiNGY1YjNiAAAAw2p1bWIAAABAanVtZGNib3IAEQAQgAAAqgA4m3ETYzJwYS5oYXNoLmRhdGEAAAAAGGMyc2he8yPR1Mwime4RtGgqyuLQAAAAe2Nib3KlamV4Y2x1c2lvbnOBomVzdGFydBghZmxlbmd0aBlTpmRuYW1lbmp1bWJmIG1hbmlmZXN0Y2FsZ2ZzaGEyNTZkaGFzaFggwV6ctXcmhKVMSK9CwsmkjRiRSKuzJBlMGcxjr794dRNjcGFkSAAAAAAAAAAAAAADw2p1bWIAAAAnanVtZGMyY2wAEQAQgAAAqgA4m3EDYzJwYS5jbGFpbS52MgAAAAOUY2JvcqdqaW5zdGFuY2VJRHgseG1wOmlpZDpjNmU2ZDZmNi1mNGYwLTQyNmMtYTU0ZC05MzFhMDAzNzE3Njh0Y2xhaW1fZ2VuZXJhdG9yX2luZm+/ZG5hbWVyTWljcm9zb2Z0X0Rlc2lnbmVyZ3ZlcnNpb25jMi4wcG9wZXJhdGluZ19zeXN0ZW14IU1pY3Jvc29mdCBXaW5kb3dzIE5UIDEwLjAuMjAzNDguMGVhcHBJZHgkNWUyNzk1ZTMtY2U4Yy00Y2ZiLWIzMDItMzVmZTVjZDAxNTk3d29yZy5jb250ZW50YXV0aC5jMnBhX3JzZjAuODQuMf9pc2lnbmF0dXJleE1zZWxmI2p1bWJmPS9jMnBhL3VybjpjMnBhOmU0NGJiODNmLTBkMjktNGM1MS04MGNjLTUyNzI3NTBjNDA4Ny9jMnBhLnNpZ25hdHVyZXJjcmVhdGVkX2Fzc2VydGlvbnOBomN1cmx4KXNlbGYjanVtYmY9YzJwYS5hc3NlcnRpb25zL2MycGEuaGFzaC5kYXRhZGhhc2hYIGyLru2+bU6pwOtrYhY8Nx5m4nwlyvC51iNakQuulxdHc2dhdGhlcmVkX2Fzc2VydGlvbnOEomN1cmx4NHNlbGYjanVtYmY9YzJwYS5hc3NlcnRpb25zL2MycGEudGh1bWJuYWlsLmluZ3JlZGllbnRkaGFzaFggFDnoBD9tU35goOk5YwW/aumyxa/wf0SW8G8SAqENmZ2iY3VybHgtc2VsZiNqdW1iZj1jMnBhLmFzc2VydGlvbnMvYzJwYS5pbmdyZWRpZW50LnYzZGhhc2hYIEqMdqhuNpqjQp2L8gPNqbtQVCIIuTTLGaWzhjCK84F3omN1cmx4KnNlbGYjanVtYmY9YzJwYS5hc3NlcnRpb25zL2MycGEuYWN0aW9ucy52MmRoYXNoWCAL2KQcr3AWGX4iRELFoD2rOHRhrp24iYw25/fC7Vi2dqJjdXJseCxzZWxmI2p1bWJmPWMycGEuYXNzZXJ0aW9ucy9jMnBhLnNvZnQtYmluZGluZ2RoYXNoWCD+BM+y7u7fC8UGF/eLXt1i3HwtbWebZ8STp3aPCTYYOGhkYzp0aXRsZXg6Q29udGVudCBmb3IgcGFnZSBwYWdlLWNkNWRjNGJiLTQyMTgtNDdiZi1iMzM2LWVlMDI5NGU0NmRmMWNhbGdmc2hhMjU2AABDJWp1bWIAAAAoanVtZGMyY3MAEQAQgAAAqgA4m3EDYzJwYS5zaWduYXR1cmUAAABC9WNib3LShFkSwqIBOCQYIYNZBikwggYlMIIEDaADAgECAhMzAAAAnKbmwp6lNcRhAAAAAACcMA0GCSqGSIb3DQEBDAUAMFYxCzAJBgNVBAYTAlVTMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xJzAlBgNVBAMTHk1pY3Jvc29mdCBTQ0QgQ2xhaW1hbnRzIFJTQSBDQTAeFw0yNTEwMDkxODI3NDVaFw0yNjEwMDkxODI3NDVaMHQxCzAJBgNVBAYTAlVTMRMwEQYDVQQIEwpXYXNoaW5ndG9uMRAwDgYDVQQHEwdSZWRtb25kMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xHjAcBgNVBAMTFU1pY3Jvc29mdCBDb3Jwb3JhdGlvbjCCAaIwDQYJKoZIhvcNAQEBBQADggGPADCCAYoCggGBAKEBvqn5VG2r9jCTdBdWq9lNbP797rWni2f9ruWq5s6sVQH7w1z/z81FjFVi5ZV56y8JRRhLAPzAEY49pVi+H5BX9Xva3sK2HL36UIuKwjQjfmxz0pKP8DBhPit8aNRSrCvP8MAmzePBFszkWFH1pabRTHs3y0Sz7/PA/SzJH8LUTRRDh9VR7pQLFo6xgn1nKc0aGlKuTIBzp6L8NgH3pB63twFYtf5Ysip5+KfAUYe8tczoSWvCFThMkqmpg1sp21gEQiOQJaFLT7f8EBtZ8w9FxXyZSZYagRiJczQkrNCHSdieSkOdv7gQem66Z5YXTQMbjyK9dSJuhXjnWyHgGCcipF1bzsBo0xFdswcV/VKK2/E6STicqSDU4WhtqwOD9pPEMoDtYq7JqzbacUo26ZAqDzkuPsEMf6zcu8G9vufG2Y4Ja8HdbQEkuywLJ8xIU8Hdk1euYiwoO+o85TK1JyIcdVmd2ot8YNyKHf73lWpEmlPdWtYA0xtH9S2ETZauwQIDAQABo4IBTDCCAUgwGQYDVR0lAQH/BA8wDQYLKwYBBAGCN0w7AQkwDgYDVR0PAQH/BAQDAgDAMB0GA1UdDgQWBBQWbsdVjGAql2Y3oSGeejY/BYr+5TAfBgNVHSMEGDAWgBSLrZr8j3XNzg2Naa18TKRgVtm0RDBfBgNVHR8EWDBWMFSgUqBQhk5odHRwOi8vd3d3Lm1pY3Jvc29mdC5jb20vcGtpb3BzL2NybC9NaWNyb3NvZnQlMjBTQ0QlMjBDbGFpbWFudHMlMjBSU0ElMjBDQS5jcmwwbAYIKwYBBQUHAQEEYDBeMFwGCCsGAQUFBzAChlBodHRwOi8vd3d3Lm1pY3Jvc29mdC5jb20vcGtpb3BzL2NlcnRzL01pY3Jvc29mdCUyMFNDRCUyMENsYWltYW50cyUyMFJTQSUyMENBLmNydDAMBgNVHRMBAf8EAjAAMA0GCSqGSIb3DQEBDAUAA4ICAQC9/qyfFvrrbdVRKneTFRXbjM3pO4YGUJS7HrGWNZAgRRBZsS3CfB733CZ1XhMIbuOYz4U3a+v6gXkt5o2ulDaK9SqV5npZxLC7cutz5e9KUtal/sEuUM8aKXPSQfbi0Y2ocaQLvpm3LrgF0epTLR5XQa/vHI6jZrtq2H9ZB+f3GaATLpNME9087k6n7RC9YQNhLmbDA6oi/cHefi75yqgxcsjGFj1rRpAmKWME0ZBvnuR7AIY8ie20xW9jgRcPK4k8QBYEqhKQe32mVNVxXAAB9YzTYP1zUC/661OYy6FXcpNqjMHXT1Gtd8wZTieOqOXx35XyuHu5OJn+1XFC1Gifpq/RHsC86cS01101yipWd4GPypg+53GbSaoJOFp/fjLVudX57RdF9NRF0eBAeorpIRc0BcotuvEmR8En8DP70kZRznkTK1kg6m9F6y3C7w4UnoL/ZOuRbZXi9S9vM7m4lp/0K9YeDgiSp1mCJ5DnRE8eFm7vSmUqmMZDGHt83lK7mh0+5Y0uVq3947YfIsUR9yJPV1VuoA3g+cQNVxvYXo17S8zzR8VL11pWT8R4JF23c+mP0h7P+Tx+y+TX9jXbEagDBmp7EXqIkj4P1kLl12ROQf2F2dhpsqaL1VaV42hNTIWI+I9OOfV6QxKdn/cZzvwc1Fcu98stvv33mFFMlVkG1jCCBtIwggS6oAMCAQICEzMAAAAE0dbhegoiYg8AAAAAAAQwDQYJKoZIhvcNAQEMBQAwXzELMAkGA1UEBhMCVVMxHjAcBgNVBAoTFU1pY3Jvc29mdCBDb3Jwb3JhdGlvbjEwMC4GA1UEAxMnTWljcm9zb2Z0IFN1cHBseSBDaGFpbiBSU0EgUm9vdCBDQSAyMDIyMB4XDTIyMDIxNzAwNDUyNloXDTQyMDIxNzAwNTUyNlowVjELMAkGA1UEBhMCVVMxHjAcBgNVBAoTFU1pY3Jvc29mdCBDb3Jwb3JhdGlvbjEnMCUGA1UEAxMeTWljcm9zb2Z0IFNDRCBDbGFpbWFudHMgUlNBIENBMIICIjANBgkqhkiG9w0BAQEFAAOCAg8AMIICCgKCAgEA1yXi2/OON2zaBMWmrfkpk9A1AV4RX6lGln0epO0gg+gBFneRAKtMN1Hvq6zTNJGp8ITCDoxuNXFCHSJ6C3W4Gh1QXgXHBLwCTYIq+iiaWaPx/FajWxYvnEPYeCSxmRzRhQCmf6xmkOJEs2fs3nFcGJfWdMPoqUzvweNdpa8oYH2YWiXW4nz/PUxAGhKhNSw1FTD5SEjI7wz6B8gOovwjMNC/kAdLvs4hk+R+3YWwVF1n+7zd+vYmUtPg8bexX16MMx5pzRuZZfcGYpwj+hFMlQ3QV94mTB2AmuupCkDCArsqZTdo5kX48tJFd5xlSQv7FL1dutgHYGdbfdeC8z3gLKwkEUIneTNmiHOsL/319uLY6K/jlaR8a6q2jIJbMVl0D7jrotcfB5jGnjCwf0zmh1XOIjK1S4pKBPcHGBm9FfZpwqQRWg9Evf6c6OrMfcaZd4NTtS9FlNJCMf0sXZzEPXqcRg7SXI8QoGRxzOejHZZnJTsm5Ng0DuHDjvJadA4/hXytnmewXfrf2VwtAtUYCBiit5FqLVcr9J1LQz1zrtL8E3Hf3JHjrUbgY4Cx1z4yTP+601xdgTex58DrTyBucp4kWsXzlL67Zjn4TjutXFr8pXDCzWmJx88E7G7S9rBcSDldIElhiQJW5r9hUEoFlytXFqIZy0IpLYhlgTVV9WkCAwEAAaOCAY4wggGKMA4GA1UdDwEB/wQEAwIBhjAQBgkrBgEEAYI3FQEEAwIBADAdBgNVHQ4EFgQUi62a/I91zc4NjWmtfEykYFbZtEQwEQYDVR0gBAowCDAGBgRVHSAAMBkGCSsGAQQBgjcUAgQMHgoAUwB1AGIAQwBBMA8GA1UdEwEB/wQFMAMBAf8wHwYDVR0jBBgwFoAUC7NoO6/ar+5wpXbZIffMRBYH0PgwbAYDVR0fBGUwYzBhoF+gXYZbaHR0cDovL3d3dy5taWNyb3NvZnQuY29tL3BraW9wcy9jcmwvTWljcm9zb2Z0JTIwU3VwcGx5JTIwQ2hhaW4lMjBSU0ElMjBSb290JTIwQ0ElMjAyMDIyLmNybDB5BggrBgEFBQcBAQRtMGswaQYIKwYBBQUHMAKGXWh0dHA6Ly93d3cubWljcm9zb2Z0LmNvbS9wa2lvcHMvY2VydHMvTWljcm9zb2Z0JTIwU3VwcGx5JTIwQ2hhaW4lMjBSU0ElMjBSb290JTIwQ0ElMjAyMDIyLmNydDANBgkqhkiG9w0BAQwFAAOCAgEAacRHLBQEPaCfp1/dI8XZtM2ka6cyVW+7ErntzHGAn1I395p1U7VPwLFqUAFoOgv8+uWB9ABHgVfKpQ2/kKBg1owHOUPSSh86CHScSQNO0NBsCRwAPJjwpBvTiQzAE3HVx3uUa94MlhVgA2X3ARD3RMXmkKwJV8nMA5UbWKSPOrY6Ks2//TirOIZfBXyvJI5vvV3lgnYsJZjwTJehnR/6LT0ZB88bVrhb9mT31bCM7ANOP0MIZlJmPDqwnijEw+K2OGjq5oI0ezIIUEXw6AzQLnlA7OcmFXX5G+c+rt5KVzz+R/wLBq2OVN4b45k0Ixir6nPb2kk7G/bR15OYPuhEESvjgvFBOSv5RPm4QYhMUEwn8CXloGoRsU3l8vNO66xNymVIOI/NJZ2jLdAzWzEsYZTxfcy8zCvHnQj3LRcCr31jDqBPZk3/YImCd1doOOZkCjmX5Pd1XFJHDWsy3foolMxZWEwfDS5ruEnNS6oK+dO1rYqd1BADQrlWQrfysit8bqTONL7m1Mlh5N0McD8Gl8uf95BsQ7Ss8u4VUwnOSC4hwZzUMr44jWFPMzrdhbPyZCDKT8u7KgL7q6aBrEsb/9KHdJ7OKd2YNmSLJLmiOunHAf+qi3gKdQAME21e5ToLYqoZfbykvQshSx+EneODPmYhihpbp8dupzqa5GJ2UsVZBbMwggWvMIIDl6ADAgECAhBoKNVMflzavUM5rgzBWio1MA0GCSqGSIb3DQEBDAUAMF8xCzAJBgNVBAYTAlVTMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xMDAuBgNVBAMTJ01pY3Jvc29mdCBTdXBwbHkgQ2hhaW4gUlNBIFJvb3QgQ0EgMjAyMjAeFw0yMjAyMTcwMDEyMzZaFw00NzAyMTcwMDIxMDlaMF8xCzAJBgNVBAYTAlVTMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xMDAuBgNVBAMTJ01pY3Jvc29mdCBTdXBwbHkgQ2hhaW4gUlNBIFJvb3QgQ0EgMjAyMjCCAiIwDQYJKoZIhvcNAQEBBQADggIPADCCAgoCggIBAJ4lAWYZH2Q0wZ05I2IdcYtW6iXSmx/vJwGCv3fYlDODGEibUJ57lmTC0MNfRf8ynOgXF7147XWYXzoGCCscN5tGSpAKsK9Gkj4ziSr6uOcyY/Mjx27SFPsmWO7+BoRU+sEfN6rb1OxWKr9JvczrAu3GTvysGbUSNWkViRdNo2jqbB4pmgnzznohxgnRGeqPMEZpO2gEK3yKLdZjXept1jmevQY+W+4vEVsoa6dSpGheTKTqrs4jv0w2cdqBRVCOyobO/1PDuEOzJO4HeqK0+scKHXvGUjUx7AgfhICSW/ix2jnWyefliQR+UX/05mpkR0nq+Oym9qBDU/7awyMk2CXaEywqtz+U3nccTHgcavmaj+tqFXd3rUmEzhBAx5lID9WWHoCcc6E4oQNv000g0LVD5PcueA9O97y/ZdptkAtbv97qJyeZZPg5fHM91iHS7tbzUxEuVcPc6vEpV95RoXhzkAsv9cl1NuuN0m2OeV26Gjj/3xkBqNLI0dby64r1LtHMkxObnJB4ZWN5BMTxnp+MOvNkDP6YHZPij1alY1MjuG5zFkUatvd7D82kMv9a/paN4Yd423CDqCSFaSDCbRIN5Xn2KlnP1qvngeagsYgtCIwLsc/XbDavnvkDZ9lBc6mrRbhxYFgY1BYsZbrRBd6SxVAQEZDOR8z7r78jwJ8FAgMBAAGjZzBlMA4GA1UdDwEB/wQEAwIBhjAPBgNVHRMBAf8EBTADAQH/MB0GA1UdDgQWBBQLs2g7r9qv7nCldtkh98xEFgfQ+DAQBgkrBgEEAYI3FQEEAwIBADARBgNVHSAECjAIMAYGBFUdIAAwDQYJKoZIhvcNAQEMBQADggIBAEjHN///wWhX14tDZkY6Jmsv6PreaKGPR/E9NJV62lUx9JXSOF8suo+ljVExaolVaGwrQmRqhSSgUQPH3dFyWO1sHozYkcXnSRGdGXo3WB53RPvCCJhCxE3jm4oOz0BFTxuAcFmMk4HoD9XIJpWp9x93BrjK75z76Gba5Tng0tJiw6fUthiaJ5smUEpyl9WzWyqk/V8vfuZioydmDPrZGcwRHTGoAVII5lQMmWMr6tiE1LQIFu++SluIWPQGFqDrel3hx0TWuy9VViXwngzkDxLbwH+vVl3GiQ5xqVYS5LmcqGQetUeVkq7QcMiTfXxaWPEF8Uq4bHIYqa4fV5kmdGb1HQ/fXfDnN1tfuvC07+RjB34fMhhpqXBakvl5nFjUfr9yXVNGK26jmWDWhYxmdxZ2r+LFGFviXQg21mY3F2XwLs+h5bzmjQ1ltFZTXZ/Ir05uUc+IvpLqMPss53U/QmDEceeXn3PHn8rRuGwj6lAoHQ5DzPWpG0Drppjl5Q/Fki+llsfX+jwY7h0bYQP9huckQTO92PO2YHzzHIID1WCv3/QgpOSBBiJazIUzfWT45Li/gBfU+yE/Y67nj7cXROxyLjXJC9CBHelyAwlB2d8JSObNt7IcYCUZUvM9EkntnZQijnEo+MEHVHPdOAi0hY8UbKoAr0CrtYfOtjlcc/mQomdzaWdUc3QyoWl0c3RUb2tlbnOBoWN2YWxZGDMwghgvBgkqhkiG9w0BBwKgghggMIIYHAIBAzENMAsGCWCGSAFlAwQCATCBkAYLKoZIhvcNAQkQAQSggYAEfjB8AgEBBgorBgEEAYRZCgMBMDEwDQYJYIZIAWUDBAIBBQAEIMIY4dpVBkSCusUhNBdKYO5m+koHE9grtIN2E2p400/1AhDvAgp082HdSIcoxDFgzYV0GA8yMDI2MDYxNzAwMzQzOFowAwIBAQIQCLgKjIU6C71wtyIEZc1Qc6CCFDcwggWPMIIDd6ADAgECAhASc7SyEcrojEtN/HGBMgmQMA0GCSqGSIb3DQEBDAUAMFcxCzAJBgNVBAYTAlVTMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xKDAmBgNVBAMTH01pY3Jvc29mdCBDMlBBIEFMMiBSb290IENBIDIwMjUwHhcNMjUxMjE2MjA1MTMzWhcNNDUxMjE2MjA1ODUwWjBXMQswCQYDVQQGEwJVUzEeMBwGA1UEChMVTWljcm9zb2Z0IENvcnBvcmF0aW9uMSgwJgYDVQQDEx9NaWNyb3NvZnQgQzJQQSBBTDIgUm9vdCBDQSAyMDI1MIICIjANBgkqhkiG9w0BAQEFAAOCAg8AMIICCgKCAgEAzOkyePso4uEIOI2RuEdFjbye351qkiY+98FsjoJWkaH+wJLJ0lfpkMjyEJDXzi3zx9fC9M4zCs07m8VG/zHYO6gAyUWbQK5PWvBvhxBwM2QVz8yW0LQtfzLLdaNq7IuBLgOxA3LhJisVbm/JB3uaeoqfeFShXi1E9Kse0zb9oYgO3DeQVSB4x0fAg/ryCfSvaeaY1rOhSzeFIX6W6USkpfqbRohta41lAcgz4668QtkP8yLhUojRrB9rRznu6d4os1XezVHXB6/cPll5RhUKGSmCJg6mMhobOkzMteqIT8vmrSxUmvM74B1rbJfYwDICYIY9r67aCxKrAWcVI0+mL27opiOYm44LI9pdMT8w2U8r+m+luShJAb4YjaPVJPmL3GrDf1sC9nNauVla/602u8Nooxz/W5Q4m4imaq0x/N0l+Rp/p2Mi0QmL68aHAyl8PZY6PGho8yM/QWn4a5IPAeDP2rYTyq1NucoZ31lt7CGM8i72+iJBIY/0jacyLJK5VsnYi5hbpfYqMxTr5/GSMr/Vzat9jtbeh3rshUZLMpWQ/9qJkBfWKqKQPtb58w9yA8XuoI87aYvtNt9W/cL9jPkEi2DDkozeJUVXrCM6hgdZF3BcjhswXBuAnFqC+yhEYDyRU/lk+vHTgVb6ooEyJ1V9CQk1lLPqYf6PhvD0fh0CAwEAAaNXMFUwDgYDVR0PAQH/BAQDAgEGMBIGA1UdEwEB/wQIMAYBAf8CAQIwHQYDVR0OBBYEFGWTtKAk9YaoVkuLTD4YjXcJ+OrhMBAGCSsGAQQBgjcVAQQDAgEAMA0GCSqGSIb3DQEBDAUAA4ICAQB+p+gT9Gl7TaCUs+m7LSC8EjGLtgqW6vmE4HwIKxgADwRq0/ngaoER8Bc5pInXHYft1iRufqUr2Gljz5wS1c+qIRp8sIyh7YW8P1gjXp9LQRWfn6q+c1K8ohQtCk7F45skKd6bC50w9woEiNmFGDKxm5xAly8ZEPsQKlq3O6BQrHwmYLzFC/hgysIFd1hHX1LlERjN+HdL+ShRR3YCQSSQ20Q/PML4uNYv3NjdF16uIPnfpNiXsSPcdDxCVVbcLAwXCoLojw1+Zv6GZKy8nEbB3jk/0g6udAY1SUG+ddqlgWXiH0fjoOAn1WICu4wxNJIoZSYvKAG7y2Ptoge1KMHO3kwNZYcHL3D60SM1jfl4Skw1oukbtHTT/M01dPm9+pq2kXG0sgqR6xji8K/wUmlkCgEwoah0C1Oyu37uWylCWcLx41k27+z9AMr8uLuHBcWak/gmWn4B/u0v5kX37hQx8QS2zofrFtERhL7lclEuVYNmSvFOFTgPgjPtXoB52DZb7vS6iRBa6qrAALNC9AyWsLVxxR3v/Yt3kOlVzLy9d34apCqBYLfOVVRl/7V/aqdUf57CSmjpWZ5Vc52IKB9AkhLyuQQwl+OZCQRhmPSo4Pg6h/4HGetQZpIasGt1tOlOc0Dy+/YxTDm2He9Ft8oyUA5e2BomWKByLITKqj85sTCCB0MwggUroAMCAQICEzMAAAAPZ3Awm0a+4tMAAAAAAA8wDQYJKoZIhvcNAQEMBQAwZDELMAkGA1UEBhMCVVMxHjAcBgNVBAoTFU1pY3Jvc29mdCBDb3Jwb3JhdGlvbjE1MDMGA1UEAxMsTWljcm9zb2Z0IEMyUEEgVGltZSBTdGFtcCBBdXRob3JpdHkgUENBIDIwMjUwHhcNMjYwMzE2MTgzODA2WhcNMjcwMzE2MTgzODA2WjBTMQswCQYDVQQGEwJVUzEeMBwGA1UEChMVTWljcm9zb2Z0IENvcnBvcmF0aW9uMSQwIgYDVQQDExtNaWNyb3NvZnQgQzJQQSBUaW1lU3RhbXBpbmcwggIiMA0GCSqGSIb3DQEBAQUAA4ICDwAwggIKAoICAQDexdffO+5SlFM0rdAHrGs/8qhBpu1kx5/vt9VwTKhOg9++OIG2h+lYYMrW796XeALGlDmwjttDJVz9n5qA0BFVnl2QsoKW8B8M0r4cGXdblltU9vaXOwj662AJqilDuA/ypUWKYKVf0o/K9NwZ+6dmBZVQQBY9bbkmK+QzUnlFe9iM4m7k3eiqPI1Mpv0D/mJro+ACc423dKgJsVm4dU9HVJhXYzMq33UARMk2QeAK/2Bv60GJGcQzju13XlYBbYCs4Lvu82nskbsrnKcGpjjcIbiLnyS7bixT2DXYxbZKU1RwXEhbQ7mwCZ3WiBQh//w/GOQ04sqrz9y/GtIL6OPlfhPMKhOSuuxvEZH/0MpTz6JxWljUdkqnxQQ9bi5v+nhB2kZAXEGNKlStIBU3ehCqLcrBVDJvnyXJdQmg7L6IMe35czDxBxTNZa4DI6ZVbs3jc3MAzSsgidoejwsb9g8cdIG1xxQbN0XssxKrbYtcI8ivFJwJCS2nlfPnyfnOEehfXHd0tbHoZQL5peRIF7/QbzyYRTvkS6poPf8CM9UumEtnXsU5IbPZeY3myrUK/8RkHd/xNRmbPRPJSq1PgkpsMw24yxgnKByybKKvo66J3ftxyMR8IxsZV4z/utTEZmUYaoSWgKTMgqI+uIJZHYLkeu6aVZ8ZE7rJXvkW5l/pCQIDAQABo4IB/TCCAfkwDAYDVR0TAQH/BAIwADAOBgNVHQ8BAf8EBAMCBsAwFgYDVR0lAQH/BAwwCgYIKwYBBQUHAwgwHQYDVR0OBBYEFMy3JAkG3X6UzjDG07e8Q9JSt4/gMB8GA1UdIwQYMBaAFMOckrEKPp033WsLK4Glvu0TxF8ZMHEGA1UdHwRqMGgwZqBkoGKGYGh0dHA6Ly93d3cubWljcm9zb2Z0LmNvbS9wa2lvcHMvY3JsL01pY3Jvc29mdCUyMEMyUEElMjBUaW1lJTIwU3RhbXAlMjBBdXRob3JpdHklMjBQQ0ElMjAyMDI1LmNybDCBrwYIKwYBBQUHAQEEgaIwgZ8wbgYIKwYBBQUHMAKGYmh0dHA6Ly93d3cubWljcm9zb2Z0LmNvbS9wa2lvcHMvY2VydHMvTWljcm9zb2Z0JTIwQzJQQSUyMFRpbWUlMjBTdGFtcCUyMEF1dGhvcml0eSUyMFBDQSUyMDIwMjUuY3J0MC0GCCsGAQUFBzABhiFodHRwOi8vb25lb2NzcC5taWNyb3NvZnQuY29tL29jc3AwXAYDVR0gBFUwUzBRBgwrBgEEAYI3TIN9AQMwQTA/BggrBgEFBQcCARYzaHR0cDovL3d3dy5taWNyb3NvZnQuY29tL3BraW9wcy9kb2NzL3JlcG9zaXRvcnkuaHRtMA0GCSqGSIb3DQEBDAUAA4ICAQCcQqn6PfPQD9R6elB+M+yTilK1RfY1sPhMITAgZEBsLsc+DI7dJ7JUNVed5uQvBa2ZhNr5kpVrFcZudzriPHTmdL9j7Yz940XWFCDkVXFD0B+31nqHOl8AQ+iG0WcKakAbpDCplyL+aHexoucQbVBbrH0qDl0i2wp3jcyXdNfqHNVy+CmxudNoBuv+P9o8C8pv8h0l4WaoSNKX/5kYlC7yi0321kpdDkZCzsNqf0115CD3UXOk42oGkTCT8UoqdHExW9pFgsDpQDrA5xeLPo5qpRQrILRyvbLFC8C2Qi2IoaJ/nq6N8nM2tw2d7f9Sfg53CIgHKpp9zpw9uHO1TdlWWDULa2VG4wzwYZzFtK2Zz/2mEo9Wf9UX2I61j12kqWdyocK0T/06lxnjz8GbueUhz+ZH1OxnqyOqj7U11LDknWzLSQq6Ww3Dqc6V8UuaWkbe/0Tw1A5QJ/1WTnv9tGi/+/UQydpT4uQstZEufv1kz/6N9fI7OC0g3pH4SYVhX/SuX6mo091T3eAzuFALAsmn5fsPJAg81nVUUhYCDQ50yt+piSUNpP70x0unhFPDZAIkF75D9q7RuWteaoms10Fh0NUDZ5m0cc4HP5riXq4GILrKwZAjPX5DcInLcYvnFTY0nBnpjhmNjnCuP7O/ycP28geZ7i0+7c/8hvoHb0ONzTCCB1kwggVBoAMCAQICEzMAAAADrHFW7d7BMUkAAAAAAAMwDQYJKoZIhvcNAQEMBQAwVzELMAkGA1UEBhMCVVMxHjAcBgNVBAoTFU1pY3Jvc29mdCBDb3Jwb3JhdGlvbjEoMCYGA1UEAxMfTWljcm9zb2Z0IEMyUEEgQUwyIFJvb3QgQ0EgMjAyNTAeFw0yNTEyMTcwMTA4NTlaFw00MDEyMTcwMTE4NTlaMGQxCzAJBgNVBAYTAlVTMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xNTAzBgNVBAMTLE1pY3Jvc29mdCBDMlBBIFRpbWUgU3RhbXAgQXV0aG9yaXR5IFBDQSAyMDI1MIICIjANBgkqhkiG9w0BAQEFAAOCAg8AMIICCgKCAgEApfznVuOf6oJXCFgTGVPFGT1vyGEaE/xbVv5QAwQFXlzu8fagCytYuBjefYY4xgu00NMavvuIgruGK+fTav0666B8LRkmihbw9e6KeIrZGEdA9x6V4eCjd0C5NbT1DSxbLF+r6SSvT8wNZNAvG95c3Og9HqqaXO9X3rXMeLMw3kHa/uPurHRvmuGA7n8T6QsRSXCwLjeWSYn8Nb7FKI66pN6rzUO0up3DYNsjKDZyox1DMnVfnJZJi9eLL+9onS5WuFIu5pYST8Ee9VzvDZm6a6fFGrZwBDO72M3Ikmo+f6c5RLHOwlNl519sw2OWgpiuZ7n9aWTIPLcFWTbgqHEHimopwOn1Ey9xo5h0BOuWiZyTMBXsna4bsND4VtslMhot/ZOtm1lKENBU3PsIjh2UnUpX5gOtO+7hWC6qzeOgIXQBN4/wfIF3wxekh+cn9+9bjC1zfL/m4ZTHnHeAiE3+0XEWikdwsHoYrMggGugqH73XqEJ+jq7WnxLSCcoLW71w6gfbOKIv44RILJWaxHNvB0xmFUpiYRedYjOmO/vSci+fkq6SAz9bE/qctvDbIodJxUqyiIYw69SmbxLenCFVZCEPmCmgavLfr9yAoIA96/7ZfXx9oIvEgjXUz7Vc4KSAU6/gqpHE8+PaxSTIrcAqkK7ENMGXQ7YnDxc1Xn6zez0CAwEAAaOCAg8wggILMA4GA1UdDwEB/wQEAwIBBjAQBgkrBgEEAYI3FQEEAwIBADAdBgNVHQ4EFgQUw5ySsQo+nTfdawsrgaW+7RPEXxkwXAYDVR0gBFUwUzBRBgwrBgEEAYI3TIN9AQMwQTA/BggrBgEFBQcCARYzaHR0cDovL3d3dy5taWNyb3NvZnQuY29tL3BraW9wcy9kb2NzL3JlcG9zaXRvcnkuaHRtMBMGA1UdJQQMMAoGCCsGAQUFBwMIMBkGCSsGAQQBgjcUAgQMHgoAUwB1AGIAQwBBMBIGA1UdEwEB/wQIMAYBAf8CAQAwHwYDVR0jBBgwFoAUZZO0oCT1hqhWS4tMPhiNdwn46uEwYgYDVR0fBFswWTBXoFWgU4ZRaHR0cDovL3d3dy5taWNyb3NvZnQuY29tL3BraW9wcy9jcmwvTWljcm9zb2Z0JTIwQzJQQSUyMEFMMiUyMFJvb3QlMjBDQSUyMDIwMjUuY3JsMIGgBggrBgEFBQcBAQSBkzCBkDBfBggrBgEFBQcwAoZTaHR0cDovL3d3dy5taWNyb3NvZnQuY29tL3BraW9wcy9jZXJ0cy9NaWNyb3NvZnQlMjBDMlBBJTIwQUwyJTIwUm9vdCUyMENBJTIwMjAyNS5jcnQwLQYIKwYBBQUHMAGGIWh0dHA6Ly9vbmVvY3NwLm1pY3Jvc29mdC5jb20vb2NzcDANBgkqhkiG9w0BAQwFAAOCAgEAIFtVVq4UcEcLxmMCwOttujFKbGz/fV9IvLwNW/R18E9gbbXjzzk5hXXKyYG8WasfxeeIlM8Wp6a9wQoO9gdHQHpFDf/iLti4BcdVbQHTV3cPQ8Lh81NrfWkf6p+7QmWZl7QKn1Mt3V8kd6mjmQY3CUCItWTnWAySd1i+GKrL5YlkaPn1nAnIseeG/r3muJ6yH18ees4T/r0Ozft8CskDXupM/gHpKzLsz1oDHYECiymy9lynkeHohnuFDEKshmMSV2ZPKb1aZOcBIfLEiKFcG+reeALLFgdOC/m5f70O3YnuczKp+vYj3Nhd9hX6W65L7bH2USqYo0PoxLkXlLFGuFEgRlsruDVq5DYtwSfIb4JIuBzJEZpVydmCoFLGitFvCFSKUr6E7NjGGivaCjoZTTLUY2s/R/LDJp2dRdeJyaQARs1vXSc9Gmwru88KFA7wK2/U7uFIhIALSP4ZswE+IonTWKs4H5K5GJCg+4bdWDR8Tq3oscLaLLhf8pHOjEKFbYvvGE5Xux5fEg5afyROFNppzBL8mhAtouYOQxgR/W5Ku/mUeMCfaxAPHjI42aAusq1G4379jfoFMKq7TxDbGrNs+Md7IszznnwCMnuItbk2GHW4lxwU7nrUrjJrlrUZXdYaMJqOXS3jhOT1KMk6E6sMi+r28jFtbqEmlAB/pisxggM4MIIDNAIBATB7MGQxCzAJBgNVBAYTAlVTMR4wHAYDVQQKExVNaWNyb3NvZnQgQ29ycG9yYXRpb24xNTAzBgNVBAMTLE1pY3Jvc29mdCBDMlBBIFRpbWUgU3RhbXAgQXV0aG9yaXR5IFBDQSAyMDI1AhMzAAAAD2dwMJtGvuLTAAAAAAAPMAsGCWCGSAFlAwQCAaCBkzAaBgkqhkiG9w0BCQMxDQYLKoZIhvcNAQkQAQQwLwYJKoZIhvcNAQkEMSIEIPmbiTyHzC8xMOFE/q8BYtSc3xgdn4dchmXXSifdvn3XMEQGCyqGSIb3DQEJEAIvMTUwMzAxMC8wCwYJYIZIAWUDBAIBBCCCDqGeLc5pf6O6dZMrbbowsBAj28I6JFLA/4MI5q2GEjALBgkqhkiG9w0BAQEEggIAahxRig9zB+9U5OboLSrOGZw3/Nwnqrql8l7fOz9GPCCv/FEdRnA2aN5WmaUFrbqohT1OZhcWDFOOpl2jYrYaHCbLnduRHzVzfdrCgqVVVFv/qtvhEbfhwfmDTDMDZIANWEzbhZ76yKYjZAx0ikuX2fLb0cZgAzQAXDZvVAuoX7b9bHcENi3aLZg0pwQWrHXKTDAl8Zjyr/AfBUsGs42SMXYR7Wpjh2eRke9FGy+QwpJ9AANGoVboc5jIiY0d44DM+l9CXiyxw2EyYPUfB9E0kOuyQ0Jbt8aF1VssZF3XX+f4yP8k54CkDlPFVx6CkpJ3pLp1U1BHC+QbIcIKjrhzk+nAWb+vMlYsYkDnC+C2mYh8ATJGn5bGLd4sqreKMg1zNdynvUq2eizyBpY1J3GPyG+JPiTOGUsLvskGfjZ2wFiGyDQTMB+AZFFPgnzuovLJ6WhWqge8HgbKnwKHRxmRI51bxAce4nfaRkeB1U3bt6ZXYE72vPLtcecMNAw1m3Moantazm0Apd6xMyWZmBqMWkwvVNH3ca3x4Zvohvh/ywGEX6+7hGBGvt/jFGbSNQhs3iQit8mKCCOHXgnmzLZM7r+hZhqUuM1BKvINiwkQ4oFbYkYgid01QERlgsy0Ue6l9krXJa20+89SljzLtVDsf+OVscVRc99J16Ks0GL2+cpjcGFkWRZLAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAPZZAYAqgc2Ok/Tx7Hb4RzBloAOIBMm0rR88y8dV/uhByTv7+Wvj9UlfBnjDVOXe4+Apha7pq0/lrTvt8nYfRCG0c16U4qwOPNj77QUFca5548UI4vqd1iyJRww5t8bqr6Iq1LL8afCGxT74aBNB13LHhcIPHMtwEgmxNga2IL/MARoRBZzePWq5JqSA/ekJDHW7GnE4L3AXdgvZvcZ50QpmteidJN5SIWJ3X4PoVaxXooyDqS4kQcXKBL1mcIpxFKeSMiA381FibPVTBRQHsr/rAUz7sKEJ20U+w95OrcDqZWtc4hs6rzdarWdNmRBAYTQMRLBLLMI2bTMaG/8apQ8rxy/5i3AI/cH2ES8XnxyfbQULhr+62S704ghHx5hoELM6KVNWbJIc5ketErMRP7BvgGFvZjPoslsWgI2sm360wKLqwYArcmdAxMjynLlb2eiYeOoDjNZKwIDkvgfdVBPOVe6rmbu2fcQdYGceNtfAJe1cJm99hHawx0lC1Xdp0Q6JaONJSi0NAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAP+6SURBVHhe7L0HnGRHea/dl805707oCStEjiZnBCJHkREIRBCghFBEIGl3ZpOECMY2mCyMLJNzRhJIOzObJHLwNZfri40DztjG2GBgp7831qnTM7vSvZ8D4Xn4/X+nTtVbb9WpPt1L/XV6ugMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAR+F/xBEAAAAAAAAAAAD+L7mFaLFouWiVaLVobaU1UaftS0QaDwAAAAAAAAAAAIE+vfb/+gTbOtGxoruJ7it6iOj4OD5IdG/RbURbRAtE/7fwZB0AAAAAAAAAAEDFss7SztbO5lX3X3z7gacvu8/Wl6540LEXrjz+NjtXPfz2e/S48vjbbl92v63nLT5247M7W1Yc31nSuZX00yfrAAAAAAAAAAAAfm24uU+e6VdQ79RZseQRC7asecaiW216/qI7DD936d1GT1px31s9a+mDbvPsFY+6w/NWPv5uL1l38v1fufHcR75u82VPvXLgd5/zicF3vPCaoateMt1972n7R6560fTwVS+6fvAtJ39286ue8u4N5z/6DWuf/4BLVj7xrmesOO62Jy+71zEnLr1j96TFx2x89uItq5/ZWb7k8TLu3UUDOgkAAAAAAAAAAIBfFW6eMbduxZ1WPujWZ6w/6+HvHHzz86eHP3r2t7sfPecbQx982VcG3/fSrwy+W3TVmV8duPKsrw+++6w/Gn7v2X889IFz/6z74fP/avijL/+b7icu/Pvupy78h2E5Dn/igr8b+sQFfzP8yQv+fOjj53936CPnfbv7Acn1gZd9VfN1P3L2NwavPP2Lg2983v4N5z/2IyuOu92FnU0r9euvS30yAAAAAAAAAAAAv1qo8bVZdExn5eI7dFYvObazZunWpb8xevyK5z7g3I1vf94nBma2fb/7jVf3ujde2huZ2dUbuX6iN/KF7b1R0ZjqOtH1k72xvRO98enJ3vjMjt6Yal+j8f1yVEl5VGJGpyd6o3slx/XbJJ9oentv5MDO3shX9/SGv7jz3za+8wXXrT79uN1L7rn18Z3uhnupSdhZ0rm1zPP2NtdOZ72IH5AAAAAAAAAAAIBfWNS8OtoTckMLtm58/PIn3u2iNRc+7op1lz/j99fsfvJb1732Ge/b8Obn793ygbP/pPv57f8+NrOzN7ZvZ29cj2m8TU82kvPRmUmXmXFSl2ac6oCcH9CjnkebxsxMiNTM294bnZowk677hW29TR846y/Xv/2FN6x9/TM/vO41T71qze4nXLHmwke/c+2LH/iupfcef2Vn1dIHytxv6gck9Lr5wQgAAAAAAAAAAPhvpzap/kdn7dLxJfc65nFrL3z872z+2HlfH/7q5b3ugUt7wzM7eyP7d/VGD+7ojd2wW7SrN3ZIz0X6VNv+0L7J3sjMjkpyPq11E2a6mUm3f4dIzbpJqxvZNyF99Rjlme0WP2pH7SNlfapOyl3NJxqWumGdyzde3bvlH7+mN/zRs762+qT7butsWXlHuY7FooV2RQ5GHAAAAAAAAAAA/Jdxc8yoxfK/2y0YXvfYZXfd+rTlj7/LS9ae+fCda7c95YqNbzv9y4NXT/z7yPTO3tj+3b2xgzt7o2rE3bCjN3LDzt7IoZ29rtQN75+cHd63Y9a+cmrGnChNuWnVZDHm9GhfWZU2f4rOTbYiM+a0rOZcY9DpsSvHrhl40lfGHTukc9ojOfZIn9294ZldPxm44iXXrXvZI/esfca9z17xwGPPWbBp5TPkGu8nGrSrPTIYdwAAAAAAAAAA8B/GTZtNyztDKx5wu+du3PW039/8rlOnN7//7G9v+uDLvz/w/vP/auhTF/1z9/rtP1cTbUwNt4M7RNXxwK5eV5+gU+3r08xO1/SORlOVss6eglOpYZf1Yd7p35vT+jDpzJjTshp3+2UOagTaXHaJ9vS6e3f9fPiTl/zDwPvO+9Pup1/xF8dMvfL7A6952hdWPuJ2uztrlz5YrnaVX/S8YMwBAAAAAAAAAMD/EzdlLC0SbdGn4zqrl9+zs2Hl7RaNb7rrqife/ZQNu57x7oGPXvj94S/s6I2Ixq7f2RsXje6V82nRzI7e8NTEYXtSLb6uamZcMeT83J6q27fb6kZU+tVXNd3CaBsJU06P5Sk6NeIkxp+k01iX1UebmncjKY2N8fRrtTqX4Zkdsyp9cm5sZo+MIXOQ+Y/I/Me/vKfXffdLblj34gddtvRuY8/ubFz5EFmHW4m6orWio/1QBGYdAAAAAAAAAAAcFTWQbuqXSFcvHN/0qLXPuOeOgVc//UMDr3vWBzbsfNq7N/zWc6/bfNXp3xm5+qKfjO2d7I1P7RBN9kZFI9P6pJoc1Zib3jZrJtl+/7EHPRZzLjSWT8tpuz0xF8aaPQ0X5pxKc6fivDHl5Lz0E+kTdfrV2FotY1DKpoiNccbkOsY0/3XbeoMfOffvBn/vJV/uXvGivZte9bRPrTrh7q9dNLbxFFmTu4jUsDwSN/UDGQAAAAAAAAAAAAU1k9RsUqmptLqzbNnI0ruMPWTteY997eDHzv3D7iH92udkb1SNuBn98QU1wfQrov5jC/ZDC3puxpwew1ib2eG/urovzDnVfjXK9Ak5N82KiaZPy6WpJn3yyTnN1VVNTZhGZA5mzJk550pjrjwll9KxNF8eRflVWGs3E1Hnu11ybpOc23pjUh7bK7pOruvzE71j5NqP2XvJ3w5dduInVh53hzM7G5bdS9Zog0jRX3PVdbspkxMAAAAAAAAAAH5NOZpxtKKzYtFdFtxq85OXPug2L1hz8gNfueplj96z5pVPfsvGt7740PA1F/9En0wbn9nZG1dTyzTpxlaYW90pfTrOTTQ3y9RUcwNsXmMuTDJ7ki3LtdQ4U6npJqqNOTUIW0/PyTglvuTQssjGiq+wlq/LqqRe+6ks10Rv+Lpts93rLpkdlbIZkPq11imZt453/fbe0CfO/bOBNz3/k5te8djfXXnC3bctus3A8zsrlz1V1u8eIv4WHQAAAAAAAAAAzMuRjbmNS26z/DF3OXXja5/5vk0fPOtLmz/5yr8e+MQrfjD00Vf8bfeaS/5ldO/kz/WHHEbt78WF1Jyrvno6ZuWdvbE05FT6dVRt2+emXDHmtE+Wtd6+0tqcq7HWlRhVMeam5sqNuTDX1GhTmSGn5Tiq1JQL2d+ym5HjtIyV84xco3vVkAvF+ZjEqSEpY8yOfH7bjwc/dcE/Dl9/0T9tPTDxN6MfeulXNr7yiR9dcrfxszuLO3eM1ZwPjDkAAAAAAAAAgF9xbo4BtL6zaNFvLBxYe9yikQ33XHKP0YeveMEDL1jzupM+svljr/zroev39LrTe+xHGcb1hxkO7OiNHtAfS9g+q/K/1aYGmhpvYaapIVcrDa9pN+b6Tbwi7StS00/NufmNudCUnM9XNqkR5+rKWKo05rqarzLl5hhzkqsYc8Wci7I9obd9VqX9xg7KmhyStZHjqP5gxJde1Rv9ymU/GH7niz6z+un3uWTh+PpH2Q9ldDq3FQ2Ijva36DDrAAAAAAAAAAB+RVCj56bNno1r7rbmece9asvbTvnCpre94DPrXv/cz2z43RffuPm95/7p0LUTP+7u29Mb3rer192/o2dfAz0gR5WcD89MNsacmXMp/7rqHONNja9yLnFVW3myrVL+MIOfR4zGqiRXbcZp2c6zXY056adGXJp6VlapMVdMOYlTZS4z5lRuxvnXWsOYs6/obp/tTm+ftacED4oO7eqN3bBbjq8SXdbbMr3rp4Ofvugvup+84JtD73rR1RtOf+ibl95r60WdFUsfJqu9xhd9XvibdAAAAAAAAAAAv4LoDxIsFenfPVsncoNoVWfj6pMffM7mT170naH9r+oNf2F3r7tXTaowq/bt6g3LUTU4vW3WnzarDTh/ss0lZavbYVITzZ+MU0m9yc/VOOsvm5mWxpuOk5K8agJq2fLqvDRG5tiVsilNOTPmtD1iJL48aRf53JxzY87qNFavd0piVFoOg86lppzM0Z72m/QnAKWf/diFPj14SOJvkJyHZO1u2CO6TM4v7w0fuLTX/dKremNf2/2j7ofP+OKq5973sgVjmx4tq75ZpIbpcpG+JvraAAAAAAAAAADALylq9BztiavlnRWL77z49gNPX/GYu5y77vRH7F71/Ie8fPWLHza5/jXPvXrwmm0/GZna3RsTuXnlZpY9KSfnQ1OTs0NT22fVxComm9Q3xpweXfnknBtpERNKc64xz1zFpJOYOcac9rO/Yae52n3TmKvrijRWr8OOel6rz5hTTakktjbm4uhfw/W4xpjb5U/0HdjVG94v67N/56yZcwf39EYOXdobUYPu4GW2hluuveQft7zv9IObX/OMd68+6b47F/3GyKmddctOlNflN0RL7BWaH77eCgAAAAAAAADwC87Rv7a6YeXtlj/uri/a8tpnvXfok+f/z6GZiR8PXHPRvwx94sK/715zyY+6Uzt+rqaTGl9myu3faRrer+dq1LlZp7+2Wr6Ouk/KUqdG3HzG3HxSUyuNvTTEaqPPjbJ5zDTJ6WWdg0rKGhfK8/LkncpypLJ/GnKVrK/IDDlRGnSm/DGJpt3mbnNqjDl/qm9Xr3tA1umga+Sg1IuGD+yYHdo78bPBa1/544Gpi346uv/iH3Q/+bJvrtv++I8tufvY6Z0lnVvHqzQfGHMAAAAAAAAAAL9EDHaWLbzvwqE1j1h4m4GHLDv+Ds9cc+pDL97w6pM+NPDhC/5MTaixA6JDu/xrmGokHZjsDe3bPjs0o19X3WFGk/9NOTfJ3CgTzahRpV/pVBMsjLl9aqqFSWWScznW5lcx52RsM8/U7EpDTHPUhpooDbcyrhlzXrb5RN80DT1X5om2yNGYciqZT2nfFdKyaNplxpyZcaJ4as6elguDzsbWXHGd9nfq7FzGi7+/1z24w9ZT13VEZMeDov36N+q2yfhyvP6ivx1443M/rj8SseCWmx7dWbHoLvLa6Y9E6FeNjwZmHQAAAAAAAADAfzFHfyouWDi28SFrX/zg3958xSnXrn/HKdese+up+ze8/YxvDL7/3O+NXDvxr2MHdvuviOqTXip9suvAzt7QvslZf1LOjSb72qgpymZ6iaYnzViyp+ekvphuaVSFihFm5lXESP85xlzmreRmWRhrqpxL1DVP8bl8LJXGaN+2PG/ElPrIlWWZU1eubdiOrmLQhUo+zWNmoRtzo/t223XmeqWR2N036TKDU9bMnjLUtdjRG7rmFT/Z9MGz/nLwk2d/Z+hdL/z86jMf9rtL7r31jDDojvS1ZH39+ZEIAAAAAAAAAID/RtSg0R8OWC8aEukPCizubN18l1Uve/ieTZ87508G9k/2hq7b0Rvdq18h3dUb37+rN3ZAtbs3vG+n/pDD7KiadKYw0EL5lJuaTGmqFRNNTaqWMSfnZsqptH9zLDkzjxlkmiPzRc5QGnFmtpkBFm16bnV956HGnNN2Nc/8b+O5ieZ9rBzjFjOuT8NHNOayjyjHy+trGXNxrTqW5LNxY3y99u70tlk350T7fe2Gprb1BqcvltfqFX878LbnfnrF0+517sLxjcfJ67lFtFKkr7H+UASGHAAAAAAAAADAfwE3/WTU0s7WBXcYfMrqk+43ufaiJ/7OyrMeuXv1BSe8ad0bX3hg8OpLfqKGkhpy+ffbzCiLp+SG9+2YY8zpjzikMedmU/TbF+ZbmE1mWIUxp6ZXGnP+lFgcw7RyRb0pzCpRGlepNMf8XGJV0ieNMDPDapX8asKFEVfapb6URZLLTLU02eSoUhOuSMYdlmvKebih5v3TlPO8MV5cY45tc5FyvzGXRmZtzI3smzB1VfqV4QOTveG9k70t7z/7exvfdsoXNr/+2e9Z95KHXLrsgbc5e8HQ6mfJq317f9Hn5WY9SQkAAAAAAAAAADePo5ktt+is6AwsfvDtnrThVc/8vS2f3/YXQ/sv6w1cN/nz7rWX/Fv3uol/705NHFYTzP9GWxpKcTygJtLO3qgcU/4knZtyajCVv5tm/SNHmE36dc/amFPp01/+9+g8d2PM6TFzNDqaMZeGmBthcjyS0hTTo8nrclx/ci/mIvGWU00ykRlzct5vzJU52Dw0X46VeTJ/jKdtEZttVqfnUm/S9RJ1ZybkXKXrJtqvppzLjLmZHbPDe3f8bOj6nT/vzuzpjc5M/vPYZ8/7wy27Tnj/0nuOP09e9zF/+ecFYw4AAAAAAAAA4D+BYxeuX/GoxbcafPLi+97yScufdI8XrDj5geesufgp79j0+2d/uzulJs7u3th+lZprO3pmfpl22hNcakr5029uWplxpUZcMebkaH29rZhZlanWmFZqMrmyLePMkMr+klOP9kRYxoXMdCuSOjm6oaU5vN3VGF2N+RbnasqZMRd1ojTNVG7MxfXqvDTftIxlX0uNMaQ8HPLxYm4an2PGmvm4Op7nL/PQPjafjKnqU2bMyXrNTKjhJkeXPjFnP7hxQA06Ode5yus4vv8yidsjMbt6w59++feGfuekD6x57n0uWnLbocd3Nqw4vrNw4QPlvhj122NeeIoOAAAAAAAAAOBmcDQTZcmy3xg/cWDyKe8bevdLD2y+8owDG68442ub3nH6N7e899w/Gf7sth+O77vUTbV9auqEEWWGkhpeUi/Sr7SO7fOvq5p5Z6acKn6h1Yy5RpqjGFua18wmN5nM4DLDSY0lN9oyxhT9G2Mu6kU2x5uS5dYxVG6MuTnW1lyjrs6j866kZpuqGHOqepymzuIiT/MEns5BzmtZXdRr37otcvVL16r5qmttbKYpp+Xdsu575HW6VHRZb/jqiX8d+OjLvz/08fP+9+BVp01vfu0zP77ymff+7c7wqsfJ/bHMb5M56Neg+Zt0AAAAAAAAAAA3k0WizZ0lnVt1lnVGOqs6G5ff/9aPGbjsme/qTm3/4fDUjt7IdTt7Y1M7e+P7dpvRNr7fDbTu9OTh7szkrBs7avA0hpKaa2bISXwac2bEmdycMxNuXxhS0q5Hf4LOyy43stzkknHUbFJTyQwnjzHTyeaU48vcdD5qjkW9GX8an+fSZnVZH+O42ZXmV+ZrNJ8xp8q5HNmYU0lsGUOPYZ5V12O56rGiLmVtNrfInW2Z15TXErk1VuRrreehXIs46mvpr5e8zgf3mPQHJjTH+Jd398avvuA7Gy541O8sufPo4+WeGQwdI9ogwpADAAAAAAAAAKi4OV8tXLn4zmNP2XT6o9+w+bIT37X+FU/87Y27n/GBgSvP/J+jUzt64zM7e2P7XG6A+d92U3WnJ8yY8x9q0Bg31YoRFzJDL8puBFUKE6t9dCPJZOZSYyipSdQytKTN59XkTPPKVOrbMbVa45myv5Rb7VEvuYpK206bZzH15hhlMtdK7evwvnUuzz2PMdcvye85sxyydl2fbOubg9TbuqlsbXSe8lqmcXpQdEheM9UNe6yue/UrfzR01ZlfG/yd53540zmP+a11z3rQG5be45Z7OiuWPELuowV+O81BDTu+3goAAAAAAAAA0M+iOwzca+OeZ7x9+OCuHw/P7OqNTO3Sp+Nmx6Z3zY7rk1P7dplZ5iaOG1kmLc/ssK9HZps9TRdP1RXpVyTDmPO+bgLNZ5QVMyyVhpKaS/pUnppZUl/MLMmR5lgrT+aydq3z8dLsstxaJ2WVx+m5xzfzmK8c+bRcxlLTzMtpoPmvrbZ/3MHzxHWUer0e72P9M4eM4ddX1UV/jfcxRbFGulZ1Ocex9bO2Zsw5xlxKxtFfzK2NOX+6Uc9397Ye3NMblxj7OuyB7T+/1bdf3Ru75hV/vubZ9391Z+Xio/2CKwAAAAAAAADArzQ39XTccGfVsgcsuN3mJy196B2et/zJ93jh0ifc/eRVpz3iNVvecca3xu2rqrt6W/ftNqkBk6aaGWtqSIk0pphuMzsbRZ0/HecaTUl9Y5rtbCT1ZoiF3LhSwylNpCirwSTtbkz50craJ1QMNotxkyxlxlKrzfu40RZlUTPPrFNDLEwxNS2jr8VYXz33HHOMOZmzGXM2/5hvqG2UeZu1a3/NF8dWOWJKn8wR6+PtUY56M/HsvIqPclmLkF17vtYHQgd3a+ysrp+W1awbV3NOv+KqT0xObbO/6Tf0ntMPrT/jYbsW3+/Ypy6+1ZZnddYse7rcb3f0225ebs6TnAAAAAAAAAAAvzQc7Q/vL194zJbjV7zkIZeue8vzPrH+/Wd/deN7z//uxivO+p8D7z3nT0Y/d/EPt+7f1RtTA+bATnu6rf7qaWPE7eqN5w86zFSmnGqfKtpaxlyYPpHLjKyUmUMuf1otziWfm0uN2ZRPgOV5U+/93CiL/qkYx4259jy0LueQppfnybnUbbsaM0vrTXEeyhztvz9XGWqi+sm2Rnod0Z7KMfr6m2Jc62vX36xLe33cHMyySX/Uohq7GHNSrz/W4eacvo7y+qVBZ6+lrImuoX3NVbXDjmMHda129QY+/op/2vKBc/9k+MPnfG30s+d/e9POJ12z6A7dF8l9t9JvvzlgzAEAAAAAAADAryQLRYOdxYvv0Fm95FadoVW3XXz8HZ666vxHvWbju0770ubPbfvJ0PW7e93pPb3uvt1mzNgTawf065eTh8142R91cjSDzgwbV/ml1Rk15HbY0b7Suk9NGjfn3PgKU06VdaLGdHKzKI0lk31dVerUOCqGkptLblr112V9Jesb5RirGTuvTcpHMOb86bHoK2Vvk7Jct48R/aJvqh7Tn3bzfmnSWYyacGaIaTmldX5dmavO7+M3SkPQz+P6dd1MXqdtwyZfp2IW6ry0n40r5bwmnZMacyo9tzXQ1z4N1mrddJwDk/L66tdZ9Rdeve+I3EvdKb3ePb1jvnV5b+C3T/zQsvttfWpn6dLxzuLObeWe1Cfotoj4kQgAAAAAAAAA+KXk5jxltHzx7UZOWH/GI1+/fvczr1gz+fR3rnvNcz+z8c2nfmXwoxf+/ei0GjD6FUU1aVRu6LgxN3HYDJl9arCp8eZHN9yao5txqvgbc5UxZ1ITx56wSjOnrf6vZrbMOTWS7KgGkUjm5uXGvMpjv0rOUP+4lmtO3RFk/TVvzjXXS8rSXl+D/R23LEuMS8ttNfH9bdK/yuHyebTrJDbmZuWsD1POzzWXHFValvq2MVfNvTUPea30aCablitJXPO675CyGnJqzE3IcUJyqeQe2K9Gr2h6Z2/LB1723U2/+9xPbbrsKVdsfOXj37/8oXf6nc665Y+T+/NIPxKR8DQdAAAAAAAAAPzCo+aFmhyLRIu1Qll4p+H7bpx4ypsHr9n+d0NTO3vDe3f1Rqf2mMGipksaMuWJKTVb1HzZr+ZLPA2XsrYwZcKYSbkpVz8p57Ky5tovOecz5iSvzUElecwckjxuymm9GlRHMObUXMr6Vpur5JUxtC0NtBy7zEXjq/q5auf13J6rHK3s4+VTeuX8/0VqzKXk3PM0Y5Q6nZ/NQcupZv2aPNFH18H6VvNuzd3z+d+X87LmKaacyF5zO8q5qLzWMq7+Iq+NLTEer/X6N+j29LrTO3tDB2Scr7+6N3rD7h+sPuP43+5sWnHXuFUVvXf1HlZhxgEAAAAAAADALww3ZVQs6KxafNvFtxs6YfWT7nbO6mfd98zlT77XKSvPedRvb3zbKV9VU6T8Qup+NUvcxEmzZmh6ctZMmMpQcUOuMlhMacBVap17XP136VxSP89Tc2YGmdTMCQNJjq2nzmKubUm7HL0980k58jV5o83q6vGqdj1GDp/jbtOIPvElddlH16EeRw2tPGbZFCaYlXNOdm2uEifSODcfs0+TJ+Ozzdrt2kNxniable3c+1i7jVeNaTF+TRYrynzWJ4w5fdLNrleupUjam78916gxYeP1lzH913llHnJf2ROT+quuKlnX4X2TvaH923uD7zvtS+svePRrlj/hLqcuecCx595ieN2pnYULHyX38rF2Px8ZnqIDAAAAAAAAgP9Sjm5ErOgMrHjMbzxn8+ue+56hz13y54Ofvugftrzvpd8d+ug53+9ee9G/6B/ot78BlkpDRk2bPIaB0xgrarqE5NylbSmNDZkxo0ZNHNWMEdmTcqYqRs6LESZ1bhr5+MWMMkVbztX6RFljY75+rlJDycueN8cJ02leZYzL5hrmXDHm6rYo1znmGnM6txgz51ZdV+mXasU3eUp7Kq89pXWyRvlUofW3WD/3eTS5rV36NcZcc226DtZejLmQnoea11vLeV7fDyEZ0zUp6yVHu/dEB1Vy7x2So2jo6gt/uPlD5/zF0Gde/mfHfHX3D7pXnf7NlU+591sXDKx5qtzR+vfnjgTGHAAAAAAAAAD8h3NzzAb9mupQZ/HiO3Y2r7jLwlsPPWDJg2/7+KVPv8+LVk889arNH3r5X41N7+qN7xPtV2NsZ9scObDTVcydXXGUtkqN2eZGjH9dtTLmNE6O+USWxqmBVww4Gbtfahq5MZfjS9li2+P3G1gm7S8qBpPGxlc25zw5J2WbRzlXA0r7eJse++dpP/6g0nMz5rysKuaVtoWaeenYEZN1On85NnP1ura0TtpNTXwZM86zv5WzrTWWSvvrWmS8r1+dt/1Uneb381I2SZuZcB5fTDobQ65b2os5p2U7j/vBFOcyvsleAz3K3FQHZT6ioX3bD+tx9JC036BzmOwNzWzrjXx9d6+79xX/e+PEE9++/AG3fXFny9onyL1+N9EtRWrUla9oAwAAAAAAAAD8R6Km3M35hcq1S+44/NjVLz5u95rLn/yONb914kfW/ubJn1//+hdMDbz77P89et32H6shN37An1Ybs19VFdn5Tq8TmcG0LyXtITV33HBRE07aTH5eGy9poqWsjxkxnq82sVLFTJpXmme+esmp5lD0L0aTGkz9R1HbbKxkbZ6nf15pxJkZp5K6chSZSRVllfWxvHrejF+Pp3X1cV7JWubTbZZD66px5kjHy7hy9HXL/i7PV8ZRlRx63uRyo66/v0he6zT2WnlaitdfYmvZvaDSe0dyq3z9pax/Y27/5OzIQTkXjd6g2u0G3YGJ3pbPvPyHmz505h8Nfej0/YNXvnhq7SkPunLhbYZe3lm68GFy76/zt8C88BQdAAAAAAAAAPyHsFC0VLRStFq0SqQsX3K/rY9YfckTfmv9h876o01Xb/vJ8LT+KqaabbvDcFMjRM2kyVk1QUxhjrSMuWKeuLHTnKckTvK2TTnRvjDhol/LlNN4y93Mo1bb1OlX5msrzcM0dtKASwNJY0qdqOlbmU8qbeubj81JVRtzcZ7GXP1EmbWFbE5Rl/Mp48gx51fq56lz4+soxlyM0ZpD3b+vb6Mqn8lzeE4/L8p6a9MxIoe81i1jTu4H76NlVZpyKr0fvN7vkTiqMadllayL3YOyriq7L+2JOdWO3vDBydkRKXdVB3b3hqb39Lpfurx37P957eGRD5x2w6qn3fOyBQOrHy3vAX16Tn8kQk1sfYLu5pjZAAAAAAAAAACFm3o6bkNn6cIHLhrf8LxlD77dOSuf98Btq1543PkrXvigC1dddMIVG97y4i8OfWHbj81M0ieP1BiyJ5F2WFkNp+7MxGFrl3KacbWKYaLmisiMNjVXSn0ozBY3YDQ2671sCtPFTSk3cDSmtIvMBIty+Zpln5pYOWq89dGjnrvMfOtrT2PMTKSqvZhRFudzMFU/7lDqpI/NUevVlFNVuXweleI6tVyMuSo+52P1KevbX6/XLYq+bsA1MRaX+TRGz61e1kwVr3nmLnlMvgYlZ9b3jdFci+f2/HKU193G0KOVQ1V9c89oDj2GJJetZ+s8jGG5V8f0hyAOqXb0hg9snx3eP3G4e8Nkr3tI8h2S1+CGS3sjBy/tDU/t6g1de8k/Dr/vtP1bLnvKVauffe/XL7r15tM7S5Y8Tt4ntxctsXfM/PAUHQAAAAAAAADMy5ENgy2r7736hHtMDL71lGsHP/eKb2/5ws4fbPrkRT/Y/JHz/nroU6/8h+612/91ZHrbz/PvxrnhkibNjmKC9MufmHPpeZowJjXGTH5eDJY0X7KtxPmTchab+Synm1yNySPnkSvHaoydtrLdZDnmmkVmTtVtZjxFm8aVNp2HH60sbeW8/LDDEYy5jKvHMqOuUhhRWk7Dq463uZj8unx+ep3NXFsx2i/rNV8erV7HzLK3NbHRd17pPGp5DjMs9dzKmqcdM2cMnXNr3nEdUadxtXKtc11zPnmf6NNyZg6rcWz3sP8gSfegSnLrUWLcJN3dG75+4mcDn3nFj4auu+jftn5r90/GPnH2t9ad9ogrF9164GR5t+gTdEcCUw4AAAAAAAAAjop+XfXYzprld180uObuy+53q6dseNmjLtv0zpdMD1y7/V8H9+7udaf29Eb3qXa76aFGjhkXbtgMzUzOpvGlR/9qa78qY04NETVGVNGvGG6aW47+lUQ91/Zoq+r9b81FX8nTmFWaV2JE3jfHyvM0dDxfKZcxmjjPW/UrdZJPr1/bpJyGUZ57ez4VF+PXfcOUM2Mu5pxxagTlV1rLGGYguTJHM6aU47weP9vLtWl7rGG5vlpxncUsq+R541xjIt6Pki/a++dY5qHt0cf7SZuo9USh1lt7XItJ46ReXyOTG6hmokZdHWevtUryuBrjM8fPeyHvS/9xCP2V4DDnpK2rZvO+ydnh6cnDPjfNI5Lxhq+f7OnXt4c/dfGfbn71M9+z4ol3uXDBMRtP7CxadC95H91WNCDSr7oeCcw6AAAAAAAAgF8zjmwGLF26dekDbvX8Ta94zO9uec0z3rvpTS+c3nTl6d8e+PB5fzNy3fafj83s7o3v178fF2bHPjUy0oxpDBU1nlxumJmBln2KtG+W1USRoxompW9oRuK0f+bSY5gxTX1lMEm+lOdzMyZNmCyXeO0refI8vwapMuNHj6bKWFJpjObJfJGzMZjCiDODyg02U2l35XzcoJM6Gz/atZ+2VX37+6uaOTbya9B2nVMo22P98jo1zgy9uLb62K+mT19sVacqcxbZHCPWx6nmJX2sr5RNkcfrsq/I1l8k8y7SuHIdEhuqX+f5lHlb9TGum4spPfdcXTWctSwxeS9aDikPfvaCH21535nf7X7gzG9233vqjete9vCrFt99/BWdlUsfIu+qNf7mmhf9CjnmHAAAAAAAAMCvGWoGLBdtDC3prF06vuIJdz954+XPfP+Wj13w193rtvdGpybDhNjRUwNrbL+ci7r7ts/q11b9q6re5iaG1+l5baK5gVY/mRQGR5geHnsEWX/P50/JSf7QTRtzmjvGs3EblXg1dlIRX7c1JpAaQ24Gef523tJHy6Z4sqoy5aw+j6FmTmHM1bK+IYvpaxe5aeXzSqOqda7qi+s35UoeU1/fPmW/liGXxyjb2h9Q6XXreY6h/aRcraU/MVfn0TafS32NpU95PZqy9/c8KrtvpM8cSYzfD33SOs1jufz+tXvY7m/Npfd+nvv9Njy9fXZ4etvs6L54j8xM9Eauv0S0rXfsjXt+Ov4HZx1a86z771lwzMbHybvrNvIeWyHSH05RHe0pOgAAAAAAAAD4JUZNt5t+Cmfj8rsve/BtX7r+pQ9/7eqzHr5z1VmPeM3anc/42KZ3nfmdkb2TvbEZ/YVVNzoao0KNkMlZ1VwTTuWmhRlltYFW2kOSJ02UljRejhrTqktFXZNL60Q2F2kzY6fqW4wYbWuMHovNGMutx6puHs01jOp5SEyMn2O46VMZc1lvZlVjtKWyX0tq4lVGnsXGfHxOIm2r6pq1UkVd9Lf4jKnbot7XrspXrsnl4+0Q6bHqF205jsbWhlz/dbiJ1pxbXaiJiZwmqavL+qus8cus/rrk+CrvP1fZf56YfP3LPdC+p8dCdm/L2umxO719tjs1OWuG3L4JM+YsdmayN3r9Jb2hj5//94NXnrp/4Lef/f51L334G5bd/1YXdLaseWFnYec4efdt8DfhvPAUHQAAAAAAAMAvOUff2K/sbFp54n1euvE9px0a2r+zt/nqi346+MkLf9i95qJ/G7nukp92904cVvPGnk4rBoVql//NONPOoqa9kvWfp17k5kiaIKk0lPrqte6IuTQ+FSaRyPvuatQybLJdJDn9a50qPy9tfWoZQCbN7cfGWJPzKDem200Zc02fOarMLJOOVcYPSX83uqq6vC4pu5koR4tTxfVqnI4dffLa7PpsTk1b09731d5WXi/7XKN/rfJ3+DSuuqaQ9rdxtGwxmlPrpb1IY6I807R5nOaWc1GudS1fZx+jv+1Ir/3c+03udxl3TGLtflTJmuhTcyMmNecm7em6kaltP5f300+Grr345yM37vrp2L6L/mTTm066fukj77Czs2rZA+RdeKRfQsaUAwAAAAAAAPgVYYFosLN80d0WHrv5kcuOv90zFz/sjk9e9rT7nrl690mfGLh64meje9VoiCeCZiZ7IyozuvLrepOzjQkXplyYEyk1KKxd6kdNaVq47OurUW5MD6+zsoyv8Y0ZouZRxJpp5kZS6b9fymaySH08uVfMHBnLZLG7XBprdV5vkpz+N+RcaWaloeVqzJvM3+SJNonTuaTSfPL6iDmw29R8LTXMumK8ZT8/L23WnvmkbPK8un45Ts7TxrPzuMaUxNnc9TrtWvWava3E6/VnXcwjx/b+sU6WR8+93Y7WT9eo3U/Lnq9fEafrEuUc2yQxnjfWWXNYHq+z+pivlbV/5OnPW8as1jLvAzvPcuT119HLeQ/aPSnlYkxLeWxa3jMqWUt936j0q6359daxA1J/UHPJXKb39LrTu/XXXX809JEzZtaf/4jXL73b+PM665Y+UN6f9xHdSqRfLT8SmHUAAAAAAAAAv4SsWXibgYesfNFDtm/43ed8dP1VL/36+jed+qUNb33Jl7d85IK/7E5NHlbDwb+eqL9CqX8/zk0bf4rIDSs1JtKcs3OTGhJuqHm7m0VpuLW0v6krJkuce/8mjysNJJmbziHPo18xW+aVGjBq1EhZ5qLlHNdMH5XkdFOurz7OU8VYq/Km3AjSuLg2KZv0b6tZfNWnGHASa6ZbZcxFm61Rv+qnzOq5aZtdm57LsW4L1dfkCmNt3rZG3t/n5PmjPuaRqtc352B11q9S9LVr7Wv3OtfcOaikTVTqcsyoK/FVzpaybxm3KdtrpG0W48rcphmVx/h9Kn3SdJ6W+jTmZE2br7rq35vzv8Oov+rqf4dRr11e70OX9ob2Tv5s8BMX/t3Qpy7409FPnf3VLa97+kdXPuUeb1gwuO4kea+O+lt2Xo70dB0AAAAAAAAA/AKwWKR/t2pMNNpZ0RlY3F135+UPuf0TV7700XvWvenFh7Zcs322e/2u3ui0Pk20u5cmjT4VNzSzbdbNOTdvTGqImQEhMgNOjQftE9L2lpmmZodIym54uLLdTZW+NsvheYtBF0ojrjGU/Lypl3IxWkLF6JF2PVeDpW9OacrludVpfHWusnnXfS1vn6xNY+LaJY//6IH3tRg11yoDyg03PdZPdWX/PhUzS+dXG1RSb2PGOFlfS/ukYg21f5H08zbP26iZk46R9TmPMh89ar3G21Hrmr42L8vh8a2/ORftbWMuZPm1X6Ny7RFbz6tVX6m051ytXo9e9rWTcmsML9vXZFV2rjESG216n/q9r/es36tmzEXZNSljuDE3tE/eX/v0bzJK/4Pymh/cY7mHpuT1+Mqlkmf79zZPPvWqpfc+9gWdlYvvKO9f/RXXTaK1In4kAgAAAAAAAOC/Gf0a2019lW3TgqF1j199wt13br7oiW/ZeOETXr/uFSe8dc3EMz+4/ndOOTj4sZf/3ei0Gw5qtLkJoebCzt7w1OTs0PT22TQg7CkhNclUEmtPB9VGhJa1r0kNi1AaHcUAqcyZWqVP5pPYmIvPocndnMvRcjZy06XJ62aXGy4lTs2UGDNVm1R+rnn0WOWKY93Xc+s4laEWbXW7mnDZ3+P7FG3eHl9t1Zx9ee06JN6vpTGObN5RN58ZateT5xbXtBXFOG6u6Xgxrir7Zb31qdptTlqv8jY3rbRftIlsXpW8Ltap5IuxRGXOkafpFzFSX14rq8scjVqvTS2bnyryz6Nct/ylV5uP3pN2X8acqzp7f1RqvtqtxpxLjW415Yb27XBjztYuXuuDl8o4u3tbrr7oJxvf87I/2vz2U69Zd8kT3rb00be7+BbHbDi9s7DzYHlfq0l3JHiKDgAAAAAAAOAXgGVL7zD8kIHtT3jj6Odf8WfD+yd7A5+/+PDwFy75Wff67T8fFY3ojzlMq/nlap4Yqw2wtsxw0D6ifLLNZH1UblTUplwaL2pCmGmU0vOsy7jsmzllvPZ8cpw4lz5tuVGV5ksZI5T1LomXHMV8CfXXNbmyrDFylLaSu2WgeVvdXtqkb9OnT1lvbZU5J+eqXNO8TlUx5kq9xMi55ajmUV/bEZU59Vj1tXK0+bW36/NarN36xVwzR+mrcXJNesxzU8w58uV1uEEYsrqMj5hsE/W/Vqlc2/76oipHI633Nssr92Atvyc9Tr/GOteUk3q7Hv+at6rc1yLra+P7a2vXVObpbcMzO2YHp7YfHty7rdc9NCnvtwv/cvBdJ88sP+Guk531y+4b73EAAAAAAAAA+C/gpp6OWyW6/cINKx+67G7jz1jzxHu8ZPWJ9zt97QWP/51NV7zkayN7J3pjMxM9+xtX+9XUkuO+iZ79oINoaO/2w0NT22a7Uq9PjTVfvVMjoTk3Q0GNBzXl9Kk5LYfZ0JTD1DCDwU0Iqw8DwowIOXdzwuPTxPD4ejzRjBqGaoj4E21pzKWJV4yNyK1yY65pq+dh44bsKau++eYY+fRcY/b4NWRcGjc5dj1+5s/rLvUim7cdd/XGDuzujdnTcXJukrZSVrmJ1fTXvnVOn5/J5uRzyzXNGD2Wa6nkeTTeldfkave3Y9Q38ZmjUTEONS77lDZRmI3Zz46WJ2KlX7km65v1PlazfjG2zdlfc4/3vJY7xvKnFdUEi3KMn3kyfzHe4lzvdbsPzHyOst4fGiMqJlyWSz8ZPww7rfcfglB5jKsZPyXvv9nhfdtn/UdLJnrdA/J+tL9HJ2U16D5/yb90rzp9av2Fj/mtpQ+6zQsWDq5/ZGfpwofIe//WoiX6QXAE+JEIAAAAAAAAgP8f6NfTjvwVtVVLbr38uNudsmnnU9++5b0v3b/lg+d/d/OVp3974L0v/e7oZy/8p62y6de/d6V/38o2/GHApTFXpHWV3CBzI8ENB+3jSjNC61sxtdQETOPDjrskTk2TqMt6NVLy3MbWfD4HM0RkvDTnyvxKDjdhTPH3yspXWQ/oUesytq1i/uS5/S29dpudRz4zcazc9CkGjxpAEVfaqvbSJn29Ts2h2iDSHNJWTLmmvumr55LbrstzpfqNOWuPebSupahpLyrjNOXW+NHP8mf/qt3nq9enc4icqYwxs0yPEV/yZJxei8Y253WO1hgW52qusakzhTHXqhN5Hlc9Rj3vOUZdv4oBF6Zgkd6j2u4qplxK2lwyfq6lzkEl7WkA2ns1/ibdmK7X9Zf8dOiT5/3DwMfP+/7gJ8/75sDvvfCz68582FsX32rLs+RTYMg/DOagphxfbwUAAAAAAAD4D0A32PoH4G/TWdI5trNmyS0X/sbo8Sufe5+Xrn/tiR/c+IHz/mL487KJ37urd8zMrt64bPTtD9Cr4TQzOTtqRlmoGGxqIrjM9OkzKVLFdMh+Gm91aiqowVHFmPpyphFSmxC1st+MysewHGpUSNmfmNOYqC/9fHzTgV297gGJ13o1Y8yoy7jol2PkuY1XnUtsbfLYmJbDr9GMHKkva2P1OlZjsukY3u7n3i6KuVidGi3R5pJ6m7O2NXlVuXZ6Pf5jCXGsrq825nJNmutwaXvLuIq6RtnWjF9iVRFX7ger8/qyNqVv5qxitd6uz9ub9dR2j7O5WtnjrT7afKymb+bU9jnXWtr6FHny+nzOTVu+rjZWPPVW6qu5WEwx5mTMlNxDRXLuZlxzn7mZLeOWOeyWc5WOL6+ZtDdm4I7e8Iw/QTcmr7XpoPSR9evu1Sdepf26i/9s0+ufceXSR97p+Z11K/VHIrqiW4kGRUd7ig4AAAAAAAAAKm7O180WLOiue/SaE+/96k2TT3jHut1PeeuaV534wXWvf/71m6962R+PXLv9J+Oyyd8qG3w15Wojrisb/PyKqEk29UV1fZoQRW4iuFHQKNtLXWVuuNToatqLEaK5wuDQJ9TqmEZhZMxT3zxNpP1Vu0J57nU2hhpdetQ66d+YJnle10c54vN67Dyu369Xy5FXzyOu3+BxeZwZSQdUcl6rjo84H7uRrbPEuBkVayh93dyqcmh85PFxI2+fmvUS6blK7oEsN7Ht/LamUW/HMp63l7lV88y2WtZm1+7Xa3FVvvnmUcaqxva+XvZ2z5+yHNW5qeRs5tgvz6exbZX2vnpXkzdl95Ssa30f5/upnFvOyphLIzCk78futLxv5X1iOqCS+kPS79Du3tjBS3vDX5j88cC7X/btLVe86Notr3v6FRvOedgblh13m9d01i97hnxeqIk/HzxFBwAAAAAAAHAUdOOcm+cFWqEsPmbdndZe+ITXDU5t+6vhA3t6Q3t1476rNyYb+/I3tHLjb+aPm3IqM93kOLZfn6LzTb8ZBWEAuGkQcVZ280DzFcMi6trtIYuJsedRGhsll+joxtx8bWog1dIcKjWO6rLPJQ0gG0tjTd5vbl5vz3i/njjGmvo1x7nVZ1xeV1WfsZGvmGn2uoT64t1o8ro0nXKtm/WLXOU883gO79POWyvXx6TnKjPm+tdF+2duL7fbRTZHb2/NL+r61cTIuRpMJTZyVbnLWNYeKu1Vn+xXx4nq/mU+pb+vUT0fiynXrP2qWCnb62HnGac5PE9d1xhrsp5pzKkitjHmvM+ovneLMZd9td7fi/Ze3T/Z6KDUH5IYfXLOdFlvTJRfjR7/xp7Dx351299uuvjx71q4dfMj4qMj0c+Tm/MfAAAAAAAAAAB+pbmpDfJi+d9tFwyve9zyh93xtJVPvddLVjz+Hs9ZdeajLtvw5hfd0N27q6dPxo3b19p04y5HkRtSO3vD0xOHh/ftsK/AmfbJhl6VBl0tM8AqpSlhfbOseT13y3jZF8ZOqBWr51kfKkZG5pXx0mTL8e0pv3Le9NV4N96kvp6btasRo2OG2ZRtJp9PY9zVbU3O5hpi/tYeRxvD481AKf2qthLjr4Pnyzxy3WGmaXu+XtlmivmncZNrZXEl9zzKWCkX80jzR6726+WyuWdOVbkHpH7G5bHRrnm1j+Twe6xpK7lFddnP4+/JmWmU9Zor5+X951ynqP165LFPcT3zthXFOJbXz0u+GCvXz+ql7Guj7T4vn5ur/qppmb8eSx9vszwaq8eqzupLncTbHDN/1OmTc5bb8/p7dVLq5HhgR294//ZZ+xqzGu+6viJ9em5c5Dl39wY/dP4fbd7xzLetevzdT1l8l9GTFmxadbJ8rtxPtMI+YeaS/zEAAAAAAAAA4FeeoxtzyzrdFcfd/qTNlz3jysFPvvx/bf74hX+98aqzvjP4/nO/N/LZ7f+8df8e+zVPf/JKN+i6GXelkdDe1KequpmQlBsjLFQZXx4fOfVYldNgKIo+aXC4CaE5m3PtZ2WJSVMujQ4d2+fj5eyTMjOjHsNy6Dw0px/LucyxFWexcTQ18/dr1rjs1+7bX5f9ctwyx1Z/z5/l5nVJea7mKHWSS9fbc8a51Oe55/J4Uxk3+lu7H8vrVNWboo+9Vtk/6yrlddl8yphVnj7NGaeqK/UlV7vNr1vbPcbHD2m/nKuVG5W8N6WSK2Rr42W7v/UY9aq8h/zeV3m9tc24NJ/3i1ya32LkOMeQi/pSrlTmqGpifE1y/D7le15NORtf5b/wawbdgct6w1dP/OvARy/8m8EPvPzbY5+7+NvDv/fCG5Y+9LYT9rcp50c/j/h6KwAAAAAAAPzKcbSnUJaKjuksWnTPztrlv7FoZOM9lt//mMesetq9X7xh4qlXbfrAOX8xfN323tj0jt7Wfbt7W2XjvTWMgNZXIlPSNpaqjK3yN6riXJ+Qqp+U8/Ymzp+yCxNA69Mo0LHTMFAzQs7VRGhMiHjyLHNITHkirphv2i+lplyfMafjhXJ+md+NP7/OzOEmird5e5abuNK/5PF565jZx6Xx+pVg76uxFmOxrjRtctxyTaI0SUr/kjPWTiX90mDRMXMdvV3z6DiRz+Tj+tjNfO0pv3idPE/m8Njyes1p89xt5Rh6rZo749r9zQiK+6wocrTGCJX523k1jsVpn1CJ81y+rnGufct59M1fdpXXysvNa1arzKca1yTlHNty12WV3o8hf720j37ddHe87nqufXKM7Ou57X6e078d4/OTuWve+fKY+efy/qkc16/bzDi9Z2d2znrZv9o+tv/S3rjllz5TE71jvrinN/q+M/atPvHeZ3e2rLxjGHR3FumvuS4UzQdP0QEAAAAAAMAvPUd7Om5g+Z3GX7Ll5Se8Z+B3nvfhjZc/630bX3XSpzb+5snXD1515ne61170b/YV1Pwaav7x91Rs5luyepFu4E3Sx+TnbhioUeaml9X35wg1ZkJI6+p6KfuTZ24Y2Lkc/Sm5lJ97XfaN+GzTOWW8GXqh6OM5XcVcybnkfOrzqGv6ZT4p98VZTpWem/HkRknT168vy834UmdxUg7Vhkkdp7/YWgy/HE+k8Rlj663SPCVnE2tzjfn6V2/TmPPrSrVMoMxZxmjGKa+HlOt5qNzUlLYqj98TnqPEVWOVNrnWVr7MISpr3a/IX/pYv1TGNW2t/Hrev95RV+bUN5aNJ/mK5sRp2ZUxdV2Jqefc169R5kx5vd0nZpyl+uIrc8/+5lxK+4bSlHMzzuVtO6XOPy/yl1zH5Xz4oxf87fA7Tj245dXPvmrLtqdcsfbp93/9gqG1J8jn0HL/OJqDfnbxFB0AAAAAAAD8SqA/4qBPyJW/8bT8/rd5wqbfe/Gnhr7+6t7g1K5e9zrZTE/t7I3LBnzMNuI7zEAb1l9m3Cdtttne6TIzQQ0GNzVaMtOlUXkirmz4PW+aAHP6h4pJUKuul3JtzJlxpO0Wk2ZcmEjZ19pSUi/zyB8fsPmq2ZSKODfFVDFWmiGqGNuU+a0t+0puK+tRx8l6PWq+KEc+NUv0WOrniamVMapittRGXDmPPhbvMhPF6mPN9TxUx6lybX0uvl7FoBO14vV658jHt9dVVPcpxpS26dhSzrHqfE1cxFTzLG1qiIUplnE+nsf7a9GnKr/FW30eK0lbO28oxixrHgZrmVPJofGaJ+Zr59qeeas4rctjyZOS/qrsU/o1fZvzfmkflc7P1fSp+qUpN92U9Yda7ChxqjTjGkOulsanoS+fKbIexxzY09sqa6NP047dsKc3dmjyrzbuOOFNC2+96YHxkaSoSbdYhCEHAAAAAAAAvzTcnK97rV8wsvEJK0+4+yWrnn//81c+535nrZp44pUb3n/2n3Vl81022bLJbzbavhnv7p04rGX9hcbyVdWU9ZFYlcaL0thKE8aNAzViGhWzyeK9vxsGUZa8aRD4H7EPI0hzWx9XbUwV88jG87ZmDO0b7aVeclUGYVHGlGPOzdtcXteYI/Pk0b4hH9/nYPn1qPMuudtyI6np7zm8rWW0pSS+XsNGfSaV5rFx5bWS85y/v1Z+nnW1fN1Fem5zCVm+PmUfKbti7OyTKvExrp03cc26i6Qt78l6jMzTzFuO1ZOHLo2py349fk/NlyfiK2V+nYONX+agObM9zLgjGnPeJ/NbLm23GI31Nlf0DZX3Y9Q3a9X0T3lsnldjV/09rm3M2d+v06PE2XsuZMac9c98OYb0T820pcb+mPTzX3WVshpysibjN+zpbb1BfyhCxpXPnc1XnvG1DZNPevPq59z7jGWPveNLF91m4LTOwoX3lc+rI3299WhPAAMAAAAAAAD8t3CTG9VFtx645/pXnfjOzQd2/9Pmayd/rH8/bvCTL//77nXbfqKbZjMDbBOdm/c41811Ss7nKjfpKu9jm/cwGIrpUYyrMHgyLmI9Xsd1lTYZN59oaz2ZFX2KaRUqRktVbzFhinm8mg/zzKOS9/FjMT8yt8nrbK523je/kI+nuWL8Sk3eucp+rRwyb20z06fPnPN53AxpThnbX+fmNfJ197Ip4prr0PnPXf9yHrL7INuKdOz+OlH2i/Zy75S6SlLXf3818jHz3vG1qNbH+uhRz7Uc6yllfx1Ckdfmn/HWp8lruS1fxs9tN9l7ytXEhqTs1+F9fY569Pa8ltIW8nzZHv0zR5ZNPsZ88tcnJffRvj5jTt9vEler6e9x5f2uOfqMubGiMObyc0Pmq8bc2EHRIZnHwT06/mz3uh3/NvTJi/5h4EPnfm9k5qLvD3/8jC+veMJdXykfWcP+yTUH/azDmAMAAAAAAID/No62KdWnSdaLbtNZvexeC2+95fiFdxq+78K7jz5g+YuO37HuD87+P4PXywZ7WjfTssk+IGU1jOQ4NDM568ZAmi+66Y6Nddlg73DN7PA2UbNBd0NAy83mXfNoTjVAauNKj2FUaVnHNKNBYz2fmwBRJ6rjNK8ZE9Y/zAMtWw43tbxPXdbxm/hiyES7GxUyZjW2zy/iVPk1Xnsiyo+lzWLzGlVS18oR16jnMoaP7+PZU0tRzrmZCaeK/p7DY6xPZc6VuWtMxHp/P69fD1+36BPXoWN4WevjvPSJPFof+XK80qZ1dtRzrc9xVFInaq4hFLEmi4m8mUvLoczRtDX5S7vWmzK26a+xKY+t56OvhZ5Hnsxb+usaz13nvEabT1W29uhXx6vKa17itazxVTlymKyfHiOP1UV99LF+lrPp337KLcbSsSPG5mn1kbf0lXWQftY/zzWf9G3dr2rE2fXp50hIv/ouR/8KvMs/N+Qoedyc03w7ZkcPSq6DksvuuT29MdPuXndqUjTR2/L2539h9TPvfYb+GE1n3ar7d5YufIh8pumPRRzpKToAAAAAAACA/zL0b8UdyZxbvGjjmruvfNzdzlw38cQ3r3nLKZ9b/frnXbN217M+ufHNp35l8HPb/kW/tuq/uKiGnGs4Nt79plLzww2V1JQLY661ubeNeyrqIo+bIG5a+ZNjbbWMOVVdlj42V81h9Zq3Gk/roy0NFpOWi7K9rZJDlNfh1xJj2Pg5XiOPa/c3xXz9WmUMPc/6EtfMqc6Zmr++6VsrzTmbT98cfS2kj8nbTSVXo/raVVZnsc25KfPl+ZFUxm3k6xH3lsTYfIu0LsaTcn1f1cqYEqt5TFmuYnT+KauPuegx4pp5ZZvX+RyaviYzYV2l3cp+nv1a+VPRXt9HJW8qx8u8olwfr9fX2V9rz6HXq2rHtsZKRUyJjbXIsbTd3zdSX0w971veT3XO6GPKOpHlK8ac5E9jTs91vJyzmsFmCMf7Xo96b9i8dvQ2f+jc769/wykHNv7m8z7RfdcZUxtf+YSPLzh2ywvl822lf8wBAAAAAAAA/PezSLRatE60RLRp6Z1GHrTq5OPOW/eGUz638RMX/d3A9Tt7Q1OyaZ7ebV9TG57e2Rvcu312eGpytjaPdONtm+vclOumWtrdgOs346LOyirZdMsGPU2JVp1t3HWz7aZMGnO+Ic9NeZ8iTylnfzlvmVY255y/tu0KM8qVccWYq3I0xlhK5lqeGIt5q+oxsq6o7h+SOI9vm3Mqbc8xm/HbOX0Mj+lvc82td2Muvo7Y12b5bE5911XGcNUGj6mKr5XX199/jjJOJfdSsw5pvjRj2nw0Ts9zXG23mPb4pjQiTZ7H++e1VbGaL1Tmo/EpObfXprRpfx/b1kz7Zc763rA2H7vO78emjylim/wuz9WYqn7NUS7tcW5PpcVRX+uoz2vwso4V4+h5nzJfmUPE5hj2vpHPCDftmzz5fmr+/py2R47oa+f2ueHnc56YU0m9fe1d10znbXPwzxEz5eTzpTszacf8kQl9gm5Qvxb7rd/sbX7DyR9aeqehh8nnnP6HiWWijSI16vRzEAAAAAAAAOA/jJv7N5QGFm/d9NS1T7/7tg2nHXfJ2rMeuXPNxU94x5rXnXTNpg+f/+cj+/RvOOnm1ze++vW0YdHg1PbZob0Th/XvttlmumzCU1IvRzXfvOyb5qxvziM+N9k2VqicR4xusnXjrX0l3swxKdeGzfznlaJfazxRMQsyb2nL874+8dVNN62a2Kwv5yqde/SzXHqe12jlrPN6Nxz6VLW7qaF9tFwfG2W//Ippo2ZeHufzqjXXfNFj9Nc+omIStaSx7bqjXktdruta9XGP9N9jGSfK+ZaxLCbqpNw2tHbLa6SKawtlXJMr6kubnnt+M4VzXvNec5Wj1MlRVb8e2SfPRRpnsdlW2pucHiO5zeRzk82MrHLNKa0/gsKYM+U15PhSLrlSkbP0qVXP1XKJ9PWaTxHX3KMRH/ltDG2T2HxaLutT7b9HqdeSc9Z+8rqY1JwLg05fKx1rZqK3+X0v/eNNu578znVnPOSiFS++/2WLH3jM9s7q5Y+Vz0E16AAAAAAAAAD+Q7lJY27pHcYesvHVJ141MLXtbzd/7uJ/3vLpl/9g8NqLfjR03bZ/HZrZ9tOh/dtmh/fr11TV1Oozufb5xrdlsMWmOsu5YdYn5Tw22mPT3erXt1Fv6qJd+psRGPG1EVXPy0yTYp5IXc7brkFiM6dJxhHZ02eZ12Kyfv54N1iirHPN8rzKWM+R45Rr1DHruohtjRttLQPRyvWxkffbZU9oZY6mvjn3XG2lGdOur/pKPzVFcu5Hk5lMR9I88aabaldFTGMe6XnM02J83m7eZLzUHcGUq/O2rj/62/VaDlmzuJ/KOHV/jS05tJ8r+9rfFLR2jY2+ZbyIsXG1PWRtHmOyeqlTY87MuawPlXM34UZCaciZKad1VvY+dp0xr7zeksdivD7n4tLzmIvNKcui6n3ebqvPs65RuYb4fHDpeUjaylNzJs+TT+Dae94+Z/SzaUK0XeYoOjDRGz00KZ8f2/69e82F/zx87YU/vOU3L+/d8sZdf735/Me+qTO0/n7xkQgAAAAAAADw/8TRTLilorHOskX3Xji05uFLfmPrExb+xtYHLbn/rR+96qzH/Oam95713eEp3djKhldNi0M7e91DO3rDchw+6H87bnBmctZNqjS43JwoRpgZZrkp1nKz8bbNc27Srb06j/5mEll+kfQxo6ycN2OZuRR505AalvphyVvHla+5Zlnzm7JdNvOhNBq8PcbsU2OaiMIQmdMv5uM5mnLGZQ5du7I+2SZ9XVEvsr5aV/pGzpJHzvVY+tZjenzT1+MstqqfE1+MHjmXWL0n6hz1vBtjxGPmW69G1WsTOby/jqXStc2cki9khpC1V4o+ZR7W18umyK/jer55cqgizq9Vpf29X/ZvcnifrCvrpn1Cc/Jb3+hvfUVx39h59tVzjalyeF3Tt/RX2dz66nKeWm+mm9alGZcxashFuxw9R5SrcXze2R7XEGW//swXZesvZXk/p3mW5/kjEL5e7fsv19+VeUKSxz8fPI99vVWOdo0q6+913i73lX3uxNE0Ke2TMtaEGnKH1Zgblc+18Rt2uaQ8sn+yN/Dhs7654RWPfdWy42/3zMW3H356Z/WKR8rn5W1E+lXXo3FznkgGAAAAAACAXwN0g6i/pHokBhfcbuiENWc95PJNb3nuhzdcdfrBdb9zyvSG1z7v+oErz/yj7jWX/Mu4bHT11w7NcDooG2iVlIfjybDh6R2zzSY6FRvhshmWOt082wY6pJv3jJcY+yqqxpa+Ui8xaXS0pPWpqPNcvsE3SS435ZoYVxpBnj/VetqpVe/9ynxCdUwt/4qorJWe23j90vn50fvkURXjt6Ttbjg08XUfl+bs71PibExX1hczReLr+lo5Z41rP9E1jzKf9m21ZT7JpdeX/ePYb8qpisFSS/NUMRbXHzNPnF9nzCHrrc41Xw7rF31tLqaom69N+2ROaXNpHm8reSs1eaK9GJ/aFsrzVl2MbYpceaz6NGPpOFWbHLMtn5KzeK23uLl1HqvSXLU8n9bX117mZv2jLpTnRTPz3Zt13ywfXfX12TXG501+9jSSe01kf9PS3vPbZ8cO7uyNHdrZG79BdOMu0W7Rnt7I5175T1veddofbnnXS24Yuuacb21504mfWnK/Y06Xz80R//g8Ivq5izkHAAAAAAAALfTpOP17SZtEazoDK2+/7DF3PnH1nqe8acMHz/vDLdftPtyduqw3vm+PSza5Y7LBHlNDRs0mk2x4i/GUm2A5tw22b3jVVDOjxc61vq2yqc/NvNT50zPaV877zLFiDlUqxpi1xaZe22RO+UTacCj7FxNPj7Yhjxx27vKcfg2lzfqneRDnGXtE6fponM4njjZHPW+riY2+uX46P1MYDiW2L97MLr8euyaVxthYHlNfgxkZUmfrX8er1BzS17gex9pFZszFPLJfrexj4/S3haRczLnIXV+r3xfN9bYUMbWatiauP6aOK3U5H1H2zfnnunifiI8+Zji1ziWm9Pe6Ei/ytjxWY+i5KXKobH1D1r+JN2VeO+8/Rjn61GOVeVTKtjK2xblaxlxVX5dtftHu+WJsk7dZXptz1nusS+u9zX/0Ie8zzRV9rX/Tt1aZo5WjzuKbOlN83pTPH/vFaLnfrN6NubEDIjXmDu7ojR+S4w3S78DE4fFDu3rjB1V7/GnhaYn/8q7ewDuf9+kVj7/jczobl9xGPlGPkc/SIdF6kf5QDj8UAQAAAAAA8GvIzX06Y3zxnUdOXP3c+29be97Dd6zbfsKb1r762R9f/84zvjZ47bYfjhy8rDdy4NJetxhvsnlVs6Ay4rK+tKvKJlsNFt307jCTzSTn+XW12rhrlOZMY9BYrEpyl7Ip8pmafu15eWzdbjGqyOeSHBGvT/+ZtJzKfhajxkGYB3oe9WmieIzGy/zLuc+nZYaF/FxiTFnv5/X1NblUu3yszGG5U97HzDmttzafr8dKH8tRqxo3z1vqi63k83D5GKGIb+bVtNXrZDGmnHfWa+5q7OjXmC9eV8vaTdo+f0zK8licnvu9ZtJxbQyNyVg9enxLmcPmF+ei1rm1R11KY6zdjSM7t/oqxpT1GhN1MaeSQ+ehUhM1y0VNnMUeUb5u+iSsPQ2rKm1hekldM2bUxXnr+uzc83m7ly0mzo+uue8Rm4eN0fTXz5HSnm2mnK+r5ImvvNqc8uuvlcoayOeb6aAadKJD0r5/4nCadWrKjR/c3Rs/tMeOgx84+3sbXvOsj23c8eQ3bZ584ltWP+0euxeMbXphZ2Hn/vIZu8Y/aueFp+gAAAAAAAB+jVmw7MG3e9LG15783s2fvujPt3x6+z8OXT/50+6BXbMjMzt+1t2/8/CIbDq7B1S6MZWNrG2u58o24KYwVczkaOSmnP4xdTde6mPLmNP+Wl+p5Ch5a4OnHVvqNbZs0r2ujrG4uj4MoVKv/cyck3Y5d2Mr+3nedp3LjBFRnmvOLJd+Gp99olzqTBJnyrZmzrYGKTk3s6jMx3N7vCtjfe20PaV13q8ly9fkbJT1tep6KWtfG9+PTS6vS2Vb3jNNDp2Tv77ZtxhLqewr11AMoYwvbZl7btt8svxybM8tjJ5Wfh1TjqVeFHEWq2OWWI3zc5untXtM05791TySctVvrrQ9YvQ8Y6s+No+4/3JOPq/sK+f9T4+FMneum6nUazlNrqquGF9yXslylDk1c/TcWq5yhMp1tBRxMabF5TjSbveTGmoaW+qrmDK/ypjT2D4zzmTj5Rz9b2e6ORdPz9kTdDtkbVWTIjfuxs2os7w/H53Z8e/dqd2Huze+qje+/9K/Hfm9l1y96ol3u6izftl95LNWn5wDAAAAAACAX1Fu6qmLhaKRzrKF91swtuFJKx5wq2ctf8CtHrficXd7/poLTrhi83te9ifd6V298X2XymZzT2/0kGxmQ/qrlEMzO2f1b8YdaVOvsq+37pNNa1FuwsNUU+kfVtdfZtW6MOOKYZcx0qcYXlbWOm3zXGY2aVuce13EV+0WI5tr/2po5NE2NX7KucRFfROjR61XSV/NYaZknEedmXYlzuVPEsp6aHvUtWO0PmRjzyOpb0yRpj6v0a5LzyPW2z1nHVPWJvpbDo2JsUtc1e4xetR8efTrbOeONptjqOpT5qfSOVTzaPLLNUo575+S1+bdrEExurSvtkUea88cVteU1Rz23NG3kvbpr/N5ttUYatGnism2Zu6paLM+laKfqr023q/kKteQ5VCJ81w+Rrb3n6uhtNtUxzRj6XtY2xo1r0etbHP5HLQuYuOzwPtlu8uvN8o2ds5Rjxkn9TNybj/SIOdhkLXusxyrSOtcc+7HlLV7v+ybP2JRZEactJV+TVuONSZ9/KlB+VxL6ftbjlk3OjM5OzqjT9FJvD49d8Olost74ze+WmJe1Ru4evs/bL7i1M+vf/njXrfsMbd7WWdg3WPlk1h/zVV/KGKlfjAfAZ6iAwAAAAAA+CVDN3FH+zGHDUtuP/DoDacd96rBtzzvmi3vO/OPNr7llC9vevOLvzL4nnP+dOTqi380Pq3GWmw+ZaNpppxsNtWY8ydPdCOrknY71ueq2pSrjDk1wuTczBlRGkaput7K0sck/W0DbjlS2U/bm75tNfFuqtXGkrRruZxXsnrvnzH+t/MyRyra1awrhp1cq5lyWpY+Jq+3PBlT6v1oJkV/nRzL2lp9W818ZR2yXvNEfTEoTHq9fs3e1+evY+WTcz5uo6xLk7Fp8xxNbqlL1e2aQ89b86jn7TFFrZh4/US2Vpa/3T/HsXWyHE1d01adR/9UY8boWH4v1eOkyn0d+fx1qWNyfq6Mt/OYW6O6T5ab2Iw/Ulvd3twbVX0ca7XNuchnfdVE0/pafX3NlGrHZP806sq55XS163w9NF+p07KdV69nGHIueU1UEmv3YdXPpPF6LWU9I0cly5uvV46nqvJp3HxzLuuo69YyNlVhxkl5TGLVtNO/tylzmtWvso4d2mMaP3Rpb/yg6rLe2N4d/z7w0fP+ZuCDZ//pyKfO++PBt5187epTHvDGRbfedLJ8Jt/SP5rnRT/LMeYAAAAAAAB+SVklGhaNdJZ1RhYdu+U+Kx9152euf/njX7/l9878WvcLk72R67b3xqYme2PTk/rUh2i7bIrluE+0f4cZMmbO6YYzNvj+d6fUvPNNqR1D9sfUVbrhzWOobJptw50b6TTY2nKzScvRR8ZzI0fr8xixkasx5DJG5Zvs7K8alnZVGlVN/5DGaFuUzVwyU07KalCFOVdyRp2em3HTZ8xl3rwO2/RLfNvsknNrqyV1uqaqedqb6/V5WJ4Yq4yn0ra+ejM8LF7HlnMtF0Wc5VTJ2DbfKEdMy8gIpXmR/coYJs+XZksxXCxG87n8ycn6tW/ayzWX+0DX2ueX57Vaa5b5ZMzGjKnro63M2evy3rX7WNpM1j/qLDbrvL7O1S/P3eRvVLdX9dlPr7PkaOaT7UXRr67Tr2NmuRlfTbXadNM4f49bWepM0rcx9vzoc23Kzd9r8+twcy3iYq7+Osa5vR567nO1vvrEXNbJud8Hcp9EHutrc/A4e6/Zayttcmzf31G2do3TmJDmiqPPLepMmU/zy/XZdbtsLSTGP/P8P15o36wbl3Ual3gz5/Q/ZOzfMds9MOm/6KqfCTKPMbmm0b3ymXvdRG/sYxf86eDlz37Pykf/xmkLN617oHxO649EbBFtFi0TAQAAAAAAwC8oN+8JisWL77j0rltftO6Fx+1ev+0Jr1+355nvXnfZSZ/a9NaXfLn7qVf+YGxKNorTO9yUm56QjfGEbIhV/hXT3EDbRtQ2574BVZUNsko2prlBtT62KVft8KNudiWfH12NqdTIYsJwKcaLxetGWY5ab7FNf43PTXj9C6sl3uQbbX2qLf9O3LBJ+po8xmR5PLfXex834/zcc4Uyd0s5flMuRlTEpHlnppLW69ppvbXpUdfe6/36s58eM2fOX+er543KWlu/iK/kr0G0S/7+/tZWjZs55qpq0z5ytHlHm48l5yaP8TVMaWw9ftwHEasyUyfkhp2/7pmjrJvlijwlf0j7R7mJ9fhaue4m6eP3czWf0q/u3+Qpr5tIy6WfSerlWPL3yfoftb1RK7fVqREU7Vk/Ty6bZ3k9QlW7G3D9yvb+8yZHfgaU935dp9LYnLNeZ+kfddInTf1S15LH2vrascmd1+DX3qcj1ktOa2uP03xWNPK5ynVLP1fMQTQ24597/hmpJmYjfa92903O+vvc71m9l8amd/bG5bO3+7mLfzzwrjP+15a3vPC6gd95zgfWvejBly+559bzO2uXnyCf3l3/EJ8XnqIDAAAAAAD4hWd5Z3D5E+96yvq3PvtTGz59zvc3f+aV/9bdK5tC3UjqJlI3k7axjKc4bFPsG8fcIOsmtGw6U7LZTIMuN9fzbVhbP+LQUhgnVZ1/Za1p181yGma+YZYNctZrfPRv5DndUPNybrjtXMsyT99o+7lKjTk35/RcTaRdzdN0/X2tvTq3XPPUm/rLrsaYkzVKUy5yqMyo0KPW27n002sp/TI+8+ncZc30GnK+Ie+ncdk3+oS8v8Z4u5uS1VHrq3Gzf5HlVh2hPerN6MhcZU7zxEm9K+6BWhpn9XKdeV/ZdcsamnK9NFfkifaiyNOO72sPlVxa1jobM9ptjLgW6xvnIZ9D5o+4PpW26lj3P5KauBw3clq5Muaq62nnyHqJqRX99H1dPyE214jrP29ylM8AfepNVdepNLZar3wdcp5u5IeyrijHiXPtK8fMXV9DExuKunnjrJz59Z5RRX2+N+39pvONa49cOb5/Pu52qSG3f49p1KTn/llpn5faxz5fw6DTz+IZyTm9u9fdO9kbmLrkn4c/c/bXN7z80W9ecOyWJ8qn+Eb/MAcAAAAAAID/TvTJiJt6OmKos3LpgxfdYfjkVY+864uWPerOJ6567oMuXLPjaR/a9P6zv9+9fns8GScbwn2iYmrs6HWnJg6PTG+fbUy6Pmm9bigtXsqq2LQ2G3Rti7JtWGWzq2aGSH8psUjrRX5UsyNj3PhIYyalc7TNstbL0c+lTY+WT+O8zevjvNS5WVfMNzXlKmOuP7ZVF/Ixw7CKmNy4F2NONucZa/HW19uybl6pOSdHlZ7nOroREPWa13JnviZnuf5Yr9Y6Zqz197LHuqy/5Y/5x1itNj1mn1KOsUQ5j5SZYdq/KOdaz0XP677tthxX1dxb2tbObfeZ1aeq/CK/F/r6WD4dr84dR203ZYwcS130V1l99s12j/W8lcqc5DjTzM/HDJkR5uPZmEVh9PS9z3Jt2rGq2jBrYnLMnIfNyeqjHOftPG5A+Zo2f0Mu29Kct3irr8oRa/H5N+xKfczB4r3czEvWRd7T/h8EdI1inapyuZ64Pp3LnCcFo1zPKedjX9vV64mYfN/meyXv/ZI/3+PyOpRxNE/ElzFiXdSMM2Nu3x4bZ8ykP6Tj0v8A0t27fXZ4atus/9mA/IzVNd1j8x+6btvPNv/BaV/b9OqnXrnm2fd+xYJbbj6hs2Lpw+Qz/k4i/dMER4Kn6AAAAAAAAP6TuCljbsnC2ww8ZN05j3r1wIfOvnHzx1/x5xt+7/Q/2nTlS//3wEfO/evutRf/2L6eOj1pxo2bKvrLp7r5VKNONqCiNOD6Dbr+zbIZc7FpTXPON8mxSZU42yCrSaTjSV83dFQ6tpdrY85U2htp3mIIRdnqqtg0bLxd6iJOY8rTb1n3f6N6zAMqN6/sXK85N+26UZdYk7VFXzvXtWlUzLIiXUdf10axjlHO/E2+GEOV9bEO/vrOM7blaObv15DHbG/K/UZk5vSyr7uvvedp5hR1LUlOyWUGicZJnc8vlfU5h0Zu6OhR2xvNV9eWzMPMP18Xq6vz6nyK9LzO6fU556w3ZR+rD2WdyPq22jSf50z19zmi1LAL066RziFV16ua2Po6m+vWsqv0mXdOYTJlvRpZFpeGlppIbs5ljpbK58XctvL66Nyyvs5Rf9aYMqaZv8bbnKOsc822/rW2uDnnjcnm78eYU2jOvWnv82qMUBlf8ukc8km5tjG3yyXX5J+l8nmr/3FEj/rnAuxzTz+DNY+067ymd/x88LMX/fPAxy/4+4GPnv9nA+968dSmix//riV3G39JZ3HntvGZPx8YcwAAAAAAAP9BHG1ztVw0KNraWbfiTkvv2D1u2cPu+OQ15z7m8o2/d+ZXh6+TTeP1shHUr6xOyYZvarts9Lb3ujMTosne0NS2WT2q0rzIJ+hMtkkU5SaySDet0WaxuomUOtu0+ibWN6oa5zKTSDeeerQNb2OS2ObX6jVGyxrf9Cnmj+QrT8FZWWJjPDeBNMbbsy6NpHxSzp6Wi/4to87qwoCq8vtTNFFX1fsmPcZPU07K3i7lopyLSPKYIRBtPpYr69LgtHzSx9ZWxk1jIg0F62d5Pc6OOcdcN1s7r8v8OZ8SG7l8Pk2c1WW9jZWqrifXOsdIlXat1xh/nb1N7wtXuRaL9XFyfG/LvH7uZoWrjnOzxuOtj117jGeq7lGbi0jmnmuR8zFln2wrub2tiYnxop8dcy7WpvXt9iz7XCKmxGp7FW/1zbndExaj534tTa4s65NrYZTtr4w5m6/EaE4bQ8+9zesyb5xXZlVjzDXyuWV+N+XKmJrX5GP6j0DIubwepa28Nh6n86+vzcrS5nMWRY68BlWJjxjLle+daPN7W+8h72uyNr3fmjFLvNV5rlwn7+9qYtOYm6fO1iWV5pyuS65N9R85puUeVMn9qj+u48bcZG94evusyvPnfSrzm97dG967s7flU5f81cBvPf/Da595n3MWjm18iHz+66+4qkm3VbRadCQw6gAAAAAAAP4v0aceVEdiZPFduk9a9/zjLt50+bPesf53X/CpNa95zmc3vO30Lw998uJ/GJ3ZI5s/2QzuneyNTsnGT0256YlQbApbm0NVbIZlM+jmnJ7LUdtr9cW2N6mVpL1W+cqpbThFUWfGnLVFWWPkOKzGoZZFuVFt/g5cv6Jec1p85DFFTB6jnOacKufs/fuluSO/yI0z38y72mXPpfONOsnhZkMtb69VzBTLH/XW19c655j5i+mg5xFn1x7lpi7KlrPOodfbXLPltLyNmnm0x8h1NUM0ynV9fZ79fA7t8XNOfi2Vsi7HL/PoO6/q23PQcX1sv5/jXtX+R5O8zpmvzl3ObX5+Xt/75f0Qbdne1Me1aFvpF+cZU9ebfB1qZVyRnft9Y+ZUlkOZX+ddjK1STkNJY5s8fo1pKKnptsvUmE7zKXPWivmZvK6eSzMPn+PcY86lic/2jCnKuaeirqUq35Fk8+mL9/uwLbv3ozz/eoRyHdOgDFOulFV6n8rnnX0Oy3l3avus1VX3sM1l/6W94c/v/Pct7z/3u+vfftr+zW98/kc2XvzYty5/yl3f2Blad6b8m3B7/6dhXniKDgAAAAAA4P8R3UwtFC2xM2Xzklsue8Qdnrp+15Pesul9p39zYGbycFc35Qf3iHRDqBu+MNGm1ZTTr6+qITdp0q+yNsac/n0jNdpUvkksT82ltC7zWazGhWxjOr9yY5ubZ3+KK0wTNUC0rBtRqfc/9q7HrJ+cY8yl+aW/oprlIovTo55rvCuf7PL6OFbSX2/1X3D1TXZuuk117iinYdKUM07rUlWc5HXTSc5b0rE0JqVr1qxnqa/6eH1K5xVtel7idP2y7MrrsfNWjphv1W7XEudNn+gX5/ma5DqXc2vX8as6ez0jT65xtdblOm0e0RZqzeFmyOObsXMu5b7WOB0r4lO67nY01fe3vBZRbtqbculfxRVJzpakLuNVNr8Yt+SQNWik9blGrjJOiek77zfmrH/mj3xF+vRbPAEnsRrn+aPdcriplD9oYAZTGk4tNeMdXTlGXnfWaVnXvf3ZknMtdWWcbO9TtGfuXLeUjp0528ocfm45Utmu66g5Sr2fW90BWQOVrEVZK4sJmSmnknmb5LW3zzsfs7nWuM5KPi+9j6O/vgYyln6ltntA7vMv7+6Nf3PPj7Zc8Zxrlz7wli/Sfx38Hwljkeho/3EHAAAAAADg156b8wTDgs6yRfdaetetL1p5wj3OXP6s+5y28vSH7V49+eQ/2HTlS742PHXJT0ZulM3bDaJDsoE7KDowKZroDU9vO6xfjxqNvytn0qfnRGrY+ZMbYcyZ+jeEKd80p9GR9bXh0N/H8mi7blz72tJAsXLZoHpuN1XUkGtMOZPkSqPGzbT4Kmppz5gsR70pxo6NtI2rsbUyRtta8v7W3t9HVDbfWtYn3ezrtXE+R7IWLeUaRbuN4W2lj9aFsl+5DouJcsTYnObpW66hX5Enr6/JmbnkddFj1ae1rhFX6uQ+stc04/TcyhFnsd6vzN3KERPqbysmq8aVulo6z9R8dVGfY8vR70u/t+vXobwe1Xl/XV3f3N8yRirHtPOIqfrq2tRzyvtBr7nOm7m9rlqTqDdZvz7V9XP6+31k52YM6bGOj3qNyWPE1n2zvcRXsY1ibJ1DtutY1sfnVkwqM6x8vZr5+JhartfEcmR81Sevs8jMSjfNytNrZjDGvDNvxJfra8XoMeZUK+K9j8ZXsusTlWsTTcs4qrouTbtW7rgnpV0/l/UJuq59vVWu84BIP+MPSduNu3pjX5TjDZO9gc+9/K8H3vr8j6879biLl9772OcsGFz7nM7ChfpV16P9miumHQAAAAAA/Npz08bcis6WVc+7/wUbP3jGFzd88py/2Pi+l35388fO/6uhz1z490PXXfQvw/u2/bx7g2z0b5AN3CGRbNpUbs6J9k/KJk/NtzDmwpzzr7dKXL25tQ2uSDaq9SZxfsmGMDe1JbaRbZ5zw1tvwkNtE8dzNuZOnylXKZ9wm2vKqdS0qepLe5o5sVnW8aOuVnt+qszh524UuEqfsjmXcv7NuTzvU27+G/kalZi+MUqdqe5TjxHlmKPPyTX/urfP559rXSevjR6rPrkmGWNrYfUieT3b69asX65ZPcccq5Wj1M+dm7b31zWKudbSeJlD3rP97Xn/5v3abm/Gz9fK176vzsYQ5ftHpXOpx1X1X0+p97z5umbZjKwjvu51/6yrjyJr9xxZV3KH/P2gMX4dfi1VrLZV8f3K/vPlNpW8Wo7znGOpUxNK1snWT8d3ZQ6fe9SlzNSKPvYZEvXRx6SmXDzJpipfzw3jTNU/bzuv2vLafHw997H8Ps22eWSxKp2naDok5bYx5+dWl7ExRl3v5rFIjTn9DzCHpF4+90dv3N0bu1F/xXXy34c+/8ofbv7kBX85+NlzvzX8wVP3r3nG/S7vrF52n/iXZD4w5gAAAAAA4NeKmzLhlolGOosW3aUzsvZBi2658e6dY7bcaemT7vqitZefdM3Q3u2yaZvojenmTDf9+eTEQdkgyiZtaP/E4aH9k7PlRwmkPn/l0345NTa//tVWPUpfe2JOjkXSpseWoRBl20zHuZWlr5R9w5hxYZxY2etNskG1jW9rY7sjTLlGjaHT/gqrqrRJn2LMab21a1yMb3XaliZe9Is8Zbw8j4109vHra5RmUemn7XU+e0pOry+uVa8zpetv8jGyvi7beeT2nNke+WLN0pCwcunbxPs8tRx1VVzLUJDXt/310syh5Xas1uW62LqWProuvg62RrpWUl+3lxxR9rYm3udb96nObT4+p2b9tF7a7Vi3ybnK2jS/l7Mu64tkDh6n8/L4XNeyDlXO5ilIr2vWXiR97DWwa0zpeUrOa6Ml+tmY1j+l7dqW+bzsc4sYO9e+oejnc9B2rfNxalnuyJXj+lHPvZz1PrbKc8+Xy+u1PWJK35xTlCO/95HzGVWUrV/0n0c5l+ybY7l0TZvPnn5jrqxPfN3TlHWWJ85tHjpOyN7DHmfjxHVZnF1Lldv6+r3oXwX395+25dfZM9ZzSFvO09bAj/Y+jDq/dumr62T5pa+tmUv72K9g6z14cFdv+MCO2ZFD0naD6EZpv3FPb+yLl5r034WhQxO9gQ+f9qV15x2/Z+Gdhh7WWbxYfyDiLiL9kQj94aAjwd+hAwAAAACAX1lu6o9vDy2+29iT15z18EvXvfnkj6/5zed9dtW2p39w/eufNzX8wfP+YmyfmnKTstGbkM3mpMn+xpBJDajGlGsMId/IuTEn0g1tmHL+NdZ6cxvSc9skysYxjirfyIeqelf2k7nYJjPO69i+vsUgyk1pmiaVWv1Eadyo7KusWlfatI+fz1HVzyWxZs5IXzNedGPdzFn75JjNucRpWfrb/LUsa2v9cxMea57nXuf1nstzNGVp02NsxJvrzbimzV6/omzvi+2X5qsVa57rMKc951DJYiV/Gls+lh41T67X3Jx53tTF+orytfV2navmbNR+Eknr6nOXmSlajtevP0fLzNNx8ljGd7XWVftl7njdvK6R5c6y9Om/tv662nz217p9bnVVLjeuIq41hp/72NJPY6N/Gld+n3j/0lbH2NFz5fX2r6fljxhXzCfb++pLv4wzRUzUW58woJpxIyZk9daW+fy8brejfUb19YlzVzVXkxpxIZtL25jL2FYfzVnml6abn/tY2jeO1bX4fZbmnNZ7XouN6y+fd3K0e6PUSVyrTs8b6TqaMSf59Jdw1ZxTuTkn81JzTo769NyYPUW3qzf0uQv+YdOVL/nmxjc9+1ObL3/Ku9a9+EFvXHDLTSfLvzVd/ydnDvrvE0/RAQAAAADArzz6Qw4rRGtE+oMOqzuj6++w+JF3eMqq7U98y/r3vOzbW/bKpmp6T294n35FSTdpshndP9HrmhlXG3IpiQmlOaRfecqyGXMqyWVmXK19IR1DzQTbJKoqU0H7ai7ZFNom1to9piXJYZvYPNfNacZqOc7NHNHY3IjquUnbtC7PRdlPVNpDasypWqaexvVLYnMeGmPjSb3lTGOnr3+Zb6hstjNPVefxsjZm5kguXaNQuy76V+1a15gakSdjLb5pa9Y/26s4Kzf17ddJZOut1yjzDdXtXtfXR1Xy63Xqa9ooc/iaNjkzv8lMhjhae7XO2t66Bpc9gWTl5npqla8mljrN48ocvoaiei4mH9/ND733df19/vYesf6+fn7Pt+W5oyx97D5PVXU2Tqy5182Vve6RM/v6vdCc59qm8h4ww0ti/HWOsranom1elThfo6wv6yltJiuHIqbu63PIelcz1+gT8b7e3tbkaKseo8RV/bwuXzORXaOXsz3nW+YlcmNO6/2rrFa2cx/LYyI2zjVv5tT7Jss5dp57Lj/6fdsox/BYqauvJcrlCblsU01rvZa13qVjjqkkpxpzY/pEdDHnJFYNukOyNnIck/OxQ7slRuO0r9xDUxJz3Sv/fuB1z3zfsgfd8hny7846kaJPz+nT2gvsDAAAAAAA4Jecm/M1oI0LhtY+ceVj73bxmlMeevGqU46/aNU5j/3N1duf+p6NV5z2taEvbPuJbiR1k6dfxdSvdZoJsn+iNzSzbba7f1LKeq71KTUF5lFlztUbTZe0q2mQX1+tlIZCMRVs46mSzV/fhrWO1c1k6RP93DCRtlBpK5L6kBscXpdlN3CqmMxV1bUUY9ZPsjXzjzbNG7nr+ubc+2Vdraatbs+NeJUvxqjrTDm2tqWszc2AYgqEmrjqWvLcFOfWFu2iNBlUc8ar6nzd2m21LCbGtNdP5xRjlHpb95CU7TWt7ovWPWXjiSKHK+cd1xBju/Q8lO36d8Psa4pank+aq8qX48Tc6vxutOh1eX2aLm7W6LG/rxpDrjqPq6mz6474dky2xXnmn0fl9aniy+tax+Z9IG0+71D0mSONs2OUa0m9qa9+Ts40jqrXp/RNad8syzX4fwiI/iq9rroux6pjQvn31kq+bLPrr9c45hK53DCMvymnhq4Zc405NzdX5KjXNmLL/EzRZsp++Zr1Sz4fTHPbSn/77GyrGHUSY/OUMU1yj48WY07aQ8P7t83q11hHVfbZr09Y74inpvf09IcnNv3BOf97w6ue9aH1px6/beVj73zaotsOvqizdOED5d+lVf7PEwAAAAAAwC83N2nMLTx28303TT75HUPXXvLXmz73yn/e+LEL/mbgUxf9YOhzF/3T8Ocv+behqe0/N7MnN2+2eZWNlxkHIt0Q2gbNz+2oko1Y1td15ck5bbONYEraLbcbc2aa1FJDIDe8MWZR5PAYn2Oqyd/EuhFT5avk16n1Ka33suevFX0krxtBkTuVdTpumjS5iY760l7lqtuaTbjXZ5uXGzOubvu/NeZs3GwvMbrpl9cojNSMLzG6gdcxMj439FX/7KNyg8NjmhyNcnxdN6vLHNHuaq4311bj6qPlsPNYf8sn0nshZK9N3B/5WuU8fd59KteW5aivVf6GmJZF+uRQyl73Or+Xy7VWaky2VNWefe1c5xyaL7bIr7OJ6Ysz06WqK/mjXKm8RlW8G0Nezjg3n7wtzaiScx6VeD3PPrWsvSm31j/zVKZRquTMc+vvdfMZc/mUbonR+DJmE2exRU18f4xL+maeks8NuZQ+Odeac72eVS7t69dcHbNs5xqXR5e+XnlsFP3mq6+eiitrWuR9tG95zWRck9zjjSTGPuNlbeQa7O+Q2jrJZ7v+8I9JxpvZOTvy+Z3/3r1m4scDH7ngb4c+dt4fD7/3xftXP/deE511K+4c/0QBAAAAAAD80nBTJtxq0a0765bdf8ntBh678M6DD1x6z2OOX3Xa8ZduvOrM7wztnbAN15huuGyTJ5tUNTZMO3tD0xOHbfNmm1/diOaG1OPcfOt/Yk5jsyxSs64Ydr6pKxtJ2RTbhn9ajtOTpmE5H57xYzFQNFb6uumim0k5atnkc/N5yrnm0/iUtmvfOldV722imK8ZPaG8Hh9TJOuQZe9b54v+NjeJsXaRGjRpcklZv7bqBpqvgY9dy9utHG3W18ZQRV/L67KcWo5Ym59eq11PKK8981WxPobnTVPOjTmvU3lcHqXNXkNVnUvbo86kfbM9Y5t2e73ydYm4NGRaY2tubc9+dvSxss3aUxHnY6RkDDMatF7XohrTDAUdK/ppnRxNEp9lz6nzj7nJmo8e3C39Q1oWjajs9XD514g9b86rmByqHM+OaWhEncjmE8p7zt6D/fd6qHkfpLLN68ua6Lm2W95mLL9vpF2l+aJvyRPxTV7t69dj1xvXnW1lPrb2EZ9xuZ5R74r+0W65qzFzTs11RJ+Iy1gve+6ylhJfnnwT5ZzquZd1qOcReVzy+sj1WJyui15faZNcpY/cD9VXWHOO9nXW+ok5nWfdz8aVXFav84l51flCvhZylOuwOejR1rqaU1xLvU4+TvaRsq5HtlWyeVkOn2c9dvkMsvMmRlX+pqisz/D09lk15uxrsDKvsWk5qmTcMfnsH/zCxYc3veUFU2tPffC2JfcYfXhn69oHddYtf5z826W/5jqk/5ABAAAAAAD8IqKm3NH/Js+qxbdd9bA7vWjj9ie9bdMVL5lZ/4YXTK2//KRrNl9x6jeHPnPhD+1Jtfiaqv79uP6vpA7PbJ8tm03biB5BVR+TbeR8g+abOa3PutjISUyzYdfjZCWvbxkHegz5ZjTOrS0l55avqst2mUPLvKlUTIhKrRjpb2O25lSN35KPY5JNq8pNLpVfe8lveSOn9dE2l2+MPS7bvG8ocufGuGmP+Myda6+5Uhab5VTUteQ5Uzau5VVJXTVHU6mL+qNK5qavlZatb5/KHHK8tjTGytoeyjz9salWflX0cdNDjlJn5oRerxztXA0LMy3qfhGT6x9PyZkxd8hVnpozSR9R2yCVeptv330UY5Txsl7k86uk93l9rxfp2srR2mK8qk2P5b6r2v29KeVqHvOuW6oVq31TWXe0tfe4Zkw5tmL0Nakkdar58qVBp+VWn1ATK9eanw+izNm096nMq29ucd8eqW+/kWZmmuWoJfUap9LYrJeyv5cll65P5ivSfpU5Z2NKn1gD/+xs1sNkcR7r66fv4+a9nHMqc7Cn5BqVdSx5VHlt2Z5zTcW/DTIf12RvTDS6b6KR/HszdkC/5ir/1nz6vL/Z/I5Tvrr5t0++ZvS9Z+7betXph9aecJ/f6axZ8Yj41+xo3Jw/3wAAAAAAAPCfylLRBtEW0ZrO0qVjC+8wdP/lT7/XqWtf/eyPbPr4K/5ueO/uXnevbCinJ3sjU6KZCdmUTfSGpv1vxhVTzjbqorpsm9k8174iKevGq8SG6dAYD7UyRsppUklMMeCKIVeZcqrI11JsOOu6MoaOq+etTWkjM2+yPmMzz7xxuRn183rMYrJU/Up/jbWy9E1jJs61PtszR26Qs08+CWb9TRlf5VCjJ1TH6JxLPptfNce8HiuHrM3r3USdW1+rmauXm9gwEKTeN/jeZiaBziNfk3z95PVtvU5FUq/tlrNZ56Kqj8/DYzxW+3jf2ngo8yyK+bXmHIaClvNcxmqk93v0j3H8yULJJfK/taV/7H539eSctJlBJ3NQ6Tzj+kouO9f3k9e5+SKaiaPNT+pzriYpx7zaRnTmd+X7qNUWc/B10WO26xg6lo+XdY2a17TI5uV9yj0icbne866/5jZ5vN9zXrbXLXKk/Pq9vc5Z8um5XH/eF801NKpjy2eZyNe7ahO11iznmm26Xtk/6ubk17oy7zSvwkiLazJpfYlr1IqTsXOOLq83c87idTyvyzXw+TfrkXnmxElbE+dtRXLvlXXOOdWKuMaYk7nNmauuUa6zqxhz8m9PGnP6Zw7078+Ny/tk68FLe1sP7JE26X+j5J66+M82nv/o1ywc23w/+RduXP5tGxGtFekPFy0SAQAAAAAA/KejTwLcnKcBti695cDzVj3t3petPvOhu1dd8PDXrrroCVesueykT29+99n/Z3RG/+i2bA6nZaOVG8vYXA5PT9pTce1N1XzKjVZjzKm0rdnsT/rmNYySWmYA6AYxJf3SGGgU/W1+HuNxUc5621hGH8nlyhx9KvWepyj6+PwkZ6jJF2PaRjTKkqeUS/tcWV9TbGSLedZseFX9/Uof2fQ2c5SymTpytBzz9VF5+xHjdDzL66rNgPzqqpbr8XPORVLfXL/Xaa5WjCnHa2TXY0ZA83r0x5S4eO3mtuvrH8o6G6svZ/9cKrmJkGZCex28Pc4lTxpfjTkn0vpU5DNjLr/KemhPb0xUnpxrPT1Xz0POda52nX49xagxxXmZj8aHbF6q7Nush80nym1FTNzvvr5z+1r/OkeudVXXem2yj85LjnbPZnxRzl3jXPO9/v7aS1njb6L/kZTrN1+bKddNX1c5V80bJyrrYedyf0kfN7La8jgfN1+3fK3Ka5fXZH2a+pxr3adfGluO/Yo59K9nM88mps45r9RAlGPL1O4bLz+T2n0lt16bXV+cq+I+bd4/as65Rqa3z47JmupXXv0XXnf1xuX9My7vl3F5/4zfsKc3tPfin25++wtv2LjzSW/fdMnj3rjyGff+rVuMbD69s3DhcfJvXv6i65HgKToAAAAAAPivY9n9b/XkTe84+VObZrb/y8bPX/Jvmz73ih8NX7v9xyPXTvy4e93kT7tTk7O2SSob8ZBtTnOD2mxU2+fN0fukMefmnH2NNdoaY07GKtJzV8sEk5y1OeDGisg2vyqJsTnHBlE3c6aszz5StvyeryWtb51Hrhjb56T9G/nG06XnuXFu5tsfPzdHW7qhbsr1pndurErGKxtgkWxa3ZhTeVu/yZNmXMZ4XJ1TJXUxbtPum/188svK1qaxkkevPc+tTvqY6rqqXJRxbdnY8ZoWZWwcc23KWldtrrgH+tra8X1zCeX1zZU/+aPxjcmir7300/eIjNd8Ja+R37sxnhpzasbdsMd1457eiErKIze4Qdc9qK+lxjZzsveE5KnHzeuwOZS56FEUfw/MVV1zyPL01bkkNsbKNa7bfSwpp7Itxyj1/l6q66xfnOf7ppGva8kX6h+/dZ596v5armP6lGs131jzyT/XVPO3q/w1yfPquueRr3vzWrlBJecyH5uXnpfriDY7arz2jxiRr0XEZH896nnUl3WRWIvvf19pfamLPqY6JuvUkBNZPn/fa06TzjfHEtm59qnqmtc610FjXFaO+7U258b1KOtpf3dO+owdVKlBp0/PuUEn9+vhkZmJfx+69uJ/7R6c7I1/a8/hkXe/6ODqE+65s7Nq2X3ln741/i8gAAAAAADAfww35+k4fUrgLgsGVj962X2OffqSe299+LJH3OUZqy5+wjs2f/wVfzs0HZsq3QjZ03GxuZIN0ND05OzQ1MRh2zzFhqn9BJBv3G2zWm+yUrYplc3pTHzlNM59AxZtITuXPm6MZbvmzBgvu3ys7FuMuewjKm1Wp+15XTFGjFXnc/Wfiyyn98k63/h6uRkn2i0+ynmeZVGJ13lIW9m4hnL9yiZWY0KZo/llw4ipFUbOXMMtY3xMb5ejxlqfUMRbu45p43p9PilX8ll8bNAjPvvkGNnXFO3eFuVQM047V1tZH2tXqTadyhraudZXbbbOGp/tGqvrHoprzPNmThKrR1snueZYLx8rjpZP3g8yF9N0alLOXf43GiVO+5sxJ7luFH1xj2lUj3quMoNO4g5qvN8zeT3N+1DV1KlxYXVmyMm89Cjn/SrXFuVyDaF8L+Rae78YR9osRs/1Ps57OWK8rPVVHj2WPLWa+pxDnluOfH/re7hPpV8crY+NFdeWyrbIZX1srLlt1l5yx/qV86bN1z3XrVE9nsVlfVXnkvoou4EmZVsvV5M/cth5lCvV+fzzQ03jkOWNGMvbfo/a2FWO9tp6fd4PrXHj78XVa+yvtSrroix5WsZdnGdczrHcr1aWvir5N8mNuZ29cZE/PSd1kke/zmqfgfLe6O6bPNzdN3FYvx4+Ju+X/Ip4d3p3b+D6bT/d/K7TZjac/cjLVzz8ji9bMLTuWfJvopp0txQtEx0NnqIDAAAAAICjopuGW3hxXv5HZ92KO696yr3O3fyGkz+y8cMXfGvDO8744oa3vPjQlg+c893udZM/sc2gbaTjyTXRcDnXr63qjznohr9RnvuGWhQb75Rv4rzN8pgx57lLn1A9bso3iXEseURSdjPNxyl9Ineqye0bwcxVYjRPlOt5uzx2jvribDMbZR8jlDHzlftjQm5Q+QZ23napK9didbLxTIU5lJtck9W1c9i51sc4lk9zp/ryuGnXnKfcCNScc9tUJZ+OYeeZR+sqRZxeb11veWIdSl1VbtS8TqVODZWoa8wdrYvXOttynVvy+bvkPGXrVp1bncRUSoPUXxPNL1IDYbrS1KQcJxuDTmO0jxpzN6oZd2lv9EuX9Ua//KrQZb0RPZd6e4Lu4G6bh/ULqUHhUrNC3pN61Hprl9x6DBNDNfd65ajSc523zT/qLUbfK+21tvW096CqOc/3VMbla5Qq/bVPLauv2vsl80+zrDGN5J6K8nx97Z7To57b9YiyLfv11XtbjiM5NUakZlEx5lQRm2ta3oeh+fKWtqrOJfWqNKMsVupFrdx1fdZVyveLH6t6PVreaJdzfz01tpHnaNfpGjfn7by1alP+SGqZcjp+qspt+etjrre8Fq37Ws/lNTcdUEmdHOXfkVnVyEF57Wod2K2/2H148JOv+LvN7z37f3U/9fI/HvvQmV9Z9+IHvXXB1g0nyb+RXf+n8ohgzAEAAAAAwM1muWizaMi0ed2dl/7G6PGrnnu/c9e/4XnXbN6746dD+2Sjv/9ykWz27YkK3TjJ5sU2SfXG1zdFrafj8txUbY5UksPNOOkXR633TXOz2fVNum7GpF7OVc0PODQGnW3YVRqr85Ojb+qkXepzU5fxNucS433dwIt8Uu85vK7kq8eqFfG+SfSjbzKlbBtBKata8dV5XV+d53zruqZNJTn6+jRt3j+NiLI5LnNpZE+0hZmU6+UGk27MXTmOtUubqlyjjRVxITef8jX2uvZ85Tzye46mr4+pOd30NePXXktfj+b6XJazT027xMe9ZPdWXaf3VMQ38liLt3nm2nnZ5mk54npirh4r/eTYfC04pOf6BJv+HTjTblc+RSd57H2iT6Dqj6b0a1rmInO1e0vXVf++3A1qwr2qN/oleX9++fLeyFde1Rv5suhLl/VGbry015WYYZ2bmUQiyd3V/PpEUSrfnzq+XVdKr0WOdk0hvWa9Frv2Zm0aEzLWTd8vqRJX1+d55FNFzuYYcToPnY/1qeOkXeatys+e5rXTczmasl3Xwcv2msfr7vmk3a7VX898TT3f3Birt5xNfs/p8Uc05mT+5T0YKjlFZRyVxTZ9Sw6Tz6G8Zpk32j1fFa9tpngiTo0xy+F5XM18/LM42jVPqj+f9Wn629rFOmf+dmwo73kp2zHKc56m03wxdn4GldfH1laOWrZ5ZZ2U7b6W94qU7Um5kN4b9u9OfCaNST79eqtdgz5ZKhrat212eGbbrL3H9l0qbfLvnvS1e/S6V/zV5lc/9Q9WHH/rkzubVj1Q/s28tUh/DEl/KGKJCAAAAAAAYA439XTcMYuP2Xji2qfffdvGcx95+fqJJ12xbvJp71v/m8/5wsCHz/vzMdmUuHJDJ5sf24TLhlQ3KrIhso2kbnpMvuE1403LunluadJjrKx5dFOkGzXdlEnZNnhVvOWTMfry1zG2QVfp5kpl5WzTDVWcx5yzn28m/dxzxLWUY9M/zaDSJ8ayDV0q2mwDmtJry+sr5oWsYSWPjXlm/lrSx8aOcew16JduSst506esTV++Jpfna3I2bfb69mm+HB6bm2Yp1+16zSmL83bfXHs/Hbc9TtapwpRr9VX5tbm8rrnGvN5mPsNyzzQmjffJ+Hbfeepk3GIw6DpnfdyXGatHl+TXPnEd+dqYAVmZcqMH98QvqobSnND7PH/RWH/d+PPbZ8eum+yNXz/Z23r9jt7WqZ29rft298YPXtbb+sXLe+NfeU1v/Guv7Y1/XfQN0dfk/Cuv7o1/8VW9sUNhps+Ipnf1hq+fnO2qwSc5RqfkPWrmnD455wbGWBhK5d7VY5TzWvK1K69xJX+v+BqUdTLpua+7r0+oavN27+evq5Qzd8nTtPnrGWWNrWQmWdXnZklfI5Neh8qvs8zZlOeiHH+escy86qs7qmJ9M/ec9mJA6WuhufV1yTG8LnNkfXkNtT76Njnmk/SpZbk1r49RJLFejvpy7nX+HpO2as6p1nj2mSj3paldzmtJNeuiZZevSyO75jyX+9rqTJLb+obs+mIORV6v/0aoKTc8o098xxOq8rrrWuq/gcPXb/vZpnef/u1Nv/2cTw//9okfXHf2Q9+06O7dszorFj5S/i3VXyo/EvrvME/RAQAAAADAHFavfNDtn7751c98z8DHzv/uwGde+c/6VI5uRrp7t/28K5sT/9pPo/x7Vc0m1TfCZdOj5Zkw30Rlo2TyuqLo5xvB2CDVearcttHT+lbfkMzHJe2qrJOjzTPam2Pmz2OquZ76+tJE8PiQzdnn3Yzrss2ctuVRr8s2nn7uJpVvOOtrdqOibxxVyS9tEpNr5TmyLc+bOr8Gv568vlY+66fq6xtxjUHR6Ehtma9soLNON9kpyW/xsbG2zXW09eerDZLheM1MkT+vzdUev1yr5oh4NeaKORd1eS1tNX1KneTx10nmapI6uR/TgDBFHz9vxijSucjrPnxIyvrjDKkbVLt6w6LuIZE+UWdzkGvQ9+LUjt64aGxav5I32RuXtRq7YU9v7MuX9ka+cXlv9I9e0+t+53W9kf/z+t7Id0V/8vre6B//Zm/kf72mN/SHr+oNf0P0FYm9cY/NwZ6Ym5rsda+X/PYUnr5XRTKemz27/b7VcqxprXwdSp3EqsprUV6rvN+q+87i4hhxnkNj+uU5yzgW57L1LK+lam6O+cyym5LfX/H6a7m6f5rx4jzGSNV5fB1VWm6P4e3z1Otro5JyjpfrkPea33sa52NoOV8rq9M1ihxZXz4rIr7JEW0R4+dNDivr2Cqtq8u16jqLiTWzOUvePnMu51NUzDiRGsh29PnUynVvxtB8UtZ1CbkJp2VZt2LMyRhFfk0mmW9b+rnqGtmv74cJ137VdqmfkLnp09nbZuX88NiBbb0xqR//0u7DWz969hc3n/Hw1y0c2fwI+TdVnzwHAAAAAIBfQ27uf42/9YKhtSesuP+tT152/B2euvLE+7107SVP/v1N7zn7j7tT/rSMfe1HNzy2QREdUENj4vDwvu2zbspFm22Q5KiboD7ZhrrIN0jFwCjyTa61W05VbLaKPEezAY5+MUZu1rzcPs852hiS2zfdcW7tzXkzn3q+UVZZfz2POjn3jbdI5pwqG0ndNEpMOZqkfEDiTPWGMOIjrx7LPES56a7bvdzk9vGrfFFfckSbrkmzVs21uCoTIucT516WeO0vsa06zaV1RR6nOZs6yR1ln5fnz/O6vZlDaL46k16D9vfrqNsyd12nr/Gwqq6ztdKxU01drWJe2LxVmk/GD7XufzUEpvvqbSxRGHPDh3b3hm9Q7ekN37inN3ho1+zgwZ2zatTpDzbYPSJxozfu6m390u7eMV+R9+XXJntDX3xlb/PnXvaPG9/2/O8MvPbZX9pw/uP2rn3hQ69ZfdKDVFevOfnB1647/eFT61/+hIObX3fiNza/+5T/s+Wac/564EvbfjL8LRnr67t73S/Kdao5qK/djD4RpIaclGUd/T7Z3dOv98l7YLbc0zL/ch8W+bpoW65H3g95vdmm76eirIt7xF5/6VMr77M8uqJNFevq8drf8xRlXI5VzdPKcV3FeArlfeDtntdy61HHy3H1qGOXcpWnzCPGUNXt0be+jlb/UObOz0uvl5xyzPe5v6fjGK+Pj9eMbfOwfqnM5bHer93meVVe34zV1Oc6tY1MVZPHcraMud1xz6g8jymeklNTrjl3lRhV5LU1kRy5NrY+uZZFWueq79Na/rro+s8j+zMJjTk3PL3tsH+9Vd4zIv0Bia0y53E5dvde8u8Dv//iGzdPPvHNG57/gEsW3W70OZ2lSx8q/97eWaQ/onQkeIoOAAAAAOBXiJvzf/AXLHvYbZ+18U3P/vimj5/3xxt+79RvbLzy9P85/OHz/3Lk6kt+lF9lc1NON0J6zLIrN4Km3Ozk5sf61hubSnLebKLz6LIxZHOTX3VsNk7aT47Sv3+T2p/DlXVyrDb9liPK2l7y1tLcVu9tlsPmHe2lf19ekW76TGVNQhajxyibpKwbUDm2ZH0iJses8mfZ56/KvCo5F+lrkyrjRXxd7+sU15txpthgz6PMU1T6+Jq14zVvO3du2IvpMSdPxoSs3ePquvk0d/wjKO4dNedKneRtjS11Pp9c22quNp+YkyruSTPhKtnfgatl95FI87eMuT297g2X9ro3XtobOrR7dvjQrln7m3BfubQ3/vXLelv/52W90a/LuFdf0Bv+rVN6Q+c8qbfy+Lv/28LhLV/vLF3x3s7Cxa+R9/TZoheJThG9UPSSTucW53c6C3fcYsnKNy1es/4jS0YGZ5bf57Z/ueXMx/RG3/yC3jHXntsb+cr23tavy333Ff0VV7mmgzvcGJHrzaeW5BpntS7/pmS5x8u6RJ2uiZovsq6+Tvl6VMdYK5PVi8x003LGVrI2yWX3V+TVc2kz5ZrWbbUiT2nvl8XpNbTrG1Nunja7fyRvpeaa51GVa44xl6rrW8oxfFyv0zm5/P3sr0eRvRYSp21Wbsav++YYGtvu6/1dUV+327rGPKSuyev19jpK3RxjrpI+jdnqZ3kaE86U91tVV8aLvn7P+Nrk69JaU12zVBlfy41sDfvqyr1l8r9fmk+T2p9g0KPck2rM6dPjY6Jx/TXX/Tt+3r3mwn8c/PBL/3zoky/7y+GPnfbVjZOP+/1l97vVWZ3Fi9WcWyCaD4w5AAAAAIBfUZaKNom2dlYuvn1ny8o7drasOWbJ/W/9mFXbn/quzddO/HhkandPf41xXDYj43K0X2eUsm5UZDMyOzyzY1aNstbGzFRt2iLevz4km5V6MySyDWVK4nSzUza0ei7KjZxviKVOlBt13yRFv6grTzuVXFreafX2NcWI06PlspxNfu/rY9v4ci3eR899Xt7flXHN3PrapK6sUYzlcTqmHHVsG6O5vtbTcnV+iwlJuay11kf/nKvn9bLn9X7N65X5ciwve47qGqLOpeNov8jbl9vL2qe/Xztn07c/NvPk2mRMxudaxbnNP+Ll3KSvd5Q9xlXuh9ImfeLeyDrLPRP1Nk5eU46RdanKEIic8445rZJNvB6n5HqKZCO/V2RfFZU2uVY35mScg3t6I4cu7Y0e8l9P7X5Z9NVX9cb+8PLe2JcnfzT6+6d+f+NLHvmdRXc+9pudtZu+3ums/Hans+yLnc6Sj92is+Q1ss9/fqez8GHyHr+9aEw0KtJfiRwX3Una7tfpLH5i5xZLXtK5xdJLOrdY/k7pe02ns/xrC9es/6MldzzmG+uf9+CvDf7mSV/ufuDMb4x85sLvDU5d/EOd28iB+CVXU65DY5j4mjTrkV8TbNZGrtPWWNZAjvn+bX0WqKrX3mJNWtY6Xy+NaZt0msvVxEifOHrZY60ucuU8TGWMRnlePouizucd/ayvK1/7Jof2adbA+pW18vr8vGzny/4hXTer9zZdR1+zyCHX1ryX9b3u53rUGD+PupyTxXt/j6kV8Zpb2yrlGDau9rU4Pc9+EWeSuUqdr332bXLak3J6/6gxF338M1nl5psZcKm837Q+jvlVeNWc18bO877Q+vb1NHNqynn9vi6uue/pCck1Gf++NU+Xju53Y25kZvvhsYNqzqnk39JD0nZwQnJc0hv85Pl/v+WNJ316zXPu8/IFt93yuM7izu3kvTkS71X9oYij/d1XAAAAAAD4BUf/C/vR/k/95oXjmx+54qn3PH/NxON/d83lJ7531YVPfueanc/68MY/OOePRvSHHGSDZEacbZ50o6GbFt2c+GZFNy2+ecmYSv31fZvesslJ6WbT6jRGY6sYyeHSjVqUJc6l5b5cWm+bJymH7G+H9fWxfpJLVdeVtpCbMXWb5Kj6ZFzmMvXV+brEJi/Uzhm5rE3iqg1mblJ9A5l56nbPX86Lok77R51vNOO85NSyxNaq5lmkcxKlOZWvgbfra6OS61CVtb65iuuPY616Xq3XoozVVutrqCm7H5pxVHr93tZIn+pqrs+V62/xcZ22frIW/WZAkeXXfDGuGm8q/bttqr2yMVddv921VzQ10Ru6bvvhob3bZ3Veowf011QvtV9O7X7r8l73O6/ujX9tT2/sijN6qx9z7+/eYs36D3RusXRHp7Po1M6Cxc9csGDpyQsWLHpOp7PgifL+1l+C1E3+sGiNaFlIDXn9hWXd9OuvLG8V3VFy3KvTWfhIyfNUzSNhL5C650muZ3cWLDhhUXfLizY+60FvHfvAWX94y2++tjf6xcvt67RpznVlHXRdhqd3zKr82lPVezpU3p/y2rfact1Sdm8dSXnvSS45tu8nWWPN3aqrpfHZJ+czX3tfnMrGqpR9rb/MJ67D+khdcy3NvVHHW5/++yb6tvqHMq+1lTWr+7rsfd5Xl/LP5uqzwCRtobrdVfWf73VKZS6dj+Wqj5Usvs7v8rYqrlKJUyPO3ndqzvXHRN086+GvVaMj32/z91fVa2/K97XKXo/KmLMfhdC4bfY3WMfsq60uO5f7dvi67T/f8oGXfW/jm1+4b9Mbn/vpTRc/7u0rH3vnPZ3Nq0+X9+U9REf6N1zrMe0AAAAAAH6J0P8D33xNZvnyoSX3veWj1lzw2Nesv+L0Q1uumfjXwRndWF8mmxLVHtuc6N+Q8j9yndJNi2zYZIPiGzc/b9pDshmxjUy2x0alMeb82Nrg6AbG+lXtUdaNcOYqZoyMUza1tonyjVTWaU7rG/n9q4me09ojT8lr/ZqcKr8GNWd8jFJnamJLXZWvrrO1knIx1KIt+9fK/s0vs3qsKtfcjbRQ5PPXJM5bRlGVI/tnW91H22tlP5U+vRdPRtZ5bc7xmhcjS+utLdc6VN0brfrcKNu5j1evh9VZvc7RXwOTjSN95om3JyPzPF5/U9b1K+cgakwVUX29sd7FmAuTsvzNq1Aaqvma29w1p35dNZ+S0yfk0pC7bkI252rIyfELWpY2ae9O75ztHth1uPvlPT8f/NLunw1ffdHPNu08qbf8QXfvddas/94tFq98/4IFap7Z03Bqtv1ns1A+Qh676NjuO0Z2nPS/bz2145+O+eZvmkE3tG/XYdHs8H79aqtcv12vXL/+outeee9PTdovutrTRPqVP3kt/P2va+Nl6xNr35LdN3qsJPeKt/vr7/eMvr4pNeVkHIvrr89yvMZ5Xu6P/vamzudT13tbmY/M3/OE4ry0l3vJ+9X3ZHkvV9eYOWx9LMZl/UJNveetVX/WeI4mPj8HPCbGbrXNlbdLXOt1as/NZNehyvO6LmSxPr7J5l+pjg2VWHufhTlXtbukruSJ8UP5Hi+vi6kdY9I8rTq/Rn2d7D/uxNo30ntaZOac399uyuk9qJrwc32Czp6ia34sQmP1dR6W+dqTujLG6PUXfX/krc+6dtlDbn1OZ4V9xTXRf8P133K+1goAAAAA8EvIMQvG1j12xXG3OWnVY+72vJWnPGTbqlc++R3r33LqwcGrt/1Qn45zqSGnmx3Z9OjmRM2BmUl7+qUYS6Kxoti0ycbCFedpwoQRUzZvsgmxjUhsRrQ+N6ZmnFmbb5yaDZPU2cbOzzPejRlXbsrs3PqIdKzMb+N5m7VHLpP1z7b+OM/bVhWvx7pN1kSV61TX1W2Zo5HW9atq17nYfFTSVso+h6Y9+qpZJEdXxGZMpYw5WpvJzCapr1TWqa6rpBtYV8SJWnFZH6+3r32/Isb6xOsb0l9gLZLY1lgam+Wor9tynTyurea10vi+9jIPka6JqnydM77aWRl09vROjqvXaL90uqM3dN22WTXkup+f6A2bJntDX5joDUrd4N6J3pj033pwx+zge8/83vozH/6Fpfc65t2dtavf2Fmw5I23WLTiHQsXr965YMHSZ8v7+g6iRfYO/69huNNZ8JTOLRZfvPS2w3s2nnzc24dfe/Inur//0i8Ofm7bD4Z0jaZFU7t6w3tlzffKaygaUZVfds3PgFiTOa+9nOfrbGsndXlepHWSR8rlnqpz2Pvd3/8+3qTUqaLN6qSftccxlOPl65bvB6tLVfV5PbVyHpbP6jzWryfKGpfx1ifa9VpUlqMZI+Pn9I/z1vs4xqvla6196nrpI2OlqZ/yz5i2amPO5mn5mnn49WbO6FevVeSdM45K82u8xbXj6z7tOv03St5fWWf9M09Kc9XnWdfIr6W9lnWcrVdeY1yn9dH2aNP7unxNPYw5k66JHNWMHpOyyg07ec9PbTs8sm+7xMvngOac0c8K/fd3j50PfPq8v9n0ppM/u/78R7xm+XG3eeGC0Q0ndZYtfpq8B+8oOtpTdJh2AAAAAAC/gCxa9cBbnrBlx5PeueWKFx/aeOXp39rysQv+cujTF/1t99rJfxqZ3vHTsYOX9kYP7oknEHSDoIoNkWxA3GiKDVzINxq6oZPNhm3uYiOisk1znvsGx5SbwjyKcpPjxpxKNyleLjGS3zfGaorkucSV+irG+si4tjmSupJPjpErN3+2uQrlPBrFBlRl49ZjZ71rzuYwy1lvirEsvlZ/XMrb0xSam6+ZT15Df3sx5SxG6oo83l/Xdl1ps/6Ro356T+TXHOW6X5yr0izL85xrmiqlvkjrsq3faJurxpSTc1GJzfj6XNS0V+sg15fr21rnUH+btcfa6v1g5WLEqeQ9ZAadH0fNqIu1krHtnpyS6/vCttmxz0/2tl4jm/TPS/3nd/WGZnbZ35Ub3ndxb9OVL+6teu59/+F/DG/6dOcWC8/tdBY+XN7Hx4qGRLcU6d+L2yJaKTrSH4z/zyD/LqXOY6PoTgs2rXnamifd/TWjV572lbEb9xweVXNhanevq8bc9ZNyrRO9ketE+pTg1IQ/OafSJwfzPZmvvcnXKl+3+eT3Xny2RLn0t/e7q5gj+9qmnJstImvTo6uVvzov0vtKVdVZ3xgrx9R5aO7SHuVGHlPmO6fd+5R++vnVKsfRrkGVnzl+bvOIcr/qPi59rzfy9cyYjKuPMb4q5pDXYO3WX2JLjugX75fmc8XPTdVciup2USu+1Om/V/FejLHmGnNeb2r9jUPX3LUUzVkDkV6jxklZ5XVZL0pzrjbmTPo5G5Kc+R+x8kk6j5HPAY2zWP8cGd677Sdbrn75P27+xHl/PfyRl/7R1o+e+dUNFz324wuO3aI/4KLvwflQY+5Iph0AAAAAAPwHc6T/Kq4b5/FFWzbce9ExG+/RGdtwu2UPue2T11zw+LcMfOjcP9Wvpo4dEB2U//OvOiQbmUN61P9Sv3NWNlezzeZMNzm6UfCn48rGQTYSKvt1Vtlg5I9BmGxj4Wo2NLrp8Q1KS7GRyw2nGnMmO/dNUOYwU0Q3X2UDJrI5Sg6Tn/vm0PPlpkiPPr62yXmobK6ynPlFc8eKc1k7N+i8TfvVZk0xuqLd5mTz0vNGZSOsT1XFk1Vzvpolcc0c41zqizlQxWUfa6/Oy1yivrWO2m6KuURZx8t+zQY6rsWu19dX51CuT8ulzZXGmfeZR9G/5C7mit8D3l9j3XxzEy5za0zVFsrrK+utZTvX/FEWpflc1zXyr2Oa7H5UxXnO1fJFTr0n5HWzJ+jUiJP3U/cG0Y1SL8ehA5Oz3YMSq18Fln5j0zt643snRJO9sb0Sf+2ufx3+8IXf2/jG539t1Qsfsm/RXca+2lm4+EbZX1/Z6Sx6qbyf9atsC/XNfRPoZ8J/xtMymvNom/31okcsHN88uf6Fx71v9Koz/7D72Yt/MHjd9p8P6d/Q+8Jkb+S6+Nru1IQ/NSdroF91LZ8B+trb6z1Zyn4vSJ28b1N+70S7yfvaPSRtqWK0WB8dw49WtnqN83rrEzmavHnuau4jPVbnVqe5RDZmjqMxLh9Tpe0e559JHl/KEWc5IsbO9Xr6ctQq163jRdlzSHxen8nfE/7DJq5sK58Bem7z9usylTh932Rbpf55SF1+bvhnhyv/XSj1rc86zRtt2r/k0vomRxPvqj/f8vqba4j5a4zOL67D82u5UX0N+drl61vmYTm03LwWdl9p2e7nUK595NPzcu32erjMoDNDTg061/DMttmRA1Ivnxejh0Q3yL+9N8hnhpTHpO6Y/Tt/OPz6kz646qF3en5n5WJ9WvYuonuL1Kj/rzToAQAAAAB+7dGN8pH+T/iWVY+7y/M3vOpZ717/+pOvXjfx1PdvfPVJnx34/TP/cOSabf86LhuNMTUI5P/kj6hZoL+mKFKTTjYks24S+YbCN2q+ibFNhWw01JCzjUYt22T4JsU2LLbx0L5R1jrpa5tFOTdlnG6Co679a6ka45sk3yjVGymf382R9dVxbLymzuvrmNiEHUHFhKnKzXyauLJesYb9Y6XKRlj7mHa7dLNab1j7Ff0tr9VFTj2mIsalfbJ/tFuM1/nY/XPxOo9t983r8fGzztfEypWKMVfXz9O/6Rf3QUjP3RRr6kxW33cu/c3Ek7nWyvm36/VeV8V5HkP5tKj9CITdtyGZS3ttde5VfjXo5L3kptzu3vAXd/eGDu2eHTo4eVjN7zFpGxeNHRJ9UTbZX5uQ+It66171jO8tutcx7+msWHKOvH8fK3pCp6NfWVt4vJTvJFonujn8Z5hyydFy61dp9QcmdK737gytPmPdKQ/+5NZrX/kTNRSGpvQHLrb5j1tM69/UmrDPje6U/kCEfgboOvrrXd775dzXvf/zoNTV0vd4fq6YJK+OUyk/C0pZVD67RHXevC/zHm29D4ritY855X3RzDGk48W4TdnbiuFjyhgvlzlFudx3Vueq60pbyvplvd7bMtfqvs62OUaVXVNdFzHVGtRt/vne5EvZ54rE2L8f0leV75mUrYFKyuXzJ86PKOujx7o+xrS+rrIOEW/ziLLnSbVj7TPB5DmtTq/RyvEaxX3SvJ59sterludy5b+bIonNv0WXxtzIQZG8d1T6H8/GD+0RXSr/Puw4vOWD53x3y1tfcN3Qb5703qEdT//EsofeWT47lp4k7z39YZf50Pfuf+ZnAwAAAADArz36JIv+H/LVovVLHnTsE9bsePqHhq6T/7N/3c7e+PXyf+qndvTGp3f6E26yGdEn5vxJrTCBRFrWv2vlBpFuYER1WTYTxYSTPOXJudhclM2H1GW53SabDNvY+KbFNqGxSfENzk7ZlFQbn6gr6ttEeb/IoRuw2JRZXVW2cxnPN0pNue6fx6a/jNEnrc9Nem7ErU3nI8dmQ6nH2PxZu+fNuftaZh+Ji37+N/1c+aMC1l/PTXEe+dqb5qZ+XsW8W4o+Pr4c7X7Q83pObbkJ4WuUxonmsPPIV9pFzRNuXp9jW0w59zzeT+P9PvB+2b99nCMdR/oX5TxizjZ/ub48V2OufgLOn7Sr42W9o90NBymHmWHrqccsq6LfsIwxbMbcnl73xkt73S9e2hu58bLe6I3y3pLzkRt2/Wz0S7t+PPDF7T/e/LGzfr7ujOP/vnPrwW91OgveJ+/d54rU3LopfpE22Eeai34Wnbb0HmP7h3772T8cufr8fxqe2vbjgesv+dnQ3osPqzGXnwnltY+j1rfe+9qW66zn+npkW8bm+3vGv0bY+tGH/jZr97qMcaMm8ld587y8Zyr5nFVxf2nZ4v2+sX4qHTvyNHOR3Hm0GG/L+8quUcuWQ+OirppXibOy5/N4Vz1mc6/KPGNu7TGyzY/Zt1xv9PPPLa3Lo7Y1KjmrvOWz0MpzP1fKWsV5xuRnZL+auUVdFWf/Rul4Ni9vy3jrU8vmH/1NXp+vX11XpJ8BFiOfN3IfqfI1yde1qLxWqWoeNq6vi0ni/d9JbQvp199V+h/N7Kl2fXJuT29cPkdUY4fkM0beM2Nf2d271f+6vLd5++Pf0xla9wB/+xn6FfclIr7WCgAAAADwH8SRN+MrF99+1RPu+uKVpz945//H3n/A+XVUZ+P4ZXfVZUmWLe3u97tNxYBptjEt1IQAhtBLsAMEeOlgMOCGi7S7klbNNi3tTSAQSkICJPTiKu2uurtNC4HQ/m8SEl5II4EQe+//eU6Zmfvd1cLvhcgGzvns87kzZ86cme8tc2aevWXx8x81uuKCp7+3//2v/9uRPVvqkT36uNzI9Hg9bJN/X7xgITejX1zNSHksSmaDdQxMG5JetlyM6DYvFm3BYXV18cYyXeCwTBcxumjRhYvrsl7tsDCSem6HBQ7gi2Mpt7zorJ6juWjKPsSedaU+daov4fYK+rYybJ3AbCzmiroO3/dp8UiwvoM2DThZp3nx02jDtvP5YJsF3D758X7Qt0DzQtLZ4nDWvhEfOHZYGDKf9qVtHUqceV73u9Yl8u9wkqPz7jjaa30eN/fjdoCdQ65vlOO35HNByWdC7pYTlOWoV+TTHXMJ1me0V0JICLbJvqOetonjdWi7LJyHboCfG3fVI5/bXY98aXe95tNv+PvjXvKYz3Sf3P6DasnS36mqhVuxdj4X27NwJfPLqj+J3F2IOfZjvr6cUlVdZ1ddC7d0bejdtOpFj3ln3/tedf3gwfEftg+N1+2DJM+w3w7iGOMcI/w8yMcYxxZ53f88BzL00VYlRxJJktC0VeQy9YU0tjo+lWD7KJdzG1vq2AcBzgVLe1/zdZPPFT8Pm2OgpUuwHfttmma/WE/rqp+mroSUsa7k6cfA81Dqah25duYE6nmdxrluaf+9ktatjF2wKceaZD8H0ljHOm5vSH0sdHrHdhNy7XJr8HMk1e+AtkG/zOd6hJYBovdys3U9fPg+8utcQL2A+5TnHLd+HDrKuTW/fq4c7XxJ/XbALkHG3y0z+mgr77adqEeOALwDl4Qd73zft63u+9DrvnzC6DP+YPnTTz2759R1F1QrlrwB1+CjgaN9rZmE3d1lLAkJCQkJCQkJCQn5uZCjTqCXPuvUl63+gxcdWfPZi3/Y99Hzvzvw0Qu/N/CZS344uHesHtw7Wg9NjddD08QW/c88Fhic/Mudc4ZyESTAgqGRFnABkBcRvpBIeZbJAqZYKCZ4XhcymtcFTVqQwkdezBQ6+rQ2yrza6gKnJFXyImj2wif5cEhbRV3TSd0OlPV0MWfgi7pJxth+Sm2Z7xK+6PKFoCy8rF5j/7s9/RZ5gbRBZD+um8uPgO1aea7f4QN2c281Lful0SbzeozSbzboftpS651sfiwV2S73wfdrScoJIWdtii+cLwIvt7zXnY3iuBqEmBNybnZZieY+IsrfhjbZrhE9qR8Cs+N75m7eXQ99+bJ6+Eu76pFPXVAf/+on1/cY7ttTdS04G5ct3xm3DlgL8DFVgne9/iLd3bIU4AvqjwP4qOtDF508dNnw2178tXvevqMeuW2XvIuvfXhL3TJiToBj7ndaEnrO6Jb7uJOMK6HH4GgwUq4T8KvXpaYdegeforwOPK3jhOsM6RxRpPHP9J3t6m9ye+pzHdHDTmySr8Im1acv5gugL96H/PvdJ4Fy03sdjhOzznv5XSy3LfI6ljTHrVn1CjRIJtYXn4T2T9MGuWs3j3kC5juIOa/r0N8AvaW1Dfo2/yyzuql9t/d8slUfzf1VwvenQs5Rpu04pS37WKDRH4e11TgPBUUadXQ/YMu75/go/CF935w8Go99Nrx/Bz+w8oP+T1z0f9ufufjvNxwa+/eB977ky4sfftI2XHf8YMxcEsRcSEhISEhISEhIyE8gc92Rwsn0KizhB6velesW/ur9zly15cyPDfCLh8DQns3pJettoHXtpXe092yeUWJOIV9/s4n/MCf9afJviwRZPDiKxQOBhYksGlE/2VqZLlS0rFwwia2XSX2mtX5aVPni19Lqw/xYvVlpg/vRdLEglYVXxwLIyrSdXJYW2GLPtNZ3cifZNuywSHJSjmmD+i78l2WNdOErATqgJNFyWstKn5oufittpT8dC1nxybpaXwko/m4vt98qyHYlZJ+KHuCxkfo8bnbsUtrLnGRh/wjq1bbRNn2bjZB4hKRpk8kZP/Z+t5wuhlUn56SXFzZ6Fxv01p58nIHwNNsXHfbXrMU/9qPszwzfx22+G43kxjT8ysvetQ+DB7BAvml3Pfi5y2ZaH73gW8ef+/QDCx9y76uq1auPVAuXfqrqWXIerl++sH0+4TX/80rQse/zvYD+8dWSZW9d8rB7f6T3/Kd/YvCDr7ux/7rN/9C6fkfdOrRNgeOhx4/HHvsXEOICY5eQa9xOjwna+4jZxFwjvx/b4tFWJUHMr0DzaQx0WJnb+TmaYHk5J9L5YbbWj5QXP5b3c6WR13YUOa31aQfATv1aHbct2knnrpQVPsRGfWhdnLviz8qBfI7rWCHjhUHHF60nNvjt5fWgdX0sNFhZtqFPbCVvvsWXwgk4IeHS1suLMrNP45U9Ys6ttmX9lz5qP7Vd1Jc7YdWH/w75feZTYH2TfWf7R/Yl25B0eaxsnEFZ/icB9jP9GvTDMPCbUJQlP3Z8AP7zTI6XtE1f0MGW0H+iMQ3dQcTxg2hzevROf7XE4PR2xHj8vinkr5+oW+9/2YHjfuv086rBVY+uli9+LK6/RwDzPS4fRF1ISEhISEhISEhIh8y1QF9Y9S6/33HPeNBLl5/9hLeu/p2X7m995KLvDNndcYOT2E5iIYo80b5u88wQ8nLX3BQm83L3nE7+/c65ErIosUWALB5cx0l/CVlEcOtoLlaaZdBx0YJ6uhCkrdZPtqm9vLghysW5k0YlGj4sLXZYwHDhnBZbtuCatYi1OkknCzu2k9tq+JA6GeViTxenpV2u29ALin1c6MvFbF7EOtg/t7f6pmd/k13hL0HqzqH3PqBMf7PCy/1xViW2zNag+0cXjlmP4ybwcq2X03pcmVZ7tCdQOyXmrA4BOyXmND0n5JyED2x9gdwk5whvh7C75mThnzFwcLvqUd6mDeDv95P9grrcT62prTOtyS13Dk5BPw09tvwq8fqbttfrbhqvWx95Xb3ylY/+r3v0Hv/Rqup+Ka7Zh1U9PY+rqp7HI30/YCUv5HmE1/3P6wL5x/W9VVUL+AXJhwEPWbCx7+y+S572mY037bpz+KbL6l4cg14crxaOZ8L01rp/cnxmAONXm5gex9aAdMtQEnQl5Hzj+CPnkSOfPzqWledwgcJOAF0jX+jlGkda6sh4p3D/bFfSODe1npWXSHbqJ/kz2zSGFrbaRtN+LhzNb6duTvC32e+TOGG/N41ruC5kvDKd7guH+7A6SW9ltKUe114G86ovIcQc4Tr4EEKO+6Twq+SVQ9tstgskH9gP5svzorOt2mve0cjLceGYw+Nh5WiXEL8FpC/2u2Qsoi0hx7QEjoulyzjNcUbBL6EjhvP4TyPG4zxnnSGMR8PA0D7GpW11/ycu+M6ad7/ihrV/+KLP9v3+CyeXnvXg91THL32qXIlzC0n1n9exJyQkJCQkJCQkJOSYyIkLTu592PLnPPg1q7af+YnWp9/03eEpTMRJuE2O1e29m2d4V8kg8oNYtJKgI1GnxBy2Qs6RmFP4YlBfPI1FACf/snjglnku2vICoQGzV6hdWqjIAod6bN0GbfgCRhaTDrEDuAXKsvSIYMNvXtA4fBFT2qWFltVhG1I2q+/Us/+e1zrZF3SsD8his/QLKCnHrZYnG7ErfJZ6QPZv+u2uMz8CXZBJ3c597rZeH1DCALZztCXlhlLf9GH7lRA7tXViTo6NtOFl5lNIuXyMUpnpvL/NY6bH123V3pBs5ob/Fu0z0rJv8HsAJ+ZaOB9alnZyTs8JJduUcNuOBX4JLvhJzCk51yb4lWLa7ocO9YexHZmeALBfsPBtT4/f2T40+h8Dey/+v+2/OPc7Kzf95nd7HnGff6iWrfx2VS2ZqqpFr8Y1u0ov3V9qmY+s660WLHrD4sfe7+DgO17yd0PXbP67/n1b/61v7/gP+vduvaON/d3CGNfey382APyng/zjgRjTtBF0TsS1ZFuMHwkd51FxvukYZucV9HodW15s9Hz2ulmPLe2kTqEv2lQd/St8DExpwOulfjTatnaST9jQr9Sjr85+0abM6xjQRNY12yfMb6nnWOBA3RLqS8t86+mmXyvjVsqs3OyFkMP1log5+hCojcPJO8nDh94pZ/4J2vB4SJ5bTXtbbisQ34Zko2juU4WOIx06OyZlrJL61t/Ub0dJOFo/pH9yTIFpbvUY+z/QfPzPv2erfshJyhnPkcY1MDy9RXSI/zNKzhG6H4aPoPyGTd8/4ZIn/l41vOo0u/74uDkfoz/ae+hCQkJCQkJCQkJCfmll9mNsixZtWPakU5+/8uzH7jp+x1kfb33s3L8bOsAvG5JgG8Vkfqxu7xm9c2Cy+diqknFOzjkwmeddc/IIXoFi0q+LDtUr+eF2lk5lXIxoPpErXLgIuOigH4I2njYkuwxf2NBnWlSbTuv5Ysbyhc1cflOfyvKjQH+3+y/a8EUUUD6qpXdTqc7Tmp8HsvjUdG7D0tS7Dvs5HweH2smi0zDrd7quqKd6+m/qBYX9LB8leHwFKE9APt0Bp/DjLOV295v7SPWsTvJd3JGX29cFcKdOgbTvM1uUi1/ss0zMKdJdcwI+Igm/BI6X3B1H8o1kHLYk4lqHiW1178EtM33A4MEd2M87hLjjduTgznrdTbvrdbfvrgdv2FqvHn/6oQUPWv/WauWyi3DZbq6qnourrsWXVN2LX4ar9lS9eH/phaTcfI/mPgjFr6wWLHjjoo39bzzx1Y/7o/Z7X3ekdWDiv3r3ba/7eJfcXmDPeN133aaZ/ms33dnm3cBC1OFc4tgnxNwYjjmBNMbFDJ4HCp4n6byT87M4DwV2neIcSdcfxzw5rxVyrpuuPOfngpM2OkZhrJbx2s7X5ItQXW7HbTKa+ma7jXYMXlbaeNpJxNz+7Dqd9mncsrolvGw+G+pkv3bYOvL17dA6yZfp9W46HUd1HLDywrfYWb5sP9sWOkD3r/kTnacdZg/I/k91LM/9KDB9Yd+A+DaUep5PBGPyXOCxAOTuOdkqSMIpGcePOwE4v4b3j9UDU6N3Dh1ArMf5PYzxdQRjmn4wYmvd/95X33LCJU/+/dUvffT5q55x+qbu1hre1XtvvRTnlLiDLiQkJCQkJCQk5JdSmhPhZVXvsqed8oKVE8/5UO8HXvulgc9e+L2hA6P14MHNWCBgEo7JON+9JBN3TNJJyPG/5yU5J3fQJcw96ecCobEY5WLRFwxuw7SUa14WdMj7YiSTKAb3Z4tIsfNtaWdwXwJpX/1nPx0o2xf4wkrzuYwLLbeZG7L47NSLn3KBRmRyTheBmvZyXxzOQlpQWj3pE5HbKNtnWvNNWyflZHEt/cu/U/a92ervKe7wkDYtXcL8+LFL+1L06qtz4altIW8km6NBHpi/+f0D2C/yHqbCpmFf6HQfY2u/0e+UI+QOuWlii5C6ilyWttI/tretbqPdFo5JK5FywJHtdRsYBIb45cNDO2o+ZrnxC5cDO+vBz15Yr3nT0+ru+637dtWz6DJcoacAiwA+orrCwA8eUBeL2vmFhB3v1uE+474iHrHw1MHLW29/0d8MXb9tpn1kW90/OVa39myq+669dKbv2ktm2nsw9u0dtTvm+L65cRx3J+gKdBJz5TkkwPlq6XSd4tzwa0+uP5wzDRKMMF15zs8JOe9gJ+MUxt5OYi6VZZ3XzTr3RzvaQ4c6MgZbmfowP+6zKHebMu+6uWw7kfYN9x11vv86ymUMs62X6bXKfcn9apjLfhasDW8HOq/XvGvOypNfreO2ZX3vC8F+uP98Xmj9VM90uQ3YOqBPaYHtf0B8enpO0FbhsZSxWN5XeTTAxv+h5nfSEX7nnAD+BCTkhJQD+JEIIeb40YhtMyMHtvxoaO+mfx+4+sLvjhy49HtD73nJNQseuu63cd31AHNJjGEhISEhISEhISG/NOKL0lIWVisXbVj6mw966YrRZ/5p/8fP+/+tOzSOyfYYFiVY4AHtfaMz7enRGU72uYjkJD39J52Td5/Ic2KPRSsf+8qTfywQTM8PQqQFgoF2+mL7Erqw0AUFtrTzxUZazEDnCxNuTeeLEidFNE2ovS5mfLEDHX1Le7l9WZR6PStjOrdvCzPT5fZtgTUHZAHHRZ3DdaYvfaU6QrIhzcVhuUj0eoZh6BVMK0pirtEu2hAiwLaSNj/lgrKsJ3qWA7pPHWqX0mbjPnRRyK3/Luqxj4WoQBn3m9mmOrL/WU+h+pxP5UB+hFTLxLfkHdB7O2ybj8vyzjnsH4Wm/YXpkhfb5u/M5wtJN7aNdIL2xc9j/1CDn0Py7jn4beF4kJxrH9xRDxzaXg8cBo7geN6wQ+6MW3/7jh+2P3XeN1af/9TDix6y8ZPVqlUfr7oXX1t1LfkdrGWfYNfqfELyKRa3TeE+OdpddPyC6xnVwoU7lz5k3R+vPufxf772/S+/pW/PRf/Zu/fSum/PJXV7z6a6xQ/cTI7V/ZObZwi9Y04JufJc5LmWSTkC505KAzgX5gTqdiL7tPOrsNXz3Ov7OZrPQ70G6If2bmeQvJbLuEZ78zkLPH+TD6+j0HzWpzFa/FGv8HGGyHVnl5XwMS2NJcW+87ISopcyg+0PAcsNpT3h17PmqUc50qlfnXVZnvybH/ZNbKwO9aJTNNp0W68L5BjCenmMkeNiNt6H1E6JZJ/3r4L5Uqf/VNMPmpTnrae13M8HOYaGRMyhj0rOba1H0KYAfRBgfOOXW4cPM0aN3zly/US97oaJev2N2+v1N+2oR27E+Lf3on9b8+YzP3TcU+//4mrNUj7iyvdg8gMRnUTdXHOUkJCQkJCQkJCQkF8o4YS3uVA94bh7H/eSx5y74ornXbn2w6/9m6G9l/xghKTcobF68JASc/JVtn3jM7744Yuh5YurmMQLpktiTiFkG4GJvEz2Jd1BzJldtm1CFw22yOACROALG5Rh63pdEKNeoWsueMwednMSc2W7UpYhv4FbWyD5Qk3bKOxcJ31RWwH0aZEm0P1YQu3Mj9l5HSHkgPK9QbpQzMjknBJ0qYx2DbCfaBO/yb+aS7u8X4mOPMsbcDuWFeWet7K0zy2v4DHS45TtShvW6aynOrkjSY6XoiTm3K/XS0B/5NzA4jF/sdD6SxQ6vaOuA9gP5fkid8jhXCWUgDN4vtDrnXN8tHUC+e3oH44j00cm6vbN2+uhL+6u133+snr9J95Ur3jZr3/7HsP9H626ei6wDzhg8drzqKpacDrSfXKtzi+xoJ0t8+0TjoP9wAOAk7jtuVd7c+/bX/S1wRsn6n6Mfb2Tm+q+vZvlAxB9ezbf2Tc1emdrP98vt6Vu8VwooOcsYESwXhcKyeP8mQtOfEgefhuAP+p9/PM2yvoOKWvUL8s87Tb02WEvfguwfvKhukadwqeXZZ2B11enTuxzXn5X0VcijUkC6MSP5zGGlRB7z2td9TO3vZfxusxjnKbpx/vgdZMP6NS/2mm7XtfqMC39pa3XtTLPiz1R2vtYM8f+8HoC5h2so/Z+DP34yPmS0kq6CQpSrmmfdQr6BOA77V9APggBPbcjyJfEnOAQ7A6OzowcQfr6rULOESMA5hH/3f+Zi/5pzXtecdPaXc95/9JnnjZaLVvEca7zIzXzkekhISEhISEhISEhv4CyrOpd/vxHvWr5775gb/+B8TuHDo0KIeekXPvA6MzQIXtM5SAn3QAWBIlMwiKrQcxxMo+tEG6c7EtadZq2u+asXO6uEz0gekVaLNiiw9MCW4zIe5tMJ1v2h4tiLhJNp+XFgqcTxQJGFyP0WbRlyAtUXXDpAs8WSm7nCyXrny/A/IuqvrjJi7K84Cn1Wlfr++KuJOZUR4JH63h/CCfm5BEs17sNt4bOhZe3R/Al5/kLhGZjvqQvnmaZ6QTuR/YB9qEcn+a+VBvVM510ab9pvXSMrVyPvyI/QliSItlejykJOysz39JH7L8GKYf96F9N1TvosIU+f8CBaYD7xftHsG2cs/o4K/qCc7aELIDl4yjYdzz/uS+nt9XtQ9vr1vVbf9B34NJ/bH3wtV9rXfjMv1v68Pv/fXXCqhuqhcv+ouqSDzmcrBfnnBLE208v85F1p1SLF/3xyqc84nsjf/yqf2l/4o1f77vyTd/pve7SH7X2j9X9ODcxTtZ9ALct5PmFX/3Kr50b5dbOPT139PyR8Yxb5JlOpJycu1rukLyd01qPdkV9lnGLtuQaElutJ9dBaSNQvVx/ySb7ytemtlHWy3XcR7Os1HWOF+X44P3rzJc6IvtAXuo7vIx6pKFLYxxhNqkebUwv/rwNKTN/hrIN2ZduA2jMsy3tJe++zF/SW1sC8+t5a0P8m17PEdcBjX2Z/Xvb7s/tdf/7PtT9WO5Xjkcae1EmOmuH9cTWfWi92WM/28UW+WHEhiYxZ1sj54Y5XziMtJFzwwLYAMOHd8LX9rp3cvRH/X9+9k2rX/mYy6r+lc/GdUdi/GiPuIaEhISEhISEhIT8wkjnQvQe1crF65Y97cEvWD76zD9d+9E3flMWAAdH6wHBGNLjmLRvFmLOv5qp0Im6LhZsIWAT/gQh3oxgS4/2GZh2uI62JDO8jkMWDgpZbBBckKDtuVD2xxct8yE9Bgl777u227TzxU4D0p4tthI6dLaAyu+Lc332kxaFR60LCJnkPkroAjCjqWNdX5AOe5p+52pf9M0y70Pp0+G2+c4Tzauu85hBX2DO4yZERt7P7qNxPvD8sHcd6nHqLC/g9a19XWTDN87hxt1x3LcJStIpTOf1BJamT1vECnAuk6TrmxyfaU3hvJqCjh8M2IvfMwm7QzvqwZt310OHt9drd/72NxY/9J5/Ui1f+pqqe8Erqq5F53Z3L3551b3wWbgu+SGH5XKFzi1BzP30Mh8xt7Kqup9cdfVcWC1YdPaCkbWvXnv2k/5seHrsn1u37Kr7Dk3U/Th/+g4B2JKoU2JOzzE9d3ku+/lisHNaQNJDwHMd4NbSej5lEoX5XM/s/ZwmzEauG2snjVdeJvYG+hDSpSMNqD3bod8OX6nc6omuKEuADtdIGiPQpzxemM+GH9WJHvWbOtYp6jE9x/6U8UvaMXidwk7HONYvfBi8f3OOwYT58LL0e6SMW5ZbusNO0yw3pHQm1Zq/y/0otF8O2liZ+Uk+5JjZfqWd6eSccTTiLe0c7qd5PARoR3+v7j/dj1v1sVYCZQrocD0IMF8YwfXBD0AMHyZJBxxB+gjmEYfYf/Zrom596pJ/7nvXy69v/+4LPrPkMfd+O669M4BOcq4biDEvJCQkJCQkJCTkF1b6VjztgS9YsfU5H1zzF+d8deDqS78/xBeGy0ce9Cus/qW1Jimnk29P68SdE3pfAEBnpBvvgMsEHCfjbuM6SzuEmMt6gSwWFL7g4OIkkSyAEju+pa3aiS9ujwJ9DJI21mZqp2grwcsLnbTHfpSYS6dokGaFn6MvCIs0iSLUn03Omb9Z+exDFlQoG2a55HPbqX2B2pdl3n7pV9BZBzr5Dab3RZ+mDfIbCZQJVFfqG8fPfOg54WScoVFm6ZS39lwn8Ha0r3p3XAmUkYQ7uL0ePAQYOSd30ImeULJO7OFDSV3U4x1xU0hPQod0i+DHAnCOrzu8rV6PBenAJ99Yr77gyfXC0zfcUS1c8jFcf0/Ty1BkIcAPOHDLhWk8xnXXCfc93z3HY6HS3X3W8ic94OC6T55bb/zcFfXgLZfVfYcn6rU4Z/pxHrQK8CMfnddMPp+5LfJ23uaxBXk71+ck5gRml2zNxq4n8Ss2SEtZrufXQspbeeoD0zZOi09uxVZtxM7LvI75ymmOJ8U4gT7N2g/mQ/2orlnm+VwnlftYUehyWwDblvGoaafjFfValsecor758uOXYD4Es8qsPOnUZ+k36Qo/Oj5hHHFdsjF/svV+ZeR2tZ76AeTYYX/K8VRILLbj6ShjrBxTSRf1Wc9hbXkfEhmHMpJxvFMuE3OE64u75wT8xx5JOZzTfGftAWAKsWjvxMzANeN3DOwbq9dfv/kHy5/1oD/C1XZ/veiSBCkXEhISEhISEhLyCysLFz9w+LErzj3jd9d+6JxvkJAbnh7D5JuE3Cgm8ON1e2rzjLx/DAs+TqoHCUy4y4WWLzqok4m8LQRksi/EnMMm/j7hNzsB8r4oSO8OKxZ5aeHBNgAnV0piTvRcZEg93Uod8yd+WVbYqZ421ClSmyhLCxn3xXZKG/PXWLwQXEBRVyzQs02TOHNIudVTOMlWgnbY8pFYs8sfefA66qeE9sd8eN7Kcr/MFr8nQfIdNtwW+yDte4EeE9+Xug9RLvsXadtfYuv7VOD7uIS3gXQ6BoQScuKr8J+QylynaW9Ljpv0mb+F2440j9nB7TjfCSXmhJw7hC3JOtniNxbHdnh6a71+CpjEAnQPibkdsNlVj9y8q25fd/EdKy9+2tcXPHjjVHXcce/HZfcnXV1L/7jqXvBSXIN8t9mPE5JEsTA9NsI7c45GiK6ruhadXy0/7iOLH37Sn5/4+ie8v/2Bsw+29m35Hr+yq+Tt9rrNLc8jO9d4rfiXLP0x/yE+Is5ynpd+bhfnuIw5Qj572s5ph53XUied21pXr0PPsw+sk+2lrthrmfuQtNihLPn3PMu0XNqjH6lvZdQxTRu/jqxcx58yr9B67E/2NTewr2R/GTi+yG/UtPq3NlK7LLO862RrPpDX/lpZB3QsBWxc9XaSjesSmvUVtCdsHGXa7HWMtG2Rl3LaYuvjL3U+rktf6Btl3n/5DY39yK3uW4HlBWan+11J31Qmdb2+pr3NMiaU5JsAfXTIuS4wnduIPevj/Of4PT0+MzA1OjM8uaUe3gsdtoOT43Vr+pJ68OOv/+qysx74lqqn5+G45jrvGo6xMCQkJCQkJCQk5BdKjl84cPwDlv32w954/OXPvXLguov+c3D/pZioj9aD08SYTNzTRx0wWZdJNe9ksgn6LHAizwl9WtT5AqADsxYCCllkoK68TB92sihkmfgzG7QjiyBM9hWFLvlU/+6PvuQ9ZPTJMrOVRaS0Q8CO9lLfAH3Z57ToLG0Abd/7w4UTFyhFHihtMsGWkRdduW4i0hxcJGLRL+mCmMswHesXYF7tc7nasP9mB72AZehr028uT/lGXf423Q8KLjJ1X8rxl33HPPc996HB96lD9GqnYBuAHUMB9H58BX5uOKyOHk+ztzK9M0WPxVFxkICNgGQcfh9JuMPb6vb1hH5BdeDItrp1cOsMjwcXnuvgdz3aWH94e33SLVfUG2+4vB762EX1Cec9pe46efBfqu6FH8Oa8jW47k4B1gO8I2QYWAKE3H1kvkX/MmAj8EBgHXDKwvu0L27/zgtvvecX3loP33J53Tq8s27znMG1JueckB/jdXtydEbfv4lraxrjavEPi3z3ZwHR2TnvkHO5POepRxtybZWAjY3Tfj3k60i3co2YPx8jCb+m/Pop837NuQ+Wa71cX3Rlmwk+Ziikn14HKMtmw+tyy307G2k8K38z056XNG3MDvB9pb/T7Axp7CvI90Y/XJcAfdneLGgffSxtEHMo1/En+1Lboh/FeO8+tf+AHA87VlLmx8lgeQHtOlGWm29Batsg/VGCrZOQUxgph/6UECI66W0+gfN7iP8EnMJcY2pzPTg5ijS205uAS+v10xf9ywnnPOHPqsU9j8Z1Vsp8xHlISEhISEhISEjI3V6aC87Vy++z/PkPPXvFzud+uvXxN/z94P7NmMBvrtv7RuvW5OaZwX1j8hiVLCIxiZb/rvvkHYu+TNJZHnol5ZC3yX5aNEha9dRJHpN7XQDo4kQWK2af7pgTP/SpdmlBlLbQw4/60jbFtqP9TMohb0iLD28Hdq5P5Y0y1gfk96JMoH1p+lWdLKCsj51l/BpnIsmMbMuPpmp5tlHkL6xqXm2Qpz9vi/6STvuoCyrTc0s76Ysultw+l1ldL/O6omde9epDt+n3W5tyPJN+9n5zYq58J1fSoY5+rAE62f+oI8dAj0M6zih3Iph3H6U+2p1IrOfnnJ4TzPNYed9LX4aD0JGYQ/9JzLW5PYT8YWxJxt2wvR68YUc9dP2Oevj6iXoddOsOb6/XQbf+5ok7hg9v+u7Ah1/9tyece8bnFj7q5L+pjj/+i1XP0oP36F7+4S79mAPJnLmE12YsNu++Mt9dOqdVyxZdvuRJ97+x/bYXfG7wo+d+sf+qi/6h75pLvj8wPVYTg1OKIWBg7/jM0BTGSn4QhI+pyj8hbHwFMjFHPW303FXYeco0z1/Xy7mMc5TnK68x5nEtNa43O8flmkz+CL/GPK3+0/Uj7amNjmG2pX0qt7T7Qjr13+r4WCH1vYy2Yl/ozKYkpSRvkN8kNqanHdKpjL+X7RNMuz39iC+tp/tB2yyhvnPb0r749brZh2ytzWTTsKNNUe5500n/kJbjxq3otZ60bXkf21J/qDdov5ku9zXT1CvBq8dF0+l3+r4HNHZDh/5wf9KHjKXlbyjaz3e/YQu9AH5IypF0I/mW7g4lmLb8MMFymVtwjqH/CByYxtxjz6V3Mi135qNsw/Xj/3nCyx77O9XChZ2PtbrEnXMhISEhISEhISE/V8KFJf/TnGTBfQcffNzFT3l3+8oL/nmYpBzfKQcM7h+Tx1flbjm7syMRc5hQ5wWAp0voRF8WAsjr4scWH1Kn2GJSX0IWKtTTXqALCK/bJFSyL4W2J/2z9rUPZXm2b9S1dvS3WZktfrK9+pc2CPRVFkgC2ms9ptPijORZSqudLuQUJenGxyQJSXs9+b15MZTtNZ/8ml0JXcRpf8Q+6RT6O0v77Fv7pouuRt0E05t/OX6yn2zf+H6QNrLe7XRfWh2rr/st50nWleSs73tdgBqsDtO6aNe+SZtWL9lSZ8j7zcrEp/cdEILQ+kKS7hAJuq0zwJ1CyN2wsx45vKMeObKj3njrZfVJX72iHvnrXXX7L17938ueeMreavWKrbjUzgJ+u+pe/LLu7oUvQPrpuOQeACyVi2+2BDF395b5iLnVVdXzGBzjF1QLu39z8T17X7H6nMe9u/2J1//N4PVb6tbkaN2e3FwP7h0FMLZes+mO9h6Mr5MYT6fG6wHACToft5rwczQjncsl/HowpGumE/BVjq963quP5FfabJYJvC/WN/GTbLK9+nH/WtevzVy3CbdL9jLGzEanXQLLyn0guox0TUMvYJtztOuQtjrrS5n1oSiTMdPurPPYQb2MkW6ftiWoY70ibz51TGOaOoOUc+v2c4G/SZH2rfzOQi9b2OL46d3w1p71Oe1r6Ze2mcuUlOMdc6ke4b4cnDfwnDZ4PsPnFzz3SdDZvGM/dIfgi3ee0s91F31jzRuf9O5q2aIn2PXmwuuxMacJCQkJCQkJCQkJ+bmTxU95wAvXvPvltw4fwILxAIm5UUzYlZiTx0wEOomWO+eQ5gRf/rOOiXgCJ+RlHvBFIMt08VaW64S/sUiyfKqPdhLMJtefDW9P6xT1pWw2tA5Q2gLJnyx+irsYZEt/Clk4sb+G1HfR+6KmINkELNe0l2ciDukGOWd10AchnKROJs9yG1ou+9QgeeuXwHT+KJTW4TEr9FaWwHL5LYWdlOXf1/j9vl9sX2m7Dup1n/mi0PergPWlTPsmeejlzrl0bKAzZL8A7Lktfw/9dB5PItdXG60DUCe2Cr1zD7bsk/zOiXr4kGKI2xt21Cd94fL63l97W73+4Fh9wpZn/3fPI+7zH9XKlbdXXQvHcGmdrFdYyC+pdFXd3U9d+NCN7x94x8v/fuTQ+A+GDm+p23s21S1g4LoxpDHWTgK8g66TnCN4LvKcFORzWM9dnJuGhp7XQgJ0PIftOkjXp+eTT2tL4H61TbleTJ9s/FqRcvdjZQTv0iv85Pp+fXp9Reof82aX7O23pOuUaf4uK59Vz3+7pA1F2q/p1Lc52nRIP91fo57bqF6BfpGUM2Iu9QHQ8bOoY8dv7nzW5XHXdGzHyuQfVJLP0HHP8/xdzX2rx6EEfoscS4IxXftZju/5t+S0l5XvjUt16acA74xTAk7LpB20p/MJpn2OYbC5h5xDvEP58PZ68PrL6v4jO+uNX3trvfrVj3/3PHfOhYSEhISEhISEhNytpfMuj55q5eL1y57xwBes3P7sv2pf96Z/Gz6EBSJxQP5jfad8AAKTdyfmfMGYJ9fc6gJAFirMc7Jv5b4YUJ1N/mHnCxt9rw62hOvsUStfyOnCyYgZs9FFCdJFGwrqrD/WXsMPwcm+QX0bvFyg/ZR6XPhg4UFomr4U5YLM06zvdX3xIu+CE6ItL2gaZZ1EGx9TPUqdRPLJAgw6WSxxQaRpsZE02zdY2hdW6iens+8mGgtC+03yOz3dCZSlfSP71PdHuV+xFTvV5brapp8PrhcfdlzThznseKW6hPfZ4Mejs5xp9YnzF1v9fQDOcXl80NCa3jzjbaT3JQH0PXzzznrDVy6vN0xv+k5ry28eXP6Y+3+4OvGE91cLlvxu14Jlb+3qWXax3jn1E935FnfH/fzLfMewDZxVLe7ZtvQ+Qzv6zn7ie4Y/9LrPDx3EOXbtJiHp2ryTbopj7ljdf93mO+X1ARxrcY6mMTXBzlk/nwvIuS/Xasd1QHtH8se0XgfNdrKtjo3c8ppBmV1Tfu1Sr+O6pcv65Rgrtk3fjWvvKPDfUGLW7+mAjjmeR51yX1DH/nsMSHYZjXZSW1pX4xXLMhI5Bui4zXz5DxWvX/SjE0VZKrf+5HLzk8oNqY8/DnP/XtHL8dd8PibqX8817pdc5sfO40PeKtL55IQcfOv7FFkGpK3aCXlHoC/6KC3OfX6l9dDWunVg7M7WwW13DNywqx645S11/40763UHNn9j2XMe+vaqR945t1Kusiyc5xztjtaQkJCQkJCQkJCQu1w6J6snrHjer7x29Xtfcbg9dckPBw5tnhk6PFYPHsaEGIvGoQNjmCgbMcdHSkhcyGIxwwk6Tu6dMEmTckAXdlquCzja6MKpAUz6laxx2GKuSP9kxJyibF/6Nge42NB0uZ0bYovFkKAsg14WRrJYYr+o1zZ1caNlswg2LmSkjpY1FnJEScwxTTtDw04WRIQvjko9dEiXxFynH68joI71HWaTF2m2td85J1CW9o0c52Kf83h4WhbsZV1ta05SzqDvm2vqcn3A+pt/n/5mPz6ih50sHAn0T/O6MJTzBQtFfmykf3J0pjU5PiM2XERO8fjtqIdv3V2v/8IV9YZrxuq+S8+qF5+y4fpq8dItVdXzRFxPDwL4IYcRw/FAkG4h/KBHHzAA8KuSpy85bf0fbPzgG/77njfvrAcObcf5xi9eAzz3rhu9s7WXrw/QMaTxWKCdvwI7r/WrweV5D6A8XxOWdvC6lGvT8/m6bIxb1AvUXssAXrsl4UaYnYzxXt/sS5tUZn3y35LrZPjv6cRctp0ox4ZGXWtXiblsk9G0S5B2nYRrIhNzqCvjNn3k8Tz3RfO5Lw7WtfqdZWX7LJuj/WTzY2HHrwHoeVzKYwOkcwvI6aJcoP1J56NB/LpPIeUIlBmEjHPYOSNjMPzLPz4krXMQfmhn8HoC18mNO+qBm3fXg7e+pR68eUd9r8/v/v6yJ5/2AYy9v4prqpT5HjUPCQkJCQkJCQkJuZvJ8IkPPOGyF/xp/8GxeugGTIKPYGF4cOzOgQPjM3zh8jAWLwJOlI2Y048/YELNPP+rLWBeF2ZOvvidR/4fcZ2s6yQ8ETRI+8v9MzGXiS8ScbxDSl/+r/BFokDslLCTMuatzNtKi43OhaTbJ6iuXCSltNRXnd8twTpiw4US+yxps+XvM2hdLm6UmJO01ZFHVrmAIzEnQFljYZcXd5JOdqpXf2yT/tgOFjWpz9RZXU97/wobXwg6VK9lapfr6bHhby32m/9WzxdlaT/YvlY/lhffClnYWntJz2NBu8KHnAMNnzxeDtY3YP8psD9kf+Z9pr/Hfgc/DCHnCfzivBZM4XwToHwvyrDdeGhnveHA5n8d/MjZXz/xkt+4ZfGj7nnjPU5Y8eWqe9Herq7FW7EwfASupgV6Uc0pXCgGQffLJ/Md86dVS5Z8bOUzH/z1gbc9/5bBv3jDrf2fuvjbJOj6p3g+4lqY3obzHZDrSs9bnr+zrtk0Lmg50bhGmZdznteVnu9+/fBa0vEZ47XAriuxzWmta3q5TlHG685txAfT1ibHYStL13CyAfw6B3Tc4W/R8Ut/F7bSpsL/cSC2hT6V/0RlbKco87aob9hpGeH9LP8h4/s47Wv3C/j43SzvqCfjkqLZTme9zrrY2n7Q2FH015D3s+Z1v/G3MU3CSwmwVIe2QCbVaGt1ZL9oWn6btIW8QT+so+03CDnm6RPngN4hp8edY21JzKU4LW3QJ99Vx3YI1McchK8LGLwe4Id2SMwd2XHHyC1X1MM3XV6P3L6z3rj/ku8uvFcbY3A1qJdVSEhISEhISEhIyM+XLF76ayc9eeXu536U/5EWYu56LAgPjc4MHByfGTq0pR4mOcfJMSfJnNQTMuHWdJ6IGzmHtCwMkPYFQloolHmWYxLuEFJNFnpcoOTFSCrH5F1JOqtnE3puk74DaeFBiF/z5zZlH1K+qMdFCOpI3cJWdSwrgIVFY+HbCdj4IkoWMCmvxFy+M66AkHIE29C023MRnv1p+9qOtqcLG0WjH6UtIb61H5IWHX4n6xl0waX19Dfm3yqLb9l3tm+SraWtLOULSBuy9f2mfck2PBY8b4p2OpFstb4uQvX3CLCfnLTIOrRH8Hei3tA00nvHZwb2jN85PI3ySfidArAgXPe5HfXI57fX7Y+9vl5+5sO+UPWufmdVLXglrp2nVtXC36qq7mcj/VCglxfUPBKPVv1yynzHfBjnz1OA51c9PU9ccFL/q1oTv3nlvb/0lnrwpt04P3lN7hAM4Pxt7ds6k8eH4r1e6UMtdg0L9Pri+e3wa1iulfJ6snGP6eYYXdZVZB3KScrZNZxh16z5K6G+3Q/SXoc+OsBruOl3NlKf5ygTlGX0W5a5ztoqx1Ld5vImOG74PlZb0XG/ug1JuQ5ironSX4bsjw7bNM6KjerYtzzOGViXsH2t+8V+f9FGZ1xIxwBQvyxDO/TP32RoxBjP81jOgh5ntl3eISdgXZT5/MH7Wvosf5P8XuxHEnNDhwH70M7gkR13jty0q15302VIb6tHbhr7jxM3PfvT1fHLOS7fF1gsV1dISEhISEhISEjIz4H09txzzaNWvO7xW47//RcfGDgwesfQ9ZgkH8Ek+TAmzvwKWgcxJ19ek8k78rItJ+L5zrm8KFBIXoDyZM/JuEIWkEAbbfDrm0LUOFgONIk5BRcCTszJYsTa0va0nvpXX774Ke0aaYPUsXpSR+qZLX0irygWEFhUpMUZbTphtrrftI7UFdII6WLBrXm1UcCGdcovpdLW04SUM63t6eKLbTDPMvwG7BPtj+WZ9voF/Hh4XnwwjXpOyrm/7JPIPpJO9nFpk6H7ldA2fZ+qDmkeCzuOzWPk74Cj7+JcSn3Gb7c75QbkpeET8iXVAZzTmt9Wt3GOt3Bek1wY2retJiHHR1WH92PBd2T3ncO37P5Re3p05sQrfmtm0WPv9Y/VquNur6rud+C6eRrARxFDQv5f5Whk3fJqyaLzV531yK/ee3K8PukLb7tj4MD2H/bt2/rf/Ye2zQwe3oExeTvGZ5yrOIflwyMYB4ZtHEjXu0H+0VGMW4rONFHeKYe6vKbl2lK7dF15PbnuzVbK/RrkNqfpT8d6pOW6JaxMrmsr45Y+ir7LuGb+Wc/TJbQvRylLbc0DaYdbGzMwDqW4JmOZ2sjdbckO4wfTjbEOOvPn9gIvbwC+ZGwq7AH5LclHtnd9Ho+1f7NILNYl/NjYftF97H75O/gbWcd0Xg6IX7Fhnj7N1srUTvVyp5zdbSx3wLG9EvAr9axcoWW+1bYJ1Ge7qS32D77ZV8Y5Iea218NHttcjN5CcQ/pGANsRoH3tRT9YPzn+7eNe9MjPVAu6X47rqF8vp5CQkJCQkJCQkJC7nzQXg8sWnLLqlY/ddPwfveRw/6cu+M7QYUyIj2DSzO0hbA9trYeNnNOFBLY2aXdiTibVNskWosQXd8xLuS0qZk3cVZ/KDf7+sM53iEmZtOG+SzTtdPFAoB3q0F8h2GRBUZYfBewb+2t58cuFSsfCcRa8vuR1YZEWTXO0Xy48/O6KxmKEerkDhlC9L4oafouyRAwCbpMXWbqVBXHx+7y++1MS039zLnd73c/qJy0Ckx6wug2d6M2PoPALSJuw8d8n9uZfYecN323oaJC8BOvz3LRzFOeskHFHiIm6DfQf2jLD7SBxGG3h3JaXi6POyPU76w1ffGu9/guX1/3veNl3Fp9x2sFqzaoPV12L/qiqesaqatG5VdX9TFw5w3oBzStxd1zI0YTnRbcm55SH4XzbVK1a8XuLH3LSxJoLn/zHA5+84PbB23brVyl55xDO3yGc18M4hwUHt/N6mdGvKfMa0mshER+8RgvodaPXmOf9OpZ6vOaSD8Cv0cZ1TXvUlWtZ6zbHZQP9F+mUL9sj6FvGWPdZlHXWF0Bv/WzqM0RX+M/+FCmeSRnGQNHbmCl6H0MB+4dJ7luxn0XHMgNJtwStm+qbvvzHi+7bAvApQJr+Nc+05nObjkIvtoTtM8u7bfP8yHopg58EL5f9rPkGcSZbL2da5wRJX6SHS/tiS8yONbplH2RrsXD40Hac68D1SANDBK4Dnv/yyo39m+v1V17wrZ57t7bgGopHWkNCQkJCQkJCQu620iAKeu7V/8Q1f/CST/Xt2YRJsJFvBSEnCz7Jq07KZfKuE2uZXMuk3BZBRsppHpNtljHPyXdh5+Win4UtRyHmdJGV4PkOuzTZF2jbMtHHwqDTxu3KfNJLO5Y2+OJlFryMW0nrosIXN7oQ6lgMUZ/ssMDANtfThci85JzZprz8PoX2y33p8XKbZGu/TfLJVtOJWLO8wu0Vsk/sbgk51jwerifcR4cu7TOD6qxds2u2UZ4zQHpZuOedoFNyLYF++cJwknJYwA1cv53vJapbhyZmhrDl3RdC2qFs5Iu763V/e1k9fGCsXrv9+XXPI+4zUy1bdrjqWri9qrqfikvlfsDqAvGYVMj/pBwHtIATgUVVT8+vrHzc/f9o43UX/2DDV66oB27bWbeEbMY1S5Lu4Hac7xgf5NrmFiiuzXSNCoprByh1Op4i7eOmobx+Z13TglxHx4EC9C1lBmuHaa1n7cg44G1Cn9L8TZ6Gf9ZNgI79py/3Oxesbuc4nGIZAb2Xqy6PiSX8Hyjqx8C+UmflmYSDXuB5h+oEVq9ByjnYT/rGVsd2pmFv/U9tEqk/bl8eC82nfppNA1Ymv1HAfWJlPJfMf0nIOXw/dxJz/ugqMYuYS+3o78qxIevdJsVCIeZIxCHtwBju+7s9vbkenrr031e89DG/Xy2oTuWFFBISEhISEhISEnK3l2VPfMCLTvjAa7/YnhzDxHock2AubkbvHDm0tRbI3RhGzHExkSbPecItE22bmPtCgBNvXfChjHlLa5mWq87qYFs+psq06gqYndvnLf2yHfbLJvrony9E0oTfbQCv4wsgWXRI3dLG/HtbQPrt7hPQd+IppK6Umz9fjHhe6lDfYWeQfen5YtGmC49MzKXfm+zZj+L3mE788bdhy7ak31avvPtCykyvyG0o0dfRJo6PoNClfcVjx7zovQ310cibrrH/rR79yPkk54HmB6bHFbwrk+X+4nDTKzln/eLvJCkn2F4PHcGC7sgOnMs76kHecXQ9cNuOuvWV3Xf0T2/6xxN3nPX5xY8/barqXztdLTjuhqp76V91dS3eXFU9v4bLZIVeLXNKfPUv5GclP+Yuuu5nd605/s9XPPfhh1pvfsG17b96/fV9n77gm317Nv2gLWMFzmlcR35dt6e2zgihwpfuywckeK0onMgur1UZ34rrN12LvC55PXF8STpuqfcy1GV9lokPQq9T8WugztuWvNURH+LP0oaSIBM7B69/AdNlm5onGo9WWj0hlcSf+k7+rL3cBxsrJeZZ28g3+iP2BqmTbXKeduarBI+XAGO6AMdNoGmv5/s6H1fd783jYD65lf6jXGD6hOzX+532P+C/q+xrIiptq3Zmy7TUzcfCdXznoRJ4Bcy3+td8+c+mFLeKfSMxD+lEzB3EWM5Ht0nGGSGHc2qGkOOOvrSvu6Ruv/81h3oeehLGb3nvJ8ntkJCQkJCQkJCQkLutLFv+nIe88vh3vPz2gekxLNgALLDa+zfrl1gJuXNOiTl55A86fcwIsEl4+vgDwMWRTtQxSXZI3nSFXYkGAZdQEnWKRj33w60sEAyY+DcWJqZ3m3LrcJsS6reJtGDxRYTlhZhjeaqPMkeyV9sM6pqQhc4c+hLy27ydOfzIAie1Tzvbus1cesk71E73Idsqgd8qx5W/t7nvZB/JsWnqBWyjA7rQzH75W3yhKYAvX1jrIl8hi3oh4+ycLdBGGSF1xf/WmREs6IZ9oXfjznroy5fVw1+7vB7au6k+/sJn/aDn5JGpavFS3hn3TCXiFj0N28fj+jgN4HuKenixHEWClAv5WQnPpfnOpxGcir8OnFH19Dx66X3XvWJg07M/sGHf2D8O3bBDv97Ka0WunW11a+/oHULCdSAR2czLmFVCr9d0DRLFGNO4Pq0sj7Nsm9eqQ69D7ZNBdAbTeZtC/oif7FMJIIBb9sFsE3wsEqCe5FVXfliAfS3rCTkk6dy+jqVaJu0JIZT75DZKLGVd3keE7qe5gXIZhwzQNf7ZksoKXfKNfNkex0zoNa16gdVTMF9C9XPtR4njVlYSZ0qqWXmC67WM+7ss5/tn9e44r4stfAlZV/g/Ksr9YNA7yp2cgw8Ctnoc7dyV443zanK07rt69D/WvO/1f9013Hc5rpuH6eUTEhISEhISEhIScveShUBvz73av7bi/Ce/9YQ/efWXB6ZH6/bUqExwhw6MA/zIA7Z8pxzJuENYyB0E+C6ukpgrJ8ZAWoTZIkkWSiyTBVK29QVbJlK4PTr51iTn6JdlXJSobwHyCZi0OznnCxdfxJR5rWd12T/pI8usbSPcHGIH8HfrwoCLCejYjth6G1xU5TL2p7zjMKdpA5gtFzGuS8SY5RVsx8C+lD7k7gNg/4S2LXA7t0Ga+bJvKV1ui7T5kt8l+x56OQ4E09TxGOhW7axc2lI/zd/hfmErfaAN0rKP3Td94nwjhEhgmjrA81MGyyuZSz/Yl9gOTXGf8g65HT8auGni3weu3fTdtb/78n9f+rSH/+s9BltfrZYsn+zqWTpeVQseopfGvBIkXMixlvnIupXVssUvXPXEh3xsw3te+5UN1479XWtyyw96rx2/sz3NsZKP9uEam8L1Qtidc+m6gY0+eqjwa1DHB72GBDIOED4e6WsG/FUDWsY0r3det068oQ3Jm74om3XHHCAkT2oLeks7saNjLm113JA+2jjh45KPTTqWc6twG68nvt2H98PakzaN/FECrbBnmYxfmu+s2xizbdzU8c58mM9EMsmYrTZSJ7VpOmmDdTUtbXaUuV7753WZVnisaui5T4v9mPe97+umvdoq/PFUyct+1jwhH4YSe60jW/fJ3y3wtvJv0XaVdEvEpJByuu3sk/wm6QORj3N7avPM4L6Juveasfqk295er3jlk26purufa9dLSEhISEhISEhIyN1KVvSsWfHEE89/0ltPeOfLbmp9/MLvDU5twmJtFJPucUx4DUbQCTFnkMc1HZwU2wRbyS0ufrgg00WQgGmH5E3HhRIXj0KmZHJOoeV8dFXActkyX/jg1tuxOmmhggl85wKG/ZW06XxRkgmlDP9tToJJ2vUO02dwv3ha28+LWtWlxQnTRVmjP7IAsjTsFN4PzyukXkM3ATvC2ke5/maDpAs92kqQOq5vQvqC/cRtQ08d95/k9RxIfS+QfGNRlvoyH/x4CnwxD984X9I5MkXg3Lh29M7BSZRN4fdPba3bk9Ahz7s2TrptZ73htm312ne9+O+WPPkBH65OXDVaLVy4qepZvK2ra9mF3d2LX15VPY/GNXG8Xhrzyo+7oykk5GctPN/4uPTR5OSq6n5BtXzJa5ecOvKGvvOf8r71n730W8MHd9X9e7fUrUlcS5O4LoA2ro/+vaMzJO0SWWfXtBMrs69FXGMEr1u5duFHUNjYmOljdL5eCwhJp2kp87GC1zRh44SMh54ugTaVmGnqfYzQvPksyktk2xLsf2FH//NhTruyTfQxQXUyNsKOSMSbxAAj5gqd6J2YAtI4S3/SluqZ9zItN/g+8u0c/aHeia4SWo50Qraf5UPeX9i0dcgX28Untx1tye9DXbZXQPVGzHGfQMdzLZUTRR+8LUlLDMK5N4Vze3LsThLO7b3b6uF9l9e97zrnc9WqpfyCdinx+oGQkJCQkJCQkJC7haxd8oh7XTD4mXO/2H/tJfUw3y83uVmJuelxTHTHMOmdg5jzRRoXIgmceGMrixwlTJhOZBnTkjcdJtKeF6ItEXNuk4mXBjFn+bQYEdumP4GV+4ReJvdex8s4uUe/mzrWdbD+bKievw/bWTbcF4TnrX3YZX+qywsV13sfuC912/Rv/U3pjOy/0Bkx1yiX3+w612s6L7qyfQnpixxPS5dl1EkZfrsdA1/sZ5hf+w1pkUqYjei4ZV58qi99JxYX9PCb3pOF84N3/EyO49zdUg9fu3lm3XU4T6/DeTs5Ua87fFm97tbLUT5an7DrrHrJr59SVytXXlNV3c/Huc/3xS0A+N4hfsRhFbAMmO9x1ZCQu6OQXOBHSFZKjtLT84Slj3vAJ+/5mQvre3/+zfXQkV11exLXk5DWE3X/3vEZEteDewG5iw7XD66rYVxzBMkOvQ5xnWE8a1yzB7fWbULGGML0Yu/juI7dOg7wurXxnLFB4oPqMmCHtpwkS2Oe55OOxIzlC2ic8XxRx9BZx8cpb5O/syxP7c8HjpUNnftjmZULNF+OfQ0Sjr+J5Fyhc4Iqj9VaX/umutJ/o9zBeqVdo47mcwwqYPW5zWQbt1quv51AmufKUYg5AWzVr/bHfQjst5dwklJ+u/3+BjFn7bBt+he95YUgNKJQ4wS2ONfbiAft977qQPeGtc+wqyMkJCQkJCQkJCTkbiWtVS9+zJvbt2y7Y2DPJZjUbsbi7VIs2DbNCDEn5Bwg5NwWgS9CfCGQFws6CfeFji96dMGkCx9ZtGHr74wTHbcJKOd/uZFO5WKTF3VO0ql96Vt9dn7BVftbLGzSJL8ok8k90oLsLy84Wd/hev4+qyM2hNbNvglvT+v7goiQxZeXd5YRyOeFjPnDgji9iLwD2leCaejgXxfO2rbqvbywxUJGoXn2Q8sMtq8T+QZd3l+01TrMa1mxDwm25f2wdtuy2Hc7O1cI6OTxOG4lTYzVLeKAQurzXNk3JnfOKSk3Vg/vGa/X8yMl149/v/WhN/z16oufPb34jFOvqVonfrJauPTKqnvJR6tq0etx3g/r6T+vxN0UIXd3me8uuuUoelW1esVnVz7lgZP948/+2OCfveFI+8rN3+2f4qOtuE6nAFw7A5P2bsap0Zk05vA6l2vUtrxLzu6UlutSiDmWGeRa5/jM61Kh5LlC4gFh17pc7zKuED6+I23wcVUBHZDyLGN7Xt/rcXySMn0kV0kh1tEx1MeohMJeYP41jimSb2+/QSbRHr5Fp3WTb+7D1K7q8v4EZGxmnexP7xTLKNvK42ceQx36uwxuI+2xvEgLaON9nA2PO5pHWsguK2Ne2mjuo3TOUC/5jPSOOW9XoP108tGJOP3NRdrspB2r6361XQfbQB2Btid1eI5P4jzdO173//lrv7zilY99W7V88WNxbfQBfI1HSEhISEhISEhIyN1C+k8490m7eycv+s/B6VEsVDZjcbMZE1reMTeGCTmhxJwsdorFUWPxZFudrOsixxdMQuQQ1BmceBPyrQMlMZcJOC7csGVdg5Jy5lfSWqdcsBDaX07UuUDhRB96mfCbHpA8bVnHfPhvai6CaOu65u/331bWzW34tgkn5nTRYrC6vsAQJF8Ot+/Ma33pK9L6dT/7fczL1soF0HHr+1L259ExNzFHsF1HsyzXZ1tqo3nfX3q+yDkDNAm58bpFCCHHLRZY+0ZnuMDneTk4hfMTi6718Lfx9h31ui/tqAevPK9e8cpf+3q1bs07qwULXlJV3Xx86TEAH1Pl9t4A74z7cRKkXMjdXXiOHu085Z2f9wRIRDwU18Lpi04ZPK/9uy+6+aTbLq+HDm+vB3jHnH2FW+5im948I/+AwTUqhIePLemfAQVhLiiub46bBiXmkBZSjn5V3yDmvE1CxnfPwxa+ZVyDXwHyOk4UOiC1yToExycp0zaEKJLfkOvMDbZnbXod0wuZJlstExSEGTEnMWdjbdlGg5hDuW+TL8aDTlKqAxp7mvCxV306WGZ6AfLQiR/Rl30rgHIn5wjaOhEm9bxcbBiniljV0BV6AX2r/4xMwimg69i3Cq+rfr2vs9uBDyPnpK+MVVM4H4E1nzr/+4OfecPXFz7u/h9ATDgT1wTvkA4JCQkJCQkJCQm5W0jv8Rf+xo61nzr/34b2Y4G2n+TcaD08PQbwbjkj5gBZ6MgCRifIslBimhNjTpgFXBCpzhdMJZnjCygl3uyuCl9cGZyYS3mD5FlX/BiSDe+gYnku80m9L1RkgWKTfNHRhpN3t/G61lf5DfI7bVuC5dj6b8/7xMpkm9sj9O4AbU8WH4ZZixHznRY/ojf/Bvkt3paUuZ1Cfo+nseAhvEz7ZuBv5Z1ylu/ch7Mg9h26hp6+2I73V/MlZJ9bPb+Lxs8LgRFwJOTk7hsBjzvPLYJkwmg9NDVejxzZesfITVv+rXXlud8+Ydtz/m7Jk077j2rt6m9WXQv+BIuvZ+P85iOq80mQbyG/aDLfXXQnVauWvfn4Fz3yW+s/+oZ/Gbj6om/1X3nhd/uuvviHvCtVPujD65LXLa9fIeRwvdr4VI5/clcrr2Nevw6/hm1cl7EccYQobYSkIglHO7/GGQ+ocx8cQ8SW/bGtA35STMFWxhAbg0q7dJe35Js+JF4xza3byVbLZAxDufaFZaZjuYzbtLNtGr+tTxzj3F7a098k5a6X/Wl1BEj7vnZ/HDOTXYa2kaFxbO60jruWL/1IOfR2XDxP5NijaR+zdb9AzzR0iXBLZdT71nwku6JtI+L4O+XdeuXjq5IvbA3aF85BrD0H7+hLbZgftgUwtvHO0PbU1rpv7+Z6+MZt9QnvesW/Vycs24lroVcviZCQkJCQkJCQkJBjL513WCxbfs7jxvo+dd6/Dh0YxeSdd0+Myd1yJTGX3gMkixsDFys2SW7oCVskyaLG07YAyNDFWKfe77ZwMk7TpoO/BPh0Qq7xaCz6U0In9pa3xUNerGDLPG2KOlKPk3vaM2313FfSNcpyXuvnPOELOIEtNLSMPgt4HanPvOq0X3P8PqLDz6zFnBFzok91mVak41VA3tNT2AioE7A8Q+0B+iKkLeiJ1F6G9DnV1+PM3ybH9QCONc6t/qnRmdaesTvVN869qTEssMbqgcNb6pEvTNQbvrijbn3wdT9Y/KyHTFV9qy+vuhacXVUL39RVLTrP7pLboKf4UaXzWggJ+UUQknJHI+aWVlXP41B8XrVsyesX3HfgtSec9+T3DO659FsDN07ULYz3rYMYkw9t00dVD+F6RVrg44dd0zoG27Vrd8bpOG2EW6H3O+b8Ovc4ofGhA6YXmw5CbFYdtyvGEh2HlMCRMdZtANbRcRVIfnM5kesR1p6Xex0nz4S4NEhZcx/lNgDLsyxvOV5qWkhQ5gvIeMltCeoI+f2abtSzMVj0BPbH3HZaLmW2P13XCSXFSHwhb/1XuL9Cl+ponBsW0s1jnsKJOCHmLDb5XXMeqwjfV0pUalr7ov5zH5lWvxncBzinJ7fM9O/dMtOa3la3cC4OHt58R8992ttxLZR3TUcsCAkJCQkJCQkJuQvluOrE5ec8YaLvUxf86yCJObljQgm5RMxxMmuLrQY5x4myT4xlUl8gTfKZzhP/BOh1gVUs4gS+uFOShjol3kxv9WYTcwXEpnMx4mmFEk22lQWATvIb9WSiz20HOnXms4TWNyAvCxJZhBiYT/Zsp0Cq29Rpv2w/pbz7mMNPCVns2BdapV65Rbkcowxpf5+WzQbLCtiHGFIeNtImbdEv72eGnSNynPz4+3HnMeaW/nAs926eae8d08fmbkP/v7yr3nAY5W97Ub3wCQ+6ozrhhM9XPQu34Ex+gJ7QIlxw8UX4fJQvFlshIVlI2C0ClkqOsmjRU1ec8aBPnXzNWH3Pv7m8Hrx5R92P67R1CNfcEZJz2zJ4PRO4zn28zWN4voYlXxJy2Kotr3lFGgNcX8DLBGgvjZWNOkUsItL4o5A6ks51yvI8bhY+ZtXV8Urg9kLEcUwlPG9lgOwfwuoLrMzL01bGSAP9cdx02Fib/DmkT0101vO65T5p2qnO9QLZP3kflHAiLJFhrN+A+QSEvJM6ak9irpOcc/IsE3HcogxpJ+oI2UdEsX9TX6wdBfuV/Q7uR6zgPtg3Ubencd7yzrn92+v+Scxnjoz/y+LH3n8bzvwT9AJIErEiJCQkJCQkJCTkmAsnoQt7Th341ePGnvGu/isv+v4gX6oPDO7nXXJj9TAWPtwOTit88ZUXMXkS7wsqJVpsESRl2GJhwHxamFmZL3hKvdhRJ4Avgdq5rdgnnS0qSrgvy3sftU3AJ/uc/KdFjC4EygWBknWa18dkuOjQhYU8KmN+Zr8PqUhLXuvkhZz5pI6Ar9R3pqFTWB+YTuVA8dsEZbnVZRtaN+s0X0L7l8g2oiTj5Lg5+Zmhbdk+JuSuGB5nPa6+j5ptAamNbEtf4pePGwnBh3KCX4qcxLE/ggXVFy+rez/zpq+vvvRZ1yx5/Kkf7h7s+6tq2bI/rRYsfXdX19JxnMd8b9xPIrHwCvlllfnuouutFix4Zdfa4z+w/LGnfrD3omd8YOCDrz/Yv3f0O30YR1qHgMMTdYtpXMetad59ND7Da5XkGz++4pAPSBj8OleCDkDsENLOxm4fD3wc1zzSaXw4OjwWzbKV8anIJ78Z3p6Or9DZOHg0JH9mm0hC7AvxUZJzgHwkg/qyfRn/aJ+heS1L/WA9qct6Om7qeGswe+1XRmde4hn9W77sh7RvfRFf1p6Ax6tEWdd8er/S1vtk0MdIjTwzDB+cEMhjpk6emb1CCTn+00rjZIbuT6bRDuF1rF/aR+ab0Bg2gXOOpBz9bK9bPCbXb/3PVW/4jfchhjwB5z7/idMtV0FISEhISEhISEjIXSAnViuWPGTlS371ouN/7/l7Bq699AdKzI1jYq3EnL5bjqQcFl3cIj/rcVZDnsjrYklIHZkwW5rki9uIHcuVgEs6QyLmuPCag5jT+kThwyboCWKn6XICL5CFhKGYxDcXCtDJ5F/TSswZmJcy83sAbcldF5rPvgnWt7wvLMrFhcH75u2lBVhH+ZyAbU5bv23bidKnItdNv0fIOdV1knIK2/++jwv4/vSF1wC2GeYX9YR8m+bjquMz/VNY6E9trVt7cSz3jNfDaH/DjbvrDbfvqAeuPrdeMfqUme57D3ykWrDghThvHww8rKoWnI7tacB9gDVASEjI0YWk9NGIad5duh54EMCPRZyy6NThCwZ+/8WH7/k3b6sHb7y87sU13Ytrt7V/ou7fs+XO/j3jd/bj+m0ZKT8whXGaSMQc0wSuaWz1Tjod0318z+MW0jaGy/hj+c4Y40SfkH2MDYakNzv3o3lPZ2ibAMfYQn90wI/884UoiTnzQRRju/wuotG+jotOmMnY6Dobh92fg/XUTwHqfRwlOO6W5QW0vVKnbXWi4Y+Q/ZvzZf+977O20i+FxskmMUcIMWfk3GxiDjZHIeZ0v3iebVAHWL+0j8xnyG/lXXP7tiNvOLAD9bfjnN38w773vfJgz/2GdlRVDz+M0g+EhISEhISEhISE3CVy8pIHbnht37te8dkTPnj2twYnR/9bXvydHmVV6N1yuk06WXBwUcQJPNO61YUI9Jb3xZKTNkqgGdLEn3kt94WAknIKt5fJuZUroJcJO7ee5gQ9I/fBdVyIaH3Ji31zQl/aJhtA75bTMn+UJttaH9hP5hvl2NpCQhcd1s+06MgLFO2XLiq0XVtgSJr1mEaZtKFpL/f94nr3SduyjJC+GfL+oK1CiTk+/mNE3DQW5IDc1SY6/FYB6spdM6iPOvpOOvweh7SBLRZchPYX+wb11k1uq0em0BdsB/bA/xT8Htl6R+vwln/r/ez5/3D8tmd9dfEZp37tHq3eb1cLll1XVQtehXO289GjTuEdQXFXXEjITybzXS/36VqxfNuypzzoyNAfv/wLI1de8uXe68a+23vd+B0DU7iWpydkTGiRjNs7Vg9MEqMC+cqrAOMhv4iJmCH/ZPFxw8Z0jgXp0Vgbz3XMVIiO4zdjjtQHhAjkViF36Xm5+fU7tDk+JX8yFplfa9PHOyeOyjFdx0z6Y134BDQG8OMQHOcN9OOwMd39C8yXjLXSnuu0bR2rCx+0NftGnWRj/t2P7VMlKw1Wln6z+LMOZMwAAP/0SURBVKFO/Smgb9ip3tv3cikze0fZ10TAwSbFFdFrvYYNYeSct+n9I/w1D5KGXglQt0MfEvGJNHXSL+2b9JVwe4lFxHb4ISEHHNpZDx3eWbf2jd8xcvvE91a+/Ek3VouXjuFcP1VP+ZCQkJCQkJCQkJBjLw8+7smnXTGw55Kv9F51wQ+G9o3d6cTc0FGIOW5VrwsBXSx52hYIsjDKiwXNYzJNYOKsi60iLfCJtaad/JmfmPP6sLXtrDJrxyftojc/vrhwsq0Jn9QDZtewL3Vep9QJtL4sdlwn/+3XPki5LeR0MWF9I+AvIy8+GkCduUBfvDtNt8izfYeU+x0LzXoCb8v2m3x1EcegRfDOFzsmndAFofcLdWTxpEQcXx7Pd1P5+5O4/0ZgO7IX+/LabfXIgcvr9be9tR667Yp67Qdf/19Lz3zYzdXA6ndVPT38gMP/qqoFr62qhc/D+cp3yB3tUTyX+e4KCgkJacp8xNwK4JG4Bl9YLVl45oJ797+yb+w5f7Xx9rf+cODIZfXaya2827VuTwJ7xur+ay+dIdp7N9eDe0nOIWaQnCNxJ+ODEmh6BzQB3QGMHxwvLK3jIiDjucUOAvUVjEPYSjwq9YZkj/EHeY1H9Gd+xXeRxngkY5JAySMdK80OaPRDdJael5wr8kAa1x3iN0PbIfi7Pa1jsfaRfjQtZeZD9ht/p9Qr6yrK35HbmDstW/n92V5gukYstb5Iv0znRJqQaW5XIBFzhuSD5Z7GvtOywpeRcSyftR8J1E/9SqBfArFOiDng4I566NDOevj6XTznZk76+tvq/j9+w/e7+ta8F+f5o+SMDwkJCQkJCQkJCTnm0lM9ZsWLH/lnI1+4/IetyUswud08w0dZh4yY0/fLFcTclG2p42KEkEUBdXmBkPQGXTRh8iwTe7MR2yLPRY5MslWvd8xhK3ZNG068fTJOvZBHsnWf5scWEoRO/pEuJ/L0hcl+Scxle1so+ARf0mqj9qYTH6b3dNLZIor2rksLNi4yzEci52DrQB3+Lq2vtu6rAfPjpJf45J1pdkeCtqHtEUNcqJTEnPv2NMG75ex4OfEmpJwRc1rGbQHf92jLv+TYRh+UkEP72PJOiGEcw2Es1Ifha/1N2+v1X3xLPbxnS33itjPrBQ8/+V+rVStvr7oW/R7OzqcAq+Q8VYk74UJCjq3wemsS4YsX/69lz3zIres/e1G9/vNXyLvn+q7bVLeuBfZsqgeQHuB2z+Z6cA/iBck5XO9Koikxxzuq013VRsYpMWdkDMYRHWtZJ4P/DCr/IdTQwz6XETp+SfyhH/qTcZCwPOtwTBRY20xjDNRxX20IiT8J1GEr8ch/A9P0bSDBlMZ6gmOrjo86vqMO8+xj2Yb0V3U+Lkt99qf0gW0z5rGepxXe9xJlWWdattam5z1uSLqE66VMIceA29JebJjW/Zr3t9aVmFn4K/UOJ+a0b9yyv9ZHyTNdtClpg8Q6knLAoR1ytxyJuaHD2+sNX3pLvW7PZXXPfUauxJnNd82FhISEhISEhISE3AWyqOfxq9/0lM+s+9Kb6/b0JkxexzGJJbjQITHHRQ/JOCPkHImYK+9+4KKLCwNbIDDNRUYCJsyETJytHOm54WVGBKW8l2X4o5Zu40RSIpCsXl4gqZ3qdBGgsHIvKyf3R4XXQZo+yjLz6WSXt+MLD9roIgNbJ+aAvPBwaH2+uFoJyCbcXhcv5kOIOfrKaPTNwTLx4XaoB32G9kH2pxNzQONRsmI/y50v8NO/f8tMW9rfjv7wjgX8bj7qCgwf2V4Pf+kKnmv/d83mZ0wuetg931mduPKPqgUL3ldVC95eVQsvrarup+HsbOlJGhIScoxlPhL83ii+uGv18g8tfeRJ7zrhDU94T+8HXnlD79TFP+q96pK6dfXF9cC1l9btazfV7etG69Z1m2daezbPyJghxJzGB4GRWWm8w1bIGI5NyAuB5wSc10m6zjLaA4wxNiYlHaD+oZfxv7BHez5WJ0CfyCPacHwTn+rXx8UM9Sl+kx/Uxxioj2Wqzu1T24T1V/tc9AuQcdsh+4RjKvUG2hQo25b6RNFv3VfqO8PaTnm2ochtF7aWTnFMAB0ghGqqY2AZtum4FrYSF8QXoftb/2nkaYXHtdwe4P6tzcb+kGNv5fIPKPgsiLnBIztnSM4N37y7Pum2360XP/L+0zivf0NP75CQkJCQkJCQkJBjLcsWPaFv15lXrrt9dz04PVoPH9yCibABk/DZj7LqYkh0Mkk3Qg5pgUzwqdeFgC8KdPKPybJMmDl59nKbSDPvk+qOdL4rwPWe9jK3zWVKyhG5zBcEeocDdJLmxN1h5bSVOqZPE/zOrfqQBYzrJZ19JF0BWWgQZZ2SVHO/0gcH9EAnKZd+u/sBlJTDYoSwRUlqbw54fYU+etpsG8B+TMQc3xfFx9cInA9C2sGGhJw8PrsfPvZtnZE2p6Gb1H0xcvuOevBrV+D3b617dz2/XvDge91QLVlyLs7CewF88fZ9gZOBDUAvwJfRh4SE3L1kOcCPRPCDK7xOT114/5Fdfe942T+O3Lqzbh0YrXuvurDuu/oiknJ137Wb7uzbu3nGxwodUywmYCxOkDzHTBs7bSz3u7GlXKAxyMms2eB4xbEfafGT4aSctl+UebsdkLGa5YXPMm6JT6kPf/QtW6+PuoCO96orx/VG++7f8mJj9RvgeI3tnDHC2nDbpE/9nr0/Mooy8a/I7WpZ6SPFMWuX+ZJ403qWT/rZ9r5/JI/4px+GUL2jjI2pTw7ps/1W65vOM1hGGyXlhg9tr4cO76iHjuyqh2/YXY/ceJnkT/7qH9bLnv4r11RVzxk4l0NCQkJCQkJCQkLuAlm26Iz+K553zYZbdwn5RmKOjxrOIub8q6xAepyoWCT5QisvWCwtE2SzZVrytNFyuduNaaB55xsgk21Loz0lgGxbQG203ZSm3nxp+0jb4kkWUGavCx2mbRIvE3ndig3SiaAr4ESZLgq4dXut43a6UNFyse2A2htk4ZGhfaF/lIl/v2tOy1OafuiPdaSuknLaDyPmSJZZPfWl9lrX9A24P/VNyP7G8W9NjQGjBqR9wY39zT4Pw3aI5N30lrr/8LaZ1k1bvj9w9YV/13vFC7669BkPufUew303V0uWHqi6Fk/gDLyfnojzSjzCGhJy10u3beeSx1bHLX3vcY9/0JcHdr/gtoEPvuZzvZ889x/XXHXxTIuPuU5jrECcaB0gMFYIMK5wi5ij45yOzY2PKsi4Ynfl8h9AkjfYP4RmkXQcr2QscuRxVMcz2qC+tS+6El4P6eSjiCXapsHKRWf99/HSx2O904t51TfKCOZNnwA9+9Kp1zpNvft2/7TxMgH7jr6W+6Qx1he27ld80J+ANtnOfSTSzLaS5r5O+1vTemec2osu7Q+Wc2sQPWExC0ivXEhlANNl3uqnuCbwPrPcibkd9fDhnULMDd1wWT1y4+WiP/kr/7te8bInXF31BDEXEhISEhISEhJyV8kKJeY23orJ6j4l5hp3zBGy8DGQnEv5vEDhxN+3kpYFjE7kuWgRcGHgiwTJl8QaUZQBukiwPPVmryAZ5Au2wi61qUgLPIP6dd+dtgaf6BdwXSoTu3JxYIuCDl3yWYI+sE1t059B/Xv92bZ6h1xedJTEnOiknt31Ni/Ur/bH22uiXHjJQgfHTu+Yw74H+vZsnuHXGPXRVviZQh/2sh876oFbLq8Hv/jmunXN5nrlOU/+avfJI39RLV86XnUveCHW98+pqoVn4ux7OFC+Q+5oEqRcSMhdL813zTWlVVU9T8S1/YJqUfdTF4ysfWXveU/9xMbbdtftI9vrvsnxuh/jdevQVnkn3YB9CIYY4rsnOdZgLBrG2JziDsdpizX+0QiPOxoPXJfH9zSmdSKNrYVN0s2Bsq7BY4zHEU3PLvfxuDmemp2Qd17GremtjtTr2DZgdWbp50PaX7PbykC/ZD922KCfAs9bGZFiQ/ot2BblrsvEnNbhtvHbU77014GSjCvTBHzk/rGvTTgxN8Q75uxR1qHrd9frbry8Hjm4s773l3+/Xv2mp11dLQpiLiQkJCQkJCQk5K6SZYvO6L38eVdtuHW3EHNDiZjjJHpLWiglMs7ummsskIpyfY8QJ/jU68RbFzC0IaCzOrQr75DzCXa6s84XEg17JeQSKZfKUFfS6kP8+MQ/+VfIXRqFXblgEdhCQcrMhy6IOMnPZWlhUCwQZukAb0fb9/rap/yYrtpwEdSsT52VAUrEGegbZfmRVtVLmwkoM5ukw6KGW++v6knmlaAfvWuBPqUN7Cd/lHVgElt+iXHvlrp13Wjdd92l/GjIncM3bfnvwQOX/vfad76wXvr8X7nzHsP9X6oWL/uzqlrwIpxt99STLiQk5BdEjkaaL66WLHjtcU87/fMbrrygPulLb6kHjuz87zVTY3eu3Tc+0zqMceXQdkGbW398EWNNInIwnuvXVzHe+Fifxv0cA5RQsrHVIOSP55kWwIbo1PtrBJKdjY2Et2vjdY4fuk12AvMPHzLeyxiqfmUMT/69zNJWX+o4LCb5uD8nOB4z7hW6Oev4Pir20/w2RT/QT4kPblfom0CZ/xbmsXUiTtKSJymnx7jzt2tefWk80rT7zu/qA7i1tNiJD+8j8xkaD+3uu+Idc8Mk5m64vF53YEd9ry/9Tr1qy7OurpYFMRcSEhISEhISEnJXiRFzG2/eXQ9Nk5jzCbWjg5jjYmi6QKFPixiZ3HPxUk7mmxP+skwWPLbIkXzhR5Ht5oK26zaKNPEHJF/CF1asa/XFB8qa0D75BL8TaeFQLhI6IAsy+vLfYu0SmWT0srJtbdf9dkIIM6alLyTetsGP5xX6qKuRctRZPYX5hl4XMdAZxC8fhZV31GHRDPvW9JYZLhaHp7bUw3vH66FrN9dt4sBYve7GXfXITZfVa97x0q8seeppf9k1sPryasmyt1fdS36v6lp8SVUtPAtnGt8dd7RH4WZ/+TEkJOTuLrxu53u89YFV18KLuo4/7g8XPXDd5See8/g/HvzwG460r9/xg9ahCb2DiY8VHt4hY1J7cuzOganxmSGOMxgP+dVmjjdDkwC2StIhzpCok/ij4ybHbtZJhB7GMyGEBMW47+XQ5XJFIo0sZri9jNXSho/Jc5SbjadVD8CfjqXcMm91C7jey6QN+00Skwyz7iy3tL/awe3SP6gKaN+RZj2kZ/u1dgXM62+Q34H+p/3EGJLsMrx+iiViizJsZT9L2oB0Scwlf5K3ul6nqOfEXHqsNe1T96X1vL+MWf76Bo95JTHH826ExNzBXfU9v/C2evXOM6/i+3btvA0JCQkJCQkJCQk5xrJs0RlrL3veVRtuIjE3Xg8f4iTYwAmxEGac3CvxxoVRg5gjqC+gE3vCJ+w578iTeU6qUcZJdtLlhUOC6+eAL5pK+IReFjyzytie1U0+mO6wk3KmdcIvk36mTZcWEWmhUOgETX+dv0kXVQUxRxT27sfbTn0Q6GJDdOhLJuE68+ZP9oXVtffPDeAYq0/fog7Sbejbh5AGhJiDfWv/lhl+VXVozzgWyuP1hhsm6vVfwgLn8Oa6/3dfVC952oN+WK1Z/SGs05+Fs+oEA18QPwisAZYB8y3iQ0JCfrHkOGAAaEt6QfWgxaev3zXy3ld//T5f+7165Pa3YIzZUQ+QmDuIcWbP+J0Dk5tn/K5siUf8RwDJuSnkhZzTsgSPSxwvDXrXHccujqGGVOZxzcbCAqpr2muMQZqElum8XMfpcuxWpDL403HW4L4L6NiryD4B+03umzEixQmxye25vdRxGyHo4FO2pc7TitnEHNPWtvSr2Ecddg6vL2XJVrdz7VOF6v13E7OOCfWWzh+F6CTmaKN2vr8l7f9kQjtC0jkxd3BHPXSIxNzueuSGy+p1SG+89Yq69y0vuKpauSyIuZCQkJCQkJCQkLtIjJhbf+OuenBqrEnMAXonGyfxBQHniyIslpycE73AJuhSh2BeffgEXCfwutU0Fwd5gZAf7ywWJFaW7KV+069O5JvQiTsn6jbRB7Qt+tD63nbDD8GFjaV10q9bWVSIzmALBV8ccCv1WFb4z4srv1OuwPTs3yp+CN65Zpj9tVVbfBC27wd45xzScgddOh7Zny9ghIBDv/M2Q34T3/2ExY08LoV+rb9xe73+tu3/NbT3Td/u2/WbX1r6zNNvuce6tVPVkuWTVfeSD1fVgpfgjFqtJ9a8Eu+MCwn5xZQfdxfdr3edsOL3lv3Gade1dpx57cgnzr91YHLsuzJOTfLDQiTkxuqhqVFgrB7cOzozPElybrweZjl0QtIh5ggwLqUPEcEH4cScAGNwuotO8hk6HnJsxFZg4yQgj6syPkg5UcYKheQ5Xgs0Pqre6rh/xgGMseWdczJOW1rKuTWfAv4W2hT+Z8cH3Up/xEbrNPyYbYoN88D7mvtDvcULljNPP2grE4DsT9GO2Nq+5O8v9qvYmd+0z8VGdZ3EXONYIeaRmCvL9R9LqMt5iuxbwmIit+w7YbpEzB3kP5R21SPX767XH9lVr7thd93+o5dcU61Z8SQ7R0NCQkJCQkJCQkKOsQgxd9ZV6zE5HZzEoufHEXPyOBFh5BygxJySc7oA0Do5z4m36tIEP0En9iX0Lq8yX+rKMvWR/NmCoglM4mWCrnZqm327D4dM+jt13Lo/W1DohL/Qy6LA4O2JfQFfyBhmkXOiY5napwWIQEk5QaHzRYi2aRBCzkA/0KVFkNXVO+PQnhFx/fu2zrQAJQCxGGI/uPBFev0XsIj52hX18J431cdf8OTv9dxv8Jpq6eIdVbXglViDPwXAgqaHdxvcC1gg59X8EsRcSMgvrvyYj0RUj8J48fiqp+eRi+7beu3gW5+376S/vqweOrKtbu0ZrQcmN9fDJOD4DstrR+8c2ovYQlJOMKZpu4NuCLGIKP8x5OM8x18ldrC18S/lASegiFQHkLE6lWEcbMSLjExOMTYC1AnMT+GfyGN5qStsy/r8HYwDjTYs7XZSx2C2Zfuapn1hNx+sP6lf7GunTeqP9qOz3GNM7kNG9st8jkcsI3ma8oZMzDHONcuknwUxp6Qn4THR8oyHphs6kO+YG3Zi7vDuet31u+qBd73i2u7e45+qp2hISEhISEhISEjIsZZli85Ys4vEHCanJOYKUk7QScwRHeRcSczJYkIWBOVCghNx6rRcJuk2qRcyStIFMZXKHWZnvpt2Cl9QlHlO4HVybjrWkXqGwocsHmCbwAm962lT+pZFBe0VbEPbUjut4368nv1+pr0fgN4tx32INKB6reN9EeKNODhRt5EnfMHRgD+6sw9bI+akHwZdrGBfEji2/EJim0SfPfYji9dptM2+HBq/Y/j2iR+s27/p33r/4CXfX/aCx9Rd6/q/Uy1admVVLbwYZ86DAT6qNpeQeAvyLSQkxOVo48GJXauWj68465H/sPHTb6rX37TjP/uvufj7vZ+56EftazfVgyTl9ozVA3tHNW3EnJJzRspJ/GFsMcj4m8dPGadtDMxkj+ttrGU8SGM3dEynPH1yXMY4zbgmaYB6iYtonzq366zvaeIg2iWQpo22iXpSrvU9Ts0m5rR9sbE6Gltop0hl0i5trD8J+XfL7zSkvPlMscfK9TdjK/1SeD39HRpbfH/L/vVyh5SrjT5SrDZyPLB1e29b7BCfGnfKSRxTP4mYI+Yg5mgvW49vQszpe+aGD++sR3i33GHg+p31wLtfPdndXv1cnI/xrtOQkJCQkJCQkJC7QISYO/OqdULMjQoZx/9Ay1YmvTrRTwsDI+OUiNP0wDQWS1LOCTsn4Zy0s45BFgpa1ijHJNwn+26vhJsiT+pNL7A6qYx95CTc+it52xYQO4f0g/3RNgScxGPbsHO4b4EtEDzdYVsuPhLK8vQ7CCflCOQFtNF+CMkmC0ouNEjKKRpkHN/NJAsXW5BYHSXmrD3bn/6IcP/06ExreuzOgUPsJ+rArj0Jm8Pb6vYXsVD58q566APnfHfVb/3qZPf61u9XS5eOVj2LdndVi7fa46oPB1bJ+TO3cHETC5yQkBAKSbn5xoPHVF09Y12rlr158f3Xbep73ZPfte6T5/8N/ynUf83meuCaTULMDewdk38etfZsnhmYQuyRd84R4+mRVn1slTACTsZPjr3NcVjyHKdLnRBZ3BbAeOnwMVsJMh1TRW9bSaNOWVbqO32LbZm3ejL2y9gNfeFDCTuNs9lWx3SPKRLb5PdomRBzqQ2LDfRv0Bhl5bSXurk8wfxrGwboO21dV+7v9Ju8DG3NOk7SjyZc3yzT+oS/I1X0QngiBhrkVQzUy+9lXCQxR5JPiTn5MitJuUO7hKQbeu/Z+7sH1zC2eVyLfyyFhISEhISEhIQcQyExt/NMececEnNY3GAyr4/6aFon+bYgMSJJyDlJY7FkxJxO2H0CrpN3mczLhF/LdLKuOpmsywLE7MxW76DzdtUmE3OmS7AJeOekvWGj8IVCWlgQrpNJ/By2DvctCw6FLsK0ni4AkKeu7E+yN9/svxBwSsrJXXKyH13P35qhvw8LCwHbJDTPDzPwpekJ1Ek/ANY1n+6rZeifHp9pT47OaDn834xj/zdvrQdwDvT971fWy5/zyB/1DPbeUPUs3IQz5IHAQmAl0AvwHXLLgZ/kkdWQkJCQ+YQECMcTji38QMw9qsU9v7r8Uff7i5M/dlF98u1XCNHSd83mun3d5pqEXGsvxi6ScrxjDuOZEzzD2I4AwxgjE6BnuY7RHeO9wPOzIYSW1+G4LbA4aDHK41SC1Sn1+r469Sc+i7TkrU/ZJ9L4XRqHCj+NPuR8io20Q73Sn6StTf5e8V0g7QPph9p32uQYrVuPnZJGudfxdCfSbyp0jXf/oQ/zEXOSd6It/Rb4s7vkOkk5IeaglzKJh4yLAOJjfpxVH2ldd2BnPcz8n732yIKR3nNw/g0BPUD8YykkJCQkJCQkJOQYCom5HWdete6mXXV7chSTWky2Scb5BFgm65yIk3hzkIgjIafEXL6DzibvMvlmHZ286wLB8ihTUAc7LiQMOvlX+5Q2fw6f2BN58o08twKmaat+ysWG/h7zldqyvNRRpIWHl3Fy721xK5P9rBNIXv2IjaXFX7JBHn65kMofe+AjvErQaV7LWpaXOlJ3opYvrhopxzvlhIzjVw2xFWDRIXfVwa5/anxG7r6b2gZ/W0UnpNzeLXX/dWNYsO6sT/oCFr37xv9p7dtfcMuSsx66r/uk1lS1fPnHq+7Ff1ZVC0dxdvDOuB8nXFjH3QUhISH/X4TEx9E+ErEQRS++x/HHfWLFE07Z3970jE+TOGl/+uJ/7L92tG7xzjmMaUNTExi/lHThmDiMMXCYWwf0JOYG093DHI/Lcb0TLCtQjOFenuJhWQ5IzPEty9zGwbKGD6YVnTFLIGO/2jViIyFx1whCh9Vr2DNNvfnXNmZD9Na+2He2RxtDqpPSqFPYup3H4tJvKgNmkXEHkT+oj63mMiPwHLQxW8JjbycpJ8QcgXNA7yIn6E/z+g+tHULQrduPcwbnzdBfvP6mno2ti3DunQzEF8RDQkJCQkJCQkKOsTgxd+POus33+PgkGhNhmUB7HpP/TnBB0NDBzhc/7scn7iWoS5N92Pu2sRgQW/pqwif2PqHXBY+mM0p7+LE6iRxzzOFf29U6zGtd+AR0gePpsr3ZkAWJ90PqWRr9cQLOf6e/Yy7d4WZI5Jy9L04g//nfXg8cwuLi8A4l5pAWCDlHYm5r3Zocnxncu2VmZC8WM5NbcWyBKfTtpt31+i+/uR46PFH3/t6L6kWPvO/+avnysaq7+0ysRZ6GM+IxwCOAU4A1QEhISMjPWuYj9KnfCPw68OhqQfWQBRta57S2PXfqvl96Sz1002UYy7YIMTd0YKfc/TSI8W8EY9/Qvq0zJORGMEaPYLwkhqcxBjLOIC3jucQahY/zPjbn2NXUMe0xTeDju4E6H+eTDcZu0UtdRfaveimzemVbCUWcUBRknMD0VlfbhT/otA1svV/Shrej6axzPZDasjx9CGDv9VIdpAt7t5E4mGwKPVDeJZdAQo3Hr6GnH0W2Uzg5p7EWdQtSLhFzBMoUapuIOWI/zpl929GfiXr4z193c8+GNu8QfwDAd6cGMRcSEhISEhISEnIMZVnPGWsmfvOqdXyUlcQcH7tJE2ksftKkXNNDJOASoBcoKUdkYs6IOvpxO7mzTtM64YetLJCKul4uNvSlk3n973jOi44EmZWVCw8l0rye6tzO4eXqR8s0z764Lus1jXKD+5S0+0zv5fN+ZbsE6PRdb9wij9+pRJ3dOUdgP7UESs7xTrfy0VY+wtM+PFG3jxA7dHt4e90Wcg4LD/gdntpaj1yH33LVlpmhvTt/NLx/979jQfTd3ne96tsrXvHYf+263/B/VyuW31J1LdyKs+A0PRmOKnFXXEhIyP+0cIw52iOEA9XSxaPHP+3BX9j4wdd/7577J/5+aHLbP/dfO/qfGBfvHMb4p3fMkfjRfyiRjBsmjJgbSsQc40tBbEks0jHY8xoDkJbHUAnUM7t0RznG4RQTpI7pPQ1f2hZ1CqnPfGoXadRRP2xHkeylvtlan2fHX/UpvsSOsP6zPeuv+KZO2tB2U6wq2pK2LS31rU7uJ/ZjWa+zLpAfH86Q48K6viWMSPO74VQPeyC/51bttYy2ap/ugJPjoOSbxD9AiTneGefEHG1RR+YMiJP74YOEnNx1Cf1fnHPrwnu2xnGenQ7wPXNBzIWEhISEhISEhBxDWdZzxglGzLX3bsbkNU/GCVngFHldBOjCQCbiMhnPZWmCbguHZEMUxBztm/XzBJ7wRYVP/H0xUCI/PsptAdOrr+yDZZ3tEOLfyLxkS9BeoGnNa1mnrS8gvE4qE53qFZZnGRdP+P1NKDHX5ruUEuzuOQL12kbM9R/aNkNibuAI754DDvG//+gHv66L373+tu31hi/vqlt/dd4/rvxfj93TffLQO6rlS0erroVv6qpIyC14Lc4A3iHH98fNJ1wsBzEXEhLyPynzEXMkSh6JMev11XHLzl90cvsNa89+4h8NfeT8m4ZvvUzGVsaSYX64iOPr1OjMwJ6xO4f4yGtBzOmdcwDGUo9FMu4yTni60CUSbi4IOaRpHe+1jgLxUd7Lqr68rQzU41id2gA8L1tFwyfqKYygE1uD/y7WExvvE/LuX9pw5PIMbads35FiVqeeW2lb+1aWzYLERk973lAQc4nAk/1rYHz2dCdoJ0Qc4qsRc4yFs96/6tiPPCDvJZxCP3CODP/5ObctPKl/G86xhwB8jyrfMxcSEhISEhISEhJyjGTZojNO2PYce5R1MybguqDwCXwi5nxSz8m3LQzygkOJOkljwaN6zftiokHKEfClE/kCPmlnmZWnxUDnZJzkmxNwnWVSnn1xqz5cNxfKOp6nL6Slnrbn5XmBon2UNOxK8k4gOgO/IsfFA2zoi+99k7vgBPZ7AZJy7ckxQUswriBJx/fFsT7vjuOjrId3GiHHBc1EPXzrrnroG1fUg59D+Z++sl7+gkf+czXcN2kfcngssBbgApjv0eEjO4uBuDsgJCTk7iwk7DhWccxS6el59NJH3PN31/3pK/9/G75yRT106w6Mjxg3p0brwcnNM4N7R2fky61CvmCLsTMRWIw9ogPS2Nsch4Wkwrjt0FjTzCfIeK912oifRIphgMShcss+FHFDwDIB0qmcPuHb/Cgs/lqc1nbcxuqyDfclbSCd2qCt99mhfhp1CO+b1J0N1mV78htly3yHD4HFUs8nv9QDiZjbijim+sZddKabE0bM5S3AmFiScwf4mCyhpJx8DIJE7ST84hwY/sA5ty3c2N6Os+phwAk8u3iKhYSEhISEhISEhBwbEWKOd8ztrNt7NmPCysm+AZP/WXfMdUAXBbpQEDIuEXMKvjNNSLmCmBNb8+8LggRMzJs6m/xjAt5cCGES3jnZN1slvpr6VMY0++F58au+pMzyyVb8sx23y77896RFlNcVO7Nlmj64lYUD83y0hm1oO/pYKx+lwu/F/nJirrVXX3TeD/Ret3mG5BzfE8cvqcpjWcDAJNJHttfDf3153dp/6bdOfMvz9i951kM+3H3foY/dY/Xxn6q6l/9hVS06F0ea743j1w9/nMSdcSEhIXcXISF3tLvollRV91O7Tlzx5sUPu/e7Tzz3Nz7Q/ugbb+4/uOn7/Xs31609o7V8vXUS4zXiDz8Y0eKddNgqdLzVDynoOC7kluU9/nD8ljHdx2+A/2DRsVzRIOckHnJr8Fhneo0/1CtImrG+xhUr85gitrrNPtWX+yljsW5Zt9DTv/fNyjKRR11ZXuZpi99nsUx0TrwZ3F76WNTTugaPf4JSV5bpXXJCwhmUSHOd2jb8u50cCyPmDpGUM3KuwNDB7Qp5vxx0xD7ojZgb+otzPrfwXq1dOKceCfCfV0HMhYSEhISEhISEHENpEHObMGkew+R7DFtdTMjknpNwmXzPDX1kx8CFjk/aCeQTMecLHvGbfTcXBIBMwD1vE3FMwBMpZxN0IdOQTgsAQ7L3tEHKvF+uF7+0NXg7UoY6Al0AeFluy/YBFznMs57b2WKmJPV0MaeLgmQntpmYS4+xTo4K+rG47L9utO67ZtOdbWzb12DBec2Y3D03cPNEPfyVK+qTbt1Zt37vZT9c9Kv3/0y1asU5OKoPAvgSay4ymL4XcCIQi42QkJCfJ+E/Co72zwLe6dsG+KGak4AHLDx5eNPI77ziq/f+/OX18I3b69Z1Y4hro4g9/JLrFo6pM/5xnQGOoQTjkscnh8Qzi0E+rhsZp49M5rzGB4sZErMMXp/xzv05SjvCyDn1AVhMoW0iwQTZT9arrccl74fEpk6/VpbS3k7Sl8g+k++SmGNe4La6daR6aD/H0TLfhJNwTs7NTcxpvBT/bpcIOJQdhZjjXXP+gQn5ii/92B1zQt4KMTdwGc6hRwO9QMTKkJCQkJCQkJCQYyj8Kus2+yprQczJAkDIM11UJBIKkMdNOFG2vC4WbPHhxJzoDb7w8XzhTxYGptOPIiDNCbgsKGjDdphXSLlPzGVyzrT5sjK110m7bN0+tV/4MDvxA6S0lClcr0Sa5rW+ttlJzGl7sOVW/CnS+29IzMnW9AL1wb7JwpHk3F4Ai0ouLPuu3lz3XTVa906N/6i1f+v3Wldt+vv+97zuH1a84oyvLjjtnrd1H3/8lVXXojfiiK7XAzuvxF1xISEhP68yH1l3StfqlX+04um/8u11f/iq72z49MV/27rq0n9ae+UlP+Sdx7xTjq8P4Hs725OIS0LOYQwXIC4xTkms0nglccrGfyHhBBi7hQAqdBYXxJZ1GA+sPsd0JeYU2bfpLI6keCJ1NB6kmCVlqMc6bit6xgyLHdQ5Eed9YV7sPXYqpH+elxjM/YJ9ImC/qGe/CZZ7Pc9be6LTdqTM2pWYJn3yfmDOgP2kRBvz3Ge639J8wnT6cQeSckbM2f4Vv+ZT/bo9tjwm2A7wmACJPKWNlGWf8jEI+ccZ6k7BlxBzr/v8gnu334zz59eAfiCIuZCQkJCQkJCQkGMoJOYmzrxq3Q1OzNliwSCT7bmQSDPNpwWHE3MGTdtCJOV1Qp8XBbYgQF6JOYVPxssJ+VzEXFp8iF8rs0l7JsoytN9eVyFlkma9QteB7MvaZFoWKSXQrqXFj/fDYYsGXzCwrI3FggB1+ie3zLQnsXjko6vXjNatz15cD+G4rLtlvD7xT176L8uee/rVPRv6Lq+WL99UdfEDDgteWlULfwtbflFuiR7YeSWIuZCQkJ9XmY+YW1VV3U+qqp5zq6WLXrlg/dpXrz3n8e9fv2fzt9ffeEXdv2dL3XvdKMZYvi4A8WYvgLG2PbUFutEZkjTlXd8kofTObIMTPnPB4obHlxwj6Mfio6EzxiY7Q5nXuFbkC53ba5vYWtws+6KxKNdRFP2jfl5kW22/qJt0QCNm5zgq4Ec5sI+yXtNO1KUPPhASF5Vw03Kvw3aRFmQfiRwF5DiQmCvIOfF3aKIe8jvpEGfdxxCPNzD0wXO+sOBeg2/F+fM4oAUEMRcSEhISEhISEnIMhcTcjk5iDpNVgS4a0oTf0wV0cl4sOLioAZyEczgxJwsd6qQeJtqziDnkSbQJdDKuE3JFg5iTMviU8sKnlGHibXBdiVzu/nN7SqI17eeEEXL6m7JeF0mG5K8JuWNOXk6tiwXmB/iSar6Yeho2WDT2Xbu5bk+P1Ruv31X3ve919ZLffFhdrTn+QNXVcyGO3KnAGmABDyOECwk+2hWkW0hIyC+r8H10HAszsbJw4XOWPvzeV4+8/+z65M9dUQ/ffFndt3esbl87Vg/uxfg8tU2IuQG5gw7xxOLXUAKJIx+7OZ4bjAwSpDLUtbE/x6QcH1OcdKR/hMEOSHHT6npsSfHWy2nrOtmyPY03Do9H4kvqaNptE6i3smxr9iVSHdq5Xuv475W2kfZXOEheSDmi0MmW5BsJOe1vSczpXXBOzGk5ocQc0uLf/Jitf1xJ0ocQT/2ORiPk+DEI+SCEv2cOfoYQZ3m35PCHXv/FBScPvA1nyxMAPh7tcTUkJCQkJCQkJCTkGIgQc8+9ekSIOfv4gy0aZBHBvEy8DbKYsDQnxthK3uzLuw18gt+sjzoyydd6JOQ0TZ3BSDld5Hi52+R8ulOurGvwibxM4n3CT2BC3yTK1E7qSZvZvrzzLWO2ThYJ9jEGB99fo7aFvflPX2eVxcOOeuDwRN06uG2mdWBiZmD/dthuqftJkt4M3595438tPfNXvl6duPpGrDD/FOvOs3HUHgLM9yVVknNHe2F6SEhIyC+acDw82pg3xHGza9WS9y0+dd17TnjF49438CevmWp/+tJvtfZuqVt7xvWdc/K+MaQlxmncG5gen9EvXjMWcay3mJNiVAesnLaelrzFRAV8F3HSY6e+g9XS1Fu9ZhxlH4p+OKR9prU8xRzmJeaqLwX0iVTztPv1+uo3Pd6a2mKZwX1YXZanmMuYyHQD1AEst3xJzEm+BMpSXalj9ZEv6/i+Z15iu5FzDv2nlz3KKl9oLYg5ErFOzN1n6HdwrpwBDABBzIWEhISEhISEhBxDWbbojLU7nnu13jGnxJz/N18WCLJAITCB5cRcdJrWCTgn67nO3MRc4UPyOomXib9t06TfJtnlQqNRbulMypVlGTpJ9zQn8wT7W+rUTmypT3mzL0m1o0LtnJyTRco8xJz49wXDYSwYDmORAGDhMDNwcOudA0e21b0Hx+qhL+2s1/zvl9Xd9xv5dtW96K+wTuBHHR4IrAZWAD+OeIs750JCQn5ZZL7xbikwAvCDONye0jWy5pzjz/2Nz667fuIHQ7fulo/ptPbwHXSIYYxVB0jQjc4gpikxJ+M6wfilMUzixixYWQLyrNMJuytPgfaIzvfbGXIcdajfRvzztrFleYo5Qp5B5z4SqGfdbJf6nMrmJubUP/2ZHeA6jZ+0QYzzdILp6MPyc94xR72gqJPK1K4s09+OeGpb6vPdjErK+Z3pQsjZo6w8psNyhyS2f3nul4yYeyIwCAQxFxISEhISEhIScgxl1h1z9vGHjq+zCunGiTcXDdwCehcBdZycm50Tc9Dr5N3rFIuNoky2RGNBg7RMvFnGCb9O+n1RoI/VZr08/mrQemonOkljgu7wiXuqY7pkw/osp97qW53k29JaBpgvqbuPPrS+knPaRmrH6/MdOIexQDi8ox46vBPYVQ9ev71uXY99+PktP1r+2if+fbV29aGqWvwBrBFehiM1rAdsTgkSLiQkJKQp/OfF0cbGe1ZLFr1xycM2fHzN2NMPtD/5+i/2X3PB/1n78XP/qe/qi38wdGBMyR8Z0zGGM1aRvPKtjPc61mu6hMUniW/ISx0DiT+7M66RFhgxWOoQN9VPjrtlW4x/kpa4QjuF2Emb7Af74HXMxn5L7quWSbnUUbs22iVS/6WO29JHthWgHwLGPNt/KdYy7tk+475NRJwQaCUxp2RdJt9U57FT3j3HOqmc7WiM9ZgsxJzVlY89MC3t0A559gEYnob9lBJzi+47+Hs4L54EDAFBzIWEhISEhISEhBxDEWLurKtHbthVt/cWxJwtFJyYKwk5YgiLAF24IM/JuqEk5ghZJFhe9JLmRF3zaULvugIywS4n9gllnQI+KU8LFNNhAi6Qcp2sy4Td6skCI5WZL7N3n02YD6vrftQX9Qoh5rhw8XKrI//d53/yj0zU7YPbZwYObZ8ZPLKj7r9xoh780s56yW8/6svVsuPe01Uten1VLXwWjtK9gPkWCkHMhYSEhDRlPmJuOXAa8PRqYfezuodWv/y4Mx96We8f/K9PtvZe8vctxL7W9KjEEpJPranNMzle5THeY0aGxTqijG9MM/bxTrlpxFLZKjIR53FX461sG4RcEzm2WBtFWbIhYOexKsUs9mUusF4n2cb0HBCyz2xTe9gHAtk/KOc+YXsC318k5hy0ga4g0pyUKx9lldhqaUEqYx3bwr/eMQd7lEucLeD+nagjhJhDnB75y/P+etF9h38f58OTgSDmQkJCQkJCQkJCjrE4MXfjruYdc4ZEzNnEm9CJM++Yw6QWkDvnZKKOiTkmuvm/7zah9ro+mXcd7TmxNzv9+AN0hjQJ50TbbMQO9YhSl2ylXi5TnU7ak43AJuppAeHlmLBzgi+TfLNlefJn9aRu1kuZtOFlivT74Mft2ihv8465I9vrAd4pd8P2unVkSz3y5d31spc+7jvVsmXvr6ruZ+Po8CXUi+U4hYSEhIT8TwkJvNMWrFvzut5zn/rJDddd/I11X9xVt6bH6vbkKMZxQuOVfqCIsaAJxsQGMcc7zeRuM4J1EUeFiFMkYo7lcqdck5jrjLsOJ8bK2OkEWrIRqD7H3AytqzZKrlHPup5Xf+53LritEnpa34m5FG9Fr9B9pvExE3NF3u5oc1JOiTn1478z7W8pp73rkE4xG1veGcd6TFMvdbWtdAcdMDSN/L5t9fBHzvvrRfcb/gOcA08BeHd6EHMhISEhISEhISHHUEjM7TJijnfMFaScEHPzLBAa4OScE3ZZaJTAhNgm5jI5Lyf9BRLZRsDf0cEJum8VabJOWF4m5clGJ+du03yptE7o81f3OHFX++SbW/crei3rhNcv32NTwu3a9MNFBb8QxzvljkzU62+//L+Oe/3Tv1Mdt+KTXV2LXoMjs14P0CyJjzqEhISE/L/JfOPnIuD0atniFy08qXXhCS//9feu/+h5nxu6deedfYhL/XvHdQzH2M5/rmhMUegHIhBbEJs0HhbkGuoqkJa75ZyI6wDKFeqDcc7jq5B+RV5irftl7LG01Ee51i1il/Qtp52YS2nJq05iGeM2YfZiZ34dSZ/stK7C9ovtH423xf4qoHn7EmvSo78ddTqRfdBO+6DEnG7zHKBZT4k5xbC8emKiHvno+X+96H5D/xvH/6kA30MYxFxISEhISEhISMgxFCfmbup8lPXHEHO+KJCFQU7/eGKus1yRSDkC/n4y6MS7MSn3bZqUE76IUr2+cNrysqBS5Ik7ffjWYOXqi/nZkN8oPpSYk//Ks57UQR6LgDbBPG0O8gusE/WGz+2uV1/2wv9TrTz+M1W15Lyq6nk4jgw/8BASEhIScmykB+C4yw/sLKkWVA9adP/htwy+43/9072/9OZ6/W1vqfsnt9Z9/JIr4lp7H8d0wN5ZJjFCHo0kivjImJjiopJw7QJNYs7sLMZ5vCXhl0g/Qy7P7SUfrjcoaZjzSsZ53ZyXepZ321TH2nSoTutJ+6I3HSCxEPskxVrZKlTnW/tiatIpKff/mZgDPA6X/fC74xxOyjE9LL9zWz3y0fO+vOh+I3+I4/40IIi5kJCQkJCQkJCQYyydd8yV75jDYiETc9xyAp51upjAZLyhQz4tNFjmk3idvDPfhk0btp1we5lYu62X+WLAJ+GwkYm4PLZK6CRcJt5cIBEyEVed5q0u8roIQJoLBynLd7ipL6sjZbTTemmyX4DlvtU2OfnP4H/kB2URh7w/6ss87Eeun6hHrhmtu04avqWqFo/iiPDLgfyKYCl8zCreIxcSEhLys5f57qJ7fNeqVe9Z+uj739w7+ozDwx954y2tqzd9a+2Vm/+jPclxfDvGd/6DBeO8fUhAtoh/Q4hzQ1PAJOKWfP0Tccpj4xTiGsCtADon1/xOtRxTFEJWUWfxVOOu11GUcdftlNRDXddJOeMQtmYj8c3TEp+Iolzu2KON2aEvWmb9FH+aljwhsY9bS8s/xPQ3OKnGL6WSlGu8Y47lDrNTG5JphOksPtNO28j9TX2QMvOLNgg5VuKDd8zBH/LDHz33y4vuP/JHON5PB9YBQcyFhISEhISEhIQcQzFibt2Nu+qByU2YpHJiD9g2E3GY8CYgLwsMTvCJbDcnMdeJBjGnaWnTymVi7bZWLl+GkwUBoRNxIeXmIOaUTHPkiXleJLgO29KWaZ/Muy+zV9954ZGhdf3uuETIYREhEMJPSbk2Fy5YtLX2bJlpXzd+J3/Xxtsuq5e+6HF11bPk+qp78QtxRBbqgQkJCQkJOQYy3z892lXV8zh552dPzxP5kYgTXv34922YGv/WutvfVremJ+r+SR3bh6Yw7k8hFk2Nycci2pOjM/IeOSHrEB+FnEMcEzJuTMH3ygly/CvjGf954zHI7yKTO8RYBttmXKZ/89GpB5SYQ10n0ZBnG0K0FfD64iPptU1PN/poZVon573fTWiMzYSbbjtjsJBthV1J3CUfApSbrb9blunUB08XdRmzGaNJzI0gLg8jPssdc/cffgeO9zMAvkYiiLmQkJCQkJCQkJBjKE7M3bSzHuAdcwdIgCkJxq2QbUCa4HOhwUUG9ULCmY42lneiLk3wZdHRLHcyTsk5y0ub8MWJdLl4ILzMJ/iccCdiTiffOhnXibdOwnVCnib3LLf6QqQ1oD7TRJ5b19Ne9Np+2ZbU7cwTiZjbjjy2KBNijou0vUjvxWLs0I6Zvve/rq5Wrfz3asHSd+FoPEQPSkhISEjIXShHI+uWVEsWnrnoIRs/3P+23/7muoNbvjtw087/7t0zdmffVZvuaO3ZNKMfcRjTOElwzJ/CdpJplBkxx7TkGR8B/aAR41+OMx6bNIbpNscp2KOOxmVry+Olw3QJXlfK1Ucm3jTdqM+86LQ/Yo+0+9K45/1pohE7LZ2IN/8t5e9jHZQr4WbpZMf6Zi9xlja0tXpSntuR9gXsv6WlLssyMbdu34QScx8//28W3W/dO3F8+RX0IOZCQkJCQkJCQkKOsZCY2213zO0dxeSVBBkm85yACzHHSbkuHBIpJ1vokU6gDSfrCUWZlbOuEnNb7P06vBuueUecApNnIeU4iXZdCZ1o+4Q/oZh054m8bzkpV6hdJ1Sf/BYQH7LF7xJ/RIefQi9EnBFzQtAZMdeiHe+qmNxa9183Wq+b3vmDBY+8301V1f07wJNwNE7Qg5IkHl8NCQkJObbCcbdbk3PKRhQ/v1q28OKee/VtPv7Fj3xH7x+/ZE//tW/6P31TmzG2b8Y4j1iKGEdCbmAS8W1yrG7t2TxDDCBNDBIk7GhTxEaJNRJTLG4RcseY6hMs1gqBJvUBibtF3pDiMuKnplnffJg/vRPP89lW47rGOS9vxmH0TbaeN8Cfp9UnoWRajqOKkoxzGwHLCjvxZWklK9Umladt7lPqA8skrXOEdMcctx89/28WPWDdH+PY8mvoG4Ag5kJCQkJCQkJCQo6hLFt0xgklMScEGybyNimXx1/SZB9pQnQdk3+pRz0mv6kOwYk9FwZlXok55tOjrHMSc57vBCfXaINbm4xzoi2Tb066ueXE3CbpaXGDMm51Ym71xBdh5Smfkf0gn8g5bUcgerdhGyTlUE4wLdhWt2g3rXfLDR7CguD3X//P1ZJFb8NR4HvlQkJCQkLu/rIE4D9RVgF8H+hpXcMnnrvqnDOuGTkw9q+DX9hV90+P1n3XbapbQsIpMdd/3eaZ/us2zcjd0ga9i85iI+OdI8UVxBOSciUxl0g1QmOx/tMLvgSoL/7ML3Vux7gp6aI+fAoxJrFN84rcH9pKLPS4a3FY04ybGg8lL/1We9F5v/FbnHDTWFkA5alMoPoGMQe9tGV5LSuJOesT0j4fKKE+mKYP1ttWEHPnlcTcRiBeKRESEhISEhISEnIMhcScPMpqxBwJMiHSOBkfr4eKtP8nXif/xcSfW4FO4Jv2nNjTB+tYHnXkLjnfJvgddBmdk+sGMecTcuQ17YsDndgLAdfYOvLkXupSx35LevaE3qHteLuFHfJCzPkCioRcAkk6JeraqMPHWfsmN9Ubb9xVH/e0X/0OjsC5eiCS8EXk872MPCQkJCTk2ArvoptvXD6lWrLojQvvN/yuE155xseG/uzsG1qfPPfLaz9y7rd7P3PRj9p7Lq1bezfX/XuI0bqNWCsEHYm7TmLOY5zEFGwJxhvGGY9DDtgTUh+xVQk621pZsmEMFmJN836Xmt6pRkCH+CZEl+m0T96u1vW0bKVPjLcaD1O/0F+Jw9xa2ok01Re2Yu8o7JhmPdqLnW1ZhnjqH3PIxFzeP7LvBOyX+pI8y5Dn/mQbw9P0P1Gv+/h5X1l4/xG+TuI5QBBzISEhISEhISEhx1jsHXMjN+6s25P8Kisn9JzAG5mWwMk8J/0Kn/hnUs7zOukvbb1cFgdC1Oldcvo4q5JyLG+Qcpj4yzvZZCLNCbYuBBTwnwgynWzrhB1lnIDLBF4n5KmOlFnabMu6vvj4cfA7CnRyj60DC4REzMkdc0rG6Xa7bNvYttD31sGxeuTKzfU9Bvq+W1U9l+AoLNaDIcIFYDzCGhISEnL3kvnG5eOB+wEPrXp6HtndPuFFK596+q6hP3z5nqED4//au3esXnvVJXX/tZvq/r2jde91m2ZaJOdIzNk/rDRuanzR+JRjnMcZj1cJsNf6Rbw1JH8GicOM64yfBn6xNftzXfYvJGGKjVoucdC2AiG/NJ18CRFmYDzGVv1avmFbAHm9C84A20zM0UZ9CTGX7HL91I+ks/akHvJSrnMD6hvE3Ckj766q7t/EMQxiLiQkJCQkJCQk5BiLE3M37FBirvPjD07QccvJfnofDvWctBvcThYYmvYFRybmsi7dIec29Gl2StyRnPOJNnwKIQY7EnQCTrI1rZPtPBHXhYJNviXv5VpP0zZZZ5r1ZfGhbfikXn2ove6PsgzwtBBw8HVwAnlFulPu0Pa6fWhC0Dq4ve7jfrh5e7368hfX1eLF/7era/E4jsKJejBCQkJCQn5OZL5/opDYeVjPut5LV5/92E8PffgNXxk5suuOvuvG67VXXkpirm7tHatbk+N1a2qr/BOqbWRRIpEKSNwiJF5ZTGM8Q0xirCxjqMRZ2SJeEYxbtJEYZmXUSTzjlj4Bsx1CX9KjrqzD2CiwOilt9STeet8UQqhBr6SakmdCzPlvEzuWUUdkf5mUm8j1vUx8QVcSc1Kmdo1+iF8tU1vVyz/vrG4m5s7/ysJT1r8bx+y5wEl2/EJCQkJCQkJCQkKOkSxbdMZaPsp6w856YJIff/AJvE3CZSKuk30h5kpwsu9lhibJVhBvLKfeyvUuOeodqhNg8SB3zFk/ygm/LFI4uRZouRNuOhHnpNvsuLV6Ym86n5gLvJx9sLbExus0oG3ldjIaPpknWSeE3Pa6RVIO6Ee6Dz6Gbt5dL33Rr9dV97Lv9AQxFxISEvLzKCTl5nu8dTXwiGrhwud0tY5/zYrfevifDn/onK+tu/3yundqrO67ZnPdniSxxvePIl4IEbaVMXNGCDKkeUeb3tVGMO5kODHnMVSJOIvFFp/5jtjkg+Xl1vX0S5g+E3NmK0Scxj9ti3m2TR37YVvxY/68n5JmTFTCzdMKpIVQ87oKId8MHk8zCWdb/iPMyrQet8ybH4vTug8NZbv0AQxj33Ord8yt/xMcrzOBewJBzIWEhISEhISEhBxDcWKOH38gMcfJNyfXNsHVCb0RbZzsC6gDjHAT0s2gJJzqhWQzIk5suEVePvwgNk04Madp6GQhgEk7FwmcbBOJOORkXtM62SZsYg44CSdpmbRn3dGIuZQ+CqR8DmJulk/mEzG3o24d3l73C3bUfYdQDn3Pw+5XV13H/Z+ursWX4ih0fo01JCQkJOTnW3oAfhxCpafncQtPbn2w/ZYX1ve8aUe9/gtvrfv3T9S9U1v0i91IM870T4/O6B3oFmcl/ll8kvijGEh3t2vM1Bict0LKCZSYUh8sty10qgfYhunFPrWpdomYI5iGrcdExj6JgZLPdTwWO2nm5FrSMw19I74CSsppPa+j9dwPbBrEXAn14f0Rck+2AOt5nYMT4sOJufUfP/+ri09d956q6j4LRyqIuZCQkJCQkJCQkGMs8xJzmKQzLWSZQv8Tb1sj3ERv0DviOHknKUeijXnqy7IOSJuwxVbfLWdbTrBlEaB9kQm3gPVyOusLSP8Bbn3yLlvVu43b6+KHyAsFJQdZ30A96xSgT7G3crdVYm6b3jEnpNzOug/ov35H3frsRXXXunbdtWDl7VXXojfgKPDOipCQkJCQn2+Z7y46knQvqpYv++CiU4Y/u/Ilj/lo/x+/fF/ryou+0j85+h/9/HrrvrG6fWAMMYVxdUxiqhJrjMW6lVhIck7SiLsCpBGvlIhDHcTZRMyxHn2In5xWEq8DbCfZal7imtgD0o7Vpd4gcVLg8Vxjp8bFHHdTLBV4XLa0QYk4tC/wvJXB3svcT267AMuKfIOY4ysnSmIOPo2Ye29VdT8Px+heQBBzISEhISEhISEhx1CMmFvvxBz/E28Ta52UcyKeiTe96023QsQZcjnqJiJOy+QuOEmj7GjoJOYIpDMxpxPwTvjEe069TNzzhH+WrU3qBXO0ocScLUBsou+EXAlZHHAriwyFEnPwwcdYDwNHSMztqFu37apXvfNV/12tWvUf3d0r91TdC16Go7BKD0ZISEhIyM+5HO29c7yDbj3wSOChwIO7Wytfuur5D3vP4KfP/ebwV3YjVozLV7vb05sRT5WYY0xtT22eISGWybISGltpK6Sc1OG2Scxp/ZyeRcoR8JeJOYI6t2WMs3wHctzUmKnI+oa9xcwcmzvzRsbB1gm5Zhn12Y/GXstbG0lX6BMxx/fAJmJOdes+ecHfLj5tw/uMmLs3EMRcSEhISEhISEjIMZT0jrld9cDeJjFHJMKN/5W3Cf/ANBYMgJNyviggeEecfwBCCDcuKpDWx1dNV4CPuurjqyTwjIwD5IusBPNiy7yVW9oXJqmMdpyUy2Scd7JxEs480pyQe5ltNU09sM+2Ui+XS/vy6JDVE38GpqlnHds6Bg6iDGgd2obFFuDE3Bd310t3v/A/qqXLvrVgwYo/x/yfL5teoQcjJCQkJOQXTOa7i+6EauHCsxbcd/C9K173+EODHz3vqwM3Tfxb33UX/Wfv1W+6Qz7I5LETsYjvX5W4XKB8lFXulCMpxy10hJBSiJFO7JHwchJO61p99yn1GO9oW9hTz7xssw/69TjtdkriZTRjKnSMk4i3fgfcIGJl4/FUKy99iJ3YKyQmCxij4dfmLtJeR17re72jEHOnrn9/VXW/AMckiLmQkJCQkJCQkJBjLMsWnbFmx1lXrS+IOZ2wc6LthJzqlJgbE2JOIBP6EqiDBYFP0CUP+Lvj/N1yc5Z5fWnbJvAy4S/TRsw5Cj/qy/SwV0KuACfk1DMtdprOCwBDZz2HLACUgCt1qufWyL+U31q3Scxh8q/k3M6679D2euBLl9XLxs/612rJ8r9e0L38nVgIPB1H4Tg9GCEhISEhv2BCUm6+j0TcE3HgadXChc/jRyKOf/bD3zH0nlcdWf/57f81cAix8brNiJ2jiDlKzEkMRJxGPJ2ReIZ8JtR02wRikiERbVZnbmLO7ZnuIOLmQBmDRSf+HYyrmmbMlLTESYPkHU668cvmRTnA9uWON4nRWZ8gMZe/h7D0LGKOsPoA29F3zE3U6z554d8ufuDGP8Vx+G0cj5OBIOZCQkJCQkJCQkKOoTgxx0dZhZjDJJ2TdZu0OzEnaUzg5yfmCNjaBN3T+q45lJG0c5it3iXHxYbVLSbSeiecgf0RW7VJsDrihz5NL/U7kP3SxvIyyc9QPbeaTrCFg5NurvdJvpJy9KF+xSYRc7xzbnvdh/TQF3bXy7c893vV0qW39PQsezsWAk/CUViuByMkJCQk5JdIeDcdH3HNRFBPz6MWntS6/MTXP/mLJ107+oONX3pr3Y/Y0T+FmEtCzuMRtnLHmWwBxLUSTpwJIUeize2ATMxZDKWdx06JuaYzSB3PI577B6DUTmOh2ms7jKF6N5zGTI+LHi8TWI+gTdIXxJzplVSjD62jPi3tecDjv8RfI+fcRok92rnviXp4P3Bwol7/qQu+tvi0jX+GePxCHIH7AEHMhYSEhISEhISEHEP5ccRcItwU+igrMWZkWxPyFdY0YTewTGyZLvIo07vckHZi7ihQ4o1pJfky3Mfc0DKd/PsE3fN5QcCtkmv62Kvq0mRf9EUdQBcFnNybzuxSG8i3gRZ0LSHoJuo+LBKGvrCrXj723O9Uixcf6OpaOoFV2K/xKOjBCAkJCQn5JRDeQXe0d9HxIxFnIEZsWTDc+7vHPeeh7x/8k9ccHrpm7B969yLuMr7s05jEVzC0p7fOyN1zvPuLJBzinn8QQu+QM2KtSA8JYWVAjBSYTSbmNP6nctEpnJhLkLpHAeMiY6hty7vW0msmHCmuepo+SPIdhZiDXSL/AMZ8+Q38fek30lbzuh9oyzqdxNyGD1RV94uw74OYCwkJCQkJCQkJOcbS+SirEGROlDWJOb1jDmkh1kjM8c455jFZN8JtTmLOfZhtemcOypxsOyoxhwm1E2xNQi6Tcp130ck7ZyzdIOZksq8T86xTNBYHpsuYg5gr7EQn28I/+wC0oO8XbKt72bcv7KyP2/Sb364WLd7b1bV4E44AXwTOhVhISEhISAjvoOsF+HXQYeCUBYMnvrF3829ed6/br/hR/xTiGuLo8PQEYum2uj21daY9NT4zOK2xh0SZPn5aEHEWlzztZZlcs7oE06LTeJ9t5oHXnRNKvuk/vMo420HKOSSOlmnrM+rro6z0o35pUxJziYiTPNOex1bytNE6fIR1JD3KKsTcn1dV94uxv+8LBDEXEhISEhISEhJyDEWIuTOvWn/D7npgzygm2Zk00/+WW54TdMEWK++AkG7jWCToooGPr7rO/cgEXog58yHlpT9NN8i3/QqWORGXiEDPY7KdCTOdgGue21yuk3WdzM+6K84XAkxTxzqN8jmQytGGbFEfeqnLdqFvId2PdD/89+L3DH5+Z738kmf/n2rhwk90VYtehyNwOrBEjkVISEhIyC+rzHcX3YZq1ZLzVr/s165b/5k3/U3fpy/6XmtqfGZ433bEQsQzxEGSdJlsG0/kG++O07vNjKRDOhN00AsRR5i96NSPx3CNzQqJcbBPPop0I/7JFnmPt06qFXaJgCshdbTc3ynnd8s5MeckneTNXttg//T3yu+WvOml3PzA5zAgxBzy6z594dcXn7bhL6qq+yXY1/cDgpgLCQkJCQkJCQk5hkJibuLMq9Y5MZfulOOENpNqiWQriTSftBeYi5hzOCk3LzEHXScpp8QcdUbEdUAe7fkJkCfwulDQu9pMh8m5/PeetqYrkf2gTkqjjJiXmCPG634sXvqmxuqh23fUKy585jeqBQv+vKoW+GMzi+VYhISEhIT8sgpJuaMRcwuA+yJuvGzJo+/7tuF3vnxv/6cv/Cf9RxXvnOOX0hHDJjfP6F1uJOYcTlAp6aakFYkthZNwDtWpXmOsxWaD2uX6CWX8o42lO0k38W/ovGNOYrPUoZ2SZnp3G+t6vtlW8mW/Ueoi9vo/6ZIN9Ar1l4g5bNd96sKvL37ghg8WxNwi7vSQkJCQkJCQkJCQYyOJmNtVt/dsxoRbSTCZ4HJyL5NwI86EVOOk3ybrQqyRcMtpIeYATSuyHevTn/oUvyTmWI60PAZLUC8g6Qa9EHNKyrWxIPAFgmJrQdhxcq7wibiQb/gtrlc4MWcTddPrwqGE++AWfWZ6li/Ti5+8WKCe/W6hXy38lhZ+V9/kaD186476uPOe+pWqewG/yPpMHIF1QCwCQkJCQkJ+nKwGHrP41PVj6/7o5ZN9n7zgu60pxBzG2b2IqVMk6BC3p8cQuxm/lZjTeG7EHNKZmFN9CSHOGsQc4llKa/xXYs6IMoL+JH5a/KONpUUPaHu6lbgLncbdIi8+1I/6zaB+rrb423Is1nYTMUed2akfbWeQH2ICRvg7kF/3qQu+sfiBGz+EmPxS7N/7AxGTQ0JCQkJCQkJCjqHYo6wjNxoxJySYkmYJRpwpuUYdJruEkWolMadEm+qJ5iOtVsYt/YlPtaWd3GmHMt4txzIl24x0wwSb75JzXUnC+aRc3zWneSXZ8sRd4AsAm/iLreRNxzpA9unQ9kt4nxKSH4XYFMRcP35bL4m5W3bUx73hKX9d9fS8rap6Hocj0AJ4N0RISEhISEgpcz3eyneSPmbpgzfuWPfe1x1ae9UlP2zxn2F7galRIeeEmENMJTEnMRzxiBBSzNKSNwKuRCcxJ2nReb7QFRDSy2Jr0ok+x1Yh1ySNcom1zXircZhlCiXTvI7WT+Vij3rsj6VzXerQX/mdqmv4OagYxu+gjsTcotM3friqul+GfRvEXEhISEhISEhIyDGWoxBzxFzEnBNsMkk3Uq1ZnvUNiC8rk62X5bSScgTS6IeTcE6CCfFmOgEn5QlWzrRMzDmJx1Zgk3RM1v3ONl0E0M5hE3ag6Vd9z4L0qyw3fwT8ie6AE3NjdT9+X+/k5nr4lol6+Tm/8cWqq+dy7P1HAWuBIOZCQkJCQn5S4Z1zz1xxxun/e91fnfulNZ8671/lTvVJJeZ4B90Qgbgj8RrxaC7oHfGIWbIlKTc3WSdxzdKMfXMRc4QQX408tkVsFWIukWuItQTsNH5quZJoLMfW7UXfAWtD+mNp8StbxmSFtmV9K8D8cHqUNRFzL8d+fQAQxFxISEhISEhISMgxFCPm0qOsDWIOE3z58iqQiDcn5kjIYeJrxFom8VBGu8JekPKc2JN0c2zRu+U8DfALrfoIK20NmGzLHXM22VY7J8Usjcl3gpQBsOMCxMs5QXcbSbudoJOU03YUc+nZJ82Lnj6xlcWC6EjKban7C2Ju6Jbt9bLXnPG5qqtnAnv/wcAqgF/hCwkJCQkJmUvmev9cf9Xd/fTVz3vMO9sfOffLrT1bhIwbnETs3Lt5xok5Id8Qi9IjrYTpVO9xy4g5L28gxzaB5TU+N8uyPxJg2AoRxm2HnmmUeSwWOytzQk9JOS8HrI7aWr6E2eg/4QjTAXx0deggY7PWT8Tc/gm+Y+4bix648S+rqvsV2K+nAEHMhYSEhISEhISEHEMpiLmBn+BRVoJ5LcOE3cg5z5d3wBFzfQQiEXGyhT3TAk0nsC9YJIgek20n5RIR1sgDmHwLpEyhiw+drPsEXfSS1rLSvgn1q/1qlmmfCsxqG/Xgnx9/6Mfv6MfvWWvE3NKzn3B71dWzBXv/NGAFEMRcSEhISMjR5GgfhlhdHbf0ZYNve/H1Q/xCKx9lnRyvW9dtvrNdfAgik3C6ldgtacbGpk7vomPMM52AdnOjk5hrQOKsEWkp7XrNZ/LMyDZL6913pQ4oSLVUNgeclEu+D5KYU8JvLmJu/Wfe9M1FDzrpr6qq65XYp0HMhYSEhISEhISEHGNJxNzOemBvx1dZsfV31AhRZqSbknBmw7zdBSd2RtQJATdl5NuU3XWHvJJzY0LEaZp1FUqAEdATWCQIwcUyTLbl7rRiwp3KbVsSZ5KnL2zThN0m6VJOUo6+zFb8Y4Ghd8a5P8VcX30t7dKiARP9VA7fSsxtNWJuS72WiyYSc68+47aqq2sMe5+PzBwHBDEXEhISEvL/Ig9Y+vD7vPvkG3fVralNGmct5urXWnnnO+M5iTnEaYAfRxAwfjlok9IaMz02l8QcCa10JxuxD3EvlWc/HguFcLP4WMZKEmskyhRWJqRZ00b8FvmyjvaF9QCLydI2yhMxZ++Tm5OYm2b9BjH3KuzPIOZCQkJCQkJCQkKOsSxbdMbaHWdetT4Rc3liPedjLdO6la+9MS3EnJJyncScQ8g5knFYLGgagG1+fJUkHIA0oTrVe9oJMiXnbKIOnZNjvghIYB1Bh17QWa9sM5dxst+sxwl9pw62RvBx0eB6eew2EXN8nHVr3TuN/XtrIuY2Y+/fj0cA6OahCAkJCQkJ+THCD0KUsqpauPD5a8950kfaHz33G2uvveQH8jV0iceMz2OITWMWz0nMMbYp5OulRV50EssyJIYm4k2R3hPXYZPjrsZEjYtFHYmRCiXdStAOWyHgcj0BdaKfo8ygbfI3IM04LfMEqwc4oTf7jrltQswtOf2kjxgxdyoQxFxISEhISEhISMgxlETM7RJirpyop4k8Jvj5DjmF3Eknk/6ynAsBTI5/DDEnOtg6/J1yPqEvSbJMlhkp55BJN/oipJiWJ5QLhFLfgNeboy2px31gE/4ELCjYbqmTib62w0WDL1AaxBy2/VgE9HIfJWKuZxP2fhBzISEhISE/jSwGBqq1K5/dv+XM953wqfP/SWOlxlaN5+MSz7n1+C4x3uNo0pFAcxINZYhlcicaUdgogcYt9VbusbNMi63CiTmtY/6TroD5bUD66TYdZYS3aXkl5ugrQ4g5lkkbRyXmXo19GcRcSEhISEhISEjIMZbyHXOdxBwmunJnHAkl2SI/zTvlNJ2IOcAfP/U76oSkMzgZ1yDrOIkG0sccZBHhaSXI2vCVCDNMqEvSTBYD2GZijjZWLjY6UVdfVtfslEzT+vpI6hwQPzbp98k9JvDlZF8n/7mOlMsCQYm5lhByBTEH3eCtO+qlr3kiiTneMXd/YDkQxFxISEhIyE8jJ658+oMv6v/M+d/qOziGmISYehBgTD/ghFznliBhhdjOrZBXSlY1HhH1uAqIDuUNAo2xkmUCxkOLtcmevjWGpvqml5iZdIXetgK2JfULHduUtLahbarOY3OyBfy3yZdg6QtoEHMP2vjRIOZCQkJCQkJCQkLuGinumBskMcdJbppgGxlHws3T0yTmuNXyTmKOd8slYs713HYgE246gXdiTog0TKKlrCTmDGmBwMk360pat7PKMXFXso5p1XESr4+ewp6wRcEsFH50EaKLBp3QQ8fJPe24AEj1TAeIb7TTQjnRf3Ci7uUC6fad9dKzn8iPP/Adc3yXTbxjLiQkJCTkp5Wli+4zfN66yUu/1Xt4cz1wEPHrEHAQMZrEHONPg5BrQu+UOwoxh63HWSW6UM44aOUZKPfYyXrUmZ9MzJXQsmxb6GyraNrPDbQpvwV1GZeL2CzlErszMZfvmJsIYi4kJCQkJCQkJOQuFifmbtxdD+4hMYeJOwk3mWA7CVfokM5kHfOqIwGnZBzrad0GKcdylPljrg3CTSbUsMFWiTQDbT1NO9Hlu+1cJ3fkWXupPibicqec5DE5TwsElpWwibtN2hPh1gEpw8JA/gsvE37aeR3kvdzs5eMPWBgJMXeImKj7oBsSYu5Jt1c9C/lV1gcCK4Eg5kJCQkJCfhpZUK1Ycc66vzzvawM3IbaRlCuIueadcwbESkci5khWMV5KzNQyiaESL1mPeo17QuCZvdYxe8ZB1hEfGhvljjWvg3ipxJ6WCZkm9ppObQD5sVevizTrM+9+xJf2UWK86CxtdVN8PzCBfYJ6QCbmLvrmkgfdM4i5kJCQkJCQkJCQu0iOSswpmu+Sw0Q3EW85LUQYyTHmHahXknJCzImd2jop5wSbvGeuzHcglRV1M6wN+sBEXGETfdlaXZmko282gVeojU/aZeJuugxdHAhkYm862+pkv1m3QcwdVGIu3TH32id9rupZuA17/0HAKiCIuZCQkJCQn0a6qwULXrtm629/deMXd9UDh8bqoUOIbSTm+EirxD6kfVvEZQVsDSXJJnG7gBJ2GekOu0IvcRD5kpgTGIGm9ZRY07IOO7HN8ZRINtALIYe0EnPUoUzisW2FhATmiPFKzE3UQ0AQcyEhISEhISEhIXcPEWLurKPfMWeQ/DTvnqPewYm9TdaNmBuYAhIxp2VzE3MdRBwn0GXekIg216G+11W9+qJ/Pg6b7TEZJzgp97rQy+IkTdiJzkm71WnAFwgsNxtuRe/1bGu2SszxMdZtQszJo6xs+/Zd9dLXkZjrmcDefzBwPBDEXEhISEjITyOLqq4Fr178zEd/7l5f2V23jozWgwcRlw4hHtm75vzOdBJzZVxWIA4aNIYTFrcLJDvEOSXYnCTLOr/7TWMw+uDorIv4qOQc826H9FxxWPyj3Oo0iTmitEU/C2LOfWqsbhJzg8iv+8xF31wUH38ICQkJCQkJCQm5y4TE3M6z9B1zRswJEYcJrZNyJTFXQibq3E6j3NAg5opyWQRIGoBeCDUSdJInueZb09Pe8no3nUPtxLdMup2IK/0QmJy7jevMXif4TSjhpvDJv07iCZY30aYedkx7/cH9mPBTh3ybNkbM9RsxtxZtGzH3+aqnZzv2/kOAIOZCQkJCQn5aWYi/53efvP6qDTeP/6Dv0KYZEnPtQ/wnEWKjxD+LxRZDMynXRIqZsNN/uhFejngnhBvSRQxVgi1DCDmzVZ3FyqKO15NHXC2WKoFGoE5Kaz2p63qrr/A23YZ90/5JjBawHmP2XMTcm7656PR4x1xISEhISEhISMhdJSUxt3e0HsZkVt4hx62RdErMNcm5RLIJrJxblsvk3SfzWBDY1if3nPAnAk7yJNQKUG/pWWWiUx9OtJXEXGeZAHmW68JE0ZzUc/LOCbtO3JPedE7GyeKAeZS1fRFAvdlygt8k5vRR1pKYG7p9d73sdU8uibnVQBBzISEhISE/jSyoqkVPqVaufNfgn7/2a62pS7/Pd6np3dtjiFmIV7PulNOtxHvJE4hnQqhp7Cxjt5c7MdeMpR4/FUKQma9sw3qKUpeIOYH5STC91yvyTRS2gLeT4rqhQcztn4AuiLmQkJCQkJCQkJC7WkjM7XreVRv4KOvkWD2MSbyScjpRz3fQGelmRJoScZy06+RdJ/g6efe8T+iVgLMyq+8EmxNzqjMCjelUR23lC63MWz2BT7wx8fe6uSwj+RV71cnEXSbwNlmXxYT5k0UFJ/BWzseBCrtsYz5YV8qdmONiiO+W22KkXH6UVYi51z+FxNwO7P2HAkHMhYSEhIT8tII4suChVc/CN51w0dOvaV114d+1GP8kXjE+I1Yhfso/0PaNaRpxUWI8YxPyumVMUzCW0ibFdcZOlsEmxVLRWx2PkxIraZN95TipMZJQEs/rOsyO5Bm2eiee6qUebATevvUpxWMpc5vsj0TcEGM0IL6NmBs6sJ2Psn5r4ekbSMy9BvsxiLmQkJCQkJCQkJBjLMsWndHrxNxeEnNbMTk3yKSdE3VOfG1CL5N6h03KASHhoCvJNyfmMjmHRQIBeyXmzN7SiUBjuqirYF1Lsz0BJt4J3q4iLRySLcDJukzUbfIuW5206yIBeZn4O6y8QcypD4XaiZ5b5g1KzDkpVxJzu5yY24m9/yvACcACHoqQkJCQkJD/R+kG7o3NC477rUe8v/XxN36ln7GUcVCIOcZFi+HTSswpKefxHGBaiC4CMU7qlDFU9Q6JpwbOGVKcZOwsyhRlHFUoMbcNdak3O7MhnJjTcrUVwo2Q3+XQvjRgdskXYjAJOf4DLRFz8vGHWcTcaUAQcyEhISEhISEhIcdQEjF3WT20d7xJzBGYjM9FzKX3ztmE3skzvmOuvAsuEWtI5zvfnHyDndlKHv4EsHd4XW6dsNN61LMugQk4t+JLoQuBnBeIjhN1nayXk/ZykZAm9CTkDImAMx868VddiZKY63zHXCLmXvfkL1Q9Pbuw9x8OnAgEMRcSEhIS8tNIF9AP/OqSJz7gst4PvfbWFuK2EFcHxjQuIoYyhg8RyPMfbx7DE6hjHYlzHj8t3koZYyDjHeYHhMXEBjGX0lqmsZVxmmBakT4cIWj6KCF3zRGwS/E59dP7yldM6NZtxI71EMMzMbddtomY46Osn73oW4tO3/CxIOZCQkJCQkJCQkLuGiExd9nz5yTmeNeZkHIEJ8GcyBdQYo5Q4kzuahNijpN4QyLfDJYviTgn5rIN6ykBV95B10ncJfs50HyvDXU6YZfFgaBcPBSTfpYhLShIuUTMyUQ/2/gCo7N+ScyVX2UdzsTcZdj7jwTWAEHMhYSEhIT8NHIPYCEwuOihG1+z9r0v39ck5ki82Z1yghy/JYY3AD3jI+N+gSYxp0ikGWOjxMOy3HUKjb1Z53fEiR/oZd4hZW5nkLjL+kUMZl/EP/tlr7OwPh6NmCMJ1yTmsOUdc0LMbSQxdzb2XxBzISEhISEhISEhx1iEmJv7jrn5iLnZpByBibF8ldUm96abC0rM0Q75lLY86jn5pnfYabki19U6rK8o801iTif5Ts6V0MWDkWycyIte86rrTKNcgDTtZBFh6aLs6MQc3zH3G1+senoux95/FLAWCGIuJCQkJOSnERJzlAXd920/Z+07X3x1a78ScgP7RyVW53fLAYirKQ1o/HVYfGeMZOwvwdha3PXmpJna2lZ0BrdLMB3rJLit2zMOK7I/xlmNscxnYg5p9KlJzBFuq200ibntSsz5O+auvPhbC4OYCwkJCQkJCQkJuctkHmJOYUQcJ8HMy6Sck3MlwEiycYKfSDgj5mSiT0LN9bQr6gixlsqwZR5lorO0LBAkb/WY9zsAgDbTrIOtI9lKfwFO0DlxB5TU07I86eciwAg2TuBFT5Kt1PsEX7eEfpWVdrYwIKRM67YPEpmYa2ER0Id2C2LuCuz9xwC9QBBzISEhISE/E+kZPvHXTvy9F3yitU+JubYQc4zBRtRNY2uxNoOxk7FbYzWR/kkn0Jjqj6N6DNXYp7FTCbdS77p8d5zeIQe9xMsOeEwVaCzN8TXX8/5ILJf+EJYnEPfFP+2k3jzEHNJCzD1o48erquu12H0PBIKYCwkJCQkJCQkJOYbij7LecFk9NHkUYg6YTczZZB6Te+Zlki9g2sqR97veOok5t/U74uhLyTqzORp84i1APWylnhNz1Bd2QszZBJ3/TVfCj2U2aU/QhYYvInRBQJ0uDHxB4AsIJeY07Xqxt7p8DHZuYm5nveyNT/lS1dPzZiyffhVHIIi5kJCQkJCfnQyuedTatz/vYwNOzB0Yla0ScyTljKBzWOx0Qs6h8Z8xDnmkNb4i3pXEnMdLh8VN1bOu6pvEnNvRh9knf7mOo9EO0kSK9YT3TWA+xIZo+mp8/KG4Y27Rg04KYi4kJCQkJCQkJOQukoKYGyQxh4msEHKY2GraiDmZ1CM9DSCtH38wUo4TewfKBwlLJ/KtyCuYJxlnxFqxTZNtQP9Db0gTb/qnDpNuQOpCp6AP+C5t8Vs4Uc861rMypPWxV074DT6BL/O0w1b8iC9N6yKBCwfa8b/xuuUC4scQc2+pqp7H4gj0AUHMhYSEhIT8bKQg5iQWH8iPspKU86+y6ldYCY21hP7zzWObxki5Aw3pFD89/glItnHOYLES8Djp5QKPpQWUkHPQb2Gf4qn7UzvtF/JpXlDEe7tDnq+ScH+pTvKF+NxBzK2/8pJvLXnQSZ8IYi4kJCQkJCQkJOSukWWLzlh72fOuWn/jZfUgH2XFBPYnI+aoM0yzTMtJuslCgCABJzojy4SMM52AdrotybnSxhcGs2CknIATcrNXP/BBnYH2MrEv8krssa/mSybv5eLAYYsC6mgjk3v6s7T4pg3zXEgoRHeQpFx+xxyJuaHP7aqXveGpf1319LytqnoehyPQAoKYCwkJCQn52cjQmkee+Nbf+miTmGO8IymX47WTcVqmablDjnHRYmGKiQ1oHBQgrXfXW5nFTbdzkq70JTC9+1G9xdUUTzXtdegz1Ud/BX63H3Xmp+13ygusL+7LiDneNTeSiLmLv7UkHmUNCQkJCQkJCQm5y0SIuednYg4TWZ9kZ1KOk14l45offiBsgm9bJdVIuJFgmwO0mQXqjZxD/cbjrKl9nZRLG0VeJ+hEUYdAvQyzs7yScUgDeQGSJ/+OPJG3PBcN5cJB0l63LNOJvxJzE3U/tv7xByHmzhNi7u1V1fN4HIE2wC/phYSEhISE/PSybs2jTnz78z7W3sdHWMdqfcccYjTvjuNWwDjqMZHxMOs03xEPG9DYl9Kd5R4fgUSQQd+JJilHmF+PsYZkzy37xn43+sd+Y2u+SmLOfehjtPbPs0TMsWwiiLmQkJCQkJCQkJC7WEjM7X7B1etvvHxuYo4TdZusax4TXdORaNP/vjOvk30h5pycMzJOypMdgQkztzK5Nlts/WMQcscb4Y+oiK1PynXirRNxlLlO/DDvfrOuLWmto1Bf0hdLz0nM+SLBFxfUy2Tfba1cdNzqAkAWEiTmDm2rW4cm6n5sm8Tc079cdXf/blX1nIEjMAAEMRcSEhIS8rORdWsedYITc9NKzAnxRmKOsdjiscdZjaVWLvFTY6XqNS+xEnFOHllNsdHio8VYj6ceNz02pvgp/nwLn7SzMombRR3JO6QOdB7Xkx+2Q3vtk+tSm/RzkD631UPY5o8/+KOsLJuo11150beWnL4hvsoaEhISEhISEhJyFwmJuV3Pu3qDEHNjsz/+gIlwg5zj5NvyiZiTSb4iEXOJoLO6CTa55qSck2vWkXpWl35dR8gEnLaoY5NyracTcvWX/aaFgdgqSMwJzFeazIsvSxeTf0fyxQk/J/nU+4RfbFjHFhK0wzYTc4ARcy0h5rbPRcw9EUcgiLmQkJCQkJ+dkJh76/M+1p4qiTnGySJmS7zVOCvxXGIj0hJfNc56OuUR4+R9chb3JHYyDlqs9ZhJmxQ/kda46W1kaMzMkHjqMdVjKcvMl7TDbQGtq/WS3uqLDxJzJOVIxnUQc/wd1BXE3Guw94KYCwkJCQkJCQkJOcaSiLnd9eDkWD180CbeDkxyCZmUY1IsxBwm9DKRFzLNJvkyuWeeepQnkk1tM3yCrWX61VaFLhKgt//cOzHnbevEX/OiFx+qV5je7KU92lq5E3MykfffJFvAFgGyEGD/DFqGNIE0XyotkHLackHgtoUPLATaJOYOKzHXOrS97kO9oc/trped95S4Yy4kJCQk5H9GBgtibmqs1jvnGFOLjz8wb3FXY6nFTsYzplM+l2l89Vhn8VFioOaV6LI4KGWAfThC7elD42kuRx7zDo2zVpc4SBKN5Brjq9WTtsy26E+ywTb54z/HSMoJGUciTvvGd8zRt3wAQt4xR2LuTd9ccvqGj1ZV16ux904FgpgLCQkJCQkJCQk5hkJibsdZcxBznMRyoq2QifU+pDlBN+ItTextct95l5wTYp0EmRBohuY75WyRYP+110k8t+7H6jPPPlle/LN+sld/SS9gHSuXhYLnC3+yuNCJv+iKthx8d00i5mSSr3XKrUCIua1KzB3OxNzw53fVy5WY+x0j5uIdcyEhISEhPzsxYq41ublul8Tc9JgQdUSTmMPWYqfGVY+b1BuYFxuLldwipmle03onHWFlgMddxs9UxwF7j5eJVLMYKuSZ+ZL6Vpbb9f54mW7FnoSc+GT92cScvgN2oh7e53fMvembi07f8JEg5kJCQkJCQkJCQu4aScQcP/5gj7IamPYvszrJJV9lBfSuOZvQdxBzbit3rflkHBByqzHpN3BinfLqU+rLpDvXla3bM0//0ka2U9APy9SHtqk+ygWC5nPa60s69UftE2zyL3fLyaRf4WluBQe3KjFnj7I2ibmnlR9/4FdZg5gLCQkJCfnZyI8j5vyOOcZtxlsh3zzmFrHPytMjrlLOOIl4J1sAcS2lS1AvsZL1iq3py3gpdo1tjqeStnqSd/+pP4D5VN+wQfylbelLHl0Vsk7vmCP4jjkScyOfedM3Fz5QiLlXYe+dAgQxFxISEhISEhIScgxlLmIOk169W25Lgkx+MVGXL7M6MZfAybunMTE2UoyTcJlAs4z5o8AXAkLycZEgaa/j9dVXmoiXNqK3tgS08S1tUG71vazRN097mbdp5W4rCwub/GdiDmXFAsAXA20sAJSYIylXEHO3766Xn/vUL1U9PW+tqp7H4gj0AwvkWISEhISEhPy0QmLuLUbMTY7WrenNDWIuP8qKvBNzEvMUOT5rTC7LNL4j3klM1NjYBO2wldjI+FlC4yeJMyfSJC/2OY4KPMY6CVimBepT2it9wU4fjW3663yU1T/+wK0Qc6et/8uq6n4F9l4QcyEhISEhISEhIcdYjkbMCY5OzA1hUs+JvUMm90bU+SReJuIyude66c4105Vl6t8WAeLP61i5570vnJgnAk31JbxuWV/KJG39EB+d9uZXJv5q01h8+CKiIObkMRyBPnrjeRJz7c475m7f5cTcm6uq51dxBHqBIOZCQkJCQn42YsRc/96CmJtibDVizh5lJUEnsdvJt5T2975azGSMtNiqMRwxroiNHncT+LoIg9rx6+iWFvIMIHnGWMk4Sj8STy2mOkTvoJ3ZEuLP24Et75KDP323nNZxos4h8bnjUVa5c+7TF3xj4WkbnZh7ABDEXEhISEhISEhIyDEUEnPbz7p6w00k5kbtTjkFCaryi6wCpIWc4wTeSLlMzHHin+19Ui8QvfqUiXyh5zaRcoT44QJB9dKGpGFLv6xf+PAFg/cv+WZZx1bT2g+Z9Dfq0KfrfRIPWHvSb070i3JZMBgRx/fYcLIv78Y5RJCcM2Lu4Pa6D/WEmHvD079Y9fRcjr3/KGANEMRcSEhISMjPRgaPl0dZ+/duqluJmNNHWBMx50SdxFeNvR5vNQYXMVNiHtJEZ/wUAkzLJJ/0Co+f/DK6xm+toyjiqJBuFm/Fp9YTfylPG9Oxb6kds5E4zHhc1JEy0xP7EZsZq4HhfXz/3HYh5haduvHDVdX9cuy9IOZCQkJCQkJCQkKOsSxbdMaa7Wddvf7G3fWAEHOY5PrkGhPfJjEHHaDEHO+ag05gk3pO8J1o4wQc2wTa2YSePrI+LwhK27QwEJ9apv2w+oUPzVuZtAG95LO9tms68aU+yrzaceJu8Ml8sRiQOwFSWssbX3/j1om5w01irhd1hJh749O+UPX07MLefzhwIhDEXEhISEjIz0aMmOubzMQc3zU3i5gTXSbjJBZKDLf4y1gqsFgp8ZJxUmOux1klwJAu4mMiy8RubmJOiDKJo6yDrcVaJdCIpi/37bHbfTkB1yDmDErMaR36bO+fqNuM1cCQEXMjJOZOW09i7mXYe/cHgpgLCQkJCQkJCQk5hjIPMcdHWIeBoURaqS5BCC1MjoVUKyf16iPdgUYb5GUy7/kE1JMXSxOlXssa/7UXaF90kg29oEin9qzPhEz8dZIvevNXwv0qMWd2jrRwgB77RxcB6pOQL9janXJO0gmKO+b6hZjD/rx9V73sDU/5fNXTswN7/2HAaqCHhyIkJCQkJOSnFhJzbyExt7luTXUSc6P1ICCPthbEnPyTTWLh7PgoYHx14i3FWYfFTCfBEhFGeLltzUbtECcFyDeIOQPSKf4amj7dTxP6DzKFx2q1JzEHNIi5iXrkU+d/Y9GpQcyFhISEhISEhITcVWKPspKY00dZOfHeIttEzHHyi61OyDXPsjyJz8TcwDQmv0JyUa9pTuClvk/mizpalok5rVfaaFraSj4MtJU6TBtSe+yDoTEx528wO8l7W9ne/TfrW1knMYeJfyLmZLJvC42SmEO6H2W9aGf4tp31snOe/Lmqp2cb9v6DgFVAEHMhISEhIT8b6SDm+hMxR0LOiTkj6uxuuRRnJR5qTNRtgTmJOaQNmTwr9YWdxE6LqSTOJF6azoi5ph+Ls54XXbNd17e9HEjEnMVj90H7tvSD+SYxt/iB6z4UxFxISEhISEhISMhdI07M3UBibqyDmMt3yDlp5uSYT+AzYabIpBbTGVqXZcw7KefEW067L9GXd8txAm71G+2a39llqktlHWiWsw3W78AsO+jKyT8n98UCwNEg5g7OQcy97sm3Vz09W7D3HwisBIKYCwkJCQn52cgcxFyLRJwRcnyUlVsl6xB7S2KujHfIN1DG21nwcoD/wLJ80nNuIXETth43G++Cy/EzxVMh69S/bt2f6pIe9Um4tdmu+7PHWNWftUMbsaeOxBz12+qRT13wjcUP3BjEXEhISEhISEhIyF0kfJR1x5lXrbuBj7JmYo6TXU5a0yOrdkdbeadcA6LHhDcRWpggQ18i65yI08UAv/7WtrTYiR+kO4g5mYhLuUIJQ52YM1+W+W/QeiiTrUEm6T7xb9YXFLap3CCkW5rs+4Tf0wqx4aKAX2XF4kMfZZ3QR1mdmOvqGsPePwU4DghiLiQkJCTkZyOdxNw+J+YUs4g5xN/OuCnxjrGzERuLuCrxTuNfjpkaP0tiztNOjglB5qScQeIl0x4/HYVf7Zv6V7AcfpimT2zbKCc0Rjs8NrsNfWs/hqc5x9km75gzYu6l2HtBzIWEhISEhISEhBxjScTcrlnEHCfqiYjbP55JurlgdeYl5hxSpouB9rQTc6Ud29b2newT3+I3Q4hDgfsuypm3vqe2oZcJvUzKbbJe2Iud/Q6BlUuduXQ+2XdwkWB+uchoH9oqd8u1ZDuhX2WVR1l/47aqq2sUe58LgOVAEHMhISEhIT8bMWKuv5OYKz7+wHfMtf1RVhJzEvsQByU+ExrnJO4x3YiPSpwp0VYSaAbRZ7guAXFRILESfhJRxzxtWIZt8of+yNyk8GvxNvnGdhYxZ21oXKaNvmOOZcw7MTf86Qu/HsRcSEhISEhISEjIXSdGzK2/YVfjUVbCiTi9Y85JOUxo+ZJoASbzDp+0s1zs8h1wDi3jhDrrSMyVNkLCdcCJuTQp9zT0mZjTvOsVqpdHcplm/5D2yX7yw7zXNXCyLpN5TuJZbvD25b/1acKv6VnE3MEtdQtbhRFzt+6ol53zJBJzm7H3uQBYBnTzUISEhISEhPzUMhcxJyScEnH+FVYh5uyfZGVcTrGYMY+QdI6hjbiIsjJGamyETYqVDviETqE2us1x04k5JdG0vsBsUz7ptb77Lx9llfaPSsypbnh6G+YG2H4qiLmQkJCQkJCQkJC7UspHWSfnJ+ac7BqaVswi5gRq48Qc00q2eRkAXz75L4k5TrR1kq/1Sts8EbctYTapj2LXuSXJVv4G1eVJv/mXtgj7mAO2MpnnlhN5q+MTfs2bDdBYXDDfScxB14d2hm7dXi9VYm4T9v79eASAIOZCQkJCQn42IsTcWc1HWYWEOwoxt1/jdSckZjIuWizWGMp8EROR97THUCHHOtJzE3M5nWOp6uWddO7X7DwWazsO1ZX+/e47eUzW/Ipv2DkxR5CY429RYm7DB42YY1wOYi4kJCQkJCQkJOQYSuNR1vxVVp2AjzcIL787rUGEzQImxJaWib3kVafvjMNE2cpKlBNtmWCXoE1RrjZqp3n0xyBtyxZ6YlZ6Nhp9wGSdxFyjLZnEW5uc9Lve8toXrZsWAELMkZAjMbel7oeul/sziLmQkJCQkP9JITH31rM+1je1SYi51r7NNV8ZIcQc746Tf5wVEGKOyPFQwNgmMZtgnjETsa4k5hyMjRL/NO35MkaKbj/jK+t3+JA6Wl9h7QP6zzD6Mn/JBkh1M/TOO7cv4jJsm8QcbIWYe9PXF58axFxISEhISEhISMhdJf4o642768E9o/UQJuectCtpxffKWRoTciXkSM4pQVeSWxmc6GIL+zSptjIh5oo75BxOdCl5pmiQcgQXBG7nZUmn2/zIKnTcij/Pe5p6BX9H6hu3tMOknn7Ur/ZfJ/u2mOCkX8p0ki9lya5YACRijh9+2Fr3w64Xi6Kh27b7o6xBzIWEhISE/OzlqMScwmN8k5QjNBZqzEVMkzS3jHHUMV4i/s0i5iw2SvzjNqMZI6GDbYOcm1VudUWH9vjPQs///9n7D3DLjupKHL/TWWplqcPL3UoEmyREjsZBJgnwAAJhG/9sMDkK5e7X6XW3BA6A7QEzHudADkIIWiB17lYCE2zAYQYM2B7GY8/477HHM5g+/7XW3ruqzn23W3hoXmN7r+9bX1Xt2lXn3HvfO3vXunXuYUn/hvPaiLdW59xGHY99mCN+Y44swtxHro4dcz+Jdy+FuUQikUgkEonEAqMV5vZs7EYLcyAS5HJrK5Jb251mia6ELPkjcUeiayIXiXr4OYeFOfrIr2er9kLaOR59xU9l9NkYO0cX6Ngnf84fc/n5y8faI48H2mKE9WYhweS+JP0s6YNjuK0Ic2j3hTnfMfeZ7d3Jr3zKp12Y+x5+AmAKc4lEIpE4PqAw95bLP7iWotzeTd34/ln/2Yi+MMeHOkmYU5wEPRaGUDaBGDhRYqH3M25KmGNMjHjJuBd1sompjIlN/LSYyb6hGAqb2TkmbKDf0mq0+BoCIMfquEHdvlrnnWAZ87If5TxhDm17+EMKc4lEIpFIJBKJE4WeMOe3srrIZSIW2xTmTJxjn+2ca+r0VRvJLoU5XwDQFqKYEmPVabOxkWyHrRXtLPm3fjHamiMY/vZNv51DnFPMj34eI+YJu+YKn/nsj2nsntCPFOaU8PcXBeOwjeM9pTC3mjvmPrszhblEIpFIfOfgwtwYhbl9Lswh3vWFuUaUI9XPmMg245mzjbeMn4qdiHMUyTwuyub1Mp/qsIWQ1ghi6kNJu9Xh39itjmOVehzP5xKtL+JyT5hzcY4xWMJizIuydysrz5tlvZU1hblEIpFIJBKJxAlAK8z1nsrKhNcELolwxxLmlKizjcR3njAXCbrNaSId603b/Xirjd1uY3OR2tmmc7Hz0TGCPE7ZBWBzFL9mjB0/7JzD5lFZ/Jj0Wz38NYYLgLaPyT3pib8WB2Fnwu/juBjgE+LGUfJ35kyY2yxhbuWrn5K3siYSiUTiO4MQ5vZxpxzi6oFNQ8LcJtDqEY8ZUyMuWpwmPX6GXfGT9RrrFBfd3+o+xhlCmn1xBVsTM2tsxZjGHpTYJruP13zRz/lsvEraDqFfwlz4hDBX5+8Lc3ZuMx+5cliYW6b3MZFIJBKJRCKRWBBQmJsLYW744Q+NCId2vU2UtP75RGLctpkQM5kufWAkyrAdjfSbRtJst5uSMVczj/xahq+dY9jrnEY7d++nL+kJfjlfUsk+7d7nib4tAJjYWz36zDeS/y0mzMHeE+Y+s7Nb+dqnUZjbiHf/AeApYApziUQikTg+mFr1uLN/nsIcBbkQ5hibbKdcxO/YMaf4GvGTfuxzWtykvfHTXMaIefLVXD7PEOWDWBjjYuw8KpY6G5vGut2ENxfmgh6T+0Qcxvj+sUcJc1ekMJdIJBKJRCKROIFYufySs+f8qax8+IMn7paMh4hlCXf722yt8EWaYEb27RK6mMSXPpDJMdrHIv0ozJXEnPNormYe1YfIc6R/Y6tzVpbzd38Jc0rceRwej/SEXyXanuzbj0vXRYB81V/PVwsCF+bGUI5hjjX7TJg75bVPpzA3i3c/hLkl/CgSiUQikfi24cLcOHfMIZ6PuzBnwlqN6+2OuYiHil/qa2Mn7OqPOnzCt9D73V7mckb8LP4eO9t6G2tHjmWMPUQ27WJrhTnGcYpyLM1W5zRRjuOUY6CcuenKLy3LW1kTiUQikUgkEicMQ8KcblkNYQ5JbRG5UBZhzuuWSDNRtyS4JvDep3rYwX3Rb7aWkdDX22nAMgftzTyy2eJCu/jgN0zzM07y9lgd21h8/JiFkbw7p4swx8Tek3km/3G7DNveH0l/UAuDQ1u6cfhKnJMwh2NTmHtdCnOJRCKR+A4hhDn9tlwIc4x3NQ4qNjY75iSAMXaxzr4e2YdS41FnTC7+iIUqrT1MzcdxjIkRJz3Gqo+l2oylHlPB3rzuEzG4CnM+V9Mm+aWYRDkJdLDFMTVnjd325V9PmPspvHuMyynMJRKJRCKRSCQWEBTmtj1nvjDHRBo04cvrSNBHCnMk+oywFZqtCmqgC2TDib8SeCb7To3n/Go343vHMsb5tIy+3jFI9vU4ZGPyDk57Eh/CWwhzvIWViwDRk/veeKcWCIewOIA/xbmxg3PdGryWFOYSiUQi8R0FhTn+xpwLc+U35hjfm5hdhTnEK5SFHjMrGddQ0tfjcwhxIXxFu0cdw/s8tsqu0uYr4lrEU8VUF9CasSa6WZ/5xbjoc8JfT2N1IfBot7JyDuYzrOs35i5a/+4U5hKJRCKRSCQSJwYhzN1ThTn9ODQTcLAV5kLsMvHLSibZ6lcfklw9/AF2tvlgBglxLvaFDfWS8PscStI5fogcQz+bH6Rv2wZDLOydr/piXp8jjoUEPerlPORr5G2t9ptzvgCgnxJ9CnMofbdcJPfq51g/XvGH7wT6Jc65MDfz2XIra/7GXCKRSCSOP0KY062s3D1WhTl7+IPF43orK2MY+93mcZF29UUspS/r+000iz4TvND22GrxsM6hcYqh7I950c82ybhaYir7OIax1eKr2uwPWxt3fR7W49ZV7pijTQ9hUn/rg5is48StrHPduo9c+eUVF51PYe5FePdSmEskEolEIpFILDB6wtxGJLn3LswxITZbCGKg95WnshZiPpALASvdz8v45j0WAibIRb0yxs0f75QAaLSdfajHuWn+WrfdcGFvjuE2JeykFgBhM7uVVo9Fhc6/nYt12rljDn7jvKUVyf9a9M185oZu5WtSmEskEonEdwhDO+biN+aOJsy1pH/EsRonOdZ9GEvVdsK3xs46Jhjjyq2sjY+OIZuJcrabLWxmVx/tjMeKyRaHFWPpw3lU5/xGu5WVOwVZGsNfX5b5vPEbcynMJRKJRCKRSCROLI62Y45Evew+U8kkuBXrhoU5JL4uzJkQh2TYhTlR88KmpN/GW4Jv7Z4oFzbZOV+f82x+zGjbOdr5Fp84V0/m2VeOR5Z+WwCU1yabU7ewMsG3BYRYzpPnz7Z9Yz8hYY6/MUdhblu3FucnYe7Veiorb2V9IJi3siYSiUTi+MGFufH9FORMmLMY18T3aIsWwyRiDcdD0MQ48+kLc+Zb46R/oRXt8AFDmCNrn7dld/HN20Zr27H6/RGDNY/aJI/B10FR7hjCnH6SIoS5uW7dLVd+eflDz39PCnOJRCKRSCQSiRODlcsvWbXtsl3nNr8xFwIak/d5u8+QkFOwMsGONhfp5A9SmFNZRTizo112tcFHiTYT9uES/aIfT3bQ5zoWW2GO43Sevogwm7EnuInmX8jFQCwA1A7GosD6eV5aIGC8nbefKxjCnB7+AI5hEbAW7+3MZ2/oVr7yqZ8dLFmyGe/+g8HTwBTmEolEInF8MLXqcee0wtzBTR6bGN8Rq1phDnXa9Ft0xeYxjaWIWKcY57FSMdx92A4iNsYTzoMmlrGMtglsRWjjebGPpeYYpglwNt7mtjbtNj58qzA3nzHHsDA3qVtZr0phLpFIJBKJRCJxAkFhbo7C3I39W1mRjJMS5UQkvkyAUfaEOfaxrnZj4245lMGwmTiHBFlkH0omzZ6UF1/N4/Tj0j7c35u/pZ9TPU+fR+dujDapBUFQyX9D2YbsIcyJdv4xviwAJMzZwx/GD811a3k+n9nZrXzlUz43WLJkK979i8DTwRTmEolEInF8MEKY4++t9WM1631hTvWwua/iazuWdcZe9jHeqQ+cFy8ZB+lT42MIauEbO+Bot2NwXO0XWx/RBTmVMcZYhLkR4pwdm3F5SJjD/LyVNYW5RCKRSCQSicSJQyPMTc3bMWe7y46+Y8594M92TwSD3cS3IPr4UAinCWooMZfocynhB3msSd5mEnWf28YZzW79hfRv5uJCQg+HcH/Rk3p7wIPZ4phGT/ppd187ji8ESNZpY7JPP9Xp67ZDWAAcpiC3rRtX2QhzL38yhbltePcfCp4BpjCXSCQSieODEOb2jRbm4gs3E+XIqFt/2X2OesRa8WCtl7jHMmJfa6cvhbKweSy1eUwQMz/22diYS2SdsVhxl2zmKP0cY+epcxV9xzrrKunHOseP2DGH+ZtbWV+Md48/MZHCXCKRSCQSiURiATFvxxwSXE/ITWyrQhzLGYpc3i7CV5DjaBORCKMM8a0IZbKxzxLoklxHf7RVNsl3mc/HanydI5L6egwbEwl7nKN8mkVAb5zqvnAAy9xaFFgCH7ba53XOqQWD23rCHDnXrcUiYfrTPWHuYjCFuUQikUgcPwz/xpyEOca/zWDEasbREcJcS/iazWKixVGrFzLeBdG2OsagbFn6m7HFn8cJW4m3tHs8dhvFNOv3sRrDsX3yt13tPOaTfSbMzeG1zGnuFOYSiUQikUgkEicWfPjD3GW71vM35nbPIkm1ZDyErBkkviOFueFkHW3ZuCOuJPRIhNFvdH/a0F/tloCXJL3YnD5HLBRMnDPyGJaU27xxPsGyyCg21JGUi42fnRMXABTWfCEAxvkU+9CCQH3R1rjG72jCnO2Y+4PBkiXb8e4/HDwTTGEukUgkEscHLsyNlYc/zHZ8GIIJc/zt2FaMq3X1q26xVXEwYqvbLB9gnLP4pxgour9ichNrw0d9xoix8sc4MfoUZz3+i43tkAlpPF6J5348CXJR5/HimF6PBzdxx9ykC3MzEub48IervnzSxee/N4W5RCKRSCQSicSJQU+Y29hNF2HOBLcZ1YMmyEVfSdYL0ZYwF3Yv5du0ZUN9iCUxjzb8RbZ9nC0gjHWsJ+Ma0/fRAiH8mJwrWTe7KD/Y/TjzzqGljy990fZFA8sgE//erax3VGHuZBPmduDdfwR4FpjCXCKRSCSOD0YJc0V4MzL20VZ2yjXCXcROxT2Pj4rjION/xECj+zF3kMBm7b5PnxZTzd8Iu8fWENJsZxzIMV6PUrHX5yrHxOsoT2KFrfzOnM/H3XYcH7eyTmjHHG328IeTLkphLpFIJBKJRCJxonBUYe5opChn7AltpB7u0NpRNn71m3fa0c++hpasN7biy9KPgXZ/oWB9SvJJHafPMl+h+WtMzKX5ho7v7ZbDCwgT4rze2zHXCnNgK8y94of/cLBkyQ149x8Fng0u5UeRSCQSicS3jakzJcyt3T/b8XbWEOb4YIe+MMcYOVqYU5xEXDRau8TcQvdTbETd8we1j8EST8mIqR5PQ3gzepv+7hNlfz6eK4W4Rphj6XWNOYowx114eiqrCXM/jXcvhblEIpFIJBKJxALDf2OOwtxUI8wF9WAHJfLRDnEObdpE/4adT2LVgx2QBMs+qnT6eCX9vQSb/dEHws8EtPANm8/jrPPA1+eK12B1JOVi48+x7Rg/rtk4ny8IRD8PLT68Hwm+LSRYh83bso28lRXvI4W5Vz6ZwtyNePcfA64CU5hLJBKJxPHBMYQ5CXEuvlHIUp3xTXGewpzHOtkQ2xgnI1YqP/D4F3708XbEUhvHvuqvhy2RjKuMmW5vRTnFzhJPq82Ohzbnj37aWPfzqMIcf1NvC8idgmi3whwFuYNzEuZ0O2sKc4lEIpFIJBKJ7woUYe7Gbmr3bF+YQ6Ibvx0XVKK9z4W6IZowZwlyScojqe/VQZ9fCbwn6Jbcgz426nUusPGz8X0q0S+0hJ7ULStej8WGxoStlI2NiwTvi+S/LDzUDz8uDA5ZPYQ6LS6Y+B+mONcIcxirHXOvfsrnB0uWvAnv/uPA1WAKc4lEIpE4PpAw94IPju2b7fhkVhPmGMNCiDNhrrZRlzBnMS7ineKcx0vFvp4wZ4x6jFM7/L2tL/M8XhbxLcY1dvUVW42nxcfnY91ibnuuI25lFd2XMRlzxY65EOZYX3fL1X+2/KLz3jcYLHoJ3r0HgSnMJRKJRCKRSCQWEEcV5iwJ1u2rSNZ74ty+zd30scQ5lqIn3k1/iHS9pN+PpXb4+znEPBxnvjYP7VEGbRznMiEuxLhSV2JvLGOi7b7tHBL2mNCrH3aNZ7/59IQ5/bA0/Y0mzCHpPzzn4pwLc3wq62ueRmHuZ/DuPx5MYS6RSCQSxw9FmNvUTSBWjx/gb80h5rkQp1tWh9qKiRFfFfNA1j32Kh9gfqA4SJELdtXpG2PN1/x9DErlEe6r+MixGmfUnLJ5HHYbd9nZTrvor31xzMmD3BVnr6PevsqHPAQZj+GvL9BSmEskEolEIpFIfDeiFeZun1UCXcQnJOVFmIsySGGOvymnPibLVhc13hJoJdJu09xM2lGKsCuB93EhllmC72XMUeYzO/3a46hPNm8jIbfE3oS5vp/76hhuY/LP0ln8fR4dE4uSUtKmPkvwleSTnCds2jHnwtyhKsyd8pqnhzD3BHANmMJcIpFIJI4PxkyY4245CnMTBzYp7jFmWuyLL9ZcpKMdVExlXPT4KMpuY+2LO9b7sbLEZVIx3fxtDOoS9OrYiKt91vhJ6gu1IZ/ecYKKx0a7bZW0+EyaMEfWNsU4+425Odi2dTMfuerPll103vsHg0Uvxbv3YDCFuUQikUgkEonEAqIR5iZv34hElQl7JO0mpo0U5sAizEUSTpvqlUrcaffx9o06EudC7/d6Seq9v8xR5oOd9L6Wmo/nwTYScLJ8S9/6NJSNZPLudZ5rqfs8tgCgX2UsIEYLc/Bpb2UNYe4zO7tTXv/0LwyWLPnZwWDJE/EJpDCXSCQSieMHCXOX61bWKswx3tnOuDbG2465WkbsDSHPyFiK+AWWWCpyHGn1Ejt9rL6MU6xkWcdGnIz4aqzxkyy73IO94zWUEDdUh3+wfytrFeamJMzxGPyNuSv/bPlF51KYexnevRTmEolEIpFIJBILDApzOy7bde49FOa4Yy6EOWNpoxwW5kxoY8ldaSiROLcsSTj9vH+ecCcRjnWbyxYNUXcf+iKhtuQcdmeZR3aj9bkN1G0w4adjNeOausYU0u59Pk8csyb8TOyN/PFokgl+LCoozPUe/lCEuRu6U15/6RcGi5b83GCw5En4BNaCKcwlEolE4vjgaMIcYrl+S4473hljEdclyLmwFbFPwlxpe0yHTxHXEEtLv+o1bkqYUztsVrYxVjvZFStRIlYa47ZVZ+x0L/HX5+G5BTW/9YkRnw/BNiTSTeEYnKvuoGuEuZspzJ33gRTmEolEIpFIJBInBq0wx9+YY6KLhJfJLkW5exPmyBDbWlFOIhzpfdEf9WD7rXwk22qzHj6NvyXhtS+Es9rvx3Hb8I65YY46ButK+GMhwuQ+jov5jJbcmzBX6/OFuTmJc2Mo1+A1zXzmhu7U1136xcGSJW8eDJZ8Pz6BMTCFuUQikUgcH7TCHOJ3T5jTz1BYXLcntMKu2GZljbGsIyYi3llMRVs+82NmtbVjgyP8GZcbSpRr4mdPmNPYej7lGMoRWqIPc4mNMFd2zIHxm3NB7fZHue6WK7/iwtzL8e49BExhLpFIJBKJRCKxgAhh7pMuzDXJucQ4CXImzCmZB6soh4TaSybMJrxZMl0EPtL77Ddj/Ftx+SJZxtiSYLOUP49nNn0b7/OHYFboiXfMRcYxrO7kMXVe7DOf8I9jau72XNiGr0Q5kueiY5pddDGuL8xxjPXx4Q8U5sZQHzvYCHOvlTD3lsFgyQ/iE5gAl+mzSCQSiUTi24Ue/tAX5vTwhyLMbZovzLFU/LJ6xL8qxlnd4jDim8e6UQKa0ccwfiuGcx6jxVDOBzup2GnUl17NvDFPjcPeLrGbr68+9IFjJyjMOc2GOu0YO455tWuOeQFjPuwS5h563gdTmEskEolEIpFInBgcS5hD0i5RjqRY5vUizMGHZV98Y9Jc66NucSXnJ9dNvaF2zjVliHSWqFtSTVqCX5N7S+r7IqAYvt7WnJzLqeOoXue2hUh/vOwuxhVhTn0sq22UMHcKd8wtXvzWwWDJD+ETSGEukUgkEscPQ8KcnsqquFbjOGm3siLuKe43sc1pdu9r4p/iq2Id4vlwjA0/lCWWqoy6z+/1MqfHTo1TP2Jo8Y/xzbmor9rr01hdjDtkrMIcfeqOOb0+xGT2T1dh7hV491KYSyQSiUQikUgsMFyYW1+EOSTsvmOt7JYjldAjKUadNpE+sEmck38V3VoRbpQ4Z8k+kmWMlfCmNhLoqIe9R0ukLVlnos2EGzYxFgpM7Gv9WIuGcg7OOB/2acHgvqXe2Di3CXP8fTkm+fWY6lO//cachDneyor5pz/L35h7xh8NFi/+xcFgyQ/jE5gEU5hLJBKJxPEBhbmfv/yDY3vtN+bGD5ow137Bxnr8vlxPmENZYqvHxeHYZzEWsVxk2/uLT21HfDX252/jt81j8bPGUfdrxveOxbma+Sm8SYCLccPCHOoS5hizGb+RU0iY++iVX1l+8XkfGgwWvRLv3kVgCnOJRCKRSCQSiQXEMYQ5lkV8Y9mj+7R9IZ4hOTbxDUTSawk0+2GXuGZtsghzLJlU+3hLuPtlzFPGlrnNpzASeyb53s9jtgm8zcl5bEycZ8yhelkc+Fxet4Tf7cXmPmC0J4owt7UR5nZ2K02Y+6XBYPGT8QlMgSnMJRKJROL4oAhzG7uJfZu68YOzJsxJgGO8tvhNscqEOcZsK2UXaz3iqH2xtk3UGI+bx6IdA/Wgx9KIl+EXc7exO2JzP277PPTzOU3Ao/BmPirdXkQ6kXF5ri/MwX/6lqtSmEskEolEIpFInEA0D3+YGnkrq4tuIxh+PTsS3borrqX1SZQj0SYt6fcS9t6tqiiD7TwlQUeiHfa+XyT3td/mjrE+noIgyv78wZjDyKSebQlwrB+y9jDtt3eMPWGOt7LiuNOf2dGdcsWlfzxYvPg/DAaLn4JPYBpMYS6RSCQSxwcuzI1rx1wIc4h1/OJNX4S5MAdSmLPfXmWdsc/iopH12o7YHsJc9Qu2Yxlbfc6WEVtJxVfzJTW37Na2uN7GbZ9fc3EOq1fhzXx6vzfHkvOoPSTM7UM/coPpW6766vKLz79pMFj0Krx7KcwlEolEIpFIJBYYxxTmXGwjkdTOOMMWyTN9ix+T3OinyCbafHW3HBNlm99uHfWEG30mzDU2HM+S+RjnYws5xuaxucyvfpPvdvb7mOIrm80Zi4Vapz0WCBTaUHIhwGT+EPqY2GthYGOi334XB8k/2n1hbpt+Y27qM9u7U95AYW7p2waDxU/DJzADpjCXSCQSieODqVUmzPE35oaEOYuDfWHOYixiNMr+l2r0JS2uy0d1i5HyUdxGX6G1ObfNDx/NazvtRLYZP3vHibqT85Q5vGzmM2HOzwMxVnGZfmjrd+Z8rF43yfgMThycAz1+uzA389Grvrr8YRd8eDBY9Gq8ew8FU5hLJBKJRCKRSCwgjiXMgSGyhSgnIsmt35wjIRbNj8m0xniSHgm6zQe/sLVEn6iEvibjRibUnlj7cfp1sJmrl9i7T7DMWXzD7vPzOE4m+zqG2kziWfdknnUJc+5fxsOGRcdRhTkuhrhj7g3P+JPB0qW/PBgsvhSfwDowhblEIpFIHB+Uhz9s7Cb211tZdVupxz3GR7vNlLRYWXbOMfbNYx3XCnMl/h5tHP3clwxhrrVFrI+20ecL34bTLsRF22IxX1/UnRhvgiRp8Tse/kBO74dNwtzVKcwlEolEIpFIJE4gjinM+W/MkbCbMOel7EyykRgXf1DCHGntSNiVtKOvCnObnbW/FebkV+ZmGQx/a9tcUdo4HafxD9bj1Ho7X5vQK+Hn+GKzpD6+pbcdc94HP7ttBrYQ5tw2Rh7aArow99mdIcy9YzBY/Ax8AuvBFOYSiUQicXwgYe55LsxtRiwyYU4iHMoQ3xT7aEOcUpxEWXIA2dzOkrE16o0wZ7E64qnbQIl3rNMvGP2NTYKaxlssNR87VhHwWiKWSpjztglyPofafepLMjBi9zxhDue5/mNXf/Wkh19482Cw6DV49y4GU5hLJBKJRCKRSCwgWmFuz6wSdiXtSs43K0k3Ya4KbipldxsS80q2zWYimSXYIgUztE1A42/d2O00IykfJNaoW8LuiTdtPo/NjbrGsGSfzy+bU0m59c+j+hofMBL6YsNCoAhyTVkSfxyjL8zRtrUbx7xjKMdYujDHhz+c8oZnhjD3THwCFOZyEZBIJBKJ44PejrnmqayMdSU2ImahjJ1m/Vhpvv24a/V5MdSpvjLecoOeT+lzwQ60MT5ONs7vhN1ueQUlxnGnnZUWj20+xWHN5fWGFN8sVsPfd9nZl2YuzDFeo0xhLpFIJBKJRCJxYjFix1xQiTES5CLGiV4PP7aRsLNUMi3RDP0Ux4K9vrAfQ5TD/CIXA6r3+4s4F34No8/EvxhzFF+WSN77jETe257MK7GnXWO8jn4l/zwn1l2UI+cJc/CnMDfz2Ru6U1OYSyQSicR3CkWYm3Vhzm9l5a5tj6MRA223GeuwRdyTb8OIq/Rh7CUxpkfFwBH2oM/df7iDx2cdM1jPr4wJYW6UKHcMTrjwxvqxhLmZW6/92vJHXPiRwWDRa/HupTCXSCQSiUQikVhgDO2YmznkgltLJMkqkYzrAQ6RNJcEut7yGkl7X5hDm+zZgu08rc3YLg6ib+SuOJ8nFhA8hi0k6lzl23cm9vzWniUTdpW1HUk925NK5j2xD3JsI8IV/8bGRdA8YQ4LpJnPpTCXSCQSie8gps583DmtMHdg1uImYjV/QqI8/AGxibS45yXtQcbZEkvNpi/MvG2xNeIo6zGWPj5ePtavHW+Ms2FjrPZ6xGfbQWdzxLx9Uc6OZ4Ib6ozRmrMffynAqV38zJdiXRHm/FbWFOYSiUQikUgkEicWRZi7Yb4w1whu1kYC7Il63T3HtiX5IczNu5WV/k616edUIt60jWaLJLtHLQoi4Z/PeYsB2b3uSX6fTPKDaPMYKEOIs7on9k3ZF+ZwzPBzG58KJ2EO9hDmVlOY++yNo4S5/I25RCKRSBwfSJh7wQfH9m0qwhzLyQObEIOrMBcimGId42fEUS8jfpe4y75oy8diZo2vfVv4trZRjAdCVHHO+6Ie4pvaFn+tDUqwM78Sk1s24yKWm20OOckcfFKYSyQSiUQikUicaLTC3O5ZPdyhCHMHN9vDHpCAK2FGkmsCm9dpV3tImPO+SOqD1eaLAhEJs9f7ghqTavpyMRBts1myT992HmPti3lQZ1LOsZ64y65EnfTknsk6feSHOUQf0/jxoQ89f/Tp23n51TkkzPGhD7Dr4Q/oM2HOd8wt1lNZ+fCHfCprIpFIJI4fdCvrsDDHuIiYibbitmKjx0fGL8ZP2tBf4yjbrNd+E+msbXEX40XMiXkkrBUbfW1M6xcxOXgsYU477MImO9ryd0q0MzL2cpzFYcZjo83lcZs8NNdNgSbMzbkwdx8Kc6/Du/cwMIW5RCKRSCQSicQC4ig75uzpq85IuJHcSoxDOYNEe4bJfVB21qufEvp9XAhYnfaozxfVmMBHEm82JdnR9jIWA1E3/3jC69CCQrREXXW8lhDdaqJvSb+OVWzGmuRbQi9bUHar22/WkK0wt3WeMKdbWT97Q3daFeYuxSeQwlwikUgkjh9CmNvf/425IsyhVIxuWIS2iK0R40gJa+GLeghtR423DdlffNyGuEjaba1O2Si0NaKbM0S1iMPtHC15rtHXi9Oy2XiKcqOFOe2YS2EukUgkEolEInECEMLcJ+f/xlwrzsmG5DZEtxmUFOe+NWHOxLlWmOv/CLUn7LSzP8a2Ni9rom02+vaFORPnevMG8RpMnGO9T0vcR9hEXwzQFoRdRL0Ic/ALu9ohzOGYa2FbvW9TN/MZ7pj7kT8eLF76tsFg8dPxCcyAKcwlEolE4viAwtxbX/DBMe6UQ6ydiKeyKu6aMFdjpNlMtDPbtyTMuTgXX6j1+udxqA9x0XbGIbYWm8VZ7oCznXMuzOm34UxUC5+jMuYCRwlzdhurCXOkhDne0rorhblEIpFIJBKJxIlET5jbVIU5Jseo94Q5JLoS5ES0kYzHwyCOJsyNJPoooMUtM7TZQqD2y4cJNfvCp/h5su1t8+8vNGKc9UeyjjpeR5nb7Srb5N6Tf7sNppZmsznjx6X1A9Ow2265yiLM4X3Ub8zBb82+Td26Iswt/g+DweKn4hOYBlOYSyQSicTxwZAwNx7C3MEQ4EyE0xdZipdGi78WE40W74owxzwAdt56WnfNuW/EXLY5jzPygcgh1Ea81E44zkPCHsJbfdCD2YpIJzttxji3Np6TOn8cK+J0sfv8EwfnEJt919x+1FOYSyQSiUQikUiccFCY23nZrvM+dWM3tXdTN3MYSbBEOeOMkmJPqpHcSpALMulW4t3UKcyRsEViPo/sCzbtVkyzhBsJtdvoF32j/MNW+nptS9bntz1ZR0lxrV0cKPFX8l+pfozn3BLm0CblJ24rpL3cyor3jqWEOf7G3BXP+qPB4sW/OBgsfjI+gSkwhblEIpFIHB+4MLd2/2w3XoQ5xq8Q5aoY16fFxxLzPN6ZHfQ8oBdDQfm437AwZz4YpzyC+YRT9hpvS+wtu+OqrYpztBn754g6bUE/n/gCjTbO0+6Y062sB2zH3AyFuYenMJdIJBKJRCKROFGQMPe8Xed96o3d9Dxhrtkxx8QaCW7ZKYcybn0JYU71fUiAizBndUvOrT/qNh/rllgruUafyLraNj4SfSX7PCa/4R/yb+fRXE3d2M5hbVEJOxN5Ju++IHC72TAm6mqbjQ93MGGOx7I+LjKOLszZb8ytlzD3TApzbx0MllyCT2ACTGEukUgkEscH/htzEuYQd0KYi91xU/s26WcmJshGlFPcjJjLmBYxM+ImYpnio/rNv8RGUv4so+7jQPtyz4m4KGEOZYhwZadcK8z5jroqzEUf4izHaw4cB/1sxxdltlPOhDmL2fDzsfOEuUPbXZi7IIW5RCKRSCQSicQJwsrll6y+4Xm7zv3kG7upvZt7whx3y1VhzhLp8rtyTK6VdJtQ1tstR7Lfx0RiXoU5T6jhV22+MIh2r+7+Pk98u1/mjflaaky/vzd/6xv9DdVWwu/kcWVzhm/0yR7CnC0Q2oc/jMNOYc52zD3ji4MlS94yGCz5QXwCKcwlEolE4vghHv6wrxHmFKtchHNhbnI/7B5P2/gY9cKIpZhD8XIovkbMbOMt6zbO2lWIg00lbIiLxloP4a3OE+PI6De7HRfHKXE4dsmxzphslCjn7AtznGuuW/fRa7+2/GEX3pLCXCKRSCQSiUTixEDC3OX2G3PcMUdRDolsK8rV21cpzEUdZB2UMKc2xu5Hcs0S7Uiq2VbCTl8R7VJ3xoJhyG4Lg37ib7Z7EeYKGx+fy44f/X2WRF4l/MWo1/4qzNm38/ZNvX2LLxsoYQ7vp4Q5LAjW4DVKmHvDMynMvXkwWPID+ATGwaX6LBKJRCKR+HbRCHPtU1knD2xCDNxkO+YUb1Ey9ir+Mj62cZf+pPUzjsZPWvS+cFN9dFvzeZt5hQlqsHm9CnJRhvAGarecjQ3/VpQjq2Bo9WBvpxzHzBPmWJ9DbjKn/kaYez3ePQpzK/Q+JhKJRCKRSCQSCwIX5s67+4ZuWg9/sGSZwtw65/COuWArzplYB1srzDEpV2Je+y2htoTdyAWBJf1FmNtHoq5v9N2P/RyLucOmhLyQiTn7nKVOXy81h/v6HLKFj/fLhwk9j8eSopzbbXFDXyf6Q5jrP5WVu+TqjrmxEOY+p9+Y+8JgydKfHwyWfD8+gTEwhblEIpFIHB8cU5hDvC00YS6+GKt2Z7F7nI1cAPFSbdpVt7YJbB432S96f7HXeohxGuNCXBHpGmGujbkxtgpvThxLcbrpm0Q+I0GO86GcPmgPe5hohTkc59yPXPu1FQ+78KODwaIr8O49HExhLpFIJBKJRCKxgKAw98bLd513z432G3MjhTnaLCE3UW6LkYk4yr4wR5qtFeaUtDOpZuJMwm5sEn+3lXEU6KLuc/R224WN85GwhchWdsWBJsyB4aPz8HYZy7bZbE76oMRr1yLA/WxxQ8JGOyhhDm37DR+jhLlmx1xPmHuDhLmfGwyWPAmfwFowhblEIpFIHB9ImLv8g2PlN+bs6awU2uYJbwc2DdkR9yK+ervEZdpkb0rGWcQ80kQ52Fi6rbC1yY/+VZgrAp3q1g6xLuJqG3fJmEdtnE/bp34Kc5ynYe9W1v2cf65b/9Fr/nzFw+7zsRTmEolEIpFIJBInBiuXX7KmFeYowIEU6EKckzCHpFfCG8oQ5qLeCnMhnE2XJB9EH6mk2+shyJmQRh8uECh42SJA1I45m6P2IeF2xuKg9IlNX/THGNWtHQuEMi581OfjYoEBhh8XNyHElQWA2uy38SyHb2UNYW7mD27oTr3yR0KY+z58AinMJRKJROL4wYW5tfs3duP7N3XjB2znXAhwEZ+1g86FuYjDjJuKk4hpFsNR30eb98tu/WKJpxEvYWPslM3iZ9jDVnbEFcEMNu1oQ10lcw36UjijmGZxlWNL3NWuNxsb/bUPvupn23yOKcx97Jo/X/6wCyjMvQHv3iPAFOYSiUQikUgkEguIYWHu0BaJcxTjKM6tc4EuhLf+Laz223L1N+iYsCMBVp8n7yxpF5lYW738ro3XW2FOfmV8XSjEHCaume98ot+PU2zFn2N9bvnQ1+2Fta+9hTVoopzTFyBhNx+bo79jbmu3FouE1XiNFOZOuVrC3M8OBkueiE9gDZjCXCKRSCSOD44qzG1C/HOBDSXjq35nrleOImIbS8bpEk+diqd9UhyrgpyRvjW2ehl10YSz2gbl48IciXbUW//SX8gx6JMwV+czuwlz5Ax/egPtdR+75s9XXHz+rsFg0ZV491KYSyQSiUQikUgsMCjM3XD5rvM/+cZuZu/mIsRptxzqo35jLm5VDWGu5bAwFwKZJezWZwl9k+RzkSCGHxjj2V/qGM/E3G2xMIi5Q1yzeu0v/hhv/eZvrPYYKyqR9/FgCG89YW4e6zz9p7KSc91q+Mz8wY3dKbZjLoW5RCKRSBx/jBLm/CmsIcyZ4Manso4Q5ny3eiVjJkrGadYVMz0+enxt42orhkU8pW8vvrZ18WjCnPVJcCv+JrzFbjir23HDrr5hYe7ANsRm+505Po11Rr8xt71b97Fr/2LFwy64dTBYdBXevUeCKcwlEolEIpFIJBYQLsyd98k3dtNDwpztnDMWYU5lX4xrd9TZgx+sHSKdknZP5q0kLeFX3YW5SPBLwq+FgNVtLBNr85no+VVWkc7aFMzKvGDpb2yi+1fWOVraLjn0NdSDH1jXGCP9Jg5x11wV5tZgvAlz+o25EObyVtZEIpFIHD+U35gLYW5jN6Ensc6CLKsYN0qYo4hn/ta2uMl4XIU5sxlrjIwY6H2Ii8dkEzPjttUQ42K+2t4mqo2xZZzmokAXghziL33pI1pf9E8caIU52ue6dbuu+4sVj7jw44PBoqvx7qUwl0gkEolEIpFYYPDhD75jjsKcBDkku8PiXDz8oRXkRPWZ73xhDom1kvj5HE74izCHdhXM3NbS+48mzFUyaTdhrvh6KYGP9aD8v0Xi9daE3xjCXCvOqV2EOe6c29atxuuZ+dwNLswtyd+YSyQSicTxR+/hD/OFOT2NtUfG2iiNVbADEfuKMKc6xTKWxnlxEgwBLQS4Um+peRhHKbgF3aZ5PKbCl9RtrdEfNhfdqjBH8c3jsWi22j9XWG5l3XXdXyw3Ye4avHuPAlOYSyQSiUQikUgsICjM7bx817n33NhN7fGHPyAhDkqYYx0JLgW3KtrRDpvsVlebiS5L2EyAs5KkaKe6/5B0CGM9YQ60ByzYQxZKoj/cp4VDkOKb2cNPt6eSOBfrd7vOi31W5zhbhLDf5qj+Nt5o/ibMkbAdQpIP9kU58ys75mAPYW4N5q9PZV2St7ImEolE4vhjapWEufH9/G05E+bGdXvqbDe5bxYxkHRRTg9/MBFOsdhjYXs7q+I24yv7Ed8kniHGxQMdGPNCaAvhLMiHPIiIgUGKZCqLGOfiGUsfpzmDjK+KsfTxfreZyAe793O8YrLmoI/Pfch2ydlvzLGNc2h2zC1/+H0+kcJcIpFIJBKJROLEYOXyS1bt9Ic/UJhDEqzfj2MJhgBXdtChbsJbsPqqvb/aQ5AL0q56EeaY6KOtZN+SfrOZPagkHOPIIsyR8A1Gf48YGyKb2M7rbZun9WnqRxXmmOw7/Tds+uIc6xgPSphDuRYLgyrMPfPzLsw9AZ9ACnOJRCKROH4IYW4ff1tuhDAnca6/Q460359jfOzbGS/18xOIl/OEOcY90e1HYQhyPYbQFnOgXohj6rg6HuOvzwW/esyGMU5tinEmztltrfWYVZjzp7Ie3N6t562sKcwlEolEIpFIJE4YijD3Ru2YM0GtPtTBxDjUWaoewhvoSXP42gMhPAlHe1iUa20l4adNyb4l/2Yzu/XZfJGkFzGNhG9PlOOYqIcvS/dju53b5mFJH7dpnLeLMGd9IhYfPXGuCHNukz8XA5jj4OZuHPajCHM/g3f/8eBqMIW5RCKRSBwfFGGOt7DOjhTmJnVrK39jjvZKE8HAnp1x2+L0sDBXBbEar0dxpDAnWgwthK+oGMzjcl47juaCj8bQZ2hcFe1MjLOdc6hTAPTjmTBn1A7/g9u7c6swdy3evRTmEolEIpFIJBILDApzOy7Tjrmp3bNIVJF4gyG2UYyLutoqaQMjaW4T9laYY6JcGL7GNuHv1znO2BPEvI/UQmLINuxnPu7Xowl0JtQ1dopwogtzsjPZ51zV1iMWIi3LooKLAbTHD26WKGd0Ye6zRZh7E979FOYSiUQicXzhwtxYT5ijEGeCnHbLoeRuumD8xlx8URZxuI3FrTAnIY5Un7cZ6xH/RI+fNQdo7UP0MayXXXS0e0y245st5q/jm/OLvkIT5Wx+F+ZclDNhDkxhLpFIJBKJRCJxwtEKc3tmkaRaUt4X4vrCnIl1jWCnZN4S9JHCHBLrwn1OjgFNSGMZdfRhbJtsi94nQY1sbMN+sfvtaLQ5ULY2nKdsPbsl/ryldVickx3vgxE2sC4GzDau35br/8bczGd2dqdUYe5xYApziUQikTh+CGFu/0bErlaYM0FuinUX5IowR6L/2MKcxT6JZop51pYNsa7HsLcMO0qxqZsPRblGmCMpznm9zOH+Np7nQZuVPUqUoyCH/kNkFebsVlb+7lzvVtYU5hKJRCKRSCQSJwAuzPHhD5MU5iTKgUhyQ4iToNaKdUhyh3fSlSTehbnSZuKM8SJvpWmFOf7WHEmhTLfZuB/GSSxD3UQ4F+NA7XhzW/hagh5jws/nweuoSbzb5AeylH2rlTpe9YnFgPq9r4pzIcpxLNpDi4IJLAIozA0//GHm00WYeyPefQpzq8AU5hKJRCJxfFAe/nBvwhwfDmEPiLAdc/VBECJiWxHjFOMs3iqeNm3FSdW9Xeoca+N7/qzHvGqDFNGGd7exT3TfENnkb3PoOLR5m1Q8jrFFmEOZwlwikUgkEolE4rsSRZi7oZvkraxIaFvBTaIbEnQy2kWYc99IwG1XHBNkS5ItUUZiTLGLlDBnglzsoIs+iWnR9jFVDKvCnP0unDGS8OKP+jxhrvURUecxyGJzluPFGPOxH5D2+QtdlBPR5kIAdc3jtrJjDnWWa7AQaoS53DGXSCQSieOPEcJcuV2VwlwjxA2zPBTCY6ziOGObYpzFxSKueTtoMf8o7d74lpYztPGTHCnMNf1l3pY+f1+YM0rUa4Q50p4WG8LchSnMJRKJRCKRSCROEIowt3O0MIcEXcIckvSw9XfMxW2r1ncsYW5KohyS/Z4wZwl2K8wp+aYd9RDCtKuOpCjnZU3I/Riq+4JC8zU+6uexnDouGeNoC7/Wl8evfn1hDiUXAKWMMS7M8Tfm4DuGci1sJsztyB1ziUQikfjOYcRvzNmu9PiNuXgqK6gdci7Kod7umGvjuO2YM5vinMc+xWvZzLf499pO+itONn4SzWrs7Alpmsfmkq3011hcYzZ9WFr8tXhs7TIfxb4Q5cAZCnOHtnfnfvz6v2x2zD0aTGEukUgkEolEIrGACGHu7p3+G3MuyB2TQ09nFc0uYc5FNRPX2GYdibjTkn6zlYSaHG6DNfn2MapbO/pahjAXO+eM6AuOGCPiPPukLfotsSft1lWnt/lkVpUkfdVPYY6iXF+Ym6Ywd2UR5h4LpjCXSCQSieOHkcIcBTd/KqsLcxLhEC/r78oZJcqRiGMhrJk4FjYKXShFxjzrD98qivVtEuEYIzHO6Db6kx4/5zPmQr34RnxucgGfI740K/EbLMKcP5W1CHNouzB3WwpziUQikUgkEokTAwpzc88xYa7ZMRdJ9rAg17Ivzvk4CXFIhPcjkVZpibjmg62Ic0z+YeuJcW0dLMk22+zr+TbJOFn6bF4T1ujjfmqDjX+vTjEO59kT5dxH8/piIMS4YGu3xYKR7d6trLCt2be5m/709u6UK5+Rt7ImEolE4jsDCnM/f/kHx0OY27/hKMKcfZElQa6IcVG3HKAIa4qDaJOM6VEvcc985Q/b8NgQ4EJc41gTzEjetkrRrMbQPltftDXWeG/CXMwxUpjTU1l7wtx1ePdSmEskEolEIpFILDBGCHOWUCORBSXUiY0Ah4S9ZSTwrTAn8Q2lCXM+p9uLMEfhTH6WYIf4FjvdrN/rpZ9zmn/02e2tbJst5i39bR2+4yhF1beIRZhrOK4SfTy2FgNoH/LS21HvC3MYg/dslDA3RWHuKglzP4N3/wngGjCFuUQikUgcH7gwV3bM7d/gD38IYY63s3K3XN0xx7ha4nNQdu9THGRpuUER5mj3uBd9EsLk09TlZ/4x1sa5Twhn3lbsxXEZmxX/y/iYF23OcRBxHaXF6hiL0m2aD/73Ksw94j63DwaLrse79xgwhblEIpFIJBKJxAKiCHM39IS5EOPsIQ8u0NFOIlk2Itn1hD0S+laYC0aS37dVAc0WANWvCmqVJTFnkh5+aFvyPt/fvkWPduMDWxHmSh9pyfxooh/vgejCXCT7QQlztPl5cozdyhrC3Ba/lXV7d8rVz/jCYMmSnx8MljwJn8AYmMJcIpFIJI4PJMw974Nj+7hTbnjHnIly3BVnvzE3JMx5XVTd+zy2Kd4j1llZ7dFndfb3Y2LPD3GxCHseO83mO+fQbmNzjCvjOS/Gqn0UYa7YOL/P3RfmtnczB6w89xMb/uuKR95n92CwaAPevRTmEolEIpFIJBILDAlz/I25EOYsSQ8Rzh7y4Em028pOOdSHE/mpffBFIk0/9YMhwtnDHzjGk320QzRTgo1SbdpBPcXVfZlgW1KOJJsimmi2IrrFeLdzUUEhLs6tFetix1z4W/LvVPK+Db6e4PMYeB9aYc6EOMzLkgsAlW7zY7UPfzBhDu/vZ3Z0p171jC8OFi9+62Cw5IfwCUyAy/RZJBKJRCLx7WLqTAlza/dyp1wrzIFHFeZYRjz3uMk6bajTbmTbYl3kAPL3PhvHPuQCiqnRtlJ1xMq+OOc2+Rvb2Bz13lwYw/pEK8x5HDaRzmgxGnMfGiXMbS/C3EmPvO+ewWDRRrx7/O3XFOYSiUQikUgkEguIRpizp7LW5DvEOYlsTJxV3zxCmLPSBDAkwLQF0aeEv7WB5hs0sSyENVsUtDQ/JeecT6wJfyTtEvQ0BxcJ7GNpc8pO34bjTV0LAZJJfdBtuqU1bP6gB+vnWC+5SIBNx9V59oW5tSiLMHfNM/9osHjxLw0Gi5+MT2AKTGEukUgkEscHRZi7vuPvzM0T5iTOMT42whxilMVyj7keywo9nlq7xl+L8dbPsRb3TWSz3XAxr5f0R18lbCOEuegTMTbifLG1DGGOYxSf3RZxmXOP3DG3HTmPC3OPut/ewWDRLN49/vZrCnOJRCKRSCQSiQXEkDA3VYQ5sibSIcxZ4u2Mvki4mawPCXNMqHt+I/qrMOeJN+qRiFdhzvsxjyXjkbi7MMa+f6Yw19LmacjkHiXFOwl4SvZxvGFhrimLMMc2j+vC3FrU1+IcVu8twtyfDJYufftgsPhSfALrwBTmEolEInF80AhzY3s3dmMjb2WlKBfCnMWtiNOKuWErbYunFo8t1sU4o/WbzUU50EQ6m0dl8Tc/lcwv6Ctu05jSB1oMN9axTglwPFaUjM+sGzVPzHuUHXPn3Tb79ZMefb99g8GiTXj3UphLJBKJRCKRSCwwXJhbd/fObsKFOUuwmdAyyXbCHrenahfd/rpzLuxKpiXOYWy5DRV2tgsbH4wxW/VV8g173HIqkc3L2vaxnrzLX/P4eM0BP86heWhjnfajsYpt4yj5+3BqH4XycV+NRxkLiTh+7JijMLcGtiFh7h2DweJn4hNYDy7XZ5FIJBKJxLcLF+bW7N3gwtzGblwx2XfM+VNZe7eyOmPHXIhoRUijnT7o6wlziH2tb/GXKBbtiI/0rbGyln0ffcmlmGp9Edd53IivVnq/+6oEJw6xzn7ajcfaMXfebRu/ftKj779/MFiyGe/e48EU5hKJRCKRSCQSCwgKczsu27XuLgpzGz0pZwLOhHq+MNcjEuN54hwSZwlxze/DWSJ/LDZ+mKMV5cR9sIPRLuM84S6/N9cScxjrOM2r+Y16+IPXlcz7N+wS2+6FsZOuCHOcQ+eCOoljSZjDfPOEuWuf+aeDpUt/ZTBY/O/xCZwHpjCXSCQSieMDCXOXmzC3b1bCnMVOE+PsVlYX5mAPYa6Ia4hXvR1uqDOWiqw3wlzZ6RZt+jSsbRPluPveRDubo7KKaCLmpF3xFcdUjI86Y7nH73Y8Y7PGxUOawCLOSZSrwtzUoe3duhDmbp/9+kmPkTC3Be9eCnOJRCKRSCQSiQVGEeZ2uDDHRD0S7M36zTn+rpz9BswWtCnC2cMdKMgxeQ9hTuKcC3MS+CTOIWGOkom0BDaS9WqXvxYH3se2+5gfyLr6Ijm3hFu73UjZgvDTfBzLY1QbOY52K8yZ4NbWMTeT+0NI5MGwcSediXJep11z+LnwOJybx2h2zJVbWT+9vTv1mmf958HSpb82GCx7Lj6BC8AU5hKJRCJxfNDumOPDHw6EMGeUMKe46sKc7BbzWdeXa4hdVvcYirbZGOMZ78JGsc3Jftroz3GYT2NoQ7w0sQxt969zMG5a/LSYbrE44rvid1B9zAOMPEb469ZVsifMIYZzDhfmJg7OdROjhbkDg8GSrXj3ngCepPcxkUgkEolEIpFYELgwt163sm5UYq5EHUluT5gTQ5Tzp64iSQ4RzvpYZ9KOkkk/yKewcrdbS4lzFNlE2JhcNwzxzZJ7H9e0S4LubJP7wujn63FbSd5LX0Ml8L4gYJ1JvYQ5T+6dvZ1yQc0BP19YWLt5+APqIcxNfXp7t/LaZ/6XwdKlvzkYLHs+PoH7gCnMJRKJROL4YOzMx51JYW7fhm5s/ybEoFnEJsbXTYinjMvDwpzFTIv7VmfMDKGu7WuFOfki9pXfkSs+Ni5yg6DF4vA3SjQrtH6VFNh6dhDnEV+u6Rx1bmxH3AUxLp6ebsKcxenhHXOTh7YjZ5lzYW7j1096zH1TmEskEolEIpFInCDcmzDnjMS67JbzdivMWRvJNm1M+rUAIM2vJ84xmUZfX5hzu0raR7MIa2HDcSNp79WH+0El7rD1knjRE3hP4pXIDyf3oO2SM3Gu+Ho9bu+x9mb4YEGE1zOmsu6YW3nNM780WLr0tweDZT+KT+B+YN42k0gkEonjgyLM8fflhoW5uJUVsZltxXfWI4Y3ddnbNvsZ4ymqeZsxEmWP9Oe4mJt+PVZhbr44ByLGKvYWW51XrwNz1rzBhbkoOVaxm2M8dnMu7n7vCXNz3cyBOZz/XHfe7o3/7aRH5Y65RCKRSCQSicSJQk+Ym0WSbIluJNAmwnlyjaS4/V25krSr3xPuEOY8YY7bWM2G5DjEOe+fv3uuz1aEC0Gt9tvc8Q26se+vY9BG9hJ9owlr3AlXWcW6hrSrr46JOhcMbNucOKba3DHHBZEJc+2trCuvljD3u4PB0h/HJ3B/MIW5RCKRSBwfTA0Lcxu7ce2Oc2EOpcVo2zFXhC6UlgNYHKsCWLRBxnES9RDZal/4O9vxoOUBVi+MsfPowh1jLGIwbfbFncX6OkcT91nSN8ZEjCb9ZylClJsCuWPOhLkN/+2kR+WOuUQikUgkEonEiQKFuZ2X7VofD384aMm0iW3cLcfE29tIfEkT5mhj2+2o925lhV0JuQtzpIQ51iXOWV8rzNVEu7KKbEzIUeIY830sURfVbkk7So5jcl4ENKP1jRbmmNgfVZgj0bZ54Nur0x/HLbeyUpjb3K3eu7Gb/n0Kc89IYS6RSCQS3xlQmHuL38q6b0iYkyjnglyxsW5kLFdsVSwLe7TZhzjXtEcLc8N1Y+QQra2MHUkT5nROJObrC3MoYY8YLypm45yGOHVom8Q4CXOH57opsAhzt6cwl0gkEolEIpE4kaAwd0PzVNYizCGB9t+X6++YQ5sMe2k3whxL9M1LytVvjKS97pbr+5JFWCMxX7CfuDMZdx9nzK15NIY+lqwX0c0XEyHgjeO8+8KclzFWdo4JtnaUnvzrnDgnj+vC3NphYe6qZ37ZhbkX4hNIYS6RSCQSxw8hzMVTWUOYQyxqhbkJlOK+Ks6V2K34Rh+UiqVO1j3GRVsxHfXIE2x8MPoiN2Db+hRD3b/GcLOxr423cSydE+n+cQzN5THY4rfHbN2+2gpz23vCHG9tPffjG76ewlwikUgkEolE4sRBwtzzdCvr5J7ZTk9fRWIrzhPmnG7T7ayg9fuY4R1zqIvo4zft1uc20nfPjRTmZHO7xnupYzHJR9Ld+A8vBIKW4INK0r30hD8WBnxKq+2KY1LvYyS4kbR7ki8f9xNpdx/112O2wtwavA4Jc5/e3p2cO+YSiUQi8Z0Cn8r6lss/uHbfhm687JhjbHRhTmXE1+CwzeNmQ4lqHm/rl2D9/uLjttIHW+kfxTJva49Y3dqMrYCneE2/NkZ7TFZcJtsdc6DdygpbCHOPlDC3Be/e48EU5hKJRCKRSCQSCwgKc298nm5lpTA3fbAmz2QR5UgkwcPCnKh+T7jjNpdI2ElPrCWmSZwLH3Deba1mpyjXE+s4tzPmGBbnLDl34nVEPcQzJudRRsIfSX15mAPHyMf9fEwk9+Hfo3yM7bx8SMS8HXNVmPudwWDpj+ETyIc/JBKJROL4oQhztlOOD38YFubKl2ce302Ya8U5xEDEr8rwQ52xGj4mzlUfywMYA8NW6z1hzuNpjJEP5wEZO2NMO97sTZu+8sd5RL/PazGb3Io4bBz58If9diurCXP3298IcxmTE4lEIpFIJBILCAlzl+86964buqk9m7pySyqSWqtX0Y2sYpyxiHUqkRi76BbtQvlWMa3Yy644T/BL2XK+CKfkHGXM2bPzWJ6gK0lHPxmCmYl20ceEHok75glhLvxr26gddF7nIodingQ9XwToeBgTx6HNhLnNJsztm+2mPzOXwlwikUgkvnMYEubG9VRWxj0X5hCPWmFOHBLlIm6X2Mox8uvbVNIP8c5+D85iYJs3DDPidI+ar8bQfjx1tjb4DwtzpPpKbOcXZI0wJ44S5q7/+kmPut++wWDJZrx7KcwlEolEIpFIJBYYIczdvbOb4o65Q0iclWBv0e45iW/OaSS8fWHOkvsQ5kpijbK2rW7sC2zy+RaEuSK+4ZiytSWoRD/8vT1KmGuFNiX3eH0leR8pzNX+YfLW13G8/mMJc7ZjbqsJc2ibMLe9O/kqCXO/PRgsewE+gfuCuQhIJBKJxPHBMYU5xmuUFNnIEn+HhDnELIur1i7xVXXEubC5XX2IgUWYgz1yh2GGf4/wN6GN49EeiqcRk4st/HF+w+KcfCO+MwfgXPcizK2QMLdoE969x4EZkxOJRCKRSCQSCwgX5uI35nrCHJLaIsx5wj2NhDyEuSLK+S2nIb4pKUd/JO1mxxgX2OTjfUafB8m1lfDn+OLLOdn2RBs+TMQjmdeOuqCObedgO+M8SXdfS9jZZ4xEXg9/QH+lt9XnxyUj2acd5xDCnI7FOf3Y4TsG21qUEub2z3ZTFObecOl/GSxd+huDwbLn4xO4EFyuzyKRSCQSiW8X97JjLgS68vAH1EOUUxxGvLIYbrE9fmKC8TRireryibhX47bo/kF9WVbadiwT1IKwa0zEZfhHjPYYa7HcWOfn63BqHH051vpLvB8S5vQbc/vtN+bWf/y6/7riUffbm8JcIpFIJBKJROLEYFiYQzIbwhwT9CLOIbEtdSTERhfS9rF0GxJmI5JiJs7Fz/vcz+zVJxYFtjCgr/kr8WY/yyDaJQEv45mox7E5zsfKD4k8x/i4mrBbm9QTWdE/T5gTa8LfH1NtRxPmeAvNGMq1h7Z2q/m7Pp/d3p185aX/ebB06a8NBsuei0/gAjCFuUQikUgcH0zxqayNMHdwVjEpdsnFU1lHsQhzxYZxDSOGV9a4Nxx/SyxHDCzCHOfUefB8jBIGS9vGW5ytsVRxtplb5+NjzLfShDnzO6Yw5w9/oDB3kglzs3j3HgumMJdIJBKJRCKRWEAUYe4GE+aUlIMHrSwPg/Cyt1OOpCgHxi2uddeaJ877aHPSTpvT6kiYVcYigMelr3E4ATdfG2NJeIxnafOXcc4Q5VSnLRJ1JvtuMyEuxLimrr7qFwuEYVGuJP++cJDfIfjBNoZy7PC2bg1sU5/b3p30hkv/dLB06a8MBot/BJ/AuWAKc4lEIpE4PpgnzG1CvGKMMlGMwlwR3iiSMT4y9qK02O1+qpMY6z6K6b6DrsRwxT0b2+YARtRLfIQ/j6eycqQwh7lNmPN46uP1OkTWfQz6Ikarj3HZ++YJcwebW1kP2K2s6z9+/X896ZH33TMYLNqIdy+FuUQikUgkEonEAoPC3A3P27X+nhu6yd3cMcdk3BJy7ZBDgms75izJnlbyHoQNCfq0C3NqMyn3uiXyLZEcs2zEuprwV/I4wzaR5xGLgLb0PrZtQRB265sIX6ctUPpt+ohtnWMxb0+YcxZhDlTy39Q1JxcKh7Z241gIjB/e2o0dnuvW8DbhP9jRnXTV0/9ksHTpOwaDxc/EJ7AeTGEukUgkEscHw7eyztsxZzQBzmKubKqH3duIf5WwDQlzNSZHO+Ig6rQr1lZbfElm7Pf16fNqDrRDYEPbXgtLtzX2lpxHsVg+FOXqjrkhYe4vVzziPrsHg0Ub8O49BkxhLpFIJBKJRCKxgKAwt/N5u8795I0S5mKnHBPyGZLCHJJa+405CnNMvNEvDtfBfSHMMblHYtxQCTv6KcyVNpNn9Zs/jxH+wwzRzRYCtKGMfrajH6yinNlblm/QybBRiMMc84Q5UvbmdlbU+fty0VbyP1zH3BOHt3bjh+e68Tu2gdu7NYdwblWY++XBYPGl+ATWgcv0WSQSiUQi8e2CwtxbL//g2v3Nb8wpPrnoxpgdbdDirpWtMFfiK1ifoN7Y6N/E4cgDGAdtLsZZays2oi9iqtkYf2t/nx6fFavR9nhdYy3r3u++NV7XY9R47zvmKMohLk+B9VbW6/9yxcPvc9tgsOg6vHuPBlOYSyQSiUQikUgsICjM7bisJ8xFos6dchLmUCrJ1u44inPcIdfSknF9i+7CW4h1FMjKLjqKds66ow3E3MMJPzmhhYDVTWxDHaUtBsymuT0ptz5va2762xgxEvS27fV2h5zo463OHXL8cWm27YEP2kVXSJGO71scDzbMTWFu4g4Kc+T2bi0WBVN/eEN30lWX/vFg6dK3DwaLn45PYAZMYS6RSCQSxwexY27/bBXmGGMltjE2R3w2G2Oq2d3GErbyBZvH9PCxuMsYzLr5yp9zhZ/3R18R+lRnfK3x12h9IaiRJZ4XG2IqY6tit/kabY4Sk4f6TMAbJcyhfnCuW7frur9Y8fD7fGIwWHQt3r1HgSnMJRKJRCKRSCQWECOEud6trM5IwCXM6dZVE+RUxq2sTN4lvJmvknAl8EiKva52wyKgKZmPMSDmmGy/ocfxIzEXYYukvRDJelkMDI3RObgQF0l8T6TDeQwn9ZyDZe/hD07aZFcfzhG2Mhd9uHhwYW7iTnJHt5YLgc/f0K28WsLc21KYSyQSicRxhwtzY0PCXAhnJq4xVjJ2WbyvMdjibsTSiOVteTRhjtRcmtPnHckm9rZxE3OR8/1J8wthbjgmH530xdihW1nt4Q9+KyuFuUdc+PHBYNE1ePdSmEskEolEIpFILDB6wtzG5jfmXHxDYssyEnAT4tzGBJ2inAtzJs4hYR8S5iKxrkKc2SSWFb/N2iE3EYKcl7aIcOGLfiptXAhysROuJ8wFD6LtfZH8K7lXvWljThPlKLZ5Qs8y6iNYxDnUbS4k/TEvj0dhjqLcXeTObuyO7d3052/sVl79zBDmnoZPIIW5RCKRSBw/FGGuuZWVsZSxHaUe8qQYOkKYY/xV3ewWyyNWso14SrIe/dHntN1xTu+rt6Nayd+tLfEStvgJCZ6DxVRQx7a6xWzUOQdYYjGOEXWbH32YW5SdvhxjwtzEwSrMzeyfQ87DHXPX/vnyh11462Cw6Cq8e48EU5hLJBKJRCKRSCwgVi6/5Oy5y3ade48Jc3XHHBN3Cm5Ga5v4FjvlbIecJeaFFOa8Hkl16UOSXMQ0r4fIpuSaYpyP6Y1jm0k3Sn2b7qxztaSvLRp6Ni0CgmYvybxKb/sxzA5b0xZjQQBbuZ3VfaeQ9Jfz4QKh7Jjb3o3ftbNbeycWAV94Y7fymmelMJdIJBKJ7wzKraytMMeYaXG9fOHlVLxUneVoRhw+mjCnOuJesavP5i59hTamMI6j+Gk2nq/Zwu5x1VlicsPSJyJGwzaOccPC3MQh49R++405CXMPv8+uwWDRlXj3HgGmMJdIJBKJRCKRWEBQmNtWhbn2qawS5lAvwhxsIdTNF+bQZtkIdW1CLyI5rmKatUOYO9qiYHiukqyDda755LlGgh8sCwIm7+5Xk/qmHsehzetK+tkX38KHXX7mO6Ufkmbyj7kPgfEbc3eCIcx9sRXmluetrIlEIpE4vrhXYW5Tf2dcxG/GMsa0QovLJQ6j33IDL71udsZXo8VE9mHe8PO+tp/Ul3I+h9H6+zaQsRdjSdXpM0yP7/cuzKHkb76GMHfrtV9b/rALPzoYLLoC797DwRTmEolEIpFIJBILiGMJcxLlfHec2ky0h3fMsQSZwDPR5643F+fausYxcde37bDThvmGFwC2CDA/inb2zTzr7sc5w4/z9eZgQk+iHqRdNptH/qWPCXsk9TEWpF19ntA3DP9I9tv+EOYkzjHxP7zNnsraCnNlx9zyt7swl09lTSQSicTxQwhz+0KY22RCV4nTJsxNsFTsBRHX+vEUlC/pbfqgVAx3f8VqxUAfX2Ki98vmPIg4SWHN24qnOHb/4UneX+YFKbjJ3+Ou2tHnZL/qLJ3hL7/5wty0C3MzFOYefuFHBoNFr8O79zBwud7HRCKRSCQSiURiQSBh7jn1N+biG3UkuX1hjnYm4dbf2zEnsu0JPNtMqkvdaMKc1ZnkR1kXANVuffSn4IW5OJ/83J8+nE+JfLVZYt+wSe7n+YJcGLBPibsIG8aMFOaU6Bv5Tfxw/70Jc2MoZ774pu7ka5/5J4OlS38Zuf+l+ARSmEskEonE8cM8Yc52zJXffnPBjcIchTHZGQub2GgMf2tbXoB4xz7VGasZ/5q26hZvrd6nhLkQ0hpWH+/TMVjHXPJvYrILc7ZDzsZpHm+bMOf+LHXMVpjzhz+EMHfLNV9dfvGFNw8Gi16Dd+9iMIW5RCKRSCQSicQCgg9/mLts13mfvLGbaoU5JuEo7amsaLswR0GuPJWVohtopQtze5EYu13JO+uaqxHm1GfH6Sf/xhD7tBAQbZz5IsH2uhYIMafa4QN6HxN2E/DqOC1Iip8n9FpcmH9fmLN+Jf9K8JnwN8KcEn5j3KajRYS+kd9ahbm7TZjjrawpzCUSiUTiO4Z4+MPe2W58X334w7GEuXJLKdsRH1Wnn7cR9yweW6yMNvskxLGMesRd94vxosfMIqAxvqoPdfgFw1bGKS67H6m4XH2sz+bUMeRDQQ51xGTF5Z4wx6eybu9mbrn6q8svPv+mwWDRq/DuXQSmMJdIJBKJRCKRWEAcVZgDUZcAF20kynoiKxJ6s1GoI71fyT4SYwprTKpJ2CJZn1ZCX/uKT1BjbSEQttYnxhYbk3HOiURciwVPztknoc1tJtixHX2V1d9KJfXeNvGN5LzGupAw2gLDvokX+WRWr/NWmRDmJu7aYcLcF27sVl6n35h7+2CwOIW5RCKRSBxfxI65vRtdmNuImMa4areullgtm9H6aa+2So+PThPmyMZecodow0/x0eojWWJozIe46mWZv6FiM8tgOx6MdonRnK/YPS47uVMudsytu+Waryy/+PwPDQaLXol3L4W5RCKRSCQSicQCI4S5u2/opm6nMLcJibUxRLkQ4WInW7UPCXNgT5jzxD/ENBPmSEuyezvcwAksCMhqc1/Q5sZcmtPmi2/kbX6bOwS4VphTgs+2SvZVtgl/L/kH9dRVtmXneNSbpF8PeFDCH3RhDlTyH8Kcnsy6oxtDqd+Yu+6ZfzxYXp7KOg2mMJdIJBKJ44OyY86EubH9jTDncdRiKW3DhL0nztGv9WUd8c7jZdvfCnO2aw51xEa7FRV1Z4mrbfyMWN1jjMFxWWJei8/IFThHjOVuuDIX46/HacxR4jXJ2BzxGTRhbpsJcw89/4ODwaKX4917CJjCXCKRSCQSiURiASFh7jm7zpUwt2FImKvinN2qakIck2/ZkBjzVlcl4LIjcWayT1+NgW2YFMs8qZdoxjrHMdnGuFHCnOjjaa+JvwlvtgCwsibydoze/CRsottisaG2j7XEv8+YryT8SObj1ph2QVAWFEz8eSvrIRPmxu/Y0Y0dnuumP39Dd8q1z/yjwdKlvzgYLH4yPoEpMIW5RCKRSBwfjBDmxhlfXZgrX3Ip3lq7knaQca9Hjqc/6xYPFRMpmtGOeskDWHeWmKh6jDHW2Bl9rW8dw2PEcUyUIy0ec7we5sCSbcRdxl61NdZ84ouzecIc2us+cvWfLX/o+R8YDBa9DO/eg8EU5hKJRCKRSCQSC4hGmJuUMOeimmjCnEQ5tRthDn76/Tm2UZaEvCT1HOd2li0llpmopuQbPhLAcJy+MEffGFPb6uMxxRDnXJhTP+elT52fLEIb+kod52nJvtnNDwl96bd2WSCUW1mHhTlL+gvpI2GOu+Zs59wY2tN/eEN3ytWXfnGwdPFbB4MlP4RPYAJMYS6RSCQSxwftwx/2zfaEOcXYloyB7OvZLBb2yfFNn2JiY0O75AJRVwwO3/msohnb9DUxrfrYeBPYQJ5D29ZYxuOGw8IcyhKjXZAjuVNOu/hRUphbdtEF7x8MFr0U714Kc4lEIpFIJBKJBcbK5ZecPfecXevv2dlN6lbWEOHI+cIcWR8IMSzMIfndB6rOcW4PQc4fDEEWf5RGWxgUYU5+SJpJJs/yBXUsliHExfHMXgQ60fyrwOaLiCHbPGGOi4PmeO1ioU3y9S19I8zpqaxO2UKYUznXrcXiY/pzO7tTrrz0i4MlS94yGCz5AXwC42AKc4lEIpE4PijC3AYJc+NFmPMYu892zpWHP7idcbAnzEVMVFxEPzjctvHmq/jrMZm0GO1zcKzHUYuZtIGo6wuvefE6bC6weV+Zg3bOoz730RdnnL8lj4XzQiwm+eAHfqFWH/4Qwtx57xsMFr0E796DwBTmEolEIpFIJBILiNgxdw93zA0Jc0ja+aAHE+ZC/Gr6iy2IBFhCmrdDkHNO7kViLsHN+8FI2vvf2rtPI8oVeqIu9sb7fJ6wR39N8rl4IN2HpduDbcJv41nCB3UtELSAsNJ2ysW377Qh0W/EObZbYW7s4Fy3FvPNfGZHd8obJMy9uRHmluqzSCQSiUTi28W8hz/wqayMeYyBFOVcmIvYRzL2qu7xsRcnPYZKiKvtOt78FZ9ZH0Ebj3pDE+VI1EvcjfnNZnbWfS6W9C/jWO9z/u+/gozZEub8CzWw/MbczVd+edlDzn/vYLD4p/HuPRBMYS6RSCQSiUQisYCgMLftsl3n3XOjP/yhFd4ozLk4h4TYxLa2320ti5jGunMvfDlur7UlvA2NKwsD0X1aYQ7JeY8UzTSuUvMoie/3DS8eRK/XPpTeVxYD4cO2FgFM7klL6iW+kfRF37Awp1tZUY6jbwy2tXhNM5/e3q284ulfHCxalLeyJhKJROL4o3crayPMHdiEuDZfmIsYSFocdaLeE+K46439iGV1bPW3WO70OF+oHXM+Pub32GkCG+oed+sx2aa92mJsEeKausVn0uoWtxmPw27tEOcozDFWz9x85ZeXmzD3Yrx7KcwlEolEIpFIJBYYEuZG75iLBz+EMGdPYB3FGIPE3MU02ns75SjOlVtZfQHgQlwcTwm7+hvSNkQdxxN4cqL1Y9JOW9CPVfq8PxgLCyX33q9kHvP3hDol+DjmUWnfvI8W5raYMIf3YOZT27uVr3/aHw0WLfqFwWDJJfgEJsEU5hKJRCJxfBAPf2iEOYt5FOYs7hZhzsvSF2T8K3GShF8Ia4Xe9lgZds0Z5DjaWj9nia2Kt0YT4vqM+NzOUWwh0PXo81KIoyCHONz7Ms2/XJuKHXMfSWEukUgkEolEInEiEcKcHv5Qhbkiyjm/NWEO3GfCmX5UmXWKZkWYQykRjcm80304pifIBWkfooQ51ZFwo6Qwp4S9oRYCvjBo7X3WBUMk+UVo04KANl8UeLLfF+PmswpztjDgbrlxjB/Dua7ds7lb98m57pTXPe2PB4sWvQ2LgKfjE5gBU5hLJBKJxPFBEeZmEQM3deP7KcxFzLVYXYQ57aBjrKQw14pziIeMiREj6TsszHm/iWWIgW5XTPb4a+Ic+zFWfpVVkGOdx+AcoPpZGms/57J5im2eKEda/FXMljCH2DwkzLHNPIUPazJh7ty8lTWRSCQSiUQicYLAhz9se/au9aOEOdRnoo7EOjiDRJlkUhs2E+WqMCeyTnENdi0GvC0b/JnA13odG8m9JfhM7FHXHNYfNB8k1vSLhD7sGudJfPE1Krn3/mpngu4LACb0XBAwqSdZ93ZZDDT9UypNkCvCXPGnMLe5G8O5r929qVt/z7bu1Cue/qeDRYP/hEXAv8cncB6Yi4BEIpFIHB+4MDe+twpztrPc46jiKeMjH/5gYpyEOcVEo2IsYp1EMMZNxLEqzCG+IbbZg5+ijtLjaQhoLWMu0WOntb3EPPYE9LDTz/oj7tJmczXzhRiHOr8EI1txrhXmuDtuUg9/MM7s39ZNlx1z574nd8wlEolEIpFIJE4MtGPu2fNuZQ1hzkS5vjBHziD5NXEuhDkm+86mbTvfKLpVYa7XjnE4RhnH5BylKH+ve3/pK6R/sN+nb+rn2XyxoEWI27UYIJngm6BWFw9OJvre1+t3jhTmcE58Gt743i3d2ttnu3V3b+1OvvLpXxosGvw2FgEvwCdwX3CFPotEIpFIJL5dxI65vbN2K2u7Yw7xrghz6Athru2XTxNP5++YY1/tLyKZt0uMbVj9MK7Hvj0EuJatMGfEfFFnnJUwh+OESCdbpWK075IzUc52zfWEuQenMJdIJBKJRCKROFGIW1k/eeP8p7JKmKPw5uLcPtSRzFdxbiv8/ffk3KbfmHOhTd++s45S87F0oc3qtNd+Ju5VmCPNlw+BsD63zSPsntSzPSzeKaFnXcn8iMUCGN/62yKAiwPU+e06KaHN6kr0UTfSx4l67zfmNBd8Me/E3s3d+J7N3Rje33V3bOlO3fqcrw2WLf7wYLD01fgELgZP0meRSCQSicS3i1aY2+s75hjvGANZIuZWYa6JhYh/sQuuvwOOhJ/Gw+YxuUfG2Yi1oM0JxhiNt7kiPjJ2Th1Em4yYSZ8yV8zR2EgJcVHnPNamKKc6d8k5y5PTXYwLYS5uZZ06OKensqYwl0gkEolEIpE4cdCtrJftWt8IcxThJJCpPizMmQDXJwU6K5XQw2aCG5Jk1S25DluIc7E4KES//JGc2zyghD6rRz9Zkn75+5hSGnuJPsewBC3hH6b7y49j6vhCJPflW3i0xVggwCZhTkJejIEd5z6xd0s3vntzN377rG4NXvW2F359cMYpeweDZZvwCTyen4I+i0QikUgkvl1MrXrc2T9/+QfXDgtzTn4ZZjHXRblRcbFpl9tg3TbBuBZ+wZgDjNteg7LpWCz7vqIEtYixHB/+HO/t4od6+LR9TX+IcibM2bzzhbk55CyM2T1h7kV49x4ApjCXSCQSiUQikVhAhDB3z43d5O6N3TSSWolsoAlzIco1whyJ5Jqk2FaEuQMmoLXCHMu6GHAb+M8V5noLACbrw/5DbAW1WAz0kngfW1j86csxYDOHqEWD1Yswx7k88bcdc9Yf78PkXvjswfneTm7qpvZs7Cbf/+q/Hcys+QJy/3dgIXApPoVT7cNIJBKJROLbRAhze0yYG+v9xpztUI8vzOK20HlCm1PimMaCip2IccPCHOy2087aRZSL+cPGudy/RwlqmJNEW6Sv1+ljt6iGr88fvu4T8xRhrqnbbawhzM2ZMMcd7kWYW5/CXCKRSCQSiUTiBEG3sl6269x7buymds92M0hsQ5gTkRyTTLglxiFBHxbmJELJPwQpE6WKCMfk3OeQDWUrzA0/iEH9xYdtJNbe5gLB/N2H/jyuM2yFOK95whzno69ofm1/uY2WdbeJZWHABUS124KgndNfP1mEuU3dxG7wtuu78/Zv/MaSx3/P3w0GKz4+GCz9KXwKZ9iHkUgkEonEtwkX5uJWVgpz/K1TPXUV5RTKKsx5nIs2aHXYS2lknA9hrtgY74fiaW/O1hbzh3BHH/dVDFUc9XbLiL1eVxv14XkiRreMeePhDxMHh4Q5lCnMJRKJRCKRSCROLCjM7TBhbnL3bDeNBFa/G0cimZ7HfSNsSMxNxKMg5dRONyTFIhYCKItYBbbCXIhzNdG3Uj5Rx5wsTZgzRt8xhTlQCXuUYNy2qjHh49QYziOyDntJ8INVmKt+PHZleZ17wT14bbeDEuau68771PZu5Uue3A2WLfvCosHyK/ApnGUfRiKRSCQS3yYozLUPfzgQt7KGMGdflNluOZLxj/1B9KnfS2cR5sSwoR4Mu2Ki9c8rMf98UY5s5gm76PHWY27E4hhf5gqfEdTc2kFHYQ6MW1lTmEskEolEIpFIfFfAhTm7lXXTkDC3tZtBIk0WEW5fU3cfE+XIrS7IueBF8cwpoSrEOtZjYaC2LwbCX/ZmcSAiqfakvy/MYU4cf1iY4yIjFhpK2llXEk/CL+j+PG4R56JPCT3J22u4OGgWCFpE0DfGca5K3X6LOaf2busm98BvN44vznbr79rWrfmPL+sGZ5zyV4NFK2bxKaQwl0gkEonjg6lVjzvnLZd/cHwvb2FthTnjKGFOfajzJywUE2GrcZV0G+Mb2+q3tvo1R1NXe8imNuIl6qQEMxHzoNSDIGCPUixxl3M4VQ8fxHj1b8brtPq43+oabTJubT2aMLfiovXvTmEukUgkEolEInFi4MLcebyVdc+mbgaJKwU3PnHVGCJciG9GCXboNxGPybmTYlTUnUzErY+CFdu2OIgkXYl62CTI9fuMSKobmhA27FPZ3wHX2FAfJcyFf29M4xe/Kadv6pXc024+ZSxfG19/y714zXtR7sGCB5zei3Lvpu6+B+e6pQ+6z18PFi3bgE/hFPswhH/nTCQSiUTin49GmIsdcxbrLL4WYc7bEQdJE9tYmihXif7oQ/yNdowrsZOMPoyzmN3ajk4dm7F1mOpn6XXGYZRG+/KNIlzL0v8tCHMzN1/5pRUPXv+uwWAxf1oihblEIpFIJBKJxAJj5fJLVu+8vApzSF61aw4swlwQyXUR5tQOYQ5JOmwmvnkpevLe9qnfFwMU4dCO3W26nXXeTjkSCfUI9hYCQ4sLCWU8NqjSbdaOOYb8ncWuRYCxiHk9Yc77YhxeS4yN1zq9F+8PhTkS7Rke98Cm7v6fubE761XP/pvBsmXb8CmssQ8jkUgkEolvE+2trPFUVsUni7EmzDFmBmvsM2EuxLiG7D9ggl7smIt4144X2Sc2cbrYKkM8a21FmGObglrUMZfmc3uMHSnMuRinfs0BHrK5KMzFU1lnDsx104e2t8LcT+Ld+15wmd7HRCKRSCQSiURiQUBhbsdRhDkSia0Jc3wIRN0pp91ySM6NTNSN2jEWbfjIxiQ+bC5YxYLAbkvlYsHG9n9rjuNsbCt6KUlH2S4EirjndrEk8pbMm81Z/Lw/2jh+7Wvoc8XCgIyFRPjEosfKeK14f/bi/WO5f5sWAjNYEKy/a0d37oeu+5+DNWf8HD6F+9qHkUgkEonEt4kizG1shDnGJv7GHB/84A9/YAxT7Kq0mH0UIg+w+M5Y6DaP27HrvcTiofhIAY1t5gUaq77qH7ZCCnJkL/aCLrRFHB7H2HGf33bKgewPn5in7JgD5+2YuyqFuUQikUgkEonECYRuZf1WhDkQyW8V5mz3V4hzlqwjYaYY5fVic6rd/M4cRay4/bNHJunyRyItPyTXst87Lcln6WRiDmoR0JSFrS/pC4zSDh9fFLTjYwERvrYD0BYnVrfXG6Lc9EGKctu7dYd3dFOHt3YX3LH1b5c97SG/icXAU/FJ8HfmcjGQSCQSiW8PLsytpTDHW1mLMGexqdzKGnGMdVAxumn3ibEU5tDfF+YYxy1u9oQ59FWiHcIc8oo4bvW1vnmkb8sRwpzNzbodQwJdT5gjMfYowhyf0Dp0KyuFubyVNZFIJBKJRCKxgJAw97xd593zxm5qt//GXEMJcKAJc2gjyS42F+X4QIgiwrmw1t6uqp1j6qcoZyzJfiu4sd62Gyr59uRfJRh9Wkg4S5Lfjo1+9jFxVxk2jkGpY7fz0q8dy3PmwsBsGsPjxHyaw8eLbOPcQpTjTrmD27uZwzu6dXfs6Gbu2on6lv89/Rsv+v1FU2f/EhYET8Onsco+lIL8rblEIpFI/PMQwtw+/425owlz5AHEZH7phLqJbu6HtsXafryLWB7xk3Guxl0fJ38boziqsWgPM/rU73OUfhwDOQjrRVyTDX2oKw6TPIbqKPXQB3v4Q9zWWoW5bRg34qms/MLspiu+5A9/yN+YSyQSiUQikUicALgwd/49b+ymh3bMTSOhrcIc2sOkKIfEuhXm4lbVYAhURtstJzt8W5bdceSI/mB/oTDESNKV6Ietzlv7muS/Pe4QI/G3eenniwH2ycfmsnlh02tsKB8uekyY42/ZTB8G79zezdy9s5u8e+s3L/zsjn86Z8dlXx6cvGIOn8a0fSgFi8EU5xKJRCLxrWPqTLuVNYQ5f/iDRLd5whzsHucsZjXxc5jFpx83i/jW2ApjTBxjBDV2nh3H8HqNxbBJmKs2HVt1xGERdYlxFOYo0rGOcccQ5tbd/IZ4KmsKc4lEIpFIJBKJEwD9xtxlLsxtni/MkSHEBZGUW8lE34S2Wt9qpdurOIfEWPQ2+o1MvjkGpdvqrjWMZx/Y9lVhrCUT+0jMfTyS8lHCXH8RUPuHWZJ8+un8ttqxaZMPjwOfEAvLa3VqDBJ/7ZizW1lDmJu8c/uRmU/d0M18alt34R+8tVvxjIfvw6Lg3/unkkgkEonE/xtCmNtrO+UmDmxCLGJctphlwlyNg3b7qvUZ2Q56jHTfiIXRH/F4fkx2cn76x/gRnCfMUUhz8U39zlHCHH9fTr8xhzmKOMedc/AxYc5uXx0W5ngLK2PzpIS5K7+8/MHnvgcx+EV491KYSyQSiUQikUgsMCjMzbkwt3tz/1ZWJMHlyaxIcrU7DmQSH785ZyKcPRiCO+f4e2pFmBNNdKuETcJb3BID4hit+FZpY1qbCWuepKvd2NyuJF+s4yiUmajHBYAvAtp+ULfx9M7D5jVf2NEXiwE7fnMMCXHm13u9/EY+iAXA1OHt3eThuSOTd8x9c+aTO7uZT72xW//ZHd36j135l4sumv7tweLFP4JPZa19OAWLwNw5l0gkEol7x9SZjzvnLS/4IHfLVWGOMSlEOX/4A4l4WOqixzLFZY/RQcU9EHG+7lz3L6xk936P6W0sjHktBtPH463XK2GjkEY2dou7tKMtYY6vKxj9zka402/KcS79xhyFublu4tDcfGHuIee+dzBY/GK8ew8EU5hLJBKJRCKRSCwgeCvrtst2nf/JG7tp/caci25IgCXMIaHVjrmwI7EWm3orzNlvyHkSX9gm6GzbAqD1KQk4607ZG1ryT7HME/GeDT6NrbfAoI2LCBfmmMi3faIfvxXm4lhlXpYNeYx6vCBeI+bosxXm5rrpO7Z303ft6Gbu2dmtw/s+88kbupnPbP/GzK1v+B//7sLJD2BxQHGuBUW5FOYSiUQice9ohTnFLgpxiEWIWSLs8QVZxOKol1iGGFfis5clJrsoZxwtzBVxjnX4GN2P8ZPzD4lvRhtrNFuNu7AVYS7ImN62Qd8xZ20X5sQqzPVvZb36z5Y/5Nz3DQaLXoJ370FgCnOJRCKRSCQSiQWEhLnnaMccH/5AAc52xH0rwhzsao8S5qrwFnNFW8k9FwVaGLgN8ymBJ2UzfyX6WgSYCGaJui8gWHpSHgm9fDQOdhfi6hzWF3MZ3UdzcQ4fz74Yz3Y5hvcFZeM5cwHidXAKCwAuOkjdxhrCHBcDvJ31rh3d9D07u5l7bjBx7lM3duf9Mcr3veq/LVq39m34ZB4PjlocpECXSCQSiaNDt7K+4INj+6swx7g1hTjXE+YiBjNG0k6WWFhjNuNniHptTLbY2MRdMGJjtGMuiWTRJxvm9Bg5RbHNY6eR5wob+jSWIpuENrRhk9DG+eTjZDv8Wvp85he3so4Q5i469/2DwaKX4t17MJjCXCKRSCQSiURiAVGEuRu7qd2z/yxhbiaEOST3w8KcJfJHpyXtXCC4jYmzf6MepDjXtkX6DVHfmKOvtyDwhUNZPPRY/Xv2mNPbZc7mWLWvLkqCZXeAkn1QbatPi40wdyd4twlz3DG37lNv6tZ9+o3d9Kc2/93Yf3zh3YvWr/5ZfDpPAPkAiBb5QIhEIpFIHB3l4Q98IitjWP835vQACAlzjGcU3KJ08U2+iF+yG9v6vbEXW9v4qRiK+b1ehbkaR0ufl6REPdH7KbJFCZoA52XU2d+UcXsshTn93hzFuRDmbrnmK8svOu8Dg8Gil+PdewiYwlwikUgkEolEYgHhwty599zYTd6bMOeJeuljnWSSrxJjJcwhAWZCrqSfCX0k/KybraXsTJzn2TEXjlNsnmDT1wQwa4cAJxGNgpxEuWqXXzNHK+SZvc5l/THe6zqW08eUsbCZCOf0RQbr6ot+CnPxLX0R5nZ0M3ffoFtaZz55Y7fuU2/spj65/cjMZ7b848w7X/r5JRefS3Hu+8Cl+qwSiUQikbg3uDC3dh+fysqnk5owF7vNKcrFF2jzhbmwewxWH2MZSpLxso2Zqnu/fBEfMX8v/h6FI4U55iBiPUYIbXZbKkpQfSy9rf5Sb8d63Y8xgdgscY5xmeeOct0tV1GY+9BgsOiVePcuAlOYSyQSiUQikUgsIEKY+xSFuY11Z5zTnsqKOpJaCXJIZO2prLAhASclrJU6E3nSbErU9c28UXbMU/ppDz9QYtiQrQhcGodEuu0bRYwvQh3OU0l501+TfPZxvjpnjFNbfbRF28uW9GnJ5B8MMa7wEBnCHOi/Mzd9104X6Py2VhH2A9f/3eTvvOieUy9/8K8OTln2XHxSa+wDGwk+HCKRSCQSiRHC3KxiG4W58kWZhDnGWhfjvCyCHWhx2mJbfBkWcdi+IPP+iIejqDhr41jqN14bmxhxUnWLoRLd3K/Ecc5Be9MXwhyfwBrCXMRhMm57VRvjTdzrC3Mzt1z11eUXnX8TQumr8O6lMJdIJBKJRCKRWGBImHv2rvM+xVtZN0qImyfM0Sa7iXNM1kOUuzdhjgxRjjRb7Tc7StbBeEiDLRiqvSTwje1YtG/sUTIhH+6DrQhzw31cvJQ2knbZjLINj8E5ifSNOhYAPVFONFGOAl1PmLuTRPvO7d3MXXPdDMp1d/AWm9kjE4eu+6fJA2/4u9XXPfWDi8fPeIZ/YqOQt7YmEolEwlBuZTVBjjvm5glzinVBi2eKyxGTizBncY7xl/EvYrzq3u7FxGHKj+OtbufR2hkvScZJq5uI5v5Oa7MPJSh7Uz+2MOfj6NPbMWelhLmHnv/hwWDRq/HuPRRMYS6RSCQSiUQisYDwHXMmzM02wpyJcNwtZzvmwsZEvGHx9RJJ+rCfiWwhytVx1gf64kACmLeHd81pMYB5lWA3dvXBFgxbEdKadthU+lxh52JBYp6OP59lvI8zWlvfwGuh4sRCwHbIkXO24Cg75lBSmAOn7tgO7lB9hu2DW4+sO4xFAt7rGZzP9L5NeH82dmtvuuLr51z1I7950qPOvXyw+tTHDJYseRQ+uXXg0W5xzae4JhKJxL9VDAlzE9oxx1iJOAzygU0mzPEWV68j/tlvvlpdbY5h3HMhjizCnOaz9nC8lB9iofq9PtzPssTtKCWYsTSbbkONNuIi54p2EezKE1hho508BB8w2mUO+cy/lXXmlqu/uvziC24eDBa9Bu/exWAKc4lEIpFIJBKJBcTK5Zes3nbZrvM+GcJcCGv223JFmGtYkm0l1sbSj0S3FeaUhDcJPhNh/VZd2y9ykYAE21ntpCXtJYlv+uwcLOEuhE0imjNEtWDYOa+Jakfr77P049x7x2vpC4AqzLV0YY4lhTgJdNu76UPbuxnYRC4ScIzpvZu6aXweU7dt6sZvueof177viq+v/Y8vuufsrc963xk/9difXXz/NdxBd7Z9iPNAUS5vb00kEol/iwhhbu/GbpwPgDgwixiGOHwsYQ5xJ+J0FeZIi38qFbcR38Dqb/WIkVHviW/yAd1WSdsQGxGt1Fub2+uOuRDm2DYhbhRjjnnC3AHE3V3Xfs2Fudfi3UthLpFIJBKJRCKxwOCOubkQ5jb1hTkkrRLmkEBX4Y2JttUjQWfiXUQ50tttQk+Wfo3hPEiKWff+EOfapN36qr8x5rV2SdZBCW1kmcMWCy3r/Gg3/j36MSrtXJX8u08cUzvi4mEPeP/EQ/bN/3xhjjvjrJyGjU9qJSXIYbyJclg47Zntpm7f0E3fvlHl1CeuA6/ppndf243/5ku/ePbrfviXTnni/S4brFp+Pj7FMTDKXFAkEonEv2XEb8zt2dCN753txvdvRPwyUe5bFuZYpzgnkc7tjMOMu/6Ap4jPJf6ibfHT43jEbfmwzlhqVL20PbayjbgpKp7W2Go2+sAXdRPi6I9jsVSb9q3dONqk+Zg/d9DZLrpt3QTjM2Myb2U9MNetu/Xar614+AUfGQwWvQ7v3sPAjKOJRCKRSCQSiQUEhbkdRxPmTJQzYc6SeSNFNtpCZKN/I7x5O5L8Ps2/T9hBPTFOtCS79Ylk/ti0xN2SdxC2UeRCpNadPEahzRHU8Zu+mDtssXhgWW/F8XqIcodZUpAzu8aQGDdNxnuHBZAJc7Zjbpqi3O3Xd1O3XQde201//Lpu8gNv+Lu173jRZ87ZftlNZ77+qf/h7Cue8ktnvuIHf3HpQ9b/OD7RYz0kIsDddHmrayKRSPxrxJAwNyFhzuK47YILMc4Yddn5cw4uxAU5ti0ltHlbVKysMbKlYj7JeqG1+/4xHkQM1RdcrQ2UcEe6EDfM2Dk3jrEhyFVhzuIxRbkgY24R5h5x4S0pzCUSiUQikUgkTgx4K2sR5trfmLNbWE2Y26yE3gS5IWEOZPIsIY5UAm5Jd5vY9+hjgkyo+S3+sDDHRLz4YNy90+ZSgk7CVmlJvYlx34IwJ5uVcc4xt+ZvxvSFOdhUugCH5L/cuqp69Plr4/u1j4Jc5cxevOd7KM6BuynQbcRnU8lddDO3z0q4m/zExm9y18PE3tn/c9bO5753ycPXXeKfbCKRSCT+LSKEub2NMFduTSU32U447ZhDvdhB3yFXYh9jXSnpw3oTD9Gu8bQfJyW+hZ1lY7fYXm0l/iqGWnwsNlACW9Mf44pYx/r/kzC3LYW5RCKRSCQSicQJRm/HXF+YkziHRLwnyCFpr8KclZG8S5yj0ITk2cazpPhk/ZHs1wTdfCM5Dz9L2snoR/Ic48D4HbrRv0fX+kXJ8UzqLWn/VsbG+VmbyTvpNpVRB5n0D5c9YY4LgrmyMCh+mFuvl7sT9uL92kthDu+56iTFuU26rXXy9g1HdHsrOI2FFh8MIUq8w+u4fXO39p2v+/LpVzztbSt++Ht/9KTvv99lK37ge1689H5jPzY45aQfwSfNxcaZ+sxHI3+TLpFIJP41oCfM8YmsIcyZIMfYYQIcY6CJc6rLB7FNdo9PbJfyGMJcExPbuF7bHstJxMB4+mr5+YfwYdz0tvlibpECG+ZDWXfMoV7icxPj5Rv0mOvxdwKx2Gi71bVjble5lTV/Yy6RSCQSiUQicQJwL8JcEeGQkItI2EcJc2GL5PqowlzzTTxLJtVFiHP7vQlzpIlqtng4FuOb/Fg4mFhnY/XkV1+A9Ag/q3vCP1KYCxuohcIwuRiowlwR6LTo4GIAPnztFOVcmAuaKBd1vN98EIRKvv+o83MApw74TgfYp/Zs6cZvuf7vx37v9V9Z/asv+9ya333lF8be+6o/WvNrL75j9dZnfWjlE+9z9WDl0gf5pz4KeXtrIpFI/GtACHP7mt+Yk+jWF+bYNlHOhTnGE5aK1d72uNiWw8JcqXtMjPhNWxvHxaFYWYQ5UjvP5/tIiKM4x2NEW/NvRUy3kscyYa4V5Sp7wtxhI4U5/ubrzEevyaeyJhKJRCKRSCROIMqtrG/sCXN8EEErzMUtLkrqZbOE3Rg21plcDwtzJPr4g9EuRJnwRdIfNvdrhbpegu9zl3Hqd7b2wrBxZ5wJcibKealFh/X36vTHnFoA0IZj6zZY0drluEz2I/GXzdpkEeL8B6dtpxyJul4XSp6jFkB4bRTiYjG0187JbP6ei3wP0ObvBIlcVG3uxvdsOKIxe9De4+Mp1sF/+vANPNd/WPP2F992+gsef82S81Y/ZrBixbn45B8IXgiezj+DRCKRSPwrQSvM7TNhjl9EmTBnu60lwCnmWRxR7KEP44xikccbr1s8pM1jdcRB1a0dcbuNyYrton0pVb6Y0m41j4mMmbKHj/XHWBPmaENsRqmdc6zH8cH4uQrObTvstnTjaMdtraOFuTnMi/KWa766/KEXfHgwWPRqvHsPBVOYSyQSiUQikUgsIFyYO/9TwzvmmieyOk1cq+1qpzBn4hx9ZMfYkpAjYQ7RzYSoWBDQn8n2EOXLhLsyxDol6SjncZ7//DnKvDj+BM6jFex6xHxWj2Ox9MQfdtZt4cC6+5OxeAiGKNfaOJbEXGUB1AhpxaZFUm3roRA4dry39v7CTwsqLKw4fi8WW2DsftAxDmzHOezs1nx4w5+v/o1XHDrn5378XWdvefavn3PVD/3eikevv36wTALd0ZA76BKJROJfGuYJcxsQ8yjGzYKNMOexwuIx24wbiC1DLMKc0+I26oxDijMt3adnAxH/oh4xNYS4upO8X2/HlzGYx74MY9vOwfqaWNzS/XvCHMifl6BYOCFh7uqvLr/4/JsGg0Wvwrt3EZjCXCKRSCQSiURiAVGEuTfNu5VVD35gHWVQIlG02RfCEYW5to++SJaVtDOJFy1htySfC4GhpL74ul/DvjBHwt6y5xtzhc3HRBvHl8AGFjFuHtvj1JJ9LKcowvmxdesMKWGOhL/Tbl+NuYx2fjhPvncjRDkjF0mYi3V/XzjGPpvhW4Q5vtLee1DH54MntndTd+zoJu/a2U198oZu6nM3dtOf39GNv+8Vnz7lGQ98Of4KTrU/Bglxy8DFaiUSiUTiXx56vzG3sRvfR2GOopzRhLn6JU4h4rLiiMeWGmMszqifbcVHlIwzTaxUfI54xRgVZCxS3XwqG2GOopzEuOG6+bYx1GKqHb/S47DIuOykLzkkzJHT+zE/jr3ulmu+svyi8z80GCx6Bd69h4ApzCUSiUQikUgkFhCNMMenfPaFuc111xxtLEfQkvcQhrwN/5qEN/Xi74TNEvog2rBLNPOFgBYD3i9f7y8Mu45Vxx+N9uAHr8NfybvqTiwq+sIcGXUrezvm6M/S5yp+XIzIz+nHjNcewtuk38bao/uKavO9s9cnqm1+fO9twYTFFko+QXdi/+wRcvIwxt+JY9+FRcndWISQ95Cw3bHx7895y/NvOvXHHn3FqZc++MWnP/XBb1i27pzn46/iPvbHcUzkb9IlEonEdyNGCHOT3FFNYW5vcJNs8cWOxZBgxBbjsDBnsQglibp98WYxSTFLfR6rIg62sdAZ/YqVIcZpx1x/nNWtT2Np9zJ8FItFxuGIyWD060sy2yFnwhzmkzDHHXNXfWXZRed9AGGNX1SlMJdIJBKJRCKRWGCsXH7Jmh2X7bpghDBHgWeGRKJ9LGGOPLYw17AnKIW99ffEHuwLc2Tt65F2n39kf482X+yM4ziJad5vdti00LD+IrSVEsfSjjmztcKc+TR+XtrCBfYh6n3QosfZ9kfby/K+Rn0U0a/ddLFg4a20d2D8neBd4N3wuXtrN30Xz23TNyY+etVfj33gdV+e/NhVfzm97/q/WfvmH/3IysdeQHHu3nbNpTCXSCQS342YOvNxZ77leUWYmzimMOe3tTLOikMxpcSmexPm2Eap+Gf1acUhE9qKmDZMj1WtMBe/x2p2L0Wzt9TcmosxeATRb+JcX5ijUKdzxzFnPnzVny178LnvR1h7Kd69B4MpzCUSiUQikUgkFhCNMMdbWacPWrJdhDkk4qSJdLy1lbdKGk2EY5+VZCTwIaTVuUAkwaLbWyGtLAhgD5uRi4Hwb21mtz5P3pv5KOoZ0WZJ/6hrEWJ+kcyXtmxM1kn0oR0CnhYbfH+U5NNm7Al5Is4FZSXbRvbrdYiNXefGuZpzc/ZsMU7kXF7nPMVuixuRi4/DsJF34DjgxKHNRybB6Tu2ddOH7LcE9ZngdY3vvv5/nbX5mb9+0uMv+JGl91vzyCXrVz1+cNpJj8BfCn9353wwHxaRSCQS382YJ8xtLMKcyhDm2I7bWhFnFIcZYzzexJdtRZhjXKdd8cZimcW1iG0oGYMQSwoZj3qCG+h+PV/4jBLmVOpY9G/Glj634RwUy8H24Q8S5uRv85owh3ojzE19+Mo/W/aQc983GCx6Cd49Pr08hblEIpFIJBKJxALiGL8xZ2JaFeKGGcKcJepRelKPsTVxb8lE2hLtSP6N8wWpPq0/5g/hre/TkH1H85EAZvX4Vt0EryCTdU/mm7YtFJp29JWFiVGvUYzz9T73DQGN7fA1H/oG0eZTbFnXoqixkxrr72k7ny9sjHNWcgFyCH2F+FxFfI4HQQmwdsvy5O0bvrnmd172x2f93OW7zv6Z533onLdc/pHVm57xO6f/+CN/ael9xrib4P72h3NMcCdd7qZLJBKJEwEJc3Er66wLc4jRqGuXXJT+O3P2e3OM4REvWVq7xBkSMY8lv2CrbY9JajMGgRLM+mzFtuhXzCptE9iibSJd1PtUzKad/SHMyW622CVnt7T6OPlTmDOyrnPH+Jmbr/zysoec+97BYPGL8e7xgUgpzCUSiUQikUgkFhArl1+yasdlu877/VHCXCPCOaf3bzIbk3gKc0rmmagbJTQhCS6iHFnGs24+RWiiSCbxrApoJZEPn5aYL+rHFubIYwlz3EVXhbm+OMdkPWhtvQ4m9o2t7JSjzUudO9+DhkVM0/FH+dj70j9P2Pn+orTXwrb51zlsXLxfWqBwsSFSlJvrpg9t66bRNtqtRdwVacRnoh0QWLjt3Xhkcs+GI/xRcH7GenIfFnLTXLTcvb2b2H39X636+efffPIP3/8lg9NXnOt/PStA7qA7GeRDI3h7ayKRSCROJMpvzG2UMDfuwlwIctox56JcT5hrY7HiT5DxhbHG6vOFOYtbJRYhztS4hH6SNqcJarQ3MWvIz4Q50u3OEq99HrObOGe3rFpZdsrpGLSbIDfupY67j+PmTJh78LnvGQwWvwjv3gPAFOYSiUQikUgkEguIYwhzFG2KKKfk3RL3KtLBRzYT5Yow50m8knWVnuwr6Xf7MWgLACTSI/v6dt6qSvEtqGNQyPLj9YQ5H2uinBPHMlGOCTyTd9RFJva0N32lHvYYC8YxyrH6lN9Qf3md7sO52baFDum+Eua87tT7q7FGW/DYYkO36JCxY459MbcYx47Pcrab2LPhyMTtG76p3yDavxGfJf4W+Pnvs0UYb/8Zu+nKPzv7Z1/woVNf+QM7T33uQ684+Yfuv2nRzFmvHyxe/DT8JYVYdzTkLrpEIpFYCEytetzZP3/5B8caYW6i7JTz35hjTFec7AtzEdPVZqxxli/sEA8kzEUsaWJVxBh+oVPjTdAEMotVZisxz6l45v0SznyeNtYVysfHMe5pfmMIb+ZT6xTqxsEqzHGubd26m6/88oqLzn93CnOJRCKRSCQSiRODEOY+9cbyG3MmyhmrMNfQbdHH3XPaRdck8VVsYsJOccyEsiIuyT6CSJp53JF9vhgYHtsX5mDzRUVh+HIs/UXaQcxl4hzn9bldnDu6MBd1Y9wW22PTL59hP7T1OmM+LW7w+hoW30aUs/fWxzi1qOEig9SPWoPRJsNHbBZYFOZcfNVOueB+inNYrB3wEsec2LP5yMTHZ/9x7S3X/+3am6/5q/FPXPvXk4c3/u+xd77oi6e/6LG/vHjm7Gfhr2m1/VGNRApziUQisRCgMPdzl39Qvy8XwpzENxflijDnVJ/Hd8QbE+aMtnPbaHGDwluNKe2XSGGLHdrRlh+FsxDaKJiFXeLZKLrv0eh+NofNbQLcKGHOfOPW1hTmEolEIpFIJBLfXfDfmDvvnht7D38QKdiIqEdyjrLujguGuONJPW0YH6JUCHk1wUc9+ilC0SfqrSAFv5rwk9Y3T7iCnx7yoLIV6YwmxLk4xuOoNFGO40Mwq0k72kzk1fZxOF65bRW0BYPPxeMO0/v4Y9nDfTHO6HNqbrNx7nitIl+H160PpS80bPeBCXDTh+YkzLEeCxQ7T5tL43AcCa9q2+dA2mcanzfI94e3t+6dPVJvW+a5YK69W7vxvVu6cb5Pd2/t1rz3FV8469qnvuPkpz7w5cvuv/a5g1NWPBF/WY8EucBZxT+zRCKRSCwQxs969Nk/+/wP9IQ5/sYcxbi9/AIGsZ51iXOwk7zO89rfu95bDC2xm7EC8UQ74hh7GI+aWByxi6QPBbqIT4pDvbiFOuZr45751DHtfGKvj3NaOW9nXTlOjYUi/CZIb4cwN3PTVV9yYe6n8O6lMJdIJBKJRCKRWGDEjjkKc7cjWWeC7Al42U3FNhLz6SDatBeRzRN6azs13pPusEVyD4b9uBLn0QpzIc4VYS7oopwJbnV8EeGQ2JNM8IutjDW7LRQwjozxeA+q33BfY+vRjmELCrPFIiTGFqov2JwHFxikbl8dYvhg8SRRrjA+R/s84jONz1s7I6JNAW//JpzDrFHvHwgbf1R87Uev/Nux97zyj9f+xk/fPf6rLzq8Zu45N531/Ef+6rJz11yDvzAKdIlEIpH4zqG/G3ny7Eec9abnvW98D38/tHn4Qwhz/J05xEa1WeI6H2SMN2Gubxeb+DOfrZ2xJ37b1GKRiWOIWy7K6QnnGFPjmo8tvk38avpqbIOP6Havy4a2xntZWMYY9butKNfddPWXVjz4vHe5MPe9YApziUQikUgkEokFhAtz599zYzctYa5JwF2ssaQcdQlzTNxNsJHIhrq+UWdd/vymnKyJNkUl82UJG1jEJtj0rbzXJaZFX4+t/Wg+Nl6ET5Q2b/h4ny8INAZlnyZ6WWl+IbhZ4m+MxUQ5tuZu/cxe6HO17O/EM1tZhIwYq8UMd6rJ399fX4BQmIsFiBZGXPyUuWw++4xAzBXCXBHg2O/24qc6P18Kc84DxvF9G46wjFuh+LfAHZczd2Mxtuf6/9/4W1+4/9SnPvi6wVmnUZxbCZ4GrgX5wIilYCKRSCSOM5bMnP2ks3/u+TeN7bEdcybMxXXaGUIdrtshwpU4EDEA+YDlBB4HvG473YOMN1ZXXFKfxyDaPT5ZbPK41YtjqIs+R+Pbzq0+tG289UlcU9vKiH1V/PO56Bd9B+pPPfA3VKf281bWq7+04iIJcz+Jt4/CHB9mlEgkEolEIpFILBDm7ZizBL0m6bXdijVt8t5SPwpdRCCUTJZRSlySOMbk2tugCVkuojmjz8a6X9M/7NdSothQW5Q/xrGORQDJBUFbr+N4jrDHWIwrfeX1GG2xEH0NG59C+TaUnfM3Ni5YVPc5UGfbbOwfRVuItPVYlNSx/hmxDLI9TPcPAXX+52x/C3yKKxZ7R0wkxeKOfzdcxLEfC8CJA5u6tR97w9fPedtP7D7z6qe+47TLH7n5pEsecMOiyXNmsfh5Mv7yKM4dC/mbdIlEIvGtI66Xy5Y9ePo5Z/3Cj946vtuEucl9G3VdNkHOfkuUwpyJdbhmO+2Lt7jO2zVfX+J4zLCd0rR7jIh44Wzj3by+hhafmrbm9raLZhoX9eJfY53N2faF3caU+diWzZ5UHsLcJJ9avm8bYtZct57CnO2YS2EukUgkEolEInECUIS5G4aEOU/KVa+2owpzkbxHwi671Yuo1VKJuIle/8/CHOn98mnqrU2kKMidc6zzuOX4TRt1se3z86t9TPRZj3PjQiD6vHT7PJaFgrPxLbfINrvhYs5Y2LQLEJXR1iLD2la3hUlZ0Gg++4wqaUeJY/Q/R/Mn29te7dYm0v8OVIIqbQfdlHbPmTg3vmfDkbHbr/vG2o9f949jt2/4+4l9m/5u6ve3H5l436u+esplD3vz4LSTH+Z/gYlEIpH49rEI5M7kC1c+5UGvP/uXX3hocjeuzXvBnjBH8c1K66Pdr+VguZW1iQtVmGMb8SHosaLGLKfbW4aP2hGviq3GvRK7ZI+SfS62lbFse4yLuUrdfZ2cZ5pCHEU5kAId2zP7YUe7EebyVtZEIpFIJBKJxAmAC3PnSpjbOCTMIUmPNssQcpDo2rfoSHiRhMc36lWYMz8m6GxXAaohfMlW/CqC27AvfdgXdaeEOR9TbJGMN/XS56y+KJ094Sz8dH71uKQtEFxICzvqsXgI0csWKPU4vXm9tPfG7eyXX9Rr2277wdxlweELF7ad6vdxcS4S6Nwm6rgsrW2fJz8/1FG24/V5so/2IsrRH38HKO0Hwo3aMafPz3fRHeDv0NmuOb3WfdvQhwXRHTu76U/f0I2/85WfOv2lT9h+0qPWP2fZ+aueMzh1GZ/o+jjwvuCp/LM8CnIXXSKRSIzGYnB6sHz5U8589Q+99Zxff9EfTO0ZFub4RFb/jTkKc/tMmNPPUTghGoWMAAD/9ElEQVTjixjZFKfAiBceM3j7p9mMjCmlHXUfW+OOjycVy5wev6JdhTkvW7Y2xkLVh8fXerHBd5Qwt24/bdu79R+5hg9/CGEuH/6QSCQSiUQikVhghDB3985ushXmJMTZ7ie1Pcm2JN1FOCW8FHDMJqJuiXgk46S3lezTp9oo3MSONNXp4333Rvq2olzsOuv54Jyiv7UbMZb9YIhVlaPHahECWh/qtLutLEzYdn+N5fyNLdiKYZXm2/6+XdzeKt9YnPiiJMiFivW1c8W5BOPYtV4XUUYer36WtLE02m2r+FtoFnHxu3P1bwbtA7M4f/AgS34+JOfe1o3v29aNffTav1n7rld/fs3vvOTOte96yWdX33jZvlOf/pC3Lx4/nYuiC+wPcyRSmEskEonRWAI+cPH0OT+96k3Pfe/Zv/fSL03xNla/lbX8xpxEORfmUPZFORfm2us9221cUFzhTjOLN5URmxg3Kmvc6c/Ri1MU00JQo9BGG+fy2NaKcaT5ovTj9cayHm2nzeW3spKH5rrpQ9u7mQMU6LZ362+55ksrHnphPpU1kUgkEolEInGCQGFu23N2rb/LhDmJKxJZkDwrSfd2JOieXLfCnNF8+gm4J8WRnPOHlvVjy94GTfjycVgYxC2n9nt0Yavzxpi2L9o2V99HD1fAuZj4ZgsNjUWp40afn2uIe+aP0sfbGJ67+WpOzV/H1veDr88oH/l7nf6FdYwEOPlWfy0u/NZW+WnRYQuUVpCrixbzs/PpfwZ2PvW4UdcYnodzvjAHW/t5NJ9VvJ/xd2ILOrZNrNOuOZQixo/t4Y46jt3aje/d0o2jnLxrezf5qR3d+M3Xfu2czc9+98lPuM/LBuec8kT8ZU6BZ4Or+VcK8jatRCKRSIwGH6jzQyc94f5vHXvfK/9g9c1v+J8U5kTfKWftumOuffgDr9/BiPXlGq94AKK0hzsx5qB0hiA3zVtDuZsudtQxfgRjDsYV1CP+KT55bJPYhrrFMLOR0xTSKKxxt5uOTV/6sO5+MY9KHKfMwz76hTCHmHNouwtzKCXMXftlF+ZehPcwhblEIpFIJBKJxAIjhLk7d5gwJ2GFZAIdgkvYvK3kPJJ0T+S9Hcl71CX+BEPgKXRhjSJPa8Mxwl76fa6ewEW7+qvgFX3RFnEuZOnjGB4DlF2s845imSuohD/IdiVfe7/fiXFHna8sXoKwUZTToqMyBLnYKVAWNg11/GZu+zzMdxT7Ox/8s2qoefAe6/OLtuiLNtQr7e8k+oofxo7vjYdFwOZzTezdBm7u1u7Z+H9Xv/d1f7LqbT926znbnvmu05//sJ9ZctHkVYPTVrwAf6X3sz/WkaBgl6JdIpH4t45lg2XLLl99/TM/Nr7vmn8Y3339EQpy9htyFOcoxrkgF2yEubh2G9trOtiLC4wZ/RgiUmhr6fYSLxiXnO04xS3GtyKgoV7ItnGaVL8z4lb4oV3jII8DO23h3whzUxLmdnTrDu6QQLf+oynMJRKJRCKRSCROJNodc7e1wpyRt7ZWYY4inNuQ+BaBjvZSZzJMejIOlvkoxoS4U0Q1EAsEa5s9RLNhYU6iFuYNYUt29Xtfwzh2Fbuafo7xY8RCoTevc7jeY1kAxCKgz7LgaOljR845yr9doByyttEXGu7H97wIe3z/Vdq89y7KcceDzUVqTEu8tz36/D2fwvg74WLPqM9VO+i81K2ubLO036JjfWzvhiPje67/5uSBDd26e/D3dce1/2v8va/4wjlXP/VdKx687ifwl7re/mATiUQiMQLLF609++Uz733NH04c3IDrL3fA2/U1rsl86MM04i1p4hz6Q5wr129cy3Wt9zrJOFKu86THkBJvar2wiSmFJUbC3pIxrie0BT3WOSXOqV79LU5af52zHqOMP9Teyrq9mz5MYc4EuvUfve7Lyx96/ntSmEskEolEIpFInBiMuJWVIool4i7AMRFnm/XCJmkHi5+ScvONZLwm+RTSYAOrqAb2hLkqXLWsyb3tbGM95iiCHkm7+5UknXUfZzRRTsIcfZtzbectZYxRu/qVhQfHgzqu13uLi55f9ScpCFY7/WPRYWNDiJs8REY/6P46jx7hQ/q89rnZfLGo6dP6CvVeYaxor7+8Xp9bi7By3uELypdjfLHHz9UXfq04p78FF+jG917/zYkDG1HH397BWbzWWSweN4rr7tzSjd30+i+dc8Oz33Xajz12w7KLpp+7eOKs5w6WLXs2/nIfCPI3le4N+Zt0iUTiXyPm/ebmKY+438vHb33DV3htnTmIa+xBXHP9Zyb0pRqu0fbUVdhcnON1O3bN2fWb/bzeW0w38tpPG+eyvkrEBP1MRbUpHkT88bhR4wz6GHfE2NFmdTJiU7VxTKV9mcQ646D12y5ylCNY5vXbYXUr62ET5tYf3NHNcOcchbmLz3/vYLD4p/E2MrakMJdIJBKJRCKRWECsXHLJqm3P3rXurp3dxLAwB2oXXMMqzJEu3IHqL4l6ZZvkV2HORK6gBB3R/JnUVxHNBS+R4hqSbW/PH492z9/GFP9C+hmLH86/1GX3uRtfUcepvva+cAHgx+U8tGlR4eTiwO3lOG1bbBYR8kfbb2WttAVG6RfjvW1px+vTxvB8WmGu2IMc27xnhUNzxfmK8rH3v5QjCT+U7e4Me69njVhMTmExyd0d0+JsN/bRq/9+9U2v/8tVH3rNn57zzpd88pzf+qlPn7HtGR9fetHMK/DXy9+guzekMJdIJP41gte23m38pz79oS9a9f5X/Gc+yXvmEK6lhxGfUVecxnXaxDlcw/06PElxbviarDZ85E8iZigu0GbX8XnxpvQbI2awtHjUsIlrIZpZrGxFubCZn2Jjw7LTGzyWKEfGMSjMadec75abuWNnt/7Qzm4d6ufuuu7Pll98wfvwdr4Eb+ODwBTmEolEIpFIJBILiFEPfxCRUDOZR1Ks35PxJLkk+Ur0mcizZPIdST2TcaszuW+TeBNtaDOBq/42HOygRDPZTIDTAxCiBPsPYrB5NJePN5vb3cfm97Lxt4cQWJ/GUpAMlnGs+3kFOb4c08f7efYXA/b+lYVFtHn+LYvdFhxlQaJ5zGeizOE+7NNxjXXhw3q0Y14n7ZrfFjTzRDmOKfMY4/VxvL0m2r0u+jm4X4iXavt7a+djfwvD1N8G3su47dVufZ3txvduOEKhjrvqRNgm9sx2Y6iPfWprN/6Fue6MX3r+TSd9330vH5x00tRgxWAd/pK/F5wBzwDzN+cSicS/LZyz/D6nv+YHt6++6XVfW3d4czcDTlOYO7S1m8G1ewbX99gBV8Q40a/JvE6L0Tbf6fLQJvOxGBtt0nIEfUHHtmIH6LEh4ol2uXkMq2JbFeCmD1E8MxEt4qA9DKIKb70Y5X7R14vBmhP9jZ92zLkwN9UIczMmzH3lpIvP/wBCx8vxTj4ETGEukUgkEolEIrGAkDB32a71d+/spnZvlPAmUYmJcpAJdyGT4hDjWlpSH8JYJPzWDlp/EbdIjA0RTKIcGTYl2aiDUfaFOaON97qOYQKR5uEYlqT3VXIMSvkO0/1Z9+OX4x6NsSjw8670hQLKHrV4wHvXvOe2oDD257J6jB1li8+C51I/qyFiXvtcfdEyxP6DIIxxHKO/Jr1eO8bwrsT4XHQeQbRbVp9YDPLvhbe7VmGOt7fqd+j8NljOPY5yHO/Vmg9f8eVz3vZjHzlz89Pfftbs03/9zJ9+/C8vf+D4ywdLl16Mv+pjCXPzbv9KJBKJf0EYvn6dPDh56UVnveDRV5z5iz/2sYlPXPs/1h3e0s3cgev54a3djAtztvOtid3N9bjdRdfae9S4pj/mCUbMQb0Ic2EjPa6VuuwumkVMatn0xxwW22JsZY1NXscY7TBnv89nwhwf/LCjEeaslDD3sAs+hNDBndgpzCUSiUQikUgkFhghzHHH3O5ZJK5IkP2ba9WZMCu5RkLOJFdshTnYlYibiBWCS7WzTbvRxLcQv9h2qs4+E3cKcawQxiZwDn1hbiv8SdbRH3P7fPRphbn+eRiLTedDPxsbc5kdpZ+DFgacayRjgRCLhKDZbIFRFxm2OMF8jTinnQRBHxfHbefn+NZH88GnPbd54hz8YwEkDotwrQ3HCnscgyyfRxyHx3DfsNn7hjreQy7gQoArY4of+/je8m/FaCIcqAdDzOJY5EZ8Bihp528m4e/SiPnvxMLqMzd205/Z8Y2xt/3kx894+kWvHZx5CnfPnQUu1t94inGJROJfL84/9SkPev3Y77xs//jHr/7r6f0b/8+6O7YWYc5iOMjrNK675PAXa2rjWkzqmryX7Ybojxjf2koZ9HjAWMV6xMCeKEcyFkY8VDxqbKWv9fNxOgbt1rbYZ8ctsRZ9vV12Psc8Ye7wzm7dwZ26rfW8Xdd+9aSLL7hpMFj0KryfF4EpzCUSiUQikUgkFhBlx9wN3eSeIWEOCW4R5pAQx7ftwV4iz3ok7eyT0MJkHoTNfKLu9kLrK0JZsTPBrz4mmsEvqDHoi1J2m6MIbGzHOJ87xsdCo/SjHOdYzqf+mBMlk35vDwtRNofZ7ZytT3PGYkELBCsLfXGhubSYMB9bdHAx4gsSHi+O1SxAzLcuWrQA8rrafp7BEN1Ylt0JTf8wdQwvxeZ12es22uuu7fhM1EbZLt7a99Hec9rtb8W4qVBPbAUn9pswZ79FF+IcqEUnXoNuTbqhG7t1w39d+7af+siZr7nkhtNf8MjrT370uS8brD71Wfgrfww4pr/30aBol7e/JhKJf3kYP+vRazY/+x0Tn7j+f667e2s3c+fWbt0d27qZwyDi+AxFKVy/9ZMUuP6GSBe24Xg+tRfXZ5LxsqH64jq9z8bx2t6L536MNhZFbCvxTnHL6T4tI7aVL5CG4lSJR87ShzFxnAmMKbvmnFWY81tZwXV8+APKc3dd+9XlD7/gZoSB1+Ad5a7rFOYSiUQikUgkEgsICnM7Ltt17idbYc6p5JhJryXjkcAbPVFvk3aRvigpVkXSPopM4htKRMN4kWNpZ+Ld9Ne6E/OEIFfsmsPqPeEN9vBr2RPuULbCnM2BUmJSf7wtENo57D2SeNU7b5TyHc1YoGjR4qXsIcqRPrfm90WP+XAMFx1WbxdDRjuHvq0uePqiXt+nnF/b5vFb+vHm2dv32j+f3muQ3d5z1XWM6LPPQ8KcaG3z4/GM9mAM1PF3OnV4DovPHd3U7Vv+cfzD1/73iZuu/dr0bdd/fepDr/ujNTuf/eHTLn3Q1sE5pzwRf+2n2B/9PKQwl0gk/qWB16wzTvn+773snJ95/oemDm89su4uCnPbuhkJc7guSpjzW1lx3SxiXGGN6SWe62nauL6yXsh+p/wxlmN4XfZrfFyjKxEbIqaRaEc8ibhX2r0+i2klRnmbfZqT1Hw1fsV8cSz7XdYqzCnm6ffl/DfmwGk+jfXgdj2V9dxbr/va8odf8BG8pa/Fe5rCXCKRSCQSiURigUFhbu6yXed98sZuSsIckm2/rTIe9KDEV8n4lm4GSfg0EvUqzG1CEs8f6I+2JesS5pioK5m38SqjLjK5BukHmtDm44YoAU79TtlCcIsxtoCoCwXvK35ht1syuXjg7bG8lZLiWqXNHwuMEOZCgBO9b5i2u62+JrP76wy/eN18X33hEJRNCw7Uy1y0+Tw+d1mQaI4+2x0G4d+OKcJcs7Cp52ttnU/Uh8ab4Gd1OyfrL5+Hl1ZHH2ifPd8fts0nxsXfgr1W2n2BqIdBVHGuPJjE27oFlp/dgdkj04e3devIZoeItbf8w8RvvPTw2a/8/htXPvaCy5eec/pDB8uWPQB/+Q8F+dCIXIAlEol/iZg66TEX/MjZm5/1q2vf/ao/WXfYd8rdwZ1hW47U66EJczOK5yjB6f12jSR5/W2FOV6n63Wb12GwubXVRDmLAyWOi6hzPtojHrHNWOLxRvEC9vhipcREEfYh/4hR1sfx7LO5+20rQ5gzol/2OfVrxxzqkwe3d5MS5rZ36w7Qtr0796PXfm35xRfeMhgsej3e14eBGRcSiUQikUgkEguIlcsvWbPjebsuoDDH35hDsszEOpL2+EY9knf+SDSFOUvgKciNEOacJrwgaRY5B0oJc95mEs6EW37HIOYyoYdlsIo/JpjZAiJsIeYUO8oQ3ApjQcBjqL/vFwsGE+aOQfezhYq9PutDqT5fKLCMuVSHrxYOtmCJ96MsYDSP2WKcjiG2/k5+djhOK8wFY557YzkPzWHz9M+tpZ2LXk/7mfA91+uvrH8DYYvPJ2x2ju1n1vs8yQNO+kuU4+tle+MRCsp6AmGhC3M4z4kPX/3fxn/7lb8//vYXfWLV3GXvXHXN09951nMf+auLJ87+CfwHnGP/CMdE/j5dIpE4ETjqtWfp+rMffs7ss96x9kNX/NfpfbP/pIc84JpHhijXF+YipoO41haBDddTi+u4nrbkdRal2Ahzdh3367/X7dqNesQGHKvSYlyJJ7AZeR0nUW9Y40vlPLtipsfJMh9pYtwwdXzulpMwZyVFupn9fH/munW7rv3z5Q+78KODwaIr8NY+HFyhNzmRSCQSiUQikVgQrFx+yeqtl+06/9NvtIc/ILllwl0TeNT3b0bybbvk2qe3aTfTXqfbmLjbN+0oPWEvokwR5ewYJelG39EYooyVbDvV5/04ryKqua+JQ04l/z5GflFvjiO70wUgCnLy8/FaBDRjCqNPr8deVyxUqt1KkX1u0wKDi4amXwsiMN6nMhdtmr++d0GeQ7sYkjinz5LEMTSHMcbEzjn7zIeoeWyOcl5u1wJIbZ6jnVdPmEO9nC9L+pbj1774fGJcfI7zRDn3j3PTa1Abf5s4F+3s5KIUi9C4DXvq0JYj3DmyDpzBgmxmH/o4ju3f39Gtu2fL/1215dnvXv6AySf7fwLBW11PBpeBeWtrIpH4bgEFuiVgXJdWrnz+o167+l2v+i8zuA5qNxyuf4gDR3gdDFHOhLkgro24fpIRh22nHEqRsZ5ts+nazWs5KWHOYzx9/PprMaBe1+fHD49DLF2Ys5iHMS1hr2zm6bFvb8dNkJrL4ukotsIc2xLmGBdwXinMJRKJRCKRSCROLJYvueSsa5916/mf+9lu4vaNlvQiwTZBzuk745iU191yo6kkngm6kvUg257gN0l8yzbpDlsIMyGgqSw2LzFnEYLQLmML3bdt+5xk+BfiPGxXVjuGNpxXc34SkGgH63nTx+eIPtjUdzRyweALjNavvn82Xxyj0t6z6NcYzBMM0a0uiMgYa7RFWtuPBUzb9vFa2BR7jLfjsx6vdd77OcJX/jxf+Yz6vEbM49TfEFnm9deA11t3hdgidOoghTm0eWsXF6fkASzM+J7uxd/qgU3d+Adf+6dnb/mRXz3t8ke85tRnP/SKk3/gftcvHjvzBfiveAh4bwuz3EWXSCQWEhODM055wrJHrHvGyhc++qozfu4Fn5i4bcM3ZvZvwnUQsRnXv8l9m77J6+A60cS4liHMSYAjETsLiw3X1vIlmvUVga65ZtdYV2NAvS5b/IkvfsQmpihOwa8Xu3ws22IcZ5jRjzHz6fEK13s+udsEO7PZ78vNYSxJQa4R5j52zZ8vf9j5KcwlEolEIpFIJE4Qliy55IxX/NCt533xzd34bRuUiDP5LjvkWC/CHOh9o0Q5sgpzLJEoe4KtNpL96B9mT7hxW9kZ522zhbDWCGexaODxw69H8y92zWH1YX9baNC3mT9eR3N+PWFOCxinXkedN15XWUy01ALFyHa8/qAteMj63lQ2x6EvbViYtKJcLIj6glplFeXCpz0PnhdKnhvPUX3Wbz7+2vx1jmL7flTa6zHW9zDYfiYttUAsjHPkayC34HWS7SIUi1I+tfUw3gftorN++91ElJz34xv+9/hN1/z38Q9f87Wpe7b93cw92/5mzc8/7+ZTvv9+rxgsX36e/4ccDSnMJRKJhcIZS6bP+sHTX/y4G89624/devZ7Xvu18U9s+Ob03k0lTvOapt1zuL6ZKDdfmNPvy/n1T9dSxMF54lwryrlPG7dps9iOknS7+jB3uS7j+KOomOd+Fk8wDucW4lyJQc28PcY4+PcFOS8lylmbwlzEWXsqK2NYcytrCHO3XPO15Ref3/7GXApziUQikUgkEokFBIW5n/yBj5/3x2/txijMMZFFkh9Cm3bL8Tfk9s6ibeLcNL+hF9kPPyTLwUieKcJEAh1JfivMSSxrhDQbY76trQhmbZvCGseXNvqGFg/6prz0+7GafolCmsd9cL71N+d8/tbfX0u0OSYWJiE4xbnb62UdZfSRWCiE2BWCmO1EM5sWFm4TyxzzOSzY6X32BQjb2rFAxvEalgWSxrJsj2VtzQkfnlPMb6/f/XTseF/4nlQfm5d+4VPfp3j/WtrfTYxH3ct2HrH5PPS+6jzoY20T6ey1V2EOtFtb8dqD+LtFyd9k4uKMOye0i+KuG7upO7b+7epffuGHTv3xR77+pMecd9mSqbN+aHDyUj4o4nvAtfyP4b/NUZBiXSKR+OfiaNeNleDU4NQVjzv50fd72Vmv++G3nPPrL7pnzcev/wZ/D5bkz0mM797wTxN7Nn6Tt6LOIG7N4PpIAa4IcfrSgtfHIGy8huqaanGw/VLLru927dV13m312st6XKOtHdf2uGYzhvS/+GnsHlPIcs3XdTna3h/HRL/9rIQLbaBiNf05F2PZIVBti1m8rVW3ttKf/eDRhDnupF73kWu/tvyh+VTWRCKRSCQSicQJw5IfPvU5j7vt3D/+xW6Mt7IyGVaiTjFuswlyeriDCXMU6oK93XNMylsycS4JttskrGB+1AtdbDGfJhkXG+EsbPDXbrdhtj6ewJstRLZ28VEXIWVM8fcx8vc+9Td1+uj1kXyNfL08bzv39nXEAkTUIoKsixUtHEQuGNq22Xrjh8jj1bYtPrQA4eIHNrIcT+y31e9jTQgk7bgifONc5FdeF218bf66+T64zewxb7xH7K/v+TA1R5nXWObVWGfvM8Q4HxuM11x2Z2ARJuJ1xI46e9owFq8suYtOgh3t9NvRje/Z+o01H7zqz9b8zivvXPvuVxxa8ys/vuus1//wr5z8pPtdPzh9xZPwD3Oa/d+MRP4uXSKR+OeAotxiqw5hxWD9yQ+/4KVj1z/n/ePvft0frvrdV//R2Puv/O8zt89203sYny1GT+zZeGQS5BPTJcrhOiphTqQQ1xD9ZbccGF9Q2XU1rq24HpIYL8a1WX0cZ+322juPjDGj7CSvt1636z2vwZg7+qOPx1H/CDI2OTlfCHJB+725sEf8M+pLmIMU57Z3M/v4xcy2buYjV311+UPPuxmX8NfgnU9hLpFIJBKJRCKx4Pi+kx77gPec+8W3duMS4SjGbeom98yKU3s2ok1R7ijCHBL6vjAXibUl+Eyie320oV4EF+6iiz7eQuNt69tiAlksGFyAY90EORPQQlCTHccrohyOUYQ2p82Fuh8/voXXN/Fqgz6vjjmC5XgcA4ZAFa85Fi70bRcaJhD1Fyx1Di4YWLbz+mJC9fnUfKrTZ5sWGBKkmnoZXxYq7Xj2W72IcTzvqLufSr121Huvy8o43xDSZOe8PJ7b+p8B5zK2fxt2bDtmOzY+s7jdqi/OuT/I16z3wY9N2m2ssB2a89+aw2K1lFik0ofH2r/5CI5/xHZVbMfYuW7iMBZwv7+zm/zUtv819u5X3HX6S5+wfcl9xy7B/8wqkAu3s8EzQD4sIpFIJL4dLAXPGJw8GB+cvvy8lT/0PS9c+ws/8bF1d+7oJvYg5t6+tVu3e0s3s3tzx1tYKcgp9u7FdQwlhbl1uCYWYU51Y/lZiiCvpbp+4lrK6ymvq2iXa2pzXY22ru3ysWt1xB1j4w+2cU6+nLfpt+sz5mI/yp4w5/4RD8wH8Ygsbff1GBb9Zadcj34uGk8f0nbN6VZWtKdvueoryx5y3ocGg0WvwGdwEZjCXCKRSCQSiURiQfGIpRdMvnX8A1d9ZWzf7DfGb93wzSksAviE1sndG7qp3Rs7tsvOOQp3SOTtFlajRBMk0nGbTCTfkVjbAqBJ+D3xVtnc3spy+PdsJOb4wqG3O04+VaARKQCRGldt0V/tFIi8D+dXfFvymD6/LRKMPWGJr1GvFT7NcYYZ74cWAKK1WxGuZfjHoqO0G7YLHxOjjEcX8miP/iDsGhPHddsQ7XWg3tDeu2D1FWNBxLE9vz7Nn3Ure9RCrSHfR2f73pod48Hp/XgPymvj6zLGLUwt4zeXTJjDPPs2fXNiPxa6B/E3zl10hzHnHTj/w/jM79jSjd92zd+s+e2fPrhq57//9TNe/n1bz7r8UVtP/77vuXHx2StfiP+hGftXOia4MyZvdU0k/m3j6NeA1SsftPIpD3zN2a+75M1nXfmUnzv7xud/YM17Xve1qT2z3cyezd06xF+W04jBjL/cJcdST1MFtWOOxDXRxDizGy1ux5cbJc6yrmtoy7h2N0KYrtHG2o5rbe2bz3qd1rWaNlxzI67VWGB9de5qszF+PW+u66zXL5G2lS/YjJy3nY8Mv7khYQ7lh6/6s2UPPvf9g8Gil+KTeDCYwlwikUgkEolEYkFx38GZp738rOuec/Pam9/w5bUfueb/Tu3epCe0TpK7Qe6cK8IcyAR/f6WEOSTCPWFOibgnxpGUk9wVJyGFbfQN7ZDTQoGlC2BlVxxsfTYCGcnjgCG+9cQ2Juk4VmtXKfsI/6DOw867+sZx/bXxtbIPjNek1yUfYywuqjDHRUUdp2//QZuLjIVHtPus40fYe4TdF0BGjmv6aeNxi4/bnDo3Uq/Hzy9eX/SJ3kf666hs/Vr6XKiX9482EW3N0/qbn/qa97b0aS4nX0/7+tv3g783hzJ2ypXbX/VUQ/xtS5jD3zPKSZSTd+Azv2PzkYkDm74xsfv6f5jYveHvpvZv/Jtz/2DnN87/9I5/WrvtuR9Zft/xp/v/0rGQolwikTjadWDZKc99xCvG3/+6z07dOdeN33rtP0x89Or/NXW7P3UVtN94xbUJrIIbSKFObGy4NpowF74sGa95rQRhC3Gu2EhdY+v1m9dKillhi2usbPvNXlnHF2LO9notG6/D3t/GiNr2Po6JehuzwoZzMNqYsvO9KVtG/Ku3ss51M/wyB+X0h6740ooHr3/3YLD4p/BZPABMYS6RSCQSiUQisaA4E3zwSY/7nteu+uWf3LX6A6//h8nbZ7uJ2zaaOOfC3ORecpNYhDkl01GCvCVFCTAYiXhL7Y4j0TePm63UONatrZ1xrIc9+tR2KrG35Hu+wFYT8yLAleNYe6QoBxa75q/2YLtIqK+70hYDfdqOrtoeuSgJESl8mmP1hb1R5FgfT+HJ/Qt9cUKf2NVWx7C/zlVeZ3lN8InX5n1xzvW8+4ul9vXZa4rXEmNrn84nbmsqdR8X/RqL0o9fPr8yr/sFD5E4nyHGudpijeOweHXG3/P4/tkjk4cx/g6079jazWDBPH0X3qM74U/bJzH20Oxfn7X5GW9f/pjzL11yn7Xft3jdGc8cLF/yg/ifejg4DR5rgZdiXSLxbxErBuuXnHv2Dyx/5PpLlzxo+rErnvaQHz/7Z378Q7wWcdfbelzT1uH6ZU9ZRYzds/GfJvbOHqm74OhnvqSEN/0MhQlwLW2XO65XrHM8GD8bYbGZpV0D63XZWK+7vF7Sh9d2o8WBoF2PNQfbMd7Z1lv2x7qfrstmL76MKx5bajxwG+oS5eAfMTvqRjuOXes5hwlz2ilXhLkr//OyB8z89mCw+Mfw6dwPzJ8nSCQSiUQikUicAJx92iPO2vyst696z6v/TqLcbRu6ids39HbMFVFOCT7oiXyhEvemjaS4TyTIPWHORbYgfEw0Y522+X7Dv1Em4lhGS+b7fdVWEncdY8gW7Pn7sUbN2y4+9LprW7ZviZzbyHZvrM8pO45nC6f5jIWRtW3hIfr4VpTri1G2qDE/s5tfzBVzt68XfjyW6vMZ89vCDbYYyz71e0lqbtbttQd7T+dzkU5tZ/Hz9y1uXY7j9MhzaUlBLoh2+17UnXOYH3ORE/tmj3CHnX6n7jAWtHegvAuL4buxoKNAB9v4x6/9P6vf+Yo/WfUrP7V3zW+/5MCa3/jJO866+pL3nfTEC3cOzjj5UvxnrbZ/sJHgwyJSnEsk/vWA/8/39j/975Y97vxnn/mm5/7eqve+9tPn/IcX71v9tpccnnz/lV9bj2vNelzzSD6gJoiYe6S9PbXctuqUWLcXfhLm2G5EOdpZJ3ltQ38R5iTOOXUtBuFj13va2uspS14v3a4xVho5vj9muD5MzcFxPlZ+ui6zr9osVs2nruvhA/+IBRHXg/F6FJsOhjC3vVt3ACV/U/SDV/zpsges+/XBYPFl+HwuBFOYSyQSiUQikUgsCIYXEKee8ZNP2rzmptf+7fit15kwt3sDf8dGu+Xi4Q9kCGclkUfCK9FFiTsFErObnyXKlnyjH4sCUrenYtHQim7yZR9K+205axvNx5Jsp44B3wOwo14ScF8kRMJeknOOF9GOUmM4h/m337xLmHN7HDPGaJzmx+vy114WEaDeExELCK/HuZW2z633Sv08htnjfVU9Xg8YY01As7ltXtaD0c8yBDfYneUc1e771Tngp3MheQ5GO1acm421xZGPaeYX/bxbtq9H9Hn0PuB84n0RuSgt7y3Pl/Q++ukc6e91zYfPp4z3c+K4Q3ONMMdbmVD6nLqtFePJslD0Y6tP4ty2buZO+N+Jv5HDs0cmDm08Ms3bXum7B8fajXnu2t6t/9yOb06+/1WfO/OVT/qFZQ+cet7glGX3x/8Xn+aaD4tIJP7tgPF1Bcj/+fHBGStmBqcMVi95yNTjT7n+Ge9Y+4lN/3ty33ZcX3DNwLVp/eG5bh1/23L/pm9O7Z89wuuOHuKA0nbPgbxGIQaFGDeDGEpKiPOHQNhv0OGa5EJdxNAizJH6ggzHYh/avHZazInrqBPH1LXX23Hdlz3aiD3qc//CuPZGXMBrsOs8x8Rx/Pg6to0rJc8z2j6Xrtu4ftuuZ/TR5v6cRz9xgddUY/zmbpx2HYN+PB8X5g5VYW7yva/74tL7T78NnxN/lmAdyAdxJBKJRCKRSCQS33HME+ZOfs5j56b2Xve/19xyTTd522w3dtv1R8Zvv/5IT5jjb91IsGqEOSXIzrCBTJRtEYB6QwpylZtg44IhFhCWTItF2EM/2hLniq2yCHM8plPJOhJ/PXWVPmRvbrc11HxK3us8mr/HOrY9Fl+7LUI4nu/DEOl/DLaLECPsorXjvKK/95ADsY7Vzi+ci3y0mDkKh8a1Cx32DR9bx49z9r52PtvVNuo9iNcSjDmDTZ/maPqiHXOj7M/pfjovr7dszm+eMBfvnfrrvEZv006Bjk9wlTgH3lGFucmD+HvE35/+LrnQ3YPzIHdd/7fj73z5p9b80o/edPaVT3nbKU95yIbl95t8zWDlkniq69HAXXSJROK7H4yfx/5/PWkwuezB655xyo8/ZuOZGy/95VNee8mbT7vi6b+26h0v+ezUXoptvAbhmsLrC4X/IsxtljDHazl3zIUwV29ldRGOpBDntB1yDdGvuOlUnMF1MlhjEK+XKP26F9dPXQMbf10vNc7H0i/a4U+bM9olRsgex7R40vqPovo5tkfaGj+frxfjwXGSdvaHr675JsjNHAAP7+im3/2azy+93+Sb8YnxZwgmwBTmEolEIpFIJBInBGtXPPEBO6cOb/y71Tdf2U1+YmM3vntDN64dcxTPuFPOhDkJESHMUbRQyeS7ScRpY9LNhNgFNxGLBUuaN9mOORJtE+aMJbFuRLj4jZxotzRhjkTbWRcPWzEXE/ahpJ3jhljmVPLuc7nNfHh8P5b8jOU1Y8FQ6ugvdbDMrQWMjyu2vq+RPvRtWf0ouh1NnFMfzuXehLmYN8aJ5TOED8q2X+cZ5xz2dr5mTtuZ4ONEjglWu5Vm12v24/d8uADz+eP8+vPZeekz9zF2/Hpu4tF2zMnP5hWH6069pxhri2gcTw+JwN/EAf5P8P+jLnwnb9/0zfHbNn1j+iAWfXft/KeZ3Zv/cvK3X33P6S98/M8smTnn+/D/1oriLY5mTyQS/3KwiLvjlj/iwqecee2z3r7q3a//0vituC58HHHvo9f949QnNnxjcs/sEf5WHK8X9oTobd0Mri/rWAZxbWp3zGlXr0S5RpgjEUf7whyui7APM65PlfALxjUP1LWTbHwZu2iLuKq5eP31fhtn5XBdbbyecn3WeYw+rtp4vT07xxay7dQY0OeK+D7OOF3q/L1QO46R1/tGmLtjZzf5ztd+dtn5E9vxuT0KPAdcwg8xkUgkEolEIpFYaKxZcv74pumbX/cXa3Zd041/YmM3dvuGbmzPRiS3s93E/lkkvpuQJFdhTruFDjLhtQS9ZSTKdVFAIc7EOIpyanvZLhzISKitzXm4oHCyHX5MvgsbcY3HV7JeE/ajc75PLAZsMdLvK+JfYyv+DUMs0vysizZn9bu3OYxapICtSBTUrgosVkJIK4St2mOxwwVJpS1umrb8nI2tjB+i2ejHOsegLTZzFhuJ11rGNWX0saRN7XaOoPf5OL2nPlbvJeeS3ce29SFKZGNdr83mly/GlN+ac8Z51vnsfW+PrePv23yEDNF0Bou/mYPb8ZnPGQ/t/Mba33v1XWde/dS3rnzS9/zE4vuuefpg4vTnDlYseSL+/8bs3/CYSNEukTgx4P/esf7/zhysXPqgJdOnf//JD1l/6alPu+jFp77gsVec/vqnvn3Vf3zJ56Y+vqmb+fiWbvr2Td30bsRR/71WCvhTe7cc4S2qVYSz0sQ6lt7G9ad3qz1IUc6+tAqxzuoWJ73uvoozYlzLSLNLXEO7dw0FLT6izttVZWMb/cP+0S59MS/Y+BV/Z2nj9fG6XK7DaEcMCjuv/+SEaLFY8TiOg9dquYPlAtoppzyhCnN6TQf4pcwc3tvt3Qyuy+vu3NFN/c4r71mybu1V+BzvA54ELuaHmkgkEolEIpFILDRWDc5Y+frVO370M3wS5epd13UTt89243s3IrHdhAR4FgktFhMHNyE5xoKilEx4jUWgYBJMcuEhugC3F/V5wpz77A9f9gfrXFWYa/yKKEdaMl6OzYQfCwRjY28ZC4NRfSAXGMO2kcKcFiJ1odErOUe8L263xQtt1qc53L9l+Nuixe1R9uy+mAm7+nxRw0WIFiOYS/5u10Kn1kX5mW/bLuMbxnHM3/2G5iflD7uOrTH1tZb5vN/OL9jM4Yy+3jmEv7djvuE5RrGIc2TMBw4Lc6I+M6f3h4+Jo7aADqE0FtYzh0i28Rr2bvrm+M1X/c3Ye1/zpYl3vfL3J3/35XeP/8pP3n32Tzz+l5asPeMJ/n94LNybOJBIJL4z4P/d0W9bXb3ygadd+tDXrNn67N9a+44X3bb2t17x+bW/9vI/Wvubr/rS5M3X/O3M7Zu6GcTTaXKPC3Pc2bZn4xE9XRXXlfp7cna9sB1ztLOk2I9ri8cOu6W12TGHWDkszLWiXJ/1OqbY5HVrN9dWXtN5bddvyNk1Un3wDz/FsuHS+3ksi5Hu67S5h2x4reS8mETSB+/LME2YQxnHYU6A11wIG29ljX573ThG/MacC3PruWPu1156eNHU2S/DJ7nGPtD8SYFEIpFIJBKJxIkBf5z+kpWP+56fnfzdV9x91rte89eTuzd143wa6/7NSGZRHmyFOSS8qE+w1A46JL8sI0kGy6437pRrhTmWLsrZN9qok+HPsRTbkFzHbbAmzCG5bmyWcJtf8eUigAk4k/1I+tVGH9vqp5/5lkVJz26MOarYaPPZa+UYS/bl63PJ7j5xTCuj7fNqfmvHXHHMMkcsQthGqXEsW5vsXLxwwWF26xta3PgYE9qGfHRrJ+1cgHmfaHPa67L+cnyfk+PKvN7H2z1j7nnH5OuN903vhb8PnJP9Po6/51aOVeYnUW/mjveE4+sxGrpv9HGeVuSL/nacfdb8XIx1gYk6/sbY5sKYvlosYx4T5kD+Dh1JG94H/N8c0a2v4mb0g/ibndmzuVuHBfl597yxm7nl2v985ssu2b703LGLBycPxgfLlt0X/4vng5PgmWA+LCKRWFgcSwBfCY6Dk6itXfI90489/YVPfMOqX3rRRyduuf6vp/ds6dbt3dKtR4yaoQC3Z6M4+YkN35y8feMRCnMzJOJgeXADrym41pigT1HOridl961KtnGN8uuPrkGMjYqP3mYcwnVFv3eptl2nRL/m6jrYu/56W74cE74ch2uYdsxF23zlo7pfO6Pfj1l9vS2aj/r8ultijOysW9uu59a2uGOcxLU14mIIc3FtZj4QItwEXr8JczU3sNeFOfbPYV7bzbyewtzhnd3kf/zpvYvXnPk8fKZ5C2sikUgkEolE4oSCP3Z89uCkkx591ku+/43n/PpLvzB++6yJarp9FfVDWGSIm5EYD1EJP0qKeEHuits7a5QwR252Yc5JYY078ki0JeYpwfZ64xcJeF+Ya8gFCZN/XzQUDvlpEdHayuKESXw9Tixa1Kd+p9vK4oSvPebU+1Dn7LPay3yyeb0c1/25ONECBTbVuThp7MFYvJRxdSFTdoSRxc857AObzqVH2P11tmMobrE/2n3SbrRFFGwxn15nzOcMX9Vtjiru+Tj5hs3sRrxnop8TbTGv/Onj84juQzbin6jj2GcUn8UoxuIzjhE77wq5Qw5kHf5HYicdf8R9Gv8r0yjXYax2w+C9GN+14e/X/KeX37Fq5/N//awtz377mZue9RunX/6oNy19wNQrBicteSz+L0/Xf+hocGdH7qJLJI4f+P909FsZT1t+wYpHnvfCM178xG1nvuHpbzrzhst/b9UvvmT/+Lve8JXpPZu+uX7/tm49/q/XIU7pVtPYHbd79ggZgtzM3vp01b4w56IciWvEME08a+lfWono5zVWwhxZbf3rm9UVk+AzfH2L63ShjutzoG3z8foG8jobPjGP5qYv63W86rgexvU2bPNK9fvcpMaA/LIDfeSxhDmeo9p4b4e/tLNzNmHu3BDm3vai2xafc+ql/gknEolEIpFIJBInHsu+d/JHxn7rpQfPufVaJLCbtCuOolxPmCviHBJelCZiUdTiIsFKu33Vxbl9JvJZG6ULbpZMuzjHtug2JdVmU7Kt5NrLsPUYiTeoJL9ZHDTUIoL1mI+24lfnpa3HGMfFDNm0tcDpsR5fi6Oog8Uf48s4P2bvuFycaIECm9dNsKp9XLxp4SJ63Rc+EonahY3f1lqoOUxAsn4uWHh+IPqizl0GrT/LurPBxw6xFcFUxrx6vVbXfCzbuo+fv+uu32/nEvPx/XHfds5yzvBTnzFs9iAIL8NGX5/TPov4rPo0O/xBE+SitHO335cDuQDEefDH2ynGmSDHBbgtwtcdRv8dfDLgTns64D03dDNf/Nlu6lPb/mrNr77otlOe8/DrB+NnUpw7w/47E4nECcSZK37oAS84680/+t41t17z52O3bv6H6X3bu/UHKPKgxP+/Ce4murXkLjqJceLm2kdfxAK7bZWiXFxPnLDNI/xth5wT81TiuiR6XxDj4noWQt2oa5td+3AdbBjXOjLEuKDdHhp9dqx2PtmjH69N1+Dizz47p+LD0o/5rQtzLsSxTrvOm/lDZXk9Ou85vI9z3bp927rz7ryhG3vzC28dnLaST8xOJBKJRCKRSCROCOb9ds6SVaf98PQ7X3bPOXuu78b3b+wmDm3uxg5u/KbEucMU6Db3hDkKcpZYU9Ry6lZVox764MKc1RsRLvxbKokeQS402Bd0eyw82nZJ9hu7BL4R44/GWCSIo/q5eGAd/bHQEGMBMOSvBY/Oi35exvjGh/ZYgEksipKLNCxMglyohPgW4pDN3dAXNNbPH72ui5yyyCu+4e/1sBfWvnhfeoumIG2ln/7WDpaxPX/3CdsQq6gWpG87V233z6npIyXEse79vffTfcpnNJ86XnxmhRgL1s/BhDm+31z8SZiDXxXmuPgGcR7Th+FzB3jn9m7mrp3gDd00efcN3fieDX+15jdfetvZ1176Cyt/5KFXLHvg1PMGq055Nv5F+Xt0a/XPOhq5iy6ROD64YNl5ay47+Un3e9mKJz/gBSt//NGvP23Ls3939Tuv+PLUXvxv78f/LX+rDP/z5bfheE3Yu/nI1J7NeqjDOnKfUULcXt8lF8T1pBBjW1qcANHXJ/pEa/fiGep2OyvbvF4FG5+htglXxlHxqxyP53S0spzv/HZcL3X99Hrw6G0KaKSPi+uzX6NNmAtSmKNY5306b9TxuvTa+H6wHYSfPrc9W7oLPvWz3VmbL9s1WLniB/wzTyQSiUQikUgkvgtw0rLnrt3+gs+t+/wbuzV7r+8mD2/pxg5t/ObkHZu1a67snJMwR1HOd8lpF5vTd8tVMY7CnN3S2hPm2jFiXQz0qOTaiXZ/IeH11obkPOpkvTUW7cav78N56znEIsEYfkzsrSyLlWHSH2XMU8g5wFiAmD/7orTXwj4tyrAAMbEINq9L+AkxCXWKS9+qMBf1YkN/XQCSQz7FPtSPen1fGv8yzsj+Oq7OE2N7/l72xbdhckwzzue0uYy9eXtj0Bd1vHeTEuLAQ3N4LymihQ/f/+Yz8vntc/PP0D8zHUvEPF63z8EW6FY3YY7vcd0RY9Rv0R0mMf5O8I7t3fRdWCzevaObBvH/9I/jH9vw1+Mfue4vxj92zZcm3/Oae9b83I/edurTL3rT4IwVj/f/1lFIYS6ROA446ZH3ee7qX/ipXWtv2/D1Ve963Z+s+vVX/vGa917xlxO7rv/7ad9JbF8a2DVat6PyfxzXh0rujEO5F0Tsmd4bD2wgh4Q5ktcRJ685nJN2xYYe0c9yKKbRZvGH/ZXRH3217XHP5zDf1mdoHp7X0do652FiTpR2HXV/Z2sb7tN729gL8V73RTknXwf7Weq8wXhvQJ2rC3Ss64nZt2/uLvj0m7uzrn3WrYPlS37IP/ZEIpFIJBKJROK7At9/0sPv85vnH9z8P1bfuambuGNrN4Fy/M7N3YR2zJFI5F2YM7GCiT36y2/KsXRxjnUKc0WUAymUxUMfUDdxz5NpMBYXSqZZMrH2hUO5jVVt1nlsJOOyzyftdjzzsV16PleQdme149hK9O2beCb/Ia6FUNNbdPA8Uartc+g2Gh0b88Y4+nFu+dMW49ln84kuFoluC+Gn7QthrowDY86eONUISKzbXDan/Nxe/HxM7FbTnPJln5fsp13+YTN7j2Hn61a/zdUb28xlIhrHgRTRKKbBVndM0LcZ759FO66+V0b1+a2rFOb0Wg7NGWOc5qpzlN+w889On2t8ZuUz5Lj2vWzG6/Zhm1Ofk/tpZw1Yduu5QDdxaMuRycNbj0iou3Oum7lzO/q34++HP1h+QzfzmZ/pJt7/2k+f/rLv27Hkfmt+YHD68vMGSwcPwf/sA8BV+u9NJBL/XPA35dYOVi57wGDVGQ8ZnHnS9JKL1z3x9E3P/k9r9mz55vhtW7rpT+B/dQ+uM4wRihN27Rnfu/nI+N7ZIxJ7cG3QLjn8n6/bj/9xXC/sllV/0AN3y/mOuYhp9vAGi39FgOM1g9cW2nTtsDqpaxDG6XfktDMurkUcRx8bG9eneHiDbD5OcYl1jClxT30Yz3nDvz2e2mC5FsKnXAvZ9jqJa1ypw27nXet8fezTNd+vldGO62cbH0x0s2NYvcZTuz7Tj3bz57HkB+r9idfE185bb/WezHUTt2/qzv/Mz3dnvPqpFObyVtZEIpFIJBKJxHcVzhuctOx5Z/3kD/7a2puv/MNV+687MnHXFglzk3dwpxwTYSbGsVOOyS9LF+UacY427pQrt7VyESBhjDRRjgxh7pgsx/HxUXeaqNZnWXRgvCX1cXxrf0uMxYEn/9E2wQXnVWxWV9vHxusdJcyJPk+M55iwFXKhMtz2xczIXXJgOUf5sez3kyFcVfHKxwRpQ1nFMJbhW1mONY8cT1abvZ/2GmOcxmou9+nZWsIu8pjOMibee46zsSaUDc8DX93KOieaMNf0+/j6vmBOjqG9R35WsJfPEf7O4tObs/bH3K0wZ+Q5bTlCahfdYYwTKRyC+7bjf2lrt/Yj1/zNOb/zynvOfuvl7ztzx7N+++xNT/vwyqfc/y2D01c8yf9/jwXupMvddIl/S/hW/uaXL77vmqed8dLvf8uqX3jhh8/Y+tzfOmPLc997zn96+eenbt/Srdu9tZvZg/9FfdGEa7liCtqkRB+KalWYK7tjUdcuOPiXHXJOiXK83quO+BftQo61uVvqGgrfEJziOlSurfRjf7HX84xxRZhTyRhp9ejvHYuknbZyna3tOKYdH2Pj2odS1/I4N/o722v8yJgimk8IbPHFmAl0NZ6W2DREfZnWvgd8XXzt+7fJTmFu7PbZ7rzPvbk77YVPunWwJIW5RCKRSCQSicR3F5aApy9efcYz177px9+z5tDGb4zdvaUbvwtJ7WEk1hQrJKRtEk10i1tVNzY0gc52zZkI1xOplKS7KNcT5tp6Q184sK7EnIsjt2lOLjBYOodFuWDYR/UFdQw/R2M/6WdfWXhogWD1umip51HmVH+Qx4gxdZy9tjq/yMVK8TOGuCPxaKgv+ktb9b6fdmD0BChjfR2xOKLdbBLBGt/iz362Nc5Y/Hs2ex8Kw4/9mq/15fg4DvvZNlrb+xr/qNsOE74uvj57jcX/ULyuuW4yRDlnO38V9fy4zfzlfPh69JmhDRbhje91OTevg3qvwwd1+x061CXKkWhr55yf8yHMLWEQx/GdqWrDzvb4fvxv3bmlW/8H27GAv/qvTvv/HjvHnT72L6zbWZeDfNoydwOlGJdI2P8B/x/4/8E4t5JceuHah59xzdPeMnbbdX81jpg2vmdWT1K13W3439u98chUiWWIT4od6Od1HdTDGBCHdMsq6jPcLYf/c6tbX1+Yow3XB9Y53uuKcw1HCXNxvTFRjtdSXmdg9+urfDBXxEr505fHoH+p025+mifq6jeb4nUTJ+04PJ7NpfOIPvVbn8UDvwbq/OoYm6f6BeM62e6ELrve+NrcVuKI2+J6rPk4FnXb4c7xsImYm68ZjL6p/du7sd0U5t7arXzyw27F30EKc4lEIpFIJBKJ70o86IwXPOG37vsnP9+tOrihm7h7Wzd2cNM3Jw7MHqnCnC9W9m6UMDe5DyVJkQ42iXNiI8xhrCXVDRsBS3UuBhrajgLSFwDygS9oyTvaLNkWfUFBnzJvZUn4nWWh4Iy2yP7G1/z7bbOZr7V5PnZOpd9ZbOX1Y5wvLHoLnTj+CEp8ahYkJOdSHTYTo6LPFjs9gcj7a93pc2k+9JcF08jxtNvxyHjv7LX5+J7taK8p5kZdtMVVGSv6a/Ax7Jef7Ghz8VcWgPQLX5yrC186hgQ4ksIcShHjtYsu5gV1y6vV4z2wcdUvPj8dV4zjBmEj/bjxvlWaMFeO5ccT9V4ZJ/ZtPDKxf+MR/aajHsCC/wP+3iPLg/if27MBC/jN3dr/9OL9Z7z0SZtPftqDXrryBx9w1YoLx356sGLJ9+H/+FgPiiC+lR1FicS/BHwrf8vjg3NOecKKx97n/zv9hY+//tQXPema0173lDef/fafvntyL/6nEJ+m9WUR/tcO4P8LMW5iz8ZvTuzeeIRfMFGwk2iH+NLufCuCG64FEuZE1GUzn0rzk1ikawfrsB+FioHNWLLECVLXijqXxCiOo4/GV1/rc5/ol73xcVvE0YivxhjrxHGDPAfVdQ3D9ZLXQfdRXzNGbV7/WB6FFqejjWOXvqj7a/frpl2jtyHHqLRj4jh8r0GKcuP7th6Z2jfXjeF1rb/n57oVj/qe3fi7eLL9eSQSiUQikUgkEt9duO/S9ZO/ct8927vJz+3o1twx240fnj0ydoBCARN1CnIU5kgX5oLcMTdPmPPflsOipyfOKdnnfCDFtCgbFmEOfcFYcGhO9CmJZ1uMsTH/fB5dmIPN24WNb/i3bbO1bT8nP8fWrzBeP2hiDI9jx7LFltuGWH5zyBcjtI1apBi5OLJ+CURgEYawaIl69PVEpWaOYnf/3jj6+jFsEcTX5G3Z7PWGX5/9ucvxwN4cjV+IdCL9MEbzl2P4PFGnKKbfkSPhW4ixIcx5u7xmjcH7XGws/Vg6H//s4pj63OCjY2MsbTGXzyf7ENWn8xxi+Ttgm8dwcvccd8xxN91h/H3zISz8f9i7qVt70xV/u+p3X/nVVTe95i/HDl73t2t//UX3rHz+I28cTJ7Bp7iebP/SI5HCXOJfC+7tb/n0JfdZ86TTX/2DN579zpcfXv2Jjf9z9Qev/qs173vDX07s2vB3EtCa/zddUxVvYmcb2vqduHb3G31qv7Xjfxj0/lGs4pzb6DuCjCdxHvOoc22OBxYBjNR48y3XZ9rczjipOKX+8KPd4mjE1ujvXeNRxjWLbdV1XfNrIVnOhXP7Oco3/Kxdr/XzaeIcj+nHDRY7Sl6bxUaYwzXZjo1juDA3jvrY3q1HJvfgdd2N6/DH57ql95tJYS6RSCQSiUQi8V2D4QXNWYPFS1942lMffvP5d879t7MOXPdP+p25g0zUN3Xj3EHgQtzU3o0gyj3WNjGO4lxfmAsBzsQ5imOcy6kFAMv6kAjbjbcJCbUxxutWorK4YL22bRFhflpAqM/rWhygbJN7JvayuV2JvPuGj7fLnDGm2Or8rIcox9cjm45fGa/Zjmm0Y3E8FhA8h+YYsXipt1jCRpGGPrrVkbY+6+vzBZIvmIow5HOWuUmMK6LR/5+99wCz6yjShg8zo1HWaJRmblZwApNsjDFgm2BMMhgDJhgMNsbGARvnJGmSJig4YFhy2CV9ZDAGDA6SJkkzkixnw7Jsgm93wexPXqKxbv/1Vujuc2ckeXf5wIau53nV3dXV1X36dJ879arPvb5O2vvXNBlmF9rnxqrXwNcRjR/w/XsMkF788njZp/UjfQFCwoVy3l5tuCwEGreFPcOIOamT+SI9CDM+qaa++eSatGFijvNqC8T9xPeN7xfqxT5/ffCjfhuBscAn5ogQXl2jlAN2agtQPwy61wacoCuNra0Xt67dA2Kc995W2nO3r3aF2650hbHVbslnz9rddtHxG2cd/4RTm5/YeULWNucI2teHEmoEvMaXJMljWfZFwuFV7gL9e3C2cO5TZz6tdsKslx3+lvmXvfCahR87c3dhmD5HRqI9x3uZ9tRo957yWHed9x2BT73RHvXEmxJzfHIOZXq2A544I52QafpcsP1M8P+pgr0Mf9DDnvL4XJC9rz5Uh88OzvvPELXh8eIZJM8Q/tywZxDB+i/Rc8Ta2Hj8f2JxiueI+GU79oX+ojpcjwIkmSfKYK9tGPqclP84sfFIXRij6OV5aWXtE0Ae0GcjwONkWD6k8hzWFG24nfjneoX9h12R5hD64tZut/yBTa70kYtd87JFX6S18ki+ozNJkiRJkiRJkiRJkj+54Dt4lmYzW45fcvrxH6nd1fuTJWNXuRJ+kXWsyxU3r9mD1+gqwyAFuvA9PK6MdKRbCLlRsqG0qMSckFQaANAf0fsl5pSUM2LOfkDCE20IFOCrIXAQaD+W57LoOECxP/7tD3odE5f5D3nReVst+wDG2jCCb6QS7MT1BN+/wF8z99kIHVfUnoMXJXmEKCIdCDnYcd70CvNhwQ/rJWgyCMkmthYwxmSS1XFwpfYxMQefEmQScnMmZQOuyXzl+zAoaUZ+LaBi4sz3Jf3libmgjyHzgVTrYQ94IkyAU3KhHHzn9QSULY8x8bisD0CuLcx1AM8Xtds7MSf+0JaDW4InCWj9yAkP0XEQT30FULAOjNP+GgFBvpb3X2kYqaC0da3r+PKlP170obPvWfKRM8eWfursrYv7X/WJ+a8/crD1oI7X0d6276ObTvZ38ihJkj+3YH3iu+L2JktaVi198bzXHbVm0bpXf3TxR9+2tf0DZ04s+vC593bedPmPjfwScouey5ri1fEcMUf7Ez/qEE7E0b5jYo7aMjmnxBxAz3Tbv/K6KPrAnlaQL/+sVf+29/EZYflGYs5/rqHMeqrn8SJvzx/4l2cPEAgpGYeVWUdlD/LrP9cI/DzTPH8mc/9hDEaK2fMvfr7zM5T1BhuP1vs0b8fXYn2yT9JHz1j5WyH07WG2ahfaAPGrrNSO7leR7leJUKZ8x+1X71l5/6Y97We95AfZzNnvpfXyTFk2SZIkSZIkSZIkSZI8OmRqUL544UmF9529rXJfv1s6QoE//lBHurXLVTYTtq4lKEEH4DVWJeUCMSfknPxRr+AyEAg3EHICvBKrZBwHKKQzYo7KEswIhMiAjZSF4IN/6SOfRrBxxNA6++Of/8j3ZRkv+2qwZxuMP+qXgywO+qSNBFsEDQS9P/NFgYUELrCTfi2gy5FDIOUURhZJwBeCIG/r2+bhA0Rq4wE9I7Tl4AptvG9NqR7tOa9jNYgd1bON2UsfrOc66wvftSbwQaX1j365bxkHE3NcFt3eITbcP/w0EG78amtU9sEc2QqsTsqWZ1C7YEdlbct9wk77lnnTto19WXu97/4e8/0m8DqmdcFrydZRQFhPuj+wV2jfgZgrjaxxha2rQS4wmYBXbBHQFiaGXPHeDfXSlqv+YenGk/921nMPfmM2f+ZBtLuN4GjVNEmSx5rgR07wuvYsQlM2a1Zt5lEHnNh26cuvXfrZC+7s3LL2d6Vh2n+0p0Ceydcq9LjCaFe9ACKO95PuRULudKqWhXQjwIfuTyPmctD9K983h3bY20jFN/snXfAr8J9r1Jah/clzQP3CjtvI88I/c/0zKMDqQMwJORfZc15JLuor/o+kfFt8nuEzKqpHW9hE8ONhULu4Tp97/CxEWW1QJ3Zxn9BpX16HPHxqPZX9eKDjz0Gy9c9U6ova+uctgHajALVFOtznOoZXP1T84sX/2HrYyi9kWfMZtG4OJCRJkiRJkiRJkiRJ8qiR6U4iPG3m0w56b23z6h90but2HVtBylHAsHmtK96+pl66bXWdibmtQhDw66zAWI8rjhKQ8h/4CgqE+HvmEHigzISbkQyKsQZiTttyXgMVJuUo9UGM/sHOffiAQvRMkEEf2UmAgD/6Sad59kuwACAPac8+1F58aX8K6VPGJESK2FlgJZjqi0/CIajggAUI5RhWF9cHsovKGpQYMcQ6teNgKQezixHaI9iJ7aVvs8nXxRBSDH3n4W2QV0hdIOakPrLhNpr3ZUPsO4aNIRrLBDAgxJz54nbqi+w4kGPItfP1mx0INibmpF7GaDbBnsHjtbZmS+2oH6kPayEXvNs6wbpAwOzXeFxn9QJ8x1xlhDDc7Ypb19SLW9bswZfWl7eTDoHrDup3x5ArT653naPrfrfsS5c8sOxDp39lcc/L3zXruQe/NWubexzt8RUEEBuNkifpkyT508t0n0mxLGvpmPuCWUeueOu8k552QdulL93YPnDKpxe9/5ydnV+9+hfVkXWuNkrPADyPQWbjBx1AzG3t3oMTp0KEEfR5Ks/SPgHpmECj57V8zlCe9yXl0S5CvFftRJz854Xq/R5XnS+Lv0aA6Juy/+GP2hu5Zs9jeSZbHs+HgEbyzdryZ9qUur3D5sfGLn6m2sFvbkz6vDSd2En/vh38UZ2cjrM01FtexkzzgjLGwc9V9IFUn9MK+Q8fyuM/J8bW0d8i1JbWQWELPSe3df964RWvuClrm/NWWj8HEObzSkqSJEmSJEmSJEmS5FEsh2TNM06fd9KRny9/84rvLRnpcp1bulxxy1rXeevqemnzWlfa3OXK9AcvyDkQBHiltTwip9/k5BxIKJxUoBSklBJlABNzRsgxKSen47zeyDkNWBjclv4wpzRPzMFfgOgQSEibqcEA/bGOOs7TH/cNwQYHAlZGEON9iD+x0/64Lq+fktdrl9dwoFdQ3ggiG1c8plAX10MXghOASSgjpzg4UdJK6zmA4Xaw0bzW5eCJJeQVlOcAi/2hfeSjAdbvdPB+OR+Qq/e+qIzU24le5kHqfTu2NfsAq59CypFtfCKQfZI9Y8LyqIcfSvnkm+gYsX1DPu5/Wug95LHS/d8raM1JXkkBAq8XrG3sC0r55M4IUirTnoOe9wN+yRV9YdyTQ666cz1d1/qHy+P9vy1ODjxcvnPw99VvXHxn+/kv+Jts8fxX0D7H6614hT1Jkkeb7JUgblmx9Nj2c1+wqeMT525feuOl/7TsG2t/Ufzm2t+Uv9nz68rW3j/g9FqN95HsITyHPSEFfVSHfSlkHJ6jESlnerLhk3Dajn/sQcF7EXVsg/YEpPb8RsoQmxj+cy0HGR/2u9iJH/+ZQWV57uC5JOAyjRP19tnFwKm5hpNzopc0/1klessb+LMG14G0oS4HHoOOFeA+Y8T2WsaYFUbKxQSdB81L/nNdxsT96RzIZwg967cBlIcfuh68yiqvsa4hm3W/nP2iw66j5ZNOyiVJkiRJkiRJkiR5zMgCwsHZrFlnLTj12M+VtnY/uJj+uAU5V7qtyxVvW+OKt691hdvW1CtKzpWHhSTg11LxOiu+D2u8xxW3UdkAok7JBSHlKOUy6dGOykiNgPDBCv9RTnn8YY6ACoGT/qEuQZcEXmzvAw5CTIgB2iaQZBJMSPATAojgD3kECOJP9OEaPCnnfVMeY/DjIB3VMTGneQCBgw9gWCdpGIfq1U6CEbXDiSi2t7FTYMJkk5QDSUdlgrTRvIHbB5/sR4MchvrIBYAKq5NgCG2DTiB9xf5AHtpYfcp68wN9jGBv45f+Sad9yikXK2tqfihvxNz0r5WaXeSTT8dRHU7IMdTO/MbtfX8yh75/pL6uAWhrtrrObI1IEC6AzhMHSHkd2rqyPSMAQefzaA//1Fdpe2+9NLGuXt1JgerOIVfdvd5V79rgqndSeseAq912xT8uueSF78sWz3sV7fMK7/gkSf58sq9Tmguz1uzJLYX5x8xYvvgZ845/wmsWXvbS65Z+9JxdnZu79+Azozre72rbaF3T89L2UWm0q14a6arLd8XZXkIq4B9xUFvZO7JHkfd62k+eaGO95a0f3Xdsi/aSNhJznnCz/sk+6DBWqQs6bcegMkPL8OufR/pMUvCromQDgotT8iMwv4B+FtHY7XPOYH2EuQiwcYittPVjM1u6/mAn4+CxQK9jC2MlG+jpGWqvwiJlmF+Axxv6sjnmPjEP9KyWzxE9ga3PaR4n/p4YwWusXeSn+5dNB5e6aD3hR0JM9ncyM0mSJEmSJEmSJEmSP7vgJM3Ts/lzr2y/8ISt5eG+ny29ZbUr3gpijlICiLnyZiPmcGJOvmtuCjHHJ+cURrrRH83ypdqS5z/A8Ye01WvQ4/9AZ4iO/ziP9QgqGPYHvLS302oSGECPP+gRCEgd2+of+vzHvvqywCMOXHh8DL0G8s8Bg/etyI1DwfXB1gILX6f9S2AjY5Q2YWx8zdupDkBb03OQhtTsqGx1BLa1dgbzz1BbDm4E5kMILPNDAQ/60Ppg1wjzl4cnDLme0v21VzsZP+kZ4svqbbzxiUHzEfoT+CCW2gvJR/oYTMwBVG/kHNehT0u1PeZtb6TcXmBzzfBrQ9aJ3HcF1g6toamknEL3ChCIObJlP9QXxjVJ6Q4a807CLgpY7yDcOeiqdw25yu4NpFvvqt+45IH2Nz57fdbS8iza5/i+ufQKa5I/l+x97bXPe+L8k464YunGN35uyQ2n3bT0vWeNd37ign8ofvWKH8uvFffRM0BS2Y/yGQFSTgi4PlfjlPQeIOvIFnuGnrd5GDEnfuI6I+aEnLP2Cu5b8vI6a2jHJ9eivQwb2esC+Q8pQMr4IRghydSO8wre4/I88s+TBhixxcTYNMQc94nPL6Qe1Fb78M8t5BU2lvizMYxP5yAqs0+kKPu5wfhIhzEqpiPmZByRL5+PxoZnrj6zmZRjYF7kWc+kH11TcbiHnnkDruMz5/86Wzx/kFbUIllYXtJzL0mSJEmSJEmSJMmjThr/SMUXbD8za5+zZumFJ95S/trV/774a1e4wm1XMzFXun21K+PHILZ0TSHm+HTPeAzRTSHmUOZTdgFiTzZMnjWA/uBn0oLL+GNddHHQwSn/QU8BBGzwR7r+YS+BAekBq2OEYECCFiDywToqsx7+odc6jJN8hABN+m8MVix4sz45eLA6BB0ckGCMMk4hfJAaSO8RyCWvg60CPhDAWB3noSc/3r/aAJ6UQ0qBjZ1I4yCT2yPgQTv40z65TvLsR9uLjZWljn1qPb9yZIEU+1U784P2UX/m14/H69Svwuu5raSelCNbG7PNHc8fB3dRUMekHNlo0Bd+qRUp2mP+wvzGY5Z7JH5MF+xDG7vnfk1EZVu7vA50LcspVFl3WG8hL2nwJX3Ld+sRJmmed+DU3IAr7eyvl3YN1Mu7N7rqrk2ucucGV/7MeZOzjn78WtrjzyAsxYZPkuTPKPi8wTpcRmjNOtpWzjvlqAsWfexto53j6+rlEVrTYzgdt87VdI3LvtR9RHp5/ZR0hNoYlWkP1UaFiDMyLuRhR3tH9yc/N20v0b4CrCz7S/avPIOkPT9PuE5hOtjoOPLEHFLYEZB6Pfay1pm97zv4BPgZ0KALdSC35LPFn1hjf4D4B+Q5EoPqYUuQ6wAoT3q+BoyD2+bnBJDPR4PY8rWoLsxNVE+pXAPlCUbQcR6ArabyGStjk/FRivvOz2daB56YGwjPdbKpjg644tZut+LbG92CC1/2s6ylFcRcJyFJkiRJkiRJkiRJkke1TPe/x/iS5Gdni+etWXzRK24t39r9i8U3X+Y6b73KFUHQbV6Dk3N7ipvX1ksj3Q3EHKUxoAP4NVYAAYnkcWJA2uAPf02Z8BJ4Ik6DGA/WA/gjXvOkD4Qa/qjXVMFBCelgE+v3Cu4HKbWJ+mU/1if541MWGCuPOQQk3DYuK+KAhoMgBBwM1FOZSR+qY8IHba0egQkFIErEBNJJ/EsfIYDxeejJj/VveiB+LZRJOfItUD2CHfXl+1O99cEBEZdDcOT9A6STE24NxBxS9gVb1cEf2kZ6aSt92Hhy/tk+tLc6IelCObQn8OuugZgD4mvj4I/1KANyH0KZ6nwZ90ntta94rv2Y7J6TnsEBp+jCmlAd1jL2B683W3N5WMBqZKSfZ5pj/p69SQpYJ/vrIOfKuza4KqG0ayO16fpV+QNn7MyWtq+hPf5U3u1Jkvy5ZF7rE+c86+BzF5573HULLnxB//xLX3x92zVvvLXjy5c+iNdUl9PartE+q4EsV1IGe42ev3XaJ3V7BoCYgz2DntnVaYg5+c8d7EXsG9pDAO8lgZBmeLZHet6/pPdAW917DLHxYP9kp3XSJ9n4/q1M/VCZgTLstWz7O34m5vZ93B/XxeQW6XKATjAdIRfXM3QMflyWNkDGYjA/4jP3uQY95yNbyk8l5iJbtQl5ecbxZ4xfA4GY88/fccqPrHNF/CfLjj7Xeuwhv8iyliFaZYmYS5IkSZIkSZIkSfKYFfx64zHZwgXdSy95xbbyLVf+ZPEtV7gOAoi5jttX14vDOC1nJ+YoBbbJK62AkHM99Mc9yDsQcQR8cT2/wkOBCYg6JbwEGrjQH+W575TjQAGBTAS1t7aeLIOO/6hXUFkIOU1BoPm2BPIdn0DK6yWVPkUfAqS4jfrUPi34ygd0Usf2GoRIkKhltkFAk28j7VBPABEEGyZgkEfAIjrrOwQzptOUySWxER8RUCbICTT1yXUIeCwfQ/RCZiEfQX15n6Qzcg0pk0isj+1E78vaj9ip3xhqZ0E521Pqrx3zCn0MbtsQyDFQJ9fhiTn0TcFflU/OKeI8bKb4bygTwj2gOcd9Vsh9xTiDTshfLWPtQcd1YZ1J2VK5Tj4ZiP70epiYA3BybnLQVXYSdmygQHW9K09uoPKQW3nngJv1mmfeQvv7pbzTkyT588icOSccdsayT50/1nnHYL1zc88fOr951a9Km7t+XxnufhiEnH8GeGJOynaSFqjRfuITdbQfarQ3PCFHeyoP2YdGHDF0L4leIM982WvYf3Ed1+vey+mxz6Oy2Ije63iPG1BWqD/uj8q2v2WMgZwD+LkQlVlHfdizwXzJZ4tA9GQD37n+Qp2HjcnQWE/gk7u4Bi2HZ5n4lD4j/5gbnR+Ax6o6+UwOdQJqx/W4fgDPW4E8p+VZx/+5Qs+5yqSUq9sGXWFLl6vcv94tfv+ZLluy8IEsa7qM1lk6GZwkSZIkSZIkSZLkMSM4Odf4pcgLCcdkbXOvXHjKcz9W/OT5dy/96mW/XnrL1a6wtcsVR7pcYXRtHShv66I/qAnbugXj3fQHuqAM4g4n6/BDEUzMGTRIQfDDp87wRzmV+Y9y0Yc/1sWWg4JYz7b6x30Mrfd/+KuOAxtq732wz2AfQ/oSexmLtiN78RHbk577UduoT6SNXxrOQZWVORCxQErtrT0HJtLGE0OUt9NtgRCSNgz2Y3qp86SewcrmkyHBjwRAWmbE+UjH/echJJzWEaSsYyU9B9QMqWeMh7wE3WZP0HKuH70G8+XHH/VrbcXeUgriIj3fAwB51sGPpeSbyTmD9mHw/Uh5yvyyPr6Piui+xGsnR8zxGohOwShsLQtQD8h12DUFKDG3Y9CVtg/Ui+N9e8rbh2icQ662c8gtXf/ae7Lm5pNlmydJ8r+WR/K9XdWWUttxs559wBtbn/P4V8551ZFntg294fOdN6/9dXm4z1VHac/xc5L2wnZK+TSqrG163tZ5Xef2IbUh1ADaA/EvqvJ3xtGeygE6D2qPfUNthExCKnW8/zwxhed50EuKvStt4met7W1+pvi8tJF22OON+5zqonHkwHrs7+hZReDnAtuENvKaqPokhD4D/Gk57l/a+Tqkds2K/OebgD8PcQ2Rjj9zMCYer163jYvzUudherqH9hzj55sfT0jNlz1vy/Rcm0rMESYG6Z4PugKNu3jr1f8585jHj9CfMz205o4lzOPVlyRJkiRJkiRJkiTJY1QQbLURnpLNnPmyuc978lD5E2+/p5P+MF98+2rXOdbjOsa693SO9+zBL7AimCpv76E/tCk/3kOBAaVjIOYoHaE0R8rpCTr+I5+CJhBa/Mc90hj4A51SgpBhBthOD/ljnvKEQG5Euul85aD90dh8PfIMBCYxpB521o77QxtLSWfEXBy8CFBPKQcjOkY/ZrOXa2qEJ+aiQNWCJA7M1M78xD59ntpwH+aDAyDUm28JgmLiysgzsRF4Qo0hvuy0CxNzqjdSDuC2IOU0AEPfYq/XpjB73wb+UYe82XFe9QrpX8FlgdTjGuU6+TpUxyn8M8jeiDmvUxvYRrp4DmWcWqY5lr7C/fA6Xhu6hgyqgz9eV2Rra9iC2Dxk/J5I5bFRoMon5wYpYMVYBlxxrK9eGiO7UdgNuGXvfOO2bOH8V/AuT5Lkfy/7JeZmHrHixCVDr/vEsm9e+Q+LP3nBfUved87uwmcu+7fK5p4/1Eb0xxrw2YHnPj5LaM3Ls0P2lO0rv6eVmPN52NOekR9yUJBP+c8f2Veipz3gn7m632LQcz0A7VVPwN4NtmgfgfYjPwcUXCa9jSW0i6B7W+rVj0L2siH4C+2lbN/dJq+oWl0e8hlF4M8s6PK+WMcgnV6vfKbJfAW76BkVwZ5pAjyTqA0hfP4CohM9ge4bw8p6HTwG9otUQfV4vgkxByJOyLgK/+cDPesmNrgiraEV39roFva+7t5s/uwracmtJOAHrZqx/pIkSZIkSZIkSZIkeazJdEHWbMKz5zxtxbqOG86YKE4O/a59rNstowCqOEl/UE/QH+2ThAn6wx2BFQIaEG9KyFUo74GTc6gD9A99OQWgyP0xb0GLBEgSYFA+rvepBAasn86X6oRsk6DDynHwxSTcOI2N9RqI+L4lL+009W3hU8eGctSnBC6iC0QUAfkIMlbKU8Bi3w0XSB6Fby/BqhBZFqQCEriaPQdD8M3+Iz37D2XpR4JfboOAkIMhzbOd6BAwWz2XI9hY+eQcyvh+ObKNbTyMmIuuiduSD8be+mA7gSfbzJ510j/PA49H7II/u0YCl0VnfgSRPgfSUTsGyuxDywTrW2xx/6J55mvFWqA2ujZ4bWG/RGsQgarcM2lr68JggS2f9qMx2zXIfTJiDnnCOAWzIOVGyMdWWdNLbjh1OOtof7ls7SRJ/qiCX/zFaetStrj1kGzpggNajlj+3IWrX/mejm9e/dMircPK6DpXG6X9CIzQM3ZrV51/VRWfHbTmjejhPUvrGOS47UV5BoS9LXl67uEZiDa0b/K/oqo+GbYHFdhXpGfwHpR9yLbYj0pUwRb7kcdlvqgvQ25/m479qs7ANlGfEXw77Gdvb3ta+uDPAx4z/Avkcwn5yB+eKfp5JJ93QS+w9tBTytdJ/XlSTO0a2vnPPNMbMCa9H4B/TsE36hiis+cUn5hDXsuelCOEz1QqK+T+y7Otgv90mBxylR3r+fX86rYhVyQfy++9xs089gljtPZegIUYCd4GeCSnOpMkSZIkSZIkSZIkedTI3v6AxS/nPTlb2nby4re/9BPVicGfFO7c4Drpj+4iSDkl5opjXfXCiPwwRHl4rStv7XYVyleGkXYxMcen5iwQoj/AhZjTQILyAioTJLCCLelgizz/oU9g4gN/tMeBjf1RrzYG1UmgAp8oa+BC45DTBJKXH61AKnoZm/Qv40C74I/ttJ7raNzx99nlAjqC9I08jRmp1stYcT10HSB0/PVRsGkBGtvLtYoe0MAUZYb6ZETBDbdTHfsXH+ZLIAEwB4Scav0+YGSUQXQaVCsxx+QYyqwLyAeuZkftAfWV7wM2qmc7aSP5gHAt6AftYr1cF4NtVK++LYD0erUPdgK7R3yfVOfHDJ3C+pJrFNgaYGCNAXRvACbmciA99ctkHJXttIkEtZYK5BooeLVXhKnP8ghAPmgvYm0v/ptTh7PikkTMJfmfyP4IjoUtKxYdP/+Up1+19JqTP7ao/1V/u7DrlZ9c+uGz7y6PddeX07qt0XORT8iNIsV/4CgxBx0/35Rkw94F7BVW0sn6BrTMey9qQ+ud9yD8KOSZibzouay6sAflOR63888lg/WZg+47/U8G3o9I1a/0Jzpr4/s08Pi1Dv7IF9tzKnl7Dth1SFshsYzI4s8fLuPzSD5/LJXr0/6YADNb5AXBr9T55xKg/n0/UdmeT/ycBHisNDaMz+xUh+cqXyPZ8YlfLQsxR3YE61vGh/aYG5pfOzHHxNwg/b0xUK9MDOwpja51K+67wS35wPkuW7TwPlqDjc+2RMwlSZIkSZIkSZIk+QuUuTNfuPj0596yYud6V/3H692yHb1u2XiXK27vpj+28WMP+E65LlcaXusqIAO2rsWpCCp3OX6tlYk5ggUNSsIZPLFl0AAjDgQ4QNFgICZL4uBC2os9w/QECRjEt/nnYMDKGtwEO2rHgYLacl0EbgvfWvZ94xUtIRglSGsAroPHLflG2PUZ4RPqwvX6a/dQn5zm0eg3EFCWNuThp0HHBJfpETBpngNjhdhJYB2C6QBPzMEWsOALeu3Drpl9oi/A7NXWl6N6sac61Ulf4sOPjXXBBu2E2NI0ngOCnzdtz4EmSDcD9I0AaaB1AVpmPcHue3T/hZhG4Gr5PtpblFK/DK5TX1jvSMdovCAHkPo8jQFB9zCtxa09rrRZ9t6S95w2nFUTMZfkjyogPxa3HFo9uv2Kl27q+PJFD5Qm+l15gvbAeO/Dle09e6qT2A89tDfw7KcUoM+BAH1O0nqWvU9rWPei7Vfes1ynoH1jzzbe4wCeJdBjT5E/Bn+OoCywdrEuVxdBSLIIuf5lrzGBRmUPtZ3im+skzwQU8mofA9fdqPO20bgDeZXPW723w2cSpxHU3vzLtUqZ9QavF0zxhfEQGv9DIdggn69jcs7sALLjceB5hbyShSAOuS3PMeW3DzheV/zdckPyS607+1xtx3o380WHuyxrGcmy5pNkSSZJkiRJkiRJkiRJ8tgX/A/z3v6XeWHW3Hz67GcedOPS/pPvKt98yT8V7+v7fef2Na6wdTUTcZWtXQRKN691JQDE3JYuObVjr7f6QEyCMR8YcBAlgZToKI//8ff1CHAkNVhw5m0ICCDMD3xKQKHgoETrjWxTXbCDn6jMgYr5FJ33b/64vfUt/QopF+w4VXDwR4FJHLDF1+YDwAhShzza5/Wc9zrxJUC92RCUcOKg15NQEgT7fGMQrHVCbpmt2qANB84R+cUwH5bm7dg20nOdtpNAXPxYHSOqD20JbKd+uV7y7NP0KKO96riNjlGCRQkYvZ5Tm/MwbzyXSrxhPv394FSh9TLvBNzfyI/oVcd2AIJPy9M6QpnSIkBtjJTzwTB88rokjIY8/wolBbUVnJYbpjUZEXOL33facFZa8jLdy0mS7E0afxCoUWZn81sPaTmk+MJ5Jz/9vLZznn/VvCtOeOeiD75tR3G8r16b7He1nYQdtI8AWr/F0Z49xZHuPXhu4scecGIuR8oBtK5lD2FfBsizBKA6JmtoreN5GcF/zxz7QL09o+X5K58naAsflCdf5sfyYhMgfWE80i+XYeshz438c1ftpvhCqnomofYDJabks4jA/WmeYORagH7OsG2wmw7WRsoYb8j7Ol8fwz7jpCzPIh2n9mlknNmFa9Dr4usQTKkDMDejpCeE5xqtH7Iv0NqiNvUafol1pNcd9N3r3YJLX/5gNnv217Nsxrm0Lp8gyzNJkiRJkiRJkiRJkr9sAWG3gPDsbMHcCxe8+Kmf7/zI236w8lvrXfE++sN5eI0r3bLalW+52pVvW+MKt15dL9y2ul7aAoLOiDkKwDgYQxBjwZj8UW9//ANCaIXACn/II+iRP+pR3gvIVgICIdCmAARaA+R1nxjSXgIRtfM+0YfUxWPmgMbXC/gakY99x3W4Jg1W/DwoJNAToklIIakPiOZCfTK0zp80I8jpLvFnOoMnwxhahl0DPBmm7UQvAbMn0gDz4W0Ccjamx8kT9R3GESBtKQCmvLw2RTqD9o8822Mc7DvKMxrbwI/5srzeh0Y99z9NPi5H9ySul/uj90iRs4ttPeRexSfkQMyBoEPwGwJgggXQU4g5yo9QPYi54R5X2tLjiuN4lRXEXHsi5pLsS/CM3zcx1z67Ouukw9666PpTPtNx0yX/svTGi/6j44sX/Wvl1qt+Wt3eu4dJuZ2073bQ2ubviaN1PdpdZ0IOoOcgw8oMnHzT02+0/3L7mWD7Jfeco9Q/Yy0PvSLsN4DaKvyzdxr4Z0BOL6T9VJCeT86RDYP8qm8/zhygnw5Uhz0b65i803baX+w3/uyyvP/sma7/Bp18niEvvn0djUM+y7SMedR8zn8M8mH+QLbJNYR2AtHzNfg2oR3XUconffE8M/B46PlHdYWxvnplrH9PaaTX1e6/3lVvXuuaq0u/TivyRAK+Dze9tpokSZIkSZIkSZLkL1oQqDUGayuzrPmtzcWlH2571TNu63zPaf9U23L1T8oTa35THl7jCt+4inCFK952tStuWSMn5+x1VvzIAk6y4YRcIyhImwIOsiT44T/wKW+khQUHuTr9Yz9PlImvOIixPI/JgDLXa9/sw8riq7HMOg4goA86Dg65TnwCPnhDIKLjluBP7e1aFT4oBTHnyTmURW8BqrQLbbmdtyFQPgbXEYy0MgKLy+pjr0BbzmvwzKD2gPnR+thOypTnPrSOiTnxZ/3HsHHyNagvBtchj+8eQl6/0w79mx2PRfyavT8ZB3i99qH1aMvz5O0UpJc+orGZDc299EtQeyPZ4nsTr43cvYQt6hm0ZkjnyTnOkw5680HBqhBzVKYA1p+UQ4qANibmCHjdfMkNpw1nS+YnYi5JLPsiM+YSqllr9vhs/syDZlSXPG3W01e+YNarn37W/MFTvrDsa6t/hh9yWT65jk/HgYir7qR1OdlXL23vreM74mQfY0/aqTYl5QDkuUx6Luse4v1gbXWvQc97Q/YH76EI9pz2dZR6wF/U3rdTW7EPNuGZIHsdedub4bMF7UgXEXM2JhmDlANQzoPbYP9iH8OG97Ts6/Brq7AL4zN//rML1635MC6D2MuY0D5/zZ5EU79xW6lDO9EH32Jv82PPU+iFaJO8+WHgmUQ6QWgTiDmyx7OL8o3EXGmUrgvgOpprnKSb6HMH3ftuN/tlz/hOls24kNYofnjEZP/EcpIkSZIkSZIkSZIkj1HBH7qNAdwcAl4deQ7hRU2LFl4y//gnf6HzulP+Y9WWLrfyjiEmE5ZtXuM6b13tClvWuuJot+sc7ap3jnfV8Yuu/D1Ehm0UmBH897KN6qtOyCuEKJM//P0f9AwNAJCPAgkjwzjw0SDGp3E99WW/GIuyD3RQR7YS8GkbDykLSUcp7FnXaJcfCwcfrKdABAEOp7h+0nEASCn0Bg0qhUyDXayHbR5C4KmNwnyyf0Vc78kuyhvxJIFXA3JtxNbbQ0c2EliH9igLOWZ6C7S1jUcchA8oYhu0pZR9Blv8yIT9ap+MRcfANvl+AQkKzZfpxa/Vo63cC/gJdvH1hv7D9Um/5gvzTPPO9yPcN79e4/tmtpEOZBz0RkDbaTkj5jiA5+CVyghqKa0SkHpiboTWHMg5CmyLO/vc4vWvG85mzkzEXJJY9k5izMpWznpq+awFb3vue9r7XvGhhetf+6m2gVO+vHDTabcv/cxF/0TrcE91ktbbDnqu4XTcTlqLO/vx3K8DOC2HX1blPUhrl181pedjAPYFAcSNPRtRjp5hts/8/mHQ+tZ9JM9See56Yg4+vD2V4Qv7ydrrM5f3IcBtonrqT/ak7PW4bdiHoU2uL4XpGOgvBxm3J9NJx+NgQs6gOrO3PlhH9YB+5vDnFGD66cZGep+fDtaWIHl97iBVxPb2TOTnqdnouMM14JowNrlOuWdoL+1K4/LjD+ybIXNSpucVfz8mfrmX8oXh7jo/16iuc6TL1e4e+q8l575sd9PcuZtolR5DiNdwIuaSJEmSJEmSJEmS/FVLO+HNzYsX3jT/eU/5j8JVr/zlypuucAfetQnfN7SnMNpd76A/qjvGelznBP0xvnPAlXfQH+jAJAI4ITc4YMIf8/QHuf8+OuQ1+MIf7znAnoMySfM6KmvQYgGMJ8/sdB50+4HZgzSUflVH8H7NzgN2eVsLUqyOg1AOVBR6/aa3QIrBdQbUkw9cpyLUo73moSNIkKn+Ae9HwIEWAwGwkE3SVvwK8aQBsrXnVOzFVmCkmNk1ElbiK9IxtF/rg8sg5qJ67huQtkaSMUDMcRv1oe14PTXC+0NZ/cV1aK96gOeW7YNfq7PrkHGonlJpE90bhrZVWF2A3J9GGDHH3zVHQBCL17qKtI6KfKKE2hoxRynIuSrpUAYxVxqRduU7+928t79omPZoIuaSTCcgNJolyzJ77vMPPWXp+9/ytY5t3b/qHKb1tLmP1hbW+iATbhU8s/EfLJP0XJukPEg6wH5N1VImpuW0XI6YIx3D67TMOmqne4XJcM2H52qcx/NV89TWnmfyXJWy7CfxI3oC+YjJKCGNBExuadn6lrbqj8ZktnF9AK5D87gW7YOvkcdOAGGV00vefMqYAoLPSM+fO1GZIMScgP2wL1yrIH9d4te3pTHZnNizx18r5pDzBJx8Yz9op/Oon4NCyulnIsDPKIDstf+YlPOADWyHaS0waGwE/CcDTtoVNq921e8MOlqT/9TU2X4NrdFnEtoIM7BgkyRJkiRJkiRJkiT5a5H9/W/0IYS3EHqzptahWSuWfXzRiU+/rbD+dd9ZubX7t6u+dZ0r3bPRde4ccEUK7EoTA66yY1CAPP3xXxzprvMf9TjtY8BJNvwvOv+hr3/kWyBBgQHIMRAVTJJBz8GC5tVeAgcBt1UbtkN/nKoPb699wp5g34ln7VAnPoONr8uVp+o5ECNfMTFnASwHTGqTD6DUztqrDw7AOAA2W7X3OrLBiazoJEqM0L8QUExwcZmCabIXCGElxJmUpb3Ym18hq2BnKXwZpB62VjbSi0+8KbxPrhebuH+0NRseB7eRE3NcRj9ox3Wa57L507z5y9mYT7WJ7X0+zG0Yj9oQfECseYHWmw/VS/ArqdwHsQ0n5lAn5FpRA1gm5Riy/nBqTk6U0HgosBVijuq3dtVhV9i5ztXuGHAtz33KNtqb+D6mJH+dMt3JZ5OFWfvso2cdXnvj3OMPPWXBqUdf2N7/us8u/dyl/1ba3OuqI/2uhvVF68/2GUDPMnpe9+zhPcREHen1P1mq/PzAXqAUoOefvK4qp6P5hDTrBOF5Rmuf87JX/N5UPe8RGgc/H9kuet6iDnsHtuZLIc8D2aum4+c826E99lvYc0wiqR33j7KHElMEG5uB7TGuCDIWjBVl9YG9izrYeD3s1I/WxfB9mC7y4a8Bfrg/pAFCgImN9cHj43YCng8D2qAtrhUpz6vmeawE6t9/pvLzKHwmcgqijT+7UQ9IOxkj/A9QHqC/B8bINz3HykrIVbfQ+LbSuiBdYaTbVXbjWdvtZjzzwLv1xx7wfbdJkiRJkiRJkiRJkiRpEHwnUSehg7CEcCThbU3L2j+98OVHfK/yrre4A0f73UH/+G634l/e6cr3bHDFSfqjHMES/vinP9w7t3bV8V10+LEI/sGIYfqDfLjLVfCHOX48Av97jj/2fQBBAQBSBAGU4kvu+URRpIvrJZDJk2ViI6kFKNMDxJyRcwT4Yn/7Rhz4CBAQaWAUIRdMxfasN8R1OhYCB3ZMGAniVyODb8kzKMjyedSjzBBSCsGwIbaLg3Ih2TTYVYS6aaDtp7OV74oDuSYIr7IKuF9A2/IYfduojsepZUJj22nzHrieCHrNjEiHeRTigdqgz9je8hGEFG3QsR8NlpGSz33BTssxsJYZEvzilAmfKsW+0DUJYrs43FUv0nWV7l7vqrdc7bJlS7dlWXMi5v56Ze/EXLn9yYvOfk5f6RPnjnd88rz7Oj719n8p3nTZz6q3rf5djdZTjZ43NXreMInGxJus/RpDy0bKxcQcrVt+NlF7D36GEswf1Xs7QnjG0V7Bnopg+0YAW23HvmFDafTsi+1l72G/KrnkEexsv+XrFeTf8vJcolR1nPf7mtrz+OFHwddt4zSQLcYc+c21Vch//ETkI7cjvcL0PD9UZ88U9hXNRfw8MZ2/JvYpbT0RF+nD+KQcnkUA/BD02WPfDwfw98N5sk4QfKKfAQF//hP0pC/+Q67CxBz+g4Ha09paed81bu7pz3fZjDm3NWetr6NVm15ZTZIkSZIkSZIkSZJEZX+n6FZlWfNphPdkC+Z/c+6zHn/3otOfv3PZptfuLN140T+U71z3n8V7Bv6AX54sbOlyRWDrWlfe3OVKlJaG15Judb0cE3MIVPBHPgUrBiHYepWYQ4pTdNDl7djWBzrmBwGD1mmQg6DFAqgQFEkbCSqtrrHedAQOfhC0RTqGBnIeCKpgK3UxuB9fF/epOpwWUbKQQYGYEHMIfMxG+rGUwUEq8hpgWplgJBUTVdqnbwc9gnBNLdA1WMAebNQPwDr0F9lyX4a9E3NAaKt+G3zwGNCn2Vmf0FsbtQ3tIh3bK5D3EHvTYx6tf28btWHfVkeYlpgD4IvuMQfXlI8DZwbpi6QXoAxbBMIAlS0o5iCY1gHyrKe1PEK+Kcgt7hh0pfuvcYuvfbPLHtfyzax55kt0Xyb565V5WWt2SMuy+c/MOhcdOuPxpaPa3vq8q5d+8Kzt5a09rkbP2hqtoRo9S2vj9NzdJsAJt/J4Vx2vrwoJ1+9qtDaBsJ8UXMbeovVuzxCsc31G2uur9mus/tlFqT23PNEGH7Z3KM97CanZKfh5aTompIKN6MwPCCGB7dnYTtoBqBd4HfWB1D9fuKztdFwyFkDbcDvqP4KUzT6yo7bymaU+sJf9Z5PMHe9zaivPA7WjdkL0aXv44TFJnYH/E0Dbej37QFuDXm+sw0k3A5W9jxhWj1Nv/pmE8euYNGVbm9ftIOb0fuDUHNXJa/r4Dzmy39LDz7JVD1zv5l/40h9kc+fenjXNvoTW8BNlKXvZ20nQJEmSJEmSJEmSJEn+KgR/EO/rj+J5hIOyrOWYLGt+GeE1hJdn8+eeNvOwFdctuvhFY5VbL/71ir8fctX7B+mP8x5XvG2NK31ztSvfvtoVN6+uE/bwyTkQczgZRH/sS3CigYoGLfI/80rMUd6/6sfQNooQ+EQwHQUHFuwheAz9CCSoDHXBhoB+zI59IAjRAIz9iu84WLK+JEgiew5cDOhH4Mdg9qzX/q29teMgV/uK26sd13EgGYJTHocGr1OCam+jekCD3OrEAAOvI3OeSanIrwWssAXQp762ykQc50Xv28a2AAVt4s9sJG/wZADGbuPyfcIW+caxi23cVvoINqIXSJ8yvzk7QOuCL7VnTDPvBLtXFqxy8KoQIo5S0gtQ1mBYg2NrG9Y4/MCO1gNIufFB13nHUL1y10Y34+XP/G3W1PoZ2ofPl22Z5C9U9k9QLJh5QNurn3FV8X1v+dqSd59246L+139h2fvO3Fn88uU/WkFraDmtHznJRs+WCPxMoHUrr6haXvaNJ+g4JR2DyrRO/fOLn43yfJRnXnRajnyEfaH7hGD7hols1ENvdthXpmPbKFUbAfzDj9QJsD/DHmX/3JfmWW82DcScgp/ngI2R6v1YcL1mq21tbAKbA7M3Mkz2vYGJLcvr3IXPKtibP2kbl/nUnL8W6Kie7eQ5YuS/PVfk5C7qQZKJHdtSf3sDxiL/QRDBiDlfT2A/wafMP31mAJRHGX0XxnvrbDtCa2OznJo/6IF3uY7rTndZ+8JRfYW1RpjJazlJkiRJkiRJkiRJkiT7lP0FiC2EY7NZc9bNOmz5aNvbnvOvnR9+82+Wb73y98Xtq39ZGln9m85br3y4eNvVrrR5tSsPd9Ef+/iuOfyRDyIKf9wjEJEgggMMJuIikB3SXIDAekq5LYIECo5Iz7+OFwUOFnQ0Em2Sh07AfcQ67g8ge7ThOgm+9k7Mid4HTpbHWJBHO/jxNjoWBMy+D9F7H9rW101jw/0iaGO/4pvzPogLY5sWHPBRAG6EnMIHtDk7yXvSCifijJzT4IzBdQIO/JHCB88BfGh9nIcN/CgpGPqwPsWP2JKOyhasAjlijstR3TT+ZK4E5s/7QR9aNsSnVrwP6NUHAlHcGwmSw/1BarocMYc6TfFddLL+RMfBNfobx/c4DrnCXRse7vz0Bf+ZLV48kTXNWUt77jDefUn+0gXPX/yYAwgMfMUA/oME0jr3FYe9ueNT5+0ojQ266siAW0FrZiXtnxW0llbQmltO6ye8hirrVcg22Rt4dTVAX2dlmwC2pfUoz5/wXLTnlRFyAuwz9KP7jPO6zqk+7AfRM2ATg/W2x+M9SP1p3nx4X9gn0BNk3whM532QP9+Orwcgve45P6YI+eetto39EPwzjf1onfeve5/a2/MBsPmzfH5+LBXwOBj5Z5K1Ef967THUP4+BPzP/Z+D23gf5xeerfsaWMBd47hNKDHznLOaAxgobvM66tdut2r3+oeJ7z8Jr+P+cZbN6af1WZBl72T8RnSRJkiRJkiRJkiTJX6ngj+VH8gfz0ixreW6WNZ1DWJO1zd7Yemihf84LDx1Y0vWKr1Y3X/Xj2r2D9Ac8/ud8Lb7MXgK87YQJ+uOdggz+H3YNJgQg4xAIgDDrkTQKFqYScxSsUCAwHTFnsCBFTjggT/0zqB79oR7jojyfYKM8n+qzdugDoHGaD/GDOkpZL3U+sNK6ELyhTbAT25iY02tBnbYJ/eu4AK6P+uTAE32ajlIOWFFGP2pHsMDc6ziIBpQUi0k5gtkzMab9eBLMXldFW4UE1ATYaDvLc1u6Ju9Tr9EINyPmKnjFzsaAdtq39Ws6XJ8FqkaocT9almtAWdr5tgSeK52j0EfeByB22lduLtG3zC/A6wcp1Vle1oHUITDPE3MSwDKhALCt6BDcliahH3KFO9a7yp0bHpp/5gsmaZ+tJxxHe64gWy/JX4Ds7xm7iH/Q4bCVZyx487O75p9x9CVzT3vuZW09b/xC4atrfrp825BbsX3QLadnaQ2vpU70uRqtS0NlrLsOQm4qQL6ZHfKmC+C9SmtWnk30LMQzUcuybxXUXvYn9oXuMyr7vYZ9AiCvOg/YsS32E/LYo7p3vZ3VSfs84JtSAhNgSLksfYZ22Fvij5+naEfXINcmY+O6GL4uqudrk+tjHXyxH6k3/wYek/ZlvtiOy9Je/It93L9/fvHYNe+vR2zt2cJ9qI77hS/APgNhgz7p85Hz+ExkiI5fXUUdbPEfZ76dtUVZ680HxkXPbBByhW199cL2vrqf5+Euet71ulX3bHDLek/+l2xJ2zfo2bWG1vMRsqxzkr5nLkmSJEmSJEmSJEmS/6W0EuYTFhOWEUAazCEsae6Yf2b765+9o/qp890BFDge+K1rXfXuDfxrrvwLlRRglCb7XSf9QV+kgLJIgZQPMCgQkNddewhCmAlphkBI6iVYRKrBhemhs8AC8PYCIVgindqLL/EvfVi/lEewgXZRfQwOiFCPYJBtCajTvrwtl9UXIYxH2kiwJtdiNmzHr6JR3vwQZFwC36f5joJSQILnQFBJ8JxHXm/BdkMbBIim07wPpEGs4bVWq/dthMAT0gvQcdJ1ClFn+gZ/rFNbrWd/uCavJ1hf1lbt7TqM6IvbStDeMEe+rZQ54M/Nr6WWD/MvawzBKoJWKQdoIBvl+QcdABqTP2lCaXmCAt0dBPzq8Y4h13n/JszRb5sOqX06y5pPoH2FvYaTqkn+MmRfxNycGeXFRy54y3O6Oz55zkjHeM8vlt289pfLvnjlD4rf6PpZebTvIT7ppj/WUFFg3+ZOxDFhR9Cy7AU56QZyzgg2rGv+vrgGxM+b8PyxPUCp7SX1ITqp5/1BdULMCbg953WPAWxn+SjdK6S974vbin9PhqmN+c6Nz/R2PVyGDyGWUGd6hvYhkGeL901zwc9ts+V+BTwW8sPPaoDbxFAd+4C96rUP3xdfn9bF7ZBq/9aH789/plm9PHvishBzsJM8k26j9FnMP/hAvrQuPMvUP/Ut48J/ItDzilDYtq6OMRbxozWkX37v9W7Zpjf8LFvYdiM9u/BL7+k/FJIkSZIkSZIkSZIk+SPLIzlNd2CWNa1t7lg6PP95TxpdevZxt5Tfd/ruFVu7/q16/6aHyndtcp0UMHbSH/md+IOeAismNrb1u8Jobz0fXACBmJPgw4i5PDwhF514awQHT5xXHwartzpPzGmg5us0j+AIZQ2SOHBD0KJ6s5UASsvIc720Ed/UxoIl9cX6bdQ/I7T1gZ/3BZ31S8E0gmSMw0NIKyOnAkkF4MSb5L1eCSoE+J7gYp2C9MGO2rBOYD/8YL4Ypott4V/LrItSBn8fHXRGGgQbHhePzSA6sYWd6tEP7EDMKWxO4tM0gFwP5RVCJoT7kL8nMt9eT/eMA2qkdE+EGNBTnYxIz/mImAOo79LEgCsjwN0x5Eq7BMU7NrriP1zrCp+98PfZzNnvov20SrZVkr9AmUEoZq3Zk7K2OUfMfGLphLkvO/wtCy5+af+iD5w92nFb70PlEaz/QVcj2L6lZ0Cd1mGd97yuzZiUW450QqE63itkL+Sc7Rta97RGc6Qc+QI8CUc6e4YxtD9PzJFfv4eQJ50n5QCy5f2jbbF3rI3UY08hr7ByBNj7fYm2Ecw/IHvQ2oW+OFWwXq8p+I+JObVDu6gf8aN23B6p2ArgTxCeAdoPwP4Bs9PnA4A86jAOD5SBfFtPPmr/3j+gn2X+M0UxhZhjoG+ps1NxTMzpyTn4iq+D7XSM9PyqF7b310uT64WcG+1xpTvW1cu3Xv2D9jNfNJG1Lfg/WdZ8Bq3txtdXcUounZRLkiRJkiRJkiRJkuR/KY+EmMPpnqcQXkw4lvC8luKiyxa+6sivFj961s9W3b3RLf+nG1zxgWtd544BCgYQCAzQH/70B/7Yurr97z2TbEgJ8mtxEjwgSODAwv+PPlL8T7/9b3+st7y2YWiQSLYMr5fgKgbbxmWksEWAErUJgZPqtJ2Ra75PrgttoRdo31qOX2HlevaNVP2yDdpIOwu42QfrACG1hKhS8CujINAiEo1sAyJba886qlOYf7ZXH3Y6Ltcmhvep4xqXfsM4Nc/EnPqGDdsJxI+MQXSaZ0gd+7C8kXIRMeeBOrLLtSPY/Foe8y55pKqP7p8EybrGKI91iFdXDaKHnaxFEHNMzuFVsIlBVwaYlNvgCjuH6sUdQ/XSXde40v3XuzmXvnwPBbfvpv2DL0xP8tiVfT0z21sKC1684KSn9i5d9+q/6/zYuVsX/925uxZ99O33dH7l6gcrY/11I7dln9Aeoz1Rwz4Z763LXjDgtVQ5JbecsS561VVB9niN1b4fTp5JjdD9YClBnj+APYsItpd0/wQiaWpd7IOfX/ysgp36i9qhzAQQ9yX1MQkY20ka7PBclLYBU8bBe1evVXXTwffB0L6jsuW9vdfLcwDg8Rio3p4DorNUfFm9XF8g54RsVB/cRsCfGfT5Jf+5JJ9/MTEn9498NELb5nwZMUewz1wB6fVz135lGs+u8lj/HhrrntLkIP8HW/W+a1zl1st/Ov+Vh38za2ldQ8+tV9DaBimH70qMBXvhkfwNkSRJkiRJkiRJkiRJkv+h7OsP7lVZ1nTB4zqXfH3+iYd9a9mm1/+w/I0rfkYB4oOVicGflEb7f1Mcw/+897vSCIIBBAX433sKNjwsCEHwIcEEIN8xF+oCMQeEvAUkPigDyB4pdPHJvEDOILixcqRnOwt+RB8HaTk7roO9ttE6rmc01DOQV2JO7W387M/7EghhJnZTgDomoAL4JBnnw4m2QFSFfA7ky8gwC3SFFNsPKReB++HrC/6mjFX1flxKzplNjoywsdD4pC1SBb/eZ2VpK7YEXD/qeR7Uzveh8832lLc2XEfQ+8DBNKdyP8KaE8TfLSdBt5JyAIJbJeXKO9a78q4NrnzHJlfevcFV7r2G1v0mlx1+8M+yplnrae+UZAsl+QsQEBX4GoDHZbOzcutTK69YcsWLryveeOHdnWM9v8Wr/libtR2DtDaFPKe1Uy+N9dVl/9BapfVUA8iWvxOOoWQb6bGe/Wk5+DKQLyHm1A/7wnME7STltU8+wjNIYOtdnlWSt30BP7a3hFTK19u+4XZcp887wOw9QB7BBhA7JvGozhNzvr7BDnnYWF4xZRzxtZnNPmD+JW/I95HXy75HP/YsYB3V++cB62FDdUi1ffgBC5qHeC7YLjxPuA3An4WEmJhjHY0d4Gea+dFU2+Jz0j5Dg476APBZO0I+8XoqffYWyHeB+jFirraNnlvjeIZtdOX7b3i442Pnf7f1WYd8MctaLqG1fThhFiGWRMYlSZIkSZIkSZIkSfInkn398Y1fGKQ/2JtPzbKmC7PW1tUt1aVXzjxi5UULznz++4ufu/hb1b9/J36J0hVG17nC5m5X2NpNQUEPBQkABQZbevYUR7rqTMRpoIIgxE6+WVASAp/pgEBE2nLgwoGJts0RcxbYTAdpz8GPBmi5QJZ0YoNgR2z9eFlHYBvLW1l9sb6xndiHOkmtL6uL25ovCX4lKDPSSk7MDVBeCDUhvrROy0LISaBvOrERnxzYc31MylEdTvNYnhHK7Cci5TioZx/q3/LoA7bQ0bXKd9KpL9TBDu3J1q7J2sr1UToBO81zX1anetRPS8zRfPG9IDAZoKnWyT2BDa0zspFTc7q2uCzrr0hlBtsJKVekufKk3OQgv8Ja3rWeCTkm5u5e72r3X+sWf/jtLps993tZC/8aayfvoCSPBdkfATEjWzj3KbOPOej17Ze+dNPiG069cfFH33ZHxzeu/gUTyTtofRHKO2hNTRJofdL6qpfGeuvY53bSDajRGpMfbBBiTV5D1bocGTcVZufb0Tq1Zwmve5QZpCNwGfW8F2Tty16EL9070Cv8XvJ5EE1xvfjybdQfnlOy/9AWfQU7TlHHbWifAWrr/Zo92ufqoLNro3wEeVYCYh+uQ/Rsx/4ohc764P6kDWCfMfBn47LPIvGvdQbYUD3gnyG+b+gEYoO2VMbnm31mqU5SAtcJzC/IOD/3SP2YYlCbnP9e/r640kgPv6IKdI4TQM6NdtWLeF2fPqMrd11H/jb9ou0dJ3w5W9Z+Hq3tlxGeTliIhd4g6fXVJEmSJEmSJEmSJEkeBYI/zEHOzSMsILSpLstmtxw965iDPrbs/ac9VN58lavcM+iK9wy44nYKBPCLrlu6XGm4x5W29tb9ybn49BwQlYVw01TBARHpOGhhwMZAtjHINoDqKZjxgRnAQQzyFhQh6FF9VG9gP1Fb3z4an/XjAzgF2nDet0EZdgLxG6eWl7LYIWgGqI5TkFFCzDUSUsj7QNvrBLAVSJ51OfvQhok0vSYB1anO2nh4vzG0HXyon+ns5VqQUn0D/PVyO22v9nmorR+rzrkCOplH1em94ECWUgmyAbGxgDognJIzxMRcSYm56h2Eeze55RNDv2s5+Zk/zrKWW5qbZ51JO2Qp75MkjwXZFzE3I1u64IB5b3zG29vf95ZvdNy69tfFETzX1j5UHu99uLS9ty6kHGGyX0DrMybSeF/TeguEWox4/VM7Jec4r2W/5r0/SWPf/JzROtsX8qyRspA8KMv+sz3m841l5Bt8GLyObcRuOvC+U3veY9t1v0GXA+xo/JRneB/RNaGdolEXjy20b5iPHIz0AqhvRsP+5+eE5OX5IHZSRymgeq6DP+8zsoEP+nxCnp8/+nnFn2eo93lqB3B7kHFyHyxf5tegkaqd+rLPx9JYjyuCnFNSjn/UgYm6bte5eXW9uLPXVb+9ntfKvNOftz2bPfs1ur4h+IxvfH01SZIkSZIkSZIkSZIkf2YB+ba/UyR47eX12dw5H5/55NrNc0464ottV5xwy7IPnDHR8Y1L/qlzdM2vi8PdrnR7rytvpUAWQQIIOgtEIuTJNkop8ACsPkfMoY5SBCcckHIbTalsQSqCKAnOBKxjvQY1lLIt6mBDeWtr+dBe+0VqeQP8qY+cL0sVMl5A8+aLofYM0fEplCiQlB9EEFIOyLdHkD5dgKwBvbVRcJntVWe2NC4h5kyHMvRi731pXvyRjeq5jV2ntYct0tjOX5cgEBNa5jZRO+4PddI+5wP9cCrzx/c1nkfWYc5pHXlI4MwBtNrg1IsE5UrIAUbK4cceGCDmhlxppxJzuza4KmH5A9e70v85/z+yZe1fyZpm4BTKYQSQ2UkefbK/51pb1tp6SFZacFTrMatOmnXy086ce/pzr1rY//qbOm666ud4HVVep6b1RQARR2unXppYV69M0v7UPcqkGq9h7Ataj7TemIij5xSDy1jbtr7RRk7NyV4XUm46Yi5GeL6ENW/7QkgeKQtpZIQPdEiprKnsec1zWdvG4HbBL/syfZwqjHBD30zKaV7QmKfxM6x9dD3wFaFRnxu3bx9B+5E0skPZng94BvBzQUHtLOW6nJ1C9VOIOfZBerrPHnyP6PMPKZUNopc24RrDfbF7xnPNxBygfZhvEHLk23/X3FZ85tLnLU6wjxHuHXLlO4fckve88SctzzxwlD7e8YzCr7JPJ9gf+9sjSZIkSZIkSZIkSZIk+RPII/nDHP/DXiU8g3AU4YhszswTZxxcumjB6cd8pvB/zvn+8ruHXPUuCgq2UdAw3EPokldcKYiwgMJ/xw6CDaRKuoVygD8ZBxvWiV3jiTluq0FOCNBQRymDbKBDHnr82AO/EhsggS7spqYe1p/mDebX7CQf9W12WofxcaCtwWIAAnLSU3AuP4YQga+J4K/JIP7YJ9rCj9lScCdBvtZ5vdR5G0ZcF0H9MeGgYzHiwI87toW+wXfOB4gHvTb2YW3Vv/eteiEkUCdlnmML8rksOp4Hmg+bawuk5Z6onlIOvOP21JcFxmUm5vqZmJNfYsVpOcJOWtN3rHfVnRv5xNzy+29wbZeecD/tgSsI9qMP6TWwR6fs+z8dFs17wpxXHH522/pXf7D9o2ePt3/k/AeWfuS8+0tfvOzfK8O9D+FHGMoTBLyyOklrcJLW5+SAq9H6AJhUozXjX1cFaK0J9Pmiz7LG03J5gOALRDwQ9kYM7Pew5xnYG1jPEWQPqg/W6Tr37bCndF+hnnS2T3K+oYvrzd50mhcd9pjAE1jWLtKbzvoO/UtdY7++7FO5LrseGQO15dT0cv25a6Y66UvGIs+EhpShJB3lkXIflhp8XwTUAf6zjHzw55X4k+dSHn5sdA/wzMGzx76ygL9mwOw4FT/yYxWYw3Ug4Or4oaUKSLnN3VTf4zpxgv3OftfxxYvdvHNe8nC2qO3ztMJfSWjntZ4kSZIkSZIkSZIkSZLHrCCo3RfpsCprnvGWWYcu/9iic46/r/zxc39cvuny/1v6xlXfL9y6+ifLbl/7e/4OnBG8ehO+bF8CFAle7ARcXMYpEyFbonpAgxVfhzbwRYGVBXABopO+0MZ0aBeTc/BBiIIz6UPHaHXatjGIQ3Bmfm0clrJObc3GxhfIOQTIAj6ZE5U9tI2NKyDyF9tFOgYCvikwO6nfKzlHYN8K8+kDa7NrLCuMFBSCzdpqyvp8nSBuF/xOCYoZOq/+nmCNyb0QiI2ckENb0pkP6ouDY+qfSTlAiTn8IisTc3eAmNvgqjvXu+qdG92KXde5luc+YQet/RNlC3h5JCdPk/x5BP+xgF+ixsmhRdns2ZVZT6s9b96bj7pg4aY3fGXJly5/sDiK9ThAawXrgZ4PSsoVt3XXAykH4KSbnnajNeTXKK2xQMoB8nwJZV3THrbWp0e8hxjkY4oOIF/2vWq2L/wehx9CvF9CO9QhDXXhORXs4v2S64vrIjTst5AP5RwxB3C7APRtdeHZGtdrvqGdHw+B/dK1GykXdFF7AP4N6JeuXcg4gxB0wUbR2HdswyBf0edVDtzG7olA8oGY86fkuD3GAUJOSTnqmz+XRgkjlN855Aq7h37fMbzmp+0fOfPf5p9z3PdaDqz9Z5bN2JJlrW/Awo9kf5/lSZIkSZIkSZIkSZIkeZTK/v6Yn004lPByMjunadHCi2ccVDyz7UVPXd0x+Lqv1sZ6f1K7ayN/qX7HSK/rGMUrOPjCfQp+KS2M9NaLI931mGxCUGLkHBN0DDl5gmDFCCkEu5aXYJKAwMUHjtCJT/arNhIkCjGX18Em2HuiB3WwVVh/PihTHxL0iZ5tUKf1wQdSsQU4CEegxnmBBcwS8CJoE1sJtuFb/EofNA9oC73a+etgTA30eUz6mpT4bKwDME7zQVAfvh99vc8IBj9m8hX6F9+BiAh+xB71pFf/Mn7RCaa7NhoX+9Gyz8s98ONH3uw15UCb8nyah9pAD0KOUypPS8ztFGKuODmwBz8AUb37Wrf8pi6XdSx6IMuaX807IEgi5v68gvnf27OqNVs8+8g5xz3p7La3vWDNgitOvK5t3SmfWXTNW27v+NzF/1Ie7dkj61PWB747k1/FpDVeHO/Z49f7ZEC87mwvSpmeEV6vzwtfR+B2gO4pLmPd635gXdgH9syInxt+n1Db8KyTtcz+uT3B8gTeB7w3InB75HVs2PN+D6len6fx+K1OXldF/9QuAuqg92Q4bFVn/cr4ZJyN4+Br5WvGeCQv/at/tee9DFh7gsyvXLuNy/vVPOvtWrkMKPnFIN+a5zaaSv/hGWfPG+5P6+1exaSe1fGY9H7wqWjGQMD2AbInHX02VvCaKn5kaaSvzp+ZkqfPvQG34s5r3PL7rnOVyV638P1n/OfMk57+tWzZom56Jp1JW+DCrHnGm2jNHyRL30si5pIkSZIkSZIkSZIk+QsV/LGP0ygzCK1QqHTOqC0+b8mZzx+rfO7C36/aMeRW3L3JlXeuF2JuuN+VhilIoWCjNEaBzigFJEhB2CGYQWDiSbkIFKxwimCHgh9OEQxZ0KXBmQSSUidAvYBfZY3KAvgQxO2sHicWAjknOrbhcZht1N70DNhqPxq0xUGmB41bSCvJ721MgPRPiObA13EfViZ/GqQyuF4g7bTe2nPdNGNTMFEWBZWeWLBgk3yFcUvZIOSc2Ek706le+wiIxsF9kE8mApCKf8uLXdy3zHeYRw22KS/BvOZ9vZJyAF0XE3OTg64I4Dvmdqx3pd3r3fK7r3Gd3W92tNzvzJpbT9G1nuTRLp3znrDwzUdfsuzDZ29e8qVLvt9x+5rfVMYG6Xmybk9ptO9h/KIqryOsB+wBWiP8Wqrq7PVyfnVV131lAuszgE/FWZ6hBB3vUQLaqD/5NWJA9kMA6bBfkKd1Cdh+leeO6GRPkY7sheiJ2rFfQMvxvtBr82WGjNn6M+TrJe/7NfB+jMcFqI7AZJyBxxrasi3ZSCp1Uhb73Nzh2pFyO+z7sPcbkZsD7iv4k7EqdD4ljcZKc8TPQWurfYdxy3wG0lF0Mjb4A6ydllGvp+Jk7JQqGSf5QQHl7fOvSp+PlRHyMbLOFSiP/8wq37PJrfjW9W7laJdb+p7T6/Nf88xfNhcWD9MKfzuhwmtdBJ/F6QcekiRJkiRJkiRJkiT5CxaQcY/kZNBTsmzGxU2Vxe+bd+wTPrj07S/8cvn9Z+1cflv3v1fv3lgvbR90BQo6hKCj4GS0n193LYx21YvD3XX8eARQIVRHKdAFcGKAyTkJnvjHCwALinxwZjoLmBAkCbHGr7FS2QK/EEQpuI7aaB6w14nK+I48teP+tH8L4qQvgQRscV79oS23z9sD8kon5XEdVDbE1yj5qL33R7aUCpCHHa7fELVXyCusoS6+tr3BSDUh50hHY/U67Ss3dt8u2E0H3wePG9cU2vJ8WBBPYL/Uxsg50YkP6ZeuA3q007KcgkEwTXWm8/UE8scn5+i6jJxjYm6H/jorvkNxpO/hGUc+9T+zphkfy7KW43WdJ/nTy76eQbOz2TOObD2k+IZ5Rx/8mnkvftJrFpx7XF/7pjd+s/NrV/wU5EiNYKSwETlYL8XRnj28nmltCAlNa45Bdjglh/XuSTWsRwXtGSPmAuRZ4/cn90NrnfdNyHNZ68Sv5ql/v5do3cYQvdQFe00J+WcgtYEt2qE9jc2vec7jGsxW/fh66KK9xP5DXewr+BCffJpO87Gdt2efmBfJy/4mu6idzS2Q84FrxDVxOwXdlzLNZRgrge3VDyPoZC5D6k/HEfwccH0M6MRGUmnj5x9jMhu15++MI719jyCTc0rIybWDoCMbPG/G+/j0OD7jlm+mNXR7v6uNbXS1e97pqvdd54o3XfLztotfcl/r01d8LZsz7330+XpDls18B615+qzdq4CgeySf10mSJEmSJEmSJEmSJPkLlDbCKsLBhBVZS8txrQcW1iw+/QWby1+88Jcr79rgKt+6xhXvGOTXdIrDfZT2usJI157iSFe9PNbDPx5RJgg5R2BijoLdMUop8JFfFQ2Blw+UpgB6EHJCylmgxeQP+4hgOgLbkT0HbSDmtB23Vb8MlNlegjXTxfYM0sf9x0FcbgyEqbZ7A2yt37g9YAFj5JvaAPGvRPo2Oh4LMvM6BQfnU7+snoE2XC9tBcjHiHwZcvaNoGuMYIE8B9/an1yn2nMd7MTW5oZPuPC1wI8QdPYKHBNyBgrwmZQjFCYHXWFiiNPSvRvcso2n/nM2c/aNWdPMs2lN4xXuJH8e2ftrq7NnVxe86Tnd5S9cvLvjcxd/e8kHz9q97JPnf7f4lUt/VB7p+n1lgvYxsJ2eIwxbK2HNMGlH68MTdLQm8iSarjOsN9QD2CMKayvPB9jo+qQ1xb/kqr4EQtbIHoghvhnx/mNEdjyeOC82TFzFdt6PjQllPD+iMl+X+PGEF+s0D5++v3wbb0e6RuR9q53Z0pwLYjvUyX6351xu3A227NcTpuobPuHf2ijEL+1xmguD2cUI9g16nFxDm1w7mWuB1IU+MVc6Z3S/QR7aM8aTidCN99VLo931ynAPfc71uBp+ZOSe692Knde58leudIv7XudaDz/w21nTjOtplZ9AwA/P4PN1JQGftXuTRMolSZIkSZIkSZIkSfJXJggC9vU9NodnWcsVLSuWfaH95GdMdN7wpn8t3Xjh98tbrv734rbu/0LgWBrudpWthNFuJuXKmlZGjJgDKPj15JwFbQIEySEoMuRtjEjziG2jcj44jIg59gMbrY/KwX8eHFzCh5KDopMgTgLP0C9gfi3Qk1T85ODtpvMBUPBvATGXpZ2dlhNiDqB6HU/sQ9oIuMy+YmKO7gXD+lF4v3mwjfmK0Hjf4nLuGpU8sQBc+orGB7CN2MnYlZhDnudM8kLG4aQcoOWImGNsH3KFXUOuuHuoPuuEI26nNXw+4RDCAizoJH8S2Ru5gGfN4mwWkxPLsjlzinNee9RZ7R85Z6JzC93TrX2uNtrrarSmsZ5AxuEHHXCaqzDeW5d1hL1g60WAk3KemCMbPlnHJFq8xqP1lgP5sLaa57XPbXA6KiLmSCf+piHnsA+n2YsCs6M8t9eyzzfaBR3vIV9PwJ43Hc+HAM8NTk1H5ZiAEl9mH2xlv+UhNmantlZPcy1EGtVHtrbf7Tkn8xjZRLbm0wCfsJ86BvVJsO+Qi0/KNcL6y+lwAq6BnMvNc+6+RXqMkYk4er5MItW1QHW1sXW0Tmmt0udDdfcQYeCh0s2X/GzJu0/7wfxTj/le86rOO7Js5mY5Iccndfe1HxIRlyRJkiRJkiRJkiTJX7nsj5hbRDiCgP/xf102e/aFrYeWuxe88aiPdX7wtL9fde+gqzww6Mq7KFga7abAusuVCaXhHlce6aVyd7003FX3xBy+i4eCoKloCJAsjwBPUw72OA0BFoMDNannAI7RSKjFIH1U5oCQdQS8+mqgsn+1DX5gp31KPuhlnAYdr9ZJSn7MzsaNYNSCUwsE46CQ+1HAJ+nlpA/VcR+ik6BbwF9EboEoyhwAh/oA8qOwPjyhofWB8JOy6GT8gJEUMbwdX7e25+uUaw3zZ5D5MUhbAtkbMRdeZaU89THltThPzPUTBl1xO+Gu9a68uctlpc7PUGD8Alq7Kfj90wmeJ3t7pjS3rOp83rwzj+maf8VLrp178Uvf1bbxtM2dN17x/1VHaQ3x84GeE7r2jARCHt8rh9TWLa9XSsMewvqJibkI3kbBazWs17idbw8bAgg5EDIMlIFtRnSTjv3QeqQxy3Mpv0fQr5HpMl7R4bp8X34cYhfKAVP3lwK+WAe/ujfMD3xzf9LWnge8h7hOfJlfq+PrYZ+qRxm2BtShvfYhtrpPeR7kXsj+hq30IXbqz8q4x7Blfwrk2a9eI4Gf6/xcDX4Cgl6uV/PaD48Tz0UuS55/UIfzVE+QdSX31vuge1+e7Eff9fJYf706RvXDfa5GddV7NrjafRvoM261W3L9qb+a/YIn7cja2j5Ia/xKWuavy7LWV9Kz53lULvPKn14SMZckSZIkSZIkSZIkSZL/tuBXXWsUeLx6xoplH1v8jhf8e+GL57nlE90UqAzWS3etqxfG19aLW9a64pYuClrs1VYKvPAjEHxqzoIgAQdiCOQ4QJIAzgOBGKBtASHUKGiKwbbII0AL4OBQfXhfSraZTvINpBxB6iwleL82PtEJZKwW5AXENmo3LSQQZCDP0Dr4ifK4Xl9mRLYA96s+OMDUfINfCUKjdgCPgWxy/ci4+dr5ekRvZFwMqzNgrnLBfFTn6xlhbhkg5qjNlGCc+vB5LouutB2/ykp12wddaQfhrg1u8XveuiebMeP9tF7xanaSP4+AdAjEw6r2J8678AX9bZ+98NtLvtn1u87NPXt4rY7RfQQa9ks4wRbWTPzsCPWRHa9htadUXmuN9NDl1qu0jUk5gdoo5JXWQMgFHwJ7Ftk4/V5jnY4NYH+ydq0vq8Oa9nYNCOM16J6BLyoL2aZjUYRrtj7FRqB1U8Zm7dUvQ/sBUDZfcb3PC/DsC/ta6qU/1Etbfx3qW8px/1F7zC/r9gEdf7gG64d0DLXD2BTyvKT5BaiOSVQQruyH7jOVmSweobpJyt+3wVXvGqgXbrnCtW98g2s97ok/z+YsmMyylmsIL6JVvlQWe5IkSZIkSZIkSZIkSZL8z2V/p+ggywmvJbOhpvaFH591xMr3zz/5GR9asu5VNxc/f+G3qzv7f1Pe0e+Kwz2uuKXblUZ6KAjCKTQ5+WKnXwAmhyjwqVHwZeCAm4MmCp5AxoEoM1JO08bXmphM4gAOOgH3QWUGk20Yg/jyethHhJwHt1F/yMNnLoAkXZzXAI99cllThZEJdt3Q+bFzMKiwINLINfjUNmEspqNAEfDtSMdjsHJeL6dEtJ6A8UjAG/sUSF+SSlukwc6ICe4b84A89FrPPgCds3CtDfUMvZcAlfHqIsP0Zkd9mH2sw0m6wrZ19eJ4f71055Cr7hx0c097gaP1+SFap/hup1jSCZU/nuSJt0Zpm3P4jGcfdNqcEw87Y9YJTzl1ztnHDS284S3jHVu6Hy7R+sO9LTLkx1qKY931eC3m9gvZ+X1E65HroFO9rAnKx2uN8vYaqqxR1WGtErhMvgBbu1xnewap2jIZlyPnoION2MXEUdh/4lueAWpL7fgZ4sciOtZr+9wcsI2OgXyGsZEdTpuhPefh1/wIrI8YXMd2qmsYW34cQDSv3H6qXXyq1fui+yLPkKiO+6O2PL7I1o8l1HmoD7M1si0339rOg8cJO+sHemnPPqKxhXWlawvXNd5Xx8lknPSuAbvWu+X3XetW3rfRddx0yR8Wnv/i77U8oTqePa71y/SM2ZhlM85XUq5DFv60kk7HJUmSJEmSJEmSJEmS5I8mCC7mEBCE4MusDyQUCE9sXjz/9PnHPeljhevf9K8HbOtxB957javdtcGVJ/CLrhR4j6xzhVF8aXbfHn5liEEBEAVZ+F4pJuYoUOITCiDPPECmgSwT0kx+dTUO+kKg1QgfvAMg2iJ/3u808HXoE22pDw7sOIikPMCBnqYIwHOwegK1kQA8wAgrKSOYDMGjEGKUjxBIsuA3kAlIEbSbLg7iBeaz0a/5CsgHsXuFjtsICk8e4DoM8OXz0i4mHSywxr2c8kMPUX0jcnMLHfkvjvXVQc6V71nvyjdf5ZpW1VzWPOsjtC7Tibn/d4Jnwd5I/IVzTz36wkWffMdk+xcu+f6ij5737Y7PXfz9wjeu/FlhtHsPE3J0nwt873tkP1NZSCDc3wDb33Lfo1T1sm4awHuU1kcO0EX6aA3F6zfsG91PWL8NEFu0hR2B95OQQV6nvqeu4dgmgNd2A7jvxrKO3T8/UKfjYt+al+vM561vPxbS5yE2qBPCDdB+GMGf9xGBbXI6ac/QttKH2tH98H3buPhZAFAdQ2ztOr0/g/kkhPERtB73xp6N7IPvFfJiI0Qv2uCk3AB+XbyO75FbuWuDO+CB69yKeza60pcvcUsue6VrOWjlj7Js5s207C/T1+SLhMWKmYQkSZIkSZIkSZIkSZIk+X8i+z4ZEwTfp/OGbMG8985/1kE3Fy586d3FG958b/WTF9xVuXXt98vbB35f2jboyiMUAI0NUnBEeU/OARokjfXxD0fgFdgywISaEGRIpxJzgJYJvk7Bgbwi+EGdEm/QNcBO1xkxB/8hONSUdaizsoKugfvVMp8ARJsoaM0RBDG0TQgeBXaN0id85iFthVRgUGAqQb0GqDymgEb//23oWK2fPFCnfVNZIPYSqMs18DXRGBrvF+rjUzg5NIxZdBSUk76wa8B13DHwUHv/a3+ZZc0TTTNmX0zrscorM8n/S2kmVLIlc47Iyu1PyjraVs5+0ZPesGDjG27u2NLDhHx5hO4R7WWcjituA5Bfh192rst9x/2UdZGHrHleM2QnpFzQW128rljn91e0/nIIBI+00bWL1NYu6WPSOfgSSJ20FR/wh3wM0xvQ71QbqzN7wBNJsU7HjmsAmcXXwnXiK5BTWheB7c2O2jCYGAOkHjq2pTQmx8WH+o5sBDr3bBPrI2h7G5/vr6FvBq0Few5wnnR+DthOxs/j47biVzCggA0BpxeBMUD9qU/U43OnRp9Dy0epzSjZT1D+zk1u1QPXueU3X/HgsqHX3DP/1UfuaKp23k/LfCzLZvxdU9PMC2i9P5lX/vTySD8vkyRJkiRJkiRJkiRJkuS/JY8k0MB30OEE3XMIL6NA5g1Z2+zXznpC8YzF577gEytuXfPTg771bnfAve9ytclNFAgNuvLYgLMvfgdRVBrprvN30vGvuva6Cr6jjk+5CYQ4Q4BFZQUCLZy4YoKH6oS4U/KOg0b4VnuQbvAFwk0JuADRs02ksyDRgsP4u6/QN9ugzLYa9EGvdoE4EzvGdMQc2lm99avwRBrl/a/bAjweDeIpGDViTAgD84nxqD3Gp7B69uHbauDdqOeAPmqjY45JjDxEZyShXZ+kMm4/Vwy7VugpVT3bxfDXIUA7fEdZiYLrjrs2uo5vXv2j5iMP/mqWNeHXWA8jpF9j/d/JIyEZZs5+zoGnLNl48qcW/c2bbp6/+qRPtm045ZaOz138f/n745hc73ElQkzMMfmq91tIYkr5niOVeysQcpzvN+wUuTVsem3buG7MLuhA4gQ9A+uZ179C121MkAV/YiN7zdqYnZbhnyBEktbrmIMftQd5pO1sz4hfs4kQj5FtG+qsLfePvqMxaL2RYfYDKzERzmOOoXrxEfWFefbPsTDvYRxIFay3eZC6vC/oNa/32t9TnjOp5zHwNeFHGWT8YWywkf/o4VRJOSbnsA75s4M+F0Z66zgVh/8EWj6yzq0YXudWbRtyq+68wa28/2/cqu1DrvR357lZzzv0m9nMmefR+n4hfZadnGWtBD4l9yRCOxb+XiQRc0mSJEmSJEmSJEmSJPmTy/6CkBlZa+spc5/zxNurN5zpDrilx63YNvjryh0bf1eeHHy4OLyuXtlKAdMwBU6jAUzQjSsxRwG9EDdIA+w1uAAj5QgU0PvXsbgtQYMzhvoIgaD106DngBNAgEipDzS1DP8A2QayyewRuKONgE+IqH0ucG3w6W14DKq38cTwbYRAYKKAA1ODjEWIDfgItnFbQyMZsFeCYMqYA4QU0P6tvdrL9YudzJOOT8cYdID5yNsZaVrme0/5MXxf2aAr3X2NW/rut3w3mzv7SlpzeL06yR9X8NrqDEIrASQ8y4wnlJ+xoOfkj3fiOyVH1tJ96ab7RPdmW7crjncRul1pDJD7JfuS7h/fT7p/fE/za5mBvWTrnOuRin4K2F7rKS++Zd2IP23PeVmX+fWra43AxBhA9Ua82X4JPs3W1rvVQ68+oSOASPI6RfCR1wO+z2naMbg/1Cm4HNeJD98/6YW4inQKI+X2SsypLgb70PmEjfm0+RVdPCbSq420l7q8f9KhDojuK9trXTwmGbMRixFBhzYMux6ywzOHdPxZQc99/McL/tOnRuu1uq2vvvzuoYdX7l7/0PKvrf5d56a3/q7t5KNd1rbo72m5X0rLu01WeZIkSZIkSZIkSZIkSZI8umV/PxYBOYjwtiyb+Z6m0pLr5h598IbFF7/0K+UvXPS9FfducLU7cdKBAigKlgA+MTeKk2EUwI/g1xsRXAESWAm5hkBLgi0E/EyocT4mADTo1PZy8g4+KEUZoHoGl6mOIeXcqRwLDAk+2NS25p8JAQ0OpR1dA9uKvbQnG5+qH/YV4Mcb989lgdQptD6QbQhkBT5IZVv4gJ0QD9ZOgt88ptPF2BdxIESFjsHyen1+/pCHPZdpfDRPcsLR+pb2HLDDHnY8t3J/5P53u8JIV70yTHaTQ27FHde4+W87/rtZ1nyWLLkk/03Z30mfedncGU9tPaT8qgWvefqFc8845vKZr3zG2+ee98J3Lf3b875dAaFOe48J0209rghQvnO0qy6kHO6x7lMm5jSv69PuvSfVWG8pdJGe89DLWrI9EWwiO1732Cdap2vL1qD4MJ1A9ogQXGE9S1/iGz6jdrrGbW0zeN3S2ozaCnRdM7AfASnzvrKU20nZbAG/h7Re2ulY1X5aYg720LEPmguCkFr6jGQ79Yt23FbnMFcvbaVe7RT5/kwv9uaTr4P64JNvbK/XQfOJV0vlpBvZ4LSb1aGdthX/0XOdU9IB9FxnHa0rrLfCWHe9gB8bgj/8Z88wCDnCjgFXfWCjW37XgOv8zLk/ajvtmOHmVYUPZi2tA1k2451Z0+yrKH2Grvt9STodlyRJkiRJkiRJkiRJkseM4ITNMkKFMJfzs2a8ae4zD/l64bo3Ply7+XJ3wO71bsV9G1xlFwVYCKK2EkYpoB/uqgsRQwEXUj1R58F1Cgv8ffCPwA1lCdh8AO/zEuQZjGgziJ5sNKi0QLPRzvw2+pd6CSgtqJTAlPQ5XxKgcx0Hmeoj8i+w64v6IPDrrbm+4gDXdCFwD3b/fUjgr5jOVxRA+4DadHzNgrhOgu3pIMG7tNdrpvvMJNBotytibWxd50o717tVw4Ou6akHfpcC6rNpfSX578u+SIambPG8x8971dPPa7/hdZ9d8vkL/nHpjVf+fNnnLvmPwlcufbB4y5pf87rcTjDSTYk3JuN4f4a9Ea9dr8P9BUDkRGuA101cH+cbwXuH2jBQFp2tzdivB9pQajaArG+kIS91wWfwK3vBr2feExFQ520MYhv7l70EmF/YxG0EPOZYB99sq/ZcltTgCS61M7JM5tqg88B+1U7rhLjD84WgdTZGI/ysvYwvgrYVSL201fH41PJ5hLFLXkB+9B7I/QAoT2uM/1OGsc4VR3rrheFuekbQ+hvuddUdA652/0ZX2dnvlt50sZt34Ut+07yyNJ5lLVfRGj+CMI+wlIATt/MJSZIkSZIkSZIkSZIkSfKYl32dJgBJ9/asafbnWw+o3Nr+mmdv7eg7eaL80TN3VW6+/F8q2/t+U5lAcIVX5LpdESSdogyQXsg5CspiUi4mBjgYhD4K3ghGCnjSysoIPHMgew0wOfjkQBNtGuy8PwH6yttIPz4oZV9kB79qw8E1j43qGn2oX39a0NervzjvdRTMGjlHZSbUOMBtsPvvgMYrRIJByIVgQ/5xbZTna0WdgnV0zXw9KLMeeRm7tclDgnD2zddPoHsuxBxOV3a70jDpdg25yofPdVnrnB1Zc+sbZXkl2Yfs76QPCIrl2bzWx89YvvgZ81586Ovnv+25V83fcMqNi790yYNlmvPaGN1//e4uPu24vZf26NqHi2NddexDfg2dT9Bhveq6xf3DffQIa5mha4Nh9x3rIKpnG+wRzUPvT33x3rF22of5aATsFH6NKeK1bXnvx8bi+5G94Mete8NIMtNJ3mC2sX+10X4Y1lbb8H5A6nXal7dHPt+/kVneN7en8TPgk1KG+QiwOnlVVOcyVx/68WNgiF6eb4aoDz8eGmuUnwrpw98zLYuegHsXoTDaXcda5OfECPW3hT4vxmgN7qR+dvX9vvDlS364aNPrvjP7xCPuyJYs/lqWzfhklrVeTenTZdlPK/hxk33tlSRJkiRJkiRJkiRJkiR5zMocwhOyrOU4So8nnJC1tr525srCOe1vOvbjtc+d/x8H3LPBLb9nkytP9rvO0R7XubULp6SYqCvpianiiBABQsr1UNDWy5DTOwgGQRIgqKS8D6oNCOhC3geRGlBKEKxlDna1nv1aO/KtmEoQBhuGtSNYoAp4EkER2kT+APIf2/nxAtbGlynoH6eAFKBg9r9DzOWCfdNx+1inNmTriQCCBPsKC/79vAWEayfo/Pp2BHnNLtLh2pSYY0KW1kCF1gJ0tYkhN+9Nx/6eYuivU6D9KiyuJPsUvH6+91fQ22aumnXswW9sv/Tl6xddf+rnFn3g7IklHzz7zo7PXvy94pbu3+KL83E/5PVUeU0V67041l3n/UdrlYH9qHm7f2GPUJ7TqGx7gteD3HdZK1LH/UyDRmKO9VSebr0FGykLOTfVXtakEWdhP9hYYluzicH7Qvccl6M8E3GR75iYs2tmsB10uq84r+P2QF+SzxFl3FZsra3XsV7nTMvmY2o72Mg1C7ROxwpIG+1b++H+MA7UK2TeSWf/YUDIP1MAaROeVZKG8ZBOU7sGrAu8tlrc2rOH/6MG31cK0viOIVe+e4Mrbl3tll1/6s9mH3Po5mz23PW0wk8nHEufPc+i9CmEJYS9yf5I7CRJkiRJkiRJkiRJkiR5zMn+gpzFWdZ8yuyDq1/ouPDEn6387DvqK26++qe14Z6flbZ2/apjy5qHlm5eU+8c7nKFEZyIoCCZAjMmtwgI9CTwtSBX6yM0BvYCBHoKDfo8sTRBfiYoQGS/0g/Xe4IhRkQ2gJzjYFTrrB1BxovxEaiOgWsBedHoh08bKcwWga1vLz6snxD0Khmg8Pb7gRAJDbq9thdyIQTkmDuFlTFvZo98pOd7xDr4VxIiB7HBPOJe8xzwicluV9na5WoTg6788Qt+mJWW4ZW0Llo/R8kySvIIBOQcXjNfSJjFKLc/ee6rjjq9ff2bP7n0M5f9Uwm/mrxtyC3fPuBquE+6dkHC+e+MA7bp6The81o2XQxPzBHMlkH3lveG3m/uS2B7Rl5RVLtovXMd14ey+bE1Z2tN9qPmDazDdYU+rZ0nmxrrIhjBJiSSAjpKua6xzEBbsc0Tc0jVN9kBPEa9BuhR9jaw93aah57HLuB2auPngjCFmNN6BvrRuumJObVRO/FLY/BtUKfl+JqQp/nOwdcJ+BnuIfMVxh8g48CPv9C1jNLaoLXF45jo+0NxW/evO75y4W8XrT/1oVnHPOmfs1nzb6blfjWtcXx3HH7AZDpJJFySJEmSJEmSJEmSJEn+KgRkwP6Cn4OzrPn1ZLqupW3ewKwnlFa3v/rp7+nof+3tta9d8cMVd6x3pYl+VxjrcQW8soRfcOUgsN8Vx3r3UIBeR95+rU8CQBBWAQjE5YcjJKDn4I5sOcCnQFRSaq+QwBP+QpAqJ4H0RJAShFY2AiOGnKJDXVRPbbidkU4RMSd1SlqwjfpmPQWnCFpxfQxrB7+4Xqrj03KoI1u6Ppx04kAXZW3H16F5IREsGFadluNA2sgN9qttJaCnMs+R6MPcQa8+FRa0y32DndgwtuHHQBB4x8E3YUzJObzOih8JGe12KyeHXPt5L5ygNbOWgO+J2tfpl79GwV7b2+m4x2WzZxw1+8gVF7Sd8syr2858/lULL33Fu9v6T7mx4+Pv+PvKcP9Dte1DrjYxIGuD7w3uGb9ivqc40l03MpnXZbw+/ToXeAIOa1zzoS6sXduzIIPCWqE6gpFvTMDBBuuQ7bWeyoDoBGFNCXBii09t6dqy6+H1TH55fcdttR3WbkwUhTKBywTUKdBW9DR3KLNe+gplA/pCHY2J2kwZM9UL6YWyjk0hdVJv7f2+gc8IvC/Zj7UJsD4t7/cn6v386ryjb6vT/7zgPD2TDeJLfHhwW9JhjhnUFnNOvuV65Brs2QbYHEpf4rM41lcvjfbuqY5ReRSkHPofdJXdG/DKquv8+Ft/PO/0Z401r1r28ayp5V20/C/JstY36Am5RbL0p5VH8tmUJEmSJEmSJEmSJEmSJH/xgsBoJmEBASQLTvLghMMRTZ1tly9+7VFbax96208O2j7kDrzverf8no2uvHPQlShgU/KmLoElgkAL+oUQAHFQZVAgqOB6DjxhHxFzFAjGxFwILMXeSDNO1TeTc8j7YBM2U8HBaKOO2+s4ze8+fIRXVIUcsEA9QANcEA6RPmdPfiQwtro4IKYygmLug8DEhbbzgK22I1srIzUSwMPbNPjgeptjCbwZ7EPnnK6Dwd9pRuMFMbel19W297qVIz2u6fBVn6P1cQwhyVTBfpqebJiTFdvfdNQV1c9fdGfx1rU/7vjS5T8qfn31Lyu3df+qurX3d5XRvoeZWJqg+89kCwgSvZe4d7wmdE1i3era9SQc9p+Sb1jHXu8h6y+fl3vuv3+Q1wXpqT2f8IINUl0bto5sX/o1Az2vLYXaydiljDUX9AT0Ha9x1DOwH/YC8iX7BmSyEMqyH6Rd3JZ1vj+0sTrVmU0O4iPsC702TRmRHXyKbajzdpH9FDTa5GBzq/NL9ja+2AefKjbweNA2jJWvWZ9JMTEHQOdfvWdIu8oE1lwg5ni+8f2Gw7315fQ8qN1B6/PuIbd825ArfuYy13blK9zjVpUm6COjl1Y4viLhQF7p8pmCtIWQJEmSJEmSJEmSJEmSJEka5JGeUsCJqIuz9nnvm//cQ7/YeeGJo8X3nr69+LXLvlPc3f/L0j1DrjQ5oCdnEMjLq3T4AnD+EnACyAMQc/jieiOmOODkkx8g5MieoeQcBYf+JB3bNxILlILgY1JOQXoJviUA5Tx0ihCMEpSMyxNzWkd9TiXm4EtgpAADgbH2aUG/BL+xHm0EXCZ/FhRLHXypDYJgC4RRRkCNNh5apn69LcZAKfRCzGk5ggX73g/rkUpeTuoMcMqnp6hfPuGE0zG4Z7h3w4TNPW75jkFX+cwFLps7529pXazgFRJk79+h9tcrq1oWzT1+1kGl41oOrT57wRue9Y5l7zz9tsow3cPRXlej+a7RPagyIUZzjFe4cU/GevbQOqkzScf1dD/8PQRJjP1EbeL1T/uCUyPmeK3R+ib4V2DJltc8l3Vf4J4DPAb0jzL0Wq8I/at9DLTBGHVNcUp62xe8N1Dm6wDUT26/oJ2uayrbnrA2ub0H2204Gaen47hd6C/MmSH0IXsm2Ip/rUee4Ek5hl2ftPF5tRWd5mFPfgz8DEMebUyv7WO7qdD5B7gv6a9xPPGJOT9eao9rkfZIRcfw95NsDLCn9cBELq0B/hoBJeZA6uFHR5bTM75215Cr3T3oOm86/4dta0+cnPWip96WlTtHs6z1M4Qrs6zlWFrv+AXwvcneCeskSZIkSZIkSZIkSZIkSbJXWUx4POEwwrOyrPmk5o62s9pOeOoHix85/bsr7htwy7+7yVXvH3LVSQrkQMSNdLvyKGGMMI5USAJ/IoMhgb6Rc3LSB6nCAsgxCmwjAs3gTwvFxIT5jvMxoN8PPCEwLYwQ0DIFroxc4CuI+7X2sY4R2Um9BPlBh/60PQXcnmxAwOxttQw92nG96XRMCvMbYLaAEHMMkHJMzFE7AhNAm7vrIOdqExvcwitOdLQWKBjPnkSI5a8x6N4X2dA86+kHnFp415u/3PG58ycWvfMtty153zm7i1+67IfLxwZcjeZa7i3di4b7VhrvreNeg7QT4k5s/b3D2jGAbKO1C8RrmXW6n4TgJmA/USpEnejMZ9h/UZ78M7xebAWhLcZupJAQvQD0VNa6cI0RsAZzgL8oT+3l2vV0nLaTvQDSMhCXtle4jeUNWM/ql/eV2bB/aYO6PLFl12u64INJsZwNgf2I7RTA3veh5ZwNCDxDpGefCjsZ56H9Ic9EGlICj1F8+ucRQDq/PvSZaWuAiVrYsw+akzEQcn1uxY5Bt/K+Da6G/4AZudItevepbsbRh4xkM2deSOv7uCxrOV5fV8VnxDLCvk7HJVIuSZIkSZIkSZIkSZIkSR6h7ItsgOB111c2LWn/UNtrjvyHwoff9KvSl875QfX2K39Qnuz+RXGi5w+lcfkVV5BzTAJQ0CfBKwV+HEha8El6nNTh+jhQVIBIABmHXwXFDxAAOWIuIuc08EQ7zuegfaBur9BgPQeQdULYhcBcfDbq+HriIFhhvmLd9Gj0r2QDAmWaNyYQptgSoOc65BVq5wN8LcdgndkjsKf2npijwFyIObr2Uep/Sy+PYcXmda75mCf+IcuaPkJr4Km8GpJAmglthCKhnTCj5bDyc9o3nPzJjgl8P9w6VxsfcMsBmmPka3biyxNMEcmkqJFOYLpoLfFam2Yd21onCOHTy4RPvK+mI+bE3taL5OUEZSjzmvGI2tLYmNThsuQF1IbXp66zadYo63zeykHH181lkE9CQMlc5OdMbAWc1/ayH6me1rXsKYXasj3aE+JrDHn4DsidQItt2E++rQfpJR/mpNHGvmczpzefDLSlfpSYw8m2MM9hbhjW3p5HOdB9o31tz1b8iAPyFXxvHK1T1NUmaZ3eNbRn+bbeX3Z87twfzrvqZT+ccewhv81mzL+P9j6+W7LxtGyj7O8zJEmSJEmSJEmSJEmSJEmS7EMeSVC1MsuaT6D0QrzG1LSs7dx5zztkcHH/SVuq42t/u+L+9a5y76Ar7ex1xW09rjjeQ8EjBYOTCEDX1Yvb8GMRUfDJQSTVa9AogSMCRgSOPQI+iSfkXI6Y03YWiHoSkBAH30KQyck9fhWQy7CDvUDqSId6jAVQOw7+NejlMttSgM+ncdQWbX291AmpIG1jEiEGkwNmZzrKC2mAemun9dSPBO9UNhvOi40F/9D74F79SjuBD+Q1qBdijvL6GivfDw7Ye/GdZ/Xaxy/Fa6z/lmUzN9G9fwKvhL98eST7YUZWbHvhgtc+c92iq17+rgXnH79uwepX/d3Sj5//9yDgVtD8LlcSTk7KUV5T1Af0K9ZxChvT2RqM1098L2PIOsSesL0hkHUT6WzNMlCmeoK8lg4gL2Wxj2xRtj55ndm4ZB2FMSKvYBtdZ8hH65Jtab8Igk7mSWDrVb6Dj+aw8UcyNOW5Qpn8SJ70PG61Y5hfbUfguSPYPMj16pjgQ2HXHZ9u471meS3LNZNO66zs6+J6A7eFXusJNl8yZ7Q/jZwzHZdV7/ey+GJ/uA5O5TlVGu2uFwmlUbqH+FGXrTjpTJiktXbvBrdq96CrfePy+rw3HTOeFRevz7Km87Ks5eqsaebb9YQcvj9uX/JI9kySJEmSJEmSJEmSJEmSJPkfCgIuvLJkX+5tcmA2b86a9pc//f7qh97mVo6udavu2+BWfPsaV757yBVByuH0EKHE+X4+JcIkAWBEG35oID4hNy2EtMNJD98egTSCz20UfBI46KYANZByMVAvCIG6EHMSxGod+UVdHLx7H2QX2zO0bEQCEwIAjyO0tf4wXtbRGH2g7uutrG3Rv5YDUUB20Fu9woL8XNBPtjlwvbVBMC+vszJ5wWQE2dD1V0fk2mqbe37b/qYX/l+6z1/Lmme9jdIa3/UkWVacf3DbZS/b1Lml54edNHeFW3t+X97S89vySM8fap4sAXFiRJKkfBpuCikH5NcQg9thrZA/qgfy91T3gaZYh75sgM7sGhATc6wjOybnqCy/vBrsxBY20MfrSMaVHx/Gq/B2Atubvp5g5Jz5iOcgtJuK4EN8hr2paCybPdoatI6vSW2YFNMx8bWQnV23EJaUx/4CVA+wHSPScZ+WBsTt2M6PR/oUvfhjPUg5JuNi6PryBJ35QUrjBPgrAwgYN56hW+lZi/09Oegq92xylZ20Dr96qeu4/ETXtKLyr1nW1E+r+2BZ5Cx43uOHgdL3SSZJkiRJkiRJkiRJkiTJn1ge6QmIZ2dZS2/WOudjs1ZVPr745KO/Vhh67bbqjRf/44rdg7+u3rfJle5Y7woUOBYp0ARKINdwckPT6giCRT3BMdKtoPxoA5TIC2QBAl5KGchrgN4Ab+PbAaQnCNmm9WyjerTVQDfnT9v4PtUeAfUUMiFq5/vnvCCQC6YL+UbyIRAX0Ef1auMDeJACPO584M/Q6xEoKcegwB4pXwv5HV3najvJ5ycv/vHjioWvUEx+Lt3nJxPw6uZfouxvnS/O5sw4vPmgZa+c+7xD3zD7BU981byzXrB20XvP3VHaOkDzhVdWB+U+Y45prZfGe/cYgSL3CgSK5nEPGbI2+HQc5p3vhQH3WMBlvnfRvaS1JN8dB/JFEK9f04X7T3nYxuA2efupxJzUY+1KG/EX1pGM1eCJPr0GGXewzV8/bGidAb6ffJ1vE8H7MhvYU9upxFyYD9GpfeSDQTq+JoUQYxiTpmRj1y2nCmke9F5wXuvYD/QxSOfngO3UX87OxgI7tVGfqGc9XmMl8PfMYR1NArSmACbnqIyUdYM0rv66/cALzy2/Wk24Y8iV7xh6eNkXL/73xeteu2vucU/9WtPixZ/IshmfzrKZG+lZ/jxd80mSJEmSJEmSJEmSJEmS5DEi+P45fBE4fhgArzq+KJs9+x0LnnXIp2sDr//eys2r3cp7NrjaA9e6Igg6CnZLwxTcbqFAc4TSkV5X3NJdx0mO8pYeD7xqxbphSpW486fmlFSwwNVggW0j5JQcgnSckJOTcXZyToL5SIdgHnqFBOsgAbRM9kwCqH1MBhhREIgFaeNfkVXfXk/BuBANomPSQOsbiQOvnwZxEJ+bD9RHfpgAIF+BHLA6+yEIajem49814BavPfnfsqxpgO5pfILmL032S0DPKLQ9re20Y9d0fPLcLctuuvy7iz963p3LPvaOb5VvXvuT2vgQ3d8BV6N1LetN5rQ0vk5/YZXqPPTEHPSYY8CvJ0ljYknWh9TxGqByTNoIiUb7wAA9tQ/7Q3Vob/qcfaj3ebazMuxQbrSdHmGN01gNOh+21sOaj8C2dt1UTylfM66dr3+qjdg16gk0XqRh/HYNQJi7XBuAxiH7JtgIGufcIARdTMw17j/vL4b6s2vwba19XLZ688V5agfCd3JAMegqOwZddXLIlalcpnxpYgCkHIHq0M/OQVe7a5NbtXs9PUevdktueN3PZj77wNuy1hZ8d9xLCXh+g3h/ImEpIUmSJEmSJEmSJEmSJEmSPMoFRMa+Xm1alWXNp2dNsz7S8vjKjiXnPv/fC+8/7Z+LX3zHd8u3rf2P8mj/L8tb+x8ub6UAd0uvK28GuhVdnFaALYStXa4yHJ2gw6/AjvdSwKkklw96kQ+w00ZGbHgirQFGJjBhp3n49eSABtC+TAF+TO6ZbspJH2vD9tLGfNu4jXgJ5ESEKTq0i/OCfCCPwF1BefbRCGrvyQbWUXBPqcwhBfL4vrkdopv1siO/R/fyHLmlXvZLZD2GBa9rLyCAnKhlxbbDZz2t9ry2s469cumHz9leojmrTgy65ZPr3YrJIbec5gtkW3m0Z095rLduxJzcVyHlhJgDIRfgCSq6RzEp54Ey4O3kvvF9sntH9YKInKN8nnQDdC2QX6w/X8826oPqcnqUI1+2br39dIjsBGHMYU4CTM9g27gdrl+u29uzPm8X9pvq+BoF4Vr0egxkBwQ/Ct432AMAbOK5BrS996HEnPlnnfjyexDjngahf/Ubwa4zfpXW11M/XA8/EwNMxJVor5Z3DtGeXe/Kuyjdud7VaG2uoDVa273RVe7ZuKe8u/fXha9f+NOF3a99sPmIA3dmM+Z+jh7fl9Mafxphb4Ln+1/qPk+SJEmSJEmSJEmSJEmSx7zsj5jD9xI9IcuaX8IEXZa9PZsx47QZqzouaHv9sz7U8b4z765ODD1c2b3JlSYGXREn6G7tIqx1xdvXusKta+rFW9bsYaJuy1pPzvF3zcXE3HaCvuJlqExAp3UG2FJg68k0Cpo9YYdglwNeKTO5QAh6IwAa7NXOiABvx+XQxnzbiT3Jkx8Kro1UkLGoPYCgnlIJ1KXMBB7ba6BOfpl0IL9Icyd4tL0RAXjNLSYFGPAHHdWV6R6UJ3HKq39Pcby/Xt1NQf5XLndZtfMXWdZyCd/RII9lYm5/Y5+VzZlzxJxnHHDGogtfvGnJNW/4TPu6kz+35G/OGC1+7cof1XYMuNqOQVfbOeCqO/pdDa8TYt2N9ezx64PuT/zdcXaf+ccdKJ8j50iP9cFrhu5bIHxRxv3BfRdbuX/RfSP4e4l1oGvBv96qwJoLhBLWn5Zha+B6ac/kzxRbGw/1qWsr9KNtCWxD4HkgW3+NmAfTc96uR65JoPZWVnuv39Z4ylDz8MWQNv46dPxejxTXpnkbs12bEWmeiNNxsC2VaW8o0Fbg54BtGvqDLwLPGX8H3IDXWV9cx23hz/wgr6nm7R5xf8hv73WlyT5X3LHOdU721QuTfXvKOza40s4hV969wS2/71q3/O9vcMvvvcYVvnmlW3DRC//xcQd2fiFrallHa/z18lzOnk5YhEW/F0nEXJIkSZIkSZIkSZIkSZI8xsVIkJjAm014SVOp/b3tbznmW4VPnP2LyniXK+9a54qTva440uWKt69xpVvWuvKthNtxck6IuTKIOZyaw/fMUXBa3U6YoACdybgIUwg5JTsoqPXkAAXGnphTWEDNwTtBgn/RxfVCvAHkmxFsrI34NNLA7BTqB0G/1Mf2oY5JAYWNd6/EXAQm5lDPfiT451ffmBCQ9kGPV+Fw8maQX4GrTODEzZCr3rneLV3/Rke37ntZ08yL6J7t7xcZ/xKkDSfkFr7x2Rctef9bbl5881U/LW7FuqF5Gl9Xp/mqV3eAkCPslFOF8v1edN9oLuUeyn30pBzAdVE+Bun4fvO6kXXq1wdANoAnqwhM7KAd8mSTWwsKXgNR2UPXSH6tAOYrysOn5hvH420A7zvYyXoVW1uz4ZpBsMmJTAH8WpupMLscMafwvjkN/e8NOcIrAtdTHzK3pKN+A3FmEFshyAIpJ7/SKvOIvW52Qu6JD/sREDudGsP75ntI49DUz+sY+unl7+VEWiRdcRul9HzDD+rI66s0B7vWu+oD17jqP73TLb9vveu46TI3/7wX/TarFf8xy2Z8Jsua30xrPP7xlv2R1EmSJEmSJEmSJEmSJEmS5DEmjzTQ6yScSOjK5s1+16zDl//dvFce/oEFZx/7/iXvetPW4u1X/qw43s0EXfmba10FJN1mwtZuPS3HAWpdfo2QAlL+EnSAAnkQdQjmESSTnZ2sM2KO9RQEAzkCQMtMElCgzKdxNEhm0oR9qB0TNYDVS19Sh37UjvwIcSB2Fmhb/0xGoC+tl1N21lbqJXCnMmx1bKaTYN6uSYgBBvrS+kAAUFlhPpgoMGIO30u1Y4jyQ658z3pXuvWqX7Y+7yn/kGVNH9DTNTNw4x5jsr/1CLLxgGzp/GNaj1j56vlnPPvytkuPv3bBxtfdvPTzl/6wMtrnamM077ifNF/yxfqDmEP8qEO9ipNzVK6R3r4zLn9SjnRox/kG2D1F6hHWmKwbu3fSRtYD7quQOZLX+4x1gPVl95/gT3PFiNZJiX+ps+H7GjU14sn7B3QsQExUxbA1K9cZykhlDVsd6fQaYzu5fk3V3tvFxBzPLa1dX6dtMAYDXYPPK/w4AfaNsqbUn78W5BViF65XyLgIDfPqfRjUB35R1YPGnjtFhzFgvAq7X0zGjYGU63FFSoHO4e56kdZmcYQAG7zKyr+Efa0r3HbVg209rxptff6hX8ja2z9Le/c9WdOsnubmWafSWj+IV/30kk7HJUmSJEmSJEmSJEmSJMlfkcwiLCOsIhxCOJQAsq428wnldyxdd/J91S1XuepEr6vh1NtItytt7XIlnJgb76GAloLj7X11/h40BgW1DArM4xNzIOUYQsrZiSQBgnvBlFcLNVAPgb61V11MzI0BQowFO+TRbioQuFvekwwA+7a+G0DjYuLC5xHYS9Cf9y3+hWjRerNnMoB0DPMjpBy+OL68Y4jJOf6uqvs3umXXnHpvNn/udVnW8gK6LyXCvl5bfrQKiIa9j3t2Vp797ANeu7D7pHcv/OjbRpd84fJ/X/rFi39QvOny/6xsXvtbvo+4t7jPSq7ghByTcR5UVmJOTshF9xX2mGfV5e9puDfxPWRb1kV6smUdA/dViRxO5T7zSSu1t7IRc3zaikE6Wh9GJhmMBBJCL+S5zvxNQTRGbif6cE1aNw3CvExjZ+253vRStnY5UJ1hij8af/AhsPGHPqYB1wXb/cKufzpQPfet/TEJxxBiDml+LNLG7gOTcqMg5QijlB/pdYXbu/cUt3a7At3Twp3k4+711L7HdX70LDfzuU/els2cje+Owy+r4lXVAwk4JVckzCMkSZIkSZIkSZIkSZIkSZK/Utk3SRLkkKyl9Zp5Rz7hW8sufMl3Cte/8c7yFy7658rta39Snli3B9+DVrbXt3ZSfhelhArlKzspAMZ3y4H88qRcIzEHkkSIEgvumZhDOUIczD9SYs6TfmTLdTkfhkAWxH3xCT1fboAF7jmIjxD4a0BPiAkVb+9P6pDe+wAph7kbFExu4O+nqt6/0c1547NvpXvxSsJj6RTNvsY6n1AlrMoWzDyg5Sml58157dPPmt/9mo8s/tRF3y2OYu4HaS3QnGH+sG7oXhVHevYU8aMONJd8SovmkIk4gEk6zaOO5lTuod5blBnxvVb4e6D3CPdKgTZx2esUQsqKDd/niAQy8K+G0nUIYmIu0tP1WZ6JICqLL03VT+w3hr8GtZfx0Pji64SdTyVv85K3IdhcTFcXtZsC1Hl7baPXsXdyLrZvAPk04ntvyPsi2JwZ6Llgc+ivif0qmJSTPNfRuuI9yv3aveh1hZGuenG0xxWGu11xS48rbe1xRaor0DNv2Xj375Z+6fwfLb7qpP8759mH/iCbOXc4y1qupjX+FCz4fQj2yWNpXydJkiRJkiRJkiRJkiRJkj+SPJJgEL+IeTQB34f0mmze7NfPfcbBmwprX7P9gNG+h1fdd71bft91rnrXRle6Y8gV7xh0xV1Drkx5/s4vBLf4hcyR7np1rM8BIM+qhBrl+Qv4CUbQyRfyg/iQNJycCzYchFOQLcQeiDcKoo2Yi5BrE5ECOT8RRE9+CGYj5am2gCcBYh2N2QgNX09jCXYS/BsRUML8MAEAImDAVSaFlCtNDtX5Vx3vv4YJzuanHXALzT9Oyz1WBGtr78Rva3bIrMNqb20/9/kb2vte9b62a990Y/s73zq87G8v+Fbxtu7/4tckmVwDKdKr6HPFse56yRNzujYAmsMclJgTAsbuSbg3dr/8fQOozhNA+7HPQfvh+0rlvSNPvgkBR3qkU+wkBUBI2hqS9lGebaRs4/VlTjEmArXhfcH6+Hoarw1luw65LiaM6X6In2DPr7z6dntHPB6+lkjHc2bzZzqeE0VsS8jNner4mtWPt4tBfnjOIn9+fNq3EXH2Ojl/VyGBXyvHXqXrL4z11Qtbux8ubaW+h/EcI3t61nXu6HMdN1/u5l/6ih+3HFCmfdqyiXA5LfJX0Up/MqGd1/zeJRFzSZIkSZIkSZIkSZIkSZLkvyVHNs2ZM9R2/BPvrm485f9bdevVvzvwnk1uxd0bHyrvXv/7zh0DDxUnBx4ube+vI6DlYH6UgmAEs/i11xEKbKnM5BwF6kB4pVUC/qnEHEB6C6gpMJ96Yk7JN9J7Yg4pyuRbiDntw/xEsJM+oS9KEbRTOpXACEF+DiArGCAJlCig/vN25Av1EwExKYdgv7pjoyvfsd6Vv32tK332fJe1zfsiBfvP1vl/rAkIuhbJksyeXZ5/wmFnLHv/W76y9Burf7B0a88fSuODNN/rab6HaG5BBMl94LmkORPYfdB7pGujRvPH3zfHaQTcOwP7Eh37BXSd5O6l3jsB8lK/V0T2ntSbAl0HDCHTAjGnyNmLnehlzVrZSLlp0TBe6Gy+sAbtWmV/iT4gtGN7huQN2McxMWe2AsvHdQJe75zGuvz4MG7TeVJuGuSJOYD0CiHi5FdaPSFnqSK298C9A/kGgIgDIYcfXsHr5FwepDrKwzds8GMsOwf21Ia7Hyp+7rw9Cy9+sWs6uPBLWtm30VLHa6uHE9qw1JMkSZIkSZIkSZIkSZIkSZL/qYBM2dspjvkZf9dZy+VZU+v6pvLiG+Yd94R3LTrt2GuXXf3KDxXefcaNpS9dek9xvO+XTDaN9bvq1nWuuoUCaUavK27ullN023odgwJtOwEXn5pj4oVgZJoRavHrqZ6YUxLDEy4EJuXYTnwLoSd23pcSA9wfEwWq1/4NpjeShAkT8g1Y3wj0hayBnRAcIB48ucDtzUYIAZzKkRM6IOUGXXn3eibmKvde75bfe4Nre8fLfp1lMz5M846AP5ZHyymbfZ+Oy7Jyy7K5L5r7jBVvWvCyJ7950fkv6G3feMqNyz5/yfcr43j1dCj3+ikTXTRnpfG+Ov+oCOVl/kOd3C+9N9TWEE4+0fwyOQefEfQkHvvAvaD7KIjvTeiDQXm+l6xXO77nMWBDetx3ssuTUeLP/Htyacz6xtoRP0I2SVlSKvOpQYP44LHYOM0/+uXxwgb1lif4PSLA2g5txU7QqLMy+ab2gZiL7Hhtmw3yqLdrIPC4qF7nxrfjei3DRu15bqg9SDV/f2iu4vkC+LQgUuovEHBYN2Ecko8hepkTGbv1X8Er+fjhBn6VHOQ47cNJWpuTG9zynRtd7e5rXPVbtC939rqlf/OGf2499pBPZfPnrs+y5vfS/rw2a571NlrrRxHm8KqfXtLpuCRJkiRJkiRJkiRJkiRJ8r8W/CLoQgJ+hABfZL6CgC8z78ja5jxt7tNWvmHxhS96X/HLF3+rct81rrR7SILqLd2usrmX0h4h5oa76tVxkHLyvXMI5pmYowA+RkzKMTiYRh4km5SFgBMfFvQbOZfTI29gHwQlDXKIyIMYEtAHcmCKTc6H2gM0jlwKgIyYIOCEjp7SKd8x6Cq711Pwv8mV7r7mD7Wbe3/V9KSV9zU1zeqj+X0i4dEo+yTmZj6+evyy1Se9t/jp83Z0/J8LHui86dIflG9f/fPKcM9vy9v76vzLvQoh1DB//Xq/lXxDmkO4dzHxhtcuA2EHiM5exwT4fpOPcN9kPdh9EZ34YMLG7DW1+xjaa73WeftIlwPWQLSGGA02wTf85eu4Ph6bgsnDXL/5ssHmzvzLeLSO/DBph+vjazSdtbd2lqc+YsKL2+Rh4+X1rmV/HeyHENUFYg6kG0FJOX6GMDlHdoC3JyBl9AtoTD5vdbCDP/aP7/jTdDultOZKO/pccSelu8jvTlonO9a72p2b3PL73ukO+NYNbuXkkOv8xLlu3uuehROsN9LSPp4wm7CUsJyAH8zBq//hdGiSJEmSJEmSJEmSJEmSJEnyR5R9naKLpZDNaT1p9hGrNi299KXfKHz0rXct/cgZ2xZ/7K07l33xwn8rbllbrwz3uOpIt6uOEZicE1KukZjzBAwAG4IF/FbGK7FATLgFqM4TBvky+wXpADIihm8fwEQACAFqD+RIQCDno4GE4HYC/0od2cmrrANCzO1UYu6ODa6y+xpX3L3pV4uveNVk1jIT31n1QppXBP6PBtnXGliazcgOzxYtOCrrXHTozGMOPGHBpS+9fsmn3vH3lbEBt5yutbZjkADig64Zv947SXOyrbcOCCkX0EjMxYQQIKRRIOX4+8GYkDOonom5uI3ce7t3/r5GOrGzPgmUj+9nWFNUx8QW6iSFLfTe3sOIJVlDOaAuso3nIf+arIzFwASct7E+JZX2MhaArwc61Cn4WqL6RmIutEMe86J7httTPxExF/s18Dg0bQT71zGZLibmZG5IT/tbiDnK69jYXtNAwOn165igCyn5BMZ6XXG8h9BLespv73PFCSXn8GM1+DXpuwfd8geudau+805X/frlP1zSddJdM59x8Hg2b/7fZ9ms0axp5tm01vEfFPuSdDouSZIkSZIkSZIkSZIkSZL8UeWRBpmzCDhBclSWNZ+QLZ5/Usuqxc+f+/xDT1828JovF0a79xRGul2ZUBpZUy+PdNWrFCw3EnP44n/RUT4CgnzOU5ANks2IOU/caeAuxJmSZ1oWYkRgxMEUgJjQvAT5oQ2XGShbPkY/EyRCmIhOTurk80I+UAq7Cf1uK7zGih/NwHfM3XOdK27p+8mMZx3yIZrH5xHmER4NJ3GwBpolO40UFh678Jzj3rXsQ6ffumjDKV9sHzrl5mV/e859lW+s/cWK7UrMTQ66Kn+HF6BkGs3DtKD5jCFEEd1rneNA7OwbclqL8t7e2ut9xv3Re2y6uL218/fYIy7LPUbe2uUJNYL1Qf1xv7k6rdeyjEP6Zp8RjIiTcSrUXtqrba4OedVpOfRrOusjX7b7MdUmzkdgW9l/5tuuKw+5tlgXiLlor9ipOR1rjBwhZzCykPOSsh+QcWMg5bqZmOsc7aqX6dnDpB9eGd5N475/o1tx/5CrjV7pll57qmt50orbsqYZF9HqPi5rbn0F4STKP57X+/4lEXNJkiRJkiRJkiRJkiRJkuRPJvsLQmfNfeqKywqfOOf7ndspWB7ucpWxLlelILky3kPBu3zfXPiOuamkHJA7FUd5/yorBfMg8oSYE1LAyD45JRdIj5gc8VASg0kf1Zl9rm0M1huCPpA4Ri4gL2SBkQ1cBx1IDBBU9t1WuwZd7b7rXcen3vFgtnj+lZg3mT4WzPGjJdjHiSEQhjO5tGh2ecH5L+xd9o2uH3ZupWvaOuCWj61zy+l6ASZ3+Lv06FoVXAZZhzoP3GfNI50CvUdULyBfPj89jADKkVQoI8W9wf1Q+PrYh9rKPZV7C10oN9Tr+KYQc4D1g3S6Os3bOMR3GIP0q31zP9JXqBeb/NgjWBv2raA58O0YIe9PJHq91TXaK3y/tu9E78ecg1wLfDCBRvD7pQG8Z6boFOTLQ3VxP1InxJyclutxpZFeVx6h584Y5Xetc4U7+/eUd3Y/1HnTpb9deNEJP8mesOJnWTb3rixrvYJWN17bn04eTfsxSZIkSZIkSZIkSZIkSZL8FcsjDVBPmn/qs26qTvT+vOOWqx+q8KusQI/zPwSxHcQMBfTbhSQRIiaQMhLwkw6gfPjxB22nZICQckiVMLDAnvL+xBtsrZ5gxJCRCzGZYDaxTk7hgEzR76zy/g3ig7+EnlMlCWwsXKa6SaqbGKiXJvvr1d1DbtU9m9yintf8ImtqBTEXy5+aCNj7jzrMzA6Y8dTlp8976zHdc0995oXzznp+/8J3njFaGh5w1bF+V2PYvdA5neBXDuvl7evq8oMNNN8g5nKn5vQ+oA1B7iHKck+4zMSS3aMBmWfKB8KJ8mhP5Thv9UwEIaWxAXyv+J7IeuAyfDPEB6c+r30x1I79CvI2SKU+7gf33uytz0BmYU1RGWPjMZIPbw/f1ofqtA/xE/XN9gT2IXXSXlI/D2jHfaEcdMGvIdiZb743ZJPfN+hLYHYyprieUvOFfjkfI6+zPcN7TMsMmjMh3SJgf0NHzxBDYbyrzoTcKEi5HoeTcmV8p9zdQ654Z69b9KUL9sx7zVH3ZIsWfoyW/eqsadbaphlzL8yylmfqip9OcHo0EXNJkiRJkiRJkiRJkiRJkuQxIThZdeysYw5ct/Qjb93W+eWLf8TfMbcNpBwBJ+QYIF40TwG8ETGWD8QbgfLyww8S/AtRJzoh5oRo8eQA6RmsN4R6D08gGImwNwQiwEiBfL35EIQTPbDVPOyor9K2/nppvL++/J5NbtX4oGt98dN+nmUtPTRnbTx7fx7ZKzE3+7hDT178t2fdvnS09zdLvnrVTws3XvWfpVu7flUaXVev4NVkujacguS5pnvJJJKRNxOC/PfBgdwRsA3Ny9T7EiO+R4Z47jGvwS4HsjWiKcaU9oyonQL3i8kxhvQd/JiN1Pk2OZ97AdYm1hDBk2kGP+aob9ORf09eMsI1AcFHBB5T3lbIsIaxe59aZ4jaNSLuh3XmSyF1Df6ia7W9IXndVyhTOy7za60yT+ILqeW1TNdQojVVBKjPztHeOtrxswOvUe8ecuWJXrfkE+e7uac8x2WLF/8HLff30p57Pi1vPKtwUnUJYS5hX786nCRJkiRJkiRJkiRJkiRJkjwmpJXwpBlPLp655NrXfb7zs+/4np2UqykRV5sQeIKOguz45BsTbQisAdYTxijwB6zM9Uq4UZ7JDQvWLcCHnsqBTIjzBE8gANLWyLRABFhqdVIvp3XEzkg3AxNzANutc0WPARrnILUZcLX7rnWVz1zqsiVL7slmzL6A5gy/9PinkH2d/FmULZjx9JZDii+ccXDhiJnHPP6EttWv/uiy27p/XcIvnW4bpDmnaxgboLml69zas6c00luXOaf5JMg9UwKJ5oKhhBxOzBkCOSf3gslYpFgP0LFe7w3Z5YH5xtxTntogHxNCDKoTiC/zbzpbH3YPmbzi12XDK7PcF+p8X1I2P2Eskb210XbcVwRPNAE8b1jrMn9mE3xROfIPxMSctwf4etSn6nwdj9XyhOjard7Dk2dIpQ+xye+duA/OR2M0WJ2dJGVY3zRWIeby+wp1fu/o981xHXSKIl0LUBjvreMXeHEatbh9ULCD1ubu9W7FnRtc+ebLftrW96p/aD36oJ1Z68yv0tL/YlPT7PfQI+oVtNbjV8djATmXCLokSZIkSZIkSZIkSZIkSfKYFRBzT5zxlMoZi655/ec6Pnfhv1a391CgDjJuXQ4gYOJfZjUCwIg5e90PmPKrrKhXe5SN7DACwYJ/b6OpkBdUp2DbRqCt5kHq5Oo8xJeRBqyDLcETCGRjpEMRoPGXh6nd9gFXHuv79cK3HfdvWTbj01lz68k0Zzix86eQvRNzxYWHLbni+GsLnzl326L3vun2he960+3LPn/xP1bG+h723xHHBA5dB81RcaS7Hu6TpHySkep5vmkumBQDmJwLxFwNZfiLiBwmhiy1ftSXpWIr85wjymI/asN6D9JFkPVBeqRUH5NfOcCefUk+JqukDeXZTnTTgf0DNDc+r8iv5Qb4cSCv4DrSGYFGZe/LAJ9xmXzIfMV20Ilv1qs/f52cUhl9MLDeLS/wcwdbHav8cm7kl8D7gfwLyUagfcCkG4H2gfwKq4HHAdt+2jNia30LKUf7iHyDlCuN9O6pjQ1Sn+tdZdc1rnLPta5y/yZq2+3arn3DH2YcufLerGnme7Os+Sxa3UcTnqGoEPZGvv2pXyNPkiRJkiRJkiRJkiRJkiRJ/qiC18OeNvPoA9+x5H1v+dqyL138b5WJHled6CVMJefkByAEQhgosUOBuJAWgvA6q+mCjQTuqmcSTustz34FjeQI6/YBI0Om1sEPSDmD6sm+ZKAySLnCWK/rxJfPD/e54tYeV9454Do/ccH3mg4ofT7Lms+k+XqSztufUnBiaDFhGeFx2bxs2dw3H33J0i9c9I+lHTTXO+me7Ox3FcpXJnFdvXUA91C+F5DAc96r8072lMoJR8rHYMImf2Kutj0+Nacg2xxoDm2uTYcy3xPUKaQc+cn5Q9oAJaE84Hfa9g15+ENbPw6rQz+qi9JYb2vU2hpiu5w94K/NEMaRG3sj+L5gztCfkGyNNnJtVo7mo9HOfFEdIHZiz2X1L3MxIK8v67yYD94HANkbIedBvj1hFwGkXIHGz+QcUkKB1p0Ar07TuqM+qjs3udId1z5U3rbhl4UbL/tZ29pX/6rpCdXvZVnLCOFdtL9eSuu7nTCdJBIuSZIkSZIkSZIkSZIkSZL8xcnsrKXl2AWvfHrf0k++fbzjK5f+COROebLPVSeFjFsOTIKoQVnIGSHmhEzwZAAF7vgeMyF/hAiwOk8kIE92gCcQQCiQnxwRwvmoreliG4AJBQJIC4YSD6ozksQTczQ2ea1V/DABoW3xih2flqNrKIz0uOKWHlcY7nKlXUNuyRUn786aWvBLkAfznMkXzP+xBaeCpice5s88eN7Ln3L2oq4T3znv/OPXzTvneYOLbnjTbZWxnt/WdtH92EVztZOuaSddyw5cU2+9tK23XqF7JsQc3ROArm8qyIbvA+YtEHOAEXPyoxAg56JXRzHHOs8813xPwtzKfSXAFm3YHmXY5dv6+0gAmSaEmpYN0Ed+pV78M6hs455igzz6ACIdrl/Wl407mhezQVtFrJtSh349pGx18XXYOOL64Au6uJ3W+TYNdqr3dnw9+fk3W+7b8nwfcU9lHDIWAt8bAuYkOilXGqV9o/uW/ZtegROmIOYK1Ban5DpHe/YwSYf9NkE2u6m/+4Zc7a5B1/HpC38y99VHbs2Wtb2PlvymLGu5NGue8WZKX0QrvSoLflpJxFySJEmSJEmSJEmSJEmSJPmLkznZzJknLjznuA8Xb7nqu8u+ceUvyzh1RajtWOdqIOcY+yfmhHDbBzHHJAGVyY6hdUIgSF5sQAxIm9iH9Adb8xWByR1FQ50QQUAfE28A/PuTQajjtnlirrC5yxW2rWWyce5zD99Mc3WiTNn/M9kb8fC42S9/0hvaP3D6cGF8wJXpHpQ3r/1FdbTrN5WJ3j18Us6wg+rofoWTUCDacBLOTsfpCTnMJ8PmWOYVMHKLYfc7IueMyAmEjs2vtLfUAzawZUIIZZ1vI4IY8JP31Vg2YksQbGwc9j15XGZ7adPYNvYb7BSYk+n0BF67jXqy935839q/v07UR32bnY4zPxazC/o84nrzE9XTfQx7LSrzqTntn9sJMWfrJMw/IG34NJwSb0bMlWgv+DKhCJCuSHuev5eR+sAr4AV8lyEIuTsGXOneDa5y54ArfvECt+D057isvX2C1vQaAn5ZFUTcHEILAadCZxAS+ZYkSZIkSZIkSZIkSZIkSf5qZF42d+6blnSd/I3yjsGfd2xZ81B5Z58Qc5P9fFJu+eQ6IWgmNKUgnok5nMTyBE8fE3NCysmvf8avsgpZQHkK3qUsOiGEQG40lmEvth4os17JA4L/UQcai0CJtghMPGieiTnui1LqiwE9t4uIuS1d9eLmLlfc2eMqX7jMNS9aehvNFb73Kpa9n3D738m8rHPBkbOPO+h1rc99/Emtxz/l1AVXvfqThVvX/Gr5xJDck539rkYAKVfc3rOnNNFXBylXoTq8yion5QRMxtHcgJDD/HFqc03w98Lq+f4q+L4jDxInkHPxyTm5J0DIiy8heoT4kTLrfV7qZJxq6+sNakeQ+95YlrYyDsV0PiN7JqG4Tsed6296yBxhHcb6/LrMjYP7D3MhfanexqTjMxuBjk3H5MfI90ny0t7s4FP9YWw6Pn8/uR3Z4bQhE3MDDE/MKfA9iuwT4Da0H+zE3Ggg4gIh18MoEDqGaa+wDaB7aTf5vGfQdWy+8ucL+19158xnHvzNbG7blixr/VzWNPvyLGt5Fq1zfL/l3mRvJHWSJEmSJEmSJEmSJEmSJEnyFyXzs7Z553Rc8+YdtW+/kwLxHlfB95XtWOeW7xBizk7MVQEKug2emKPgHSfl+LQcMC0xFxEbHPgHHZNyIDk8uYCykBBSJ/kcUdEIGg+TcgCV5XvkYj8BXE/92Ok5sTU95ekaOm/vqhfxGutEr1uy9rWO5mlUX7WL5X9CHOy/TXnBgXMveP6G9o+efV/7h8+5f/F73raj87OXfq803PMHJst29Mn3yekJueL2vjoIOb4/BCbl7BVWAHMLYg73jfKA3Beaj+he8PwDfG8VOFGlCKfnVKekkM29J7xiMIkk824wcsoTZQZtY+RW0BtQpzoAY6fUxs2EI3yrbeg77kv6mA42Tj9eq+P5ifQ2Z1aeDn7MNh6k6k/HxUC/Cukv0vlxx9B2lA9zJHWsVwJOXvtthJBybOtJOeQjG25LPu0aeW/LngCEkOt2BWBE0LF5bb000iN7ZycIuQFX2NblOt97ums96qBvZS2tg7SqX0L75/lZNuMoyh9AwPclpl9TTZIkSZIkSZIkSZIkSZIkf/XS1rRs4aXFj5zz3eq3rqfgutdV8J1lO9fxqSw7MYfUE3PblAQBwQPgpBwF7ULMgfDpyxFzFtwbKWevshohZMQckx8Akw1xPcr7gZIfgCfZ2C/pyKfZ5Qg4gGzYjvsiPU4JjaxzxS29rjDR54qbu13r8w5zWfa4L2dZy3E6Z38sATExn1DM2matpHQBlZbMfOlh5yx6/9vuL1Pf1WGaDxoP/wLmNpAjPXtK27vrfCpuBwGpAt8JKKcaaX71Rx9witCIOSHEtEzXy/cmBteRDewU/sRcBE/WMZEE0NimwEijSMeEUJT3QBl68QedkE2qVx/clstar/dViCoaK9oqQnvLh778GLQ/6VP9YF4sb/U8N6Lzqeri/qVe5tHGa5D+FDYWrZN8VGaIjvun+gC01za+HNUpsSb7SGwEdiIOp+OorOA543qaP2oje5bKSHUv41rj11iLhMIo7Y+t3YQuIevo2bBsuOu/ln7yvH9feNnL/2XG4Qf/LGud+y9Z04y/oX3zbCz2/Ug6HZckSZIkSZIkSZIkSZIkSf4qpa2l2H5V6dMX/lvl/utcCT8QwMScvC4ppJwScxzMg7iQ1yLjE3FCtOG0nJUR2KNeTtAxkQEbghFzvk7JDCGM0AYEAyCkgJERBhBqnDcCA8QC5UE4lBhS50mSuK2Cy+qH/QE0Xv7OrGG8ytrjCpO9e5Z96OyHsjnzvp9lM95Fc/UMmbJHLPs9ETSjuuRpC17zzPPbL3v5B+a8/YXvnv/WF1yzZOiN3yx+9apfVLb2uAqNpTIC0lPmrDzSs6c03lu3k2sx7HsAhZzD3BioLcFeZ/WkKuaY50KvX+fLk3jaR/xaLEPLTPBwfYDcE/Er90VgpJT5kH6VTIIt62FH9kwUGcTO/AgBJnWyRlCOoHaBqIrrqByNISCMYzoIySXwOsyTz8d1NoekI8SkmYzN7EiHfuOxos7g24FIEzLNIP3FcyB5HgPKrEdbhZ6MC37stJxC2woxR3mQwAD7lv1SGO2tl0dpf4z0U77PFcZof9CaKYEcnuhxpa9f5Oaf/dz7mgqLrqdl/RZa2ec3Nc1+R5a1HE/lDl7s+5ZEzCVJkiRJkiRJkiRJkiRJkr9KaWupLV1d/OplPyzdf60ScxSgKzFnP/7Ap7GU3OEgngN5yitAygkxp/DEnNUrQC55Yk51SmZ4m8ayJysEU4g5hZByIB7UPoa11TQGv9aK78Ya6RVs7XOd2/pdx/a+h+af/tzv0hx9PmuecQalB/KM/XFkRtbRtmLea496x5J3nT5W3trjaltp3m7reqhy65o/lDd31asjPa4ySvMK8JzhFWGQa7gPuB8ET8TlSbmpxBzpDHTNQu7Q9TORE5V57tEGwLxaPg/MaY6U03sQyCOF6hsRz7/0F9Vb2wjcH/qFLeuQWnkqvF+6vim6uC+F1U0LW28A26J/mlPVxb7FRmDjZZtcX9MAem6n+b3aoG+AbJk4o7zB7KIyn5RTAs4IOuSrpMf3BHIKG7I3Uq7Ka43GHF036paP0LOA9kl5+3p6Rmx0td2DrvC1S9z8C1/kHrei8Gta0h+mdf10Wd4s+CEH/KhDemU1SZIkSZIkSZIkSZIkSZIk2YssbF1VWFu85aofle69xpW2g5jry73KGkgeEDzRqSsK3BlKGnlSTok5ef2S6jzJBoJJSCYmmkCaMDEEAgBlkHtiz3pugzqyU4IAAJFmeSYfjMwgMDFndTxmqteTXKy3tmwnEGKuzxWH+1yJUNy6zhV2D7mOb655qPmg8jdojs4mHExYgAnbi+yLfAA5UW0ptj97zpErT5h/0uGnzzn5GefNP/WY1Us2vuFrlW9e/YsazUtttMdVR7oI3YQeV9rStae0tWtPBd/7h/kFMQfQnMTkXA1gYi6cnpPTjVKfh9bx3GEudI5pDmTedf65Xnx4QMd+ZT6NlONTdaZnnwCV9d6EOsubTQTuQ+uQcns7zaX9aT3r+D6aHXQKvo4oz5Drsr5sPDFs/ci60DHpfMgaVh2PD/5lntifgteT+oivMbSN9GxH647y/KMj6lf6D2MojVMd52Gv5BpfM5VBpAFsD73USXvRWRvpj+oZGIfmyU6gfdB8FUe76kK00zUyQUfrhp4BtZ2DbsWdm9zKnde4wgfP/tc5Jx1xa7ak7cNZ1vzxrGnm+7Lm5lfTOscvq+5P0um4JEmSJEmSJEmSJEmSJEmSRGVh68GlNcXbVz9YvHsjBe8UjOPHBQj84w9MvJBOSTkQC0zwKNnRiOq2Xoac5kFQL3oE+fYF8qEughId3tc26PIIZEVEsiiEeDBCQm1jsgOp6QlMTrAtiLk+foW1c3N3vUAojZD+7vVu2d+c8busaeYHaI4Ok6nap4CY2xvhsHjOsw48of2ql21a/K7Tv7rsU+ffV/z8Rf9U+sxF/1D56hUPVkfWPlTd1kPXbQQcTsr1MAlXHumq49QcykbOyfwIYrItJuCmvH4KvRJ2MTEnEPLJE6HR/E4L8gXExJZB9OK3sR3bMAmk94jue/wDHeKD8uwnAuw1b/fMXtU0f9wfj138CmElfhsRxivjZDuvk3Yx2B9SlO36eJyU8r1o8Iu8wv9asNowyEaAvIL0TNJR3sg6Wcsg5jTPBJuSaREpx4TcuCHU+R960PZhrmgNROB6HgOluCcjPXtwaq5GwGvstd0DbsV9g+6Au2jtfeHtbt6bnvOHrH3hV7Ks+c20tvHdiMATCQUCSOgkSZIkSZIkSZIkSZIkSZIkySOUttYn19YUt6x9sHDHRgrO+1xlRy+Tccsn1jH49dWImGOAkGiA/OInBe+AEiQgS1CH1L483sgfI+XYDv0CVrctkBtGeBix4b8TLrIxYoUJDy6LbSA/lHxQ8DUwEWHEXA+flitt6XXFiQFX2TXk5r/x6B/R/KwmzOWZEgH5tq8TP3MIlaw1e3w2a1YtW9X+xNkvf/Ip8y552buWfuqi+6rbB/+wctd6t3LnkFu5Y9CtmBykOQah0lcvj/bske/5onlUYqQ6KphyYg6g62PQteVIOQDEXI6cI5scMWegOkDvlZ9XvicRyAcDdeozEEwBjfcqBtdpv6zT+851Vs9tFT4vdaFMdiCeQEAhz4A/jN/8mi3yMaQPD+j0eoJex2OI58KuW8tCamre/LDe6gzaHkB7BvIKrTNiLoxfiDkm5xqIObl+lEmPfAzVcx1DfOVs6LoY0GtddXTA1Wi9raBydXvv7ysja35c/Oo7vrd0w2v+ac5zD707m9d2Z5bNvDlrmnkurfNlvOKnCkjq9PpqkiRJkiRJkiRJkiRJkiRJ8gikbdYRq9YUR3ofLOzcQIF/n6tMKjG3fR2DAnRXAbaBDBIYGecBHeuVmAO5wCQJ5e0VViZOFGyvxACXQRBIGfWB3KByRIQw1Ca2s/axnRAcBGrPr7hOgNwwoA5+lZgDabiV0i2Euza54m3drmlV7TtZ1nQJzRF+OfWRSnXOM1a9bVH/Kz7UNvTKD81f8+qPtl93+s3LPnnh3dUtPT9buWuDW7Fr0C037CBMgphbV6+M9dVrNF4GjSkm5piQo7J9N58gzBEIIX69WAlUvr4JgSfpckQd7I2kozLBiDImlMi/3CfoBGIH/9Ke58/mUhHqpkLaY86pzPewoU7HYjp+vVPzgUgL4Drzo2NHGs8L22BMWo79S51cg40b9UZ2+r7GpS9uz3ZqbzqCr0de5y6uiwk6fy0RxIaAvNcLEccwgs2IOZsDjM0DfQNaVmIuT8YRxhRkWxrp3VOlfA2guuXkcwWtyVX3rHfV2674+cILXri15YCO67OmmZdlWfOpWfPs12fNra+idf5kAr5DbjrZH3mdJEmSJEmSJEmSJEmSJEmSJFFZOOuog9d0but/sDC53jGxM9nDZFyN8gIj5nqnJ+ZQ9joh4oRgAygPUi5HzJEeBAW3UzsuWxv4iUmNBrCNtuf+92K7HegPmFCCjqH123Aaqc8VaZylkXWuk/wWdw38bunga/4ze1zr17Nsxuk0R4tkqqYIyIfcyaDZLzjkVe0fOvUbiyf76h2Tg66yc72r3bHRVe/AF+YPutquIcKAYCeBiTnCBOVpXDUaD4g5PjkHYs4D80jXipTnJ4LNPxATc8DeiDnqA8gRcwby6aE6bxO1DXMreq5TCOmJdBro/WN4PY1Dx8Jl8hG+dw0+rT4gtCWQL1tP08HsGst2DeIfdeo/7gtkFvcnem9v/rjvaDw8bzJ3YT6CLlyP1sMO7dV38J8n5ISUE2JO5pdAY2PYdzryvIqOx8725B9jVPjveAThO0z7Gz/qgH53DtZX3r3erRpb6zq7X/OblkPKI1lTSxct62cRlvACT5IkSZIkSZIkSZIkSZIkSZL8UWXhnOc+cU1he/+DxYlBV8H3yU2AmOtxNSPiImLOk3OeEBKdEDlIJS8EAAAiKZByRp4IgaFkAetU7+2glzrWIfXt+rWdtrd6SnPfP2cEB07KTQwwMcfkHBN0ZAMygoB8kdLC6DrXsXvAdXz9qgfnHvvE22huLicczrM0vRyUdcx/xZzjnvT2uSc87bRZJx5+1vw1r/zY4i9c8h9MrkwOuOrOQVfdRfNKKO8iHco7qd/Jvnptx4CrkQ1OzOG7/IyY4xNbdF1AOCFn80jIEZ2i83PoiTkA10hluqd8es7rYavzx+QQjYuJHpk3zCXPKfXP9wLQOmmj9pR6EsmgNp5oY+IJerJlwkh88g9uIM/9CsyH+WVwXX6M3o77EX8GWw+Sok5Sfw2ws3aE0B+V47lg/5pXkovbsB3KZBvXs436Rb+a9+Qb5p3vkZTjPsUvtc9BSDhPynH/aBPZaL/cH36kQX+ogb9rjtr5X17dTmsObUd694g9+aO1XhnBKUxabzvWu8qOgT8U/u7t/9J2ynOGmyvFm7Jsxg2Et2VZyzG0zttkuU8r6bXVJEmSJEmSJEmSJEmSJEmS5H8hC+e88KlrChMDDxYQ0DOR00OBfa+rbetjGPlWwQ8UMIQkMn0ZUGLEk3N6gkdIEqsTcsITFCAUmMRAWXRm54mmqG38a6yCUBdsKNX6mOgQUg7kHGBlJekm+5mY69jW7zrv3+SWvuesf8wWtL0vy1peSPOzt5NCC2YfufKUjqHXfbbzxsu+3/6Z87+/+OPnfafzC5f/sLy567dCkq1z1R00pwBeVwVRB+zoZ0JOoKfl7MQcQOPg10zpeu075+y0XI6YszLqbb5w3QqeH9xPBuYjgtlNpzPofFuZ59XmU1MB1Vuexh3yoa1vH8G/akyADyO9jIwT0krL0LNfsTfwDyXovTfSrxGBqNNyPE+xTxuD5ZFa/5xXvY2F9ELYWd7aTEX+/qh97IPHQcD61NNxcloOdVYWu1BHa4qJNoH9AASXtw8yKsC2QVcdo/KW/j3VUSrDlnxg3R2EH3n58vlu3pue9aPHLcYPOjRdSGv7eMIKwkICSLm9vbKaJEmSJEmSJEmSJEmSJEmSJPlfysI5Lz9iTWFy8MHiGAJ6EAk4FZcn5oR8m46Yo3ojhxhGFpEfkCFKSBghFwgTqlPEZIn4QTtty/UYUx+/cjqF9LB6AgiaPDEHgOyg61JUJoGBiJyjdoQi5Tt2DbnSXde6heee8C9Z1vS3zc2tb6T5OZJwCGE54fHZsvnPajmk8zlzX/SkNyxcc9LHlnzxwv8sbO1yVZw+GqY50++D47EzySn9g3BByj/AQGNgUo7yTMjFKY0X4B9p2AZEJ+c8yD9D5pLnHdfrEc8T1fE8NCC2xWmunE7B8x7KnqDCnHKeUmrHbaGLobaxPxmP+mgA11O7UB+l0HM/mmobf7/5nsscyNrKg/tH6mFjkvHn5oPzMv5c3kPGFF8jENfFY2cbXZ++D9Uz4Ievi1KsESbhUA7kW4D45tdTlYTLIbaj9pXtQ4zl44NuJWPArdi+wVW3DP2y9NlLv7Ns4LU7Z57wtPuzOfMnab1/NMtm4AcdnkDYm6TvjkuSJEmSJEmSJEmSJEmSJEn+yLJwziuUmBtZR4F8HwX1U0/MCdnWQ0F/TMyhDkSHEHNCjkAPKAmhJ7aYlDPAVu2FOAGU4GDiBHVoL3VCbAgJE0PaC+kh/sSHBxMcSAkTID4IIOaYJBuIyLl+V8T3r9290a28fZ1bcPSTfpplrXc1z5j76ay1daBp2aJL5h59yFXtF7zobxZf/4YvLVz/+psWbXr9rUs/fs53Srde/fvKSJerjALdrjpGYAJT54GuPf7FVAb1F4g4vGoIwg7lCKRjgm4bga6PyTm6Rp4bgE/LSZ71uF6dhzCHAiOfwq+yYm4FMk9kZzq2BfkT1TOMFCI0tGfyyduZTvM8DtXBr8J0vg4+TI9Ux4Fy/CMeZo9rnXqCEnqZEwOvn4Y6Ixw9qC/pT/N6rbn1Y3rO6xwowtzIdcQI6xN1CmqP+8o2TKAZrE+8Xj3AkGuOEJ2Q8yA9n/ykdVMY76uDaC5PDNJ9HnTLGQNu1a5Bd8DdG9yBk0OuMPDm77YcunxD1tz8Str/J1P6mixreR7lDyXs67VVSCLmkiRJkiRJkiRJkiRJkiRJkj+iLJzzyiPXFCZAzIHE6XOBmOvhk3NVpCDlQMAZ4cTEXAQm56wckRAKI4k8jDyhvNQreaE6ITKUWGFiA2ScnZKScqOfHNkCEoNJDvJjZMqEAnnUUZ5/HGGS0p39buXd17rqB8922cIZe2Z0LPnD3Gcf9pNFF594X+cHTh9e+sUL7lx6y9U/Lw2Tr5FuVx2lOQERR6iMdbny2FoaB+VBymEeCFWAxsWn3gAah5BwQsx5ogxpRMohz0AdQNfC82jzY4jJObvuOM/APZU5QD+WZ39mQ2WkotP5sboGWJ3VS0rzp2VvZ3mMz3RoqzBdqKMUek1ju9wrr9P0FRDNjSJcB8ahOtTl2hF0XjwRxnkt5+oo1TGwb9RhzswP61FWW16rAthAL/c13N88OadkHPelp9/UX/hOOUCISaBIYyjSGi5NDioh18+vqVYn++vVHQMPVXf2/6Y6vuY3hfef7mYce+jvs+ZZH6N9/wzZ/kmSJEmSJEmSJEmSJEmSJEmSP6csnPfqoyJibp0r44cexntdjYklkHPhlJwRHHwqi1PRC4GGVPLyGiYIBZAV2oZ9a1mJE/FlpAUBeSVOxD/sQGz0udLYOoKVRWc2vj0hR64oqQEdEz2kY9IHeuTxKusk2e0adNU7NtYXffLc+tyzj3bLP3qpW/6tTa5jZ/d/Lf3qVT9a9o21/1UdHXDLR/vd8rFeQh/ND34Eg+ZktKsOYBz4Pjj7cQwDfmVVIEQMn4YzUg4wYg6vuFKe65iYk1NUQt5E10c++do1lXkIdTIfpEcZ88PXDKgfPsGIOrSJ5oQgOpQFdu/M3tc3wNdHY/LQujAG9JMH+9E+47E0QvqjepBU3C7YecIWawjXDVuvi8aBdIpO+mToeFhvZYyL+yJonsuUNx+yJtXGdApZ19KG9wbuLfsQgs7yQswJrA8GtyfgdXNCkdZfkcZfJF2R6nG6Dq+t1rYPutqOIbfiro1uxQOb3LL/c/635p3yrE81FRb3Z034QYeW67Ks+WTse9n++5R0Oi5JkiRJkiRJkiRJkiRJkiT5fywL573uWWsK2/qZmGOiYTsIOT3xNS7EHMpGugUyTFPNg+gINiAvDFJvZU/MIYUOhAPqjXzgerJjWD7WCaSNgv0GTDllxddFeU+qaB6v/U2SPX6gYdeAW353v1v1wAZXeGDQtW/rri8d6f1DaWvv7yuj6x5aju/oGu93y6lfvOIrpwklD5JOiLq9gO3t5JyQbSBn/Mm5vYC/l07tbf78NRtIx3NPsFTmEdcv9Qy7fn11lecGfj2kXvQK9SH3S+s0tXwoq10EPuHo67UN9WWEXClHZOEaZRyx/d4gfUsblLkf9Gt18TjiOkqxduR16NgXga87GgfnVcdl0clYpW3wIe38dewF/GupIOE8oI/zDRgDqA8FE9R0f0HKlVA3SsCY8EMid25wKx+4hnS9buGVr/hFU3EpTse9hDCPsIBQJrQT0g86JEmSJEmSJEmSJEmSJEmSJI8CWTjvlKPXdIz3PVgc6WXy8RXr1AAAalZJREFUoQJiDt+NxqfhcFpOiDmAfwQCZI+SG0KmSR7EhxBzgBEWYstkGvIgMJBnUkTbcDsQGwa1NztNpV1ATMyJP9ITQMJMIeYU4l+IDM4b2TIh39FVnux3nZPrGMUdKA86fB+dvGqKNn11moO6vJba52qUAsvJF07FeRIOKYhNzJvXKTEHUP/A9MRcqDedJ4YwZvKXB+k1L/NleVwrQGVOLU916stOEpp/3weAsubNr69rAPsiyLyrLUFet2wgwagvux8g5oScQ53pg7/pAd9x39LG9w099Sdrk3Q6FmvLbaBn27wvXis2FzoeHq/NDeelvxj5sVAbJdrCyVEFn2wzEo4Q/XCDIPhijAnsV3hLwCgAHYHGzt9Hd+cGVxhZ/Z+d7z/jvgUnHvXN5o6lX8+aWj+WZa2n0B6fL1t9iqQfc0iSJEmSJEmSJEmSJEmSJEn+zLKw7U3HrukY7vKvsk5HzBnhZmQGEx9McgjRwWAyRMkOKguM/FCYvSdDojqCJ46YUAGieq1DO+/bbKK8EUFTfsUVfQJMkkT9A3ildaLflSb7XHEH5Xcg389EnfySqxJmRpwRagDNlwf5k5NxGKMRlBFoTDZ+wJNvCpl7Avn1eV/Wej9uA/nD9Rv0l1r5mqhe5iUC6wGZk1w/08H3g/kTHzldpAdk7o0ke6QQMksQ9J7Mo2vP6RShfwWX0SaM0dqJHdohnZrPrysCXbuQl9S3zQUIUk+S7gsDEbTMpBvlCUbWTcF4Q55B/Y7SdQ/31kujNLe4v0zUYXwg5Na7MqF041Vu9suOGsnmz7+a9vRzCM/KspZjKV1JaCbsTRIxlyRJkiRJkiRJkiRJkiRJkvwZZWHbaceu6dy89sFSfGIOp+SYmAOpJMScnJZDuS8imQKpIWgkRjTvyRAqg2yI2pk+htgbVG++fT72KWmMHDGnvrhvgMck7Rj2eien/XyCjk/KTcppOX7tVAFyzr4LjkH2+P44+R45PR3HRBzmCifm7OSczh36NgKG2k5PzKEcdEwIIs9tFXwd8BeuGcScXZPNVzwnPH8M8eFJJ0Y0BkPcHwDfmg/zKr49Iarwc79PGCGnoD6tznxgXDkdga8ZOu7fIHVcrzZip+C6YBfWR7gmD4wD/TJkDAy6DyE/kAcIOIK183YGuj47GWevs8avtFb59BzZ0BgNTMDh15IJSMvDNO6tXQ9XRrt/Wxm58tfL3nvKg3Nedti/ZQsW3pFlLV20nw+SbT1F0um4JEmSJEmSJEmSJEmSJEmS5FEoC9vf8tw1nbcpMcfEG0g5Jea0zKQcygr5hVYhPsIpsXVcZr0RC6yXMkiRnI2SIEaU+NNehNjOSBkP9RtgftWfggkiTbkNv/qnebRDPj5h5skoJcsm8IMMg5IyKQcijlIm5FCH1Ig1nIDDibno1Bz5lWunMs+l5ckGBAwTNWHc3o8Rc6i3MXGdwF+njdv6YlCdpd5OwGW1s3ljIhIp6uDbA2XUG7Rf75/aGWhe5SQXyoFQa/Rr5ZgwtWtrzO8V1id88HVZmdrGIJ2tOauX12oxVhmjXxdqY9eE9tKfjgvQa7B7ICcn///27jtMkqu+939p4uYk7e5M9/TM7K4QIglkMizJIpgcRBLJJBNMEAIkJKTdyRsUEMIXJ362jOWLjREWyQQh7U6eVU5IILLxfbiOJF+MAe2c3zecU3W6Z1Zg46Dwfj3P96mqc06dqu6e+efznOrWkFZL/0a8qpVyUtKn5T/mEM+f8QBuYLaqMqSzvjF9VHWxX+5xYNJr8OBwGNTVrPMjYdvVe8PAFaM/2bD75GvaH7rjw0XXyjOLovMdRVv3bxdFx075f17h/9ZLtMUCAAAAAAB3IRs2vl6Duff+nT4qp+GRhkrpF0ctUMqqDEL02EI7H6+BlIciPkcZisR9O059Fnqk8TqumtMrHsvYMkBqLenPA6A0V5o7jUnjPISRysO/1or3pq/FVsBp+GblAVy5Oi61L3ifBTUW5sl5Uimca5rX3r/UJv3xPcnvuwz3YkClwVwZFsY2C+Z0X8aXgZKen+bLKr0v/jnEYxlfnqMlc1nwZiFgXrG9LGmz82U/zuGlx3nF9nS9bE5bSSZtFpbG/tay12y1fL8Ga+XrkW35ucpx+Zq0L/3tZPfloZxsU5V93l/u2/nVvXiAWr0O+0EO+fw1dEvhXKMM56pArqnSuXMxhJP5ys8ylj/CKvc4Nbw4MKO//jtqtV3mPfaGibD9qnPC1ote8bOuRx//5WLlygvlf/epUqukdCWc/rDDSqk7e2wVAAAAAADcxazZ+LonnLX1s2f9XV8M5tLjlxbOybE/lqmBkq+S88AjC5p0jI6d0XAhziH7Hg7pcdyPYz040eNY0uZz6nke+KVjqzIo8SrbW/tS0JGPk21aNWchTvwS/aagzO4p7sv96OuwYE7mS4Gcr2TzMEXDFw3m+u1752KbhipynbzS3Gm1oT0GXF5T+9MYv1+fQ+fTY3ktUv76UoCj2+o1pkBJ59B9n1v7dM7Ulsb5OT62umaac0kwZ9eVkmulQM3P8fNbK32+ZZXjs2vEuZqCuWx8a5WfWTpObfaeSltLv1YKucq/KSlfKSfbMpST46ZwTsbFbT5fGciVVb0WD+TSVkoDt7RKLu6XfdYWz0vjrPSHL0YWU9/AzHjYJtfZJtfZLufsuGaPbIe+X//Im7687o1POtS+o/cTRVvXHxVF10RRtD9D/nfv7EcdWB0HAAAAAMBdUOt3Ta1c/7rHv2fLp8/4v/pIon7Xmn3H3IwGc/rdaB7OpbCnCn28yjYLfpqPLeywY+mLwYeHR/FcPSe2e8UAS/flvLJ9VkMdDS/0Gtl43U8Vw40qzErjdJv2Y03LtbVk30rG5qXhWBnMpUoBnFYKXJqCOTlP7q8pmCvfi3jd7Di9R03XtfPT64yvReZuen1yvfJ90Gp5bT5nNX/1+lvGxfPTtf0acRsrBUbansbpvPkYG2fXy8bYPTSPSZ9hWXnfkvuUsj4Pyiws07Kxvu/tHtaV8+j9671mryOdX87RdJ3YV4Z0aVzql/vUrcxThnPxPVkSvs1Vj7OmGtDSR6Htcej4txL70kq6vqnRw33TY4uDMqeHcuPhPtfuC8ffdl4YOPiexfWvffwtbces+92iaH+D/K8+TupBUidI1aQ6pZbT+j8OAAAAAADuorase9OTh7Z++oy/1xDCgzlfJeePs/p+GfZklUKSKuDKyvo01PBjDzpSe3ae9ll/c/k5sWLoYoGUzeNhSVkWxHhwYgFQ2ZfmiPvZnHcezHlA1hTMyTU8lNPyMEYfZU3BnF67DO50rM3lryNdJ13fj1uvmZ2rr1OO/fXE16bXiPPb+yC19HXG0uPUl96vsnyMXlP70/VtbD6n7su1fJuNyeeOZXPo3Gnf2lvuQ9ttvlRVu9WS+5SS8y0sS3NZu2698lCuHu+zrNSu42ItmV/K+3R/6Tj/IYZY8fNIoVwZzKVwLg/mYhinoZwHc9JvVfXp+AGpQS39nGdGDg/MDP/rcbO7v7fj46f9aM3LH/e3xTEbDhRFx/uLov3Z8n+62f9dlyCEAwAAAADgbmvTmgdsOP3Z52/99Bn/6AHQiJSulGsJ5rR9Vvti6X6sKljR/fTIZmxLY2K/hVL6KKk9TurHdl5LedgTj3U/Lw1N0r4FJTE0iW0pWEnhkI2Pc2p7eW/apsfxPAvHmsrbPJSTiiFcuUoqrYIqr+9lwZqOt9fu1/PX6tf09yRe1+bOztVr6vnxuFxdFuesStt1jI6Vyl5H+X7rebYfX2e6D3tPsor3oz8EYefb9b18Dt33cUs+i9iXrpeuY68zlbQ3n5deW3VsY+w8XQU34mGc7lvF+bXiPfg4PZatvBcezEmfbmN7KgvesvN8ntQX555OJcdpP42VSvfc9Kuqct1U5S/2phDuUHM1pOrzY4tyr4sa1g1K2Yq6Q3vCwHXnhuNu2hcGP/jqr6563AMuLVavvKgo2t5dFF0vKYqOJ8t/6aDUkcI3gjkAAAAAAO6GuqWO7n5Q3zM27X3JxZv/+j3f0xDIApj4KGsK5axSINcSylkok8pCmCqYs4AmjbGAQ/a14ve8pWMPTPTY28owJ9+XSsFOtS/3q/ecKvXHMCXdl7XLPFVg5P1+j7F/mUohWRnKZUFcuY1VnqdhU6rUZtdufi3layjPrbbNwZyMzefMqzxnmYqv1T4ju763lZ9bvB///PycdJ0qMMsrO6+1z+bSbV4+f/6aq3M8VEztZcl5Wh7MefC2fDDnwV05Js2V7l3e02XnT1VeJ80vZY+zSvsRgjmt9JloleGtXMtWVMq+rnwb1LBNQ7dDXv2H9sTaazUoNaB19blh8MYLwo6bzw+Df/7OsO4Zj/p2sWbdnxdF+6vl//LBUqv1n1TojzloEb4BAAAAAHAPskHqQatOesBvH/N7r/301r8+64e6CsgDmBTMDZfHHsgNh745rRjSWGlgkQIYLQ/YGincyMZ5KCL7scr9/Lx0rGFIOidWdV481nYNYVqDGAtSdM5snJ0b2y1kiSX7FsDF/urXUj30stBtTrdSZTA3Htt9TCp/fNXDmjys82tW1/X7ivetYVI89kcm/Vo+d+y3Yzk/BaOpTc/RcbJfrbSrKr0Hvq/X0JLzY3m/j7FxNld1TzaPvi/pHN1P74/t58d+jfLaqeI17Hqxz15X7E+r43yFnGxlDtuP41NwpnPbNhufgjn/7H3O9H6k982rek35nHZu2tdrZ98z17Sv19XS+fU9j6vmbD8+5urfSeir5uy74+bGF7U0jBs4tM/DuEP7w7FXnRuOv2l/2HbFGX939PBzrul+xLGTRdfKS4uia0zq5fI/eZz9dx4ZP+oAAAAAAMA9wJaio+OJ6178yNEtl7x5euvnzvx/HmiMhPQoa/mjD3JcVWzLa2Y0VtZmK488YPGAR8MND190XL7fNFesFIRUIU3Wb+doSKL36+Pyar2X8jrpPnVMNp//wEUMt6zGqkrBl1UVulnZPF5N58X+1LdsaX8qa6vG+7leKVyyz0b6UmBlr1PH5POUc0m/vN6msvN0X/tkTHqPyvdDSq+j56bjJXNUfXrc9DlYpbHaVr2/TXPEa6f7SO0pKKsCuqrSuBSSlUFZvq/Xl9dfhbRe6X2097K8T61sXJpjWuZLgXJrxf7GtMwl1R9LA2grmWdgVr87blzexwkdc3hwdnxx24Kvlhu8YX849vbzwn2u3Ru2XPCan3Y8+D4zRceKsaJoP0X+Fx8jpY+r6g866C+tsjoOAAAAAIB7uP5izcqXbnrLUy7u/eQ7v9TzubP+VUMNX5HlodzA7KhVUxjXEsylwGdJ4NUUzGm44uVhUMtxOierFJZYMGNBiran+eO1yrCquTx00v3sOtqeVbovrTsL5SyQy/utzdvz+bTPt/HcrG9J6bl2vgZGqd330zVS0Fau9tKSvqbXaWPiXOkcbdf++Lqt9H3Q1xrfB3s/UsXxfj/xXN3ae5jPkV0326b9srJz7P2Nn6OXtqe5m8da6CbjreJx3l/N0RzMWaVrp/dL3st8398X3ddxsb+lUuC37HUtmJN93Zalx8OhrjWrpa91XD4/D+YG5vaGbQv7wuA1+w4PzI9+v/5X7/jO0Wc96/+2nTD47aKj+9NFW9euoujQQC49srocAjoAAAAAAO6hdhSb1v7W0ec897La58/4Zs8X3vtvGmQ0BXNzGhKNWpt9sb6GD1mV4YUez6TARfel7jSYG5V2rdRefbdcGfbYcdzGsTaPVTrfx1qV43U/zpGNLceVQVZVzSFbDLesrwq60pxlWb/cU6qmdg/vyrDMytv9ejHc02tZe9XWdA0pu387L/bl9yFtZQCVKp5TjsmO7f2N71FTle9rrDQunlt9Bl753Kny/tYxVXt1Df8bqvosFJM2f7zU7yEFdF4pNPNtKp0rPdJqJe+B/xCEvx/+C6rx/YvVJ21e+n551WdSOCdzxr9d++XeWHpcn5L+aampGMjFqk3tXtQakL5+madvYSLsuH5fuM+t54fG5e89vObFj1koNq27qCg6zi6KzrcXRftz5f/vEVJb7D/xyAjmAAAAAAC4hzq+bcumd20+/2VTvVO7/n7rF8/+mYdEI1K/RDBn3wfngZoft1QKNGy8zKFBT+rTY6kUwDT96EOqvN/G6za2Z5Xm8uumfT13mTFZpcdD04q45n4NcOI2ten8cd9XtWmNxK2023jfliFbWbHP5vXyEK7qt3Pi/HmV92/jZZvdR5qrDOW0Ul+sdGzvRfYeNQVz9h6n99bHVG1Suq9tca6ysnvxc6uq/k7S+XGupjHZuTL/nQdzy5UGaXpeS8n72RzOtb5PMZTT0lBOSwM1m0+vLeeWq+JSaQg3EmpaU8OhNinHk3IPU/I3IKU/aKLfU7j9mn1h+23nh8bn3xPWv/GpoTjm6NuKouNc+X/bKbVZqkv/+QAAAAAAwL3bgzoGto71/Mkbb60fGv1xz5Xn3GEBUh7MzabQaSR4qDIc7IcgdF8DDNn6cR6seWDX+uMPHuz4fhnIxHOq1UdpDh9j19R9ObcKizx88fuJc1mfbr3fx6c5pCyAkbZ0HNtSYGNzWHs838pDmzS3zR/7yjAu268eYfXQLQ/m0vXLa+q8TSu5ZC5rj9ew6+nrrl5DU38sm7ecI15nuf00XrfxdaTXlFaJWSBlY2Qbr5/eDz/P7yfdl4+r7rGcP+2Xfbqt2mw+nV+3cY7yXuRYK11T+8rwLe6nseU5tk3BmodstgrOXre+fn2fvdJKufqsloyNY+zvKQZz1ZxS6W/Yrj1soVzP5O5FDeP6Dsrf+ORQ6JPSv93+aybCtquGfrxl9CXTnQ+7z+8Xq1b9YVF0XVK0rdBHVp8m/2/r/d/uiFgdBwAAAADAvciJnccPvK922Tu+U79uz2LPlbsP+yqyESkN5jyc8xoO9qupFsJ52OIhSwxeLJhrXvXmQc5oOcaCjtkquCmDmlgafGiV56dxKSSxkj5p8+AkjovbtJ/uz66p19A5YjjVXBraaF+8Tt4XAxvbz66fxlUr2/Q6qa0K5lJQ1lo6X7lKTrdS1lYGc1rZNXV/mUrXtGObO81RVQrl8nnzStfw0Kv5PU1VhXV6Tmxv3UrZZxDHlRXnt1Au7qdzfFXa0rJgTrdxrB8vvcfq/DhvFqr5dbT8M9RfT9UwzoM5adNgTrZ5pbH6959W2/nqu3gP8vddi8Fc7eDuxdqBXYsayOlx/brxMHD9RNjxkXeEdU952DeKrq4z5X9rh5Sujruf1HapY6Q6pQAAAAAAAMzDux963If6Js/+l8b1e321moUUGnj4qjcN5Tx8isGctsc2C0aW9Hubz5FCkuZjD4bSHNqmWw1bYjAXxzetssrKApN0bOO8Un8ezllAE4OqFILl1RRUpTZ9D+x9iG3ZPebjvaq2cqWchUHSZteNwVisat54bPek+3FsPLe8l2w/Lx+X5kjXkj6tpuM4Ro/Lc+N+fE1l6CVteWDmAVl8D7PxZZXtsdL73dQm17LxsV3HyLHNO+1/T2k+u570p+tWFe8vVjlWqpzX2lN/rPK91dfvqxM9mMsCubLiqrr58dC3EEv267FsLpnTHs+eHAm1mbHQuzD6095Pvvs7R5/9gutWPfaEW45as/7LRXv3B+X/6qH+73WnWB0HAAAAAMC93KNWPu4Bf9p/zfDPG1dPBPuyewtwNPDQwGQoNKw0jNPALAUpVZhiwVisFLpoWZsFJNUYD26yMTYuBUVZf6wUzFlfDFt89ZIGMNqWxvn8Pi6O1bDG2mQbQykLvmIQ4/vel5e1W+glx3pOnKe859iWysKadFzO63M3H7dWOi/ux2O/7+ra2p+3lX1xnuZ2Pc5eXxqr2/x+UlustDIsf51lWXs1thqvn4EepzFVVavMWtplviWljy/L35MfN1+nufy15tfWbdmv52qbhstxP4WX5fuipavmZFu3VXNZ8DY/ERpSfQt7Qn1hIvTOjy3WbX9PqOl2bsKuOTC3J+y4dk/onzsrbJx43j+33afnT4r29pfL/9Jz2ouuVxRFx5Nk/2j777pzBHMAAAAAANzLPWb1kx/8F9tuHg99h8ZsBVMKuGy1XCwP5oaCB3OytYoBXRqfr5iT8jm0NCDx4zKkiWP8XOm3MRqwVGPKc8s+7/dgpjnESUGTjdP25fpS+JXCqbJknJaO1/5ZDXP8HJ8ju1+9p9hXjrHXEY+PeI0jVXZOms/mlP34uqtgKd2PVHYNW+1V9vt+0zX0WOdpavO503wpJGt9rWWl8em9lbJgTPfjHE3z2ecUS84vt1mVc+ujoBbOxfZsnqbKrpmX98vrTm1lMCevMyt71DW9P7aNj7ZqKGcr5Cbkf2BPaCzs9e2hvWHg0L7Qv7An9MyPLvZeNX64/6bxO+oHz/i3DSMnL7b/2nGHi87VlxdF+wvj/9JyCN8AAAAAAMARPXbNsx526cAtE6E+r4HGUNDvlkuhmv7CZL8+bjgdw7gYvtVl378LLgvyrFJbrDI48X4LfuIcGsqk8E3HeL8GKr6fgiBrt6DF5ylXaln5OWV4l43z4CfNpePiWB03N+Y1r/vSJmXXbgnJ/P68yvtNfVnZubofgy97nHXeqykQi5V+GMJ+2XXW98u54rXK12KvQfv1XDmWsXn4lq7tr7ulz0rPqfqq1xn3rVJgpp+Nv1fV69P96j2wsdJeBWPeZr/YG6+lQZyN0zltfPx+NqkUwNlqtzhnavPX4FW9Bj32e7cfd9Br637rNt5POY+VzFHu6zgf63NI6b0uyLmHtDScHrdgTkO5xlX7Qt/V+8PgDeeHwdvfF7Z+7O3/Z9XLHvmFomfD7xQdXfuLjpXvL9pXvUH+h/T745ajoVy77wIAAAAAACy1c80LH/VXjdv2hJqGKTMjwX74wUI1D2Ma/65gLp5jAY/2eyDiPyYRwxo93/r8emVJn9Vy7bavYUq1n/dbgJTaZD+fxwIlPbbSPikLpmQ771sPvLJK48o5ZIwcN9eYbXVc2RbDJAvajlD98+PSr6XBnLRpaV8+j5RfO82fxmX7VtJf3q+HV1r59fw1eRhVVuvrjO9P+f7KNgWmre+1bn0FXPw8msrP8TBOx8Rt2s/L5o7zlW3VHGV4avd+hLKVcHK9WH6vXvaZaU172co8u47fX01ef10+//q8zLGggdxE6D8kn4uuoLtqbxi8+fxw7DcuCAMHzglr3/q0fyu2brq8KDrfIv8z95FaK1WX2iq1UgoAAAAAAODfbeealz/2E31f0WBOwzQP0Dxg03BOaplgLn+UdblgzvpsXAxJsmDOgj0tG5eVza3buG9tsWS/DILinKldt1WQ5cepz+8lq9hXhnEayqR9C4O0PPDK5/TX5m39MsZ+kTWO0/Oa+qTdgzkdE8fFslCuDObG5RydI/XH8XEuv980fzVHFcppOOX366VjtS0bW/Zpu8+l4+ryWehqMX+9sd0qvU9yzXR9fd9lP1V+XxauxffUy9u8vM1COa1sjiXHqc0qzpvfn92/vmZ9/NT3U6VQLgVzZSCnZaGcfHb6eKs9Lqur9kYshPZzdA6ZU1c2xoBu4Ma9YeCGPf/ac/GbvrPyuQ+9qjhmw6eL9hUfKdq6T5P/F/2F1SPR1XE8ugoAAAAAAH5pO9f/1q9f1veVfaFXw7cYoHkQVX3HnIdwQ9am3zWXr5jzGmkuC9k0CIkhi8zXWlWgkrd5eOIr6nzePLDxcbrNwhfdl3atdH1btWftPra1qtBMjnXb2h/LX0t1XIZuemznVwGYVWy3lXAtfR7WtZSFTR4w+blxK1Xej92jjy8DNnndHk5qVeeUpedYX1bSZo9uxrJ5rc/f+/K9tbn9fbTPJFbTuLI/ziHn2Oe4pNJ8cVw2Pj/fPtsl4/QxVSl7zem1j8tWSo9TfzovXs9ef2xLwZyultO/J32Utndy92JtcnjR+tJ58l40rt8Ttt2yP9T/8u1hxXMe8X+KtasvLgr9UYeOR0s9Xv5XHiS13v5rlkcoBwAAAAAA/j06dq5/61Mvq39lX+iZ3B1sJdXsiAU7/bLfb+GchhpSFsrpKroYzNlKuhSgeYhWlgZrFq556LIk2JGqwrYUquixB3NN4Y+1+zleo2X40hzM+bGOz4M5rSrs8fJQLZYeW3tzmFeuGIttWr5SLh7HsKypynHL9C1XOi6GTFo+b5xfyq7fMt7a7HWme4vnZfvpvptKzveQL6umMfpey9be87jN+1rL+rV0rAZzHs7Z56rbcp44R9q3Sud7X1Mwp2XtVfiW9qsat/aqL52jFecu/y5j2Ksr5vTvdUo+Z11FJ9va1O47ti4M/aRvYddPei961b+seuJD/qHYuOE7RcfKjxVF+wviP8lyCOEAAAAAAMCvqKNj54Z3Pv2y+u2twZxWCuJSiKKBmQYd8bFWDT2mPaTLA5AyFInBnFbzCjgPUJrCm6ZARa+Vys/PQ5cytIpza7sHcbFsP42XioGUhVYx4CrDLm3T83WcnhcrXSfNUQVjcb+cy89dWr5CLq+8Pw/G0ngtC5lkbF75dZtCqPj607n266NxP42z1WRpvLbJHLYCLVXsK8M02/fzyn35W6ju19+bPHzz0nOkTyuNsf04V2y387Tf2rw9jfP7yFbBpf14z9Wjt1Vbeh3Nn0X87GSO2tTwYu/k8GJdv2NuyqtxQH/UZDw0rt8vf/Pn/Gj9u591bef9+j9adHZeULR1ndnWtvLMoug6Rf47jvSjDqpNinAOAAAAAAD8Cjo6dq4//Vm2Yq734DLB3IxXCt6qUK4K41LglpedlwVzGtL5Y6o6vwYwWSiTygKbaj8FLCko8vDG22w/zW19fp4HM14p7ElV9VVhTh5OpWt7MJefm4KfsabvllsulNPxFo7F8Y34XXIWzmmfnpNVOT6rFDZ5yZg0LjunufQ1tMxhFd+38r2L1TKvjbPy9706jpWCOdn38s9tScW+9PnZuDSH9eeVtdn4dB9jy5fcc1XjRyh9Xfr6x+01e/A3HnqnRhZr07I/ORoGpe/Y684L26eHwqax5/+84359N8n/wEXyn6Cr43ZIdUutk9JHVnUfAAAAAADgv8jKjp3rz3rWZfUv79Pv3gr1ueHgK+M0YMmDuSyMs0dYh7K+GKKlEEiDFu1rCeYsnIvBTApwPNSJK+1im8+ngZWHNRbYaHt5HPcteIl9Lefl55bzZn0pOLOQKo6zQM6qGpcCMg/jfOshm+6ncT63lY7XtnRONt5/GEL7Uumcen5z+Rzap9fW4zh3vH7qt+vYuKrP79GP05g8nLPwzObQbVbWp6+jei0eiMlWK/XF99Hf+/T5pdLz0hjfWrvs69bf2zQ2tfs8+TX9vvV75OJ3ydkPNOg9y/F8DOH03rQ/hXLaLlWbGVvsn9kj7/dEGJjdGwZkvz4t5149EfpvOS/UPvuuf1r35qd8rX2gNl+0dV1cFF1DRdH+LPlP2OL/EMviRx0AAAAAAMB/gZUrd64/57mX1b+0N66YGwr62Gl6hNVLAzYp+44u/W65oaZgzle2VaXhi52nxzOjVmVgo6X7+XHenu8vKZlHV2+lYxnnIY8GRX6cxnlYFSv1SZXhV1ZVW7znNIeFVhoIVWM8jKsCsLw84NI+HZONaxnfHM7puOa2FJT5ObHKebyvLDtHg6uW9qzKVWtL2mI4FlfEVW1xbLoX64tj5T1M85QBWywL/+Ln7Z95nCdV7M+v4dV6rKXBm5bs2zlS8jrr82OhNjcuJVvdXxgLvbF65kYXe2ZGDvfP7ZH3ZI+csy9su/r8sO2W94f+A7vDulN/4+dF3zHXFG0dFxZFuz6q+mCp46X6pFZIHQmhHAAAAAAA+C+wduXO9cPPu6z35gn/jrn5IakYTsVArlpJ1hzMpf4UaOWBmYdgUhbSNLdXQU3WZu0t+61lIVQVnKWxHuSkcCiOlbYqxPL2FBaW12+q1JeNyYIpPa6CtBiSpeOsPJjzAM1WyZXH3qZjlgRzc80r8MoVbdYWK87RfM9eR2q3igFZa5VBWyy9XhWKeaXX7o83x3Z7v7PS46x0bnsN5XWz91vbrF3HVtfxkra4n8b5arpY1qehnAZy4zGQG/c6JMeH5D2Qtvr8qI1pHNr7s8a1+/+l9onTf7T2tSf9oOg55ivFUSsOtrV1n1sUHU+Sv3xdBbccQjgAAAAAAPDfRIO5sRdcVrtpLPROeTDXN++r4MpgzsKV6th+kbXpO+ZS6JXt27FuqzCm6Vj6PUTzIMfDnNgm1fR4a1k+T3lOOY+2t5S0pVAwzW/j03489v04RoOxWBpIpaCsLG2XrT1WWY73R1U9QNNxHqDZ46vlHN5mpePy46bSsX6OXV/b0tzZGH//0r0015KgTLfy+pYP0+Q9burT8Ks6tvc6rlDM2630HvPrWnvWFu/R7iFvl7Kx1hdL29JY2dpnOi1jtNK92nmjoabB2/x46JkbW+ydGT1cn58I9QUtbZf354Y9YfCr+0LtY2///opnPnS22LD2T4q2rr1F0flmqVcWRcdJ8lff63/8y9JgjnAOAAAAAAD8N9i4dufGfSdf1nvDWOidHgq1ean0PXMzGqQ1r5hbvrL+2bSfgjjd10DGj8vgJa/sHNtPJcce7Gj5HPn+LxfMxbAr7qf+MgRLFcOwFMiV+1m/bu3cFMJZtQRzUh7KZWNsDjm2oC3fb6kYvDWNSeOy4/I1ppLxrW0ecnk1BWpa2iZVk89LH1tu7o9Bmc5jj7h6Ve3V/pJrpn39Tjh9r6U0cPP9vE3m0K2eI9ualbZ52T0vF8zJ+6DhW31uj9zThLSPLfbNT9iqufpN+0LjaxeGxufPCat+86SfFps2LRRtK/YURcdvyF/5gJSGbbpKTh9Z7ZACAAAAAAD4H7Zx7c4N+190We91ox7MzUnNjgQPxjSU0xVzKXzz1XPpe+XSmCpQif3W5+d5eCaV2m2cjq+CPQ+HqlDPAzSdQ8fp+doeAx0Lf7zPQqw4XxkQzWlwFSte28fpftXXHJDJfgq3dKtzxCqDsqbxXv3zHsLZfuzTX171x1Kz0jlS2Vxxfhnn9+37zffj53pQmObwNj/2vqZQTsfHff88ZF/fN30fpfRzLUvedw3DfBs/wziuaV4N5WTf2q1f92Of3oNVNd77pC0GcXXZaulxOi/9vfh9xFAuD+Hsuwz9vrR6p3cv9k4NHfbr6PX0V1f1++UmQu1L+8KWG0f/ZfMHXvOVlU/9tfmjjjn6QNG56kNF28p3x9VxG/0PfVltUqyOAwAAAAAA/0M2r9258fwXX9Z7zUjo1SDEVsx5CGPBmpUfe9imwZm3W5vse6UxHsClc5aumFumpN3CuTRPVikQ8iAoHlvgVI2xiqFPCsDs2K6bHacgqQy3qvHVmOocrywMs5K2pn0N4rKx+blNY6tzquDLx6f7Ks/PxpdjY5Vz2jnNfVY2nwdkXv6elaFXKmnzYCyGc/r+p7GpZD7/frm072Xh2J1U+uGG8v0uK92fzhnvreW+/Bd/PSi0e5seDVsnhxd7Dg7f0ae/rjqjodxEaNxwbui75byw9SOnhe6nPfzrxbq1Fxdtnb8df2H1QVKDUluluqSOhFAOAAAAAAD8D9q8cefGC19yWe/CkAc0+h1zc8PBf2RhmWBOwxOpKpzLwjQbU4UvHuKNWvlx3E/js/OWC+YsqLE5tdK8o9K+zBxxbArFNJTza2t7bEsBkQZIsa0qPz9dq2yXscuX9nl5MFe1l+dbf36Ol/b7GDm2cX5P1XF1Xmvgltp1XGo7ciinx/rexpL3oqmkrVpFtzSc83njsexbzcs9adm9ZZW1HTGY0/641Xvz1XK+Yq5Xqlw5Ny1tU9o25iXHPZPSPjMRBmZGflT71Lu+s3H85G93P/GEfyjWrbu+aFvxB0XR9VL5a9ZfVz0SvjsOAAAAAADcxdQ27jz6Ay+5rGd2lwcw+qusGsrlwZx+55wcpxVzFsal/RkvD+HSsYZOqS0FRn6cVrulvhTaWUl7HhzZSrzUZ3P4WAvmyuMUQHlZsCVj9F78errN+uaqYMva7TpV+XneVwZuOm5Oz5W2+TSHzhnnSXPGMT5W+9IjrnKsc5Tl4z3I8vFWsa+1XT+PdG51Pd/aeO2L59t7Yedlc2ibVva+2Xe8adn72VLapu+dVTwnzaXz6o8spIAuq8b8uPRrybFuZ8flfqtKfRbapevL9TQY7JWy7dTIYv2K4cONaRk3NRp6J8dCr8xbv2V/6Ll6LKx5x9Oua+vburdo63xjUXS/oyg6X1cUHU+Tv+TjpH7R6jiCOQAAAAAAcBfS78Fc78w5wVatzQ3Z1oI5C+fiNu1LpbAsD8DSvodbMUCyNt+38bHPg7tY8TjNmcKhMiyL7V4eJJVh0RHKrhfntHtq6fPQLI2T/lipLVVTqDYnW6kqmBuT+bTf9z3Aaylpt++gS3OkSgFX6otzW5X9VVkwp/s21ue166eK8/pjpqm0z/fL4NLea38/qmBOt9X7Xlb+PjfNGa+fBXLlKjoN5spwzkO4vPTHG+ry/tlWXpM+Mt1rNRx65W9Ogzm9dp+GcQdkzOy5of+mC2zchrEXhuKEgf9XdHZ/QP5q9VHVRH/QQQM3vi8OAAAAAADczfRv3LlRg7npc4IFY3NDsvUVcrpSzkO3FBz5fgrKylBtRvelpqVdS/ftvGyM7cfj1KfzyrFVCoFsvM9vZcfelkKjPDCykMhCMN9PfT6nj02Bod+XtjdXa1+/zFX+smoaZ/N7XxmMyXXtHtIYO9fP01CvdQ5/H+JYbY99ZVssew0yr4Vm8p7nIVl5nvSnsMzGyfVsJVrWXgZi1i8VX6cf63upc+rc3mbXkLYl77O9v3qv1bV12zcvfVZ+He8bj1svOz/dp4ZyUvrLqjXdSnvP9NBhu7b8rdWnh0Lv1FCoHZL5bzk/bPnk6f+08lWPnS8aWz5RdHR/tmjr/p2iaH+G/NVqCHdnWB0HAAAAAADuBiyYO+Wy3qkYzNmjrL5izh9ZjSGaBXVaox62xEDHQq04JgVcqc3aYxhVjrH96thDMzkux1VlQVXsK8Mim/sIY/O2NLal0rjqBxvieXKNsi8LlqzNgqi8tK/5uksqjsnb8nstz43z+PtVjdGyQM4+h6xdxy8pDcZkm/XbijZpt9VpNpfPV+1Lyfvhlbel0uvm19aQTd+PeK1YOr9f36t837TsnDhW7qe+oKXB3Hio6fhZ2U7uvkOv0zu9O9SukXNu3xtql78rrHvX00PRv/m6oq1jl/yVPkXqMVLHS22S+kXBHAAAAAAAwN3A9s0WzPUc1EdZPZhrzKfQTFfMxdKASIO52RQsaWgzHNL3wDUFc6lPz4uBVzlG96WtHGf7S4/L68RjD+XSONm2VHVfWjo2zddcaVxaFZeOq9JVbrItwyXdb64UhJXnZPvl/ZRtcZ54nN+rHetc1q73G8+3a7RU6iv7ZXzZ7/eTn1sFcxqexYrXWC6YS+9z9X7LVuez0nuUmtH3I95vOa9foymU09eczpFjW1G3MG5V1+3chI/TH3qQuWrXjYeeG0d+dMzvvfZrK05+9C1HNTZ/p+hcfW3RtmKP/IU+0v9Ql2BVHAAAAAAAuHvr2LHlsRve/9K/0mBOwyEP5jyM8yBLS/ZbgrnUruFdNS6GX9lxWvGWVtxZqGNtup+XtKXK+3Q/VbxueX5W5WOyei0bd+flj5vmbVWgVAVwVb9/35xWy1ipphV2Vnq+bz1405L9OJdW+VqXKz1Pthpalcex0qOtHobdWcVAzAK6ZfrS3GVpWyo5ju9lWkGXjlN/Ge5lry1V+fr0XvV6dh/jOn6xsRDP1/Yb94T6VydCzydODd3PfuiXinWrxov29hcVReerpV4pf56Pldpsf6hLEcoBAAAAAIC7t47jNu/ckFbMzQ2H2tzuYCvdYijmAduRgjmvMiDT4xSSpf4U6MTzqwAnjamOyyAutaV5y1AoXjcd29jqOlY25s5r+ZVyWtru4VoKsVJfFcwtrfR9clUoF+eKben15NcqX2Nr6XWXayvLw67lgrnmdg3DYpVtaZy2ZSX3WgVu2aOv8n56ZZ+Blt1X83lW0uefvezPjct1qrLvnpuWsfN7Qt81+w43btl/R8/H3h1Wv/xJoahvuaNo7/6Q/Dme4H+VAAAAAAAA9wIdx/U+buP7T7mstymYiz8AYSFLXD2Xgrm5UamR0K81G4MmC27iWH381fZjaVhjwU1q8xBH933OOIce2zze79+HJsdxvJ2T9XuIJ9t4bd/XMR6WWWA2lyqFcR6U5cGczZtK26XKa9prrUpfr5XOEa/hlY2J5XM2z5dev1Z5jVRyThmcxdfuAVncl/eqXC0Xzym36Vz7cYUx+4GFFMo1ZN/Lj9OYKrTzSqFe9Z10upWS9zQ97mpl10ztuk37Pq42NbzYmB6Xfan5CZlP9/0aO244Nxx727nh6PNf9vWuRxz7sWL1yguKtq4/amtb+b+Kov0F8ufY7X+Vd4qVcgAAAAAA4J6h44T6SRs/8IpP9U76d8zVUihXBnOx5lJ5sJRCqnL1mo7R8anSedLnYZQfWwilbVopULJ+bctCNh3XWqnfjtPYVNof55GyVWyp5jQg8+tZf9zaHHZf8Rwbp0FbNr4s7/MwMo4py8eUAaCOj3NqVa8hXkv6y9dgFY+l3cIxaUsBnFceyFXjy36dTyuGbRa8xTa7FzlOVQZyVvp9b1o63qtaTSfts1qyb+9tDOHk2Lb6/XDT2i7Xj9Unx/WDw4vbpsZDvwZzsq3P7QkDt7wvDN6yL2z9w98KK5/58FCsWXNpUbQ/V/78Nkr1SB0rpY+sdkgBAAAAAADco6VVR6u7dh5/yqY/eM0Xeyd3W/hjwZz9oENc+aYl7SmU80AurZbTfg+ffL+lrG3My8ZokKMhmodLeWhV9elW27LAKpY/UpnafT+t1ErnputoYJaCOQunYul51h/3y2MpX0mXxus2Pzfb1zFS6b78/Unj0pj4+uPc5diyv7p+6vOQLYVwVRBnW3udzeekYM5DOLmfPHSzNm8vg7mFcSnp04rj6vrdbzLGVtqlbSo91squVZM5a3IfNfkb6ZWq2Wfg1TctfxdTw6H/gLyXM/vD4DUXHa5/cfc/bHzvybd2nLhtuli98lDRvurTRdH5W/K3t87/BJfQX1xlVRwAAAAAALjH0uBjddFdHLfy5Ie+c9OHXz+vwVxttgrm/LHUESsL4zRM0nBOxnjgpiVtqWSM78e+ctxYLG/3YM2DJQ+txmRfS0On1FadXwVRVQBkYZS22Xyyb+XHfp7eR5orlt1/tp+3leWr62zlW36uVHVvUjImnePhl1xTj8v+2C771ev1MemRYO3P50iBnJ0T22yctFlIF193GdqVxzpO3j+r/Fxvq1bKjVt5OOf7Fszpo6YWxGlA51WGe6lkHg/pNJgbCb1yzV75++iZ2r2oK+fss50eDn3yN9RYGA6DX9oXGlfvDRtHTvl5+/36ryg6ut5bFB2/URTtT5d6pvzt3U+qXf8Ql6F/mwRzAAAAAADgHktXJTWKTauftuatJ73/6I/+9i09UxrMDUvtDn0zQyF9V1y5Om5OS0OvvKStDMDkWAMaC+fisbZPx5L9FFBZSFXOOSbHWtrv48r5Ytk5MreWh3NxLmuTfSs/tnPLe2gpu17Lflkatnn5yjlpi9V6b+Ujr+W5ct2WeSwcs/PkvmLpONvXrfbHc3xf+9I5sS0734M4Lzu2iivZyrFe5aOtWtn3zeXhXNr3YC6W3LcFc3r/sb8cJ+dr+Yo52epjrFbDoT4lfzcLI4v1W/bf0b8wfMeG9z77+0c9cNvfFStWXFUUHWfJ35o+qrocQjgAAAAAAHCvo6uVjm/vO/oV64aed8kxl536tZ6pIQvm6mnF3IwGcsMxmPNQyYKlGC6VwVsZgHm7BkNVMCeVApx4nMKmFGb5sQZhMcySappPw7ZYNta2Xn6NNF4qC+jKcbLNQ7r0uGr5mGs6167rfWWbVJorb/NVdfk4vZa+Dr1/addgSyu1xfttOo6vtansfJkvVuu5vjouzZPPF/vkdaawztrze5HSYM2DNt2mVXAxiEur5bRSGLegNWHVOze2aO0ytjEj26kRC+Tq14yH2tcuCLWrRsPa337qV9u39X6s6FoxXBRd7ymKzrfJ39ljpI703XEaEGsBAAAAAADca2gwd2LbcbW3bDj/lE9u+dRpf2PfFzZX/fiDhk0pmLMAbLlQTqp/Rh/79PAuhUgWpNkYObawTMaWbRoapXligBTnSpX6rS8Lxnxs6pO27Doeyvl4C/G0dJztN8+/bAAXq7ktzV+1+epBD7yqdr1WNYevANRKx/Fe4rGtekthXKymVW5Lzms5Vyu+rmqs9MU2X1EY78Hm1K0HcbavwVzrcQrrUmkod0jLg7na3NhivwZzdh157ddPhMZXzw8D07vChjOeH466/8CPihUrPiJ/Ws+Xvy39/jj9hdVjpNZIHemxVQAAAAAAgHsdXcG0s+uhx44cc/FvzW793Ol/b4+xzo+E+pz+8MNIaMi2DL20LQZzrSGVBnNpXAqRLDDT0mMNibIxZYAWjy1U0n4bk88Vx8WxZeAW98t+Gxv74ni/fpqrOvZrpuvESnNl5deP887JsVRaYbd8MKfX1etXc6XzfOWbz1U9fqrvc2zXbSxtS+0ewOl+HGdjqzn0WuVjvfG12Xmy78FcrJb5072lsE7LXktcJafVOze6WDu0N9Su8urTcE5X1M1OhMb154e+6/b+uOfP3v7t1S974g3tjd7PFyvXfLKtbdXFRdH5Kvm7Otr+wpZidRwAAAAAAIDQYO6k7ic84IItl77txq2fO+N7FvjEYE737ccfym3c12BuzoOo8hHXGT/2sbFmqn0Py5r7PMBqOdYxWXnYlCobJ/tpvIVwehz703lNY2283kM+X7rnX1D6WuPrrdpikKWhVtau86fvgbP2eK6FYvE+LDTTrYWcHpTllYdov6jSI6tV+bUttLPS+2uu1KerBf1+tU1KXosFb/qa5iZCbWb8sAVzV+8NvfMeym2/4dyw4yvnhb7LzwyrXv74vy22bvqLoq37nfKndJL8LT1atjtlq98l1yW1HL5PDgAAAAAAQHQWRfuzV7/gkRf3XHHmt7d+4awfp8AorZjzx1nzVXNaVRDlj7lKvwZisS0Fbd4Wy75jTsOxqi0FZSlIS0FbOtdDNQ+RPEjS8/wcH5udo/PZnNm5rWVjqvl8zl9Q+lqz11u1V8Gcl7dXc/u1ylBOK/allWwW3uXtsv+Lgrl0P3mbnRfnKOctX6vcm1b8Vdy+vOQzaehnEr//L62y03EDsxNhcGZC5tobts5M3NF3aN+/9l878cOej536r6te9Ji/K2o91xddaz4if0Kvlr+jfv9zWoIQDgAAAAAA4Ag0mHvh2t980qW1hZF/2nrl2T/TgMtXcmXBnP4IhOxr9c+Nes1mq+W0ZnQ7Wo7TFXRpjPVZCDQa9AclrKS9DK8sYPNQKYVbZciV+qRs/mxsKr/mMufG/dbKQ63mvuaQrWyz8uM075JKc5Zt6fVVVa1m8/L2FKr5SjbbT3PJvZT78bvfGlmlX1q174KTsV4yR3m+3LO+Hpnb7im+d+Uvu05rDYXa9HDovXLocO/B3YsanvZPSU2OhIHZPWH7jReFwesv/Onm8191Q9fjjv+zYsO6DxRtXWcX7SveUBRdL5S/oQdIHelHHTSU47FVAAAAAACAZXQV7V0vW//Wp3++dvPen/RM7lrUMCcP5jyQ0q2UtPensvYYzmnwZpWNtf0smNNj3cqxBVKyn0Isb0v7qeLYGCZ5xeM4tumcOCadv6Q/K2tPwVVLX2sQ11r5fPk1mirda3qtUs2hnI6r9q3sffdK97ZcVcHceAznWiofq6GctFX3KteT+9LSH/fQQK42tdt/VfXgrsX+g0Oh78BuGTccBm/dEwZuHg89f/zm0PXkE28v1q373aJof578zdxfaoNUp9QqKf1xB8I3AAAAAACAf6euorPztRvPfsFC31ffF2rTQ0FXYel3zJU/8KDbuWz1mx1XwZyGZ2U4F49TgJb3ezDnx02BWzy28MjaPMhaMs7mlDbZt7F5lWO0T4OoMWlPIZUflyvhtPQc65OSfb+2Xjc71q1dz6+b2vNK7SnssvNtvJ6vxykI81WCv2ww5xVXv1mfzCX7FrZp0GYh3HhoLMj+gm69+uQ4hXM6pt9K+mSu6gcjRkLv9O7Fnqndi/UpmXNyKDSu3B0aV+yy+6/dMPKzoy9727fWvvWk+faH7Zgr1q79dNHWNVYU7c+Vv5et/mezLFbHAQAAAAAA/Dt0FatWvemYc191U99XNJgbDn36ww/zGjx5WOar5EatUlhWtmm/HHtol/XLvgdbWt5mv9pqwVkc03LsoVY13ttiyGWl+14ejHng5mN8XArtUtmc2XHZLvfu5WFdGpPGl5VdM42xkMv6q/bW8utoX1YytjWY80rn6bxLy8dKxXu261swl1UWzuXVf0i2UhrY2eOuOlbnnBw+XJ/avdg4MBzqV+yWeeU9vW1P6Jk5J6x5zzO/V+zo/XjR2f22omg/uSg6nih/J/eVqkmt0D+aO8H3yQEAAAAAAPySuts2rnvr1t9/41cbt10Q6hrMLYyGulS5Ym5WtnPZijkpf5w1BVCpLx2n82SOPGgrg7MjHMfxGmClczzQqtrLPjsvq9TW0l6FZFJZX1PIpeHcrFbsL19XmjNeM/VbYBbvSbctVY3Tc1L45mWr52xfx8ZtOY/Mqz/IYPPnlfqlynv2fS851lVxMYRrXJXV1V59h2SMrq6Tcf0yvn96JNSmhkNteuTn/VcP/6D28VO/u+60p//wqPsP/N9iZffHi6LzDfK3Meh/IkdECAcAAAAAAPAr6G7bvOHtW/7kzd+o33qer5hbEsx5yGSPq1ow5AFdWS2r55oCKQvDYpvs66q58jzty/s1eJLjFMClMKsK1NI83uZzV31pfHl+OYeM1fPK/fw8Py4DsBR2xbnyEDCdu1ylgM/PTSXn6nuYtaVHXqvStpbSca0lY9Mcfi/+euw1yHUtcFuYCP1XSV29p6zGNROhdmj8sFbjKjmeGw990+Ohce25YfvXfycMXDn0k/WveMLBtp5N5xVdK4aLouOMomh/vvxdPEj/Nuwv5MgI5gAAAAAAAH4F3cXWTaduueSt3+y75Tz7lc70eGTDHmfVSiGQh0y+oiyGa1rLBnMxRJqRfS1tm9HjKpjLQ6/yHDnOQzXry8bkYVkKqXyMjo1l/T5Hmrfsi+NTyLV8+Vi/VnWN5caV+xq+pSr7YpVtHrAtLW3PysYtX/7+VFWu3psbD42FiWDh29V7Q7/VvtB/7bmhce3eUL9mItQXxsOOG84Nx//NhWHb3K6w+Zznhc4dgzcXHV275O/gwVKbpFbrH4XQ74ojeAMAAAAAAPgv1N3Wt/nULX9+6rfqN53rwZz+gIA+7hjDtipE88AqVf/MsNWA7Gv5CrjUHwOteJ6FcdafB3N+rMGSj4v7dn4Vetmx9KegKx9rx3Gc98s2zutj4nlzXtXc2TyxNCRLc5RjbF/nyUrHxSr7s1AufY9cGq+r5tJ9eLufm0K5mj5WOt0cwFm79qe+WH3Lld6LXmd+PNTnxxf7FsYXt121L2y/el8YmNsb+q46Nwx+6aJwn6//rzt6/+ytX179vEf8dXv/1s8UK9dcVhQrzpa/gZ1SvyiEI6QDAAAAAAD4T9bdNrj11K1/edq3+m7YFyycmh8NugKrX2s2BmnSbqUBkJWGWh7MedAm42akXcv6PPRKY1O/jdF+G5P6Y7ik58TjJe1xmx9bm8xTlYwvz4nb1r54XN2DjvWqgjnt0+v7mOrcWNk5emxj02OsZTgXx8qxPcoa2/PyAG7EArgjlgVzqfQ74YZDfara9xqysfE76RblNR0elP0BXSF32/vDfb/2h2HbFeNh41ue8YO22paLi/b2F8rn/lipR0ltl9ooRfAGAAAAAADw36y78z69p/Z84t3frF+3N1igpd8vN68r5qRmNZgb9RBtJgZsehxDLW/LgzkPs3y8jvOx+Qq5FHjl/Sno8uOsZEzZF6va14BLAyktD8MsELNz9NzUpyvVsrBMK189F8drUOZ96frV3FXp+FhZnz/iO+YhXBobA7ny8V/t17H5DzrIPL5ybqQlkMuDNyn9oYapIaveqd1e07tDz8Fdi3XZ75MxFpLKvgWmN+9dHLh57Htbf/+131z14sd9t32g/7vFinV/WRRdL5HPfJV/9EdESAcAAAAAAPDfoLvzfv2n9nzq9G/Wr56w4MiCuYXWR1lHLeiyVW8WLqVwzMOtMpiz/TguHqe2PIhLwZqPq4613+bSIMvGxgArVmor++S8PNzywEzbpV8q7ZfBW+qP5+u9+HE6tyq/Tz+nnDsra4tzeaXAzR8Frkraym22H89LoWF6fDWv+kwVzPWmUG5SayhsPXjOYu8V59zRd3B3aFyxKzSkfeCW/aHxrfNCz1+8LXQ/6YFTxbo140Vb92nyMZ9aFO3PlM97h3/sAAAAAAAA+J/W3X1C/6k9nznjm/Wrxj2IskdZl/uOOQ3R4nHZ5vtVMOdhlz/6mo79HA/hvD+dnx6RXS6YqwIvb8/3UzWFZKlSv8xn5+g2VmuYpveRh3Gt5dfIzm2tcq6xpbUkmMuqPKeau+nRVrl2fU5KtnrcOz1sZSvnDg6HvoNDoX7gnFD7wjmhfnBX6L9xT+j/8r6w5U/eFLpPfmQoNm+6vWhbMSqf7wP1M5Zq1w8bAAAAAAAAdx3d3Q8ZPLXn82d8s7agj1qO+a+yzlXBXBWiedhlIVtss3Y7jqFbdpyCNgvqtJr60/l6nvbF/dSn52WBlt2XHXubBWZacp5Xtl/2efDl4Zq2e4DnAVgcYxX7rH/Yqnneamz+ow5+ns9Vl2tp+Vi9brznWHW553o61tck2/KRVjsnXl+qpoGc1rzsS/XK/tap3Ys9Vw4drk9qKLcr1C4/O/RM7Qp9N06EHVJHn/38m9sfMvi/i7Wr/7ho6/pIUXTsknqifL4d/jEfEY+tAgAAAAAA/A/p7n74fU7d+oX3fEsDIQ3UfMXcaHMwJ9sqXEttGqZ5qJYHc96exqWwLR43nevnNK+Yi312XSkN4fIgrrXkvDI805Jzm49jWywP5rzurM/6U+XXyyp/BLW6VhW6+Q9CxPAthYo2xiu9H9W5UjKmPj/mtTAaalK98+OhZ2Z0sefA0OHeK4dCz5S037QvNK7dGzZ/4NVhxUm/9q/FmjW/L5/lE6QGpe4vNSClP+rASjkAAAAAAIC7qO7uR9/31J4vnPmt+uxw0B98aH2UtTE3YqvXLHybTSGa1nBoyDkapFWhnYZqqV+O9Xw91m0KouzY+z2Yi+dpn1Ucb9eOWyk7L/Z5kJWFZvE8D9jysMzb+uQ+tXxVXAzemuaRttZqGqP3EkM2K10BV4VzebiWwkRb2adhXNwvQ7p8pZyutJvWa+k8MZCbH/c6tEdqb+iVz6NHv2tufiQM3rwn1C4/7R/Wv/Npt3Q94rhbj9q46dvFUasuLYr258fP886wOg4AAAAAAOAupHvl4+9nwZz+yEBzMCdlq788ICuDuVQakkk1raTTNjlOIZmdG9s8lNPyfh8vZeNiCKZjUp+emyq1x2Ofx9tSsFaVB3N5COePqMo2jdH+rGyesvQ4VryOXVfejzx0S2XBXBoX+8vHVK38OC97zFZrWq4v77veW21qeFHn0/MH5sblPZW5J+X8q/eHxm3vD40b9oRNF7wqdJyw/cpi5cp3FUX7S4qi8zWyfbp8jrpSDgAAAAAAAHcj3StPeuCpWzWYmxoKvsJrxLa+Yk6DJA+cPKjLwrkZryXBXCwL0rTK8SmoqkKwKoDTMdKnY1Kftccqx3ml0MtWtWnpnGVVfUtK+tPjpx7IpfJ50rgjBXMpOEvfw2fBnLxPFs7Ze5fuT/riuUsr3oN+j5xcqzY9ZL+8qvdj79eUvJ+Tw6Fx1djhwVvP+1ljavwnx7z/9aH76Q8LxcbNXy/aVpwtn1vdPz4AAAAAAADcXXWvfMovCuY8jPJQToMnDc2kUpBkx7ov51p5WwqiUphWVexP46VNwyoP9HxMGWJJm81nc7b0SVVBWlapP/bpirS8r1xJp9usPDBL+3q+z1PedwziGnPjZaVwzsbHUM4qvTabQ+eM19UwrqxhubfR0Du5a7FP3/uDUpO7w8D1E2Hw6+eG2udO++eVr37cwbZjax8sVq/dX3Su+YOiWHFOUXTsjJ/dcnhcFQAAAAAA4G6ie+XTTvBHWaeGLIxrzMdQTsO4WBpKlaHTrAdrTcGcbrMwqhp75LIwy/bT+TpXS39qz/rSeal/SbX06Xe4VX1eKYBbvuIKOCsd7/v5irlUHsyl8foepYrn23W9vwrk9FhX2sm4Kdlecc5i48pdYfCa8dB3+/7Qe+A9Yd3pzwwd92tcU3R1nSWf0QOlVkn1SulKuXVSBHAAAAAAAAB3c92rn/WQU2ufO/NbfdNDoX9eAygp2Q5YQKfBnAZiWeAUwyYL4iy4Gpb9uCpN+nWMrazT4EnOsZV2GvTF821+qTLwkjmWC+a8tF22FvZpv4+pQq9U2XFr/y8I5prG276vgKvCOa8ymNPXofvlOA/cbEx6jbHSuRbG6Q87pDn0fiflPbtyJAxesycM3n5B2PLnb/n7lS991LVH9W/+TLFyzcfbihXD8vk8wj+mZRHOAQAAAAAA3I11r3vew0+tfTYL5uZHslVzXnnYVNaMhmZV2JXCumXHptJQKlYKrarQrAqyvKp58+Ct6ZppbNZfzpUfS6Vwz+eM23yOVHZv8f7StVvHSOnrsX0NGZeUzGth3Ujond692DutP+wQz5kaCvXJ4VC/cSJs/+p5Ydvl7w3r3vLMUDS2fL5ob/8t+Uw0jHtcUXQ+VLbH6Id0BARzAAAAAAAAd2Pd605++Km1z5wRg7mR0JjLS8M0D5TSarVy5ZoGXRp4WbglZQFY1m9jW+rOgrklAViaV8do+bg0V9O4NEaPy62PL8+zc72szeaPY1PpfWXBXPMPRGil+eQetGSMvZ75cSndxorvW5qvMTse6lMjoefgyGL92n0/67vpvH/detnpP1zzhif/oOP4wTuKrtVfKoqOd8rnsdY/liUI4QAAAAAAAO5hPJj71Hu+1dBgLlvpZcFcDLL8UdS0gk5/iTWGUjHsslArBVetx3nbMmWPgsqYI5aNkbGyn89XXlfPT/3W5mP8MVUfY0FevJ8UqKXjNL8+ZuoVj3WMbmekTWta9lPZuTpP/AGIednGqs2OLmp7Y35C+qVvcre9Tztu3Bd23L4vbL74dd9d8RsP/kJx9IaLijb9DrmucfkYNJR7uH8ky2qTIpwDAAAAAAC4B+le/aJHnlr/pAdz5Uq52Ww7Wz3S6gGdB3P9Gkxp0CX7GnZVwZeXBVdxTArhlqulwZwHX+W+jfFjC99SxdDN5pBxeV+aq7x+U+X3W82vq+QslGsJ5voslPNgri+W7dv5vrLOVsXpL7XqyrnZ0cVBaW9Myvsk2x03XBR23Pr+0HvpO8KqFz36B0f1HH1F0dZ5mrz3D5LqkFottUlKf+CB8A0AAAAAAOBeonPdix/19tonT/9m/8xwGCgDOF0lp6Gbl/0IxGzWJ/vWpyGXBWSyb6GVh18pjLPQLvaXAVo89jDNj211m5zj4Zu2aTAWAzbb13Pj8fRwHOdzWriW5rIxPt770jgvX0WX+nxbzh9Xy9VieUAnbdJXk7E1eW36C69aNb3O7HDonRpe1K2e2z83Lu/fWBiYGgmDcu72r5wbBq8Z+8Hm81/ztRVPPXG+2Lzpz4u27t+Tt/wd8r6f6G//stqlCOgAAAAAAADu4drXvORRb+u99F3fHJhJwZvUbBXKebhWhWxlyXhfeebhlwdzfpxWzDWFY6kt7pdBm8yVgrOq0rmjZdBWjl+u8nNaSu877es8/supcaVbup/Y5+Xz5SvnLJiTcWXNDoeeqd2LGszpr9LqHP1Tw6H/0HjY9rULwsDt+0PPJW8MXU978JeL9et+T97mV8l7/WtSJ0jdT0pXyB2JhnIEcwAAAAAAAPdw69f+5uPf3XPpu77dGsxppRBt+WCupcrHO/3YA69qDu+rxqRQTcekMCwFaHmVwVycqwrkfOWc1RHOLX8JVSqFbP7DDhrOSV+6H+2TsSkYTPdjoZzVSOiV61gop6vm9J70vMmhUL9yV+ib3n1H44axH/cc2v3DTRe+/AddTz7hq8WWjbNFW/fvFkX7s+R91sdUl0MIBwAAAAAAcK+0YsW2DW8+aaR26Tu/MzDjj68uF8xVNRxshVjW15A2qxjMpaDOA69qXFnaJpWCNq0UsKW+fIy1x/Aslbfl5de268cVcfZrqTGES7+OWpUHdqn8enEr86Wwz4I5GavhXM/U8KKunLPHdQ/sDrUrz7F5tt1+fqhdPfrTtWc+++a2Bw9+tFi7+ryi6Pztomh/pdTT5V0e8Dd7WfyoAwAAAAAAwL2M/ujAqo7apsdsOvPZF9U+/q7v2oq52RjMZd8vp+Whl4dWjdlhqxRiaVjn4ZwHY2mc7WvY1dQmpedJpVVpTX1Zf155OGcr31rPKa/tY6rHVWNpSJfV8sGcl31nXCxdKWcl89oquYO7Q+8Xd4XeeXkfvrw3DB4aDlvPf2XoePR9byvWrvlQUbS/RN7X46RS2LZCSt9rwjcAAAAAAAAYfbSy0XV83wuOHnvxh3s+8e6/11DOQrcYzOXhnIZf5Uo4C+X0e9W87JwUzMWQy1aclYFZFXpVwZdutbRfyn5UQR9NjY+n2hg/TuFbFcjFa2il+ae9z/rtHqpAzh5RjYFc/7y3pWDOQ0gN8VJI5/dYm961mO5Pf9ihJvP3HhwJvQsToe/LF4T6gaGfrD/9Od/ofNix821r1/5RUXScXRTtL5T3tN/e3eWxOg4AAAAAAADFWqn7r3jEttdtvvCVH+/91On/1K8B2+yIB3O6tUohXAzmYruHZVkwp+3W5uXhmodcR6oqnJPj/LyysqBOS0My2dpY2beyYC4L7Cy006qCOa/x0JgfD/1Sul8GdbpNc1n53LXJ4cMa+tUmR0Lt4K7QODQWtn15X2gsnB2Onjg5dD/0vv+nbfXaPy2KztfJ+/gAqT6putSRvktOEcoBAAAAAACgOLroKHauedaJ52z9ozcc6P3Uu79fhXG+cswDuBjM6fFM6vd9W+EmVfbrd6/psQV2GqR5pbDLfsU1C7/8BxR8P4VrZfAW95uO7Zx4HM9Zvsa8ynBuPDSk+ucmymrMx7KQLp43NRL6p0fDgLyO/gNDoXZgt17njoGFPf/S++kz/mHtmc/+u84Tj73tqLVrZ9qKVR8uivZT7H1cnoZwBHEAAAAAAABYYmvR3f70DS/becHW//3Wq3o/fcaP+mdGQv/MqK2YK4O5WGUgp5XCMQvgsmBO2rwvC+fS2DJI8+MUunm/tGu/Bm+xtK3qr+awvtjv4Vs8XqZsrD66Ou/BXFoxp9u+BalDUrJfmxle1ACvf3okDBwYDv1XDoXBqybCttv3h8aVZ/18/RufenPbcX2XFKtWjRdF5+uLouvFRdH+THkP8++Sa0UwBwAAAAAAgGXVizVdL930lqf8Sc9l776t97Nn/j8L5WZGLJiz75hrDeRSWVCWwjcN5bTyAE2r9TFUby+Py0pBmuxbVeFbuaJOy86PfSmUWyaYS98V52P9cVUL5XRlnIV00n7VmNfCRGhYyTmTI2FAatsNe8O2r54f+r74nrD+7OeEjgcN3FKsWvOH8fvjdvhbZwjeAAAAAAAA8B8y0L5p1RuOOed5n659/uzv9H7hvT+xR01nRqpgzlbOxcpX0c2kH3tIgVkK6VJA5u3Nj7JmbRbYVaFd2WfHHsiloM2CujjO+sr2MTnWYM4r/dpqWTLG73ncSh9X1R+B6J0dWbRA7qo9oXFItpNjYWB6NNznlvPCfb5xQdhy6du/v+oVj72p/djalcWKVR8qio7Ti6L9+fp++du2LA3o9IcdAAAAAAAAgF9oe9vWdadu3n/KZO3g7n/s/cJ7f2qh3KyWhloe0C0XzPmvsI6EaiWb7mubjLFQTdvSSroqZPNALYVsVeWPslaPqKbSYK65fEw+zu/Rf9BB7zW2yX6fhnJSFuRJ1aZHDsu5i41p/7XWY790Qbjf33wwNL54Tljzjqff0XbfvpuLzpUXFUX7y+Q9ur/UZqktUnf2ow4AAAAAAADAL+24jsFNu3v+vzfcUp8b+7feK8/+ua6C819m1VDLgzkL2Szs0vBLAzcPxTyY8/2+LMTzkE5rmWBOxi4N5bJ5rMa8ymBOV8bFUG46nR/ns37f15DNfsRBa163Mk5L+/QasrVHdQ8Oh/qU3O+N++7YduOef+7/07d8a/1rT/rWUcc1bipWrvlkW9F9blG0P0Penw3+Ni3BI6wAAAAAAAD4ldxvxQNq+/oufee3+67eE3qvPOewBm/6vXL2CGsK5mRrYZcFYClIixVDMV8hp22yb+2xP+2nkM0CttQn+3quneP7ZQCn+/mxVE3OaQ7z5P6k3Vfoydy6Sm5+LDSyqk0PH67Njiza/Adk3OR4GLjlwrD9m78bGp9677+tfcnO2bb6MRcWbV1nFEX7b0o9vSg6HivvTb/UnT2aSjAHAAAAAACA/7AHrnzowAf6v3DW9/qungj16SEL4zSU658b862uPJsfsWDOVsRpSJcFY8sHc37s3y0n+xrqacVx/r1zGrzFc5raWkoDtVi1WHmbX88fUa3LPdf111at9obG/J7QmJ0IfQdGQuPgrsXtN+w5vO3L+w7XPnVa2PCmp4aO/vpVRceKcXkfdkrp46oAAAAAAADAf4sHr3rUjg/1zw7/rH71eKjPtARz8zGYSyvmLFzT8C0vbcvbZd/aWvu9PHBrGSv7Vdim+36sK+SspuVYH2FtKQ/p/DHXmsxZmx5drM+MLw4s7AkD+r1ykzL3wng49kvvC8feeMHhYy569VdXPOVBf93Wc8xfFm3dv18UbW8tio4nyfuw0d+OI+JHHQAAAAAAAPCf6iGrnnDcxQPX7Am1Bf2BhDyYi5VWzGmYpuHaTArVYqXATfZtBVs89jYv3U8/IFGtkkv70m/nxpI2rTyUq8UQzgK61CbbXhuj7fpLq1Izo4t9U6OL/forq3MT4djb3xd2fOP8UP/YO8Kalzz+H4ueY/6iKNpfK6/78VL3ldoktU6qSwoAAAAAAAD4b3Pi2mc+5JL+G/aHXg3J9FdVNZib9VBO9xsWzGm4pv1aWdCWjq0thmxa+l1vurrO+vx75dJ5Pk7DtxjMWZ9uNXRLYZ7s2yo5bdNtCuaGQ6/cY14e1knf1HCoHxwJAwt7w8DN+3/c+9enf2XDWc+ZXvmMh1x7VH3zdUXbij8uivbXyGse9Jd+RHx3HAAAAAAAAP7Lnbju5Ede0rj5PF1t5gFZXDHX9KusWjFQS+Gb/kiE/0Jrc7uVBXNS6Qcf7EcftKTPgrfm6puW87V03yoGcha8efiWVshtmdq1qNtea5Ox+gurX9xl9zj4tQvD4K37w8b9L/1OxwmN84vu7qfLa3xqUbS/qCg6fl32dZXcan3hd4JgDgAAAAAAAP/lTlx3ymMuadxygT0Omr5Hzh5h1fAtD+a0YvDmj6V6MJceUfW+Mdn38mBOH48dr9pi8GaBnFyvDOd01ZvWTBXOeV8VzGkY16Olq+c0lDswFHqu2BV6Du7+ef/1e3/SmB3+4cYPvm6x+3kP+2mxcd1H5bXp46oAAAAAAADAXdKJ6171uEsaX7ow1KZGm4K5MnyzQG7E2i2oS2GcBWgapGUr3TR8i9WY08pCOa04rnxUNZ6vAZyX7Mv85ffMzQ6HmrRvndy1uHVq96J959xBD+S2Sl/j1v2h9sl3fmfta3/9so779o4Va1aNFUXHuUXRfoq8tl/mV1ZZHQcAAAAAAID/ESeue90TLum79UJbhWY/8pAeZU1BXArmUlBn2xTKxUDOQjktDd9iWSCnwdzSFXNloCfXbGjNDPuxzTEagzndjtqvrfZM7V7cekDqStmfmwi12y4IvVfvCZsufM1i96OO+1zR2fkqeS1HS3VKbZXSX1nVfQAAAAAAAOAu6cQNbz4pBnPDTcFc06OsFsx5cKYr5fzRVTmeGfFHUrUsSBsL9fRdchrE2aq55hVz6fHV9L1yDX2EtpzDA7ne6ZHFvplxm6t3UtquGPJfh715/OebL33z11e/8YmXtz9kx2XFuvUfLYrOt8nrGPCXs4SuiGv3XQAAAAAAAOCuorM4ceNpz7ikfuv7Qm1y2II3LftV1ljpkdYUxOkvsebHZThngVsM5lIQVwZyqTyE80rz+b4/2ipzTI0s1g8MH9YfddDHVvuungjbbr0g9E8Nh/WjL/xh+/3rf1B0dDxN7v4hUo+Tup/UkX7QQYM5HlcFAAAAAADAXUxn54kb3/PsS/pueV+oHxwOjZkxC8r6Zds/qxXDuZYwzh9prYI5e/w0BXNyXn0uK+2LlT/Gqt8dp2XfKzeloZy2y/UPyNxXyv7CxGL/deM/7v3Maf+4afTkH3X9+oN/Uqxb99miaH92vPvlEMIBAAAAAADgbsCCueddUr/pglA/4MFc//RYGJCtVlMwp2X7LaWBW1lj9vhq+SuuUnkwZz/eEAO53pmhsHVq12LPgaHDGszVJneHPrmHbdfsC9u+9r7QN7MrrH7Dkw61HdtzQbFixXuLomu0KNpfIXd9Z4+ttvkuAAAAAAAAcFe2qvPEjWc+75K+G2Iwp9/5ZqHcuFX5KKutZtPyAM4fb9V277NwzgK58VhVMGfhXBnQDdsvrfbKdTSk01VyvVfsWuy7YigMXLM37PjGhaF/blfYOPqi0PGY+/5T0d19ntzlA6U6pNbpHcd9AAAAAAAA4G5sVeeJG8583iX1a8+PwZwGbSmY88dZ/ddUR4I+xprCOH8k1SutnLMQbn48NLQsnIur5+bTY636ow7DizUdq98nd2Ao9BzYHRrXjodjv/a+UPvI2766+uU7P9P+gMbHi9XrriiKzt8tCvsuuSPRH3Xg0VUAAAAAAADcDWkwd87Jl9QXzg31g7paLgvmYijnv6iqwZz0ld8rlwdzGthJzcn+vIzXmtMaDfUFKQ3mZA4L5SaHDvcd3B1qB3aF3mvGw7bbzw/bp3aFjUMvPtx2374/LdrbnyF3dV+pRxdF50Nl26O3eQSEcgAAAAAAALib2rDqxI1DL7ykNrcnNA6OhP4Z/U45DeX0MdYYys2P26o3+w45C+U0wPPy8bLVUG7OV8r1z42F/oWJ0FiQ81LJ+bXJ0VA7KPNcuz/Ubtj3o55L3/G3G0971t+s3Hn/fy5WrzlQFO0vj3e1HEI4AAAAAAAA3INsXH3ChuEX/Wnv7EQZzGkoN6AhmwZz9iiqBmtSehxXyKVVc3psAd6M9+v3yzXmNJTbG+qzE4cbc3tCfWo89E+OhG3XTYTBr50XatNnh9Wve+Khtv5jhov2zlcXReepUq+Uuzneb2pZBHMAAAAAAAC4B9EVc6MvvKQ2PR76Jv1RVlsxNz8R+qXKR1MtfJOa1hoNjamR0DelK+jGQv+Mh3gWzM3KeXN7pCZCbXJ4UdsGbzkvHHv7+0P/FWeHDcMnh87H3PdHRZf9qMOg34TpktLvjAMAAAAAAADuBbasffTGPaf8eW1qPDRiMKePo2ow5+GcrpjTwE1Xx8VQbnok9E0Ph7qU/SCEtPfNjoc+XSmnwdzUWOifHg+DN58fdnzjfaH3o2/6m9UvedSn2rbX/qjoXvXxomj/ID/qAAAAAAAAgHu1jm1bnrLp/Fd+vHdSV8SN2C+rWjA3N2GVB3N16a/rr6nGUM7KfgxiLNSmRhZrcqyPtW675YKw7ZsfCPXLzwwbz3h2aD++/pmiaH++XK4upY+rPkhqi9SREMoBAAAAAADgHqtDanPnw7a9atPvvf7zFszZ98npNgZzuvrNHk8d9UBuSms41Ce1RvzRV2lvHBgODWkbuPW80H/jxPe2fPgNX1/9uid+qf0hA98tOlZdXRRtb5NrrbSrLsXqOAAAAAAAANyrrJd64MqnPPBdm/70rdNLgrnZiTAwE4O5GQ/meg/uXqwfHA59B7RGQt8Xh0L9wK7QuH4ibL/9fXJ8dlj1ysfOt/VtGi7a219UFO2vLYrOV8l1HmJXXF6bFMEcAAAAAAAA7jX6i/b2Z6x+6c6Ljvn4u27snRwN9Xkv/bEH+2VW/U45e1RV2vXHHnSV3JXDofGFXaF++TmhftVY2HbruaHxmXffsfGs54X2E7b/bdHWtU/mvq9fAgAAAAAAAECrE4ru7jevf8ezLtty+dnf1GCuT4O5hVH7XrnBOamZsVA/MHS4Pjm82Dgo/V8cDvUvDMu4PaHvxv2Hj/mD19+25gWP/quOwS3vK9q7f6coOiakniJzH2kFnD62qivkAAAAAAAAgHutR7etWze8cc8p8z2zo39fn9JgbsRXzM2OhYHpiTA4tSf0XTF8uP75XYsDlw+FwUPjYeDGPaH3s+8Ka9/+9H9q3167uCjanytzbZXqldoudbQU4RsAAAAAAACwvI6ntvce84c9F7/527Xr9/xYf9Shb0ZqeiQMTI2FwStHQv/l+l1yQ2Hguj1h+8L4v239/dd9a/Upj5luv1/ti0Vn14dlkpdKrbHpltJwjoAOAAAAAAAAaNb+nK5j639Z+/Q7v1+/dd8d9quq+j1yV+xeHLx8OGy/YjjsuHpP2H7TROj77DvCujc/5btt9WMulBN/Q+rXi6LjybI9VupIj61qOz/qAAAAAAAAALR4VucDBj7S+9nT/7Hvxn0/HzgwHAavHAqDB0bCjqsmFo8/tOfn/X926k83vf0ZYcXDdoSivevjcs5j/dQlCOEAAAAAAACAX9JJnQ/c9sHey9757YEb9obBqeFw7PV7wnGyf+zH3hk2nfL4r3Y0tnyyKLr+oija/qAo2vWx1Q1+6hI8tgoAAAAAAAD8kh7adp/au4/50OsnG1/a98MdN+4NfX95ajj67c8KKx64/XtF0fVRGfObUo+S0kdWV0sRvgEAAAAAAAC/okZHbf2TN77lpLM2vPsZH1753IdPtR3XuLXoWH1tUXT9cVF0vkbGDPjQJqyOAwAAAAAAAH4FugKuVmxZ/eBizYonyv6zi6L9RVLPL4qOJ8jxcXFMK75PDgAAAAAAAPgP+I+GaoRxAAAAAAAAwK+AYA4AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD/2Yri/wdBD23oFd67FQAAAABJRU5ErkJggg==" style="width:36px;height:36px;object-fit:contain" alt="Voco"></div>
    Voco
  </div>
  <div class="nav-links">
    <a href="/inbox">Panel</a>
    <a href="/privacy">Privacidad</a>
    <a href="/terminos">Términos</a>
  </div>
</nav>

<div class="hero">
  <div class="badge">Plataforma SaaS · WhatsApp Business API</div>
  <h1>Automatiza tu negocio<br>en <span>WhatsApp</span> con IA</h1>
  <p class="sub">Voco conecta tu cuenta de WhatsApp Business con un asistente de inteligencia artificial que atiende clientes, procesa pedidos y envía difusiones — sin intervención manual.</p>
  <a href="/inbox" class="cta">Acceder al panel →</a>
</div>

<div class="features">
  <div class="feat">
    <div class="feat-icon">🤖</div>
    <h3>Asistente IA 24/7</h3>
    <p>Responde consultas, gestiona pedidos y hace seguimiento automáticamente usando modelos de lenguaje avanzados.</p>
  </div>
  <div class="feat">
    <div class="feat-icon">📢</div>
    <h3>Difusiones masivas</h3>
    <p>Envía plantillas aprobadas por Meta a cientos de contactos con métricas de entrega y lectura en tiempo real.</p>
  </div>
  <div class="feat">
    <div class="feat-icon">🛍️</div>
    <h3>Catálogo y pedidos</h3>
    <p>Integra tu catálogo de productos, gestiona carritos y recibe pedidos directamente desde el chat de WhatsApp.</p>
  </div>
  <div class="feat">
    <div class="feat-icon">👥</div>
    <h3>Multi-tenant</h3>
    <p>Cada negocio tiene su propio agente, número de WhatsApp, configuración y panel de administración aislado.</p>
  </div>
  <div class="feat">
    <div class="feat-icon">📊</div>
    <h3>Panel de gestión</h3>
    <p>Bandeja de entrada, historial de conversaciones, escalaciones a agentes humanos y reportes de campaña.</p>
  </div>
  <div class="feat">
    <div class="feat-icon">🔗</div>
    <h3>Integraciones</h3>
    <p>Conecta con Shopify, HubSpot y Calendly. Sincroniza pedidos, contactos y citas sin salir de WhatsApp.</p>
  </div>
</div>

<div class="how">
  <h2>¿Cómo funciona?</h2>
  <p class="sub">Tres pasos para automatizar tu WhatsApp</p>
  <div class="steps">
    <div class="step">
      <div class="step-num">1</div>
      <h4>Conecta tu WABA</h4>
      <p>Vincula tu cuenta de WhatsApp Business API desde el panel en minutos.</p>
    </div>
    <div class="step">
      <div class="step-num">2</div>
      <h4>Configura el agente</h4>
      <p>Define el prompt, catálogo, integraciones y plantillas de tu asistente.</p>
    </div>
    <div class="step">
      <div class="step-num">3</div>
      <h4>Empieza a atender</h4>
      <p>El asistente responde 24/7. Tú supervisas y escalas cuando lo necesites.</p>
    </div>
  </div>
</div>

<footer>
  © 2026 Juan Carlos Soto López · Voco ·
  <a href="mailto:soporte@myvoco.ai">soporte@myvoco.ai</a> ·
  <a href="/privacy">Política de privacidad</a> ·
  <a href="/terminos">Términos del servicio</a> ·
  <a href="/data-deletion">Eliminación de datos</a>
</footer>

</body></html>"""
    return HTMLResponse(content=html)


@app.get("/privacy", response_class=HTMLResponse)
async def pagina_privacidad():
    """Política de privacidad pública — requisito de Meta App Review (Tech
    Provider). Servida desde myvoco.ai/privacy, no requiere sitio aparte."""
    from agent.legal import politica_privacidad_html
    return HTMLResponse(content=politica_privacidad_html())


@app.get("/terms", response_class=HTMLResponse)
async def pagina_terminos():
    """Términos del servicio públicos — requisito de Meta App Review."""
    from agent.legal import terminos_html
    return HTMLResponse(content=terminos_html())


@app.get("/terminos", response_class=HTMLResponse)
async def pagina_terminos_es():
    """Alias en español de /terms — URL configurada en Meta App Review."""
    from agent.legal import terminos_html
    return HTMLResponse(content=terminos_html())


@app.get("/data-deletion", response_class=HTMLResponse)
async def pagina_eliminacion_datos():
    """Instrucciones de eliminación de datos — requisito de Meta App Review
    (Tech Provider). Configurar en Meta: URL de instrucciones para la
    eliminación de datos → https://myvoco.ai/data-deletion"""
    from agent.legal import eliminacion_datos_html
    return HTMLResponse(content=eliminacion_datos_html())


@app.get("/health")
async def health_check_deep():
    """#77 — Healthcheck profundo: verifica conexión a BD y que el loop de
    seguimientos siga ticando (no solo que el proceso responda HTTP).
    Pensado para Railway (healthcheckPath) y monitoreo externo — a
    diferencia de "/", acá un problema real devuelve 503."""
    problemas = []

    try:
        from agent.memory import obtener_agente
        await obtener_agente(1)
    except Exception as e:
        problemas.append(f"db: {e}")

    inactivo_seg = (
        round(time.monotonic() - _ultimo_ciclo_seguimiento, 1)
        if _ultimo_ciclo_seguimiento else None
    )
    # Si nunca corrió un ciclo (deploy recién hecho, aún en warm-up) no es
    # error — el primer ciclo tarda hasta CHECK_INTERVAL_SEG en arrancar.
    if inactivo_seg is not None and inactivo_seg > WATCHDOG_UMBRAL_SEG:
        problemas.append(f"loop_seguimientos sin ticar hace {inactivo_seg}s")

    if problemas:
        return JSONResponse(status_code=503, content={
            "status": "unhealthy", "service": "voco", "problemas": problemas,
        })
    return {
        "status": "ok", "service": "voco",
        "loop_seguimientos_inactivo_seg": inactivo_seg,
    }


@app.get("/webhook")
async def webhook_verificacion(request: Request):
    resultado = await proveedor.validar_webhook(request)
    if resultado is not None:
        return resultado
    return {"status": "ok"}


def extraer_marcador_botones(respuesta: str) -> tuple[str, dict | None]:
    """Extrae datos del marcador [[BOTONES:...]] y lo elimina del texto."""
    match = re.search(r'\[\[BOTONES:(.*?)\]\]', respuesta, re.DOTALL)
    if not match:
        return respuesta, None
    try:
        datos = json.loads(match.group(1))
    except Exception:
        datos = None
    texto_limpio = re.sub(r'\s*\[\[BOTONES:.*?\]\]', '', respuesta, flags=re.DOTALL).strip()
    return texto_limpio, datos


def extraer_marcador_catalogo_cat(respuesta: str) -> tuple[str, dict | None]:
    """Extrae [[CATALOGO_CAT:{"categoria":"X"}]] — abre catálogo nativo de WhatsApp."""
    match = re.search(r'\[\[CATALOGO_CAT:(.*?)\]\]', respuesta, re.DOTALL)
    if not match:
        return respuesta, None
    try:
        datos = json.loads(match.group(1))
    except Exception:
        datos = None
    texto_limpio = re.sub(r'\s*\[\[CATALOGO_CAT:.*?\]\]', '', respuesta, flags=re.DOTALL).strip()
    return texto_limpio, datos


def extraer_marcador_producto(respuesta: str) -> tuple[str, list[str]]:
    """Extrae uno o más [[PRODUCTO:nombre|presentacion]] — envía single_product del catálogo.
    Andrea lo usa cuando menciona un producto específico para darle al cliente
    la opción de armar carrito directo desde WhatsApp (con +/- nativos)."""
    matches = re.findall(r'\[\[PRODUCTO:([^\]]+)\]\]', respuesta)
    texto_limpio = re.sub(r'\s*\[\[PRODUCTO:[^\]]+\]\]', '', respuesta).strip()
    productos = [m.strip() for m in matches if m.strip()]
    return texto_limpio, productos


def extraer_marcador_carrito(respuesta: str) -> tuple[str, list | None]:
    """Extrae [[CARRITO:[...]]] — Andrea lo emite cada vez que el carrito cambia."""
    match = re.search(r'\[\[CARRITO:(\[.*?\])\]\]', respuesta, re.DOTALL)
    if not match:
        return respuesta, None
    try:
        items = json.loads(match.group(1))
    except Exception:
        items = None
    texto_limpio = re.sub(r'\s*\[\[CARRITO:\[.*?\]\]\]', '', respuesta, flags=re.DOTALL).strip()
    return texto_limpio, items


def extraer_marcador_vaciar_carrito(respuesta: str) -> tuple[str, bool]:
    """Extrae [[VACIAR_CARRITO]] — el agente lo emite cuando el cliente pide
    vaciar/borrar/limpiar su carrito. El parser ejecuta limpiar_carrito_activo()
    y limpia el marcador del texto que llega al cliente.

    Acepta variantes: [[VACIAR_CARRITO]], [[VACIAR CARRITO]], [[LIMPIAR_CARRITO]],
    [[LIMPIAR CARRITO]] — tolerante a errores del LLM.
    """
    patron = r'\[\[(?:VACIAR|LIMPIAR)[_ ]?CARRITO\]\]'
    if not re.search(patron, respuesta, flags=re.IGNORECASE):
        return respuesta, False
    texto_limpio = re.sub(patron, '', respuesta, flags=re.IGNORECASE).strip()
    return texto_limpio, True


def extraer_marcador_mostrar_carrito(respuesta: str) -> tuple[str, bool]:
    """Extrae [[MOSTRAR_CARRITO]] — el agente lo emite cuando el cliente pide
    ver/mostrar su carrito por TEXTO (no botón). El parser ejecuta el handler
    determinístico _enviar_resumen_carrito() que lee BD y muestra resumen +
    botones reply, garantizando consistencia. Sin esto, Andrea inventaba el
    texto del carrito basado en su contexto sin reply buttons.

    Acepta variantes: [[MOSTRAR_CARRITO]], [[VER_CARRITO]], [[VER CARRITO]].
    """
    patron = r'\[\[(?:MOSTRAR|VER)[_ ]?CARRITO\]\]'
    if not re.search(patron, respuesta, flags=re.IGNORECASE):
        return respuesta, False
    texto_limpio = re.sub(patron, '', respuesta, flags=re.IGNORECASE).strip()
    return texto_limpio, True


def extraer_marcador_lista(respuesta: str) -> tuple[str, dict | None]:
    """Extrae datos del marcador [[LISTA:...]] y lo elimina del texto."""
    match = re.search(r'\[\[LISTA:(.*?)\]\]', respuesta, re.DOTALL)
    if not match:
        return respuesta, None
    try:
        datos = json.loads(match.group(1))
    except Exception:
        datos = None
    texto_limpio = re.sub(r'\s*\[\[LISTA:.*?\]\]', '', respuesta, flags=re.DOTALL).strip()
    return texto_limpio, datos


# URL base del sitio web del merchant — se usa para construir links a colecciones
# desde [[TIENDA:término]]. Multi-tenant: configurable por env var con fallback
# a Equora mientras hay un solo tenant en producción.
EQUORA_BASE = os.getenv("MERCHANT_SITE_URL", "https://equoradistribuciones.com")

# Mapeo de términos que Andrea escribe en [[TIENDA:término]] → URL de la colección
_COLECCION_MAP: dict[str, str] = {
    # Cocina
    "lavaloza": f"{EQUORA_BASE}/coleccion/cocina",
    "cocina": f"{EQUORA_BASE}/coleccion/cocina",
    # Lavandería
    "detergente": f"{EQUORA_BASE}/coleccion/lavanderia",
    "desmanchador": f"{EQUORA_BASE}/coleccion/lavanderia",
    "lavanderia": f"{EQUORA_BASE}/coleccion/lavanderia",
    "suavizante": f"{EQUORA_BASE}/coleccion/lavanderia",
    # Baños
    "bano": f"{EQUORA_BASE}/coleccion/banos",
    "banos": f"{EQUORA_BASE}/coleccion/banos",
    "desinfectante": f"{EQUORA_BASE}/coleccion/banos",
    # Hogar
    "multiusos": f"{EQUORA_BASE}/coleccion/hogar",
    "limpiavidrios": f"{EQUORA_BASE}/coleccion/hogar",
    "ambientador": f"{EQUORA_BASE}/coleccion/hogar",
    "hogar": f"{EQUORA_BASE}/coleccion/hogar",
    # Talleres / industrial
    "desengrasante": f"{EQUORA_BASE}/coleccion/talleres-e-industrial",
    "taller": f"{EQUORA_BASE}/coleccion/talleres-e-industrial",
    "industrial": f"{EQUORA_BASE}/coleccion/talleres-e-industrial",
    # Higiene personal
    "shampoo": f"{EQUORA_BASE}/coleccion/higiene-personal",
    "higiene": f"{EQUORA_BASE}/coleccion/higiene-personal",
    # Combos
    "combo": f"{EQUORA_BASE}/coleccion/combos",
    "combos": f"{EQUORA_BASE}/coleccion/combos",
    # Más vendidos
    "mas-vendidos": f"{EQUORA_BASE}/coleccion/mas-vendidos",
    "populares": f"{EQUORA_BASE}/coleccion/mas-vendidos",
}


def _utm_andrea(url: str, content: str = "") -> str:
    """
    Agrega parámetros UTM a cualquier URL de la tienda enviada por Andrea.
    El Pixel web de Meta los captura automáticamente en PageView / ViewContent.
      utm_source   = whatsapp
      utm_medium   = chatbot
      utm_campaign = andrea
      utm_content  = término del producto / categoría (si lo hay)
    """
    import urllib.parse
    params = {
        "utm_source": "whatsapp",
        "utm_medium": "chatbot",
        "utm_campaign": "andrea",
    }
    if content:
        params["utm_content"] = content[:100]  # truncar por seguridad
    sep = "&" if "?" in url else "?"
    return url + sep + urllib.parse.urlencode(params)


def _construir_url_tienda(query: str) -> str:
    """
    Convierte el término del marcador [[TIENDA:término]] en la URL más
    específica posible en equoradistribuciones.com:

    1. Si el término coincide exactamente con una clave de _COLECCION_MAP
       (ej. "desengrasante", "lavaloza", "combos") → abre la colección.
       Esto evita que el buscador de productos escoja mal entre variantes
       cuando Andrea usa un término genérico de categoría.
    2. Si el término es específico (ej. "desengrasante profesional" o
       "lavaloza antibacterial 500ml") → intenta URL de producto concreto.
    3. Si no hay handle, busca coincidencia parcial en _COLECCION_MAP.
    4. Fallback: /catalogo
    """
    if not query:
        return f"{EQUORA_BASE}/catalogo"

    q = query.lower().strip()

    # 1. Término genérico de categoría → colección directa (sin pasar por Jaccard)
    if q in _COLECCION_MAP:
        return _COLECCION_MAP[q]

    # 2. Término específico → buscar producto por handle (Jaccard)
    url_producto = obtener_url_producto(query)
    if url_producto:
        logger.info(f"[tienda-url] '{query}' → producto: {url_producto}")
        return url_producto
    else:
        logger.warning(f"[tienda-url] '{query}' → Jaccard no encontró handle, usando colección")

    # 3. Coincidencia parcial de categoría — solo si el query es una sola palabra
    #    (evita que "desengrasante profesional" caiga a la colección por "desengrasante")
    palabras_query = q.split()
    if len(palabras_query) == 1:
        for clave, url in _COLECCION_MAP.items():
            if clave in q or q in clave:
                return url

    return f"{EQUORA_BASE}/catalogo"


def extraer_marcador_tienda(respuesta: str) -> tuple[str, bool, str]:
    """
    Extrae [[TIENDA:]] o [[TIENDA:término]] del texto de Andrea.
    Retorna (texto_limpio, abrir_tienda, query_busqueda).
    Si Andrea escribe [[TIENDA:lavaloza]], la tienda abre con "lavaloza" pre-buscado.
    """
    # Forma con término de búsqueda: [[TIENDA:nombre producto]]
    m = re.search(r'\[\[TIENDA:([^\]]+)\]\]', respuesta)
    if m:
        query = m.group(1).strip()
        texto_limpio = re.sub(r'\s*\[\[TIENDA:[^\]]*\]\]', '', respuesta).strip()
        return texto_limpio, True, query
    # Forma sin término: [[TIENDA:]]
    if '[[TIENDA:]]' in respuesta:
        texto_limpio = re.sub(r'\s*\[\[TIENDA:\]\]', '', respuesta).strip()
        return texto_limpio, True, ''
    return respuesta, False, ''


async def _procesar_status_difusion(status: dict) -> None:
    """Procesa un status update de Meta (delivered/read/failed). Actualiza
    DOS tablas: difusion_mensajes (tracking de campañas) y mensajes (chulitos
    de confirmación en el chat — #79). Los mismos status updates aplican a
    ambos sin distinguir si fue mensaje regular o de difusión."""
    wamid      = status.get("id", "")
    status_val = status.get("status", "")   # sent | delivered | read | failed
    ts_str     = status.get("timestamp", "")
    errors     = status.get("errors", [])
    error_code = int(errors[0].get("code", 0)) if errors else None
    error_title = errors[0].get("title", "")   if errors else ""

    if not wamid or status_val not in ("sent", "delivered", "read", "failed"):
        return
    try:
        await actualizar_status_difusion(
            wamid=wamid,
            status=status_val,
            error_code=error_code,
            error_title=error_title,
            ts_str=ts_str,
        )
        if status_val == "failed":
            logger.info(f"[webhook-status] FAILED wamid={wamid[:20]} code={error_code} title={error_title}")
    except Exception as e:
        logger.debug(f"[webhook-status] Error actualizando difusion_mensajes {wamid[:20]}: {e}")

    # #79 — actualizar tabla mensajes para chulitos en el chat
    try:
        from agent.memory import actualizar_status_mensaje
        await actualizar_status_mensaje(wamid, status_val, error_code=error_code, error_title=error_title)
    except Exception as e:
        logger.debug(f"[webhook-status] Error actualizando mensajes {wamid[:20]}: {e}")


# ── Palabras clave de opt-out (baja de difusiones) ───────────────────────────
# Solo palabras con intención EXPLÍCITA de darse de baja.
# "No gracias", "No", "Cancelar" NO están aquí porque tienen otros contextos válidos.
PALABRAS_OPT_OUT: set[str] = {
    "stop",
    "baja",
    "dar de baja",
    "darme de baja",
    "dame de baja",
    "quitar de la lista",
    "quitarme de la lista",
    "no quiero más mensajes",
    "no quiero recibir más",
    "no quiero recibir más mensajes",
    "cancelar suscripción",
    "cancelar suscripcion",
    "desuscribir",
    "desuscribirme",
    "unsubscribe",
}


def _es_solicitud_opt_out(texto: str) -> bool:
    """
    Retorna True si el mensaje es una solicitud explícita de darse de baja.
    Incluye el botón 'Dar de baja' (quick reply de plantilla) y palabras clave exactas.
    """
    t = texto.strip().lower()
    # Coincidencia exacta primero (más segura — evita falsos positivos)
    if t in PALABRAS_OPT_OUT:
        return True
    # El botón quick reply de la plantilla llega exactamente como "Dar de baja"
    if t == "dar de baja":
        return True
    # Frases que contienen palabras clave de opt-out claras
    for palabra in PALABRAS_OPT_OUT:
        if len(palabra) > 6 and palabra in t:  # solo frases largas para evitar "baja" en otro contexto
            return True
    return False


def _es_respuesta_automatica(texto: str) -> bool:
    """
    Detecta si un mensaje parece ser una respuesta automática de WhatsApp Business.
    Estos mensajes no deben procesarse con Andrea para evitar conversaciones entre bots.

    Ejemplos típicos:
    - "Hola, somos Pizzería Pepito. ¿Qué deseas pedir?"
    - "Gracias por contactarnos, en breve te atendemos."
    - "Este es un mensaje automático. Estamos fuera de horario."
    """
    import re as _re
    t = texto.strip()
    if not t or len(t) > 500:  # Los mensajes reales suelen ser cortos en auto-reply
        return False

    tl = t.lower()

    # Patrones de respuesta automática
    patrones = [
        r'respuesta automát',
        r'mensaje automát',
        r'auto.?reply',
        r'bot\b',
        r'fuera de horario',
        r'fuera de servicio',
        r'no estamos disponibles',
        r'en este momento no (podemos|estamos|hay)',
        r'atenderemos (tu|su|el) (mensaje|consulta|pedido)',
        r'te (responderemos|contestaremos|atenderemos) (pronto|en breve|a la brevedad)',
        r'en breve (te|le) (atendemos|contactamos|respondemos)',
        r'gracias por (contactar|escribir|comunicarte|tu mensaje)',
        r'business hours',
        r'working hours',
        r'somos un (negocio|restaurante|empresa|tienda)',
        r'^hola[,\.]?\s+(somos|soy)\s+\w',        # "Hola, somos Pizzería..."
        r'^buenos\s+(días|tardes|noches)[,\.]?\s+(somos|soy)\s',
        r'pedido.*deseas\s+(pedir|ordenar)',
        r'(deseas|quieres)\s+(pedir|ordenar|comprar)',
        r'menú\s+(disponible|del día)',
    ]

    for patron in patrones:
        if _re.search(patron, tl):
            logger.info(f"[auto-reply] Mensaje ignorado (patrón: {patron[:30]}): {t[:60]}")
            return True

    return False


@app.post("/webhook")
async def webhook_handler(request: Request):
    # ── Extraer phone_number_id para routing multi-agente ─────────────────────
    _body_raw = {}
    try:
        _body_raw = await request.json()
    except Exception:
        pass

    _phone_id = ""
    try:
        _phone_id = _body_raw["entry"][0]["changes"][0]["value"]["metadata"]["phone_number_id"]
    except Exception:
        pass
    _phone_id = (_phone_id or "").strip()

    # ── Sandbox hub de Voco: el número de la plataforma se intercepta ANTES ────
    # de resolver el tenant. Ese número es propio de Voco (VOCO_SANDBOX_*), no
    # pertenece a ningún cliente; debe entrar SIEMPRE al flujo de sandbox aunque
    # exista un agente en BD reclamando ese phone_number_id (diseño viejo) o esté
    # en borrador. Antes esto vivía dentro del loop y el early-return de draft
    # más abajo mataba el mensaje antes de llegar acá.
    _voco_sb_pid = os.getenv("VOCO_SANDBOX_PHONE_NUMBER_ID", "").strip()
    _is_sandbox_msg = bool(_voco_sb_pid and _phone_id == _voco_sb_pid)
    logger.info(
        f"[sandbox] phone_id={repr(_phone_id)} sb_pid={repr(_voco_sb_pid)} "
        f"match={_is_sandbox_msg}"
    )

    _agente_actual = await _resolver_agente(_phone_id)
    _agent_id = _agente_actual.get("id", 1)

    # Si el agente está pausado o en draft, ignorar el webhook — SALVO que sea
    # un mensaje al número sandbox (ese flujo lo maneja el hub más abajo, y el
    # agente sandbox viejo suele quedar en borrador).
    if not _is_sandbox_msg and _agente_actual.get("status") in ("paused", "draft"):
        return {"status": "ok", "agente": "inactivo"}

    # Proveedor Meta específico para este agente (credenciales desde BD o env)
    _proveedor_agente = await _get_meta_para_agente(_agente_actual)

    # ── Status updates de Meta (delivery/read/failed) ─────────────────────────
    # Llegan en el mismo POST que los mensajes, en el campo "statuses"
    try:
        for entry in _body_raw.get("entry", []):
            for change in entry.get("changes", []):
                for status in change.get("value", {}).get("statuses", []):
                    asyncio.create_task(_procesar_status_difusion(status))
    except Exception:
        pass  # no interferir con el flujo normal de mensajes

    try:
        mensajes = await _proveedor_agente.parsear_webhook(request)

        for msg in mensajes:
            if msg.es_propio or not msg.texto:
                continue

            # ── Deduplicación por mensaje_id de WhatsApp ─────────────────────
            # Meta reintenta el webhook si no recibe 200 en ~20 s. Sin este
            # guard el mismo mensaje genera respuestas y tickets duplicados.
            if msg.mensaje_id:
                _ahora_w = time.time()
                for _k in [k for k, v in _wamids_procesados.items()
                            if _ahora_w - v > _WAMID_DEDUP_TTL]:
                    del _wamids_procesados[_k]
                if msg.mensaje_id in _wamids_procesados:
                    logger.warning(
                        f"[dedup-webhook] {msg.mensaje_id[:24]} ya procesado — ignorado"
                    )
                    continue
                _wamids_procesados[msg.mensaje_id] = _ahora_w

            logger.info(f"Mensaje de {msg.telefono}: {msg.texto[:80]}")

            # ── Filtro: respuestas automáticas de WhatsApp Business ───────────
            # Evita que Andrea quede atrapada respondiendo bots de otros negocios
            if _es_respuesta_automatica(msg.texto):
                continue  # ignorar silenciosamente

            # ── Sandbox hub de Voco: número central compartido entre tenants ────
            # El sandbox usa el WABA propio de Voco (VOCO_SANDBOX_*), independiente
            # del WABA de cada cliente. El código VOCO-EQ001 enruta al agente correcto.
            if _is_sandbox_msg:
                from agent.providers.meta import ProveedorMeta as _MetaProv
                _voco_sb_prov = _MetaProv(
                    access_token=os.getenv("VOCO_SANDBOX_ACCESS_TOKEN", ""),
                    phone_number_id=_voco_sb_pid,
                    verify_token=os.getenv("META_VERIFY_TOKEN", ""),
                    catalog_id="",
                )
                if msg.mensaje_id:
                    asyncio.create_task(_voco_sb_prov.marcar_leido(msg.mensaje_id))

                _sb_target_id = _sandbox_sessions.get(msg.telefono)

                # Comando SALIR: resetea la sesión para probar otro agente
                if msg.texto.strip().upper() in ("SALIR", "RESET", "CAMBIAR"):
                    _prev_id = _sandbox_sessions.pop(msg.telefono, None)
                    # También desactiva modo_humano del agente previo para que
                    # el próximo test empiece limpio (sin ir al panel)
                    if _prev_id:
                        try:
                            await set_modo_humano(msg.telefono, False, agent_id=_prev_id)
                        except Exception:
                            pass
                    await _voco_sb_prov.enviar_mensaje(msg.telefono,
                        "🔄 Sesión reseteada. Envía un código de proyecto para conectarte a otro agente."
                    )
                    continue

                if _sb_target_id is None:
                    # Validar código de proyecto
                    _code_recv = msg.texto.strip().upper().replace(" ", "")
                    _parsed_id = _sandbox_codigo_a_agent_id(_code_recv)
                    if _parsed_id:
                        from agent.memory import obtener_agente as _get_ag_sb
                        _target_ag = await _get_ag_sb(_parsed_id)
                        if _target_ag:
                            _sandbox_sessions[msg.telefono] = _parsed_id
                            _nombre_ag = _target_ag.get("agent_name", "el agente")
                            await _voco_sb_prov.enviar_mensaje(msg.telefono,
                                f"✅ Código correcto. Ahora estás conectado con *{_nombre_ag}*.\n\n"
                                f"Escribe tu primer mensaje para empezar la prueba.\n"
                                f"_Escribe SALIR para probar otro agente._"
                            )
                        else:
                            await _voco_sb_prov.enviar_mensaje(msg.telefono,
                                f"⚠️ Código no reconocido. Verifica el código en el panel de Voco."
                            )
                    else:
                        await _voco_sb_prov.enviar_mensaje(msg.telefono,
                            f"¡Hola! 👋 Soy el sandbox de Voco.\n\n"
                            f"Envía el código de tu proyecto para comenzar la prueba.\n\n"
                            f"Ejemplo: *VOCO-001*"
                        )
                    continue

                # Código ya validado: sobreescribir agente y proveedor para el flujo normal
                _agent_id = _sb_target_id
                _proveedor_agente = _voco_sb_prov

            # ── Acción de botón interno (act_*) ───────────────────────────────
            # El cliente tocó un botón de reply con ID act_* (Voco lo procesó
            # en parsear_webhook y lo serializó como __ACCION_BTN__:act_xxx).
            # Estas acciones se procesan directo sin pasar por Claude.
            if msg.texto.startswith("__ACCION_BTN__:"):
                accion = msg.texto[len("__ACCION_BTN__:"):]
                await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)
                # Guardamos en historial el título de la acción (más legible que el id)
                titulos = {
                    "act_envio_gratis":     "Alcanzar envío gratis",
                    "act_confirmar_pedido": "Confirmar pedido",
                    "act_ver_catalogo":     "Ver catálogo",
                }
                await guardar_mensaje(msg.telefono, "user", titulos.get(accion, accion), agent_id=_agent_id)
                logger.info(f"[accion-btn] {msg.telefono} eligió: {accion}")

                if accion == "act_envio_gratis" or accion == "act_ver_catalogo":
                    # Reabrir catálogo nativo de WhatsApp via product_list
                    # (catalog_message removido: cuelga WB mobile en algunos Android).
                    if hasattr(_proveedor_agente, "enviar_catalogo_productos"):
                        secciones = obtener_secciones_catalogo(None)
                        if secciones:
                            header_cat = await _catalogo_header_for_agent(_agent_id)
                            await _proveedor_agente.enviar_catalogo_productos(
                                msg.telefono, header_cat,
                                "Agrega los productos que quieras — se suman a tu pedido.",
                                secciones,
                            )
                    continue

                if accion == "act_confirmar_pedido":
                    # SIEMPRE recrear desde el carrito_activo actual. Equora
                    # reportó (10-jun-2026): tras "Ver carrito" + "Confirmar
                    # pedido" se reutilizaba un checkout_url viejo guardado de
                    # un pedido previo — productos del checkout no coincidían
                    # con el carrito actual (cobraba lo que NO pidió el cliente).
                    # El supuesto anterior ("__ORDEN_CATALOGO__ siempre actualiza
                    # el URL") no se cumple cuando el cliente solo navega con
                    # botones de carrito. forzar_recrear=True garantiza que el
                    # link de pago refleje EXACTAMENTE lo que está en BD ahora.
                    checkout_url = await _obtener_o_recrear_checkout_url(
                        msg.telefono, agent_id=_agent_id, forzar_recrear=True
                    )
                    if checkout_url:
                        # Mensaje y label del botón configurables por agente
                        # (#28 fase 2.6). El cliente personaliza en
                        # Configuración → Mensajes → "Flujo de compra".
                        from agent.mensajes import format_seguro
                        _ctx_ph = await construir_contexto_placeholders(_agent_id)
                        _ctx_ph["total"] = await _total_carrito_fmt(msg.telefono, _agent_id)
                        msg_checkout = format_seguro(
                            await obtener_mensaje(_agent_id, "cart.checkout_listo_texto"),
                            _ctx_ph,
                        )
                        boton_label = (await obtener_mensaje(_agent_id, "cart.checkout_listo_boton")).strip() or "Pagar ahora"
                        # Mensaje 1: CTA URL directo al checkout (1 click → pagar).
                        # WhatsApp no permite mezclar URL button + reply buttons,
                        # entonces los otros 2 (Agregar más / Ver carrito) van
                        # como mensaje complementario aparte (no obliga al cliente
                        # a 2 clicks para pagar, que es la fricción que Jelou
                        # evita).
                        ok_cta = False
                        if hasattr(_proveedor_agente, "enviar_cta_url"):
                            ok_cta = await _proveedor_agente.enviar_cta_url(
                                msg.telefono, msg_checkout, boton_label, checkout_url,
                            )
                        if not ok_cta:
                            await _proveedor_agente.enviar_mensaje(
                                msg.telefono, f"{msg_checkout}\n\n👉 {checkout_url}"
                            )
                        # Mensaje 2 (complementario): opciones de modificar antes
                        # de pagar — siempre disponibles.
                        if hasattr(_proveedor_agente, "enviar_botones_reply"):
                            try:
                                await _proveedor_agente.enviar_botones_reply(
                                    msg.telefono, "¿Quieres modificar algo antes de pagar?",
                                    [
                                        {"id": "act_agregar_mas", "title": "🔍 Agregar más"},
                                        {"id": "act_ver_carrito", "title": "🛒 Ver mi carrito"},
                                    ],
                                )
                            except Exception as _e_b:
                                logger.warning(f"[confirmar-pedido] botones complementarios fallaron: {_e_b}")
                    else:
                        _txt_no_enc = await obtener_mensaje(_agent_id, "error.checkout_no_encontrado")
                        if _txt_no_enc:
                            await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_no_enc)
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, _txt_no_enc, agent_id=_agent_id)
                    continue

                if accion == "act_vaciar_carrito":
                    # El cliente tocó el botón "Vaciar carrito" del seguimiento.
                    # Vaciar directo en BD (sin depender del LLM) + confirmar.
                    try:
                        await limpiar_carrito_activo(msg.telefono, agent_id=_agent_id)
                        # Limpiar también el pedido_checkout_url para que el
                        # loop de checkout-abandonado no envíe "Vimos que
                        # iniciaste tu pedido pero no terminaste el proceso"
                        # cuando el cliente acaba de decir explícitamente
                        # que no quiere el pedido. Bug reportado por Equora.
                        await limpiar_checkout_url(msg.telefono, agent_id=_agent_id)
                        logger.info(f"[act_vaciar_carrito] carrito + checkout_url vaciados para {msg.telefono}")
                    except Exception as e:
                        logger.error(f"[act_vaciar_carrito] error vaciando: {e}")
                    msg_confirmacion = (
                        "🗑 *Listo, tu carrito quedó vacío.*\n\n"
                        "Cuando quieras armar un pedido nuevo, escríbeme qué necesitas "
                        "o pídeme el catálogo 🌿"
                    )
                    await _proveedor_agente.enviar_mensaje(msg.telefono, msg_confirmacion)
                    await _guardar_con_wamid(_proveedor_agente, msg.telefono, msg_confirmacion, agent_id=_agent_id)
                    continue

                if accion == "act_ir_pago":
                    # Botón Jelou (#61): el cliente decidió pagar — enviamos
                    # cta_url al invoice_url/checkout_url guardado en BD.
                    checkout_url = await _obtener_o_recrear_checkout_url(
                        msg.telefono, agent_id=_agent_id, forzar_recrear=False,
                    )
                    if not checkout_url:
                        _txt_no_enc = await obtener_mensaje(_agent_id, "error.checkout_no_encontrado")
                        if _txt_no_enc:
                            await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_no_enc)
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, _txt_no_enc, agent_id=_agent_id)
                        continue
                    boton_label = (await obtener_mensaje(_agent_id, "cart.checkout_listo_boton")).strip() or "Pagar ahora"
                    txt = "Toca el botón para completar tu pago de forma segura 🔒"
                    ok_cta = False
                    if hasattr(_proveedor_agente, "enviar_cta_url"):
                        ok_cta = await _proveedor_agente.enviar_cta_url(
                            msg.telefono, txt, boton_label, checkout_url,
                        )
                    if not ok_cta:
                        await _proveedor_agente.enviar_mensaje(
                            msg.telefono, f"{txt}\n\n👉 {checkout_url}"
                        )
                    continue

                if accion == "act_agregar_mas":
                    # Botón Jelou (#61): cliente quiere agregar más productos —
                    # reabrimos el catálogo nativo de WhatsApp.
                    secciones = obtener_secciones_catalogo(None)
                    if secciones:
                        header_cat = await _catalogo_header_for_agent(_agent_id)
                        await _proveedor_agente.enviar_catalogo_productos(
                            msg.telefono, header_cat,
                            "Elige los productos que quieras agregar a tu pedido.",
                            secciones,
                        )
                    else:
                        await _proveedor_agente.enviar_mensaje(
                            msg.telefono,
                            "Cuéntame qué producto te gustaría agregar y te lo busco 🛒",
                        )
                    continue

                if accion == "act_ver_carrito":
                    # El cliente tocó "Ver carrito" — backend lee BD y envía resumen
                    # determinístico (NO depende de Andrea/LLM).
                    await _enviar_resumen_carrito(msg.telefono, agent_id=_agent_id)
                    continue

                # Acción desconocida: no hacer nada
                continue

            # ── Opt-out: el cliente pide darse de baja de las difusiones ──────
            if _es_solicitud_opt_out(msg.texto):
                await marcar_opt_out(msg.telefono, motivo=msg.texto.strip()[:200], agent_id=_agent_id)
                await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                respuesta_baja = (
                    "Listo ✅ Te quitamos de nuestra lista de difusiones. "
                    "No volverás a recibir mensajes masivos de nuestra parte.\n\n"
                    "Si en algún momento cambias de opinión y quieres volver a recibir "
                    "nuestras ofertas y novedades, solo escríbenos. 🌿"
                )
                await _proveedor_agente.enviar_mensaje(msg.telefono, respuesta_baja)
                await _guardar_con_wamid(_proveedor_agente, msg.telefono, respuesta_baja, agent_id=_agent_id)
                logger.info(f"[opt-out] {msg.telefono} dado de baja — motivo: {msg.texto[:50]}")
                continue  # no pasar por Claude

            # ── Flujo conversacional Calendly (agendado sin salir de WA) ─────────
            _cal_pend = await obtener_calendly_pendiente(msg.telefono, _agent_id)
            if _cal_pend:
                _paso_cal = _cal_pend.get("paso")

                if _paso_cal == "esperando_seleccion":
                    _opciones_cal  = _cal_pend.get("opciones", {})
                    _id_map_cal    = _cal_pend.get("id_map", {})
                    _titulos_map   = _cal_pend.get("titulos_map", {})
                    _texto_norm    = msg.texto.strip()
                    _iso_sel       = None
                    _titulo_sel_cal = _texto_norm

                    # 1) Match por lista_id (más fiable — evita problemas de encoding del título)
                    if msg.lista_id and msg.lista_id in _id_map_cal:
                        _iso_sel = _id_map_cal[msg.lista_id]
                        _titulo_sel_cal = _titulos_map.get(msg.lista_id, _texto_norm)
                        logger.info(f"[calendly] match por id '{msg.lista_id}' → {_iso_sel}")
                    # 2) Match por title exacto
                    elif _texto_norm in _opciones_cal:
                        _iso_sel = _opciones_cal[_texto_norm]
                        logger.info(f"[calendly] match por título exacto '{_texto_norm}'")
                    # 3) Match case-insensitive
                    else:
                        for _k_cal, _v_cal in _opciones_cal.items():
                            if _k_cal.strip().lower() == _texto_norm.lower():
                                _iso_sel = _v_cal
                                _titulo_sel_cal = _k_cal
                                logger.info(f"[calendly] match case-insensitive '{_k_cal}'")
                                break

                    if _iso_sel is None:
                        logger.warning(
                            f"[calendly] NO match: lista_id='{msg.lista_id}' "
                            f"texto='{_texto_norm}' "
                            f"opciones_keys={list(_opciones_cal.keys())[:4]} "
                            f"id_map_keys={list(_id_map_cal.keys())[:4]}"
                        )
                    if _iso_sel is not None:
                        _evt_titulo = _cal_pend.get("event_type_nombre", "Cita")
                        await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                        await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)
                        _fecha_inicio_dt = datetime.fromisoformat(_iso_sel.replace("Z", "+00:00")).replace(tzinfo=None)
                        _appt_id = await crear_appointment_pendiente(
                            _agent_id, msg.telefono, _fecha_inicio_dt, _evt_titulo
                        )
                        _cal_pend.update({
                            "paso": "esperando_datos",
                            "seleccion": _iso_sel,
                            "seleccion_titulo": _titulo_sel_cal,
                            "appointment_id": _appt_id["id"] if isinstance(_appt_id, dict) else _appt_id,
                        })
                        await guardar_calendly_pendiente(msg.telefono, _cal_pend, _agent_id)
                        _pedir_datos = (
                            f"Perfecto, reservé el horario *{_titulo_sel_cal}* 🗓\n\n"
                            "Para confirmar necesito tu nombre completo y correo electrónico.\n"
                            "Puedes escribirlos así:\n"
                            "_Juan Pérez, juan@ejemplo.com_"
                        )
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _pedir_datos)
                        await _guardar_con_wamid(_proveedor_agente, msg.telefono, _pedir_datos, agent_id=_agent_id)
                        continue
                    else:
                        # Texto no coincide — avisar sin invocar el LLM (evita bucle).
                        _cal_pend["intentos"] = _cal_pend.get("intentos", 0) + 1
                        if _cal_pend["intentos"] >= 3:
                            await limpiar_calendly_pendiente(msg.telefono, _agent_id)
                            _msg_cal_err = "No pude identificar tu selección. Escribe *agendar* para ver los horarios de nuevo."
                            await _proveedor_agente.enviar_mensaje(msg.telefono, _msg_cal_err)
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, _msg_cal_err, agent_id=_agent_id)
                        else:
                            await guardar_calendly_pendiente(msg.telefono, _cal_pend, _agent_id)
                            _msg_cal_retry = "No reconocí esa opción 🤔 Toca *Ver horarios* y elige un horario de la lista."
                            await _proveedor_agente.enviar_mensaje(msg.telefono, _msg_cal_retry)
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, _msg_cal_retry, agent_id=_agent_id)
                        continue  # nunca dejar pasar al LLM cuando hay estado Calendly activo

                elif _paso_cal == "esperando_datos":
                    _email_match = re.search(r"[\w.+\-]+@[\w\-]+\.[a-z]{2,}", msg.texto, re.IGNORECASE)
                    if _email_match:
                        await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                        await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)
                        _email_cal = _email_match.group(0)
                        _nombre_raw = msg.texto.replace(_email_cal, "").strip(" ,.-\n")
                        _cli_cal = await obtener_cliente(msg.telefono, agent_id=_agent_id)
                        _nombre_cal = (
                            _nombre_raw
                            or (_cli_cal.get("nombres", "") if _cli_cal else "")
                            or "Cliente"
                        )
                        _cal_pend.update({
                            "paso": "esperando_asunto",
                            "nombre_cal": _nombre_cal,
                            "email_cal": _email_cal,
                        })
                        await guardar_calendly_pendiente(msg.telefono, _cal_pend, _agent_id)
                        _pedir_asunto = (
                            f"Gracias, {_nombre_cal} 👋\n\n"
                            "¿Cuál es el asunto o motivo de la cita?\n"
                            "_Ej: Consulta sobre productos de limpieza_"
                        )
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _pedir_asunto)
                        await _guardar_con_wamid(_proveedor_agente, msg.telefono, _pedir_asunto, agent_id=_agent_id)
                        continue
                    # Sin email → recordar al usuario sin invocar el LLM
                    await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                    _recordatorio_cal = (
                        "Para confirmar la cita necesito tu nombre y correo 📧\n"
                        "Escríbelos así: _Juan Pérez, juan@ejemplo.com_"
                    )
                    await _proveedor_agente.enviar_mensaje(msg.telefono, _recordatorio_cal)
                    await _guardar_con_wamid(_proveedor_agente, msg.telefono, _recordatorio_cal, agent_id=_agent_id)
                    continue  # nunca dejar pasar al LLM cuando se esperan datos del invitado

                elif _paso_cal == "esperando_asunto":
                    await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                    await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)
                    _asunto_cal = msg.texto.strip() or "Sin asunto"
                    _nombre_cal = _cal_pend.get("nombre_cal", "Cliente")
                    _email_cal  = _cal_pend.get("email_cal", "")
                    _resultado_cal = await calendly_crear_cita(
                        agent_id=_agent_id,
                        event_type_uri=_cal_pend.get("event_type_uri", ""),
                        start_time_iso=_cal_pend.get("seleccion", ""),
                        invitee_email=_email_cal,
                        invitee_nombre=_nombre_cal,
                        location_kind=_cal_pend.get("location_kind"),
                    )
                    if _resultado_cal.get("ok"):
                        _appt_id_cal = _cal_pend.get("appointment_id")
                        if _appt_id_cal:
                            await confirmar_appointment(
                                _appt_id_cal, _nombre_cal, _email_cal,
                                _resultado_cal.get("uri", ""),
                                _resultado_cal.get("cancel_url", ""),
                                _resultado_cal.get("reschedule_url", ""),
                            )
                        await limpiar_calendly_pendiente(msg.telefono, _agent_id)
                        _titulo_sel = _cal_pend.get("seleccion_titulo", "Cita agendada")
                        _msg_conf = (
                            f"✅ ¡Tu cita está confirmada!\n\n"
                            f"📅 *{_titulo_sel}*\n"
                            f"👤 {_nombre_cal}\n"
                            f"📧 {_email_cal}\n"
                            f"📝 {_asunto_cal}\n\n"
                            "Recibirás un correo de confirmación de Calendly con todos los detalles."
                        )
                        if _resultado_cal.get("reschedule_url"):
                            _msg_conf += f"\n\n🔄 Reagendar: {_resultado_cal['reschedule_url']}"
                        if _resultado_cal.get("cancel_url"):
                            _msg_conf += f"\n❌ Cancelar: {_resultado_cal['cancel_url']}"
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _msg_conf)
                        await _guardar_con_wamid(_proveedor_agente, msg.telefono, _msg_conf, agent_id=_agent_id)
                        logger.info(f"[calendly] Cita confirmada para {msg.telefono} — {_email_cal} — asunto: {_asunto_cal}")
                        try:
                            from agent.memory import obtener_pipeline_activo, crear_deal
                            _pipeline_cal = await obtener_pipeline_activo(_agent_id)
                            if _pipeline_cal:
                                await crear_deal(
                                    agent_id=_agent_id,
                                    pipeline_id=_pipeline_cal["id"],
                                    cliente_telefono=msg.telefono,
                                    cliente_nombre=_nombre_cal,
                                    titulo=f"📅 {_titulo_sel} — {_asunto_cal}",
                                    stage="Calificado",
                                    source="calendly",
                                )
                        except Exception as _e_deal_cal:
                            logger.warning(f"[calendly] No se pudo crear deal en pipeline: {_e_deal_cal}")
                    else:
                        await limpiar_calendly_pendiente(msg.telefono, _agent_id)
                        _err_cal = _resultado_cal.get("error", "Error inesperado")
                        _msg_err_cal = (
                            f"Lo siento, no pude confirmar la cita: {_err_cal} 😔\n\n"
                            "Escribe *agendar* para ver otros horarios disponibles."
                        )
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _msg_err_cal)
                        await _guardar_con_wamid(_proveedor_agente, msg.telefono, _msg_err_cal, agent_id=_agent_id)
                    continue

            # ── Media entrante (imagen/video/documento/ubicación) ──────────────
            # Guardamos el marcador en historial para que el inbox lo renderice,
            # y reemplazamos el texto que pasa a Claude por el caption o un placeholder
            # genérico (Andrea no procesa el contenido binario, solo responde al caption).
            if msg.texto.startswith("__MEDIA__:"):
                try:
                    media_payload = json.loads(msg.texto[len("__MEDIA__:"):])
                except Exception:
                    media_payload = {}
                # Guardamos el marcador original en BD (el inbox lo renderiza)
                await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)
                # Si hay caption, lo usamos como mensaje "real" para Claude
                caption = (media_payload.get("caption") or "").strip()
                tipo_media = media_payload.get("tipo", "media")
                tipo_legible = {
                    "image": "imagen", "video": "video",
                    "document": "documento", "location": "ubicación",
                    "audio": "audio"
                }.get(tipo_media, "archivo")
                if caption:
                    # Reescribir msg.texto para que Claude responda al caption con contexto
                    msg = type(msg)(
                        telefono=msg.telefono,
                        texto=f"[Cliente envió {tipo_legible}] {caption}",
                        mensaje_id=msg.mensaje_id,
                        es_propio=msg.es_propio,
                    )
                else:
                    # Sin caption: Andrea reconoce que llegó media pero no continúa con Claude
                    # — el operador humano debe responder. Solo registramos.
                    logger.info(f"Media {tipo_media} de {msg.telefono} sin caption — no se invoca Claude")
                    continue

            # ── Orden directa desde catálogo nativo WhatsApp ──────────────────
            # El cliente armó su carrito en WhatsApp y confirmó — bypaseamos Claude
            # y creamos el checkout de Shopify directamente.
            if msg.texto.startswith("__ORDEN_CATALOGO__:"):
                await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)
                logger.info(f"[orden-catalogo] Recibida orden de {msg.telefono}: {msg.texto[:500]}")
                try:
                    items_raw = json.loads(msg.texto[len("__ORDEN_CATALOGO__:"):])
                    # ── Merge inteligente: solo sumar si el carrito previo está
                    # FRESCO (actividad < CARRITO_MERGE_HORAS = 2h). Si está viejo
                    # entre 2h y 4h, REEMPLAZAR — evita acumular pedidos olvidados
                    # de ayer. Si vacío o expirado (>4h), arrancar limpio.
                    carrito_previo = await obtener_carrito_activo(msg.telefono, agent_id=_agent_id)
                    es_fresco = await carrito_es_fresco_para_merge(msg.telefono, agent_id=_agent_id)
                    if carrito_previo and not es_fresco:
                        logger.info(
                            f"[orden-catalogo] Carrito previo encontrado pero NO fresco "
                            f"({len(carrito_previo)} items) — REEMPLAZANDO (no merge)"
                        )
                        # Limpiar antes para que el guardado sea solo lo nuevo
                        await limpiar_carrito_activo(msg.telefono, agent_id=_agent_id)
                        carrito_previo = []
                    # ── Construir 'productos' UNIFICANDO carrito_previo + items_raw nuevos.
                    # El método anterior descartaba la info completa del carrito_previo
                    # cuando lookup en _sku_map fallaba — por eso a veces faltaban items
                    # en el checkout. Ahora preservamos la info: si el lookup falla,
                    # usamos la info del carrito_previo que ya está completa.
                    productos = []
                    no_encontrados = []
                    ya_agregados: dict[str, dict] = {}

                    # Paso 1: items_raw (la orden NUEVA que acaba de llegar)
                    # CRÍTICO: cada item de productos lleva su retailer_id desde el inicio.
                    # Sin esto, cuando se guarde en carrito_activo el rid quedaba "" para
                    # items mergeados → siguiente merge los descartaba → bug del producto perdido.
                    for item in items_raw:
                        rid = item.get("product_retailer_id", "")
                        qty = int(item.get("quantity", 1))
                        if not rid:
                            continue
                        info = obtener_producto_por_retailer_id(rid, _agent_id)
                        if not info:
                            logger.warning(f"[orden-catalogo] rid '{rid}' no en _sku_map[{_agent_id}] — recargando catálogo")
                            await obtener_catalogo_shopify(_agent_id)
                            info = obtener_producto_por_retailer_id(rid, _agent_id)
                        if info:
                            precio = info["precio_unitario"]
                            item_full = {
                                "producto": info["producto"],
                                "presentacion": info["presentacion"],
                                "cantidad": qty,
                                "precio_unitario": precio,
                                "subtotal": precio * qty,
                                "retailer_id": rid,
                            }
                            productos.append(item_full)
                            ya_agregados[rid] = item_full
                            logger.info(f"[orden-catalogo-nuevo] rid={rid} → {info['producto']} · {info['presentacion']} × {qty}")
                        else:
                            no_encontrados.append(rid)
                            from agent.tools import _sku_map as _sku_map_all
                            logger.error(
                                f"[orden-catalogo] rid '{rid}' NO existe en _sku_map[{_agent_id}] (entradas: {len(_sku_map_all.get(_agent_id, {}))})"
                            )

                    # Paso 2: carrito_previo (items del pedido anterior, info ya completa
                    # en BD). Si un rid ya viene en items_raw, sumamos cantidades. Si NO
                    # estaba, agregamos como item NUEVO usando la info persistida —
                    # crítico: aunque _sku_map ya no lo conozca, el carrito_previo SÍ tiene
                    # la info completa (producto, presentación, precio) → checkout completo.
                    if carrito_previo:
                        logger.info(f"[orden-catalogo] Mergeando con carrito previo (fresco): {len(carrito_previo)} items")
                        for prev in carrito_previo:
                            rid = prev.get("retailer_id") or prev.get("product_retailer_id") or ""
                            qty_prev = int(prev.get("quantity", prev.get("cantidad", 1)))
                            if not rid:
                                logger.warning(f"[orden-catalogo] carrito_previo item sin retailer_id — descarto: {prev}")
                                continue
                            if rid in ya_agregados:
                                # Mismo producto en orden nueva — sumar cantidades
                                ya_agregados[rid]["cantidad"] += qty_prev
                                ya_agregados[rid]["subtotal"] = (
                                    ya_agregados[rid]["cantidad"]
                                    * ya_agregados[rid]["precio_unitario"]
                                )
                            else:
                                # Item del carrito previo NO viene en la nueva orden —
                                # preservarlo (no descartar). Si tiene info completa
                                # persistida usar esa; si no, intentar lookup último recurso.
                                prod_prev = prev.get("producto", "")
                                pres_prev = prev.get("presentacion", "")
                                precio_prev = prev.get("precio_unitario", 0) or 0
                                subt_prev = prev.get("subtotal") or (precio_prev * qty_prev)
                                if not prod_prev:
                                    info_prev = obtener_producto_por_retailer_id(rid, _agent_id)
                                    if info_prev:
                                        prod_prev = info_prev["producto"]
                                        pres_prev = info_prev["presentacion"]
                                        precio_prev = info_prev["precio_unitario"]
                                        subt_prev = precio_prev * qty_prev
                                    else:
                                        logger.warning(
                                            f"[orden-catalogo] carrito_previo rid='{rid}' sin info — descarto"
                                        )
                                        continue
                                item_prev_full = {
                                    "producto": prod_prev,
                                    "presentacion": pres_prev,
                                    "cantidad": qty_prev,
                                    "precio_unitario": precio_prev,
                                    "subtotal": subt_prev,
                                    "retailer_id": rid,   # ← preservar para persistir
                                }
                                productos.append(item_prev_full)
                                ya_agregados[rid] = item_prev_full
                                logger.info(f"[orden-catalogo-previo] rid={rid} → {prod_prev} · {pres_prev} × {qty_prev}")
                        logger.info(f"[orden-catalogo] Total después de merge: {len(productos)} items")

                    if no_encontrados:
                        logger.error(f"[orden-catalogo] IDs sin mapear ({len(no_encontrados)}/{len(items_raw)}): {no_encontrados}")

                    if productos:
                        total = sum(p["subtotal"] for p in productos)

                        # ── Validar pedido mínimo ANTES de crear checkout ──
                        try:
                            pedido_min_str = await get_config_value("PEDIDO_MINIMO", _agent_id) or os.getenv("PEDIDO_MINIMO", "0")
                            pedido_min = int(float(pedido_min_str)) if pedido_min_str else 0
                        except Exception:
                            pedido_min = 0

                        if pedido_min > 0 and total < pedido_min:
                            falta = pedido_min - total
                            # ── Guardar items en carrito_activo con info COMPLETA ──
                            # FIX RAÍZ: usamos 'productos' (que YA tiene el merge completo
                            # con retailer_ids correctos), NO 'items_raw' (solo orden nueva).
                            # Antes, items del carrito previo se guardaban con rid="" y se
                            # perdían en el próximo merge.
                            items_para_carrito = [
                                {
                                    "retailer_id":     p.get("retailer_id", ""),
                                    "quantity":        p["cantidad"],
                                    "cantidad":        p["cantidad"],
                                    "producto":        p["producto"],
                                    "presentacion":    p["presentacion"],
                                    "precio_unitario": p["precio_unitario"],
                                    "subtotal":        p["subtotal"],
                                }
                                for p in productos
                            ]
                            try:
                                await guardar_carrito_activo(msg.telefono, items_para_carrito, agent_id=_agent_id)
                                logger.info(f"[orden-catalogo] Guardado en carrito_activo: {len(items_para_carrito)} items completos")
                            except Exception as e:
                                logger.error(f"Error guardando carrito_activo: {e}")

                            # Mensaje configurable por agente
                            lineas_pedido = "\n".join(
                                f"  • {p['producto']} {p['presentacion']} × {p['cantidad']} = ${p['subtotal']:,}".replace(",", ".")
                                for p in productos
                            )
                            msg_template = await get_config_value("PEDIDO_MIN_MSG", _agent_id) or (
                                "😊 ¡Tu pedido va muy bien! Lo que tienes:\n\n"
                                "{ITEMS}\n\n"
                                "💰 *Subtotal:* ${SUBTOTAL}\n"
                                "📦 *Pedido mínimo:* ${MINIMO}\n"
                                "➕ *Te faltan:* ${FALTA} para confirmar 🎯\n\n"
                                "✅ *No te preocupes, tu pedido quedó guardado.* "
                                "Solo agrega más productos en el catálogo de abajo "
                                "y *se sumará automáticamente* a lo que ya tenías. 🛒"
                            )
                            mensaje_minimo = (msg_template
                                .replace("{SUBTOTAL}", f"{total:,}".replace(",", "."))
                                .replace("{MINIMO}",   f"{pedido_min:,}".replace(",", "."))
                                .replace("{FALTA}",    f"{falta:,}".replace(",", "."))
                                .replace("{ITEMS}",    lineas_pedido)
                            )
                            # Enviar SIEMPRE el texto primero
                            await _proveedor_agente.enviar_mensaje(msg.telefono, mensaje_minimo)
                            # Persistir en BD para que el operador vea este mensaje
                            # en el panel de Conversaciones (antes se perdía).
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, mensaje_minimo, agent_id=_agent_id)

                            # Catalog_message removido: cuelga WhatsApp Business
                            # app en algunos Android (Meta side bug). Vamos directo
                            # al product_list que es estable en todos los devices.
                            logger.info(f"[pedido-min] Enviando product_list a {msg.telefono}")
                            cat_reabierto = False
                            tiene_metodo = hasattr(_proveedor_agente, "enviar_catalogo_productos")
                            logger.info(f"[pedido-min] tiene enviar_catalogo_productos={tiene_metodo}")
                            if tiene_metodo:
                                # Async: garantiza que _fb_items esté cargado antes
                                # de armar el payload (defensa contra cold start).
                                from agent.tools import obtener_secciones_catalogo_async, _fb_items
                                secciones = await obtener_secciones_catalogo_async(None)
                                logger.info(
                                    f"[pedido-min] _fb_items={len(_fb_items)} | secciones={len(secciones)} | "
                                    f"items_en_secciones={sum(len(s.get('product_items', [])) for s in secciones)}"
                                )
                                if secciones:
                                    try:
                                        cat_reabierto = await _proveedor_agente.enviar_catalogo_productos(
                                            msg.telefono, "Agrega más productos 🌿",
                                            "Tu carrito anterior está guardado — lo nuevo se suma automáticamente.",
                                            secciones,
                                        )
                                        logger.info(f"[pedido-min] cat_reabierto={cat_reabierto}")
                                    except Exception as e_pl:
                                        logger.error(f"[pedido-min] product_list excepción: {e_pl}", exc_info=True)
                                else:
                                    logger.warning(f"[pedido-min] secciones VACÍAS — no se mandó product_list")
                            # Mensaje de cierre — depende de si el catálogo nativo se abrió o no:
                            #   ✅ OK   → texto sugerente "escribe o explora el catálogo"
                            #   ❌ Fail → ESCALAR a humano (el cliente esperaba una respuesta
                            #             concreta; sin catálogo no podemos cerrar la venta).
                            # NUNCA mandamos link a tienda web — eso rompería el flujo del
                            # carrito nativo (universos separados que no se sincronizan).
                            if cat_reabierto:
                                texto_final = (
                                    "Escríbeme qué más quieres agregar y te ayudo, "
                                    "o explora el catálogo en *Ver artículos* 👆🌿"
                                )
                                await _proveedor_agente.enviar_mensaje(msg.telefono, texto_final)
                                await _guardar_con_wamid(_proveedor_agente, msg.telefono, texto_final, agent_id=_agent_id)
                            else:
                                # Meta nativo no disponible → escalar
                                await _escalar_meta_fallo(
                                    msg.telefono,
                                    motivo_corto="post-pedido bajo mínimo",
                                    contexto_extra=(
                                        f"Cliente cerró pedido por ${total:,} pero el mínimo es "
                                        f"${pedido_min:,} (faltan ${falta:,}). Catálogo nativo no "
                                        f"se pudo enviar para completar el pedido. Items actuales: "
                                        + ", ".join(f"{p['producto']} x{p['cantidad']}" for p in productos)
                                    ).replace(",", "."),
                                    agent_id=_agent_id,
                                )
                            logger.info(f"Pedido bajo mínimo: total={total}, min={pedido_min}, falta={falta} — items guardados para acumular")
                            continue  # No crear checkout

                        datos_pedido = {"productos": productos, "total": total}
                        checkout_url = await crear_checkout_shopify(msg.telefono, datos_pedido)
                        if checkout_url:
                            await guardar_pedido_pendiente(msg.telefono, productos, agent_id=_agent_id)
                            # IMPORTANTE: NO limpiar carrito_activo aquí — el cliente puede
                            # agregar más productos para alcanzar envío gratis. El carrito
                            # se limpia cuando llega webhook orders/paid de Shopify (pago
                            # confirmado) o cuando expira el TTL de 4h.
                            #
                            # FIX RAÍZ: 'productos' YA tiene retailer_id de cada item
                            # (paso 1 y paso 2 del merge lo agregaron). Usamos eso
                            # directamente — antes buscábamos rids en items_raw lo cual
                            # solo encontraba el rid del último pedido, dejando los items
                            # mergeados con rid="" y perdiéndolos en el siguiente pedido.
                            items_carrito_full = [
                                {
                                    "retailer_id":     p.get("retailer_id", ""),
                                    "quantity":        p["cantidad"],
                                    "cantidad":        p["cantidad"],
                                    "producto":        p["producto"],
                                    "presentacion":    p["presentacion"],
                                    "precio_unitario": p["precio_unitario"],
                                    "subtotal":        p["subtotal"],
                                }
                                for p in productos
                            ]
                            try:
                                await guardar_carrito_activo(msg.telefono, items_carrito_full, agent_id=_agent_id)
                            except Exception as e_c:
                                logger.warning(f"No pude actualizar carrito_activo: {e_c}")
                            # Guardar checkout_url para que el cliente pueda confirmar después
                            try:
                                await guardar_checkout_url(msg.telefono, checkout_url, agent_id=_agent_id)
                            except Exception as e_url:
                                logger.warning(f"No pude guardar checkout_url: {e_url}")

                            # ── Detectar si alcanzó envío gratis ──────────────
                            try:
                                umbral_gratis_local = obtener_umbral_envio_gratis()
                            except Exception:
                                umbral_gratis_local = 0
                            try:
                                costo_envio_local = obtener_costo_envio()
                            except Exception:
                                costo_envio_local = 0

                            total_fmt_loc = f"{total:,}".replace(",", ".")

                            # CASO A: alcanza mínimo PERO no envío gratis
                            # Estrategia: 2 mensajes separados — uno con botón "Ver
                            # catálogo" (para agregar más) y otro con CTA URL directo
                            # al checkout. WhatsApp no soporta 2 CTA URL en un mismo
                            # mensaje, así que es la forma más limpia.
                            # IMPORTANTE: la promesa de "envío gratis" solo aplica si
                            # el cliente está en zona local. Verificamos su ciudad.
                            if umbral_gratis_local > 0 and total < umbral_gratis_local:
                                falta_gratis = umbral_gratis_local - total
                                falta_g_fmt = f"{falta_gratis:,}".replace(",", ".")
                                envio_fmt = f"{costo_envio_local:,}".replace(",", ".") if costo_envio_local else "0"

                                # Detectar ciudad del cliente
                                try:
                                    cliente_info_a = await obtener_cliente(msg.telefono, agent_id=_agent_id)
                                except Exception:
                                    cliente_info_a = None
                                ciudad_a = (cliente_info_a.get("ciudad") if cliente_info_a else "") or ""
                                ciudad_a = ciudad_a.strip().lower()
                                ciudades_loc_raw = (
                                    await get_config_value("CIUDADES_ENVIO_PROPIO", _agent_id)
                                    or os.getenv("CIUDADES_ENVIO_PROPIO", "cali")
                                )
                                ciudades_loc = {
                                    c.strip().lower() for c in ciudades_loc_raw.split(",") if c.strip()
                                }
                                es_local_a = ciudad_a in ciudades_loc if ciudad_a else None

                                # ── Mensaje 1: invitar a agregar más (configurable por agente) ──
                                # El template del cliente decide qué decir: si quiere mencionar
                                # zonas de envío gratis, descuento, etc., usa los placeholders
                                # dinámicos {total} {falta_envio_gratis} {descuento_codigo} etc.
                                # Eliminamos la rama por "es_local_a" — el cliente personaliza
                                # su mensaje según su modelo de negocio, no asumimos nada.
                                from agent.mensajes import format_seguro
                                _ctx_estado4 = await construir_contexto_placeholders(_agent_id)
                                _ctx_estado4["total"] = total_fmt_loc
                                _ctx_estado4["falta_envio_gratis"] = falta_g_fmt if falta_g_fmt else ""
                                mensaje_agregar = format_seguro(
                                    await obtener_mensaje(_agent_id, "cart.estado4_cross_sell"),
                                    _ctx_estado4,
                                )
                                # Si el agente desactivó el cross-sell, no enviamos sugerencia ni
                                # catálogo extra — saltamos directo al CTA de pago abajo.
                                if mensaje_agregar:
                                    # catalog_message removido — directo a product_list (estable en WB).
                                    cat_btn_ok = False
                                    if hasattr(_proveedor_agente, "enviar_catalogo_productos"):
                                        secciones_loc = obtener_secciones_catalogo(None)
                                        if secciones_loc:
                                            try:
                                                await _proveedor_agente.enviar_catalogo_productos(
                                                    msg.telefono, "Agrega más productos 🌿",
                                                    mensaje_agregar[:1024], secciones_loc,
                                                )
                                                cat_btn_ok = True
                                            except Exception:
                                                pass
                                    if not cat_btn_ok:
                                        await _proveedor_agente.enviar_mensaje(msg.telefono, mensaje_agregar)
                                    # Siempre persistir el texto del mensaje (sea por product_list o por
                                    # fallback). El cuerpo del product_list NO se guarda automáticamente
                                    # — sin esto el operador NO ve el "¡Pedido confirmado!" en el panel.
                                    await _guardar_con_wamid(_proveedor_agente, msg.telefono, mensaje_agregar, agent_id=_agent_id)

                                # ── Mensaje 2: CTA al checkout (texto + botón configurables) ──
                                mensaje_confirmar = format_seguro(
                                    await obtener_mensaje(_agent_id, "cart.estado4_cta_texto"),
                                    _ctx_estado4,
                                )
                                _boton_label = (
                                    await obtener_mensaje(_agent_id, "cart.checkout_listo_boton")
                                ).strip() or "Confirmar pedido"
                                if hasattr(_proveedor_agente, "enviar_cta_url"):
                                    await _proveedor_agente.enviar_cta_url(
                                        msg.telefono, mensaje_confirmar,
                                        _boton_label, checkout_url,
                                    )
                                else:
                                    await _proveedor_agente.enviar_mensaje(
                                        msg.telefono, f"{mensaje_confirmar}\n\n👉 {checkout_url}"
                                    )
                                # Persistir el mensaje en BD (panel lo necesita).
                                await _guardar_con_wamid(_proveedor_agente, msg.telefono, mensaje_confirmar, agent_id=_agent_id)
                                # Botones determinísticos del carrito (Ver/Vaciar)
                                # — el cliente SIEMPRE debe poder gestionar antes de pagar.
                                if hasattr(_proveedor_agente, "enviar_botones_reply"):
                                    try:
                                        await _proveedor_agente.enviar_botones_reply(
                                            msg.telefono, "¿O prefieres revisar tu carrito antes?",
                                            _botones_carrito_estandar(ofrece_confirmar=False),
                                        )
                                    except Exception as e:
                                        logger.warning(f"[pedido-min] botones carrito fallaron: {e}")
                                continue  # No pasar por Claude

                            # CASO B: carrito >= umbral envío gratis (o agent que no usa zonas)
                            # Mensaje + botón 100% configurables. El cliente decide qué decir
                            # según su modelo de envío (gratis a una zona, transportadora,
                            # solo pickup, etc.). Reutilizamos los mismos mensajes que el
                            # flujo de checkout directo para mantener UX consistente.
                            from agent.mensajes import format_seguro
                            _ctx_estado5 = await construir_contexto_placeholders(_agent_id)
                            _ctx_estado5["total"] = total_fmt_loc
                            texto_checkout = format_seguro(
                                await obtener_mensaje(_agent_id, "cart.checkout_listo_texto"),
                                _ctx_estado5,
                            )
                            _boton_label_b = (
                                await obtener_mensaje(_agent_id, "cart.checkout_listo_boton")
                            ).strip() or "Confirmar pedido"
                            if hasattr(_proveedor_agente, "enviar_cta_url"):
                                await _proveedor_agente.enviar_cta_url(
                                    msg.telefono, texto_checkout,
                                    _boton_label_b, checkout_url
                                )
                            else:
                                await _proveedor_agente.enviar_mensaje(
                                    msg.telefono, f"{texto_checkout}\n\n👉 {checkout_url}"
                                )
                            # Persistir el mensaje (panel lo necesita)
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, texto_checkout, agent_id=_agent_id)
                            # Botones determinísticos del carrito
                            if hasattr(_proveedor_agente, "enviar_botones_reply"):
                                try:
                                    await _proveedor_agente.enviar_botones_reply(
                                        msg.telefono, "¿O prefieres revisar tu carrito antes?",
                                        _botones_carrito_estandar(ofrece_confirmar=False),
                                    )
                                except Exception as e:
                                    logger.warning(f"[caso-B] botones carrito fallaron: {e}")
                        else:
                            _txt_err = await obtener_mensaje(_agent_id, "error.procesar_pedido_fallo")
                            if _txt_err:
                                await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_err)
                                await _guardar_con_wamid(_proveedor_agente, msg.telefono, _txt_err, agent_id=_agent_id)
                    else:
                        _txt_err2 = await obtener_mensaje(_agent_id, "error.productos_no_reconocidos")
                        if _txt_err2:
                            await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_err2)
                            await _guardar_con_wamid(_proveedor_agente, msg.telefono, _txt_err2, agent_id=_agent_id)
                except Exception as e:
                    logger.error(f"Error procesando orden catálogo: {e}")
                    _txt_exc = await obtener_mensaje(_agent_id, "error.excepcion_pedido")
                    if _txt_exc:
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_exc)
                continue  # No pasa por Claude

            # Cliente respondió → resetea timers de seguimiento
            await registrar_mensaje_usuario(msg.telefono, agent_id=_agent_id)

            # Modo humano: guardar mensaje pero no responder con Andrea
            if await get_modo_humano(msg.telefono, agent_id=_agent_id):
                await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
                logger.info(f"Modo humano activo — {msg.telefono} — Andrea no responde")
                continue

            historial = await obtener_historial(msg.telefono, agent_id=_agent_id)

            # CAPI: Lead — primera vez que este número escribe a Andrea
            if not historial:
                asyncio.create_task(capi_lead(msg.telefono))

            # ── Contexto de campaña ───────────────────────────────────────────
            # Si el cliente responde dentro de las 72 h de recibir una difusión,
            # Andrea sabe de qué producto/campaña viene y no lo pregunta de nuevo.
            contexto_campana = await obtener_campana_reciente_para_telefono(msg.telefono, agent_id=_agent_id)
            if contexto_campana:
                logger.warning(
                    f"[campaña✅] {msg.telefono} respondió a campaña "
                    f'"{contexto_campana["campaign_name"]}" '
                    f'(hace {contexto_campana["horas_ago"]}h)'
                )
            else:
                logger.warning(
                    f"[campaña❌] Sin campaña reciente en difusion_mensajes para {msg.telefono}"
                )

            # Delay natural en sandbox: hace que el agente parezca que está escribiendo
            if _is_sandbox_msg:
                await asyncio.sleep(1.5)

            respuesta = await generar_respuesta(
                msg.texto, historial, msg.telefono, contexto_campana, agent_id=_agent_id
            )

            await guardar_mensaje(msg.telefono, "user", msg.texto, agent_id=_agent_id)
            # El wamid se conoce recién tras el envío real más abajo (después
            # de procesar marcadores) — guardamos sin wamid y lo actualizamos
            # luego con actualizar_wamid_mensaje() (#79, fix: capturaba antes
            # del envío y siempre quedaba vacío).
            _msg_id_assistant = await guardar_mensaje(
                msg.telefono, "assistant", respuesta, agent_id=_agent_id,
            )
            await registrar_mensaje_asistente(msg.telefono, agent_id=_agent_id)

            # Procesar marcadores via dispatcher MARKER_HANDLERS:
            #   [[CARRITO:[...]]], [[VACIAR_CARRITO]], [[MOSTRAR_CARRITO]],
            #   [[ESCALAR:...]], [[PEDIDO:...]], [[CIERRE_CONV:]]
            #
            # Los marcadores interactivos (BOTONES, LISTA, TIENDA, PRODUCTO,
            # CATALOGO_CAT) aún se procesan inline más abajo — se migrarán
            # en la fase E del refactor.
            _marker_ctx = MarkerContext(
                respuesta=respuesta, telefono=msg.telefono, agent_id=_agent_id,
            )
            _marker_ctx = await aplicar_marcadores(_marker_ctx)
            respuesta        = _marker_ctx.respuesta
            checkout_url     = _marker_ctx.checkout_url
            checkout_fallo   = _marker_ctx.checkout_fallo
            datos_escalacion = _marker_ctx.datos_escalacion

            # [[MOSTRAR_CARRITO]] requiere envío inmediato — _enviar_resumen_carrito
            # vive en main.py (necesita acceso al proveedor de WhatsApp), por
            # eso el handler solo levanta un flag y aquí ejecutamos la acción.
            if _marker_ctx.mostrar_carrito_pendiente:
                try:
                    await _enviar_resumen_carrito(msg.telefono, agent_id=_agent_id)
                    logger.info(f"[MOSTRAR_CARRITO] resumen enviado a {msg.telefono}")
                except Exception as e:
                    logger.error(f"Error mostrando carrito de {msg.telefono}: {e}")

            if _marker_ctx.mostrar_citas_pendiente:
                try:
                    await _enviar_horarios_calendly(msg.telefono, agent_id=_agent_id)
                    respuesta = ""  # la lista interactiva ya es la respuesta; evitar doble envío
                    logger.info(f"[CITA_DISPONIBILIDAD] horarios enviados a {msg.telefono}")
                except Exception as e:
                    logger.error(f"Error enviando horarios Calendly a {msg.telefono}: {e}")

            # [[MOSTRAR_CARRITO]], [[PEDIDO:...]], [[ESCALAR:...]] y
            # [[CIERRE_CONV:]] ya fueron procesados por MARKER_HANDLERS arriba.
            # Los 5 marcadores interactivos (CATALOGO_CAT, BOTONES, LISTA,
            # TIENDA, PRODUCTO) también — solo extraemos los resultados del
            # contexto. El ENVÍO de cada uno sigue inline acá abajo porque
            # tiene dependencias de orden, fallbacks y estado local
            # (_proveedor_agente, _agent_id) que no conviene mover sin un
            # refactor mayor del bloque de envío.
            # NOTA: NO reasignar respuesta aquí — puede haber sido limpiada
            # intencionalmente (e.g. respuesta="" cuando se envió la lista de Calendly).
            datos_catalogo_cat     = _marker_ctx.datos_catalogo_cat
            datos_botones          = _marker_ctx.datos_botones
            datos_lista            = _marker_ctx.datos_lista
            abrir_tienda           = _marker_ctx.abrir_tienda
            tienda_query           = _marker_ctx.tienda_query
            productos_mencionados  = _marker_ctx.productos_mencionados

            # [[PEDIDO_CONFIRMAR]] — confirmar pedido nativo Voco y notificar al cliente
            if _marker_ctx.pedido_creado:
                try:
                    _ped = _marker_ctx.pedido_creado
                    _lineas = "\n".join(
                        f"  • {p['nombre']} ×{p['cantidad']}  ${p['precio_unitario']:,}"
                        for p in _ped.get("productos", [])
                    )
                    _msg_pedido = (
                        f"✅ *Pedido confirmado {_ped['numero']}*\n\n"
                        f"{_lineas}\n\n"
                        f"📦 Subtotal: ${_ped['subtotal']:,}\n"
                    )
                    if _ped.get("descuento", 0):
                        _msg_pedido += f"🎁 Descuento: -${_ped['descuento']:,}\n"
                    if _ped.get("costo_envio", 0):
                        _msg_pedido += f"🚚 Envío: ${_ped['costo_envio']:,}\n"
                    _msg_pedido += f"💰 *Total: ${_ped['total']:,}*"
                    if _ped.get("direccion_entrega"):
                        _dir_full = _ped["direccion_entrega"]
                        if _ped.get("direccion2_entrega"):
                            _dir_full += ", " + _ped["direccion2_entrega"]
                        if _ped.get("ciudad_entrega"):
                            _dir_full += " — " + _ped["ciudad_entrega"]
                        _msg_pedido += f"\n📍 Dirección: {_dir_full}"
                    await _proveedor_agente.enviar_mensaje(msg.telefono, _msg_pedido)
                    logger.info(
                        f"[PEDIDO_CONFIRMAR] Resumen {_ped['numero']} enviado a {msg.telefono}"
                    )
                except Exception as _e_ped:
                    logger.error(f"Error enviando confirmación de pedido: {_e_ped}")

            # Enviar texto primero
            # Excepción: si hay CTA de catálogo general (sin query), el texto de Andrea
            # se fusiona como cuerpo del CTA para que quede un único mensaje de bienvenida.
            _texto_absorbido_por_cta = abrir_tienda and not tienda_query and respuesta.strip()
            if respuesta and not _texto_absorbido_por_cta:
                await _proveedor_agente.enviar_mensaje(msg.telefono, respuesta)
                # Chulitos de confirmación (#79): ahora sí el provider ya tiene
                # el wamid real de este envío — actualizamos el mensaje guardado.
                await _vincular_wamid_post_envio(_proveedor_agente, _msg_id_assistant)

            # ── [[PRODUCTO:nombre]] — enviar uno o más productos del catálogo ──
            # Andrea lo usa cuando menciona un producto específico. Por cada
            # producto: single_product (con +/- nativos al tocar "Ver") + cta_url
            # (link directo a la tienda como alternativa).
            if productos_mencionados and hasattr(_proveedor_agente, "enviar_producto"):
                from agent.tools import _sku_map, _normalizar, obtener_url_producto
                # Construir índice (producto_norm + presentacion_norm) → retailer_id largo
                _producto_idx: dict[str, str] = {}
                for rid, info in _sku_map.get(_agent_id, {}).items():
                    if not (str(rid).isdigit() and len(str(rid)) >= 10):
                        continue  # Solo variant_id largos (los que sí están en FB Catalog)
                    prod_n = _normalizar(info.get("producto", ""))
                    pres_n = _normalizar(info.get("presentacion", ""))
                    _producto_idx[f"{prod_n}|{pres_n}"] = rid
                    # También indexar solo por producto (sin presentación) → primera variante
                    if prod_n not in _producto_idx:
                        _producto_idx[prod_n] = rid

                for nombre_prod in productos_mencionados[:3]:  # max 3 productos por mensaje
                    partes = [p.strip() for p in nombre_prod.split("|", 1)]
                    nombre = partes[0]
                    presentacion = partes[1] if len(partes) > 1 else ""
                    nombre_n = _normalizar(nombre)
                    pres_n   = _normalizar(presentacion)
                    rid_match = ""
                    if pres_n:
                        rid_match = _producto_idx.get(f"{nombre_n}|{pres_n}", "")
                    if not rid_match:
                        # Buscar coincidencia parcial por nombre
                        for k, v in _producto_idx.items():
                            if "|" in k and nombre_n in k.split("|")[0]:
                                rid_match = v
                                break
                        if not rid_match:
                            rid_match = _producto_idx.get(nombre_n, "")
                    if not rid_match:
                        logger.warning(f"[PRODUCTO] No se encontró '{nombre_prod}' en _sku_map")
                        continue
                    # 1) Enviar producto del catálogo nativo
                    try:
                        _res_prod = await _proveedor_agente.enviar_producto(msg.telefono, rid_match)
                        if isinstance(_res_prod, dict) and not _res_prod.get("ok"):
                            raise Exception(_res_prod.get("error", "single_product falló"))
                        logger.info(f"[PRODUCTO] single_product '{nombre_prod}' (rid={rid_match}) enviado")
                    except Exception as e:
                        logger.error(f"[PRODUCTO] Error enviando single_product: {e}")
                        # Registrar falla + alerta al admin
                        await _registrar_falla_catalogo(
                            "single_product", msg.telefono,
                            f"Producto '{nombre_prod}' (rid={rid_match}): {e}",
                            agent_id=_agent_id,
                        )
                        # Fallback: enviar URL del producto en la web
                        try:
                            url_web = obtener_url_producto(nombre)
                            if url_web and hasattr(_proveedor_agente, "enviar_cta_url"):
                                await _proveedor_agente.enviar_cta_url(
                                    msg.telefono,
                                    "Aquí está el producto 🌿",
                                    "Ver producto", url_web,
                                )
                                logger.info(f"[PRODUCTO→FALLBACK-WEB] CTA web enviado: {url_web}")
                        except Exception as e2:
                            logger.error(f"[PRODUCTO] Fallback web también falló: {e2}")
                        continue

            # Catálogo nativo WhatsApp (product_list con fotos reales)
            if datos_catalogo_cat and hasattr(_proveedor_agente, "enviar_catalogo_productos"):
                categoria = datos_catalogo_cat.get("categoria")
                secciones = obtener_secciones_catalogo(categoria)
                cat_enviado = False
                if secciones:
                    header = categoria or await _catalogo_header_for_agent(_agent_id)
                    cuerpo = "Selecciona los productos que quieres, ajusta las cantidades y confirma tu pedido."
                    cat_enviado = await _proveedor_agente.enviar_catalogo_productos(
                        msg.telefono, header, cuerpo, secciones
                    )
                    if cat_enviado:
                        logger.info(f"Catálogo nativo '{categoria}' enviado a {msg.telefono}")
                    else:
                        logger.warning(f"Catálogo nativo falló para {msg.telefono} — usando fallback texto")

                # Fallback: si el catálogo nativo falla (IDs no mapeados, error Meta, etc.)
                # mostramos los productos como lista interactiva de texto
                if not cat_enviado:
                    from agent.tools import _categorias_cache, ORDEN_CATEGORIAS
                    cats = {categoria: _categorias_cache[categoria]} if (
                        categoria and categoria in _categorias_cache
                    ) else {c: _categorias_cache[c] for c in ORDEN_CATEGORIAS if c in _categorias_cache}
                    lineas = []
                    for cat_name, productos_cat in cats.items():
                        lineas.append(f"*{cat_name}*")
                        for prod_title, variantes, _img in productos_cat:
                            precio_min = min(int(float(v["price"]["amount"])) for v in variantes)
                            lineas.append(f"  • {prod_title} — desde ${precio_min:,}")
                        lineas.append("")
                    if lineas:
                        _texto_cat_fb = "\n".join(lineas).strip()
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _texto_cat_fb)
                        await _guardar_con_wamid(_proveedor_agente, msg.telefono, _texto_cat_fb, agent_id=_agent_id)

            # Luego enviar mensajes interactivos
            if datos_botones and hasattr(_proveedor_agente, "enviar_botones"):
                try:
                    await _proveedor_agente.enviar_botones(
                        msg.telefono,
                        datos_botones.get("texto", ""),
                        datos_botones.get("botones", [])
                    )
                    logger.info(f"Botones enviados a {msg.telefono}")
                except Exception as e:
                    logger.error(f"Error enviando botones: {e}")

            if datos_lista and hasattr(_proveedor_agente, "enviar_lista"):
                lista_enviada = False
                try:
                    lista_enviada = await _proveedor_agente.enviar_lista(
                        msg.telefono,
                        datos_lista.get("texto", ""),
                        datos_lista.get("boton", "Ver opciones"),
                        datos_lista.get("secciones", [])
                    )
                    if lista_enviada:
                        logger.info(f"Lista enviada a {msg.telefono}")
                except Exception as e:
                    logger.error(f"Error enviando lista: {e}")
                # Fallback: si la lista interactiva falla (límites de Meta,
                # demasiados items, etc.) enviamos los productos como texto
                if not lista_enviada:
                    lineas = [datos_lista.get("texto", "")]
                    for sec in datos_lista.get("secciones", []):
                        titulo = sec.get("titulo", "")
                        if titulo:
                            lineas.append(f"\n*{titulo}*")
                        for item in sec.get("items", []):
                            t = item.get("titulo", "")
                            d = item.get("descripcion", "")
                            if t and d:
                                lineas.append(f"• *{t}* — {d}")
                            elif t:
                                lineas.append(f"• {t}")
                    fallback_text = "\n".join(l for l in lineas if l)
                    if fallback_text.strip():
                        _txt_fb = fallback_text[:4000]
                        await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_fb)
                        await _guardar_con_wamid(_proveedor_agente, msg.telefono, _txt_fb, agent_id=_agent_id)
                        logger.info(f"Fallback texto de lista enviado a {msg.telefono}")

            # Cuando Andrea usa [[TIENDA:]] PRIORIZAR catálogo nativo de WhatsApp
            # — el cliente arma su pedido sin salir de WhatsApp. Solo si el
            # catálogo nativo falla (API caída, sin catalog_id, etc.) caemos
            # al fallback de URL a la tienda web.
            if abrir_tienda:
                # Pie del catálogo — configurable por agente desde el panel
                # (Configuración → Mensajes → "Mensaje de bienvenida del catálogo").
                # Los placeholders {minimo}, {envio_gratis}, {descuento_*}, {negocio}
                # se resuelven en runtime contra la config actual del agente —
                # cuando cambia el pedido mínimo en el panel, se refleja en el
                # próximo envío sin redeploy.
                from agent.mensajes import format_seguro
                _pie_template = await obtener_mensaje(_agent_id, "cart.bienvenida_catalogo")
                _ctx_ph = await construir_contexto_placeholders(_agent_id)
                pie_tienda = format_seguro(_pie_template, _ctx_ph)

                # ── Detectar si la query es una categoría conocida ──
                # Si es categoría → product_list de esa categoría (mejor UX)
                # Si no es categoría → catalog_message (botón "Ver catálogo")
                categoria_solicitada = None
                if tienda_query:
                    from agent.tools import ORDEN_CATEGORIAS, _categorias_cache
                    q_norm = tienda_query.lower().strip()
                    # Match exacto contra términos genéricos
                    mapa_terminos = {
                        "lavaloza": "Cocina", "desengrasante": "Cocina",
                        "limpiavidrios": "Cocina", "multiusos": "Cocina",
                        "detergente": "Lavandería", "shampoo": "Lavandería",
                        "desmanchador": "Baño", "ambientador": "Baño",
                        "combos": "Combos", "combo": "Combos",
                    }
                    categoria_solicitada = mapa_terminos.get(q_norm)
                    # Verificar que la categoría exista realmente
                    if categoria_solicitada and categoria_solicitada not in _categorias_cache:
                        categoria_solicitada = None

                # Si lo que Andrea escribió en query no es categoría reconocida
                # pero suena a producto, intentar enviar single_product
                if tienda_query and not categoria_solicitada:
                    # Intentar buscar el producto específico en _sku_map
                    from agent.tools import _sku_map, _normalizar
                    q_n = _normalizar(tienda_query)
                    rid_match = ""
                    for rid, info in _sku_map.get(_agent_id, {}).items():
                        if not (str(rid).isdigit() and len(str(rid)) >= 10):
                            continue
                        prod_n = _normalizar(info.get("producto", ""))
                        if q_n in prod_n or prod_n in q_n:
                            rid_match = rid
                            break
                    if rid_match and hasattr(_proveedor_agente, "enviar_producto"):
                        try:
                            res_prod = await _proveedor_agente.enviar_producto(msg.telefono, rid_match)
                            if (isinstance(res_prod, dict) and res_prod.get("ok")) or res_prod is True:
                                logger.info(f"[TIENDA→PRODUCTO] '{tienda_query}' → producto {rid_match} enviado")
                                # Saltarse el resto del bloque tienda
                                continue
                        except Exception as e:
                            logger.warning(f"[TIENDA→PRODUCTO] error: {e}")

                # Construir texto/cuerpo del catálogo. Si el agente desactivó
                # el mensaje de bienvenida (pie_tienda == ""), se envía solo
                # la parte de invitación sin pie — limpio, sin saltos vacíos.
                _pie_sufijo = ("\n" + pie_tienda) if pie_tienda else ""
                if _texto_absorbido_por_cta:
                    cuerpo_catalogo = respuesta.strip() + ("\n\n" + pie_tienda if pie_tienda else "")
                elif tienda_query:
                    cuerpo_catalogo = "🛒 *Aquí está lo que tenemos:*" + _pie_sufijo
                else:
                    cuerpo_catalogo = (
                        "🛒 *Aquí está nuestro catálogo completo — arma tu pedido sin salir de WhatsApp:*"
                        + _pie_sufijo
                    )

                # ── INTENTO 1: product_list de la categoría (o general) ──
                enviado_catalogo_wa = False
                if hasattr(_proveedor_agente, "enviar_catalogo_productos"):
                    try:
                        secciones = obtener_secciones_catalogo(categoria_solicitada)
                        if secciones:
                            header_cat = categoria_solicitada or await _catalogo_header_for_agent(_agent_id)
                            enviado_catalogo_wa = await _proveedor_agente.enviar_catalogo_productos(
                                msg.telefono, header_cat,
                                cuerpo_catalogo[:1024],
                                secciones,
                            )
                            if enviado_catalogo_wa:
                                logger.info(f"[TIENDA→catalogo-WA] product_list enviado a {msg.telefono} (cat={categoria_solicitada})")
                                # #79 — si el texto de Andrea se fusionó en este
                                # mensaje (_texto_absorbido_por_cta), el row de
                                # historial guardado antes (_msg_id_assistant)
                                # aún no tiene wamid: actualizarlo ahora.
                                if _texto_absorbido_por_cta:
                                    await _vincular_wamid_post_envio(_proveedor_agente, _msg_id_assistant)
                    except Exception as e:
                        logger.warning(f"[TIENDA] product_list falló: {e}")

                # ── FALLBACK: catálogo en texto ───────────────────────────
                # Si product_list falla (403 de permisos, API caída, etc.),
                # enviamos el catálogo como texto plano — mantiene el servicio
                # sin depender de la API interactiva de WhatsApp.
                # Solo escalamos si el texto tampoco llega (API completamente
                # caída), lo que generalmente indica un problema de plataforma
                # que ningún humano puede resolver vía WhatsApp en ese momento.
                if not enviado_catalogo_wa:
                    await _registrar_falla_catalogo(
                        "tienda_catalogo_wa", msg.telefono,
                        f"product_list falló (query='{tienda_query}')",
                        agent_id=_agent_id,
                    )
                    logger.info(f"[TIENDA→TEXTO] catálogo WA no disponible para {msg.telefono} — enviando texto")
                    try:
                        from agent.tools import obtener_catalogo_shopify
                        _cat_texto = await obtener_catalogo_shopify(_agent_id)
                        if _cat_texto:
                            _intro = (
                                f"🛒 *Aquí nuestros productos "
                                f"({'categoría ' + tienda_query if tienda_query else 'completos'}):*\n\n"
                                + _cat_texto[:3800]
                            )
                            _ok_txt = await _proveedor_agente.enviar_mensaje(msg.telefono, _intro)
                            if _ok_txt:
                                await _guardar_con_wamid(_proveedor_agente, msg.telefono, _intro, agent_id=_agent_id)
                                logger.info(f"[TIENDA→TEXTO] catálogo texto enviado a {msg.telefono}")
                            else:
                                logger.warning(f"[TIENDA→TEXTO] texto también falló para {msg.telefono} — API bloqueada, sin escalar")
                        else:
                            logger.warning(f"[TIENDA→TEXTO] catálogo vacío para agent_id={_agent_id}")
                    except Exception as _e_cat:
                        logger.error(f"[TIENDA→TEXTO] error obteniendo catálogo texto: {_e_cat}")

            # Enviar link de checkout de Shopify si se generó Y es válido
            # (con /checkouts/ — no la home). Si es inválido, intentar
            # autoreparar antes de mandar al cliente un botón muerto.
            if checkout_url and not _es_url_checkout_valida(checkout_url):
                logger.warning(f"[PEDIDO] checkout_url corrupto: {checkout_url[:60]} — intentando recrear")
                checkout_url = await _obtener_o_recrear_checkout_url(msg.telefono, agent_id=_agent_id)
            if checkout_url:
                # Mensaje y label del botón configurables por agente (#28 fase 2.6)
                from agent.mensajes import format_seguro
                _ctx_ph = await construir_contexto_placeholders(_agent_id)
                # Inyectar {total} — primero intentamos del carrito_activo
                # (autoridad de la verdad). Si está vacío, fallback a 0.
                _ctx_ph["total"] = await _total_carrito_fmt(msg.telefono, _agent_id)
                texto_checkout = format_seguro(
                    await obtener_mensaje(_agent_id, "cart.checkout_listo_texto"),
                    _ctx_ph,
                )
                boton_label = (await obtener_mensaje(_agent_id, "cart.checkout_listo_boton")).strip() or "Confirmar pedido"
                enviado_cta = False
                if hasattr(_proveedor_agente, "enviar_cta_url"):
                    try:
                        enviado_cta = await _proveedor_agente.enviar_cta_url(
                            msg.telefono, texto_checkout, boton_label, checkout_url
                        )
                    except Exception as e:
                        logger.error(f"Error enviando cta_url: {e}")
                if not enviado_cta:
                    await _proveedor_agente.enviar_mensaje(
                        msg.telefono, f"{texto_checkout}\n\n👉 {checkout_url}"
                    )

            # Si la creación del checkout falló (stock agotado, producto no
            # mapeado, etc.) avísale al cliente en vez de quedarnos mudos
            if checkout_fallo:
                _txt_chk_fail = await obtener_mensaje(_agent_id, "error.checkout_no_generado")
                if _txt_chk_fail:
                    await _proveedor_agente.enviar_mensaje(msg.telefono, _txt_chk_fail)

            # Notificar al equipo si Andrea decidió escalar
            if datos_escalacion:
                await _notificar_escalacion(msg.telefono, datos_escalacion, agent_id=_agent_id)

            logger.info(f"Respuesta a {msg.telefono}: {respuesta}")

        return {"status": "ok"}

    except Exception as e:
        logger.error(f"Error en webhook: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════
#  INTEGRACIÓN LOVABLE — recibe estado del carrito desde equoradistribuciones.com
# ═══════════════════════════════════════════════════════════════

@app.post("/shopify/cart-update")
async def shopify_cart_update(request: Request):
    """
    La página de Lovable (equoradistribuciones.com) llama este endpoint
    cada vez que el carrito cambia (debounced 8 s en el cliente).
    Guarda el estado en BD para que el loop de seguimientos pueda:
      - Detectar carritos abandonados y enviar recuperación por WhatsApp
      - Disparar cross-sell si el total está bajo $80.000
    Payload esperado:
      { "telefono": "573001234567", "productos": [ { "producto": "...",
        "presentacion": "...", "cantidad": N, "precio_unitario": X } ] }
    """
    try:
        body = await request.json()
        telefono = (body.get("telefono") or "").strip()
        productos = body.get("productos") or []

        if not telefono:
            # Sin teléfono no podemos asociar a una conversación
            # ADVERTENCIA: la página Lovable debe incluir siempre el teléfono en el payload
            logger.warning(
                f"[cart-update] SIN teléfono — payload ignorado. "
                f"productos={len(body.get('productos') or [])}. "
                f"Revisa que Lovable incluya 'telefono' en cada cart-update."
            )
            return JSONResponse(content={"ok": True})

        if productos:
            items_para_bd = []
            total_recibido = 0
            for p in productos:
                precio = int(float(p.get("precio_unitario") or p.get("precio") or 0))
                qty = int(p.get("cantidad") or p.get("qty") or 1)
                subtotal = precio * qty
                total_recibido += subtotal
                items_para_bd.append({
                    "producto":      p.get("producto", ""),
                    "presentacion":  p.get("presentacion", ""),
                    "cantidad":      qty,
                    "precio_unitario": precio,
                    "subtotal":      subtotal,
                    # retailer_id es CRÍTICO para que el evento CAPI
                    # AddToCart matchee con el catálogo Meta (sin esto
                    # la proporción de coincidencias cae <90% y los
                    # eventos no optimizan campañas)
                    "retailer_id":   (p.get("retailer_id")
                                      or p.get("product_retailer_id")
                                      or p.get("sku") or ""),
                })
            await guardar_carrito_activo(telefono, items_para_bd)
            logger.info(
                f"[cart-update] {telefono}: {len(productos)} productos, "
                f"total=${total_recibido:,}"
            )
            # CAPI: AddToCart — cliente agregó productos en la tienda web
            asyncio.create_task(capi_add_to_cart(telefono, total_recibido, items_para_bd))
        else:
            # Carrito vacío → limpiar
            await limpiar_carrito_activo(telefono)
            logger.info(f"[cart-update] {telefono}: carrito vaciado")

        return JSONResponse(content={"ok": True})
    except Exception as e:
        logger.error(f"Error en /shopify/cart-update: {e}")
        return JSONResponse(content={"ok": True})  # siempre 200 para no bloquear al cliente




# ═══════════════════════════════════════════════════════════════
#  INBOX — Panel de administración
# ═══════════════════════════════════════════════════════════════

INBOX_COOKIE = "inbox_session"
VOCO_COOKIE = "voco_session"
COOKIE_MAX_AGE = 60 * 60 * 24 * 30  # 30 días


def _verificar_admin(token: str) -> bool:
    """Valida el token de administrador (legacy sync — mantiene compatibilidad con 47 call sites)."""
    if not ADMIN_TOKEN:
        return False
    return hmac.compare_digest(token or "", ADMIN_TOKEN)


async def _obtener_sesion_usuario(token: str) -> dict | None:
    """Retorna el dict del usuario autenticado por token de sesión, o None.
    Prueba primero el nuevo sistema de sesiones; fallback al ADMIN_TOKEN legacy."""
    if not token:
        return None
    # Nuevo sistema: sesión de usuario registrado en BD
    try:
        user = await verificar_sesion(token)
        if user:
            return user
    except Exception as e:
        logger.warning(f"[sesion] Error verificando sesión DB: {e}")
    # Fallback legacy: ADMIN_TOKEN para compatibilidad con setup de Railway existente
    if ADMIN_TOKEN and hmac.compare_digest(token, ADMIN_TOKEN):
        return {"id": 0, "email": "admin@voco.local", "rol": "admin", "nombre": "Admin", "plan": "admin"}
    # Sin ADMIN_TOKEN configurado → cualquier acceso directo es admin (solo dev)
    if not ADMIN_TOKEN:
        return {"id": 0, "email": "admin@voco.local", "rol": "admin", "nombre": "Admin", "plan": "admin"}
    return None


def _token_de_request(
    token: str = "",
    inbox_session: str = Cookie(default=""),
) -> str:
    """Extrae el token de la cookie o del query param."""
    return inbox_session or token


# ── Login / Logout ─────────────────────────────────────────────────────────

@app.get("/inbox/login", response_class=HTMLResponse)
async def inbox_login_page():
    return HTMLResponse(content=obtener_login_html())


@app.post("/inbox/login")
async def inbox_login(request: Request):
    form = await request.form()
    password = str(form.get("password", ""))
    email = str(form.get("email", "")).strip().lower()

    # Intento 1: login con email + contraseña (nuevo sistema)
    if email:
        user_data = await obtener_usuario_por_email(email)
        if user_data and user_data.get("is_active"):
            try:
                pw_match = bcrypt.checkpw(password.encode(), user_data["password_hash"].encode())
            except Exception:
                pw_match = False
            if pw_match:
                token = await crear_sesion(user_data["id"])
                response = RedirectResponse("/inbox", status_code=302)
                response.set_cookie(
                    VOCO_COOKIE, token,
                    httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax"
                )
                # Compatibilidad con endpoints legacy que leen inbox_session
                if ADMIN_TOKEN:
                    response.set_cookie(
                        INBOX_COOKIE, ADMIN_TOKEN,
                        httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax"
                    )
                return response
        return HTMLResponse(content=obtener_login_html(error=True))

    # Intento 2: login legacy con ADMIN_TOKEN (campo password)
    if _verificar_admin(password):
        response = RedirectResponse("/inbox", status_code=302)
        response.set_cookie(
            INBOX_COOKIE, ADMIN_TOKEN,
            httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax"
        )
        return response
    return HTMLResponse(content=obtener_login_html(error=True))


@app.get("/inbox/logout")
async def inbox_logout(
    voco_session: str = Cookie(default=""),
    inbox_session: str = Cookie(default=""),
):
    if voco_session:
        await cerrar_sesion(voco_session)
    response = RedirectResponse("/inbox/login", status_code=302)
    response.delete_cookie(INBOX_COOKIE)
    response.delete_cookie(VOCO_COOKIE)
    return response


# ── Registro de usuarios (Sprint 1) ─────────────────────────────────────────

@app.get("/auth/register", response_class=HTMLResponse)
async def auth_register_page():
    return HTMLResponse(content=obtener_register_html())


@app.post("/auth/register")
async def auth_register(request: Request):
    form = await request.form()
    nombre   = str(form.get("nombre", "")).strip()
    email    = str(form.get("email", "")).strip().lower()
    password = str(form.get("password", ""))
    confirm  = str(form.get("confirm_password", ""))

    # Validaciones
    if not nombre or not email or not password:
        return HTMLResponse(content=obtener_register_html(error="Todos los campos son obligatorios"))
    if len(password) < 8:
        return HTMLResponse(content=obtener_register_html(error="La contraseña debe tener al menos 8 caracteres"))
    if password != confirm:
        return HTMLResponse(content=obtener_register_html(error="Las contraseñas no coinciden"))

    # Verificar que el email no exista ya
    existente = await obtener_usuario_por_email(email)
    if existente:
        return HTMLResponse(content=obtener_register_html(error="Este email ya está registrado"))

    # Crear usuario
    try:
        pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        user = await crear_usuario(email=email, password_hash=pw_hash, nombre=nombre)
        token = await crear_sesion(user["id"])
        response = RedirectResponse("/inbox", status_code=302)
        response.set_cookie(
            VOCO_COOKIE, token,
            httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax"
        )
        return response
    except Exception as e:
        logger.error(f"[auth/register] Error: {e}")
        return HTMLResponse(content=obtener_register_html(error="Error al crear la cuenta. Intenta de nuevo."))


@app.post("/auth/logout")
async def auth_logout(
    voco_session: str = Cookie(default=""),
    inbox_session: str = Cookie(default=""),
):
    if voco_session:
        await cerrar_sesion(voco_session)
    response = RedirectResponse("/inbox/login", status_code=302)
    response.delete_cookie(INBOX_COOKIE)
    response.delete_cookie(VOCO_COOKIE)
    return response


# ── OAuth Social Login ────────────────────────────────────────────────────

GOOGLE_CLIENT_ID     = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
FACEBOOK_APP_ID      = os.getenv("FACEBOOK_APP_ID", "")
FACEBOOK_APP_SECRET  = os.getenv("FACEBOOK_APP_SECRET", "")
APP_BASE_URL         = os.getenv("APP_BASE_URL", "https://tienda.equoradistribuciones.com")


@app.get("/auth/google")
async def auth_google():
    """Redirige al flujo OAuth de Google."""
    if not GOOGLE_CLIENT_ID:
        return RedirectResponse("/inbox/login?error=google_not_configured", status_code=302)
    params = (
        f"client_id={GOOGLE_CLIENT_ID}"
        f"&redirect_uri={APP_BASE_URL}/auth/google/callback"
        "&response_type=code"
        "&scope=openid%20email%20profile"
        "&access_type=offline"
    )
    return RedirectResponse(f"https://accounts.google.com/o/oauth2/v2/auth?{params}", status_code=302)


@app.get("/auth/google/callback")
async def auth_google_callback(code: str = "", error: str = ""):
    """Callback OAuth de Google — intercambia code por token y crea sesión."""
    if error or not code:
        return RedirectResponse("/inbox/login?error=google_cancelled", status_code=302)
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            # Intercambiar code por access_token
            r = await client.post("https://oauth2.googleapis.com/token", data={
                "code": code,
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": f"{APP_BASE_URL}/auth/google/callback",
                "grant_type": "authorization_code",
            })
            tokens = r.json()
            access_token = tokens.get("access_token")
            if not access_token:
                raise ValueError(f"No access_token: {tokens}")
            # Obtener perfil del usuario
            profile = (await client.get(
                "https://www.googleapis.com/oauth2/v3/userinfo",
                headers={"Authorization": f"Bearer {access_token}"},
            )).json()
        email  = (profile.get("email") or "").lower().strip()
        nombre = profile.get("name") or profile.get("given_name") or email
        if not email:
            return RedirectResponse("/inbox/login?error=google_no_email", status_code=302)
        # Buscar o crear usuario
        user_data = await obtener_usuario_por_email(email)
        if not user_data:
            pw_hash = bcrypt.hashpw(secrets.token_urlsafe(24).encode(), bcrypt.gensalt()).decode()
            user_data = await crear_usuario(email=email, password_hash=pw_hash, nombre=nombre)
        if not user_data.get("is_active", True):
            return RedirectResponse("/inbox/login?error=cuenta_inactiva", status_code=302)
        token = await crear_sesion(user_data["id"])
        response = RedirectResponse("/inbox", status_code=302)
        response.set_cookie(VOCO_COOKIE, token, httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax")
        if ADMIN_TOKEN:
            response.set_cookie(INBOX_COOKIE, ADMIN_TOKEN, httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax")
        return response
    except Exception as e:
        logger.error(f"[OAuth Google] {e}")
        return RedirectResponse("/inbox/login?error=google_error", status_code=302)


@app.get("/auth/facebook")
async def auth_facebook():
    """Redirige al flujo OAuth de Facebook."""
    if not FACEBOOK_APP_ID:
        return RedirectResponse("/inbox/login?error=facebook_not_configured", status_code=302)
    params = (
        f"client_id={FACEBOOK_APP_ID}"
        f"&redirect_uri={APP_BASE_URL}/auth/facebook/callback"
        "&scope=email,public_profile"
        "&response_type=code"
    )
    return RedirectResponse(f"https://www.facebook.com/v21.0/dialog/oauth?{params}", status_code=302)


@app.get("/auth/facebook/callback")
async def auth_facebook_callback(code: str = "", error: str = ""):
    """Callback OAuth de Facebook — intercambia code por token y crea sesión."""
    if error or not code:
        return RedirectResponse("/inbox/login?error=facebook_cancelled", status_code=302)
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            # Intercambiar code por access_token
            r = await client.get("https://graph.facebook.com/v21.0/oauth/access_token", params={
                "client_id": FACEBOOK_APP_ID,
                "client_secret": FACEBOOK_APP_SECRET,
                "redirect_uri": f"{APP_BASE_URL}/auth/facebook/callback",
                "code": code,
            })
            tokens = r.json()
            access_token = tokens.get("access_token")
            if not access_token:
                raise ValueError(f"No access_token: {tokens}")
            # Obtener perfil
            profile = (await client.get(
                "https://graph.facebook.com/me",
                params={"fields": "id,name,email", "access_token": access_token},
            )).json()
        email  = (profile.get("email") or "").lower().strip()
        nombre = profile.get("name") or email
        if not email:
            return RedirectResponse("/inbox/login?error=facebook_no_email", status_code=302)
        user_data = await obtener_usuario_por_email(email)
        if not user_data:
            pw_hash = bcrypt.hashpw(secrets.token_urlsafe(24).encode(), bcrypt.gensalt()).decode()
            user_data = await crear_usuario(email=email, password_hash=pw_hash, nombre=nombre)
        if not user_data.get("is_active", True):
            return RedirectResponse("/inbox/login?error=cuenta_inactiva", status_code=302)
        token = await crear_sesion(user_data["id"])
        response = RedirectResponse("/inbox", status_code=302)
        response.set_cookie(VOCO_COOKIE, token, httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax")
        if ADMIN_TOKEN:
            response.set_cookie(INBOX_COOKIE, ADMIN_TOKEN, httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax")
        return response
    except Exception as e:
        logger.error(f"[OAuth Facebook] {e}")
        return RedirectResponse("/inbox/login?error=facebook_error", status_code=302)


# ── Gestión de agentes (Voco platform) ────────────────────────────────────

@app.get("/inbox/api/agents")
async def inbox_listar_agentes(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista agentes: admin ve todos, usuario solo los suyos."""
    effective_token = voco_session or inbox_session or token
    current_user = await _obtener_sesion_usuario(effective_token)
    if not current_user:
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_todos_agentes, obtener_agentes_de_usuario
    if current_user.get("rol") == "admin":
        agentes = await obtener_todos_agentes()
    else:
        agentes = await obtener_agentes_de_usuario(current_user["id"])
    return JSONResponse(content={"agents": agentes})


@app.post("/inbox/api/agents")
async def inbox_crear_agente(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Crea un nuevo agente en la plataforma. Asigna owner_id al usuario actual."""
    effective_token = voco_session or inbox_session or token
    current_user = await _obtener_sesion_usuario(effective_token)
    if not current_user:
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import crear_agente
    body = await request.json()
    slug         = (body.get("slug") or "").strip().lower().replace(" ", "-")
    name         = (body.get("name") or "").strip()
    agent_name   = (body.get("agent_name") or "Agente").strip()
    business_type = (body.get("business_type") or "productos").strip()
    phone_number_id = (body.get("phone_number_id") or "").strip()
    waba_id      = (body.get("waba_id") or "").strip()
    color        = (body.get("color") or "#6366f1").strip()
    emoji        = (body.get("emoji") or "🤖").strip()

    if not slug or not name:
        return JSONResponse(status_code=400, content={"error": "slug y name son requeridos"})

    # Admin (id=0 legacy o rol=admin) → owner_id=None (admin-owned)
    # Usuario regular → owner_id = su id
    user_id = current_user.get("id", 0)
    owner_id = None if (current_user.get("rol") == "admin" or user_id == 0) else user_id

    try:
        agente = await crear_agente(
            slug=slug, name=name, agent_name=agent_name,
            business_type=business_type, phone_number_id=phone_number_id,
            waba_id=waba_id, color=color, emoji=emoji,
            owner_id=owner_id,
        )
        return JSONResponse(content={"ok": True, "agent": agente})
    except Exception as e:
        logger.error(f"[agents] Error creando agente: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/inbox/api/agents/{agent_id_param}")
async def inbox_actualizar_agente(
    agent_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Actualiza campos de un agente existente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import actualizar_agente, obtener_agente
    body = await request.json()
    try:
        ok = await actualizar_agente(agent_id_param, **body)
        if not ok:
            return JSONResponse(status_code=404, content={"error": "Agente no encontrado"})
        # Invalidar cache de phone_number_id para que el webhook lo recargue
        if "phone_number_id" in body:
            for key in list(_phone_agent_cache.keys()):
                if _phone_agent_cache[key].get("id") == agent_id_param:
                    del _phone_agent_cache[key]
        agente_actualizado = await obtener_agente(agent_id_param)
        return JSONResponse(content={"ok": True, "agent": agente_actualizado})
    except Exception as e:
        logger.error(f"[agents] Error actualizando agente {agent_id_param}: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/inbox/api/agents/{agent_id_param}/activate")
async def inbox_activar_agente(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Activa un agente (status → active)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import actualizar_agente
    await actualizar_agente(agent_id_param, status="active")
    return JSONResponse(content={"ok": True})


# ──────────────────────────────────────────────────────────────────────────────
# #43 — Promoción post-venta (código descuento tras pago)
# ──────────────────────────────────────────────────────────────────────────────
async def _verificar_acceso_agente(user: dict, agent_id: int) -> bool:
    """True si el usuario puede modificar este agente.

    Permitido: admin (todos) o owner del agente. Usuarios internos sin
    rol admin solo pueden tocar el agente que poseen.
    """
    if not user:
        return False
    if user.get("rol") == "admin" or user.get("id", 0) == 0:
        return True  # admins ven/editan todo
    from agent.memory import obtener_agente
    agente = await obtener_agente(agent_id)
    if not agente:
        return False
    # owner_id None = admin-owned (legacy); owner_id == user.id = suyo
    return agente.get("owner_id") in (None, user.get("id"))


@app.get("/inbox/api/agents/{agent_id_param}/descuento")
async def api_obtener_descuento(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve la config de descuento post-venta del agente.

    Siempre retorna el shape completo (activo + 4 campos), aunque esté
    desactivado — el form del panel necesita rellenar los inputs.
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if not await _verificar_acceso_agente(user, agent_id_param):
        raise HTTPException(status_code=403, detail="Sin acceso a este agente")
    from agent.memory import obtener_descuento_promo_config, DESCUENTO_MENSAJE_DEFAULT
    config = await obtener_descuento_promo_config(agent_id_param)
    # Exponemos también el default para que la UI pueda mostrarlo como placeholder
    config["mensaje_default"] = DESCUENTO_MENSAJE_DEFAULT
    return JSONResponse(content=config)


# ──────────────────────────────────────────────────────────────────────────────
# #28 — Mensajes del sistema configurables por agente
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/inbox/api/agents/{agent_id_param}/mensajes")
async def api_listar_mensajes(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el catálogo COMPLETO de mensajes con el valor actual de cada
    uno y la metadata para que la UI lo pueda renderizar agrupado por
    categoría sin hacer queries adicionales.

    Response shape:
      {
        "categorias": {<slug>: {<label>, <descripcion>, <orden>}, ...},
        "mensajes":   [{<key>, <categoria>, <titulo>, <descripcion>, <cuando>,
                        <default>, <placeholders>, <content>, <personalizado>,
                        <updated_at>}, ...]
      }
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if not await _verificar_acceso_agente(user, agent_id_param):
        raise HTTPException(status_code=403, detail="Sin acceso a este agente")
    from agent.memory import listar_mensajes_agente
    from agent.mensajes import CATEGORIAS
    mensajes = await listar_mensajes_agente(agent_id_param)
    return JSONResponse(content={
        "categorias": CATEGORIAS,
        "mensajes":   mensajes,
    })


@app.put("/inbox/api/agents/{agent_id_param}/mensajes/{key}")
async def api_guardar_mensaje(
    agent_id_param: int,
    key: str,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Guarda el override del mensaje. Body: {"content": "..."}.

    Validaciones en memory.guardar_mensaje_agente: key del catálogo,
    no vacío, <=4000 chars, placeholders requeridos presentes.
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if not await _verificar_acceso_agente(user, agent_id_param):
        raise HTTPException(status_code=403, detail="Sin acceso a este agente")
    from agent.memory import guardar_mensaje_agente
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "JSON inválido"})
    content = str(body.get("content") or "")
    ok, error = await guardar_mensaje_agente(agent_id_param, key, content)
    if not ok:
        return JSONResponse(status_code=400, content={"ok": False, "error": error})
    logger.info(
        f"[mensajes] agent_id={agent_id_param} actualizó {key!r} "
        f"({len(content)} chars) — user={user.get('id')}"
    )
    return JSONResponse(content={"ok": True})


@app.patch("/inbox/api/agents/{agent_id_param}/mensajes/{key}/activo")
async def api_set_mensaje_activo(
    agent_id_param: int,
    key: str,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Activa o desactiva un mensaje sin modificar su contenido.

    Body: {"activo": bool}
    Si el mensaje del catálogo está marcado como esencial (puede_desactivarse=False),
    se rechaza con 400.
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if not await _verificar_acceso_agente(user, agent_id_param):
        raise HTTPException(status_code=403, detail="Sin acceso a este agente")
    from agent.memory import set_mensaje_activo
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "JSON inválido"})
    activo = bool(body.get("activo"))
    ok, error = await set_mensaje_activo(agent_id_param, key, activo)
    if not ok:
        return JSONResponse(status_code=400, content={"ok": False, "error": error})
    logger.info(
        f"[mensajes] agent_id={agent_id_param} {key!r} → activo={activo} "
        f"— user={user.get('id')}"
    )
    return JSONResponse(content={"ok": True, "activo": activo})


@app.delete("/inbox/api/agents/{agent_id_param}/mensajes/{key}")
async def api_restaurar_mensaje(
    agent_id_param: int,
    key: str,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Restaura el mensaje al default — borra el override del agente.

    Idempotente: si no había override, devuelve ok=true sin cambios.
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if not await _verificar_acceso_agente(user, agent_id_param):
        raise HTTPException(status_code=403, detail="Sin acceso a este agente")
    from agent.memory import restaurar_mensaje_agente
    existia = await restaurar_mensaje_agente(agent_id_param, key)
    logger.info(
        f"[mensajes] agent_id={agent_id_param} restauró {key!r} "
        f"(habia_override={existia}) — user={user.get('id')}"
    )
    return JSONResponse(content={"ok": True, "habia_override": existia})


@app.put("/inbox/api/agents/{agent_id_param}/descuento")
async def api_guardar_descuento(
    agent_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Guarda la config de descuento post-venta del agente.

    Body esperado:
      {
        "activo":  bool,
        "umbral":  int,    # COP, sin envío
        "codigo":  str,    # se normaliza a MAYÚSCULAS
        "pct":     int,    # 1-100
        "mensaje": str     # template con {codigo} {pct} {umbral} — opcional
      }

    Validaciones en memory.guardar_descuento_promo(): umbral >= 1000,
    código [A-Z0-9_-]{2,30}, pct 1-100, mensaje debe contener {codigo}.
    Si activo=false, los demás campos se persisten igual para permitir
    reactivar más tarde sin perder la configuración.
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if not await _verificar_acceso_agente(user, agent_id_param):
        raise HTTPException(status_code=403, detail="Sin acceso a este agente")
    from agent.memory import guardar_descuento_promo
    try:
        body = await request.json()
    except Exception:
        return JSONResponse(status_code=400, content={"ok": False, "error": "JSON inválido"})

    activo  = bool(body.get("activo"))
    try:
        umbral = int(body.get("umbral") or 0)
        pct    = int(body.get("pct") or 0)
    except (TypeError, ValueError):
        return JSONResponse(status_code=400, content={"ok": False, "error": "umbral y pct deben ser números enteros"})
    codigo  = str(body.get("codigo") or "").strip()
    mensaje = str(body.get("mensaje") or "").strip()

    ok, error = await guardar_descuento_promo(
        agent_id_param, activo, umbral, codigo, pct, mensaje,
    )
    if not ok:
        return JSONResponse(status_code=400, content={"ok": False, "error": error})
    logger.info(
        f"[descuento] agent_id={agent_id_param} actualizado por user={user.get('id')} "
        f"(activo={activo}, codigo={codigo}, umbral={umbral}, pct={pct})"
    )

    # ── #54 fase 3 — Auto-crear cupón en Shopify si hay Admin API ─────────────
    # Si el agente activó la promo Y tiene SHOPIFY_ADMIN_TOKEN configurado,
    # intentamos crear el discount code automáticamente en Shopify. Esto evita
    # que el merchant tenga que duplicar el código en 2 lugares (Voco + Shopify).
    # Es best-effort: si Shopify falla (ya existe, scope faltante, etc.) la
    # respuesta al panel sigue ok=True pero incluye `shopify_setup` con detalle.
    respuesta_extra: dict = {}
    if activo and codigo:
        from agent.memory import get_config_value
        store_raw = (await get_config_value("SHOPIFY_STORE", agent_id_param)
                     or os.getenv("SHOPIFY_STORE", "")).strip()
        admin_tok = (await get_config_value("SHOPIFY_ADMIN_TOKEN", agent_id_param)
                     or os.getenv("SHOPIFY_ADMIN_TOKEN", "")).strip()
        if store_raw and admin_tok:
            from agent.shopify_admin import buscar_discount_code, crear_discount_code
            store = store_raw.replace("https://", "").replace("http://", "").rstrip("/")
            try:
                existente = await buscar_discount_code(store, admin_tok, codigo)
                if existente:
                    respuesta_extra["shopify_setup"] = {
                        "ok": True,
                        "estado": "ya_existe",
                        "mensaje": f"El código '{codigo}' ya existe en Shopify — Voco lo anunciará tal como esté configurado allá.",
                    }
                    logger.info(f"[descuento] '{codigo}' ya existe en Shopify para agent {agent_id_param} — no recreo")
                else:
                    res = await crear_discount_code(store, admin_tok, codigo, pct, umbral)
                    if res.get("ok"):
                        respuesta_extra["shopify_setup"] = {
                            "ok": True,
                            "estado": "creado",
                            "mensaje": f"Cupón '{codigo}' creado automáticamente en Shopify ({pct}% desde ${umbral:,}).",
                            "price_rule_id":    res.get("price_rule_id"),
                            "discount_code_id": res.get("discount_code_id"),
                        }
                        logger.info(f"[descuento] cupón '{codigo}' creado en Shopify para agent {agent_id_param}")
                    else:
                        respuesta_extra["shopify_setup"] = {
                            "ok": False,
                            "estado": "error",
                            "mensaje": f"No pude crear el cupón en Shopify: {res.get('error', 'error desconocido')}. Créalo manualmente en Shopify Admin → Discounts.",
                        }
                        logger.warning(f"[descuento] crear cupón '{codigo}' falló: {res.get('error')}")
            except Exception as e:
                logger.error(f"[descuento] error con Shopify Admin: {e}")
                respuesta_extra["shopify_setup"] = {
                    "ok": False,
                    "estado": "error",
                    "mensaje": f"Error contactando Shopify: {e!s}. Créalo manualmente.",
                }
        # Si no hay Admin token configurado, simplemente no intentamos —
        # la respuesta normal (ok=True) sigue. El cliente puede crear el
        # cupón a mano en Shopify Admin como hasta ahora.

    return JSONResponse(content={"ok": True, **respuesta_extra})


# ──────────────────────────────────────────────────────────────────────────────
# Sprint A — Módulos por agente (modules_json)
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/inbox/api/agents/{agent_id_param}/modules")
async def inbox_obtener_modulos_agente(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el dict de módulos activos para un agente. Defaults seguros."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import get_agent_modules
    modules = await get_agent_modules(agent_id_param)
    return JSONResponse(content={"ok": True, "modules": modules})


@app.post("/inbox/api/agents/{agent_id_param}/modules")
async def inbox_guardar_modulos_agente(
    agent_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Guarda los módulos del agente. Espera body JSON con claves válidas.

    Solo procesa claves en DEFAULT_MODULES; el resto se ignora silenciosamente.
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import set_agent_modules
    try:
        body = await request.json()
    except Exception:
        body = {}
    modules_in = body.get("modules", body) if isinstance(body, dict) else {}
    saved = await set_agent_modules(agent_id_param, modules_in)
    return JSONResponse(content={"ok": True, "modules": saved})


@app.post("/inbox/api/agents/{agent_id_param}/pause")
async def inbox_pausar_agente(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Pausa un agente (status → paused). No afecta al agente Equora (id=1)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    if agent_id_param == 1:
        return JSONResponse(status_code=400, content={"error": "El agente Equora no puede pausarse desde la API"})
    from agent.memory import actualizar_agente
    await actualizar_agente(agent_id_param, status="paused")
    return JSONResponse(content={"ok": True})


@app.post("/inbox/api/agents/{agent_id_param}/clone")
async def inbox_clonar_agente(
    agent_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Clona un agente existente con un nuevo slug."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_agente, crear_agente, get_config_value, set_config_value
    body = await request.json()
    nuevo_slug = (body.get("slug") or "").strip().lower().replace(" ", "-")
    nuevo_name = (body.get("name") or "").strip()
    if not nuevo_slug:
        return JSONResponse(status_code=400, content={"error": "slug requerido"})

    original = await obtener_agente(agent_id_param)
    if not original:
        return JSONResponse(status_code=404, content={"error": "Agente no encontrado"})

    try:
        nuevo = await crear_agente(
            slug=nuevo_slug,
            name=nuevo_name or f"{original.get('name', '')} (copia)",
            agent_name=original.get("agent_name", "Agente"),
            business_type=original.get("business_type", "productos"),
            phone_number_id="",  # El nuevo agente empieza sin número asignado
            waba_id="",
            color=original.get("color", "#6366f1"),
            emoji=original.get("emoji", "🤖"),
        )
        # Clonar config_values del agente original (prompt, vars, etc.)
        nuevo_id = nuevo.get("id")
        for clave in ("SYSTEM_PROMPT", "BUSINESS_VARS", "BUSINESS_TYPE", "ACTIVE_MODULES"):
            valor = await get_config_value(clave, agent_id_param)
            if valor:
                await set_config_value(clave, valor, nuevo_id)
        return JSONResponse(content={"ok": True, "agent": nuevo})
    except Exception as e:
        logger.error(f"[agents] Error clonando agente {agent_id_param}: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/inbox/api/agents/{agent_id_param}/sandbox")
async def inbox_sandbox_info(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Retorna la configuración del sandbox hub de Voco para este agente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_agente

    # El sandbox usa el WABA propio de Voco — credenciales a nivel de plataforma
    voco_sb_phone_id = os.getenv("VOCO_SANDBOX_PHONE_NUMBER_ID", "")
    voco_sb_phone    = os.getenv("VOCO_SANDBOX_PHONE_NUMBER", "")

    if not voco_sb_phone_id or not voco_sb_phone:
        return JSONResponse(content={"configured": False, "phone_number": "", "phone_number_id": ""})

    # Código de proyecto específico para este agente (VOCO-EQ001)
    _src_agent = await obtener_agente(agent_id_param)
    _src_slug  = _src_agent.get("slug", "agente") if _src_agent else "agente"
    _sandbox_code = _sandbox_code_para_agente(_src_slug, agent_id_param)

    return JSONResponse(content={
        "configured": True,
        "phone_number": voco_sb_phone,
        "phone_number_id": voco_sb_phone_id,
        "active": True,
        "code": _sandbox_code,
    })


@app.get("/inbox/api/qr")
async def generar_qr(data: str = ""):
    """Genera un QR code como SVG a partir del parámetro ?data= ."""
    if not data:
        return Response(content="<svg/>", media_type="image/svg+xml")
    try:
        import qrcode
        import qrcode.image.svg as _svg
        from io import BytesIO
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=4,
            border=2,
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(image_factory=_svg.SvgPathImage)
        buf = BytesIO()
        img.save(buf)
        buf.seek(0)
        return Response(content=buf.read(), media_type="image/svg+xml")
    except Exception as e:
        logger.error(f"[qr] Error generando QR: {e}")
        return Response(content="<svg/>", media_type="image/svg+xml")


@app.post("/inbox/api/agents/{agent_id_param}/sandbox/activar")
async def inbox_sandbox_activar(
    agent_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Crea el agente sandbox copiando credenciales del agente origen."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import (
        obtener_agente, crear_agente, obtener_agente_por_phone_id,
        get_config_value, set_config_value,
    )

    body = await request.json()
    sandbox_phone_id = (body.get("phone_number_id") or os.getenv("META_SANDBOX_PHONE_NUMBER_ID", "")).strip()
    sandbox_phone_number = (body.get("phone_number") or os.getenv("META_SANDBOX_PHONE_NUMBER", "")).strip()

    if not sandbox_phone_id:
        return JSONResponse(status_code=400, content={"error": "phone_number_id requerido"})

    # Si ya existe un agente con ese phone_number_id no crear duplicado
    existing = await obtener_agente_por_phone_id(sandbox_phone_id)
    if existing:
        return JSONResponse(content={"ok": True, "sandbox": existing, "created": False})

    original = await obtener_agente(agent_id_param)
    if not original:
        return JSONResponse(status_code=404, content={"error": "Agente origen no encontrado"})

    access_token   = await get_config_value("META_ACCESS_TOKEN",  agent_id_param) or os.getenv("META_ACCESS_TOKEN", "")
    verify_token   = await get_config_value("META_VERIFY_TOKEN",  agent_id_param) or os.getenv("META_VERIFY_TOKEN", "")
    system_prompt  = await get_config_value("SYSTEM_PROMPT",      agent_id_param) or ""
    active_modules = await get_config_value("ACTIVE_MODULES",     agent_id_param) or ""
    business_vars  = await get_config_value("BUSINESS_VARS",      agent_id_param) or ""

    try:
        sandbox = await crear_agente(
            name=f"{original.get('name', 'Agente')} (Sandbox)",
            slug=f"{original.get('slug', 'agente')}-sandbox",
            agent_name=original.get("agent_name", "Agente"),
            business_type=original.get("business_type", "productos"),
            phone_number_id=sandbox_phone_id,
            waba_id="",
            color=original.get("color", "#6366f1"),
            emoji="🧪",
        )
        sid = sandbox["id"]
        await set_config_value("META_ACCESS_TOKEN",  access_token,   sid)
        await set_config_value("META_PHONE_NUMBER_ID", sandbox_phone_id, sid)
        await set_config_value("META_VERIFY_TOKEN",  verify_token,   sid)
        if system_prompt:
            await set_config_value("SYSTEM_PROMPT",  system_prompt,  sid)
        if active_modules:
            await set_config_value("ACTIVE_MODULES", active_modules, sid)
        if business_vars:
            await set_config_value("BUSINESS_VARS",  business_vars,  sid)

        # Guardar referencia en el agente origen para que la UI la encuentre
        await set_config_value("SANDBOX_PHONE_NUMBER_ID", sandbox_phone_id,   agent_id_param)
        await set_config_value("SANDBOX_PHONE_NUMBER",    sandbox_phone_number, agent_id_param)

        # Limpiar caché de routing para que el webhook reconozca el nuevo número
        _phone_agent_cache.pop(sandbox_phone_id, None)

        logger.info(f"[sandbox] Agente sandbox creado (id={sid}) para agent_id={agent_id_param}")
        return JSONResponse(content={"ok": True, "sandbox": sandbox, "created": True})
    except Exception as e:
        logger.error(f"[sandbox] Error activando sandbox para agente {agent_id_param}: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/inbox/api/agents/{agent_id_param}/sandbox/invite")
async def inbox_sandbox_invite(
    agent_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Envía hello_world template al número indicado para iniciar conversación de prueba."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import get_config_value
    import httpx as _httpx

    body = await request.json()
    to_phone = re.sub(r"[^\d]", "", body.get("to", ""))
    if not to_phone:
        return JSONResponse(status_code=400, content={"error": "Número de teléfono requerido"})

    sandbox_phone_id = (
        os.getenv("META_SANDBOX_PHONE_NUMBER_ID")
        or await get_config_value("SANDBOX_PHONE_NUMBER_ID", agent_id_param)
        or ""
    )
    access_token = (
        await get_config_value("META_ACCESS_TOKEN", agent_id_param)
        or os.getenv("META_ACCESS_TOKEN", "")
    )
    if not sandbox_phone_id or not access_token:
        return JSONResponse(status_code=400, content={"error": "Sandbox no configurado (falta phone_number_id o access_token)"})

    url = f"https://graph.facebook.com/v21.0/{sandbox_phone_id}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": to_phone,
        "type": "template",
        "template": {"name": "hello_world", "language": {"code": "en_US"}},
    }
    try:
        async with _httpx.AsyncClient(timeout=15) as client:
            r = await client.post(url, json=payload, headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            })
        if r.status_code == 200:
            logger.info(f"[sandbox] Invitación enviada a {to_phone} desde sandbox {sandbox_phone_id}")
            return JSONResponse(content={"ok": True})
        else:
            err = r.json().get("error", {}).get("message", r.text)
            logger.warning(f"[sandbox] Error Meta API al invitar {to_phone}: {err}")
            return JSONResponse(status_code=400, content={"error": err})
    except Exception as e:
        logger.error(f"[sandbox] Excepción al enviar invitación: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/inbox/api/agents/{agent_id_param}/stats")
async def inbox_stats_agente(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Estadísticas básicas de un agente: conversaciones, mensajes."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_metricas_internas
    try:
        data = await obtener_metricas_internas(dias=30, agent_id=agent_id_param)
        return JSONResponse(content=data)
    except Exception as e:
        logger.error(f"[agents] Error obteniendo stats del agente {agent_id_param}: {e}")
        return JSONResponse(content={"error": str(e)})


# ── Panel principal ────────────────────────────────────────────────────────

@app.get("/inbox", response_class=HTMLResponse)
async def inbox_panel(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    # Resolver token efectivo: voco_session > inbox_session > query param
    effective_token = voco_session or inbox_session or token

    # Intentar autenticación con nuevo sistema de sesiones primero
    current_user = await _obtener_sesion_usuario(effective_token)

    if not current_user:
        return RedirectResponse("/inbox/login", status_code=302)

    from agent.memory import obtener_todos_agentes, obtener_agentes_de_usuario
    if current_user.get("rol") == "admin":
        agentes = await obtener_todos_agentes()
    else:
        agentes = await obtener_agentes_de_usuario(current_user["id"])

    response = HTMLResponse(content=obtener_global_html(agentes, user=current_user))

    # Si llegó por query param con ADMIN_TOKEN legacy, persistir en cookie legacy
    if token and _verificar_admin(token):
        response.set_cookie(
            INBOX_COOKIE, ADMIN_TOKEN,
            httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax"
        )
    return response


# ── Admin: gestión de usuarios (Sprint 1) ───────────────────────────────────

@app.get("/inbox/api/admin/users")
async def admin_listar_usuarios(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista todos los usuarios registrados. Solo accesible por administradores."""
    effective_token = voco_session or inbox_session or token
    current_user = await _obtener_sesion_usuario(effective_token)
    if not current_user or current_user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Acceso restringido a administradores")
    usuarios = await listar_usuarios()
    return JSONResponse(content={"users": usuarios})


@app.post("/inbox/api/admin/users/{user_id_param}")
async def admin_actualizar_usuario(
    user_id_param: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Actualiza datos de un usuario (plan, is_active, rol). Solo para administradores."""
    effective_token = voco_session or inbox_session or token
    current_user = await _obtener_sesion_usuario(effective_token)
    if not current_user or current_user.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Acceso restringido a administradores")
    body = await request.json()
    # Solo campos seguros
    campos_seguros = {k: v for k, v in body.items() if k in ("plan", "is_active", "rol", "nombre")}
    ok = await actualizar_usuario(user_id_param, **campos_seguros)
    if not ok:
        return JSONResponse(status_code=404, content={"error": "Usuario no encontrado"})
    return JSONResponse(content={"ok": True})


# ── API endpoints ──────────────────────────────────────────────────────────

@app.get("/inbox/api/conversaciones")
async def inbox_conversaciones(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    convs = await obtener_todas_conversaciones(agent_id=agent_id)
    return JSONResponse(content=convs)


@app.get("/inbox/api/mensajes/{telefono}")
async def inbox_mensajes(
    telefono: str,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    mensajes = await obtener_historial_con_timestamps(telefono, 150, agent_id=agent_id)
    modo = await get_modo_humano(telefono, agent_id=agent_id)
    return JSONResponse(content={"mensajes": mensajes, "modo_humano": modo})


@app.post("/inbox/api/responder")
async def inbox_responder(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    telefono  = (body.get("telefono") or "").strip()
    mensaje   = (body.get("mensaje") or "").strip()
    agent_id  = int(body.get("agent_id") or 1)
    if not telefono or not mensaje:
        return JSONResponse(status_code=400, content={"error": "Faltan datos"})
    try:
        # Usar el proveedor del agente correcto
        from agent.memory import obtener_agente
        _prov = await _get_proveedor_panel(agent_id)
        # DIAGNÓSTICO: loguear qué credenciales se están usando (sin exponer
        # token completo). Útil para detectar config_value en BD que difiere
        # de env vars de Railway tras cambios recientes.
        _pn = (_prov.phone_number_id or "")
        _tk = (_prov.access_token or "")
        logger.info(
            f"[responder] agent_id={agent_id} → phone_number_id={_pn} "
            f"(len={len(_pn)}) · access_token={_tk[:12]}…{_tk[-4:] if len(_tk)>4 else ''} "
            f"(len={len(_tk)})"
        )
        ok = await _prov.enviar_mensaje(telefono, mensaje)
        if not ok:
            logger.error(f"[responder] enviar_mensaje devolvió False para {telefono} (agent_id={agent_id})")
            return JSONResponse(status_code=502, content={
                "ok": False,
                "error": f"WhatsApp rechazó el envío.\n\n"
                         f"Phone Number ID usado: {_pn}\n"
                         f"Token (primeros chars): {_tk[:8]}…\n\n"
                         f"Causas típicas: token expirado, Phone Number ID incorrecto, "
                         f"ventana de 24h cerrada, o scope insuficiente. Logs de Railway "
                         f"tienen la respuesta exacta de Meta."
            })
        await _guardar_con_wamid(_prov, telefono, mensaje, agent_id=agent_id)
        await registrar_mensaje_asistente(telefono)
        logger.info(f"Mensaje manual enviado a {telefono} desde inbox (agent_id={agent_id})")
        return JSONResponse(content={"ok": True})
    except Exception as e:
        logger.error(f"Error enviando desde inbox a {telefono}: {e}")
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


# ══════════════════════════════════════════════════════════════════════════════
# SPRINT 4 — Mensajes multimedia desde el inbox
# ══════════════════════════════════════════════════════════════════════════════

# Límites Meta WhatsApp Cloud API
_MEDIA_LIMITS = {
    "image":    5  * 1024 * 1024,   # 5 MB
    "video":    16 * 1024 * 1024,   # 16 MB
    "audio":    16 * 1024 * 1024,   # 16 MB
    "document": 100 * 1024 * 1024,  # 100 MB
}

_MEDIA_MIME_TIPO = {
    "image/jpeg":  "image",   "image/png":  "image",   "image/webp": "image",
    "video/mp4":   "video",   "video/3gpp": "video",
    "audio/mpeg":  "audio",   "audio/ogg":  "audio",   "audio/mp4":  "audio",   "audio/amr": "audio",
    "application/pdf": "document",
}


def _detectar_tipo_media(mime_type: str) -> str:
    """Detecta el tipo de media según el MIME type. Default = document."""
    mime = (mime_type or "").lower().split(";")[0].strip()
    if mime in _MEDIA_MIME_TIPO:
        return _MEDIA_MIME_TIPO[mime]
    if mime.startswith("image/"):    return "image"
    if mime.startswith("video/"):    return "video"
    if mime.startswith("audio/"):    return "audio"
    return "document"


@app.post("/inbox/api/responder/media")
async def inbox_responder_media(
    file: UploadFile = File(...),
    telefono: str = Form(...),
    caption: str = Form(""),
    agent_id: int = Form(1),
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Envía un archivo multimedia (imagen/video/documento/audio) por WhatsApp.
    El frontend sube el archivo aquí, lo subimos a Meta y enviamos el mensaje."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    telefono = telefono.strip()
    if not telefono:
        return JSONResponse(content={"ok": False, "error": "Teléfono requerido"}, status_code=400)

    contenido = await file.read()
    if not contenido:
        return JSONResponse(content={"ok": False, "error": "Archivo vacío"}, status_code=400)

    mime_type = file.content_type or "application/octet-stream"
    tipo      = _detectar_tipo_media(mime_type)

    limite = _MEDIA_LIMITS.get(tipo, 5 * 1024 * 1024)
    if len(contenido) > limite:
        mb = round(limite / 1024 / 1024)
        return JSONResponse(content={
            "ok": False,
            "error": f"Archivo demasiado grande. Máximo {mb} MB para {tipo}"
        }, status_code=400)

    try:
        _prov = await _get_proveedor_panel(agent_id)

        # 1. Subir el archivo a Meta
        media_id = await _prov.subir_media(contenido, file.filename or "archivo", mime_type)
        if not media_id:
            return JSONResponse(content={
                "ok": False, "error": "No se pudo subir el archivo a Meta"
            }, status_code=500)

        # 2. Enviar el mensaje según el tipo
        cap = caption.strip()
        ok = False
        if tipo == "image":
            ok = await _prov.enviar_imagen(telefono, media_id=media_id, caption=cap)
        elif tipo == "video":
            ok = await _prov.enviar_video(telefono, media_id=media_id, caption=cap)
        elif tipo == "audio":
            ok = await _prov.enviar_audio(telefono, media_id=media_id)
        else:  # document
            ok = await _prov.enviar_documento(
                telefono, media_id=media_id, caption=cap, filename=file.filename or ""
            )

        if not ok:
            return JSONResponse(content={
                "ok": False, "error": "Meta rechazó el envío del mensaje"
            }, status_code=500)

        # 3. Guardar marcador en historial — wamid ya disponible (#79, se envió arriba)
        marcador = json.dumps({
            "tipo":      tipo,
            "media_id":  media_id,
            "mime_type": mime_type,
            "filename":  file.filename or "",
            "caption":   cap,
        }, ensure_ascii=False)
        await _guardar_con_wamid(_prov, telefono, f"__MEDIA__:{marcador}", agent_id=agent_id)
        await registrar_mensaje_asistente(telefono)
        logger.info(f"Media {tipo} enviada a {telefono} desde inbox")
        return JSONResponse(content={
            "ok": True, "tipo": tipo, "media_id": media_id, "filename": file.filename or ""
        })
    except Exception as e:
        logger.error(f"Error enviando media a {telefono}: {e}", exc_info=True)
        return JSONResponse(content={"ok": False, "error": str(e)[:300]}, status_code=500)


@app.get("/inbox/api/media/{media_id}")
async def inbox_obtener_media(
    media_id: str,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Proxy que descarga el media de Meta y lo sirve al frontend.
    Necesario porque las URLs scontent.whatsapp.net requieren auth header."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from fastapi.responses import Response
    try:
        from agent.memory import obtener_agente
        _agente = await obtener_agente(agent_id) or {"id": 1}
        _prov = await _get_meta_para_agente(_agente)
        result = await _prov.descargar_media(media_id)
        if not result:
            raise HTTPException(status_code=404, detail="Media no encontrada o expirada")
        contenido, mime = result
        return Response(content=contenido, media_type=mime,
                        headers={"Cache-Control": "private, max-age=3600"})
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo media {media_id}: {e}")
        raise HTTPException(status_code=500, detail=str(e)[:200])


@app.post("/inbox/api/responder/ubicacion")
async def inbox_responder_ubicacion(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Envía una ubicación geográfica al cliente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    telefono  = (body.get("telefono") or "").strip()
    latitud   = body.get("latitud")
    longitud  = body.get("longitud")
    nombre    = (body.get("nombre")    or "").strip()
    direccion = (body.get("direccion") or "").strip()
    agent_id  = int(body.get("agent_id") or 1)
    if not telefono or latitud is None or longitud is None:
        return JSONResponse(content={"ok": False, "error": "Faltan datos"}, status_code=400)
    try:
        _prov = await _get_proveedor_panel(agent_id)
        ok = await _prov.enviar_ubicacion(telefono, float(latitud), float(longitud), nombre, direccion)
        if not ok:
            return JSONResponse(content={"ok": False, "error": "Meta rechazó la ubicación"}, status_code=500)
        marcador = json.dumps({
            "tipo":      "location",
            "latitud":   latitud,
            "longitud":  longitud,
            "nombre":    nombre,
            "direccion": direccion,
        }, ensure_ascii=False)
        await _guardar_con_wamid(_prov, telefono, f"__MEDIA__:{marcador}", agent_id=agent_id)
        await registrar_mensaje_asistente(telefono)
        return JSONResponse(content={"ok": True})
    except Exception as e:
        logger.error(f"Error enviando ubicación: {e}")
        return JSONResponse(content={"ok": False, "error": str(e)[:300]}, status_code=500)


@app.post("/inbox/api/responder/producto")
async def inbox_responder_producto(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Envía un producto del catálogo de Shopify al cliente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    telefono     = (body.get("telefono") or "").strip()
    retailer_id  = (body.get("retailer_id") or "").strip()
    cuerpo       = (body.get("cuerpo") or "").strip()
    agent_id     = int(body.get("agent_id") or 1)
    if not telefono or not retailer_id:
        return JSONResponse(content={"ok": False, "error": "Faltan datos"}, status_code=400)
    try:
        _prov = await _get_proveedor_panel(agent_id)
        resultado = await _prov.enviar_producto(telefono, retailer_id, cuerpo)
        if not resultado.get("ok"):
            err_msg = resultado.get("error", "Meta rechazó el producto")
            err_code = resultado.get("code", 0)
            # Mensajes más amigables según el código de Meta
            if "Object with ID" in err_msg or err_code == 100:
                err_msg = f"El SKU '{retailer_id}' no existe en tu catálogo de Meta (CATALOG_ID: {resultado.get('catalog_id','')}). Verifica que el producto esté sincronizado desde Shopify a Facebook Catalog."
            elif "catalog" in err_msg.lower():
                err_msg = f"Problema con el catálogo de Meta: {err_msg}"
            return JSONResponse(content={
                "ok": False, "error": err_msg, "detail": resultado
            }, status_code=500)
        marcador = json.dumps({
            "tipo":        "product",
            "retailer_id": retailer_id,
            "cuerpo":      cuerpo,
        }, ensure_ascii=False)
        await _guardar_con_wamid(_prov, telefono, f"__MEDIA__:{marcador}", agent_id=agent_id)
        await registrar_mensaje_asistente(telefono)
        return JSONResponse(content={"ok": True})
    except Exception as e:
        logger.error(f"Error enviando producto: {e}")
        return JSONResponse(content={"ok": False, "error": str(e)[:300]}, status_code=500)


@app.get("/inbox/api/catalogo/buscar")
async def inbox_buscar_catalogo(
    q: str = "",
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Busca productos en el catálogo de Shopify del agente por nombre (para enviar desde el inbox)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        await obtener_catalogo_shopify(agent_id)
        catalogo = obtener_catalogo_json(agent_id) or []
        from agent.tools import _sku_map as _sku_map_all, _normalizar
        _sku_map_ag = _sku_map_all.get(agent_id, {})
        rev_sku = {}
        for rid, info in _sku_map_ag.items():
            clave = (_normalizar(info.get("producto", "")), _normalizar(info.get("presentacion", "")))
            existente = rev_sku.get(clave, "")
            # Un rid es "variant_id" si es numérico y largo (>=10 dígitos)
            es_variant_id        = rid.isdigit() and len(rid) >= 10
            existente_es_variant = existente.isdigit() and len(existente) >= 10
            # Reemplazar si: no había, o el actual es variant_id y el existente no
            if not existente or (es_variant_id and not existente_es_variant):
                rev_sku[clave] = rid

        # Construir mapa de SKU "humano" para mostrar en la UI (cuando exista)
        sku_humano = {}
        for rid, info in _sku_map_ag.items():
            clave = (_normalizar(info.get("producto", "")), _normalizar(info.get("presentacion", "")))
            # SKU humano = el corto, no-numérico o numérico corto (<10 dígitos)
            if not (rid.isdigit() and len(rid) >= 10):
                if clave not in sku_humano:
                    sku_humano[clave] = rid

        q_norm = q.strip().lower()
        resultados = []
        for item in catalogo:
            titulo       = (item.get("producto") or "").strip()
            presentacion = (item.get("presentacion") or "").strip()
            if q_norm and q_norm not in (titulo + " " + presentacion).lower():
                continue
            clave = (_normalizar(titulo), _normalizar(presentacion))
            retailer_id = rev_sku.get(clave, "")
            sku_legible = sku_humano.get(clave, "")
            resultados.append({
                "retailer_id": retailer_id,
                "sku_legible": sku_legible or retailer_id,  # para mostrar al usuario
                "title":       titulo,
                "variant":     presentacion,
                "price":       item.get("precio", 0),
                "image":       item.get("imagen", ""),
                "categoria":   item.get("categoria", ""),
            })
            if len(resultados) >= 50:
                break
        return JSONResponse(content={"productos": resultados})
    except Exception as e:
        logger.error(f"Error buscando catálogo: {e}", exc_info=True)
        return JSONResponse(content={"productos": [], "error": str(e)[:200]})


# ── Carrito omnichannel — un agente humano arma/edita el carrito de un ──────
# cliente desde el panel (clientes con dificultad para usar el catálogo
# nativo de WhatsApp). Reusa el mismo carrito_activo que ya usa Andrea, así
# que lo que arma el humano se ve y se puede confirmar igual que un carrito
# armado por el bot.

@app.get("/inbox/api/carrito/{telefono}")
async def inbox_obtener_carrito(
    telefono: str,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    items = await obtener_carrito_activo(telefono, agent_id=agent_id)
    total = sum(int(it.get("subtotal", 0)) for it in items)
    return JSONResponse(content={"items": items, "total": total})


@app.post("/inbox/api/carrito/{telefono}")
async def inbox_guardar_carrito(
    telefono: str,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Reemplaza el carrito activo del cliente con la lista de items recibida
    desde el panel. No se setea retailer_id aunque el producto venga del
    catálogo nativo — este carrito lo arma un humano, no es un checkout
    nativo de WhatsApp, y el prompt de Andrea usa retailer_id para decidir
    si está "en medio de un checkout nativo" (ver brain.py)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    items_raw = body.get("items") or []
    if not isinstance(items_raw, list):
        return JSONResponse(status_code=400, content={"error": "items debe ser una lista"})

    items = []
    for it in items_raw:
        try:
            cantidad = int(it.get("cantidad", 1))
            precio_unitario = int(it.get("precio_unitario", 0))
        except (TypeError, ValueError):
            return JSONResponse(status_code=400, content={"error": "cantidad o precio_unitario inválido"})
        if cantidad <= 0 or not (it.get("producto") or "").strip():
            continue
        items.append({
            "retailer_id":     "",
            "quantity":        cantidad,
            "cantidad":        cantidad,
            "producto":        (it.get("producto") or "").strip(),
            "presentacion":    (it.get("presentacion") or "").strip(),
            "precio_unitario": precio_unitario,
            "subtotal":        precio_unitario * cantidad,
        })

    if items:
        await guardar_carrito_activo(telefono, items, agent_id=agent_id)
    else:
        await limpiar_carrito_activo(telefono, agent_id=agent_id)

    logger.info(f"[carrito-manual] Carrito de {telefono} actualizado desde inbox (agent_id={agent_id}, {len(items)} items)")

    # Avisar al cliente — mismo resumen + botón "Confirmar pedido ✅" que
    # recibiría si hubiera armado el carrito él mismo (ver _enviar_resumen_carrito,
    # llamada también desde act_ver_carrito y [[MOSTRAR_CARRITO]]). Si Meta
    # rechaza el envío (ej. ventana de 24h cerrada), el mensaje queda marcado
    # como "failed" en el chat con el motivo real — no rompemos el guardado
    # del carrito por un fallo de notificación.
    notificado = False
    try:
        notificado = await _enviar_resumen_carrito(telefono, agent_id=agent_id)
    except Exception as e:
        logger.error(f"[carrito-manual] Error notificando a {telefono}: {e}")

    total = sum(it["subtotal"] for it in items)
    return JSONResponse(content={"ok": True, "items": items, "total": total, "cliente_notificado": notificado})


@app.post("/inbox/api/modo/{telefono}")
async def inbox_modo(
    telefono: str,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    activo = bool(body.get("activo", False))
    await set_modo_humano(telefono, activo)
    logger.info(f"Modo cambiado a '{'humano' if activo else 'Andrea'}' para {telefono}")
    return JSONResponse(content={"ok": True, "modo_humano": activo})


# ══════════════════════════════════════════════════════════════════════════════
# SPRINT 1 — API de Tickets de Escalación
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/inbox/api/tickets")
async def api_listar_tickets(
    estado: str = "",
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista tickets filtrados por estado (sin_asignar|activo|pendiente|resuelto|todos)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    tickets = await obtener_tickets(agent_id, estado=estado or None)
    return JSONResponse(content={"tickets": tickets})


@app.get("/inbox/api/tickets/counts")
async def api_contar_tickets(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Conteos por estado para los badges del panel."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    counts = await contar_tickets(agent_id)
    return JSONResponse(content=counts)


@app.post("/inbox/api/tickets/{ticket_id}/tomar")
async def api_tomar_ticket(
    ticket_id: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """El agente humano toma un ticket (sin_asignar o pendiente → activo).
    Si quien toma es el admin SaaS (no un UsuarioInterno explícito), se
    crea/usa automáticamente un usuario interno con rol 'admin' para que
    el ticket quede registrado con nombre + rol."""
    sesion = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not sesion:
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    agente_humano_id = int(body.get("agente_humano_id") or 0)

    # Si el frontend no pasó un usuario interno explícito (= 0), significa que
    # está actuando el admin SaaS o un usuario sin perfil de equipo interno.
    # Buscamos/creamos un usuario interno tipo admin para él.
    if not agente_humano_id:
        # agent_id del ticket (lo necesitamos para crear el admin en el negocio correcto)
        from agent.memory import Ticket as _T, async_session as _AS
        from sqlalchemy import select as _sel
        async with _AS() as _s:
            _r = await _s.execute(_sel(_T).where(_T.id == ticket_id))
            _tk = _r.scalar_one_or_none()
        if _tk:
            try:
                admin_ui = await obtener_o_crear_admin_interno(
                    agent_id=_tk.agent_id,
                    email=sesion.get("email", "admin@voco.local"),
                    nombre=sesion.get("nombre", "") or sesion.get("email", "Administrador"),
                )
                agente_humano_id = admin_ui["id"]
            except Exception as e:
                logger.warning(f"No pude obtener/crear admin interno: {e}")

    ticket = await tomar_ticket(ticket_id, agente_humano_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    # Activar modo humano para esta conversación. Sin esto Andrea sigue
    # respondiendo al cliente mientras el humano lo atiende — el cliente
    # ve respuestas duplicadas/contradictorias. La creación via escalación
    # ya lo activa, pero si el humano toma un ticket que no pasó por esa
    # ruta (ej: reabierto manualmente, o quedó con modo humano apagado por
    # algún cambio anterior), aquí lo forzamos.
    try:
        await set_modo_humano(ticket["telefono_cliente"], True, agent_id=ticket["agent_id"])
    except Exception as e:
        logger.warning(f"No pude activar modo humano al tomar ticket {ticket_id}: {e}")
    await registrar_evento_ticket(ticket_id, "tomado", ticket.get("agente_nombre","Agente"),
                                  "Tomó la conversación")
    _push_evento_ticket(ticket["agent_id"], "ticket_tomado", ticket)
    return JSONResponse(content={"ok": True, "ticket": ticket})


@app.post("/inbox/api/tickets/{ticket_id}/pendiente")
async def api_ticket_pendiente(
    ticket_id: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Marca el ticket como pendiente (agente necesita más info)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    ticket = await marcar_ticket_pendiente(ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    await registrar_evento_ticket(ticket_id, "pendiente",
                                  ticket.get("agente_nombre","Agente"), "Marcó como pendiente")
    _push_evento_ticket(ticket["agent_id"], "ticket_pendiente", ticket)
    return JSONResponse(content={"ok": True, "ticket": ticket})


@app.post("/inbox/api/tickets/{ticket_id}/resolver")
async def api_resolver_ticket(
    ticket_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Resuelve el ticket y reactiva el bot automáticamente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    ticket = await resolver_ticket(ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    # Reactivar bot para este cliente
    try:
        await set_modo_humano(ticket["telefono_cliente"], False, agent_id=agent_id)
        logger.info(f"Bot reactivado para {ticket['telefono_cliente']} al resolver ticket {ticket_id}")
    except Exception as e:
        logger.error(f"Error reactivando bot: {e}")
    await registrar_evento_ticket(ticket_id, "resuelto",
                                  ticket.get("agente_nombre","Agente"),
                                  "Resolvió el ticket — bot reactivado")
    _push_evento_ticket(ticket["agent_id"], "ticket_resuelto", ticket)
    return JSONResponse(content={"ok": True, "ticket": ticket})


@app.get("/inbox/api/tickets/{ticket_id}/eventos")
async def api_eventos_ticket(
    ticket_id: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Auditoría: historial de eventos de un ticket."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    eventos = await obtener_eventos_ticket(ticket_id)
    return JSONResponse(content={"eventos": eventos})


@app.get("/inbox/api/sistema/capi-status")
async def api_capi_status(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Diagnostica el estado del Conversions API (CAPI) token de Meta.
    Verifica si el token está configurado, si es válido contra la API de Meta,
    y a qué pixel/dataset tiene acceso."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    pixel_id   = await get_config_value("META_PIXEL_ID", agent_id=agent_id) or ""
    capi_token = await get_config_value("META_CAPI_TOKEN", agent_id=agent_id) or ""
    test_code  = await get_config_value("META_CAPI_TEST_CODE", agent_id=agent_id) or ""
    if agent_id == 1:  # solo para el agente primario, caer a env vars como fallback
        pixel_id   = pixel_id or os.getenv("META_PIXEL_ID", "").strip()
        capi_token = capi_token or os.getenv("META_CAPI_TOKEN", "").strip()
        test_code  = test_code or os.getenv("META_CAPI_TEST_CODE", "").strip()
    api_ver     = "v21.0"

    resultado = {
        "pixel_id_configurado":   bool(pixel_id),
        "capi_token_configurado": bool(capi_token),
        "modo_prueba_activo":     bool(test_code),
        "pixel_id_partial":       f"{pixel_id[:6]}...{pixel_id[-4:]}" if len(pixel_id) > 10 else pixel_id,
        "token_partial":          f"{capi_token[:8]}...{capi_token[-4:]}" if len(capi_token) > 16 else "(vacío)",
        "estado":                 "desconocido",
        "mensaje":                "",
        "pixel_info":             {},
    }

    if not pixel_id:
        resultado["estado"] = "no_configurado"
        resultado["mensaje"] = "Falta META_PIXEL_ID en variables de entorno"
        return JSONResponse(content=resultado)
    if not capi_token:
        resultado["estado"] = "no_configurado"
        resultado["mensaje"] = "Falta META_CAPI_TOKEN en variables de entorno"
        return JSONResponse(content=resultado)

    # 1) Validar token contra /me — devuelve la app/usuario al que pertenece
    try:
        async with httpx.AsyncClient(timeout=10) as cli:
            r_me = await cli.get(
                f"https://graph.facebook.com/{api_ver}/me",
                headers={"Authorization": f"Bearer {capi_token}"},
            )
            if r_me.status_code != 200:
                err_data = {}
                try:
                    err_data = r_me.json().get("error", {})
                except Exception:
                    pass
                resultado["estado"] = "token_invalido"
                resultado["mensaje"] = (
                    f"Token rechazado por Meta — code={err_data.get('code', r_me.status_code)} "
                    f"msg={err_data.get('message', r_me.text[:200])}"
                )
                resultado["error_code"] = err_data.get("code")
                if err_data.get("code") == 190:
                    resultado["sugerencia"] = (
                        "Token expirado. Genera uno nuevo en Meta Business Manager → "
                        "Pixel → Configuración → Conversions API → Generate Access Token"
                    )
                return JSONResponse(content=resultado)
            me_data = r_me.json()
            resultado["token_owner"] = {
                "id":   me_data.get("id", ""),
                "name": me_data.get("name", ""),
            }
    except Exception as e:
        resultado["estado"] = "error_red"
        resultado["mensaje"] = f"No pude contactar Meta API: {str(e)[:200]}"
        return JSONResponse(content=resultado)

    # 2) Validar capacidad de enviar eventos al pixel
    # En vez de hacer GET (requiere ads_management), hacemos un POST de test event
    # que es exactamente para lo que sirve el token CAPI. Si Meta acepta el envío,
    # el token funciona para CAPI aunque no tenga permisos de lectura.
    # Usamos test_event_code para no contaminar el pixel real con eventos de prueba.
    try:
        url_eventos = f"https://graph.facebook.com/{api_ver}/{pixel_id}/events"
        # Evento de test mínimo (no se atribuye, solo valida que el token funciona)
        import hashlib
        test_payload = {
            "data": [{
                "event_name":   "VocoDiagnostico",
                "event_time":   int(datetime.utcnow().timestamp()),
                "action_source": "system_generated",
                "user_data":    {
                    "external_id": [hashlib.sha256(b"voco_diagnostic").hexdigest()],
                },
            }],
            "test_event_code": "TEST_VOCO_DIAGNOSTIC",  # marca como test, no afecta métricas
        }
        async with httpx.AsyncClient(timeout=10) as cli:
            r_ev = await cli.post(
                url_eventos,
                params={"access_token": capi_token},
                json=test_payload,
            )
            if r_ev.status_code == 200:
                ev_data = r_ev.json()
                resultado["estado"]   = "ok_activo"
                resultado["mensaje"]  = "✅ Token válido — Meta aceptó evento de prueba"
                resultado["eventos_recibidos"] = ev_data.get("events_received", 0)
                resultado["fbtrace_id"] = ev_data.get("fbtrace_id", "")
            else:
                err_data_p = {}
                try:
                    err_data_p = r_ev.json().get("error", {})
                except Exception:
                    pass
                err_code = err_data_p.get("code", 0)
                err_msg  = err_data_p.get("message", r_ev.text[:200])
                if err_code == 190:
                    resultado["estado"]  = "token_invalido"
                    resultado["mensaje"] = f"Token expirado/inválido: {err_msg}"
                    resultado["sugerencia"] = (
                        "Regenera el token en Events Manager → Equora Pixel → "
                        "Configuración → API de conversiones → Generar token de acceso"
                    )
                elif err_code in (100, 200, 803):
                    resultado["estado"]  = "sin_acceso_pixel"
                    resultado["mensaje"] = (
                        f"Token sin permiso sobre el pixel {pixel_id}: {err_msg}"
                    )
                    resultado["sugerencia"] = (
                        "Verifica que el token fue generado DESDE el panel del pixel "
                        "Equora Pixel en Events Manager (no desde otro pixel). "
                        "Confirma también que META_PIXEL_ID en Railway coincida con el "
                        "pixel para el que generaste el token."
                    )
                else:
                    resultado["estado"]  = "error"
                    resultado["mensaje"] = f"Meta respondió HTTP {r_ev.status_code} code={err_code}: {err_msg}"
                resultado["error_code"] = err_code
    except Exception as e:
        resultado["estado"]  = "error_red"
        resultado["mensaje"] = f"Error enviando evento de test: {str(e)[:200]}"

    return JSONResponse(content=resultado)


@app.get("/inbox/api/sistema/webhook-suscripcion")
async def api_webhook_suscripcion_status(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Verifica si nuestra app está suscrita al webhook de la WABA del agente.
    Sin esta suscripción Meta no envía los mensajes entrantes al backend, los
    clientes escriben y Andrea nunca se entera (token funciona OK pero no
    recibe nada). Pasa después de re-activar números o cambios en la WABA."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    waba_id = await get_config_value("META_WABA_ID", agent_id=agent_id) or (os.getenv("META_WABA_ID", "") if agent_id == 1 else "")
    tok     = await get_config_value("META_ACCESS_TOKEN", agent_id=agent_id) or (os.getenv("META_ACCESS_TOKEN", "") if agent_id == 1 else "")
    if not waba_id or not tok:
        return JSONResponse({"ok": False, "error": "Falta META_WABA_ID o META_ACCESS_TOKEN"})
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(
                f"https://graph.facebook.com/v21.0/{waba_id}/subscribed_apps",
                params={"access_token": tok},
            )
            if r.status_code != 200:
                return JSONResponse({"ok": False, "error": f"Meta API {r.status_code}: {r.text[:200]}"})
            data = r.json() or {}
            apps = data.get("data", []) or []
            return JSONResponse({
                "ok":         True,
                "suscrito":   len(apps) > 0,
                "total_apps": len(apps),
                "apps":       apps,
            })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


@app.post("/inbox/api/sistema/webhook-suscribir")
async def api_webhook_suscribir(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Suscribe la app Voco al webhook de la WABA. Reemplaza el paso manual
    que antes había que hacer en developers.facebook.com → WhatsApp → Configuration."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    waba_id = await get_config_value("META_WABA_ID", agent_id=agent_id) or (os.getenv("META_WABA_ID", "") if agent_id == 1 else "")
    tok     = await get_config_value("META_ACCESS_TOKEN", agent_id=agent_id) or (os.getenv("META_ACCESS_TOKEN", "") if agent_id == 1 else "")
    if not waba_id or not tok:
        return JSONResponse({"ok": False, "error": "Falta META_WABA_ID o META_ACCESS_TOKEN"})
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.post(
                f"https://graph.facebook.com/v21.0/{waba_id}/subscribed_apps",
                params={"access_token": tok},
            )
            ok = r.status_code == 200
            return JSONResponse({"ok": ok, "status": r.status_code, "detalle": r.text[:300]})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


@app.get("/inbox/api/sistema/fallas-catalogo")
async def api_fallas_catalogo(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve las últimas fallas del catálogo nativo de WhatsApp.
    Útil para diagnosticar problemas de sincronización con Meta."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    # Filtrar las últimas 24 horas
    hace_un_dia = datetime.utcnow() - timedelta(hours=24)
    recientes = [
        f for f in _catalogo_fallas
        if datetime.fromisoformat(f["at"]) > hace_un_dia
    ]
    # Contar por tipo
    por_tipo: dict[str, int] = {}
    for f in recientes:
        por_tipo[f["tipo"]] = por_tipo.get(f["tipo"], 0) + 1
    # Flag de "configurado" por agente — sin esto el panel mostraba "Catálogo funcionando"
    # para agentes nuevos que NO tienen META_CATALOG_ID, usando el catalog de otro agente.
    meta_cat_id = await get_config_value("META_CATALOG_ID", agent_id=agent_id) or ""
    if not meta_cat_id and agent_id == 1:
        meta_cat_id = os.getenv("META_CATALOG_ID", "")
    configurado = bool(meta_cat_id.strip())
    return JSONResponse(content={
        "configurado": configurado,
        "total_24h":  len(recientes),
        "por_tipo":   por_tipo,
        "ultimas":    list(recientes)[-20:],   # las 20 más recientes
        "cooldown_alerta_seg": _CATALOGO_ALERTA_COOLDOWN_SEG,
    })


@app.get("/inbox/api/equipo/stats")
async def api_stats_equipo(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Dashboard supervisor: métricas por agente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    stats = await obtener_stats_equipo(agent_id)
    return JSONResponse(content={"stats": stats})


@app.post("/inbox/api/config/auto-asignar")
async def api_toggle_autoasignar(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Activa/desactiva la asignación automática round-robin."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    activo = "1" if body.get("activo") else "0"
    await set_config_value("AUTO_ASIGNAR", activo, agent_id)
    return JSONResponse(content={"ok": True, "auto_asignar": activo == "1"})


@app.get("/inbox/api/tickets/{ticket_id}/historial")
async def api_historial_ticket(
    ticket_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Historial completo de mensajes del ticket (Andrea + humano) para el panel."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    ticket = await obtener_ticket_activo_por_telefono(agent_id, "")
    # Buscar por ticket_id directamente
    from agent.memory import Ticket as _TicketModel, async_session as _as
    from sqlalchemy import select as _sel
    async with _as() as sess:
        r = await sess.execute(_sel(_TicketModel).where(_TicketModel.id == ticket_id))
        t = r.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    mensajes = await obtener_historial_con_timestamps(t.telefono_cliente, 150, agent_id=t.agent_id)
    return JSONResponse(content={"mensajes": mensajes, "telefono": t.telefono_cliente})


# ══════════════════════════════════════════════════════════════════════════════
# Item #3 del backlog — Pipeline. Las tablas (Pipeline/Deal/DealActivity)
# existían desde Sprint A pero sin ningún endpoint que las usara — el módulo
# vivía como placeholder "en construcción" en el panel. Estos son los
# primeros endpoints CRUD.
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/inbox/api/pipeline")
async def api_obtener_pipeline(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Pipeline activo del agente (se crea con stages default si no existe)
    + todos sus deals. Pensado para cargar el kanban en una sola llamada."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    pipeline = await obtener_pipeline_activo(agent_id)
    deals = await listar_deals(agent_id, pipeline_id=pipeline["id"])
    return JSONResponse(content={"pipeline": pipeline, "deals": deals})


@app.post("/inbox/api/pipeline/{pipeline_id}/stages")
async def api_actualizar_stages_pipeline(
    pipeline_id: int,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    stages = body.get("stages") or []
    if not isinstance(stages, list):
        return JSONResponse(status_code=400, content={"error": "stages debe ser una lista"})
    pipeline = await actualizar_stages_pipeline(pipeline_id, agent_id, [str(s) for s in stages])
    if not pipeline:
        return JSONResponse(status_code=400, content={"error": "Pipeline no encontrado o lista de stages vacía"})
    return JSONResponse(content={"ok": True, "pipeline": pipeline})


@app.post("/inbox/api/deals")
async def api_crear_deal(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    telefono = (body.get("cliente_telefono") or "").strip()
    if not telefono:
        return JSONResponse(status_code=400, content={"error": "cliente_telefono es requerido"})
    pipeline = await obtener_pipeline_activo(agent_id)
    stages = pipeline["stages"] or ["Nuevo"]
    stage_inicial = body.get("stage") or stages[0]
    deal = await crear_deal(
        agent_id=agent_id,
        pipeline_id=pipeline["id"],
        cliente_telefono=telefono,
        cliente_nombre=(body.get("cliente_nombre") or "").strip(),
        titulo=(body.get("titulo") or "").strip(),
        valor_cop=body.get("valor_cop") or 0,
        source=(body.get("source") or "manual").strip(),
        stage=stage_inicial,
        autor_nombre=user.get("nombre") or user.get("email") or "Sistema",
    )
    return JSONResponse(content={"ok": True, "deal": deal})


@app.patch("/inbox/api/deals/{deal_id}")
async def api_actualizar_deal(
    deal_id: int,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    campos = {k: body[k] for k in
              ("titulo", "stage", "valor_cop", "source", "owner_id", "notas", "cliente_nombre", "cliente_email")
              if k in body}
    deal = await actualizar_deal(deal_id, agent_id, autor_nombre=user.get("nombre") or user.get("email") or "Sistema", **campos)
    if not deal:
        return JSONResponse(status_code=404, content={"error": "Deal no encontrado o sin cambios válidos"})
    return JSONResponse(content={"ok": True, "deal": deal})


@app.delete("/inbox/api/deals/{deal_id}")
async def api_eliminar_deal(
    deal_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    ok = await eliminar_deal(deal_id, agent_id)
    if not ok:
        return JSONResponse(status_code=404, content={"error": "Deal no encontrado"})
    return JSONResponse(content={"ok": True})


@app.get("/inbox/api/deals/{deal_id}/actividades")
async def api_listar_actividades_deal(
    deal_id: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    actividades = await listar_actividades_deal(deal_id)
    return JSONResponse(content={"actividades": actividades})


@app.post("/inbox/api/deals/{deal_id}/actividades")
async def api_crear_actividad_deal(
    deal_id: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    contenido = (body.get("contenido") or "").strip()
    if not contenido:
        return JSONResponse(status_code=400, content={"error": "contenido es requerido"})
    actividad = await agregar_actividad_deal(
        deal_id, tipo="note", contenido=contenido,
        autor_nombre=user.get("nombre") or user.get("email") or "Sistema",
    )
    return JSONResponse(content={"ok": True, "actividad": actividad})


# ══════════════════════════════════════════════════════════════════════════════
# Pipeline Fase 2 — Integraciones externas (Calendly, HubSpot, ...). CRUD
# genérico sobre IntegrationConfig — la tabla existía desde Sprint A sin
# ningún endpoint que la usara. Un (agent_id, tipo) por integración.
# ══════════════════════════════════════════════════════════════════════════════

TIPOS_INTEGRACION_VALIDOS = {"calendly", "hubspot", "sendgrid", "pipedrive"}


def _integracion_para_frontend(config: dict) -> dict:
    """Nunca devolvemos el token real al frontend — mismo patrón que ya usa
    Configuración para META_ACCESS_TOKEN/SHOPIFY_ADMIN_TOKEN (_ofuscar_secreto)."""
    config = dict(config)
    config["tiene_token"] = bool(config.get("api_token"))
    config["api_token"] = _ofuscar_secreto(config["api_token"]) if config.get("api_token") else ""
    return config


@app.get("/inbox/api/agents/{agent_id_param}/integrations")
async def api_listar_integraciones(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    configs = await listar_integration_configs(agent_id_param)
    return JSONResponse(content={"integraciones": [_integracion_para_frontend(c) for c in configs]})


@app.get("/inbox/api/agents/{agent_id_param}/integrations/hubspot/verify")
async def api_verificar_hubspot(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Llama a HubSpot /account-info/v3/details para confirmar qué portal está conectado."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    resultado = await hubspot_portal_info(agent_id_param)
    return JSONResponse(content=resultado)


@app.get("/inbox/api/agents/{agent_id_param}/integrations/calendly/verify")
async def api_verificar_calendly(
    agent_id_param: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Llama a Calendly /users/me para confirmar qué cuenta está conectada."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    resultado = await calendly_usuario(agent_id_param)
    return JSONResponse(content=resultado)


@app.get("/inbox/api/agents/{agent_id_param}/integrations/{tipo}")
async def api_obtener_integracion(
    agent_id_param: int,
    tipo: str,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    config = await obtener_integration_config(agent_id_param, tipo)
    if not config:
        return JSONResponse(content={
            "tipo": tipo, "configurado": False, "settings": {}, "activo": False,
            "tiene_token": False, "api_token": "",
        })
    config = _integracion_para_frontend(config)
    config["configurado"] = True
    return JSONResponse(content=config)


@app.post("/inbox/api/agents/{agent_id_param}/integrations/{tipo}")
async def api_guardar_integracion(
    agent_id_param: int,
    tipo: str,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    if tipo not in TIPOS_INTEGRACION_VALIDOS:
        return JSONResponse(status_code=400, content={"error": f"tipo inválido: {tipo}"})
    body = await request.json()

    api_token = body.get("api_token")
    if isinstance(api_token, str):
        api_token = api_token.strip()
        # Mismo saneo de caracteres invisibles que ya existe para tokens de
        # Meta/Shopify (bug real reportado al pegar tokens copiados de una web).
        for ch in ("​", "‌", "‍", "\xa0", "﻿", "‪", "‬"):
            api_token = api_token.replace(ch, "")
        if not api_token:
            api_token = None  # vacío = no tocar el token ya guardado, mismo patrón que Configuración
    elif api_token is not None:
        return JSONResponse(status_code=400, content={"error": "api_token debe ser texto"})

    settings = body.get("settings")
    if settings is not None and not isinstance(settings, dict):
        return JSONResponse(status_code=400, content={"error": "settings debe ser un objeto"})

    activo = body.get("activo")
    if activo is not None and not isinstance(activo, bool):
        return JSONResponse(status_code=400, content={"error": "activo debe ser booleano"})

    config = await guardar_integration_config(
        agent_id_param, tipo, api_token=api_token, settings=settings, activo=activo,
    )
    logger.info(f"[integraciones] '{tipo}' actualizado para agent_id={agent_id_param}")
    return JSONResponse(content={"ok": True, "integracion": _integracion_para_frontend(config)})


# ── API de Equipo (Usuarios Internos) ─────────────────────────────────────────

@app.get("/inbox/api/equipo")
async def api_listar_equipo(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista los agentes humanos de soporte del negocio."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    usuarios = await obtener_usuarios_internos(agent_id)
    return JSONResponse(content={"equipo": usuarios})


@app.post("/inbox/api/equipo")
async def api_crear_agente_equipo(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Crea un nuevo agente humano de soporte. Respeta límites por plan."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body = await request.json()
    nombre   = (body.get("nombre") or "").strip()
    email    = (body.get("email") or "").strip().lower()
    password = (body.get("password") or "").strip()
    rol      = body.get("rol", "agente")

    if not nombre or not email or not password:
        return JSONResponse(content={"ok": False, "error": "nombre, email y password son requeridos"})
    if rol not in ("agente", "supervisor", "admin"):
        return JSONResponse(content={"ok": False, "error": "Rol inválido"})

    # Verificar límite de agentes según plan del agente
    from agent.memory import Agent as _AgentModel, async_session as _as
    from sqlalchemy import select as _sel
    async with _as() as sess:
        r = await sess.execute(_sel(_AgentModel).where(_AgentModel.id == agent_id))
        agente_cfg = r.scalar_one_or_none()
    max_permitidos = getattr(agente_cfg, "max_agentes", 2) if agente_cfg else 2
    actuales = await contar_agentes_activos(agent_id)
    if actuales >= max_permitidos:
        return JSONResponse(content={
            "ok": False,
            "error": f"Límite de {max_permitidos} agentes alcanzado. Actualiza tu plan para agregar más."
        })

    try:
        nuevo = await crear_usuario_interno(agent_id, nombre, email, password, rol)
        return JSONResponse(content={"ok": True, "usuario": nuevo})
    except Exception as e:
        return JSONResponse(content={"ok": False, "error": str(e)[:200]})


@app.put("/inbox/api/equipo/{ui_id}")
async def api_actualizar_agente_equipo(
    ui_id: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Actualiza nombre, rol, activo o password de un agente interno."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    # Whitelist de campos editables. #52 agrega notif_escalaciones_wa y telefono_wa.
    campos = {k: v for k, v in body.items() if k in (
        "nombre", "rol", "activo", "password",
        "notif_escalaciones_wa", "telefono_wa",
    )}
    # Normalizar el teléfono a solo dígitos (sin '+' ni espacios) si viene
    if "telefono_wa" in campos:
        digitos = re.sub(r"\D", "", str(campos["telefono_wa"] or ""))
        campos["telefono_wa"] = digitos
    if "notif_escalaciones_wa" in campos:
        campos["notif_escalaciones_wa"] = bool(campos["notif_escalaciones_wa"])
    ok = await actualizar_usuario_interno(ui_id, **campos)
    return JSONResponse(content={"ok": ok})


@app.delete("/inbox/api/equipo/{ui_id}")
async def api_desactivar_agente_equipo(
    ui_id: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Desactiva (soft-delete) un agente interno."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    ok = await actualizar_usuario_interno(ui_id, activo=False)
    return JSONResponse(content={"ok": ok})


@app.post("/inbox/api/equipo/ping")
async def api_ping_agente(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Heartbeat del agente para marcar que está online."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    ui_id = body.get("usuario_interno_id")
    if ui_id:
        await ping_usuario_interno(ui_id)
    return JSONResponse(content={"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# SPRINT 2 — SSE (Server-Sent Events) para notificaciones en tiempo real
# ══════════════════════════════════════════════════════════════════════════════

import asyncio as _asyncio
from fastapi.responses import StreamingResponse

# { agent_id → [Queue, Queue, ...] }  — una queue por pestaña/agente conectado
_sse_queues: dict[int, list[_asyncio.Queue]] = {}


def _push_evento_ticket(agent_id: int, tipo: str, ticket: dict) -> None:
    """Envía un evento SSE a todos los suscriptores del agent_id."""
    payload = json.dumps({"tipo": tipo, "ticket": ticket})
    for q in _sse_queues.get(agent_id, []):
        try:
            q.put_nowait(payload)
        except _asyncio.QueueFull:
            pass


async def _sse_stream(agent_id: int, q: _asyncio.Queue):
    """Generador SSE con keepalive cada 25 segundos."""
    try:
        while True:
            try:
                data = await _asyncio.wait_for(q.get(), timeout=25)
                yield f"data: {data}\n\n"
            except _asyncio.TimeoutError:
                yield ": keepalive\n\n"
    except _asyncio.CancelledError:
        pass
    finally:
        # Limpiar queue al desconectar
        qs = _sse_queues.get(agent_id, [])
        if q in qs:
            qs.remove(q)


@app.get("/inbox/api/eventos")
async def api_eventos_sse(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """SSE stream de eventos de tickets para el panel de escalaciones."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    q: _asyncio.Queue = _asyncio.Queue(maxsize=50)
    _sse_queues.setdefault(agent_id, []).append(q)

    return StreamingResponse(
        _sse_stream(agent_id, q),
        media_type="text/event-stream",
        headers={
            "Cache-Control":               "no-cache",
            "X-Accel-Buffering":           "no",   # Nginx: deshabilitar buffer
            "Access-Control-Allow-Origin": "*",
        },
    )


# ── Transferir ticket entre agentes ───────────────────────────────────────────

@app.post("/inbox/api/tickets/{ticket_id}/transferir")
async def api_transferir_ticket(
    ticket_id: int,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Transfiere un ticket a otro agente humano."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    nuevo_agente_id = body.get("agente_humano_id")
    if not nuevo_agente_id:
        return JSONResponse(content={"ok": False, "error": "agente_humano_id requerido"})
    ticket = await transferir_ticket(ticket_id, nuevo_agente_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    _push_evento_ticket(agent_id, "ticket_transferido", ticket)
    return JSONResponse(content={"ok": True, "ticket": ticket})


# ── Notas internas ────────────────────────────────────────────────────────────

@app.get("/inbox/api/tickets/{ticket_id}/notas")
async def api_listar_notas(
    ticket_id: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    notas = await obtener_notas_ticket(ticket_id)
    return JSONResponse(content={"notas": notas})


@app.post("/inbox/api/tickets/{ticket_id}/notas")
async def api_crear_nota(
    ticket_id: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    contenido        = (body.get("contenido") or "").strip()
    agente_humano_id = body.get("agente_humano_id", 0)
    agente_nombre    = (body.get("agente_nombre") or "Agente").strip()
    if not contenido:
        return JSONResponse(content={"ok": False, "error": "Nota vacía"})
    nota = await crear_nota_interna(ticket_id, agente_humano_id, agente_nombre, contenido)
    return JSONResponse(content={"ok": True, "nota": nota})


# ── Templates rápidos ─────────────────────────────────────────────────────────

@app.get("/inbox/api/templates")
async def api_listar_templates(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    templates = await obtener_templates_rapidos(agent_id)
    return JSONResponse(content={"templates": templates})


@app.post("/inbox/api/templates")
async def api_crear_template(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body     = await request.json()
    titulo   = (body.get("titulo")   or "").strip()
    contenido = (body.get("contenido") or "").strip()
    orden    = int(body.get("orden", 0))
    if not titulo or not contenido:
        return JSONResponse(content={"ok": False, "error": "titulo y contenido requeridos"})
    tpl = await crear_template_rapido(agent_id, titulo, contenido, orden)
    return JSONResponse(content={"ok": True, "template": tpl})


@app.delete("/inbox/api/templates/{tpl_id}")
async def api_eliminar_template(
    tpl_id: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    ok = await eliminar_template_rapido(tpl_id)
    return JSONResponse(content={"ok": ok})


# ── Difusión / Broadcast ───────────────────────────────────────────────────

# Cache en memoria: {scontent_url → whatsapp_media_id}
# Se llena la primera vez que se usa cada plantilla; se pierde al reiniciar (Railway redeploy).
_header_media_cache: dict[str, str] = {}


async def _subir_imagen_whatsapp(image_url: str, access_token: str, phone_number_id: str) -> str | None:
    """
    Descarga la imagen del CDN de Meta y la sube a la WhatsApp Media API
    para obtener un media_id reutilizable en el componente header de templates.

    Returns:
        media_id  si fue exitoso, None si falló (el envío continuará sin header).
    """
    global _header_media_cache
    if image_url in _header_media_cache:
        return _header_media_cache[image_url]

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            # 1. Descargar la imagen (las URLs de scontent.whatsapp.net son públicas con firma)
            r_img = await client.get(image_url)
            if r_img.status_code != 200:
                logger.warning(f"[broadcast] No se pudo descargar imagen header: {r_img.status_code}")
                return None
            image_bytes = r_img.content
            content_type = r_img.headers.get("content-type", "image/jpeg").split(";")[0].strip()
            ext = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp"}.get(content_type, "jpg")

            # 2. Subir a la WhatsApp Media API
            upload_url = f"https://graph.facebook.com/v21.0/{phone_number_id}/media"
            auth_headers = {"Authorization": f"Bearer {access_token}"}
            files = {
                "file": (f"header.{ext}", image_bytes, content_type),
            }
            data = {
                "messaging_product": "whatsapp",
                "type": content_type,
            }
            r_up = await client.post(upload_url, headers=auth_headers, files=files, data=data)
            if r_up.status_code == 200:
                media_id = r_up.json().get("id")
                if media_id:
                    _header_media_cache[image_url] = media_id
                    logger.info(f"[broadcast] Imagen header subida → media_id={media_id}")
                    return media_id
            logger.warning(f"[broadcast] Error subiendo imagen: {r_up.status_code} {r_up.text[:200]}")
    except Exception as e:
        logger.warning(f"[broadcast] Excepción subiendo imagen header: {e}")
    return None

@app.get("/inbox/broadcast/templates/raw")
async def inbox_broadcast_templates_raw(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Debug: devuelve el JSON crudo de Meta para ver cómo están registradas las variables."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    access_token = os.getenv("META_ACCESS_TOKEN", "")
    waba_id      = os.getenv("META_WABA_ID", "")
    api_ver = "v21.0"
    url = (
        f"https://graph.facebook.com/{api_ver}/{waba_id}"
        f"/message_templates?fields=id,name,status,components,language&limit=50"
        f"&access_token={access_token}"
    )
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.get(url)
        return JSONResponse(content=r.json())

@app.get("/inbox/broadcast/templates")
async def inbox_broadcast_templates(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista las plantillas aprobadas en Meta para usar en difusiones."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = os.getenv("META_ACCESS_TOKEN", "")
    waba_id      = os.getenv("META_WABA_ID", "")
    logger.info(f"[broadcast] access_token={'OK' if access_token else 'FALTA'} waba_id={waba_id or 'FALTA'}")

    if not access_token or not waba_id:
        return JSONResponse(content={"templates": [], "error": "META_ACCESS_TOKEN o META_WABA_ID no configurados en Railway"})

    try:
        # Las plantillas pertenecen al WhatsApp Business Account (WABA), no al número
        api_ver  = "v21.0"
        waba_url = (
            f"https://graph.facebook.com/{api_ver}/{waba_id}"
            f"/message_templates?fields=id,name,status,components,language&limit=100"
            f"&access_token={access_token}"
        )
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(waba_url)
            logger.info(f"[broadcast] Meta templates response: {r.status_code} {r.text[:300]}")
            if r.status_code != 200:
                return JSONResponse(content={"templates": [], "error": f"Meta API {r.status_code}: {r.text[:200]}"})

            data = r.json()
            # Meta a veces devuelve {"error": ...} con status 200
            if "error" in data:
                err_msg = data["error"].get("message", str(data["error"]))
                logger.error(f"[broadcast] Meta error en templates: {err_msg}")
                return JSONResponse(content={"templates": [], "error": err_msg})

            templates_raw = data.get("data", [])

            # Mostramos todas (APPROVED, PENDING, REJECTED, PAUSED)
            templates = []
            for t in templates_raw:
                # Extraer variables del body — usamos la fuente oficial de Meta:
                # 1. body_text_named_params → variables con nombre ({{nombre}})
                # 2. body_text              → variables posicionales ({{1}})
                # 3. Fallback regex
                variables = []
                named = False
                for comp in t.get("components", []):
                    if comp.get("type", "").upper() != "BODY":
                        continue
                    example = comp.get("example", {})
                    if example.get("body_text_named_params"):
                        # Plantilla con variables nombradas
                        variables = [p["param_name"] for p in example["body_text_named_params"] if p.get("param_name")]
                        named = True
                    elif example.get("body_text"):
                        # Plantilla con variables posicionales: body_text es [[val1, val2, ...]]
                        bt = example["body_text"]
                        count = len(bt[0]) if bt and bt[0] else 0
                        variables = [str(i + 1) for i in range(count)]
                    else:
                        # Fallback: regex sobre el texto de la plantilla
                        text = comp.get("text", "")
                        found = sorted(set(re.findall(r'\{\{([^}]+)\}\}', text)))
                        variables = found
                        named = any(not v.isdigit() for v in found)
                    break
                # Extraer header: tipo y URL de imagen si existe
                header_type = None
                header_url  = None
                for comp in t.get("components", []):
                    if comp.get("type", "").upper() != "HEADER":
                        continue
                    fmt = comp.get("format", "").upper()
                    header_type = fmt
                    if fmt == "IMAGE":
                        handles = comp.get("example", {}).get("header_handle", [])
                        header_url = handles[0] if handles else None
                    break

                templates.append({
                    "id":          t.get("id", ""),       # ID de la plantilla en Meta (para editar)
                    "name":        t.get("name", ""),
                    "language":    t.get("language", "es_CO"),
                    "status":      t.get("status", ""),
                    "variables":   variables,
                    "named":       named,       # True si usa {{nombre}}, False si usa {{1}}
                    "header_type": header_type, # "IMAGE", "TEXT", "VIDEO", None
                    "header_url":  header_url,  # URL de la imagen del header (si aplica)
                    "components":  t.get("components", []),  # componentes completos para edición
                    "preview":     next(
                        (c.get("text", "") for c in t.get("components", [])
                         if c.get("type", "").upper() == "BODY"),
                        ""
                    ),
                })

            logger.info(f"[broadcast] {len(templates)} plantillas APPROVED encontradas de {len(templates_raw)} totales")
            return JSONResponse(content={"templates": templates})

    except Exception as e:
        logger.error(f"[broadcast] Excepción en inbox_broadcast_templates: {e}", exc_info=True)
        return JSONResponse(content={"templates": [], "error": f"Error interno: {str(e)}"})


@app.post("/inbox/broadcast/send")
async def inbox_broadcast_send(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """
    Envía una plantilla aprobada a una lista de destinatarios personalizados.
    Body JSON:
      { "template": "nombre_plantilla",
        "language": "es_CO",
        "recipients": [
          {"phone": "573001234567", "variables": ["Juan"]},
          {"phone": "573009876543", "variables": ["María"]}
        ]
      }
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body          = await request.json()
    template_name = body.get("template", "")
    language      = body.get("language", "es_CO")
    recipients    = body.get("recipients", [])   # [{phone, variables:[]}]
    var_names     = body.get("var_names", [])    # nombres de variables, ej: ["nombre"]
    is_named      = body.get("named", False)     # True si plantilla usa {{nombre}}, False si {{1}}
    header_type   = body.get("header_type")      # "IMAGE", "TEXT", etc.
    header_url    = body.get("header_url")       # URL de imagen del header (si aplica)
    campaign_name = body.get("campaign_name", "")  # nombre de la campaña dado por el usuario
    campaign_id   = body.get("campaign_id", "")    # ID único compartido entre todos los lotes
    agent_id_dif  = int(body.get("agent_id", 1))   # agente que envía la difusión
    # Texto del body del template — el frontend lo manda para que podamos
    # guardarlo en el historial del cliente con variables sustituidas (#71).
    # Si no viene, usamos un fallback descriptivo.
    body_text     = body.get("body_text", "")

    if not template_name or not recipients:
        return JSONResponse(content={"error": "Faltan template o recipients"}, status_code=400)

    access_token    = os.getenv("META_ACCESS_TOKEN", "")
    phone_number_id = os.getenv("META_PHONE_NUMBER_ID", "")
    if not access_token or not phone_number_id:
        return JSONResponse(content={"error": "META_ACCESS_TOKEN o META_PHONE_NUMBER_ID no configurados"}, status_code=500)

    api_ver = "v21.0"
    api_url = f"https://graph.facebook.com/{api_ver}/{phone_number_id}/messages"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

    enviados   = 0
    fallidos   = 0
    opt_outs   = 0
    errores    = []

    # Pre-cargar lista de opt-outs para filtrar sin consulta por cada destinatario
    opt_out_set: set[str] = {r["telefono"] for r in await obtener_opt_outs()}
    if opt_out_set:
        logger.info(f"[broadcast] {len(opt_out_set)} números en opt-out — serán saltados")

    # Si el template tiene IMAGE header, subir la imagen una sola vez y cachear el media_id
    header_media_id: str | None = None
    if header_type == "IMAGE" and header_url:
        header_media_id = await _subir_imagen_whatsapp(header_url, access_token, phone_number_id)
        if not header_media_id:
            logger.warning("[broadcast] No se pudo obtener media_id para el header — se enviará sin imagen")

    async with httpx.AsyncClient(timeout=15) as client:
        for dest in recipients:
            tel = "".join(filter(str.isdigit, str(dest.get("phone", ""))))
            if not tel or len(tel) < 10:
                fallidos += 1
                errores.append(f"Número inválido: {dest.get('phone')}")
                continue

            # Saltar números dados de baja
            if tel in opt_out_set:
                opt_outs += 1
                logger.info(f"[broadcast] {tel[-4:]}**** saltado — opt-out")
                continue

            # Variables específicas de este destinatario
            vars_dest = dest.get("variables", [])
            components = []

            # HEADER con media_id: usamos el ID subido vía Media API (no la URL de scontent)
            if header_media_id:
                components.append({
                    "type": "header",
                    "parameters": [{"type": "image", "image": {"id": header_media_id}}],
                })

            # BODY con variables personalizadas
            if vars_dest:
                parameters = []
                for i, v in enumerate(vars_dest):
                    param = {"type": "text", "text": str(v)}
                    # parameter_name para plantillas con variables nombradas ({{nombre}})
                    if is_named and i < len(var_names) and not var_names[i].isdigit():
                        param["parameter_name"] = var_names[i]
                    parameters.append(param)
                components.append({
                    "type": "body",
                    "parameters": parameters,
                })

            payload = {
                "messaging_product": "whatsapp",
                "to": tel,
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": language},
                    **({"components": components} if components else {}),
                },
            }
            try:
                logger.info(f"[broadcast] named={is_named} var_names={var_names} Payload → {payload}")
                r = await client.post(api_url, json=payload, headers=headers)
                if r.status_code == 200:
                    enviados += 1
                    try:
                        resp_data  = r.json()
                        wamid      = resp_data.get("messages", [{}])[0].get("id", "")
                        msg_status = resp_data.get("messages", [{}])[0].get("message_status", "?")
                    except Exception:
                        wamid, msg_status = "", "?"
                    logger.info(f"[broadcast] Enviado a {tel[-4:]}**** wamid={wamid[:20] if wamid else '?'} status={msg_status}")
                    # Guardar wamid para tracking de delivery/lectura
                    if wamid and campaign_id:
                        try:
                            await guardar_mensaje_difusion(
                                wamid=wamid,
                                campaign_id=campaign_id,
                                campaign_name=campaign_name,
                                telefono=tel,
                                agent_id=agent_id_dif,
                            )
                            logger.warning(f"[broadcast✅] wamid guardado para {tel[-4:]}**** campaign={campaign_id[:20]}")
                        except Exception as _e:
                            logger.warning(f"[broadcast⚠️] No se guardó wamid en difusion_mensajes: {_e}")
                    elif not campaign_id:
                        logger.warning(f"[broadcast⚠️] campaign_id vacío — no se guarda wamid para {tel[-4:]}****")

                    # #71 — Guardar el mensaje en historial de Conversaciones para
                    # que el merchant vea la difusión enviada en el chat del cliente.
                    # Renderizamos el body sustituyendo variables ({{1}}, {{nombre}}, etc).
                    try:
                        texto_render = body_text or ""
                        if texto_render and vars_dest:
                            for i, val in enumerate(vars_dest, start=1):
                                # Sustituir tanto posicionales {{1}} como nombradas {{nombre}}
                                texto_render = texto_render.replace("{{" + str(i) + "}}", str(val))
                                if is_named and i - 1 < len(var_names):
                                    texto_render = texto_render.replace("{{" + var_names[i-1] + "}}", str(val))
                        # Si no llegó body_text, marcador descriptivo
                        if not texto_render:
                            texto_render = f"[Plantilla: {template_name}]"
                        # Prefijo 📣 para distinguir visualmente del chat orgánico
                        texto_para_historial = f"📣 {texto_render}"
                        await guardar_mensaje(tel, "assistant", texto_para_historial, agent_id=1,
                                              wamid=wamid, status="sent" if wamid else "")
                    except Exception as _e:
                        logger.warning(f"[broadcast⚠️] No se guardó en historial mensajes: {_e}")
                else:
                    fallidos += 1
                    try:
                        err_data = r.json().get("error", {})
                        err_code = err_data.get("code", "")
                        err_msg  = err_data.get("message", r.text[:200])
                        err_detail = f"#{err_code} {err_msg} | payload_components={payload.get('template',{}).get('components','none')}"
                    except Exception:
                        err_detail = r.text[:300]
                    errores.append(f"{tel}: {err_detail}")
                    logger.warning(f"[broadcast] Falló {tel}: {r.status_code} — {r.text[:400]}")
            except Exception as e:
                fallidos += 1
                errores.append(f"{tel}: {e}")

            # Cadencia: 10 mensajes/segundo
            await asyncio.sleep(0.1)

    logger.info(
        f"[broadcast] Difusión '{template_name}': "
        f"{enviados} enviados, {fallidos} fallidos, {opt_outs} opt-outs saltados"
    )

    # Registrar resultado en BD para el historial de difusiones
    try:
        await registrar_difusion(
            template_name=template_name,
            language=language,
            destinatarios=len(recipients),
            enviados=enviados,
            fallidos=fallidos,
            errores=errores,
            campaign_name=campaign_name,
            campaign_id=campaign_id,
            agent_id=agent_id_dif,
        )
    except Exception as e:
        logger.warning(f"[broadcast] No se pudo registrar difusión en BD: {e}")

    return JSONResponse(content={
        "enviados":  enviados,
        "fallidos":  fallidos,
        "opt_outs":  opt_outs,
        "errores":   errores[:20],
    })


@app.get("/inbox/difusiones/historial")
async def inbox_difusiones_historial(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el historial de difusiones enviadas desde el inbox."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        rows = await obtener_difusiones(100, agent_id=agent_id)
        return JSONResponse(content={"difusiones": rows})
    except Exception as e:
        logger.error(f"[historial-dif] Error: {e}", exc_info=True)
        return JSONResponse(content={"difusiones": [], "error": str(e)[:300]})


@app.get("/inbox/difusiones/campana/{campaign_id:path}")
async def inbox_difusion_detalle(
    campaign_id: str,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Detalle de una campaña: breakdown de entregados, leídos, fallidos con motivo de error."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    data = await obtener_detalle_campana(campaign_id)
    return JSONResponse(content=data)


# ══════════════════════════════════════════════════════════════════════════════
# #51 — Wizard onboarding Meta WhatsApp (8 pasos guiados)
# ══════════════════════════════════════════════════════════════════════════════

# El progreso se persiste en un único registro ConfigValue como JSON,
# scopeado por agent_id — así cada agente (cada tenant) tiene su propio
# estado del wizard. Cuando cambian de usuario, ven el progreso de SU agente,
# no el de otro.
_ONBOARDING_KEY = "ONBOARDING_META_PROGRESS"
# 6 pasos en total:
#   1 Crear Business Manager
#   2 Verificar negocio
#   3 Crear WABA + agregar número (Meta lo combina en un solo modal)
#   4 Invitar Voco como socio
#   5 Generar token System User
#   6 Conectar Voco (valida + registra número + suscribe webhook automático)
_ONBOARDING_TOTAL_PASOS = 6
# Camino corto = solo 2 pasos NUEVOS (3 = crear WABA+número, 6 = conectar).
_ONBOARDING_TOTAL_PASOS_CORTO = 2


async def _resolver_agent_id_para_user(user: dict) -> int | None:
    """Devuelve el agent_id activo del usuario para el wizard.

    Estrategia: primer agente del que el user es owner. Si el user no tiene
    agentes propios (ej: cuenta recién creada), retorna None — el handler
    debe redirigir a Mis Agentes para crear uno.
    """
    if not user or not user.get("id"):
        return None
    from agent.memory import obtener_agentes_de_usuario
    agentes = await obtener_agentes_de_usuario(int(user["id"]))
    if not agentes:
        return None
    return int(agentes[0]["id"])


async def _leer_onboarding_estado(agent_id: int = 1) -> dict:
    raw = await get_config_value(_ONBOARDING_KEY, agent_id) or ""
    if not raw:
        return {"pasos": {}}
    try:
        import json as _json
        data = _json.loads(raw)
        if not isinstance(data, dict) or "pasos" not in data:
            return {"pasos": {}}
        return data
    except Exception:
        return {"pasos": {}}


async def _guardar_onboarding_estado(estado: dict, agent_id: int = 1, user: dict | None = None) -> None:
    import json as _json
    await set_config_value(
        _ONBOARDING_KEY,
        _json.dumps(estado),
        agent_id=agent_id,
        usuario_id=(user or {}).get("id"),
        usuario_email=(user or {}).get("email", ""),
    )


@app.get("/inbox/api/onboarding/estado")
async def api_onboarding_estado(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el estado del wizard del agente del usuario actual + flag
    `tiene_meta_previo` para que el frontend muestre la bifurcación si aplica."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    agent_id = await _resolver_agent_id_para_user(user)
    if agent_id is None:
        return JSONResponse({"pasos": {}, "completados": 0, "total": _ONBOARDING_TOTAL_PASOS,
                             "porcentaje": 0, "sin_agente": True})
    from agent.memory import usuario_tiene_meta_configurado
    estado = await _leer_onboarding_estado(agent_id)
    # tiene_meta_previo: el user ya completó el wizard largo en OTRO agente.
    # Si esto es true Y este agente todavía no eligió camino, mostramos
    # la pantalla de bifurcación.
    tiene_previo = await usuario_tiene_meta_configurado(int(user["id"]))
    # Es "el agente actual ya completó algo" si tiene cualquier paso marcado
    actual_tiene_progreso = any(p.get("completado") for p in estado["pasos"].values())
    # Es propiamente "otro agente" si tiene_previo Y el agente activo no es
    # el que tiene el paso 8 — chequeo simple: si este mismo tiene paso 8, no
    # es escenario de "segundo agente"
    este_ya_termino = estado["pasos"].get("8", {}).get("completado", False)
    mostrar_bifurcacion = bool(tiene_previo and not este_ya_termino and not actual_tiene_progreso)

    # Camino del wizard que el merchant eligió para ESTE agente — guardado
    # como una clave especial dentro del propio JSON de progreso.
    camino = estado.get("_camino") or ""

    completados = sum(1 for p in estado["pasos"].values() if p.get("completado"))
    total = _ONBOARDING_TOTAL_PASOS_CORTO if camino == "corto" else _ONBOARDING_TOTAL_PASOS
    return JSONResponse({
        "agent_id": agent_id,
        "pasos": estado["pasos"],
        "completados": completados,
        "total": total,
        "porcentaje": round((completados / total) * 100) if total else 0,
        "mostrar_bifurcacion": mostrar_bifurcacion,
        "camino": camino,  # '' / 'corto' / 'completo'
    })


@app.post("/inbox/api/onboarding/reset")
async def api_onboarding_reset(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Borra TODO el progreso del wizard para el agente actual. Útil cuando el
    merchant quiere recomenzar (cambió de número, se equivocó de cuenta, etc).
    El historial de cambios queda registrado vía set_config_value."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    agent_id = await _resolver_agent_id_para_user(user)
    if agent_id is None:
        return JSONResponse({"ok": False, "error": "Sin agente"}, status_code=400)
    await _guardar_onboarding_estado({"pasos": {}}, agent_id, user)
    logger.info(f"[onboarding] Wizard reiniciado para agent_id={agent_id} por user={user.get('email')}")
    return JSONResponse({"ok": True})


@app.post("/inbox/api/onboarding/camino/{nombre}")
async def api_onboarding_set_camino(
    nombre: str,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Marca qué camino eligió el merchant: 'corto' (reusa BM) o 'completo'.
    Esto define qué pasos rendea el wizard a partir de ahora para este agente."""
    if nombre not in ("corto", "completo"):
        return JSONResponse({"ok": False, "error": "Camino inválido"}, status_code=400)
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    agent_id = await _resolver_agent_id_para_user(user)
    if agent_id is None:
        return JSONResponse({"ok": False, "error": "Sin agente"}, status_code=400)
    estado = await _leer_onboarding_estado(agent_id)
    estado["_camino"] = nombre
    # Si eligió camino corto, marcamos automáticamente los pasos heredados
    # del BM existente: 1 (BM creado), 2 (verificado), 5 (Voco invitado),
    # 6 (token generado), 8 (webhook activo a nivel app)
    if nombre == "corto":
        ahora = datetime.utcnow().isoformat()
        # 1=BM, 2=verificación, 4=socio invitado, 5=token generado — todos
        # heredables del primer agente. Paso 6 (conectar+webhook+registro) sí
        # toca hacerlo de nuevo porque el número y WABA son nuevos.
        for paso_n in ("1", "2", "4", "5"):
            estado["pasos"][paso_n] = {
                "completado": True,
                "datos": {"_heredado": True},
                "fecha": ahora,
                "usuario_email": user.get("email", ""),
            }
    await _guardar_onboarding_estado(estado, agent_id, user)
    return JSONResponse({"ok": True, "camino": nombre})


@app.post("/inbox/api/onboarding/paso/{n}/completar")
async def api_onboarding_completar(
    n: int,
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Marca un paso como completado. Body opcional: {"datos": {...}} con campos
    capturados en ese paso (Business Manager ID, número de teléfono, etc)."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    if n < 0 or n > _ONBOARDING_TOTAL_PASOS:
        return JSONResponse({"ok": False, "error": f"Paso {n} fuera de rango"}, status_code=400)
    agent_id = await _resolver_agent_id_para_user(user)
    if agent_id is None:
        return JSONResponse({"ok": False, "error": "Primero crea un agente"}, status_code=400)
    try:
        body = await request.json()
    except Exception:
        body = {}
    datos = body.get("datos") or {}
    estado = await _leer_onboarding_estado(agent_id)
    estado["pasos"][str(n)] = {
        "completado": True,
        "datos": datos,
        "fecha": datetime.utcnow().isoformat(),
        "usuario_email": user.get("email", ""),
    }
    await _guardar_onboarding_estado(estado, agent_id, user)
    return JSONResponse({"ok": True})


@app.post("/inbox/api/onboarding/paso/{n}/reabrir")
async def api_onboarding_reabrir(
    n: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Reabre un paso (desmarca como completado). Útil si el merchant quiere
    volver a revisar algo."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    agent_id = await _resolver_agent_id_para_user(user)
    if agent_id is None:
        return JSONResponse({"ok": False, "error": "Sin agente"}, status_code=400)
    estado = await _leer_onboarding_estado(agent_id)
    if str(n) in estado["pasos"]:
        estado["pasos"][str(n)]["completado"] = False
        await _guardar_onboarding_estado(estado, agent_id, user)
    return JSONResponse({"ok": True})


@app.post("/inbox/api/onboarding/validar-meta")
async def api_onboarding_validar_meta(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Valida credenciales Meta haciendo 3 llamadas reales: /me, /{phone_id},
    /{waba_id}/templates. Devuelve qué pasó con cada una para feedback claro."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    tok      = (body.get("access_token") or "").strip()
    phone_id = (body.get("phone_number_id") or "").strip()
    waba_id  = (body.get("waba_id") or "").strip()
    if not tok or not phone_id:
        return JSONResponse({"ok": False, "error": "Faltan access_token y/o phone_number_id"})

    resultados = {"me": None, "phone": None, "waba": None}
    api_ver = "v21.0"
    async with httpx.AsyncClient(timeout=10) as client:
        # /me — token válido?
        try:
            r = await client.get(f"https://graph.facebook.com/{api_ver}/me",
                                 params={"access_token": tok})
            resultados["me"] = {"ok": r.status_code == 200, "detalle": r.json() if r.status_code == 200 else r.text[:200]}
        except Exception as e:
            resultados["me"] = {"ok": False, "detalle": str(e)}
        # /{phone_id} — token tiene acceso al número?
        try:
            r = await client.get(f"https://graph.facebook.com/{api_ver}/{phone_id}",
                                 params={"access_token": tok})
            resultados["phone"] = {"ok": r.status_code == 200, "detalle": r.json() if r.status_code == 200 else r.text[:200]}
        except Exception as e:
            resultados["phone"] = {"ok": False, "detalle": str(e)}
        # /{waba_id}/message_templates — token tiene scope whatsapp_business_management?
        if waba_id:
            try:
                r = await client.get(f"https://graph.facebook.com/{api_ver}/{waba_id}/message_templates",
                                     params={"access_token": tok, "limit": 1})
                resultados["waba"] = {"ok": r.status_code == 200, "detalle": r.json() if r.status_code == 200 else r.text[:200]}
            except Exception as e:
                resultados["waba"] = {"ok": False, "detalle": str(e)}

    todo_ok = resultados["me"]["ok"] and resultados["phone"]["ok"] and (not waba_id or resultados["waba"]["ok"])

    # Si las 3 validaciones pasaron, ejecutamos lo que el merchant antes hacía
    # MANUALMENTE en API Setup → "Enviar mensaje": registrar el número para
    # Cloud API (sin esto Meta no enruta los mensajes al webhook) + suscribir
    # nuestra app a los webhooks de la WABA. Todo en backend, sin que el
    # merchant tenga que entrar a developers.facebook.com.
    registro_numero = None
    webhook_subscrito = None
    if todo_ok:
        # PIN para 2FA del número. Default '000000' — Meta lo acepta como nuevo
        # PIN si el número aún no tiene 2FA. El merchant puede cambiarlo después
        # desde WhatsApp Manager si quiere.
        pin = (body.get("pin") or "000000").strip()
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                rr = await client.post(
                    f"https://graph.facebook.com/{api_ver}/{phone_id}/register",
                    headers={"Authorization": f"Bearer {tok}"},
                    json={"messaging_product": "whatsapp", "pin": pin},
                )
                # Códigos comunes:
                # 200 → registro exitoso
                # 400 con "already registered" → ya estaba registrado, OK
                # 400 con "pin mismatch" → ya tiene 2FA con otro PIN
                resp_text = rr.text[:300]
                ok = rr.status_code == 200 or "already" in resp_text.lower()
                registro_numero = {
                    "ok":      ok,
                    "status":  rr.status_code,
                    "detalle": resp_text,
                }
        except Exception as e:
            registro_numero = {"ok": False, "detalle": str(e)}

        if waba_id:
            try:
                async with httpx.AsyncClient(timeout=10) as client:
                    rs = await client.post(
                        f"https://graph.facebook.com/{api_ver}/{waba_id}/subscribed_apps",
                        params={"access_token": tok},
                    )
                    webhook_subscrito = {"ok": rs.status_code == 200,
                                         "detalle": rs.json() if rs.status_code == 200 else rs.text[:200]}
            except Exception as e:
                webhook_subscrito = {"ok": False, "detalle": str(e)}

    return JSONResponse({
        "ok": todo_ok,
        "resultados": resultados,
        "registro_numero":   registro_numero,
        "webhook_subscrito": webhook_subscrito,
    })


# Página del wizard (HTML standalone, fuera del panel principal)
@app.get("/inbox/onboarding-meta", response_class=HTMLResponse)
async def inbox_onboarding_meta(
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Página del wizard de onboarding Meta WhatsApp.
    Standalone — no usa el shell del panel para foco total en la tarea."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session)
    if not user:
        return HTMLResponse(status_code=302, headers={"Location": "/inbox/login"}, content="")
    # Si el usuario no tiene agentes, no tiene sentido abrir el wizard.
    # Lo enviamos a Mis Agentes para que cree uno primero.
    agent_id = await _resolver_agent_id_para_user(user)
    if agent_id is None:
        return HTMLResponse(status_code=302, headers={"Location": "/inbox"}, content="")
    from agent.inbox import HTML_WIZARD_ONBOARDING
    webhook_url = (os.getenv("PUBLIC_BASE_URL") or "").rstrip("/") + "/webhook"
    return HTMLResponse(content=HTML_WIZARD_ONBOARDING.replace("__WEBHOOK_URL__", webhook_url))


@app.get("/inbox/api/metricas/operacion")
async def inbox_metricas_operacion(
    dias: int = 7,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Dashboard de salud del flujo (#47): tráfico + conversión + escalación.
    Devuelve métricas del período actual vs período anterior para detectar
    tendencias. agent_name viene incluido para que la UI no hardcodee 'Andrea'."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_metricas_operacion, obtener_agente
    dias = max(1, min(int(dias or 7), 90))
    data = await obtener_metricas_operacion(agent_id=agent_id, dias=dias)
    ag = await obtener_agente(agent_id) or {}
    data["agent_name"]    = ag.get("agent_name") or "Bot"
    data["agent_business"] = ag.get("name") or ""
    return JSONResponse(content=data)


@app.get("/inbox/metricas/resumen")
async def inbox_metricas_resumen(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Métricas de WhatsApp Business: mensajes enviados, entregados, leídos (últimos 30 días)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = os.getenv("META_ACCESS_TOKEN", "")
    waba_id      = os.getenv("META_WABA_ID", "")
    phone_number_id = os.getenv("META_PHONE_NUMBER_ID", "")

    if not access_token or not waba_id:
        return JSONResponse(content={"error": "META_ACCESS_TOKEN o META_WABA_ID no configurados"})

    import datetime as dt
    hoy   = dt.date.today()
    start = int((hoy - dt.timedelta(days=30)).strftime("%s") if hasattr(hoy, "strftime") else 0)
    end   = int(hoy.strftime("%s") if hasattr(hoy, "strftime") else 0)

    # Usar timestamps de época correctamente
    import time as _time
    end_ts   = int(_time.time())
    start_ts = end_ts - 30 * 86400

    api_ver = "v21.0"
    # WABA Analytics: mensajes enviados, entregados y leídos por día
    analytics_url = (
        f"https://graph.facebook.com/{api_ver}/{waba_id}/analytics"
        f"?granularity=MONTHLY&start={start_ts}&end={end_ts}"
        f"&metric_types=['SENT','DELIVERED','READ']"
        f"&access_token={access_token}"
    )

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(analytics_url)
            data = r.json()
            if "error" in data:
                # Fallback: intentar con phone_number_id si WABA falla
                ph_url = (
                    f"https://graph.facebook.com/{api_ver}/{phone_number_id}/analytics"
                    f"?granularity=MONTHLY&start={start_ts}&end={end_ts}"
                    f"&metric_types=['SENT','DELIVERED','READ']"
                    f"&access_token={access_token}"
                )
                r2 = await client.get(ph_url)
                data = r2.json()
            if "error" in data:
                return JSONResponse(content={"error": data["error"].get("message", str(data["error"])), "raw": data})
            # Sumar totales del periodo
            puntos = data.get("data", {}).get("data_points", [])
            totales = {"sent": 0, "delivered": 0, "read": 0}
            for p in puntos:
                totales["sent"]      += p.get("sent", 0)
                totales["delivered"] += p.get("delivered", 0)
                totales["read"]      += p.get("read", 0)
            return JSONResponse(content={"resumen": totales, "puntos": puntos})
    except Exception as e:
        logger.error(f"[metricas] Error consultando Analytics API: {e}")
        return JSONResponse(content={"error": str(e)})


@app.get("/inbox/metricas/plantillas")
async def inbox_metricas_plantillas(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Analíticas por plantilla: enviados, entregados, leídos, clics en botón."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = os.getenv("META_ACCESS_TOKEN", "")
    waba_id      = os.getenv("META_WABA_ID", "")

    if not access_token or not waba_id:
        return JSONResponse(content={"error": "META_ACCESS_TOKEN o META_WABA_ID no configurados"})

    import time as _time
    end_ts   = int(_time.time())
    start_ts = end_ts - 30 * 86400
    api_ver  = "v21.0"

    url = (
        f"https://graph.facebook.com/{api_ver}/{waba_id}/template_analytics"
        f"?granularity=MONTHLY&start={start_ts}&end={end_ts}"
        f"&access_token={access_token}"
    )
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(url)
            data = r.json()
            if "error" in data:
                return JSONResponse(content={"error": data["error"].get("message", str(data["error"])), "raw": data})
            return JSONResponse(content={"analytics": data.get("data", [])})
    except Exception as e:
        logger.error(f"[metricas] Error consultando template_analytics: {e}")
        return JSONResponse(content={"error": str(e)})


@app.get("/inbox/metricas/interno")
async def inbox_metricas_interno(
    dias: int = 30,
    desde: str = "",        # ISO date opcional: "2025-05-01"
    hasta: str = "",        # ISO date opcional: "2025-05-31"
    granularidad: str = "dia",   # dia | semana | mes (para series temporales)
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Métricas calculadas desde la base de datos interna.
    Acepta rango específico con `desde`/`hasta` (formato ISO YYYY-MM-DD) o
    fallback a últimos `dias` días. Incluye series temporales con la granularidad."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        # Parsear rango de fechas si vino explícito
        desde_dt = None
        hasta_dt = None
        if desde:
            try:
                desde_dt = datetime.fromisoformat(desde.replace("Z", ""))
            except Exception:
                logger.warning(f"[metricas] 'desde' inválido: {desde}")
        if hasta:
            try:
                hasta_dt = datetime.fromisoformat(hasta.replace("Z", ""))
                # Si solo es fecha (YYYY-MM-DD), incluir hasta el final del día
                if len(hasta) <= 10:
                    hasta_dt = hasta_dt.replace(hour=23, minute=59, second=59)
            except Exception:
                logger.warning(f"[metricas] 'hasta' inválido: {hasta}")

        # Métricas agregadas
        data = await obtener_metricas_internas(
            dias=dias, desde=desde_dt, hasta=hasta_dt, agent_id=agent_id,
        )
        # Series temporales (usar fechas efectivas calculadas)
        from agent.memory import obtener_series_metricas
        desde_eff = datetime.fromisoformat(data["desde"])
        hasta_eff = datetime.fromisoformat(data["hasta"])
        series = await obtener_series_metricas(
            desde=desde_eff, hasta=hasta_eff,
            granularidad=granularidad, agent_id=agent_id,
        )
        data["series"]       = series
        data["granularidad"] = granularidad
        return JSONResponse(content=data)
    except Exception as e:
        logger.error(f"[metricas-interno] Error: {e}", exc_info=True)
        return JSONResponse(content={"error": str(e) or type(e).__name__})


@app.get("/inbox/api/opt-outs")
async def inbox_opt_outs(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista de números dados de baja de difusiones masivas."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    rows = await obtener_opt_outs(agent_id=agent_id)
    return JSONResponse(content={"opt_outs": rows, "total": len(rows)})


@app.get("/inbox/api/clientes/lookup")
async def inbox_cliente_lookup(
    telefono: str = "",
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Busca un cliente por teléfono exacto. 404 si no existe."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    if not telefono:
        raise HTTPException(status_code=400, detail="Falta telefono")
    from agent.memory import obtener_cliente
    cli = await obtener_cliente(telefono.strip(), agent_id)
    if not cli:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    nombre = ((cli.get("nombres") or "") + " " + (cli.get("apellidos") or "")).strip()
    return JSONResponse(content={**cli, "nombre": nombre})


@app.get("/inbox/api/clientes")
async def inbox_clientes(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Base de clientes con estado de engagement (activo/tibio/frío/baja)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    clientes = await obtener_clientes_con_estado(agent_id=agent_id)
    resumen = {"total": len(clientes), "activo": 0, "tibio": 0, "frio": 0, "baja": 0}
    for c in clientes:
        resumen[c["estado"]] = resumen.get(c["estado"], 0) + 1
    return JSONResponse(content={"clientes": clientes, "resumen": resumen})


@app.post("/inbox/api/clientes/edit")
async def inbox_editar_cliente(
    request: Request,
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Edita nombre/teléfono/ciudad de un cliente desde el panel.
    Si cambia el teléfono, migra todas las referencias (mensajes, escalaciones, etc).
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session):
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Payload inválido")

    telefono_original = (body.get("telefono_original") or "").strip()
    if not telefono_original:
        return JSONResponse({"ok": False, "error": "Falta telefono_original"}, status_code=400)

    datos = {
        "nombres":   (body.get("nombres") or "").strip(),
        "apellidos": (body.get("apellidos") or "").strip(),
        "direccion": (body.get("direccion") or "").strip(),
        "direccion2": (body.get("direccion2") or "").strip(),
        "ciudad":    (body.get("ciudad") or "").strip(),
    }

    agent_id = int(body.get("agent_id") or 1)
    telefono_nuevo_raw = (body.get("telefono") or "").strip()
    telefono_nuevo: str | None = None
    if telefono_nuevo_raw:
        telefono_nuevo = _normalizar_telefono(telefono_nuevo_raw)
        if not telefono_nuevo:
            return JSONResponse(
                {"ok": False, "error": "Teléfono inválido"},
                status_code=400,
            )

    resultado = await editar_cliente(
        telefono_original=telefono_original,
        datos=datos,
        agent_id=agent_id,
        telefono_nuevo=telefono_nuevo,
    )
    status = 200 if resultado["ok"] else 409
    return JSONResponse(resultado, status_code=status)


def _normalizar_telefono(raw: str) -> str | None:
    """Normaliza un número de teléfono para WhatsApp.
    Acepta formatos con espacios, guiones, signos +.
    Si comienza con 3 y tiene 10 dígitos, antepone código país 57 (Colombia)."""
    digits = re.sub(r'[^0-9]', '', str(raw))
    if not digits:
        return None
    if digits.startswith('57') and len(digits) == 12:
        return digits
    if digits.startswith('3') and len(digits) == 10:
        return '57' + digits
    if len(digits) >= 7:
        return digits  # aceptar tal cual para internacionales
    return None


@app.get("/inbox/api/clientes/import/template")
async def inbox_plantilla_csv(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Descarga la plantilla CSV con las columnas correctas y ejemplos."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from fastapi.responses import Response
    contenido = (
        "telefono,nombres,apellidos,ciudad,departamento,email,cc_nit\n"
        "3001234567,Juan,Pérez,Bogotá,Cundinamarca,juan@email.com,12345678\n"
        "3159876543,María,López,Medellín,Antioquia,maria@email.com,87654321\n"
        "573201112233,Carlos,García,Cali,Valle del Cauca,,\n"
    )
    return Response(
        content=contenido.encode("utf-8-sig"),  # BOM para Excel en español
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes.csv"},
    )


@app.post("/inbox/api/clientes/import")
async def inbox_importar_clientes(
    request: Request,
    file: UploadFile = File(...),
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Importa clientes masivamente desde un archivo CSV (campo 'file' en multipart/form-data).
    Columnas requeridas: telefono, nombres, apellidos. Opcionales: ciudad, departamento, email, cc_nit."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    contenido = await file.read()
    try:
        texto = contenido.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = contenido.decode("latin-1", errors="replace")

    reader = csv.DictReader(io.StringIO(texto))

    CAMPOS_OPCIONALES = {"nombres", "apellidos", "ciudad", "departamento", "email", "cc_nit"}

    total = 0
    insertados = 0
    actualizados = 0
    errores = 0
    filas_error: list[dict] = []

    for i, fila in enumerate(reader, start=2):  # start=2: fila 1 es cabecera
        total += 1
        fila_norm = {k.strip().lower(): (v or "").strip() for k, v in fila.items()}
        raw_tel = fila_norm.get("telefono", "")
        if not raw_tel:
            errores += 1
            filas_error.append({"fila": i, "razon": "columna 'telefono' vacía", "datos": fila_norm})
            continue

        tel = _normalizar_telefono(raw_tel)
        if not tel:
            errores += 1
            filas_error.append({"fila": i, "razon": f"número inválido: {raw_tel}", "datos": fila_norm})
            continue

        # nombres y apellidos son obligatorios
        if not fila_norm.get("nombres") or not fila_norm.get("apellidos"):
            errores += 1
            filas_error.append({"fila": i, "razon": "nombres y apellidos son obligatorios", "datos": fila_norm})
            continue

        datos = {campo: fila_norm[campo] for campo in CAMPOS_OPCIONALES if fila_norm.get(campo)}
        try:
            resultado = await guardar_cliente_import(tel, datos, agent_id)
            if resultado == "inserted":
                insertados += 1
            else:
                actualizados += 1
        except Exception as e:
            errores += 1
            filas_error.append({"fila": i, "razon": str(e), "datos": fila_norm})

    logger.info(f"[import-clientes] total={total} insertados={insertados} actualizados={actualizados} errores={errores}")
    return JSONResponse(content={
        "ok": True,
        "total": total,
        "inserted": insertados,
        "updated": actualizados,
        "errors": errores,
        "error_rows": filas_error[:20],  # máximo 20 ejemplos de errores
    })


@app.get("/inbox/api/clientes/templates")
async def inbox_listar_templates_clientes(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista plantillas aprobadas de Meta para enviar a clientes individuales."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = await get_config_value("META_ACCESS_TOKEN", agent_id) or os.getenv("META_ACCESS_TOKEN", "")
    waba_id      = await get_config_value("META_WABA_ID", agent_id)      or os.getenv("META_WABA_ID", "")

    if not access_token or not waba_id:
        return JSONResponse(content={"templates": [], "error": "META_ACCESS_TOKEN o META_WABA_ID no configurados"})

    api_ver = "v21.0"
    url = (
        f"https://graph.facebook.com/{api_ver}/{waba_id}"
        f"/message_templates?fields=id,name,status,components,language&limit=100"
        f"&access_token={access_token}"
    )
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.get(url)
        if r.status_code != 200:
            return JSONResponse(content={"templates": [], "error": f"Meta API {r.status_code}: {r.text[:200]}"})
        data = r.json()
        if "error" in data:
            return JSONResponse(content={"templates": [], "error": data["error"].get("message", str(data["error"]))})

        templates = []
        for t in data.get("data", []):
            if (t.get("status", "")).upper() != "APPROVED":
                continue
            body_text = next(
                (c.get("text", "") for c in t.get("components", []) if c.get("type", "").upper() == "BODY"),
                ""
            )
            templates.append({
                "name":     t.get("name", ""),
                "language": t.get("language", "es_CO"),
                "category": t.get("category", ""),
                "preview":  body_text,
                "components": t.get("components", []),
            })
        return JSONResponse(content={"templates": templates})
    except Exception as e:
        logger.error(f"[clientes/templates] Error: {e}", exc_info=True)
        return JSONResponse(content={"templates": [], "error": str(e)})


@app.post("/inbox/api/clientes/message")
async def inbox_enviar_mensaje_cliente(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Envía una plantilla aprobada de Meta a un cliente específico.
    Body JSON: {telefono, template_name, language_code, components?}"""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body          = await request.json()
    telefono      = str(body.get("telefono", "")).strip()
    template_name = str(body.get("template_name", "")).strip()
    language_code = str(body.get("language_code", "es_CO")).strip()
    components    = body.get("components")  # lista de componentes, puede ser None

    tel = _normalizar_telefono(telefono)
    if not tel:
        return JSONResponse(content={"ok": False, "error": f"Número inválido: {telefono}"}, status_code=400)
    if not template_name:
        return JSONResponse(content={"ok": False, "error": "template_name es requerido"}, status_code=400)

    access_token    = await get_config_value("META_ACCESS_TOKEN", agent_id)    or os.getenv("META_ACCESS_TOKEN", "")
    phone_number_id = await get_config_value("META_PHONE_NUMBER_ID", agent_id) or os.getenv("META_PHONE_NUMBER_ID", "")

    from agent.providers.meta import ProveedorMeta
    meta = ProveedorMeta(
        access_token=access_token,
        phone_number_id=phone_number_id,
    )
    resultado = await meta.enviar_plantilla(
        telefono=tel,
        template_name=template_name,
        language_code=language_code,
        components=components,
    )
    if resultado.get("ok"):
        # Registrar en historial
        try:
            _w = resultado.get("message_id") or ""
            await guardar_mensaje(tel, "assistant", f"[Plantilla: {template_name}]", agent_id=agent_id,
                                  wamid=_w, status="sent" if _w else "")
        except Exception:
            pass
        logger.info(f"[clientes/message] Plantilla '{template_name}' enviada a {tel[-4:]}**** message_id={resultado.get('message_id','?')}")
    else:
        logger.warning(f"[clientes/message] Fallo enviando plantilla a {tel[-4:]}****: {resultado.get('error')}")

    return JSONResponse(content=resultado)


@app.delete("/inbox/api/opt-outs/{telefono}")
async def inbox_revertir_opt_out(
    telefono: str,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Reactiva un número para recibir difusiones (el cliente cambió de opinión)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    await revertir_opt_out(telefono)
    logger.info(f"[opt-out] {telefono} reactivado desde inbox")
    return JSONResponse(content={"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DINÁMICA — guardar credenciales en BD + probar conexiones
# ══════════════════════════════════════════════════════════════════════════════

_CONFIG_KEYS_ALLOWED = {
    "META_ACCESS_TOKEN", "META_PHONE_NUMBER_ID", "META_WABA_ID", "META_VERIFY_TOKEN",
    "META_CATALOG_ID",
    # Conversions API (Pixel server-side) — antes solo en Railway env vars
    "META_PIXEL_ID", "META_CAPI_TOKEN", "META_CAPI_TEST_CODE",
    "ANTHROPIC_API_KEY", "AI_MODEL",
    # Shopify post-#62: solo OAuth (modelo Jelou). El cliente entrega Client
    # ID + Client Secret; Voco hace el OAuth dance y obtiene ADMIN_TOKEN
    # automáticamente. STOREFRONT_TOKEN y WEBHOOK_SECRET fueron retirados.
    "SHOPIFY_STORE", "SHOPIFY_ADMIN_TOKEN",
    "SHOPIFY_CLIENT_ID", "SHOPIFY_CLIENT_SECRET",
    # Sprint 4 — reglas del negocio configurables por agente
    "PEDIDO_MINIMO", "PEDIDO_MIN_MSG",
}

_CONFIG_META = {
    "META_ACCESS_TOKEN":       {"label": "Token de acceso",    "tipo": "secret"},
    "META_PHONE_NUMBER_ID":    {"label": "Phone Number ID",    "tipo": "plain"},
    "META_WABA_ID":            {"label": "WABA ID",            "tipo": "plain"},
    "META_VERIFY_TOKEN":       {"label": "Verify Token",       "tipo": "plain"},
    "META_CATALOG_ID":         {"label": "Catalog ID",         "tipo": "plain"},
    "META_PIXEL_ID":           {"label": "Pixel ID (CAPI)",    "tipo": "plain"},
    "META_CAPI_TOKEN":         {"label": "Token CAPI",         "tipo": "secret"},
    "META_CAPI_TEST_CODE":     {"label": "Test Event Code",    "tipo": "plain"},
    "ANTHROPIC_API_KEY":       {"label": "API Key",            "tipo": "secret"},
    "AI_MODEL":                {"label": "Modelo IA",          "tipo": "plain"},
    "SHOPIFY_STORE":           {"label": "Dominio tienda",   "tipo": "plain"},
    "SHOPIFY_ADMIN_TOKEN":     {"label": "Admin API Token",  "tipo": "secret"},
    "SHOPIFY_CLIENT_ID":       {"label": "Client ID (OAuth)",     "tipo": "plain"},
    "SHOPIFY_CLIENT_SECRET":   {"label": "Client Secret (OAuth)", "tipo": "secret"},
    "PEDIDO_MINIMO":           {"label": "Pedido mínimo (COP)", "tipo": "plain"},
    "PEDIDO_MIN_MSG":          {"label": "Mensaje pedido mínimo", "tipo": "plain"},
}


@app.get("/inbox/api/config")
async def inbox_get_config(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el estado de cada clave de configuración (sin exponer valores secretos)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    resultado = {}
    for clave, meta in _CONFIG_META.items():
        db_val  = await get_config_value(clave, agent_id=agent_id)
        # Env vars son globales de la plataforma; solo se muestran para el
        # agente primario (id=1). Los demás tenants solo ven sus propios valores de BD.
        env_val = os.getenv(clave, "") if agent_id == 1 else ""
        valor   = db_val if db_val else env_val
        if valor:
            display = "•" * 8 if meta["tipo"] == "secret" else valor
            resultado[clave] = {"configurado": True, "display": display,
                                "fuente": "db" if db_val else "env"}
        else:
            resultado[clave] = {"configurado": False, "display": "", "fuente": None}

    return JSONResponse(content=resultado)


@app.post("/inbox/api/config/save")
async def inbox_save_config(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Guarda credenciales en la BD y las inyecta en el entorno actual."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")

    body = await request.json()
    # Claves que permiten valor vacío (para resetear) — el resto solo guarda si no está vacío
    _PERMITE_VACIO = {"PEDIDO_MIN_MSG", "META_CAPI_TEST_CODE"}
    saved = []
    for clave, valor in body.items():
        if clave in _CONFIG_KEYS_ALLOWED and isinstance(valor, str):
            valor = valor.strip()
            # Sanitizar caracteres invisibles que se copian de páginas web
            # (zero-width space, non-breaking space, BOM, etc.). Estos rompen
            # httpx con 'ascii codec can't encode' al usar el valor como header.
            # Bug reportado por Equora 13-jun al pegar token de Meta Business.
            for ch in ("​", "‌", "‍", " ", "﻿", " ", " "):
                valor = valor.replace(ch, "")
            # Para campos que van como HTTP header (tokens), forzar ASCII
            if clave in {"META_ACCESS_TOKEN", "META_CAPI_TOKEN",
                         "SHOPIFY_ADMIN_TOKEN", "SHOPIFY_CLIENT_SECRET",
                         "ANTHROPIC_API_KEY"}:
                try:
                    valor.encode("ascii")
                except UnicodeEncodeError:
                    # Filtrar a solo ASCII printable
                    valor = "".join(c for c in valor if 32 <= ord(c) < 127)
                    logger.warning(f"[config] {clave} tenía caracteres no-ASCII — filtrados")
            if valor or clave in _PERMITE_VACIO:
                # Pasamos contexto de usuario para que el historial (#48) sepa quién cambió qué.
                await set_config_value(
                    clave, valor,
                    agent_id=agent_id,
                    usuario_id=user.get("id"),
                    usuario_email=user.get("email", ""),
                )
                os.environ[clave] = valor   # actualizar en tiempo real
                saved.append(clave)

    # Si se actualizó META_CATALOG_ID o META_ACCESS_TOKEN, invalidar el
    # cache del catálogo y forzar recarga inmediata desde Meta. Evita que
    # el cliente tenga que esperar el TTL (5 min) o reiniciar el server.
    if "META_CATALOG_ID" in saved or "META_ACCESS_TOKEN" in saved:
        try:
            from agent import tools as _tools
            _tools._catalog_cache.pop(agent_id, None)
            _tools._catalog_cache_at.pop(agent_id, None)
            _tools._fb_items.clear()
            _tools._sku_map.pop(agent_id, None)
            asyncio.create_task(_tools._cargar_fb_catalog(agent_id))
            asyncio.create_task(_tools.obtener_catalogo_shopify(agent_id))
            logger.info(f"[config] {saved} actualizado(s) — cache agent_id={agent_id} invalidado, catálogo recargando")
        except Exception as e:
            logger.error(f"[config] error invalidando cache: {e}")

    return JSONResponse(content={"ok": True, "saved": saved})


# ── #48 — Historial + restore de configuración por agente ──────────────────────

# Claves consideradas sensibles — al exponer historial, ofuscamos sus valores
# para no leakear tokens completos en una UI accesible a múltiples usuarios.
# El valor real se preserva en BD intacto, listo para el restore.
_CLAVES_SENSIBLES = {
    "META_ACCESS_TOKEN", "META_CAPI_TOKEN", "META_VERIFY_TOKEN",
    "SHOPIFY_ADMIN_TOKEN", "SHOPIFY_CLIENT_SECRET", "SHOPIFY_WEBHOOK_SECRET",
    "ANTHROPIC_API_KEY", "OPENAI_API_KEY",
}


def _ofuscar_secreto(v: str) -> str:
    """Muestra solo los últimos 4 caracteres. Vacío → '(vacío)'."""
    if not v:
        return "(vacío)"
    if len(v) <= 8:
        return "•" * len(v)
    return "•" * 8 + v[-4:]


@app.get("/inbox/api/config/historial")
async def inbox_historial_config(
    clave: str = "",
    limite: int = 100,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista cambios de configuración del más reciente al más viejo, opcionalmente
    filtrados por clave. Valores de claves sensibles vienen ofuscados (••••XXXX).
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_historial_config
    entries = await obtener_historial_config(
        agent_id=agent_id,
        clave=clave or None,
        limite=max(1, min(limite, 500)),
    )
    # Ofuscar tokens en la respuesta
    for e in entries:
        if e["clave"] in _CLAVES_SENSIBLES:
            e["valor_antes_display"]   = _ofuscar_secreto(e["valor_antes"])
            e["valor_despues_display"] = _ofuscar_secreto(e["valor_despues"])
        else:
            e["valor_antes_display"]   = e["valor_antes"]
            e["valor_despues_display"] = e["valor_despues"]
        # Eliminar valores crudos para no leakearlos a la UI
        e.pop("valor_antes", None)
        e.pop("valor_despues", None)
    return JSONResponse(content={"historial": entries})


@app.post("/inbox/api/config/historial/{historial_id}/restore")
async def inbox_restore_config(
    historial_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Restaura el valor_antes de una entrada de historial. Crea un nuevo entry
    de tipo 'restore' para mantener trazabilidad — no se borra el historial."""
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import restaurar_config_desde_historial
    res = await restaurar_config_desde_historial(
        historial_id=historial_id,
        agent_id=agent_id,
        usuario_id=user.get("id"),
        usuario_email=user.get("email", ""),
    )
    if not res["ok"]:
        return JSONResponse(status_code=404, content={"ok": False, "error": res["error"]})
    # Sincronizar os.environ + invalidar cache catálogo si aplica (mismo pattern
    # que /config/save, sin duplicar logica)
    clave = res["clave"]
    valor_restaurado = await get_config_value(clave, agent_id=agent_id) or ""
    os.environ[clave] = valor_restaurado
    if clave in ("META_CATALOG_ID", "META_ACCESS_TOKEN"):
        try:
            from agent import tools as _tools
            _tools._catalog_cache.pop(agent_id, None)
            _tools._catalog_cache_at.pop(agent_id, None)
            _tools._fb_items.clear()
            _tools._sku_map.pop(agent_id, None)
            asyncio.create_task(_tools._cargar_fb_catalog(agent_id))
        except Exception as e:
            logger.error(f"[config-restore] error invalidando cache: {e}")
    logger.info(f"[config-restore] {clave} restaurado por {user.get('email')} (historial_id={historial_id})")
    return JSONResponse(content={"ok": True, "clave": clave})


@app.post("/inbox/api/config/test/{service}")
async def inbox_test_config(
    service: str,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Prueba la conexión a Meta, Shopify o Anthropic con las credenciales proporcionadas."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body = await request.json()

    async def _val(clave: str) -> str:
        v = (body.get(clave) or "").strip()
        if not v:
            v = await get_config_value(clave, agent_id=agent_id) or ""
        if not v and agent_id == 1:
            v = os.getenv(clave, "")
        # Sanitización para tokens (mismo motivo que en /config/save): caracteres
        # invisibles que se copian de Meta Business rompen httpx en headers.
        if clave in {"META_ACCESS_TOKEN", "META_CAPI_TOKEN", "SHOPIFY_ADMIN_TOKEN",
                     "SHOPIFY_CLIENT_SECRET", "ANTHROPIC_API_KEY"}:
            for ch in ("​", "‌", "‍", " ", "﻿", " ", " "):
                v = v.replace(ch, "")
            try:
                v.encode("ascii")
            except UnicodeEncodeError:
                v = "".join(c for c in v if 32 <= ord(c) < 127)
        return v

    try:
        if service == "meta":
            # El test viejo solo hacía GET /me (basic auth). Eso da OK aunque
            # falten scopes whatsapp_business_*. Equora reportó bug 13-jun:
            # "Probar conexión" decía Conectado pero envíos a /messages
            # fallaban con 400. Ahora validamos los 3 endpoints reales que
            # Voco usa: /me (auth), /{phone_id}? (lectura phone),
            # /{waba_id}/message_templates (scope mgmt).
            access_token = await _val("META_ACCESS_TOKEN")
            phone_id     = await _val("META_PHONE_NUMBER_ID")
            waba_id      = await _val("META_WABA_ID")
            if not access_token:
                return JSONResponse(content={"ok": False, "error": "Token de acceso no configurado"})
            warnings: list[str] = []
            ok_total = True
            nombre = ""
            async with httpx.AsyncClient(timeout=12) as cli:
                # 1) auth básica
                r1 = await cli.get(
                    "https://graph.facebook.com/v21.0/me",
                    headers={"Authorization": f"Bearer {access_token}"},
                )
                if r1.status_code != 200:
                    err = r1.json().get("error", {}).get("message", f"HTTP {r1.status_code}")
                    return JSONResponse(content={"ok": False, "error": f"Token inválido: {err}"})
                nombre = r1.json().get("name") or r1.json().get("id") or "OK"
                # 2) scope whatsapp_business_messaging (para enviar mensajes)
                if phone_id:
                    r2 = await cli.get(
                        f"https://graph.facebook.com/v21.0/{phone_id}",
                        headers={"Authorization": f"Bearer {access_token}"},
                    )
                    if r2.status_code != 200:
                        ok_total = False
                        err = r2.json().get("error", {}).get("message", f"HTTP {r2.status_code}")
                        warnings.append(f"❌ Phone Number ID {phone_id} inaccesible: {err}. Falta scope 'whatsapp_business_messaging' o ID incorrecto.")
                # 3) scope whatsapp_business_management (para plantillas/webhooks)
                if waba_id:
                    r3 = await cli.get(
                        f"https://graph.facebook.com/v21.0/{waba_id}/message_templates?limit=1",
                        headers={"Authorization": f"Bearer {access_token}"},
                    )
                    if r3.status_code != 200:
                        ok_total = False
                        err = r3.json().get("error", {}).get("message", f"HTTP {r3.status_code}")
                        warnings.append(f"❌ WABA {waba_id} inaccesible: {err}. Falta scope 'whatsapp_business_management' o WABA ID incorrecto.")
            if ok_total:
                return JSONResponse(content={"ok": True, "msg": f"✅ Conectado · {nombre} · Phone + WABA + Templates OK"})
            return JSONResponse(content={
                "ok": False,
                "error": "Token responde a /me pero faltan scopes WhatsApp:\n\n" + "\n".join(warnings) +
                         "\n\nRegenera el token en developers.facebook.com → tu app → WhatsApp → API Setup con scopes whatsapp_business_messaging + whatsapp_business_management."
            })

        elif service == "shopify":
            # Test Admin API únicamente (#62 retiró Storefront).
            domain = (await _val("SHOPIFY_STORE")).replace("https://", "").replace("http://", "").rstrip("/")
            if not domain:
                return JSONResponse(content={"ok": False, "error": "Dominio Shopify no configurado"})
            admin_token = await _val("SHOPIFY_ADMIN_TOKEN")
            if not admin_token:
                return JSONResponse(content={"ok": False, "error": "Sin Admin token — completa OAuth con 'Instalar'"})
            from agent.shopify_admin import verificar_admin_token
            resultado = await verificar_admin_token(domain, admin_token)
            if resultado.get("ok"):
                msg = f"✅ Conectado · {resultado.get('shop_name', domain)}"
                if resultado.get("plan"):
                    msg += f" ({resultado['plan']})"
                return JSONResponse(content={"ok": True, "msg": msg, "tipo": "admin"})
            return JSONResponse(content={"ok": False, "error": resultado.get("error", "Error con Admin API")})

        elif service == "anthropic":
            api_key = await _val("ANTHROPIC_API_KEY")
            if not api_key:
                return JSONResponse(content={"ok": False, "error": "API Key no configurada"})
            async with httpx.AsyncClient(timeout=20) as cli:
                r = await cli.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                             "content-type": "application/json"},
                    json={"model": "claude-haiku-4-5", "max_tokens": 5,
                          "messages": [{"role": "user", "content": "hi"}]},
                )
            if r.status_code == 200:
                return JSONResponse(content={"ok": True, "msg": "✅ API key válida · Conexión exitosa"})
            if r.status_code == 401:
                return JSONResponse(content={"ok": False, "error": "API key inválida o revocada"})
            return JSONResponse(content={"ok": False, "error": f"Error {r.status_code}"})

    except Exception as e:
        return JSONResponse(content={"ok": False, "error": str(e)[:120]})

    return JSONResponse(content={"ok": False, "error": "Servicio no reconocido"})


# ══════════════════════════════════════════════════════════════════════════════
# #55 — OAuth Shopify (modelo Jelou/99Envíos)
# ══════════════════════════════════════════════════════════════════════════════
# Flujo OAuth Authorization Code Grant de Shopify:
#   1. Cliente da Client ID + Client Secret + dominio en Voco
#   2. Click 'Conectar con Shopify' → /oauth/shopify/start genera state + URL
#   3. Browser redirige al cliente a Shopify Admin para autorizar
#   4. Shopify redirige a /oauth/shopify/callback con code+shop+hmac+state
#   5. Voco verifica HMAC + state, intercambia code por access_token
#   6. Guarda SHOPIFY_ADMIN_TOKEN en config + redirige al panel
#
# Docs oficiales:
# https://shopify.dev/docs/apps/build/authentication-authorization/access-tokens/authorization-code-grant

# Scopes que Voco necesita. Si se agregan features, agregarlos acá y los
# clientes deben re-autorizar para recibir los nuevos permisos.
SHOPIFY_OAUTH_SCOPES = ",".join([
    # Catálogo (reemplaza Storefront API)
    "read_products",
    "write_products",
    "read_inventory",
    "write_inventory",
    # Pedidos (webhooks + confirmaciones)
    "read_orders",
    # Carrito (reemplaza permalinks /cart/c/)
    "write_draft_orders",
    # Cupones post-venta (feature #43)
    "write_discounts",
])

# State tracking en memoria (anti-CSRF). TTL 10 min.
# { state_token: {"agent_id": int, "shop": str, "user_id": int, "at": float} }
_shopify_oauth_states: dict[str, dict] = {}
_SHOPIFY_OAUTH_STATE_TTL = 600  # 10 min


def _cleanup_shopify_states() -> None:
    """Limpia states expirados (>10 min). Llamado en cada start/callback."""
    ahora = time.time()
    expirados = [s for s, info in _shopify_oauth_states.items()
                 if ahora - info.get("at", 0) > _SHOPIFY_OAUTH_STATE_TTL]
    for s in expirados:
        _shopify_oauth_states.pop(s, None)


def _shopify_oauth_verify_hmac(query_params: dict, client_secret: str) -> bool:
    """Verifica HMAC del callback de Shopify para asegurar que viene de Shopify
    y no fue tampered. Algoritmo según docs:
      1. Quitar 'hmac' y 'signature' del query
      2. Ordenar parámetros alfabéticamente
      3. Concatenar key=value separados por &
      4. Calcular HMAC-SHA256 con client_secret como key
      5. Comparar con el hmac que llegó
    """
    recibido = query_params.get("hmac", "")
    if not recibido or not client_secret:
        return False
    filtrado = {k: v for k, v in query_params.items() if k not in ("hmac", "signature")}
    mensaje = "&".join(f"{k}={v}" for k, v in sorted(filtrado.items()))
    esperado = hmac.new(
        client_secret.encode("utf-8"),
        mensaje.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(esperado, recibido)


def _shopify_callback_url(request: Request) -> str:
    """Construye la URL de callback OAuth a partir del host actual.
    Multi-tenant: funciona en Railway, dev local y futuro dominio propio."""
    base = str(request.base_url).rstrip("/")
    if base.startswith("http://") and "localhost" not in base:
        base = base.replace("http://", "https://", 1)
    return f"{base}/oauth/shopify/callback"


@app.post("/inbox/api/oauth/shopify/start")
async def api_shopify_oauth_start(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Inicia el flujo OAuth. Cliente debe haber guardado SHOPIFY_STORE +
    SHOPIFY_CLIENT_ID + SHOPIFY_CLIENT_SECRET en Configuración antes.

    Response: {ok, auth_url} — el frontend hace window.location = auth_url
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")

    try:
        body = await request.json()
    except Exception:
        body = {}

    # Permitir override desde el body (para 'probar antes de guardar') o leer del config
    from agent.memory import get_config_value
    # Para v1 usamos agent_id=1; con #45 (selector) el frontend podría pasarlo
    agent_id = int(body.get("agent_id") or 1)

    store = (body.get("store") or "").strip() or (
        await get_config_value("SHOPIFY_STORE", agent_id)
        or os.getenv("SHOPIFY_STORE", "")
    )
    client_id = (body.get("client_id") or "").strip() or (
        await get_config_value("SHOPIFY_CLIENT_ID", agent_id) or ""
    )
    client_secret = (body.get("client_secret") or "").strip() or (
        await get_config_value("SHOPIFY_CLIENT_SECRET", agent_id) or ""
    )
    store = store.replace("https://", "").replace("http://", "").rstrip("/")

    if not store or not client_id or not client_secret:
        return JSONResponse(status_code=400, content={
            "ok": False,
            "error": "Configura primero: Dominio + Client ID + Client Secret (guardar) antes de conectar.",
        })
    if not store.endswith(".myshopify.com"):
        return JSONResponse(status_code=400, content={
            "ok": False,
            "error": f"El dominio debe terminar en .myshopify.com (recibido: {store}). Usa el dominio interno, no el personalizado.",
        })

    # Generar state token aleatorio + guardar contexto
    _cleanup_shopify_states()
    state = secrets.token_urlsafe(32)
    # panel_path = la URL del panel del agente desde donde se inició el OAuth.
    # El callback la usa para redirigir de vuelta al panel correcto (no al
    # selector raíz /inbox). Validamos que empiece con / para evitar open
    # redirect a dominios externos.
    panel_path = (body.get("panel_path") or "").strip()
    if not panel_path.startswith("/"):
        panel_path = "/inbox"
    _shopify_oauth_states[state] = {
        "agent_id":      agent_id,
        "shop":          store,
        "client_id":     client_id,
        "client_secret": client_secret,
        "user_id":       user.get("id"),
        "panel_path":    panel_path,
        "at":            time.time(),
    }

    callback = _shopify_callback_url(request)
    auth_url = (
        f"https://{store}/admin/oauth/authorize"
        f"?client_id={client_id}"
        f"&scope={SHOPIFY_OAUTH_SCOPES}"
        f"&redirect_uri={callback}"
        f"&state={state}"
    )
    logger.info(f"[shopify-oauth] start agent={agent_id} shop={store} (callback={callback})")
    return JSONResponse(content={"ok": True, "auth_url": auth_url, "callback_url": callback})


@app.get("/oauth/shopify/callback")
async def shopify_oauth_callback(request: Request):
    """Callback OAuth de Shopify. NO requiere auth de Voco — viene del browser
    del cliente redirigido desde Shopify. Validamos HMAC + state para anti-CSRF.

    Recibe: ?code=X&shop=Y&hmac=Z&state=W&timestamp=N
    Devuelve: redirect al panel con ?shopify_oauth=ok|error&msg=...
    """
    qp = dict(request.query_params)
    code  = qp.get("code", "")
    shop  = qp.get("shop", "")
    state = qp.get("state", "")
    hmac_recibido = qp.get("hmac", "")

    # panel_path se resuelve desde el state cuando esté disponible (más abajo).
    # Si no, fallback a /inbox (selector raíz).
    panel_path_default = "/inbox"

    def _redirect_panel(estado: str, mensaje: str = "", path: str | None = None) -> RedirectResponse:
        import urllib.parse as _up
        msg = _up.quote(mensaje[:200])
        target = path or panel_path_default
        return RedirectResponse(
            f"{target}#configuracion?shopify_oauth={estado}&msg={msg}",
            status_code=302,
        )

    if not code or not shop or not state or not hmac_recibido:
        return _redirect_panel("error", "Parámetros incompletos del callback")

    # Recuperar contexto del state (anti-CSRF)
    _cleanup_shopify_states()
    info = _shopify_oauth_states.pop(state, None)
    if not info:
        return _redirect_panel("error", "State inválido o expirado. Reintenta el OAuth desde el panel.")

    # Shop mismatch tolerante: cuando una tienda tiene múltiples dominios
    # .myshopify.com (ej. equora-6 público + kbetje-6y interno canónico),
    # Shopify SIEMPRE devuelve el canónico interno en el callback aunque
    # el user haya iniciado con el alias público. Aceptamos el shop que
    # Shopify devuelve si es .myshopify.com válido — la seguridad la da
    # el HMAC (firmado con el client_secret de ESA app), no el dominio.
    if info["shop"] != shop:
        if not shop.endswith(".myshopify.com"):
            return _redirect_panel("error", f"Dominio inválido recibido de Shopify: {shop}")
        logger.info(
            f"[shopify-oauth] dominio canónico distinto al iniciado: "
            f"iniciado={info['shop']} canónico={shop} — adoptando canónico"
        )
    # Usar el panel_path del state para que el toast aterrice en el panel
    # del agente correcto (no en el selector raíz).
    panel_path_default = info.get("panel_path") or "/inbox"

    # Verificar HMAC (firma del client_secret)
    if not _shopify_oauth_verify_hmac(qp, info["client_secret"]):
        logger.warning(f"[shopify-oauth] HMAC inválido para shop={shop}")
        return _redirect_panel("error", "Firma HMAC inválida. Verifica el Client Secret en Configuración.")

    # Intercambiar code por access_token
    try:
        async with httpx.AsyncClient(timeout=15.0) as cli:
            r = await cli.post(
                f"https://{shop}/admin/oauth/access_token",
                json={
                    "client_id":     info["client_id"],
                    "client_secret": info["client_secret"],
                    "code":          code,
                },
            )
        if r.status_code != 200:
            logger.error(f"[shopify-oauth] intercambio code→token HTTP {r.status_code}: {r.text[:300]}")
            return _redirect_panel("error", f"Shopify rechazó el code (HTTP {r.status_code}). Reintenta.")
        access_token = (r.json() or {}).get("access_token", "")
        if not access_token:
            return _redirect_panel("error", "Shopify no devolvió access_token. Reintenta.")
    except Exception as e:
        logger.error(f"[shopify-oauth] error intercambiando code: {e}")
        return _redirect_panel("error", f"Error de red: {e!s}")

    # Guardar el token + dominio en config del agente
    try:
        from agent.memory import set_config_value
        agent_id = info["agent_id"]
        await set_config_value("SHOPIFY_STORE",       shop,         agent_id)
        await set_config_value("SHOPIFY_ADMIN_TOKEN", access_token, agent_id)
        logger.info(f"[shopify-oauth] ✅ token guardado para shop={shop} agent={agent_id}")
    except Exception as e:
        logger.error(f"[shopify-oauth] error guardando token: {e}")
        return _redirect_panel("error", f"OAuth OK pero falló al guardar: {e!s}")

    # Auto-registrar webhooks de Voco (orders/create, orders/paid, etc.).
    # Best-effort: si falla, el cliente igual queda conectado y puede reintentar
    # desde el panel — no abortamos el OAuth por esto.
    try:
        from agent.shopify_admin import sincronizar_webhooks_voco
        callback_webhook = f"{str(request.base_url).rstrip('/').replace('http://', 'https://', 1)}/shopify-webhook"
        res = await sincronizar_webhooks_voco(shop, access_token, callback_webhook)
        creados = len(res.get("creados", []))
        existentes = len(res.get("conservados", []))
        fallidos = len(res.get("fallidos", []))
        logger.info(f"[shopify-oauth] webhooks auto-registrados: {creados} nuevos, {existentes} existentes, {fallidos} fallidos")
        if fallidos:
            logger.warning(f"[shopify-oauth] webhooks fallidos: {res.get('fallidos')}")
    except Exception as e:
        logger.error(f"[shopify-oauth] no se pudieron auto-registrar webhooks (no es bloqueante): {e}")

    return _redirect_panel("ok", f"Conectado con {shop}. Tu agente ya tiene acceso a tu catálogo y pedidos.")


# ══════════════════════════════════════════════════════════════════════════════
# #54 — Sincronización programática de webhooks Shopify
# ══════════════════════════════════════════════════════════════════════════════
@app.post("/inbox/api/integraciones/shopify/sincronizar-webhooks")
async def api_shopify_sincronizar_webhooks(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Registra (o actualiza) los webhooks que Voco necesita en la tienda
    Shopify del agente, vía Admin API. Reemplaza el copy-paste manual del
    cliente en Shopify Admin → Notifications → Webhooks.

    Body opcional:
        {"store": "...", "admin_token": "..."}
    Si no se pasan, los lee del config_value/env del agente activo.

    Response:
        {ok, creados, conservados, recreados, fallidos, callback_url, error?}
    """
    user = await _obtener_sesion_usuario(voco_session or inbox_session or token)
    if not user:
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        body = await request.json()
    except Exception:
        body = {}

    # Resolver desde 3 fuentes (en orden): body explícito > config_value del
    # agente (BD, donde OAuth lo guarda) > env var (legacy). Esto soporta:
    # · Cliente probando sin guardar primero (body)
    # · Cliente ya conectado vía OAuth (config_value)
    # · Setup viejo con env vars de Railway (legacy)
    from agent.memory import get_config_value
    agent_id_int = int(body.get("agent_id") or 1)
    store = (body.get("store") or "").strip()
    if not store:
        store = (await get_config_value("SHOPIFY_STORE", agent_id_int)) or os.getenv("SHOPIFY_STORE", "")
    admin_token = (body.get("admin_token") or "").strip()
    if not admin_token:
        admin_token = (await get_config_value("SHOPIFY_ADMIN_TOKEN", agent_id_int)) or os.getenv("SHOPIFY_ADMIN_TOKEN", "")
    store = store.replace("https://", "").replace("http://", "").rstrip("/")
    if not store or not admin_token:
        return JSONResponse(status_code=400, content={
            "ok": False,
            "error": "Falta dominio Shopify o Admin API token. Conéctate primero con 'Conectar con Shopify' o configura el token en modo avanzado."
        })

    # Construir el callback URL desde el host de la request (multi-tenant ready
    # — funciona con cualquier dominio donde esté desplegado Voco)
    base = str(request.base_url).rstrip("/")
    if base.startswith("http://") and "localhost" not in base:
        # Shopify exige HTTPS para webhooks — forzar si no estamos en local
        base = base.replace("http://", "https://", 1)
    callback_url = f"{base}/shopify-webhook"

    from agent.shopify_admin import sincronizar_webhooks_voco
    res = await sincronizar_webhooks_voco(store, admin_token, callback_url)
    res["callback_url"] = callback_url
    if res.get("ok"):
        logger.info(
            f"[shopify-admin] webhooks sync para {store}: "
            f"+{len(res.get('creados', []))} ={len(res.get('conservados', []))} "
            f"~{len(res.get('recreados', []))} !{len(res.get('fallidos', []))}"
        )
    return JSONResponse(content=res)


# ══════════════════════════════════════════════════════════════════════════════
# PROMPT EDITOR — leer, guardar y mejorar el system prompt con IA
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/inbox/api/prompt")
async def inbox_get_prompt(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el prompt actual (BD → archivo) y las variables del negocio."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    import yaml as _yaml

    # Prompt: BD primero, luego archivo, luego brain.py (fallback final para agentes existentes)
    db_prompt = await get_config_value("SYSTEM_PROMPT", agent_id)
    if db_prompt:
        prompt  = db_prompt
        fuente  = "db"
    else:
        try:
            with open("config/prompts.yaml", "r", encoding="utf-8") as f:
                cfg = _yaml.safe_load(f) or {}
            prompt = cfg.get("system_prompt", "")
        except FileNotFoundError:
            prompt = ""
        fuente = "file"

    # Si sigue vacío, intentar cargar desde brain.py (captura prompts.yaml con cache y fallback)
    if not prompt:
        try:
            from agent.brain import cargar_system_prompt as _csp
            raw = await _csp(agent_id)
            # Solo usar si no es el prompt genérico por defecto
            if raw and raw != "Eres un asistente virtual. Responde en español.":
                prompt = raw
                fuente = "brain"
        except Exception:
            pass

    # Variables del negocio (JSON guardado en BD)
    vars_json = await get_config_value("BUSINESS_VARS", agent_id)
    try:
        business_vars = json.loads(vars_json) if vars_json else {}
    except Exception:
        business_vars = {}

    # Tipo de negocio
    business_type = await get_config_value("BUSINESS_TYPE", agent_id) or "productos"

    # Módulos activos
    modules_json = await get_config_value("ACTIVE_MODULES", agent_id)
    try:
        active_modules = json.loads(modules_json) if modules_json else {}
    except Exception:
        active_modules = {}

    return JSONResponse(content={
        "prompt":         prompt,
        "business_vars":  business_vars,
        "fuente":         fuente,
        "business_type":  business_type,
        "active_modules": active_modules,
    })


@app.post("/inbox/api/prompt/save")
async def inbox_save_prompt(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Guarda el prompt y las variables del negocio en la BD."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body = await request.json()
    prompt        = (body.get("prompt") or "").strip()
    business_vars = body.get("business_vars") or {}
    business_type = (body.get("business_type") or "").strip()
    active_modules = body.get("active_modules")

    if prompt:
        await set_config_value("SYSTEM_PROMPT", prompt, agent_id)
    if isinstance(business_vars, dict):
        await set_config_value("BUSINESS_VARS", json.dumps(business_vars, ensure_ascii=False), agent_id)
    if business_type:
        await set_config_value("BUSINESS_TYPE", business_type, agent_id)
    if isinstance(active_modules, dict):
        await set_config_value("ACTIVE_MODULES", json.dumps(active_modules), agent_id)

    return JSONResponse(content={"ok": True})


@app.post("/inbox/api/prompt/improve")
async def inbox_improve_prompt(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """
    Recibe el prompt actual + una instrucción del usuario en lenguaje natural
    y devuelve una versión mejorada generada por Claude.
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body        = await request.json()
    prompt      = (body.get("prompt") or "").strip()
    instruccion = (body.get("instruccion") or "").strip()
    vars_dict   = body.get("business_vars") or {}
    imagenes    = body.get("imagenes") or []   # [{media_type, data}]

    if not prompt:
        return JSONResponse(content={"ok": False, "error": "Prompt vacío"})
    if not instruccion:
        return JSONResponse(content={"ok": False, "error": "Escribe qué quieres mejorar"})

    # Contexto de variables para que Claude las conserve
    vars_context = ""
    if vars_dict:
        vars_lines = "\n".join(f"  {{{k}}} = {v}" for k, v in vars_dict.items())
        vars_context = f"\n\nVariables del negocio definidas (CONSERVARLAS intactas en el prompt):\n{vars_lines}"

    system_meta = (
        "Eres un experto en diseño de system prompts para agentes de IA conversacionales "
        "que operan por WhatsApp. Tu tarea es mejorar el prompt de acuerdo a la instrucción "
        "del usuario.\n\n"
        "REGLAS ABSOLUTAS:\n"
        "- Conserva TODA la información existente; solo mejora, agrega o ajusta lo indicado\n"
        "- Conserva todas las variables {EN_MAYUSCULAS} exactamente como aparecen\n"
        "- Responde ÚNICAMENTE con el prompt mejorado — sin explicaciones, sin markdown extra, "
        "sin texto antes o después\n"
        "- Mantén el idioma español\n"
        "- El prompt debe ser claro, específico y directamente accionable"
    )

    texto_usuario = (
        f"PROMPT ACTUAL:\n{prompt}"
        f"{vars_context}\n\n"
        f"INSTRUCCIÓN DEL USUARIO: {instruccion}"
    )

    # Construir content blocks — texto + imágenes opcionales
    content_blocks: list = []
    for img in imagenes[:5]:   # máximo 5 imágenes
        media_type = img.get("media_type", "image/jpeg")
        data       = img.get("data", "")
        if not data:
            continue
        content_blocks.append({
            "type": "image",
            "source": {
                "type":       "base64",
                "media_type": media_type,
                "data":       data,
            }
        })
    content_blocks.append({"type": "text", "text": texto_usuario})

    from anthropic import AsyncAnthropic as _Anthropic
    _client = _Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY", ""))
    # Usar modelo con visión si hay imágenes; haiku no soporta visión en todas las versiones
    modelo = os.getenv("AI_MODEL", "claude-haiku-4-5")
    if imagenes:
        modelo = "claude-sonnet-4-5"   # garantiza soporte de visión
    try:
        resp = await _client.messages.create(
            model=modelo,
            max_tokens=4096,
            system=system_meta,
            messages=[{"role": "user", "content": content_blocks}],
        )
        improved = resp.content[0].text.strip()
        return JSONResponse(content={"ok": True, "improved_prompt": improved})
    except Exception as e:
        logger.error(f"Error mejorando prompt: {e}")
        return JSONResponse(content={"ok": False, "error": str(e)[:200]})


# ══════════════════════════════════════════════════════════════════════════════
# CHAT DE PRUEBA — simula conversación con el agente desde el panel
# ══════════════════════════════════════════════════════════════════════════════

def _test_phone(agent_id: int) -> str:
    return f"__test_inbox_{agent_id}__"


@app.post("/inbox/api/chat/test")
async def inbox_chat_test(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Envía un mensaje al agente y devuelve su respuesta (sin WhatsApp real)."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    body    = await request.json()
    mensaje = (body.get("mensaje") or "").strip()
    if not mensaje:
        return JSONResponse(status_code=400, content={"error": "Mensaje vacío"})

    tel = _test_phone(agent_id)
    historial = await obtener_historial(tel)
    try:
        respuesta_raw = await generar_respuesta(mensaje, historial, telefono=None, contexto_campana=None, agent_id=agent_id)
    except Exception as e:
        logger.error(f"Error en chat de prueba: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)[:200]})

    # Procesar marcadores para extraer [[ESCALAR:...]] y otros antes de guardar
    ctx = MarkerContext(respuesta=respuesta_raw, telefono=tel, agent_id=agent_id)
    ctx = await aplicar_marcadores(ctx)
    respuesta = ctx.respuesta

    await guardar_mensaje(tel, "user", mensaje, agent_id=agent_id)
    await guardar_mensaje(tel, "assistant", respuesta, agent_id=agent_id)

    # Crear ticket de escalación si el agente lo solicitó
    escalacion = False
    if ctx.datos_escalacion:
        try:
            await _notificar_escalacion(tel, ctx.datos_escalacion, agent_id=agent_id)
            escalacion = True
        except Exception as e:
            logger.error(f"Error creando escalación de chat de prueba: {e}")

    return JSONResponse(content={"ok": True, "respuesta": respuesta, "escalacion": escalacion})


@app.get("/inbox/api/chat/test/history")
async def inbox_chat_test_history(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Devuelve el historial de la conversación de prueba."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    mensajes = await obtener_historial_con_timestamps(_test_phone(agent_id), 100, agent_id=agent_id)
    return JSONResponse(content={"mensajes": mensajes})


@app.delete("/inbox/api/chat/test/clear")
async def inbox_chat_test_clear(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Borra el historial de la conversación de prueba."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    await limpiar_historial(_test_phone(agent_id))
    return JSONResponse(content={"ok": True})


@app.post("/inbox/plantillas/subir-header")
async def inbox_subir_header_media(
    file: UploadFile = File(...),
    file_type: str = Form(...),
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """
    Sube un archivo (imagen/video/doc) a Meta vía Resumable Upload API
    y retorna el handle para usar en example.header_handle al crear una plantilla.

    Flujo Meta Resumable Upload:
      1. POST /{app_id}/uploads?file_name=...&file_length=...&file_type=...
         → {"id": "upload:{session_id}"}
      2. POST /upload:{session_id}  (body = bytes del archivo, header file_offset: 0)
         → {"h": "handle_string"}
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = os.getenv("META_ACCESS_TOKEN", "")
    app_id       = os.getenv("META_APP_ID", "")

    if not access_token:
        return JSONResponse(content={"error": "META_ACCESS_TOKEN no configurado"}, status_code=500)
    if not app_id:
        # Sin APP_ID no podemos usar Resumable Upload — informar al frontend
        # (la plantilla se crea sin example header, Meta igual la aprueba)
        return JSONResponse(content={"error": "META_APP_ID no configurado en Railway — la plantilla se creará sin imagen de ejemplo", "handle": None}, status_code=200)

    try:
        file_bytes = await file.read()
        file_name  = file.filename or "header_file"
        file_size  = len(file_bytes)
        mime_type  = file_type or file.content_type or "image/jpeg"

        api_ver = "v21.0"

        async with httpx.AsyncClient(timeout=60) as client:
            # 1. Iniciar sesión de upload
            init_url = (
                f"https://graph.facebook.com/{api_ver}/{app_id}/uploads"
                f"?file_name={file_name}&file_length={file_size}"
                f"&file_type={mime_type}&access_token={access_token}"
            )
            r1 = await client.post(init_url)
            data1 = r1.json()
            if "error" in data1:
                err = data1["error"]
                logger.error(f"[subir-header] Error iniciando upload: {err}")
                return JSONResponse(content={"error": err.get("message", str(err)), "handle": None})

            session_id = data1.get("id", "")
            if not session_id:
                return JSONResponse(content={"error": "No se obtuvo upload session_id de Meta", "handle": None})

            # 2. Subir el archivo
            upload_url = f"https://graph.facebook.com/{api_ver}/{session_id}"
            upload_headers = {
                "Authorization": f"OAuth {access_token}",
                "file_offset": "0",
                "Content-Type": mime_type,
            }
            r2 = await client.post(upload_url, content=file_bytes, headers=upload_headers)
            data2 = r2.json()
            if "error" in data2:
                err = data2["error"]
                logger.error(f"[subir-header] Error subiendo archivo: {err}")
                return JSONResponse(content={"error": err.get("message", str(err)), "handle": None})

            handle = data2.get("h")
            if not handle:
                return JSONResponse(content={"error": "Meta no devolvió un handle válido", "handle": None})

            logger.info(f"[subir-header] Archivo '{file_name}' subido → handle={handle[:30]}...")
            return JSONResponse(content={"ok": True, "handle": handle})

    except Exception as e:
        logger.error(f"[subir-header] Excepción: {e}", exc_info=True)
        return JSONResponse(content={"error": str(e), "handle": None}, status_code=500)


@app.post("/inbox/plantillas/crear")
async def inbox_plantillas_crear(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Crea una nueva plantilla en Meta y la envía a revisión.
    Body JSON: { name, category, language, header_text?, body, footer?, buttons? }
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = os.getenv("META_ACCESS_TOKEN", "")
    waba_id      = os.getenv("META_WABA_ID", "")
    if not access_token or not waba_id:
        return JSONResponse(content={"error": "META_ACCESS_TOKEN o META_WABA_ID no configurados"}, status_code=500)

    body = await request.json()
    nombre       = (body.get("name") or "").lower().replace(" ", "_")
    categoria    = body.get("category", "MARKETING")
    idioma       = body.get("language", "es_CO")
    cuerpo       = body.get("body", "")
    header_tipo   = (body.get("header_type") or "").upper()  # NONE|TEXT|IMAGE|VIDEO|DOCUMENT|LOCATION
    header_texto  = body.get("header_text", "")
    header_handle = body.get("header_handle")                # handle de Resumable Upload
    footer        = body.get("footer", "")
    buttons       = body.get("buttons", [])
    sub_category  = body.get("sub_category", "DEFAULT")      # DEFAULT|CATALOG_MESSAGE|CALL_PERMISSION_REQUEST
    catalog_format = body.get("catalog_format", "FULL")      # FULL | MULTI (solo para CATALOG_MESSAGE)
    var_type      = (body.get("var_type") or "").upper()     # NOMBRE | NUMERO (controla ejemplo de variables)
    ttl_activo    = body.get("ttl_activo", False)
    ttl_secs      = int(body.get("ttl", 43200))              # período de validez en segundos
    # Ubicación: campos de ejemplo para LOCATION header
    loc_lat  = body.get("loc_lat")
    loc_lng  = body.get("loc_lng")
    loc_name = body.get("loc_name", "")

    if not nombre or not cuerpo:
        return JSONResponse(content={"error": "Nombre y cuerpo son requeridos"}, status_code=400)

    componentes = []

    # ── HEADER ──
    # CATALOG_MESSAGE y CALL_PERMISSION_REQUEST no admiten encabezado multimedia
    if sub_category not in ("CATALOG_MESSAGE", "CALL_PERMISSION_REQUEST"):
        if header_tipo == "TEXT" and header_texto:
            componentes.append({"type": "HEADER", "format": "TEXT", "text": header_texto[:60]})
        elif header_tipo in ("IMAGE", "VIDEO", "DOCUMENT"):
            comp_hdr: dict = {"type": "HEADER", "format": header_tipo}
            if header_handle:
                comp_hdr["example"] = {"header_handle": [header_handle]}
            componentes.append(comp_hdr)
        elif header_tipo == "LOCATION":
            comp_hdr = {"type": "HEADER", "format": "LOCATION"}
            if loc_lat and loc_lng:
                comp_hdr["example"] = {
                    "header_location": {
                        "latitude": str(loc_lat),
                        "longitude": str(loc_lng),
                        "name": loc_name or "Ubicación de ejemplo",
                    }
                }
            componentes.append(comp_hdr)

    # ── BODY ──
    import re as _re
    vars_encontradas = _re.findall(r'\{\{([^}]+)\}\}', cuerpo)
    body_comp: dict = {"type": "BODY", "text": cuerpo}
    if vars_encontradas:
        example_vals = ["ejemplo"] * len(vars_encontradas)
        # var_type=NOMBRE fuerza named params aunque el usuario use {{1}} por error
        # var_type=NUMERO fuerza posicional aunque el usuario use {{nombre}}
        # Si no se especifica, auto-detectar
        named_vars = [v for v in vars_encontradas if not v.isdigit()]
        usar_named = (var_type == "NOMBRE") or (var_type != "NUMERO" and bool(named_vars))
        if usar_named:
            body_comp["example"] = {
                "body_text_named_params": [
                    {"param_name": v, "example": ["ejemplo"]} for v in vars_encontradas
                ]
            }
        else:
            body_comp["example"] = {"body_text": [example_vals]}
    componentes.append(body_comp)

    # ── FOOTER ──
    if footer:
        componentes.append({"type": "FOOTER", "text": footer[:60]})

    # ── BUTTONS ──
    if sub_category == "CATALOG_MESSAGE":
        # Catálogo: solo un botón fijo de tipo CATALOG (Meta lo muestra como "Ver catálogo")
        catalog_btn: dict = {"type": "CATALOG", "title": "Ver catálogo"}
        componentes.append({"type": "BUTTONS", "buttons": [catalog_btn]})
    elif sub_category == "CALL_PERMISSION_REQUEST":
        # Permisos de llamada: Meta genera los botones automáticamente, no enviamos ninguno
        pass
    elif buttons:
        btn_list = []
        for b in buttons[:10]:
            btype = b.get("type", "").upper()
            texto_btn = b.get("text", "")[:25]
            if btype == "URL":
                url_val = b.get("url", "")
                if url_val and texto_btn:
                    btn_obj = {"type": "URL", "text": texto_btn[:20], "url": url_val}
                    if "{{1}}" in url_val:
                        btn_obj["example"] = [url_val.replace("{{1}}", "ejemplo")]
                    btn_list.append(btn_obj)
            elif btype == "PHONE_NUMBER":
                phone = b.get("phone_number", "")
                if phone and texto_btn:
                    btn_list.append({"type": "PHONE_NUMBER", "text": texto_btn[:20], "phone_number": phone})
            elif btype == "QUICK_REPLY":
                if texto_btn:
                    btn_list.append({"type": "QUICK_REPLY", "text": texto_btn})
        if btn_list:
            componentes.append({"type": "BUTTONS", "buttons": btn_list})

    payload = {
        "name": nombre,
        "category": categoria,
        "language": idioma,
        "components": componentes,
    }
    # Subcategoría (solo si no es DEFAULT para evitar errores en Auth)
    if sub_category and sub_category != "DEFAULT":
        payload["sub_category"] = sub_category
    # Período de validez (solo Marketing)
    if ttl_activo and categoria == "MARKETING" and ttl_secs:
        payload["message_send_ttl_seconds"] = ttl_secs

    api_ver = "v21.0"
    url = f"https://graph.facebook.com/{api_ver}/{waba_id}/message_templates"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

    logger.info(f"[plantillas] Enviando a Meta — nombre={nombre} cat={categoria} sub={sub_category} componentes={len(componentes)}")
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(url, json=payload, headers=headers)
            # Parsear respuesta de Meta — puede no ser JSON en casos de error
            try:
                data = r.json()
            except Exception as json_err:
                raw = r.text[:600]
                logger.error(f"[plantillas] Meta devolvió respuesta no-JSON (status={r.status_code}): {raw}")
                return JSONResponse(
                    content={"error": f"Meta devolvió respuesta no válida (status {r.status_code}): {raw}"},
                    status_code=502,
                )
            if not isinstance(data, dict):
                logger.error(f"[plantillas] Respuesta inesperada de Meta (tipo={type(data).__name__}): {str(data)[:300]}")
                return JSONResponse(
                    content={"error": f"Respuesta inesperada de Meta: {str(data)[:200]}"},
                    status_code=502,
                )
            if r.status_code in (200, 201) or "id" in data:
                logger.info(f"[plantillas] Plantilla '{nombre}' creada — id={data.get('id')} status={data.get('status')}")
                return JSONResponse(content={"ok": True, "id": data.get("id"), "status": data.get("status")})
            else:
                err = data.get("error", {})
                err_msg = err.get("message", str(data)) if isinstance(err, dict) else str(err)
                err_details = err.get("error_data", {}) if isinstance(err, dict) else {}
                logger.error(f"[plantillas] Error de Meta (status={r.status_code} code={err.get('code') if isinstance(err,dict) else '?'}): {err_msg} | details={err_details}")
                return JSONResponse(content={"error": err_msg, "details": err_details}, status_code=400)
    except Exception as e:
        logger.error(f"[plantillas] Excepción ({type(e).__name__}): {e}", exc_info=True)
        return JSONResponse(content={"error": str(e) or type(e).__name__}, status_code=500)


# ── Borradores de plantillas (guardado local antes de enviar a Meta) ─────────

@app.get("/inbox/plantillas/borradores")
async def inbox_plantillas_borradores_list(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Lista todos los borradores de plantillas guardados localmente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    borradores = await obtener_borradores_plantillas()
    return JSONResponse(content={"borradores": borradores})


@app.post("/inbox/plantillas/borrador")
async def inbox_plantillas_borrador_guardar(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Guarda (crea o actualiza) un borrador de plantilla localmente.
    Body JSON: los mismos campos del formulario de creación.
    Si ya existe un borrador con el mismo nombre, lo sobreescribe.
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    body = await request.json()
    nombre    = (body.get("name") or "").strip()
    categoria = body.get("category", "MARKETING")
    idioma    = body.get("language", "es_CO")
    if not nombre:
        return JSONResponse(content={"error": "El nombre es requerido"}, status_code=400)
    bid = await guardar_borrador_plantilla(nombre, categoria, idioma, body)
    logger.info(f"[plantillas] Borrador guardado id={bid} nombre={nombre}")
    return JSONResponse(content={"ok": True, "id": bid})


@app.delete("/inbox/plantillas/borrador/{bid}")
async def inbox_plantillas_borrador_eliminar(
    bid: int,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Elimina un borrador local por ID."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    eliminado = await eliminar_borrador_plantilla(bid)
    if not eliminado:
        raise HTTPException(status_code=404, detail="Borrador no encontrado")
    return JSONResponse(content={"ok": True})


@app.post("/inbox/plantillas/editar")
async def inbox_plantillas_editar(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Edita una plantilla existente en Meta (solo se pueden cambiar los componentes).
    Meta devuelve la plantilla a estado PENDING tras la edición.
    Body JSON: { template_id, header_type?, header_text?, header_handle?,
                 body, footer?, buttons? }
    """
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    access_token = os.getenv("META_ACCESS_TOKEN", "")
    if not access_token:
        return JSONResponse(content={"error": "META_ACCESS_TOKEN no configurado"}, status_code=500)

    body = await request.json()
    template_id   = body.get("template_id", "")
    cuerpo        = body.get("body", "")
    header_tipo   = (body.get("header_type") or "").upper()
    header_texto  = body.get("header_text", "")
    header_handle = body.get("header_handle")
    footer        = body.get("footer", "")
    buttons       = body.get("buttons", [])

    if not template_id or not cuerpo:
        return JSONResponse(content={"error": "template_id y body son requeridos"}, status_code=400)

    componentes = []

    # HEADER
    if header_tipo == "TEXT" and header_texto:
        componentes.append({"type": "HEADER", "format": "TEXT", "text": header_texto[:60]})
    elif header_tipo in ("IMAGE", "VIDEO", "DOCUMENT"):
        comp_hdr: dict = {"type": "HEADER", "format": header_tipo}
        if header_handle:
            comp_hdr["example"] = {"header_handle": [header_handle]}
        componentes.append(comp_hdr)

    # BODY
    import re as _re2
    vars_encontradas = _re2.findall(r'\{\{([^}]+)\}\}', cuerpo)
    body_comp: dict = {"type": "BODY", "text": cuerpo}
    if vars_encontradas:
        example_vals = ["ejemplo"] * len(vars_encontradas)
        named_vars = [v for v in vars_encontradas if not v.isdigit()]
        if named_vars:
            body_comp["example"] = {
                "body_text_named_params": [
                    {"param_name": v, "example": ["ejemplo"]} for v in vars_encontradas
                ]
            }
        else:
            body_comp["example"] = {"body_text": [example_vals]}
    componentes.append(body_comp)

    # FOOTER
    if footer:
        componentes.append({"type": "FOOTER", "text": footer[:60]})

    # BUTTONS
    if buttons:
        btn_list = []
        for b in buttons[:10]:
            btype = b.get("type", "").upper()
            texto_btn = b.get("text", "")[:25]
            if btype == "URL":
                url_val = b.get("url", "")
                if url_val and texto_btn:
                    btn_obj = {"type": "URL", "text": texto_btn[:20], "url": url_val}
                    if "{{1}}" in url_val:
                        btn_obj["example"] = [url_val.replace("{{1}}", "ejemplo")]
                    btn_list.append(btn_obj)
            elif btype == "PHONE_NUMBER":
                phone = b.get("phone_number", "")
                if phone and texto_btn:
                    btn_list.append({"type": "PHONE_NUMBER", "text": texto_btn[:20], "phone_number": phone})
            elif btype == "QUICK_REPLY":
                if texto_btn:
                    btn_list.append({"type": "QUICK_REPLY", "text": texto_btn})
        if btn_list:
            componentes.append({"type": "BUTTONS", "buttons": btn_list})

    api_ver = "v21.0"
    url     = f"https://graph.facebook.com/{api_ver}/{template_id}"
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(url, json={"components": componentes}, headers=headers)
            data = r.json()
            if r.status_code == 200 and data.get("success"):
                logger.info(f"[plantillas] Plantilla editada id={template_id}")
                return JSONResponse(content={"ok": True})
            else:
                err = data.get("error", {})
                logger.error(f"[plantillas] Error editando plantilla: {err}")
                return JSONResponse(content={"error": err.get("message", str(data))}, status_code=400)
    except Exception as e:
        logger.error(f"[plantillas] Excepción editando: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# ══════════════════════════════════════════════════════════════════════════════
# CATÁLOGO NATIVO VOCO
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/inbox/api/catalogo")
async def api_catalogo_listar(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    from agent.tools import _resolver_admin_token, _resolver_store, obtener_catalogo_shopify, obtener_catalogo_json
    admin_tok  = await _resolver_admin_token(agent_id)
    store_dom  = await _resolver_store(agent_id)
    tiene_shopify = bool(admin_tok and store_dom)

    if tiene_shopify:
        # Forzar carga si el cache está vacío (primer acceso al panel)
        items = obtener_catalogo_json(agent_id)
        if not items:
            await obtener_catalogo_shopify(agent_id)
            items = obtener_catalogo_json(agent_id)
        # Normalizar al mismo esquema que CatalogoVoco para que el panel use un solo formato
        productos = [
            {
                "id": f"sh_{i}",
                "nombre": it["producto"],
                "presentacion": it["presentacion"],
                "precio": it["precio"],
                "precio_tachado": None,
                "stock": it.get("stock"),
                "disponible": True,
                "imagen_url": it.get("imagen", ""),
                "categoria": it.get("categoria", "General"),
                "descripcion": "",
                "sku": "",
            }
            for i, it in enumerate(items)
        ]
        return JSONResponse(content={
            "ok": True,
            "fuente": "shopify",
            "shopify_store": store_dom,
            "productos": productos,
        })
    else:
        from agent.memory import obtener_catalogo_voco
        productos = await obtener_catalogo_voco(agent_id)
        return JSONResponse(content={"ok": True, "fuente": "voco", "productos": productos})


@app.post("/inbox/api/catalogo")
async def api_catalogo_crear(
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import crear_producto_voco
    data = await request.json()
    if not str(data.get("nombre", "")).strip():
        return JSONResponse(status_code=400, content={"ok": False, "error": "nombre es obligatorio"})
    if not data.get("precio") and data.get("precio") != 0:
        return JSONResponse(status_code=400, content={"ok": False, "error": "precio es obligatorio"})
    producto = await crear_producto_voco(agent_id, data)
    return JSONResponse(content={"ok": True, "producto": producto})


@app.put("/inbox/api/catalogo/{producto_id}")
async def api_catalogo_actualizar(
    producto_id: int,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import actualizar_producto_voco
    data = await request.json()
    producto = await actualizar_producto_voco(producto_id, agent_id, data)
    if not producto:
        return JSONResponse(status_code=404, content={"ok": False, "error": "Producto no encontrado"})
    return JSONResponse(content={"ok": True, "producto": producto})


@app.delete("/inbox/api/catalogo/{producto_id}")
async def api_catalogo_eliminar(
    producto_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import eliminar_producto_voco
    ok = await eliminar_producto_voco(producto_id, agent_id)
    if not ok:
        return JSONResponse(status_code=404, content={"ok": False, "error": "Producto no encontrado"})
    return JSONResponse(content={"ok": True})


@app.get("/inbox/api/catalogo/template")
async def api_catalogo_template(
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Descarga el template Excel para importación masiva."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogo"
        cols = ["nombre", "categoria", "presentacion", "precio", "precio_tachado",
                "descripcion", "imagen_url", "sku", "stock", "disponible"]
        header_fill = PatternFill("solid", fgColor="16A34A")
        header_font = Font(bold=True, color="FFFFFF")
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # Fila de ejemplo
        ejemplo = ["Lavaloza Biotú", "Limpieza", "500ml", 12900, 15900,
                   "Fórmula biodegradable", "https://tudominio.com/imagen.jpg",
                   "LAV-500", 50, 1]
        for ci, val in enumerate(ejemplo, 1):
            ws.cell(row=2, column=ci, value=val)
        # Instrucciones en hoja 2
        ws2 = wb.create_sheet("Instrucciones")
        instrucciones = [
            ("Campo", "Descripción", "Obligatorio"),
            ("nombre", "Nombre del producto", "SÍ"),
            ("categoria", "Categoría (ej: Limpieza, Bebidas)", "SÍ"),
            ("presentacion", "Variante: 500ml, 1L, Pack x3, etc.", "No"),
            ("precio", "Precio en COP (sin puntos ni comas)", "SÍ"),
            ("precio_tachado", "Precio anterior para mostrar descuento", "No"),
            ("descripcion", "Descripción corta del producto", "No"),
            ("imagen_url", "URL pública directa de la imagen (JPEG/PNG/WebP, max 5MB, mín 500×500px)", "No"),
            ("sku", "Código único del producto (para catálogo nativo WhatsApp)", "No"),
            ("stock", "Cantidad disponible (dejar vacío = sin control de stock)", "No"),
            ("disponible", "1 = disponible, 0 = no disponible", "No"),
            ("", "", ""),
            ("RESTRICCIONES IMAGEN", "", ""),
            ("Formato", "JPEG, PNG o WebP únicamente", ""),
            ("Tamaño máximo", "5 MB", ""),
            ("Dimensiones mínimas", "500 × 500 px (relación 1:1 recomendada)", ""),
            ("URL válida", "Debe ser un enlace directo y público", ""),
            ("NO usar", "Google Drive, Instagram, Facebook (URLs expiradas o con auth)", ""),
            ("SÍ usar", "Tu propio sitio web, imgbb.com, imgur.com, Cloudinary", ""),
        ]
        for ri, fila in enumerate(instrucciones, 1):
            for ci, val in enumerate(fila, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                if ri == 1 or fila[0] in ("RESTRICCIONES IMAGEN",):
                    cell.font = Font(bold=True)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=catalogo_voco_template.xlsx"},
        )
    except Exception as e:
        logger.error(f"Error generando template Excel: {e}")
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.get("/inbox/api/catalogo/exportar")
async def api_catalogo_exportar(
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Exporta el catálogo actual como archivo Excel."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    try:
        from agent.memory import obtener_catalogo_voco
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        productos = await obtener_catalogo_voco(agent_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogo"
        cols = ["nombre", "categoria", "presentacion", "precio", "precio_tachado",
                "descripcion", "imagen_url", "sku", "stock", "disponible"]
        header_fill = PatternFill("solid", fgColor="16A34A")
        header_font = Font(bold=True, color="FFFFFF")
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = max(12, len(col) + 4)
        for ri, p in enumerate(productos, 2):
            ws.cell(row=ri, column=1, value=p.get("nombre", ""))
            ws.cell(row=ri, column=2, value=p.get("categoria", ""))
            ws.cell(row=ri, column=3, value=p.get("presentacion", "") or "")
            ws.cell(row=ri, column=4, value=p.get("precio", 0))
            ws.cell(row=ri, column=5, value=p.get("precio_tachado") or "")
            ws.cell(row=ri, column=6, value=p.get("descripcion", "") or "")
            ws.cell(row=ri, column=7, value=p.get("imagen_url", "") or "")
            ws.cell(row=ri, column=8, value=p.get("sku", "") or "")
            ws.cell(row=ri, column=9, value=p.get("stock") if p.get("stock") is not None else "")
            ws.cell(row=ri, column=10, value=1 if p.get("disponible") else 0)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=catalogo_voco.xlsx"},
        )
    except Exception as e:
        logger.error(f"Error exportando catálogo: {e}")
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.post("/inbox/api/catalogo/importar")
async def api_catalogo_importar(
    file: UploadFile = File(...),
    agent_id: int = Form(1),
    modo: str = Form("append"),
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Importa un archivo Excel (.xlsx) o CSV con productos al catálogo Voco.
    modo='append'  → agrega sin borrar existentes.
    modo='replace' → reemplaza todo el catálogo."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")

    contenido = await file.read()
    nombre_archivo = (file.filename or "").lower()
    productos: list[dict] = []

    try:
        if nombre_archivo.endswith(".csv"):
            import csv, io
            texto = contenido.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(texto))
            for fila in reader:
                productos.append({k.strip(): v.strip() for k, v in fila.items()})
        elif nombre_archivo.endswith(".xlsx") or nombre_archivo.endswith(".xls"):
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                fila = {headers[i]: str(row[i] if row[i] is not None else "").strip()
                        for i in range(min(len(headers), len(row)))}
                productos.append(fila)
        else:
            return JSONResponse(status_code=400,
                content={"ok": False, "error": "Formato no soportado. Usa .xlsx o .csv"})
    except Exception as e:
        return JSONResponse(status_code=400,
            content={"ok": False, "error": f"Error leyendo archivo: {e}"})

    if not productos:
        return JSONResponse(status_code=400,
            content={"ok": False, "error": "El archivo está vacío o no tiene filas de datos"})

    from agent.memory import importar_productos_voco
    resultado = await importar_productos_voco(agent_id, productos, modo=modo)
    return JSONResponse(content={"ok": True, **resultado})


# ─────────────────────────────────────────────────────────────────────────────
# Módulo Pedidos nativos Voco
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/inbox/api/pedidos/stats")
async def api_pedidos_stats(
    agent_id: int = 1,
    dias: int = 30,
    desde: str = "",
    hasta: str = "",
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_stats_pedidos
    from datetime import datetime as _dt, timedelta as _td
    _desde = desde or None
    _hasta = hasta or None
    if not _desde and not _hasta and dias:
        _hasta = _dt.utcnow().strftime("%Y-%m-%d")
        _desde = (_dt.utcnow() - _td(days=int(dias))).strftime("%Y-%m-%d")
    stats = await obtener_stats_pedidos(agent_id, desde=_desde, hasta=_hasta)
    return JSONResponse(content=stats)


@app.get("/inbox/api/pedidos")
async def api_pedidos_list(
    agent_id: int = 1,
    estado: str = "",
    telefono: str = "",
    limite: int = 50,
    offset: int = 0,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_pedidos
    pedidos = await obtener_pedidos(
        agent_id=agent_id,
        estado=estado or None,
        telefono=telefono or None,
        limite=limite,
        offset=offset,
    )
    return JSONResponse(content={"pedidos": pedidos})


@app.post("/inbox/api/pedidos")
async def api_pedidos_crear(
    request: Request,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    data = await request.json()
    from agent.memory import crear_pedido_manual, guardar_cliente
    try:
        ag = int(data.get("agent_id", 1))
        pedido = await crear_pedido_manual(ag, data)
        # Crear/actualizar cliente con los datos del pedido
        tel = str(data.get("telefono_cliente", "")).strip()
        nombre = str(data.get("nombre_cliente", "")).strip()
        if tel and nombre:
            await guardar_cliente(tel, {
                "nombres":   nombre,
                "direccion": str(data.get("direccion_entrega", ""))[:200],
                "direccion2": str(data.get("direccion2_entrega", ""))[:200],
                "ciudad":    str(data.get("ciudad_entrega", ""))[:100],
            }, agent_id=ag)
        return JSONResponse(content={"ok": True, "pedido": pedido})
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.get("/inbox/api/pedidos/{pedido_id}")
async def api_pedidos_get(
    pedido_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_pedido
    pedido = await obtener_pedido(pedido_id, agent_id)
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    return JSONResponse(content=pedido)


@app.put("/inbox/api/pedidos/{pedido_id}")
async def api_pedidos_actualizar(
    pedido_id: int,
    request: Request,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    data = await request.json()
    from agent.memory import actualizar_pedido
    try:
        pedido = await actualizar_pedido(pedido_id, agent_id, data)
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado")
        return JSONResponse(content={"ok": True, "pedido": pedido})
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})


@app.delete("/inbox/api/pedidos/{pedido_id}")
async def api_pedidos_eliminar(
    pedido_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import eliminar_pedido
    eliminado = await eliminar_pedido(pedido_id, agent_id)
    if not eliminado:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    return JSONResponse(content={"ok": True})


@app.get("/inbox/api/pedidos/{pedido_id}/remision")
async def api_pedidos_remision_pdf(
    pedido_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Genera la remisión de entrega como PDF descargable."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_pedido
    pedido = await obtener_pedido(pedido_id, agent_id)
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    pdf_bytes = await _generar_remision_pdf(pedido, agent_id)
    from fastapi.responses import Response
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="remision_{pedido["numero"]}.pdf"',
        },
    )


@app.post("/inbox/api/pedidos/{pedido_id}/remision/whatsapp")
async def api_pedidos_remision_whatsapp(
    pedido_id: int,
    agent_id: int = 1,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Genera la remisión PDF y la envía por WhatsApp al número del cliente."""
    if not await _obtener_sesion_usuario(voco_session or inbox_session or token):
        raise HTTPException(status_code=401, detail="No autorizado")
    from agent.memory import obtener_pedido
    pedido = await obtener_pedido(pedido_id, agent_id)
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    telefono = pedido.get("telefono_cliente", "")
    if not telefono:
        return JSONResponse(status_code=400, content={"ok": False, "error": "Pedido sin número de cliente"})

    pdf_bytes = await _generar_remision_pdf(pedido, agent_id)
    filename  = f"remision_{pedido['numero']}.pdf"

    try:
        proveedor = await _get_proveedor_panel(agent_id)
        media_id  = await proveedor.subir_media(pdf_bytes, filename, "application/pdf")
        if not media_id:
            return JSONResponse(status_code=502, content={"ok": False, "error": "No se pudo subir el PDF a Meta"})
        ok = await proveedor.enviar_documento(
            telefono,
            media_id=media_id,
            caption=f"📦 Remisión {pedido['numero']} — gracias por tu pedido.",
            filename=filename,
        )
        if not ok:
            return JSONResponse(status_code=502, content={"ok": False, "error": "No se pudo enviar el documento"})
        return JSONResponse(content={"ok": True, "media_id": media_id})
    except Exception as e:
        logger.error(f"Error enviando remisión WA pedido {pedido_id}: {e}")
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)[:200]})


async def _generar_remision_pdf(pedido: dict, agent_id: int) -> bytes:
    """Genera un PDF de remisión de entrega para el pedido dado."""
    from fpdf import FPDF
    from io import BytesIO
    from agent.memory import obtener_configuracion

    # Nombre del negocio desde configuración del agente
    try:
        biz_name = await obtener_configuracion("BUSINESS_NAME", agent_id=agent_id) or "Voco"
        biz_phone = await obtener_configuracion("BUSINESS_PHONE", agent_id=agent_id) or ""
        biz_address = await obtener_configuracion("BUSINESS_ADDRESS", agent_id=agent_id) or ""
    except Exception:
        biz_name, biz_phone, biz_address = "Voco", "", ""

    LABEL_PAGO = {"pendiente": "Por cobrar", "pagado": "Pagado", "cod": "Contra entrega (COD)"}
    LABEL_ESTADO = {
        "creado": "Creado", "alistado": "Alistado", "despachado": "Despachado",
        "entregado": "Entregado", "cancelado": "Cancelado",
    }

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    pdf.set_margins(20, 20, 20)

    W = 170  # ancho útil

    # ── Encabezado ──
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(22, 163, 74)  # verde
    pdf.cell(W, 8, biz_name, ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(100, 100, 100)
    if biz_phone:
        pdf.cell(W, 5, f"Tel: {biz_phone}", ln=True)
    if biz_address:
        pdf.cell(W, 5, biz_address, ln=True)
    pdf.ln(4)

    # ── Título ──
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(W, 8, "REMISION DE ENTREGA", ln=False)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, pedido["numero"], align="R", ln=True)

    pdf.set_draw_color(229, 231, 235)
    pdf.set_line_width(0.4)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(5)

    # ── Datos del pedido ──
    fecha = (pedido.get("created_at") or "")[:10] or "—"
    estado_label   = LABEL_ESTADO.get(pedido.get("estado", ""), pedido.get("estado", ""))
    pago_label     = LABEL_PAGO.get(pedido.get("estado_pago", ""), pedido.get("estado_pago", ""))

    def campo(label: str, valor: str, col2: bool = False):
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(30, 6, label.upper() + ":", ln=False)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(30, 41, 59)
        if col2:
            pdf.cell(55, 6, valor, ln=False)
        else:
            pdf.cell(W - 30, 6, valor, ln=True)

    campo("Fecha", fecha)
    campo("Estado", estado_label)
    campo("Pago", pago_label)
    pdf.ln(3)

    # ── Info cliente ──
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(22, 163, 74)
    pdf.cell(W, 7, "DATOS DEL CLIENTE", ln=True)
    pdf.set_line_width(0.3)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(3)
    campo("Nombre", pedido.get("nombre_cliente", "") or "—")
    campo("Telefono", pedido.get("telefono_cliente", "") or "—")
    _dir = pedido.get("direccion_entrega", "") or ""
    if pedido.get("direccion2_entrega"):
        _dir += ", " + pedido["direccion2_entrega"]
    if pedido.get("ciudad_entrega"):
        _dir += " — " + pedido["ciudad_entrega"]
    campo("Direccion", _dir or "—")
    if pedido.get("notas_cliente"):
        campo("Notas", pedido["notas_cliente"])
    pdf.ln(5)

    # ── Tabla productos ──
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(22, 163, 74)
    pdf.cell(W, 7, "PRODUCTOS", ln=True)
    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(3)

    # Cabecera tabla
    col_w = [85, 20, 30, 35]
    pdf.set_fill_color(241, 245, 249)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(100, 100, 100)
    for txt, w in zip(["PRODUCTO", "CANT.", "PRECIO UNIT.", "SUBTOTAL"], col_w):
        pdf.cell(w, 7, txt, border=0, fill=True, ln=False, align="C" if txt != "PRODUCTO" else "L")
    pdf.ln()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(30, 41, 59)
    for p in pedido.get("productos", []):
        nombre = str(p.get("nombre", ""))[:50]
        cant   = str(p.get("cantidad", 1))
        precio = f"${int(p.get('precio_unitario', 0)):,}".replace(",", ".")
        sub    = f"${int(p.get('subtotal', int(p.get('precio_unitario', 0)) * int(p.get('cantidad', 1)))):,}".replace(",", ".")
        row_h  = 6
        pdf.cell(col_w[0], row_h, nombre, ln=False)
        pdf.cell(col_w[1], row_h, cant,   ln=False, align="C")
        pdf.cell(col_w[2], row_h, precio, ln=False, align="R")
        pdf.cell(col_w[3], row_h, sub,    ln=False, align="R")
        pdf.ln()

    pdf.line(20, pdf.get_y(), 190, pdf.get_y())
    pdf.ln(3)

    # ── Totales ──
    def fila_total(label: str, valor: int, bold: bool = False, color=(30, 41, 59)):
        pdf.set_font("Helvetica", "B" if bold else "", 9)
        pdf.set_text_color(*color)
        pdf.cell(W - 40, 6, label, align="R", ln=False)
        pdf.cell(40, 6, f"${valor:,}".replace(",", "."), align="R", ln=True)

    fila_total("Subtotal:", pedido.get("subtotal", 0))
    if pedido.get("descuento", 0):
        fila_total("Descuento:", -pedido["descuento"])
    if pedido.get("costo_envio", 0):
        fila_total("Costo de envio:", pedido["costo_envio"])
    fila_total("TOTAL:", pedido.get("total", 0), bold=True, color=(22, 163, 74))
    pdf.ln(8)

    # ── Firmas ──
    pdf.set_draw_color(200, 200, 200)
    pdf.set_line_width(0.3)
    sig_y = pdf.get_y() + 15
    pdf.line(25, sig_y, 85, sig_y)
    pdf.line(105, sig_y, 165, sig_y)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.set_y(sig_y + 2)
    pdf.set_x(25)
    pdf.cell(60, 5, "Firma Entregado por", align="C", ln=False)
    pdf.set_x(105)
    pdf.cell(60, 5, "Firma Recibido por", align="C", ln=True)
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(W, 5, f"Generado por Voco | {pedido['numero']} | {fecha}", align="C", ln=True)

    buf = BytesIO()
    pdf_out = pdf.output(dest="S")
    if isinstance(pdf_out, str):
        pdf_out = pdf_out.encode("latin-1")
    return bytes(pdf_out)


# ── Panel por slug (ruta comodín — DEBE ser la última ruta de /inbox/*) ────
# FastAPI evalúa rutas en orden de definición: al estar aquí al final, las rutas
# específicas (/inbox/api/*, /inbox/broadcast/*, etc.) ya fueron registradas primero.

@app.get("/inbox/{slug}", response_class=HTMLResponse)
async def inbox_agente_por_slug(
    slug: str,
    token: str = "",
    inbox_session: str = Cookie(default=""),
    voco_session: str = Cookie(default=""),
):
    """Panel de inbox para un agente específico por su slug (ej: /inbox/equora)."""
    effective_token = voco_session or inbox_session or token
    current_user = await _obtener_sesion_usuario(effective_token)

    if not current_user:
        return RedirectResponse("/inbox/login", status_code=302)

    from agent.memory import obtener_agente_por_slug
    agente = await obtener_agente_por_slug(slug)
    if not agente:
        raise HTTPException(status_code=404, detail=f"Agente '{slug}' no encontrado")

    # Verificar ownership: admin ve cualquier agente; usuario solo los suyos
    if current_user.get("rol") != "admin":
        owner = agente.get("owner_id")
        if owner != current_user["id"]:
            raise HTTPException(status_code=403, detail="No tienes acceso a este agente")

    response = HTMLResponse(content=obtener_inbox_html(agente, user=current_user))
    if token and _verificar_admin(token):
        response.set_cookie(
            INBOX_COOKIE, ADMIN_TOKEN,
            httponly=True, max_age=COOKIE_MAX_AGE, samesite="lax"
        )
    return response


SHOPIFY_WEBHOOK_SECRET = os.getenv("SHOPIFY_WEBHOOK_SECRET", "")
# Cache de deduplicación: evita procesar el mismo webhook de Shopify más de una vez
# Shopify reintenta hasta 19 veces si no recibe 200 — guardamos los procesados por 10 min
_shopify_dedup: dict[str, float] = {}
_SHOPIFY_DEDUP_TTL = 600  # segundos

# Cache de deduplicación para mensajes de WhatsApp entrantes.
# Meta reintenta el webhook cuando no recibe 200 en ~20 s — sin este guard,
# el mismo mensaje genera múltiples respuestas y escalaciones.
_wamids_procesados: dict[str, float] = {}
_WAMID_DEDUP_TTL = 300  # 5 minutos

ADMIN_WHATSAPP_NUMBERS = [
    n.strip() for n in os.getenv("ADMIN_WHATSAPP_NUMBERS", "").split(",") if n.strip()
]


# ══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE ALERTAS — Tracking de fallas del catálogo nativo de WhatsApp
# ══════════════════════════════════════════════════════════════════════════════
# Cuando el catálogo nativo de WhatsApp falla (Meta API caída, token expirado,
# catalog_id inválido, etc.) y caemos al fallback web, registramos la falla.
# Si se acumulan en una ventana corta, alertamos al admin por WhatsApp.

from collections import deque

# Buffer en memoria de las últimas 100 fallas del catálogo nativo
_catalogo_fallas: deque = deque(maxlen=100)
# Timestamp de la última alerta enviada al admin — cooldown 30 min
_ultima_alerta_catalogo_at: float = 0.0
_CATALOGO_ALERTA_COOLDOWN_SEG = 30 * 60  # 30 minutos entre alertas


async def _registrar_falla_catalogo(
    tipo: str, telefono_cliente: str, motivo: str, agent_id: int = 1
) -> None:
    """Registra una falla del catálogo nativo y alerta al admin si aplica.
    `tipo` puede ser: product_list, catalog_message, single_product, etc."""
    global _ultima_alerta_catalogo_at
    ahora = time.time()
    _catalogo_fallas.append({
        "tipo":     tipo,
        "telefono": telefono_cliente,
        "motivo":   motivo[:300],
        "agent_id": agent_id,
        "at":       datetime.utcnow().isoformat(),
    })
    logger.warning(f"[catalogo-falla] {tipo} → {telefono_cliente}: {motivo[:150]}")

    # Alerta al admin: solo si pasó el cooldown (evita spam) Y hay admins
    if not ADMIN_WHATSAPP_NUMBERS:
        return
    if ahora - _ultima_alerta_catalogo_at < _CATALOGO_ALERTA_COOLDOWN_SEG:
        return
    _ultima_alerta_catalogo_at = ahora

    # Contar fallas en la última hora para dar contexto
    hace_una_hora = datetime.utcnow() - timedelta(hours=1)
    fallas_recientes = sum(
        1 for f in _catalogo_fallas
        if datetime.fromisoformat(f["at"]) > hace_una_hora
    )

    mensaje_alerta = (
        f"⚠️ *Alerta Voco — Catálogo WhatsApp*\n\n"
        f"El catálogo nativo de WhatsApp falló y se está usando el fallback web.\n\n"
        f"*Último error:* {tipo}\n"
        f"*Motivo:* {motivo[:200]}\n"
        f"*Cliente afectado:* +{telefono_cliente}\n"
        f"*Fallas en la última hora:* {fallas_recientes}\n\n"
        f"Posibles causas: token Meta expirado, catálogo no sincronizado, "
        f"Meta API caída, o un SKU específico no existe en Facebook Catalog.\n\n"
        f"Próxima alerta en {_CATALOGO_ALERTA_COOLDOWN_SEG // 60} min si sigue fallando."
    )

    _prov = await _get_meta_para_agente({"id": agent_id})
    cliente_digits = "".join(filter(str.isdigit, telefono_cliente))
    for admin in ADMIN_WHATSAPP_NUMBERS:
        admin_digits = "".join(filter(str.isdigit, admin))
        if admin_digits == cliente_digits or admin_digits.endswith(cliente_digits[-10:]):
            continue
        try:
            await _prov.enviar_mensaje(admin, mensaje_alerta)
            logger.info(f"[alerta-catalogo] Notificado al admin {admin}")
        except Exception as e:
            logger.error(f"[alerta-catalogo] Error notificando a {admin}: {e}")


async def _notificar_escalacion(telefono_cliente: str, datos: dict, agent_id: int = 1):
    """Crea un ticket en BD, pausa el bot y notifica por WhatsApp a los asesores internos.
    NUNCA envía el aviso al número del propio cliente."""
    motivo        = datos.get("motivo", "sin especificar")
    contexto      = datos.get("contexto", "")
    urgencia      = datos.get("urgencia", "normal")
    cliente_nombre = datos.get("nombre_cliente", "")

    # ── 1. Crear ticket en BD y notificar via SSE ─────────────────────────────
    ticket = None
    try:
        ticket = await crear_ticket(
            agent_id=agent_id,
            telefono_cliente=telefono_cliente,
            nombre_cliente=cliente_nombre,
            motivo=motivo,
            urgencia=urgencia,
            contexto=contexto,
        )
        # Auditoría — autor = nombre del agente del tenant (no hardcode "Andrea Bot")
        try:
            from agent.memory import obtener_agente as _obtener_ag
            _ag = await _obtener_ag(agent_id)
            _bot_label = ((_ag or {}).get("agent_name") or "Bot") + " (IA)"
        except Exception:
            _bot_label = "Bot (IA)"
        await registrar_evento_ticket(ticket["id"], "creado", _bot_label,
                                      f"Motivo: {motivo} | Urgencia: {urgencia}")
        # Auto-asignación round-robin (si está habilitada)
        try:
            auto_asig = await get_config_value("AUTO_ASIGNAR", agent_id)
            if auto_asig == "1":
                agente_id_rr = await obtener_siguiente_agente_roundrobin(agent_id)
                if agente_id_rr:
                    ticket = await tomar_ticket(ticket["id"], agente_id_rr)
                    if ticket:
                        await registrar_evento_ticket(ticket["id"], "tomado",
                            "Sistema (auto)", f"Asignado automáticamente a {ticket.get('agente_nombre','')}")
        except Exception as e_rr:
            logger.warning(f"Auto-asignación falló: {e_rr}")

        _push_evento_ticket(agent_id, "ticket_nuevo", ticket)
        logger.info(f"Ticket creado para {telefono_cliente} — motivo: {motivo}")
    except Exception as e:
        logger.error(f"Error creando ticket de escalación: {e}")

    # ── 2. Pausar el bot para esta conversación ────────────────────────────────
    try:
        await set_modo_humano(telefono_cliente, True, agent_id=agent_id)
        logger.info(f"Bot pausado para {telefono_cliente} por escalación")
    except Exception as e:
        logger.error(f"Error pausando bot: {e}")

    # ── 3. Notificar por WhatsApp a los asesores que hayan opt-in ──────────────
    # Estrategia de resolución de receptores (#52):
    #   1. Usuarios internos del agente con notif_escalaciones_wa=True (BD).
    #   2. Si no hay nadie opt-in en BD → fallback a ADMIN_WHATSAPP_NUMBERS
    #      (env var legacy). Mantiene compatibilidad con setup actual de Railway.
    #   3. Si ambos están vacíos → solo se queda en el panel (sin notif WA).
    from agent.memory import obtener_receptores_escalacion_wa, obtener_agente
    receptores: list[str] = []
    try:
        receptores = await obtener_receptores_escalacion_wa(agent_id)
    except Exception as e:
        logger.error(f"[escalacion] error leyendo receptores de BD: {e}")
    origen_receptores = "BD"
    if not receptores:
        receptores = ADMIN_WHATSAPP_NUMBERS
        origen_receptores = "env(legacy)"

    if not receptores:
        logger.info(
            f"[escalacion] sin receptores configurados (BD vacía + ADMIN_WHATSAPP_NUMBERS "
            f"vacía) para agent={agent_id} — solo notificación en panel"
        )
        return

    # Nombre del negocio (para el mensaje, en lugar del hardcode "Equora")
    nombre_negocio = ""
    try:
        agente_info = await obtener_agente(agent_id)
        if agente_info:
            nombre_negocio = agente_info.get("name") or ""
    except Exception:
        pass
    referencia_panel = f"panel de {nombre_negocio}" if nombre_negocio else "panel de Voco"

    cliente_digits = "".join(filter(str.isdigit, telefono_cliente))
    mensaje = (
        f"🚨 *Escalación de conversación*\n\n"
        f"*Motivo:* {motivo}\n"
        f"*Urgencia:* {urgencia}\n"
        f"*Cliente:* {cliente_nombre or 'desconocido'}\n"
        f"*WhatsApp cliente:* +{telefono_cliente}\n\n"
        f"*Contexto:*\n{contexto}\n\n"
        f"Entra al {referencia_panel} para tomar la conversación."
    )

    logger.info(
        f"[escalacion] notificando a {len(receptores)} receptor(es) "
        f"(origen={origen_receptores}) para agent={agent_id}"
    )
    _prov = await _get_meta_para_agente({"id": agent_id})
    for admin in receptores:
        admin_digits = "".join(filter(str.isdigit, admin))
        # Filtro de seguridad: no notificarse a uno mismo si el cliente es el
        # mismo número del receptor (caso típico al hacer pruebas con propio WA).
        if admin_digits == cliente_digits or admin_digits.endswith(cliente_digits[-10:]):
            logger.warning(f"Escalación omitida para {admin}: es el mismo número del cliente")
            continue
        try:
            await _prov.enviar_mensaje(admin, mensaje)
            logger.info(f"Escalación notificada al asesor {admin}")
        except Exception as e:
            logger.error(f"Error notificando escalación a {admin}: {e}")


async def _verificar_hmac_shopify(body: bytes, hmac_header: str, shop: str = "") -> bool:
    """Verifica el HMAC del webhook de Shopify.

    Webhooks registrados vía Admin API (OAuth) se firman con el Client Secret
    de la app — el mismo que el merchant pegó en Configuración. Voco intenta
    validar contra el client_secret de cada agente conocido hasta encontrar
    coincidencia. Fallback al SHOPIFY_WEBHOOK_SECRET legacy (env var) para
    setups antiguos manuales.
    """
    if not hmac_header:
        return False

    secrets_a_probar: list[str] = []

    # 1) Client Secret del agente conectado a este shop (post-OAuth).
    #    Multi-tenant: si hay varios agentes, recolectamos todos los
    #    client_secrets de quienes tengan ese shop configurado.
    try:
        from agent.memory import async_session, ConfigValue
        from sqlalchemy import select
        async with async_session() as s:
            # Buscar todos los agentes con SHOPIFY_CLIENT_SECRET configurado.
            # Si conocemos el shop, filtramos por SHOPIFY_STORE coincidente.
            stmt = select(ConfigValue).where(ConfigValue.clave == "SHOPIFY_CLIENT_SECRET")
            res = await s.execute(stmt)
            for cv in res.scalars().all():
                if cv.valor:
                    if shop:
                        # Verificar que el agente tenga este shop configurado
                        stmt2 = select(ConfigValue).where(
                            ConfigValue.agent_id == cv.agent_id,
                            ConfigValue.clave == "SHOPIFY_STORE",
                        )
                        store_cv = (await s.execute(stmt2)).scalar_one_or_none()
                        if store_cv and store_cv.valor and store_cv.valor.lower() == shop.lower():
                            secrets_a_probar.append(cv.valor)
                    else:
                        secrets_a_probar.append(cv.valor)
    except Exception as e:
        logger.warning(f"[hmac] no pude leer client_secrets de BD: {e}")

    # 2) Fallback legacy: env var SHOPIFY_WEBHOOK_SECRET (setups manuales viejos)
    if SHOPIFY_WEBHOOK_SECRET:
        secrets_a_probar.append(SHOPIFY_WEBHOOK_SECRET)

    if not secrets_a_probar:
        logger.warning("[hmac] sin secrets configurados — saltando verificación")
        return True

    for secret in secrets_a_probar:
        digest = base64.b64encode(
            hmac.new(secret.encode("utf-8"), body, hashlib.sha256).digest()
        ).decode()
        if hmac.compare_digest(digest, hmac_header):
            return True
    return False


# NOTA: antes había aquí una segunda definición de _normalizar_telefono que
# sobrescribía la versión inteligente del panel (línea ~5302) porque Python
# carga el módulo top-down. Resultado: webhooks de Shopify que traían el
# teléfono sin indicativo (ej: '3022888274') NO recibían el prefijo +57, y
# el mensaje "Tu pedido está listo" se enviaba al número sin código de país
# → no llegaba al cliente. La función inteligente añade '57' si detecta
# 10 dígitos empezando por '3' (caso real Lorena Camayo, 2026-06-16).


def _extraer_telefono(payload: dict) -> str | None:
    for attr in payload.get("note_attributes", []) or []:
        if attr.get("name") in ("Telefono WhatsApp", "Telefono", "phone"):
            tel = _normalizar_telefono(attr.get("value"))
            if tel:
                return tel
    cliente = payload.get("customer") or {}
    for valor in (cliente.get("phone"), payload.get("phone"),
                  (payload.get("shipping_address") or {}).get("phone"),
                  (payload.get("billing_address") or {}).get("phone")):
        tel = _normalizar_telefono(valor)
        if tel:
            return tel
    return None


def _extraer_datos_cliente(payload: dict) -> dict:
    """Construye el dict de cliente desde la orden Shopify."""
    cliente = payload.get("customer") or {}
    envio = payload.get("shipping_address") or payload.get("billing_address") or {}
    return {
        "nombres": cliente.get("first_name") or envio.get("first_name") or "",
        "apellidos": cliente.get("last_name") or envio.get("last_name") or "",
        "email": cliente.get("email") or payload.get("email") or "",
        "direccion": envio.get("address1") or "",
        "barrio": envio.get("address2") or "",
        "ciudad": envio.get("city") or "",
        "departamento": envio.get("province") or "",
        "razon_social": envio.get("company") or "",
    }


@app.post("/shopify-webhook")
async def shopify_webhook(request: Request):
    """Recibe eventos de Shopify (orders/paid, orders/create) y notifica al cliente."""
    body = await request.body()
    hmac_header = request.headers.get("X-Shopify-Hmac-Sha256", "")
    topic = request.headers.get("X-Shopify-Topic", "")
    shop_header = request.headers.get("X-Shopify-Shop-Domain", "")

    if not await _verificar_hmac_shopify(body, hmac_header, shop=shop_header):
        logger.error(f"HMAC inválido en shopify-webhook (topic={topic}, shop={shop_header})")
        raise HTTPException(status_code=401, detail="Invalid HMAC")

    try:
        payload = json.loads(body)
    except Exception as e:
        logger.error(f"Payload Shopify inválido: {e}")
        raise HTTPException(status_code=400, detail="Invalid JSON")

    numero = payload.get("order_number") or payload.get("name", "").lstrip("#")
    nombre_orden = payload.get("name") or (f"#{numero}" if numero else "")
    telefono = _extraer_telefono(payload)
    total = payload.get("total_price") or payload.get("current_total_price") or "0"

    logger.info(f"Shopify webhook {topic} — orden {nombre_orden}, tel {telefono}")

    # ── Deduplicación: ignorar reintentos de Shopify ──────────────────────────
    ahora_ts = time.time()
    # Limpiar entradas expiradas
    expiradas = [k for k, v in _shopify_dedup.items() if ahora_ts - v > _SHOPIFY_DEDUP_TTL]
    for k in expiradas:
        del _shopify_dedup[k]
    dedup_key = f"{numero}_{topic}"
    if dedup_key in _shopify_dedup:
        logger.warning(f"Webhook duplicado ignorado: {dedup_key}")
        return {"status": "ok"}
    _shopify_dedup[dedup_key] = ahora_ts
    # ─────────────────────────────────────────────────────────────────────────

    if not telefono or not numero:
        logger.warning(f"Sin teléfono o número de orden — no se notifica (telefono={telefono}, numero={numero})")
        return {"status": "ok"}

    try:
        total_int = int(float(total))
    except Exception:
        total_int = 0

    # Guardar/actualizar cliente con los datos que llegaron en la orden.
    # Solo incrementamos pedidos_realizados cuando es orders/paid (pago confirmado).
    datos_cliente = _extraer_datos_cliente(payload)
    if any(datos_cliente.values()):
        try:
            await guardar_cliente(
                telefono, datos_cliente,
                incrementar_pedidos=(topic == "orders/paid")
            )
            logger.info(f"Cliente {telefono} guardado/actualizado desde Shopify ({topic})")
        except Exception as e:
            logger.error(f"Error guardando cliente desde webhook: {e}")

    # ── checkouts/create: cliente abrió el formulario de Shopify ──────────────
    # Guardamos la checkout_url para poder reenviarla si no termina el proceso.
    # La URL del checkout viene en el campo "checkout_url" o se construye desde el token.
    if topic == "checkouts/create":
        checkout_url = (
            payload.get("checkout_url")
            or payload.get("web_url")
            or ""
        )
        if not checkout_url:
            # Fallback: construir desde el token con formato correcto de Shopify.
            # Usar SHOPIFY_STORE configurado (multi-tenant) en lugar de hardcoded.
            token = payload.get("token") or ""
            if token:
                # Resolver tienda multi-tenant: config_value primero, env legacy
                from agent.tools import _resolver_store as _rs
                tienda = await _rs()
                if tienda:
                    checkout_url = f"https://{tienda}/checkouts/cn/{token}/es"
        # Normalizar dominio: si el cliente configuró SHOPIFY_CUSTOM_DOMAIN
        # (con la ruta /checkouts correctamente apuntada), usarlo. Si no,
        # dejar el dominio myshopify.com original — funciona SIEMPRE y evita
        # el 404 cuando el dominio personalizado no tiene esa ruta.
        if checkout_url:
            import re as _re
            custom_dom = os.getenv("SHOPIFY_CUSTOM_DOMAIN", "").strip()
            if custom_dom:
                # Limpiar custom_dom (sin https:// ni trailing slash)
                custom_dom = custom_dom.replace("https://", "").replace("http://", "").rstrip("/")
                checkout_url = _re.sub(
                    r"https://[^/]*\.myshopify\.com/",
                    f"https://{custom_dom}/",
                    checkout_url,
                )
            # Quitar preview_theme_id (solo aparece en modo preview, confunde al cliente)
            checkout_url = _re.sub(r"[&?]preview_theme_id=[^&]*", "", checkout_url)
            checkout_url = checkout_url.rstrip("?&")
        if telefono and checkout_url:
            try:
                # Guardar carrito activo como pedido pendiente para abandono detection
                line_items = payload.get("line_items") or []
                if line_items:
                    productos = [
                        {
                            "producto": item.get("title", ""),
                            "presentacion": item.get("variant_title") or item.get("title", ""),
                            "cantidad": int(item.get("quantity", 1)),
                            "precio_unitario": int(float(item.get("price", 0))),
                            "subtotal": int(float(item.get("price", 0))) * int(item.get("quantity", 1)),
                        }
                        for item in line_items
                    ]
                    await guardar_pedido_pendiente(telefono, productos)
                await guardar_checkout_url(telefono, checkout_url)
                # NO vaciar carrito_activo aquí — el cliente abrió el checkout pero
                # puede abandonarlo. Solo vaciamos cuando llega orders/create o
                # orders/paid (pedido cerrado realmente).
                logger.info(f"Checkout creado para {telefono}: {checkout_url} — carrito persiste hasta orders/paid")
            except Exception as e:
                logger.error(f"Error guardando checkout create para {telefono}: {e}")
        return {"status": "ok"}

    # Cliente completó el checkout → ya no hay carrito pendiente NI activo.
    # IMPORTANTE: resolver TODOS los agent_ids bajo los cuales existe el cliente
    # y limpiar para cada uno. Sin esto, asumíamos agent_id=1 por default —
    # rompía multi-tenant y causaba el bug "checkout abandonado" enviado
    # DESPUÉS de confirmar pedido (el limpiar no aplicaba al agent correcto).
    if topic in ("orders/create", "orders/paid"):
        agent_ids = await obtener_agent_ids_por_telefono(telefono)
        if not agent_ids:
            agent_ids = [1]  # fallback legacy si el cliente no existe en BD
            logger.warning(
                f"[shopify-webhook] {telefono} no encontrado en BD — fallback agent_id=1"
            )
        for aid in agent_ids:
            try:
                await limpiar_pedido_pendiente(telefono, agent_id=aid)
                logger.info(f"[shopify-webhook] pedido_pendiente limpiado para {telefono} (agent_id={aid}) tras {topic}")
            except Exception as e:
                logger.error(f"Error limpiando pedido pendiente para {telefono} (agent_id={aid}): {e}")
            try:
                await limpiar_carrito_activo(telefono, agent_id=aid)
                logger.info(f"[shopify-webhook] carrito_activo vaciado para {telefono} (agent_id={aid}) tras {topic}")
            except Exception as e:
                logger.error(f"Error limpiando carrito_activo para {telefono} (agent_id={aid}) tras {topic}: {e}")

    # ── Sync Shopify → Voco: registrar pedido nuevo en el panel ─────────────
    if topic == "orders/create":
        try:
            from agent.memory import upsert_pedido_shopify
            shopify_order_id = str(payload.get("id", ""))
            line_items = payload.get("line_items") or []
            prods_voco = [
                {
                    "nombre":          item.get("title", ""),
                    "presentacion":    item.get("variant_title") or item.get("title", ""),
                    "cantidad":        int(item.get("quantity", 1)),
                    "precio_unitario": int(float(item.get("price", 0))),
                    "subtotal":        int(float(item.get("price", 0))) * int(item.get("quantity", 1)),
                }
                for item in line_items
            ]
            addr = payload.get("shipping_address") or {}
            dir1 = (addr.get("address1") or "").strip()
            dir2 = (addr.get("address2") or "").strip()
            ciudad = (addr.get("city") or "").strip()
            financial = payload.get("financial_status", "pending")
            ep_init = "pagado" if financial in ("paid", "voided") else "pendiente"
            try:
                sub_shopify = int(float(payload.get("subtotal_price") or 0))
            except Exception:
                sub_shopify = 0
            try:
                env_shopify = int(float(
                    (payload.get("total_shipping_price_set") or {})
                    .get("shop_money", {}).get("amount") or 0
                ))
            except Exception:
                env_shopify = 0
            nombre_cli = (
                (datos_cliente.get("nombres") or "") + " " +
                (datos_cliente.get("apellidos") or "")
            ).strip()
            data_upsert_create = {
                "numero":             nombre_orden,
                "telefono_cliente":   telefono,
                "nombre_cliente":     nombre_cli,
                "estado":             "creado",
                "estado_pago":        ep_init,
                "productos":          prods_voco,
                "subtotal":           sub_shopify,
                "costo_envio":        env_shopify,
                "total":              total_int,
                "direccion_entrega":  dir1,
                "direccion2_entrega": dir2,
                "ciudad_entrega":     ciudad,
            }
            for aid in agent_ids:
                await upsert_pedido_shopify(aid, shopify_order_id, data_upsert_create)
            logger.info(f"[shopify-sync] Pedido {nombre_orden} creado en Voco (agents={agent_ids})")
        except Exception as _e_usq:
            logger.error(f"[shopify-sync] Error creando pedido en Voco: {_e_usq}")

    if topic == "orders/fulfilled":
        # Shopify fulfillment puede traer tracking — si está, lo incluimos
        fulfillments = payload.get("fulfillments") or []
        tracking_num = ""
        tracking_url = ""
        if fulfillments:
            f0 = fulfillments[0]
            tracking_num = f0.get("tracking_number") or ""
            tracking_url = f0.get("tracking_url") or ""

        # Tracking opcional — se forma según lo que Shopify mande, se inyecta
        # como placeholder {tracking} en el template configurable.
        extra_tracking = ""
        if tracking_num:
            extra_tracking = f"\n📦 Guía: *{tracking_num}*"
        if tracking_url:
            extra_tracking += f"\n🔗 Seguimiento: {tracking_url}"

        # Sync Shopify → Voco: marcar pedido despachado
        try:
            from agent.memory import upsert_pedido_shopify
            _shopify_oid_f = str(payload.get("id", ""))
            try:
                _aids_f = agent_ids
            except NameError:
                _aids_f = await obtener_agent_ids_por_telefono(telefono) or [1]
            for _aid_f in _aids_f:
                await upsert_pedido_shopify(_aid_f, _shopify_oid_f, {"estado": "despachado", "total": total_int})
            logger.info(f"[shopify-sync] Pedido {nombre_orden} marcado despachado en Voco")
        except Exception as _e_usq_f:
            logger.error(f"[shopify-sync] Error actualizando estado despachado en Voco: {_e_usq_f}")

        # Resolver agent_id para leer el mensaje configurable del agente correcto.
        # Si no se resolvió antes en este handler, lo hacemos ahora.
        try:
            _aid_shopify = agent_ids[0]  # ya resuelto arriba en orders/create|paid
        except (NameError, IndexError):
            _aids_tmp = await obtener_agent_ids_por_telefono(telefono)
            _aid_shopify = _aids_tmp[0] if _aids_tmp else 1
        from agent.mensajes import format_seguro as _fmt_shopify
        _ctx_shopify = await construir_contexto_placeholders(_aid_shopify)
        _ctx_shopify["numero_pedido"] = nombre_orden
        _ctx_shopify["total"] = f"{total_int:,}"
        _ctx_shopify["tracking"] = extra_tracking
        mensaje = _fmt_shopify(
            await obtener_mensaje(_aid_shopify, "shopify.order_fulfilled"),
            _ctx_shopify,
        )
        if not mensaje:
            # Cliente desactivó este aviso desde el panel — no enviamos nada.
            logger.info(f"[shopify-webhook] orders/fulfilled desactivado para agent={_aid_shopify} — saltando")
            return {"status": "ok"}
    elif topic == "orders/paid":
        try:
            _aid_shopify = agent_ids[0]
        except (NameError, IndexError):
            _aids_tmp = await obtener_agent_ids_por_telefono(telefono)
            _aid_shopify = _aids_tmp[0] if _aids_tmp else 1

        # CAPI: Purchase — pago confirmado. Construir productos con
        # retailer_id desde los line_items del payload de Shopify para que
        # Meta haga match con el catálogo (crítico para optimizar campañas).
        try:
            line_items = payload.get("line_items") or []
            productos_capi = []
            for it in line_items:
                # Shopify expone sku en line_items; ese suele coincidir con
                # el retailer_id del catálogo Meta (Shopify Catalog Connector
                # sincroniza SKU → retailer_id por defecto).
                sku = (it.get("sku") or "").strip()
                if sku:
                    productos_capi.append({
                        "producto":        it.get("title", ""),
                        "presentacion":    it.get("variant_title", ""),
                        "cantidad":        int(it.get("quantity", 1)),
                        "precio_unitario": int(float(it.get("price", 0))),
                        "retailer_id":     sku,
                    })
            from agent.capi import capi_purchase
            asyncio.create_task(capi_purchase(
                telefono, total_int, productos_capi,
                order_id=str(payload.get("order_number") or payload.get("id") or ""),
            ))
        except Exception as _e_capi:
            logger.warning(f"[capi] Purchase falló (no bloqueante): {_e_capi}")

        # Sync Shopify → Voco: marcar pago confirmado
        try:
            from agent.memory import upsert_pedido_shopify
            _shopify_oid_p = str(payload.get("id", ""))
            for aid in agent_ids:
                await upsert_pedido_shopify(aid, _shopify_oid_p, {"estado_pago": "pagado", "total": total_int})
            logger.info(f"[shopify-sync] Pedido {nombre_orden} marcado pagado en Voco")
        except Exception as _e_usq_p:
            logger.error(f"[shopify-sync] Error actualizando pago en Voco: {_e_usq_p}")

        from agent.mensajes import format_seguro as _fmt_shopify
        _ctx_shopify = await construir_contexto_placeholders(_aid_shopify)
        _ctx_shopify["numero_pedido"] = nombre_orden
        _ctx_shopify["total"] = f"{total_int:,}"
        mensaje = _fmt_shopify(
            await obtener_mensaje(_aid_shopify, "shopify.order_paid"),
            _ctx_shopify,
        )
        if not mensaje:
            logger.info(f"[shopify-webhook] orders/paid desactivado para agent={_aid_shopify} — saltando")
            return {"status": "ok"}
        # ── Promoción post-venta configurable por agente ──────────────────
        # Si el cliente está bajo un agent con promoción activa Y el subtotal
        # (sin envío) supera el umbral configurado, anexamos el código de
        # descuento. CERO defaults — si el agent no lo configuró en el panel,
        # no enviamos nada. Multi-tenant: cada agent tiene su propia promo.
        subtotal_raw = (
            payload.get("subtotal_price")
            or payload.get("current_subtotal_price")
            or total  # fallback solo si Shopify no expone subtotal
        )
        try:
            subtotal_int = int(float(subtotal_raw))
        except Exception:
            subtotal_int = total_int

        # Reutilizamos los agent_ids ya resueltos arriba para limpieza de
        # carrito — evitamos otro round-trip a BD. Si por alguna razón no
        # se resolvieron, los buscamos ahora.
        try:
            promo_agent_ids = agent_ids  # se resolvió antes en este mismo handler
        except NameError:
            promo_agent_ids = await obtener_agent_ids_por_telefono(telefono)

        for aid in promo_agent_ids:
            try:
                promo = await obtener_descuento_promo(aid)
            except Exception as e:
                logger.error(f"[descuento] error leyendo promo agent={aid}: {e}")
                continue
            if not promo:
                continue  # agent sin promo activa o config inválida
            if subtotal_int < promo["umbral"]:
                logger.info(
                    f"[descuento] {telefono} (agent={aid}) subtotal ${subtotal_int:,} "
                    f"< umbral ${promo['umbral']:,} → no enviar"
                )
                continue
            # Sustituir placeholders del template del cliente
            try:
                anuncio = promo["mensaje"].format(
                    codigo=promo["codigo"],
                    pct=promo["pct"],
                    umbral=f"{promo['umbral']:,}".replace(",", "."),
                )
            except (KeyError, IndexError) as e:
                logger.error(
                    f"[descuento] template inválido para agent={aid}: {e} — "
                    f"usando default"
                )
                anuncio = (
                    f"🎁 Como agradecimiento, usa el código *{promo['codigo']}* "
                    f"en tu próxima compra y obtén *{promo['pct']}% de descuento* 😊"
                )
            mensaje += f"\n\n{anuncio}"
            logger.info(
                f"[descuento] {telefono} (agent={aid}) subtotal ${subtotal_int:,} "
                f">= ${promo['umbral']:,} → enviado código {promo['codigo']}"
            )
            break  # ya enviamos para uno — no duplicar si está en varios agents
    else:
        # orders/create (y cualquier otro topic no manejado explícitamente)
        try:
            _aid_shopify = agent_ids[0]
        except (NameError, IndexError):
            _aids_tmp = await obtener_agent_ids_por_telefono(telefono)
            _aid_shopify = _aids_tmp[0] if _aids_tmp else 1
        from agent.mensajes import format_seguro as _fmt_shopify
        _ctx_shopify = await construir_contexto_placeholders(_aid_shopify)
        _ctx_shopify["numero_pedido"] = nombre_orden
        _ctx_shopify["total"] = f"{total_int:,}"
        mensaje = _fmt_shopify(
            await obtener_mensaje(_aid_shopify, "shopify.order_created"),
            _ctx_shopify,
        )
        if not mensaje:
            logger.info(f"[shopify-webhook] orders/create desactivado para agent={_aid_shopify} — saltando")
            return {"status": "ok"}

    try:
        _prov = await _get_meta_para_agente({"id": _aid_shopify})
        await _prov.enviar_mensaje(telefono, mensaje)
        # Guardar en BD para que aparezca en el inbox
        await _guardar_con_wamid(_prov, telefono, mensaje, agent_id=_aid_shopify)
        # Suprimir follow-up y cierre: esta es una notificación automática,
        # no un mensaje conversacional — no debe activar los timers de seguimiento
        await marcar_followup_enviado(telefono)
        await marcar_cierre_enviado(telefono)
        logger.info(f"Confirmación {topic} enviada a {telefono} ({nombre_orden})")
    except Exception as e:
        logger.error(f"Error enviando confirmación a {telefono}: {e}")

    return {"status": "ok"}
